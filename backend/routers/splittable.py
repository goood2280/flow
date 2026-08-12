"""routers/splittable.py v4.1.0 - multi-prefix transposed view + history + Base features.

v4.1 (2026-04-19, adapter-engineer slice):
  - Module-level DB_BASE removed. All route handlers now call PATHS.db_root /
    PATHS.base_root at request time, so `FLOW_*` env overrides and the
    admin_settings.json `data_roots` block land without a process restart.
  - New endpoint `GET /api/splittable/features` — joins
      `<db_root>/features_et_wafer.parquet` (wafer-level ET, 750 rows)
      + `<db_root>/features_inline_agg.parquet` (wafer-level INLINE aggregate, 50 rows)
    on (lot_id, wafer_id, product) via ET-left-join (Q005 default — preserves
    wafer coverage, INLINE-side cols are null when an ET wafer has no inline
    data). Returns wide-table metadata + columns + sample rows.
  - New endpoint `GET /api/splittable/uniques` — proxies
      `<db_root>/_uniques.json` unchanged, for frontend feature-select
    autocomplete catalog.
"""
import json, datetime, io, csv as csv_mod, hashlib, logging, time, threading, os, gc, shutil, zlib
import contextlib
from pathlib import Path
import sys
from collections import OrderedDict, deque

_BACKEND_ROOT = Path(__file__).resolve().parents[1]
_APP_ROOT = _BACKEND_ROOT.parent
for _path in (_APP_ROOT, _BACKEND_ROOT):
    _raw = str(_path)
    sys.path[:] = [p for p in sys.path if p != _raw]
    sys.path.insert(0, _raw)

from fastapi import APIRouter, Depends, HTTPException, Query, Request, Response
from pydantic import BaseModel, Field
from typing import Any, List, Optional
import polars as pl
from core.paths import PATHS
from core.latest_lot_cache_format import (
    FILE_NAME as LATEST_LOT_STEP_CACHE_FILE,
    FORMAT_COLUMN as LATEST_LOT_STEP_CACHE_FORMAT_COLUMN,
    FORMAT_VERSION as LATEST_LOT_STEP_CACHE_FORMAT_VERSION,
    SOURCE_COLUMN as LATEST_LOT_STEP_CACHE_SOURCE_COLUMN,
    SOURCE_SPLITTABLE as LATEST_LOT_STEP_CACHE_SOURCE,
    normalize_product as _normalize_latest_cache_product,
)
from app_v2.shared.source_adapter import resolve_existing_root, resolve_column
from core.audit import record_user as _audit_user
from core.auth import current_user, is_page_manager, require_page_manager, require_admin
from core.domain import classify_process_area
from core import latest_lot_partitions as _latest_lot_partitions
from core import lot_list_cache as _lot_list_cache
from core import matching_cache as _matching_cache
from core import ml_table_lookup as _ml_table_lookup
from core import search_timing_log as _search_timing_log
from core import s3_sync as _s3
from core.utils import (
    _STR, is_cat, find_lot_wafer_cols, load_json, load_json_cached, save_json, safe_id,
    csv_response, csv_writer_bytes,
)
from core.splittable_sets_cache import invalidate as invalidate_splittable_sets_cache
from app_v2.modules.splittable.rulebook_repository import RulebookRepository
from app_v2.modules.splittable.rulebook_service import RulebookService

rulebook_repo = RulebookRepository()
rulebook_service = RulebookService(rulebook_repo)

# v8.8.26: override CI 매칭 진단 — 실패 경로/스키마 mismatch 를 로그로 가시화.
logger = logging.getLogger("flow.splittable")

router = APIRouter(prefix="/api/splittable", tags=["splittable"])

_DISCOVERY_CACHE_TTL_SEC = 30.0
_RGLOB_CACHE: dict[tuple[str, tuple[str, ...]], tuple[float, list[Path]]] = {}
_FIRST_DATA_FILE_CACHE: dict[tuple[str, tuple[str, ...]], tuple[float, Path | None]] = {}
_DB_ROOTS_CACHE: dict[str, tuple[float, list[Path]]] = {}
_LOT_LOOKUP_CACHE_TTL_SEC = 60.0
_LOT_LOOKUP_CACHE_MAX = 256
_LOT_LOOKUP_CACHE: dict[tuple, tuple[float, dict]] = {}

# 위 캐시들은 접근 시에만 만료를 확인하므로, 다시 조회되지 않는 key 는 메모리에
# 계속 남는다. 주기 sweep 으로 만료 항목을 정리한다.
from core import cache_sweeper as _cache_sweeper
_cache_sweeper.register_ttl_dict("splittable._RGLOB_CACHE", _RGLOB_CACHE, _DISCOVERY_CACHE_TTL_SEC, clock=time.monotonic)
_cache_sweeper.register_ttl_dict("splittable._FIRST_DATA_FILE_CACHE", _FIRST_DATA_FILE_CACHE, _DISCOVERY_CACHE_TTL_SEC, clock=time.monotonic)
_cache_sweeper.register_ttl_dict("splittable._DB_ROOTS_CACHE", _DB_ROOTS_CACHE, _DISCOVERY_CACHE_TTL_SEC, clock=time.monotonic)
_cache_sweeper.register_ttl_dict("splittable._LOT_LOOKUP_CACHE", _LOT_LOOKUP_CACHE, _LOT_LOOKUP_CACHE_TTL_SEC, clock=time.monotonic)
_CSV_ROWS_CACHE: dict[str, tuple[float, int, list[dict]]] = {}
_CSV_ROWS_CACHE_MAX = 32   # 파일 전체 행을 dict 로 들고 있으므로 개수 상한 필수
_SCHEMA_COLUMNS_CACHE: dict[str, tuple[float, int, list[str]]] = {}
_SCHEMA_COLUMNS_CACHE_MAX = 256
SPLITTABLE_VIEW_MAX_WAFERS = 3200
SPLITTABLE_MAX_WAFER_ID = 25


def _db_base() -> Path:
    """Resolve DB root at call time so runtime overrides take effect."""
    return resolve_existing_root("db", PATHS.db_root)


def _base_root() -> Path:
    """Resolve Base root at call time (env / admin_settings / default chain)."""
    return resolve_existing_root("base", PATHS.base_root)


def _path_cache_sig(path: Path | None):
    if path is None:
        return ("", 0.0, 0)
    try:
        st = path.stat()
        return (str(path.resolve()), st.st_mtime, st.st_size)
    except Exception:
        return (str(path), 0.0, 0)


def _lot_lookup_cache_sig(product: str = "") -> tuple:
    try:
        product_fp = _product_path(product) if product else None
        product_sig = _path_cache_sig(product_fp) if product_fp else ("", 0.0, 0)
        lookup_meta_sig = _path_cache_sig(_ml_table_lookup.meta_path_for(product_fp)) if product_fp else ("", 0.0, 0)
    except Exception:
        product_sig = (str(product or ""), 0.0, 0)
        lookup_meta_sig = ("", 0.0, 0)
    return (
        _path_cache_sig(_db_base()),
        _path_cache_sig(_base_root()),
        _path_cache_sig(SOURCE_CFG if "SOURCE_CFG" in globals() else None),
        product_sig,
        lookup_meta_sig,
    )


def _clone_lookup_payload(payload: dict | None) -> dict | None:
    if payload is None:
        return None
    out = {}
    for key, value in payload.items():
        if isinstance(value, list):
            out[key] = list(value)
        elif isinstance(value, dict):
            out[key] = dict(value)
        else:
            out[key] = value
    return out


def _lot_lookup_cache_get(key: tuple) -> dict | None:
    now = time.monotonic()
    cached = _LOT_LOOKUP_CACHE.get(key)
    if cached and now - cached[0] < _LOT_LOOKUP_CACHE_TTL_SEC:
        return _clone_lookup_payload(cached[1])
    if cached:
        _LOT_LOOKUP_CACHE.pop(key, None)
    return None


def _lot_lookup_cache_set(key: tuple, payload: dict) -> dict:
    if len(_LOT_LOOKUP_CACHE) >= _LOT_LOOKUP_CACHE_MAX:
        try:
            _LOT_LOOKUP_CACHE.pop(next(iter(_LOT_LOOKUP_CACHE)))
        except Exception:
            _LOT_LOOKUP_CACHE.clear()
    _LOT_LOOKUP_CACHE[key] = (time.monotonic(), _clone_lookup_payload(payload) or {})
    return payload


PLAN_DIR = PATHS.data_root / "splittable"
PLAN_DIR.mkdir(parents=True, exist_ok=True)
MATCH_CACHE_DIR = PLAN_DIR / "match_cache"
MATCH_CACHE_STATE_FILE = PLAN_DIR / "match_cache_state.json"
MATCH_CACHE_VERSION = 5
MATCH_CACHE_REFRESH_MINUTES_DEFAULT = 30
MATCH_CACHE_REFRESH_MINUTES_MIN = 30
MATCH_CACHE_REFRESH_MINUTES_MAX = 60
MATCH_CACHE_ROOT_COL = "__cache_root_lot_id"
MATCH_CACHE_WAFER_COL = "__cache_wafer_id"
MATCH_CACHE_FAB_COL = "__cache_fab_lot_id"
MATCH_CACHE_TS_COL = "__cache_ts"
# 이 행이 온 FAB 제품 폴더. 매칭은 FAB 전체를 훑으므로(랏 lineage 가 다른 제품
# 폴더에 있을 수 있다) 제품별 물량을 세려면 행마다 출처가 있어야 한다.
MATCH_CACHE_SRC_PRODUCT_COL = "__cache_src_product"
LEGACY_LATEST_LOT_STEP_CACHE_FILE = "splittable_latest_lot_step.parquet"
LATEST_LOT_STEP_CACHE_COLUMNS = [
    LATEST_LOT_STEP_CACHE_FORMAT_COLUMN,
    LATEST_LOT_STEP_CACHE_SOURCE_COLUMN,
    "product",
    # 이 랏이 실제로 속한 FAB 제품 폴더. product(=ML_TABLE 제품, 교차 폴더
    # lineage 포함)와 다르며 물량 집계는 이쪽을 봐야 한다.
    "src_product",
    "root_lot_id",
    "wafer_id",
    "lot_id",
    "step_id",
    "function_step",
    # FAB 가 주는 랏 구분(양산/엔지니어링/모니터…). 매칭 캐시에는 이미 있었지만
    # 이 캐시에 안 실려서 대시보드가 lot_type 으로 물량을 나눠 볼 수 없었다.
    # FAB 에 열이 없는 환경에서는 빈 문자열로 채운다.
    "lot_type",
    "tkout_time",
    "update_time",
]
_MATCH_CACHE_THREAD: threading.Thread | None = None
_MATCH_CACHE_STARTED = False
_MATCH_CACHE_NEXT_TICK_AT = ""
_MATCH_CACHE_STOP = threading.Event()
# 프로세스 종료용 _MATCH_CACHE_STOP 과 다르다 — 이쪽은 관리자가 "지금 이 제품만"
# 중단시키는 신호다. 중단된 제품은 부분 결과를 버리고 다음 제품으로 넘어간다.
_MATCH_CACHE_CANCEL_LOCK = threading.Lock()
_MATCH_CACHE_CANCEL: dict = {"product": "", "by": "", "at": ""}
_MATCH_CACHE_BUILD_LOCK = threading.Lock()
_MATCH_CACHE_AUTO_BUILD_MISS_TTL_SEC = 120.0
_MATCH_CACHE_AUTO_BUILD_MISS: dict[str, tuple[float, str]] = {}
_MATCH_CACHE_JOB_LOCK = threading.Lock()
_MATCH_CACHE_JOB_STATE: dict = {
    "running": False,
    "queued": False,
    "reason": "",
    "started_at": "",
    "finished_at": "",
    "current_product": "",
    "current_started_ts": 0.0,
    "total": 0,
    "done": 0,
    "ok_count": 0,
    "failed_count": 0,
    "skipped_count": 0,
    "paused": False,
    "last_error": "",
    "products": [],
    "order": [],
}
_PLAN_RISK_CACHE: dict[tuple[str, bool], dict] = {}
_PLAN_RISK_CACHE_LOCK = threading.Lock()
_PLAN_RISK_CACHE_MAX = 64
# 엔트리 = (hard_sig, soft_sig, payload, approx_bytes). hard_sig 는 즉시 무효화
# 대상(소스 ML_TABLE = 신규 lot 신호 + 사용자 편집 입력), soft_sig 는 백그라운드
# 스케줄러가 주기적으로 재기록하는 파생 캐시 — soft 만 바뀌면
# stale-while-revalidate 로 즉시 서빙한다.
# 항목 수와 별도로 바이트 예산을 둔다 — wide KNOB payload 는 개당 ~22MB
# (실측 441B/셀)라 128개 상한만으로는 수 GB 까지 자랄 수 있다.
_VIEW_CACHE: OrderedDict[tuple, tuple[tuple, tuple, dict, int]] = OrderedDict()
_VIEW_CACHE_LOCK = threading.Lock()
_VIEW_CACHE_MAX_ENTRIES_DEFAULT = 512
_VIEW_CACHE_BYTES = 0  # 현재 보유 추정치 (lock 하에서만 갱신)
_VIEW_CACHE_CELL_COST = 450  # 레거시 _cells 셀당 파이썬 객체 비용 (실측 441B)
_VIEW_CACHE_COMPACT_CELL_COST = 40  # v2 슬림 행(a/p/m) 셀당 비용 — 캐시는 이쪽만 담는다
_VIEW_CACHE_AUTO_MB_LOCK = threading.Lock()
_VIEW_CACHE_AUTO_MB_CACHE: tuple[float, float] | None = None
_VIEW_CACHE_AUTO_MB_TTL = 60.0
_VIEW_DISK_CACHE_VERSION = 2
_VIEW_DISK_CACHE_LOCK = threading.Lock()
_VIEW_PRODUCT_SIG_CACHE: OrderedDict[tuple[str, ...], tuple[float, tuple]] = OrderedDict()
_VIEW_PRODUCT_SIG_LOCK = threading.Lock()
_VIEW_PRODUCT_SIG_CACHE_MAX = 512
_VIEW_COMPUTE_LOCK = threading.Lock()
_VIEW_COMPUTE_EVENTS: dict[tuple, tuple[int, threading.Event]] = {}

# ── /view cold 계산 전용 레인 ────────────────────────────────────────────────
# 예전에는 resource_guard 의 essential 세마포어(코어수 기준 2~4슬롯)가 /view 요청
# '전체' 를 직렬화했다. 그래서 payload 캐시 HIT(수십 ms)도 앞선 cold 계산(수 초)
# 뒤에 줄을 섰고, 동시 사용자가 늘수록 HIT 응답까지 같이 느려졌다.
# 이제 미들웨어는 /view 를 self-gated 로 통과시키고(app_v2/runtime/resource_guard.py
# DEFAULT_SELF_GATED_PATHS), 실제 메모리를 쓰는 cold 계산 구간만 이 세마포어로
# 직렬화한다. 캐시 HIT·빈 결과·single-flight 대기는 여기 줄서지 않는다.
_VIEW_COLD_LANE_TLS = threading.local()


_VIEW_COLD_CONCURRENCY_MIN = 1
_VIEW_COLD_CONCURRENCY_MAX = 8


def _view_cold_lane_default() -> int:
    if _ml_table_lookup._root_ram_cache_use_dev():
        return 1
    try:
        from core.runtime_limits import effective_cpu_count
        cores = int(effective_cpu_count())
    except Exception:
        cores = 4
    # 운영 기본은 3슬롯이다. 실제 할당 코어가 3보다 적으면 그 수에 맞춰 낮춘다.
    return max(1, min(3, cores))


def _view_cold_lane_concurrency() -> int:
    """우선순위: env > 톱니바퀴 설정(운영/개발 분리) > 코어수 기반 기본값.

    호출할 때마다 다시 읽는다 — 톱니바퀴에서 저장하면 재시작 없이 다음 요청부터
    적용된다(설정 읽기는 2초 메모이즈). polars 스레드 풀과 달리 이 레인은
    런타임 조절이 가능하다."""
    # 개발 서버는 백그라운드 작업과 Flow-i에 여유를 남기기 위해 항상 1슬롯이다.
    if _ml_table_lookup._root_ram_cache_use_dev():
        return 1
    raw = os.environ.get("FLOW_SPLITTABLE_VIEW_COLD_CONCURRENCY", "").strip()
    if raw:
        try:
            return max(_VIEW_COLD_CONCURRENCY_MIN, min(_VIEW_COLD_CONCURRENCY_MAX, int(raw)))
        except Exception:
            pass
    try:
        from core import cache_settings
        saved = cache_settings.get_int_role(
            "view_cold_concurrency", _ml_table_lookup._root_ram_cache_use_dev(), None)
        if saved:
            return max(_VIEW_COLD_CONCURRENCY_MIN, min(_VIEW_COLD_CONCURRENCY_MAX, int(saved)))
    except Exception:
        pass
    return _view_cold_lane_default()


class _ResizableLane:
    """크기를 런타임에 바꿀 수 있는 동시성 게이트.

    threading.Semaphore 는 생성 시점에 크기가 고정이라 톱니바퀴로 조절할 수 없다.
    여기서는 활성 수를 직접 세고 목표 크기를 acquire 시점마다 다시 읽는다.
    대기 중에도 상한이 커질 수 있으므로 최대 1초 간격으로 깨어나 재확인한다."""

    def __init__(self, size_fn):
        self._size_fn = size_fn
        self._cond = threading.Condition()
        self._active = 0

    def acquire(self, timeout: float) -> bool:
        deadline = time.monotonic() + max(0.0, float(timeout))
        with self._cond:
            while True:
                try:
                    limit = max(1, int(self._size_fn()))
                except Exception:
                    limit = 1
                if self._active < limit:
                    self._active += 1
                    return True
                remain = deadline - time.monotonic()
                if remain <= 0:
                    return False
                self._cond.wait(min(remain, 1.0))

    def release(self) -> None:
        with self._cond:
            self._active = max(0, self._active - 1)
            self._cond.notify()

    def stats(self) -> dict:
        with self._cond:
            active = self._active
        try:
            limit = max(1, int(self._size_fn()))
        except Exception:
            limit = 1
        return {"active": active, "limit": limit}


_VIEW_COLD_SEMAPHORE = _ResizableLane(_view_cold_lane_concurrency)


def _view_cold_lane_stats() -> dict:
    out = _VIEW_COLD_SEMAPHORE.stats()
    out["default"] = _view_cold_lane_default()
    out["env_pinned"] = bool(os.environ.get("FLOW_SPLITTABLE_VIEW_COLD_CONCURRENCY", "").strip())
    return out


def _view_cold_lane_wait_sec() -> float:
    return max(1.0, min(600.0, _env_float("FLOW_SPLITTABLE_VIEW_COLD_QUEUE_TIMEOUT_SEC", 90.0)))


def _view_cold_lane_acquire(runtime_profile: dict | None = None) -> bool:
    """cold 계산 슬롯을 잡는다. 대기 시간은 runtime_profile 에 남긴다."""
    wait_started = time.perf_counter()
    ok = _VIEW_COLD_SEMAPHORE.acquire(timeout=_view_cold_lane_wait_sec())
    if runtime_profile is not None:
        runtime_profile["cold_lane_wait_ms"] = (time.perf_counter() - wait_started) * 1000.0
    if ok:
        _VIEW_COLD_LANE_TLS.held = True
    return ok


def _view_cold_lane_release() -> None:
    """획득한 스레드에서만 1회 반납 (미획득 상태 호출은 무시)."""
    if getattr(_VIEW_COLD_LANE_TLS, "held", False):
        _VIEW_COLD_LANE_TLS.held = False
        try:
            _VIEW_COLD_SEMAPHORE.release()
        except Exception:
            pass


def _view_cache_max_entries() -> int:
    try:
        n = int(float(os.environ.get("FLOW_SPLITTABLE_VIEW_CACHE_MAX_ENTRIES", "")
                      or _VIEW_CACHE_MAX_ENTRIES_DEFAULT))
    except Exception:
        n = _VIEW_CACHE_MAX_ENTRIES_DEFAULT
    return max(8, min(4096, n))


def _view_cache_auto_max_mb() -> float:
    """호스트 메모리 비례 hot 응답 예산 — 총량의 15%를 [1GB, 6GB]로 제한.

    SplitTable 읽기 순서의 첫 계층이라 pivot/lookup/FAB 연산을 모두 건너뛴다.
    30GB 호스트에서는 약 4.5GB를 사용하고 전체 cache_budget 지분 상한을 한 번
    더 지나 ET 다운로드·백그라운드 빌드 여유를 보존한다."""
    global _VIEW_CACHE_AUTO_MB_CACHE
    now = time.monotonic()
    with _VIEW_CACHE_AUTO_MB_LOCK:
        cached = _VIEW_CACHE_AUTO_MB_CACHE
        if cached is not None and now - cached[0] < _VIEW_CACHE_AUTO_MB_TTL:
            return cached[1]
    mb = 1024.0
    try:
        from core.runtime_limits import system_memory_snapshot
        total_gb = float(system_memory_snapshot().get("system_memory_total_gb") or 0.0)
        if total_gb > 0:
            mb = max(1024.0, min(6144.0, total_gb * 1024.0 * 0.15))
    except Exception:
        mb = 1024.0
    with _VIEW_CACHE_AUTO_MB_LOCK:
        _VIEW_CACHE_AUTO_MB_CACHE = (now, mb)
    return mb


def _view_cache_max_bytes() -> int:
    raw = str(os.environ.get("FLOW_SPLITTABLE_VIEW_CACHE_MAX_MB", "") or "").strip()
    if raw:
        try:
            mb = float(raw)
        except Exception:
            mb = _view_cache_auto_max_mb()
        budget = int(max(64.0, min(8192.0, mb)) * 1024 * 1024)
    else:
        configured = None
        try:
            from core import cache_settings
            configured = cache_settings.get_float_role(
                "view_mb", _ml_table_lookup._root_ram_cache_use_dev())
        except Exception:
            configured = None
        mb = configured if configured is not None and configured > 0 else _view_cache_auto_max_mb()
        budget = int(max(64.0, min(8192.0, mb)) * 1024 * 1024)
    try:
        from core import cache_budget
        # 운영자 개별 설정도 프로세스 전체 안전 풀은 우회하지 않는다.
        budget = cache_budget.capped("splittable_view_payload", budget)
    except Exception:
        pass
    return budget


def _estimate_view_payload_bytes(payload: dict) -> int:
    compact = payload.get("rows_compact")
    if compact is not None:
        cells = 0
        for r in compact:
            a = r.get("a")
            if isinstance(a, list):
                cells += len(a)
        return 8192 + cells * _VIEW_CACHE_COMPACT_CELL_COST
    cells = 0
    for r in (payload.get("rows") or []):
        c = r.get("_cells")
        if isinstance(c, dict):
            cells += len(c)
    return 8192 + cells * _VIEW_CACHE_CELL_COST


def _view_signature_digest(value: tuple) -> str:
    raw = json.dumps(value, ensure_ascii=False, separators=(",", ":"), default=str)
    return hashlib.sha256(raw.encode("utf-8")).hexdigest()


def _view_disk_cache_enabled() -> bool:
    return _env_bool("FLOW_SPLITTABLE_VIEW_DISK_CACHE", True)


def _view_disk_cache_path(key: tuple) -> Path:
    product = str(key[0] if key else "")
    product_hash = hashlib.sha1(product.encode("utf-8")).hexdigest()[:16]
    key_raw = json.dumps(key, ensure_ascii=False, separators=(",", ":"), default=str)
    key_hash = hashlib.sha256(key_raw.encode("utf-8")).hexdigest()
    return _base_root() / "cache" / "split_table_view_payload" / f"v{_VIEW_DISK_CACHE_VERSION}" / product_hash / f"{key_hash}.json.z"


def _view_disk_cache_read(key: tuple, hard_sig: tuple, soft_sig: tuple) -> tuple[str, dict | None]:
    if not _view_disk_cache_enabled():
        return "miss", None
    fp = _view_disk_cache_path(key)
    try:
        raw = zlib.decompress(fp.read_bytes())
        if _orjson is not None:
            record = _orjson.loads(raw)
        else:
            record = json.loads(raw.decode("utf-8"))
        if not isinstance(record, dict) or int(record.get("version") or 0) != _VIEW_DISK_CACHE_VERSION:
            raise ValueError("unsupported view disk cache version")
        if tuple(record.get("key") or ()) != tuple(key):
            raise ValueError("view disk cache key mismatch")
        if str(record.get("hard") or "") != _view_signature_digest(hard_sig):
            try:
                fp.unlink()
            except OSError:
                pass
            return "miss", None
        payload = record.get("payload")
        if not isinstance(payload, dict):
            raise ValueError("view disk cache payload missing")
        freshness = (
            "fresh" if str(record.get("soft") or "") == _view_signature_digest(soft_sig)
            else "stale"
        )
        return freshness, payload
    except FileNotFoundError:
        return "miss", None
    except Exception:
        logger.debug("SplitTable disk view cache read failed: %s", fp, exc_info=True)
        try:
            fp.unlink()
        except OSError:
            pass
        return "miss", None


def _view_disk_cache_write(key: tuple, hard_sig: tuple, soft_sig: tuple, payload: dict) -> None:
    if not _view_disk_cache_enabled():
        return
    try:
        record = {
            "version": _VIEW_DISK_CACHE_VERSION,
            "key": list(key),
            "hard": _view_signature_digest(hard_sig),
            "soft": _view_signature_digest(soft_sig),
            "payload": payload,
        }
        if _orjson is not None:
            raw = _orjson.dumps(
                record, default=str,
                option=_orjson.OPT_SERIALIZE_NUMPY | _orjson.OPT_NON_STR_KEYS)
        else:
            raw = json.dumps(record, ensure_ascii=False, separators=(",", ":"), default=str).encode("utf-8")
        max_mb = max(1.0, min(256.0, _env_float("FLOW_SPLITTABLE_VIEW_DISK_ENTRY_MAX_MB", 64.0)))
        if len(raw) > int(max_mb * 1024 * 1024):
            return
        compressed = zlib.compress(raw, level=1)
        fp = _view_disk_cache_path(key)
        fp.parent.mkdir(parents=True, exist_ok=True)
        tmp = fp.with_name(f"{fp.name}.{os.getpid()}.{threading.get_ident()}.tmp")
        tmp.write_bytes(compressed)
        os.replace(tmp, fp)
        # 제품별 최근 결과만 디스크에 유지한다. RAM과 달리 압축 bytes라 작지만
        # 검색 이력이 무한히 쌓이지 않도록 쓰기 시점에 저비용 상한을 적용한다.
        limit = int(max(8.0, min(512.0, _env_float("FLOW_SPLITTABLE_VIEW_DISK_MAX_PER_PRODUCT", 64.0))))
        with _VIEW_DISK_CACHE_LOCK:
            files = sorted(fp.parent.glob("*.json.z"), key=lambda p: p.stat().st_mtime, reverse=True)
            for old in files[limit:]:
                try:
                    old.unlink()
                except OSError:
                    pass
    except Exception:
        logger.debug("SplitTable disk view cache write failed", exc_info=True)


def _view_product_signature(paths: list[Path]) -> tuple:
    """공유 드라이브 stat 묶음을 짧게 메모이즈해 payload-cache HIT를 빠르게 한다.

    앱을 통한 plan/tag/config 쓰기는 `_clear_split_view_cache()`를 호출하므로 즉시
    무효화된다. 외부에서 ML_TABLE을 교체한 경우만 기본 2초 안에 감지한다.
    """
    key = tuple(str(Path(path)) for path in paths)
    now = time.monotonic()
    ttl = max(0.0, min(30.0, _env_float("FLOW_SPLITTABLE_VIEW_SIGNATURE_TTL_SEC", 2.0)))
    with _VIEW_PRODUCT_SIG_LOCK:
        cached = _VIEW_PRODUCT_SIG_CACHE.get(key)
        if cached is not None and now - cached[0] <= ttl:
            _VIEW_PRODUCT_SIG_CACHE.move_to_end(key)
            return cached[1]
    value = tuple(_path_cache_sig(path) for path in paths)
    with _VIEW_PRODUCT_SIG_LOCK:
        _VIEW_PRODUCT_SIG_CACHE[key] = (now, value)
        _VIEW_PRODUCT_SIG_CACHE.move_to_end(key)
        while len(_VIEW_PRODUCT_SIG_CACHE) > _VIEW_PRODUCT_SIG_CACHE_MAX:
            _VIEW_PRODUCT_SIG_CACHE.popitem(last=False)
    return value


def _view_compute_begin(key: tuple) -> tuple[bool, threading.Event]:
    """같은 검색 key의 동시 cold 계산을 single-flight로 병합한다."""
    with _VIEW_COMPUTE_LOCK:
        current = _VIEW_COMPUTE_EVENTS.get(key)
        if current is not None:
            return False, current[1]
        event = threading.Event()
        _VIEW_COMPUTE_EVENTS[key] = (threading.get_ident(), event)
        return True, event


def _view_compute_finish(key: tuple | None) -> None:
    # cold 레인 반납은 owner 등록 여부와 무관하게 항상 시도한다 — 모든 종료 경로가
    # 이 함수를 지나므로 여기 한 곳에서 반납하면 누수가 없다(미획득이면 no-op).
    _view_cold_lane_release()
    if key is None:
        return
    event = None
    with _VIEW_COMPUTE_LOCK:
        current = _VIEW_COMPUTE_EVENTS.get(key)
        if current is not None and current[0] == threading.get_ident():
            _VIEW_COMPUTE_EVENTS.pop(key, None)
            event = current[1]
    if event is not None:
        event.set()


def _view_compute_wait_seconds() -> float:
    return max(1.0, min(300.0, _env_float("FLOW_SPLITTABLE_VIEW_SINGLEFLIGHT_WAIT_SEC", 90.0)))
# stale hit → 백그라운드 재검증. 전역 단일 워커 + 병합 큐 — 예전 thread-per-key
# 즉시 실행은 lot_progress 재기록 직후 검색마다 풀 재계산 스레드를 띄워, 5코어의
# 전역 polars 풀을 사용자 검색과 나눠 쓰는 CPU 경쟁(연속 검색 지연)을 만들었다.
# stale 재계산도 SplitTable 응답의 일부이므로 운영 서버에서만 실행한다.
# soft dep(최신 lot/fab 라벨)의 신선도는 쿨다운(기본 3h) 간격 갱신으로 충분하고,
# 신규 lot/사용자 편집은 hard_sig 가 요청 내 동기 반영하므로 정확성과 무관하다.
# TLS.force 는 재검증 워커가 view_split 을 재진입할 때 캐시 서빙을 건너뛰고 강제
# 재계산하게 하는 플래그.
_VIEW_REVALIDATE_TLS = threading.local()
# HTTP 경로에서 타이밍 기록을 직렬화 뒤로 미루기 위한 슬롯. view_split_http 가
# 리스트를 걸어 두면 _record_search_timing 이 즉시 쓰지 않고 여기에 담고,
# 직렬화 시간을 채운 뒤 flush 한다.
_VIEW_TIMING_TLS = threading.local()
_VIEW_REVALIDATE_LOCK = threading.Lock()
_VIEW_REVALIDATE_INFLIGHT: set[tuple] = set()
_VIEW_REVALIDATE_LAST: dict[tuple, float] = {}
_VIEW_REVALIDATE_PENDING: OrderedDict[tuple, tuple[float, dict]] = OrderedDict()
_VIEW_REVALIDATE_WAKE = threading.Event()
_VIEW_REVALIDATE_THREAD: threading.Thread | None = None
_VIEW_REVALIDATE_LAST_ENQUEUE_TS = 0.0
_VIEW_REVALIDATE_PENDING_MAX = 16
_VIEW_REVALIDATE_LAST_MAX = 512
_VIEW_REVALIDATE_COOLDOWN_SEC_DEFAULT = 3 * 3600.0
_VIEW_REVALIDATE_DELAY_SEC_DEFAULT = 30.0
_VIEW_REVALIDATE_BURST_QUIET_SEC = 3.0
# HIT 경로 최적화: 의존 시그니처의 stat 중 product-독립 전역 파일(config/rulebook =
# hard, lot_progress 파생 = soft)만 짧은 TTL 로 캐시한다. 동시 다수 사용자가 매
# 요청마다 동일 전역 파일을 재-stat 하던 공유드라이브 부하를 제거. per-product 파일
# (소스 ML_TABLE/plan/tag/management/custom)은 항상 fresh stat 하므로 사용자 편집·
# 신규 lot 은 지연 없이 즉시 감지된다. 전역 config/rulebook admin 변경과 soft 파생
# 변경만 최대 TTL(≤1s) 지연 — 각각 체감 없음 / SWR 이 흡수.
_VIEW_GLOBAL_SIG_CACHE: dict = {}
_VIEW_GLOBAL_SIG_LOCK = threading.Lock()
_VIEW_GLOBAL_SIG_TTL = 1.0
_VIEW_RAW_FALLBACK_MAX_MB_DEFAULT = 16.0
_MISMATCH_NOTIFY_LOCK = threading.Lock()
_MISMATCH_NOTIFY_WAKE = threading.Event()
_MISMATCH_NOTIFY_PENDING: OrderedDict[tuple, dict] = OrderedDict()
_MISMATCH_NOTIFY_THREAD: threading.Thread | None = None
_MISMATCH_NOTIFY_DEBOUNCE_SEC = 0.2
_MISMATCH_NOTIFY_PENDING_MAX = 2000
PRODUCT_RAM_CACHE_VERSION = 1
PRODUCT_RAM_CACHE_REFRESH_MINUTES_DEFAULT = 30
PRODUCT_RAM_CACHE_REFRESH_MINUTES_MIN = 30
PRODUCT_RAM_CACHE_REFRESH_MINUTES_MAX = 240
PRODUCT_RAM_CACHE_MAX_GB_DEFAULT = 2.0
ROOT_LOT_CACHE_LIMIT_MAX = 50000
_PRODUCT_RAM_CACHE_LOCK = threading.RLock()
_PRODUCT_RAM_CACHE: dict[str, dict] = {}
_PRODUCT_RAM_CACHE_STATUS: dict[str, dict] = {}
_PRODUCT_RAM_CACHE_REFRESHING: set[str] = set()
_PRODUCT_RAM_CACHE_BUILD_LOCK = threading.Lock()
_PRODUCT_RAM_CACHE_STOP = threading.Event()
_PRODUCT_RAM_CACHE_THREAD: threading.Thread | None = None
_PRODUCT_RAM_CACHE_STARTED = False
_PRODUCT_RAM_CACHE_NEXT_TICK_AT = ""
_PRODUCT_RAM_CACHE_JOB_LOCK = threading.Lock()
_PRODUCT_RAM_CACHE_JOB_STATE: dict = {
    "running": False,
    "queued": False,
    "reason": "",
    "started_at": "",
    "finished_at": "",
    "current_product": "",
    "total": 0,
    "done": 0,
    "ok_count": 0,
    "failed_count": 0,
    "skipped_count": 0,
    "last_error": "",
    "products": [],
    "order": [],
}
LONG_PIVOT_CACHE_VERSION = 1
_LONG_PIVOT_CACHE_DIR = PLAN_DIR / "long_pivot_cache"
_LONG_PIVOT_JOB_LOCK = threading.Lock()
_LONG_PIVOT_QUEUE: deque[tuple[str, str, bool]] = deque()
_LONG_PIVOT_JOB_THREAD: threading.Thread | None = None
_LONG_PIVOT_JOB_STATE: dict = {
    "running": False,
    "queued": False,
    "current": "",
    "started_at": "",
    "finished_at": "",
    "last_error": "",
    "last_source": "",
}
PREFIX_CFG = PLAN_DIR / "prefix_config.json"
DEFAULT_PREFIXES = ["KNOB", "MASK", "INLINE", "VM", "FAB"]
PLAN_ALLOWED_PREFIXES = ["KNOB", "MASK", "FAB"]  # Only these can have plan values
CUSTOM_TAG_PREFIX = "TAG"
MANAGEMENT_ROW_PREFIX = "MGMT"
_INVALID_CUSTOM_TOKENS = {"undefined", "null"}
# v8.8.6: paste 세트 공유 저장소 — LocalStorage 대신 BE 에 올려 팀 공용 풀 + CUSTOM 탭 연동.
PASTE_SETS_FILE = PLAN_DIR / "paste_sets.json"
# v8.4.9: 엑셀 메모/태그 저장소 — wafer 단위(tag) + parameter 단위(memo) 공용.
#   scope="wafer": target_key = "{product}__{root_lot_id}__W{wafer_id}"
#   scope="param": target_key = "{product}__{root_lot_id}__W{wafer_id}__{param}"
# 각 항목은 {id, text, username, created_at} 를 보관하고 작성자/관리자만 삭제 가능.
NOTES_FILE = PLAN_DIR / "notes.json"
TRACKER_ISSUES_FILE = PATHS.data_root / "tracker" / "issues.json"
INFORMS_FILE = PATHS.data_root / "informs" / "informs.json"


def _clean_custom_token(value: Any) -> str:
    if not isinstance(value, str):
        return ""
    token = value.strip()
    if not token or token.casefold() in _INVALID_CUSTOM_TOKENS:
        return ""
    return token


def _clean_custom_column_name(value: Any, *, allow_management: bool = False) -> str:
    column = _clean_custom_token(value)
    if not column:
        return ""
    if not allow_management and column.upper().startswith(f"{MANAGEMENT_ROW_PREFIX}_"):
        return ""
    return column


def _clean_custom_columns(columns: Any, *, allow_management: bool = False) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for raw in columns if isinstance(columns, list) else []:
        column = _clean_custom_column_name(raw, allow_management=allow_management)
        if not column or column in seen:
            continue
        seen.add(column)
        out.append(column)
    return out


def _clean_custom_set_name(value: Any) -> str:
    return _clean_custom_token(value)


def _custom_file_path_for_name(value: Any) -> tuple[Path, str]:
    name = _clean_custom_set_name(value)
    safe_name = safe_id(name) if name else ""
    if not name or not safe_name or safe_name.casefold() in _INVALID_CUSTOM_TOKENS:
        raise HTTPException(400, "custom name required")
    return PLAN_DIR / f"custom_{safe_name}.json", name


def _custom_file_stem_invalid(path: Path) -> bool:
    stem = path.stem
    raw = stem[len("custom_"):] if stem.startswith("custom_") else stem
    return not _clean_custom_set_name(raw)


def _sanitize_custom_record(record: Any, path: Path | None = None, *, persist: bool = False) -> dict | None:
    if not isinstance(record, dict):
        return None
    if path is not None and _custom_file_stem_invalid(path):
        return None
    name = _clean_custom_set_name(record.get("name"))
    if not name:
        return None
    columns = _clean_custom_columns(record.get("columns") or [])
    if not columns:
        return None
    cleaned = {**record, "name": name, "columns": columns}
    if persist and path is not None and cleaned != record:
        save_json(path, cleaned)
    return cleaned


def _clean_overlay_store_data(data: Any, *, allow_management: bool = True) -> tuple[dict, bool]:
    if not isinstance(data, dict):
        return {"columns": [], "values": {}}, False
    changed = False
    cleaned_cols = []
    for raw in data.get("columns") if isinstance(data.get("columns"), list) else []:
        if not isinstance(raw, dict):
            changed = True
            continue
        column = _clean_custom_column_name(raw.get("column"), allow_management=allow_management)
        if not column:
            changed = True
            continue
        entry = {**raw, "column": column}
        if entry != raw:
            changed = True
        cleaned_cols.append(entry)
    cleaned_values = {}
    values = data.get("values") if isinstance(data.get("values"), dict) else {}
    for raw_key, value in values.items():
        parts = str(raw_key or "").split("|", 3)
        if len(parts) != 4:
            changed = True
            continue
        column = _clean_custom_column_name(parts[3], allow_management=allow_management)
        if not column:
            changed = True
            continue
        parts[3] = column
        key = "|".join(parts)
        if key != raw_key:
            changed = True
        cleaned_values[key] = value
    return {"columns": cleaned_cols, "values": cleaned_values}, changed


def _load_prefixes():
    prefixes = load_json_cached(PREFIX_CFG, DEFAULT_PREFIXES)
    if not isinstance(prefixes, list):
        prefixes = list(DEFAULT_PREFIXES)
    out = [str(p).strip().upper() for p in prefixes if str(p).strip()]
    if CUSTOM_TAG_PREFIX not in out:
        out.append(CUSTOM_TAG_PREFIX)
    return out


def _cast_cats_lazy(lf):
    """Cast Categorical to Utf8 in a LazyFrame."""
    try:
        schema = lf.collect_schema()
    except Exception:
        schema = lf.schema
    casts = [pl.col(n).cast(_STR, strict=False) for n, d in schema.items() if is_cat(d)]
    out = lf.with_columns(casts) if casts else lf
    try:
        from core.utils import filter_valid_wafer_ids_lazy
        return filter_valid_wafer_ids_lazy(out, list(schema.keys()))
    except Exception:
        return out


def _scan_cast_options():
    try:
        return pl.ScanCastOptions(categorical_to_string="allow")
    except Exception:
        return None


# ── parquet 스키마 메모이즈 ───────────────────────────────────────────────
# `collect_schema()` 는 파일 footer 를 파싱하므로 컬럼 수에 비례해 비싸다
# (4000컬럼 실측 9.2ms/회). 그런데 cold 검색 1건이 이 호출을 16~25회 하고,
# 같은 파일에 매번 같은 답을 다시 묻는다 — 넓은 테이블에서는 검색 시간의
# 대부분이 데이터 읽기(21ms)가 아니라 이 반복 조회(~164ms)였다.
# 파일 mtime+size 로 키를 잡아 내용이 바뀌면 자동 무효화된다.
_SCAN_SCHEMA_CACHE: OrderedDict[tuple, tuple] = OrderedDict()
_SCAN_SCHEMA_CACHE_LOCK = threading.Lock()
_SCAN_SCHEMA_CACHE_MAX = 256


def _scan_schema_cache_key(source, hive_partitioning) -> tuple | None:
    """(경로, mtime, size, hive) 키. stat 실패(원격/미존재)면 None = 캐시 미사용."""
    try:
        p = Path(str(source))
        st = p.stat()
        return (str(p), st.st_mtime, st.st_size, bool(hive_partitioning))
    except Exception:
        return None


def _cached_scan_schema(source, hive_partitioning=None):
    """(schema_override, column_names) 반환. schema_override 는 Categorical→String
    드리프트 보정이 필요할 때만 dict, 아니면 None (기존 계약 유지)."""
    key = _scan_schema_cache_key(source, hive_partitioning)
    if key is not None:
        with _SCAN_SCHEMA_CACHE_LOCK:
            hit = _SCAN_SCHEMA_CACHE.get(key)
            if hit is not None:
                _SCAN_SCHEMA_CACHE.move_to_end(key)
                return hit
    try:
        kwargs = {}
        if hive_partitioning is not None:
            kwargs["hive_partitioning"] = hive_partitioning
        schema = pl.scan_parquet(str(source), **kwargs).collect_schema()
    except Exception:
        return None, None
    out = {}
    changed = False
    for name, dtype in schema.items():
        if is_cat(dtype):
            out[name] = _STR
            changed = True
        else:
            out[name] = dtype
    value = (out if changed else None, list(schema.names()))
    if key is not None:
        with _SCAN_SCHEMA_CACHE_LOCK:
            _SCAN_SCHEMA_CACHE[key] = value
            _SCAN_SCHEMA_CACHE.move_to_end(key)
            while len(_SCAN_SCHEMA_CACHE) > _SCAN_SCHEMA_CACHE_MAX:
                _SCAN_SCHEMA_CACHE.popitem(last=False)
    return value


def _first_scan_schema_with_string_cats(source, hive_partitioning=None):
    if not isinstance(source, (list, tuple)) or not source:
        return None
    return _cached_scan_schema(source[0], hive_partitioning)[0]


def _scan_parquet_compat(source, **kwargs):
    """Scan parquet while accepting String/Categorical drift across partitions."""
    scan_kwargs = dict(kwargs)
    hive = scan_kwargs.get("hive_partitioning")
    names = None
    if "schema" not in scan_kwargs:
        is_multi = isinstance(source, (list, tuple))
        first = (source[0] if source else None) if is_multi else source
        if first is not None and not isinstance(first, (list, tuple)):
            schema, names = _cached_scan_schema(first, hive_partitioning=hive)
            # Categorical 드리프트 보정은 원래 계약대로 다중 파티션 스캔에만 적용.
            # 단일 파일은 드리프트가 없고, names 캐시는 양쪽 다 재사용한다.
            if schema and is_multi:
                scan_kwargs["schema"] = schema
    if hive:
        # hive 파티션 컬럼은 개별 파일 스키마에 없다 — 컬럼 목록이 불완전할 수
        # 있으므로 이 경우엔 넘기지 않고 기존처럼 lf 에서 직접 읽게 둔다.
        names = None
    opts = _scan_cast_options()
    if opts is not None and "cast_options" not in scan_kwargs:
        scan_kwargs["cast_options"] = opts
    try:
        lf = pl.scan_parquet(source, **scan_kwargs)
    except TypeError:
        scan_kwargs.pop("cast_options", None)
        lf = pl.scan_parquet(source, **scan_kwargs)
    try:
        from core.utils import filter_valid_wafer_ids_lazy
        # 방금 구한 컬럼 이름을 넘겨 두 번째 collect_schema 를 없앤다.
        # (hive 파티션 컬럼은 스캔 뒤에만 보이지만, wafer 컬럼 탐지 용도라 무관)
        return filter_valid_wafer_ids_lazy(lf, names)
    except Exception:
        return lf


import re as _re
_NUM_RE = _re.compile(r"(\d+(?:\.\d+)?)")
_PREFIX_NUM_RE = _re.compile(r"^(\d+(?:\.\d+)*)(?:[_\s-]|$)")


def _version_num_key(raw: str) -> tuple:
    try:
        return tuple(int(p) for p in str(raw).split("."))
    except Exception:
        return (float("inf"),)

def _natural_param_key(name: str):
    """v8.4.4 — prefix 뒤 숫자(정수/소수) 기준 자연 정렬 키 생성.
    예: 'KNOB_12.0_ASV_FOO' → (12.0, '_ASV_FOO', 'KNOB')
    숫자가 없으면 prefix 를 뺀 본문 natural token 기준으로 후순.
    v8.8.14: 내부 문자열 tail 도 자연 정렬(숫자/비숫자 분리)로 안정화 →
      `KNOB_10_FOO` 뒤에 `KNOB_2_FOO` 가 오는 오작동 방지.
    v9.0.3: prefix 는 정렬 tie-breaker 로만 사용.
      여러 prefix(KNOB/MASK/INLINE/VM...)를 같이 볼 때 prefix 그룹이 먼저 묶이지 않고
      prefix 를 제거한 항목명/순번 기준으로 자연정렬된다.
    v9.5.x: 요청에 따라 prefix 를 다시 최상위 정렬 기준으로 복원.
      정렬 순서: KNOB -> MASK -> FAB -> INLINE -> VM -> 기타
    v10.0.x: TAG_ 는 사용자가 공정 번호를 붙여 만든 주석 열이라 예외다.
      앞머리 숫자가 있으면 그 숫자 위치(KNOB 과 같은 rank)로 끼워 넣고,
      숫자가 없을 때만 기타로 뒤에 붙인다 — prefix 최우선 복원 때 TAG 가
      통째로 맨 뒤로 밀려 "번호 위치 유지" 기능이 죽어 있었다.
    """
    if not name: return (99, 1, (), (), "")
    raw = str(name)
    parts = raw.split("_", 1)
    pfx = parts[0] if len(parts) > 1 else ""
    rest = parts[1] if len(parts) > 1 else raw
    # split rest into natural tokens (numbers → version tuple, others → lowercased str)
    tail: list = []
    for tok in _NUM_RE.split(rest):
        if tok == "":
            continue
        if _NUM_RE.fullmatch(tok):
            tail.append(("n", _version_num_key(tok)))
        else:
            tail.append(("s", tok.lower()))
            
    _PREFIX_RANK = {
        "knob": 0,
        "mask": 1,
        "fab": 2,
        "inline": 3,
        "vm": 4,
    }
    pfx_lower = pfx.lower()
    rank = _PREFIX_RANK.get(pfx_lower, 99)

    # Only the immediate segment after the prefix is the primary process/order
    # key. Numbers buried later in the feature name must not split 1.0/2.0/2.1
    # process-order groups.
    m = _PREFIX_NUM_RE.search(rest)
    if m:
        if pfx_lower == "tag":
            rank = _PREFIX_RANK["knob"]
        return (rank, 0, _version_num_key(m.group(1)), tuple(tail), pfx_lower)
    return (rank, 1, (), tuple(tail), pfx_lower)


# v8.8.14/v9.0.7: ML_TABLE 컬럼 display rename.
#   KNOB 는 원래 feature 명(예: KNOB_1.0 STI)을 유지한다. 적용 공정 정보는
#   ppid_knob.csv + Vehicle_matching.csv detail panel 에서만 보여준다.
#   규칙:
#     KNOB_<feature>   → 표시명 유지
#     INLINE_<item_id>               → 표시명 유지
#     VM_<step_desc>_<item_id>       → 표시명 유지
#   display 이름과 원본 col 이름(`_param`) 을 그대로 보존 → plan/notes/
#   knob_meta lookup 이 깨지지 않음.
def _safe_step_segment(s: str) -> str:
    """step_desc/function_step 값에서 공백/특수문자 제거 → 컬럼명 조각으로 안전하게."""
    if not s:
        return ""
    return _re.sub(r"[^A-Za-z0-9]+", "_", str(s)).strip("_")


_RULE_ORDER_RE = _re.compile(r"^R(\d+)$", _re.I)


def _rule_order_label(raw: object, fallback_idx: int = 1) -> str:
    text = str(raw or "").strip().upper()
    if text == "RO":
        return "RO"
    m = _RULE_ORDER_RE.match(text)
    if m:
        return f"R{int(m.group(1))}"
    if text.isdigit():
        return f"R{int(text)}"
    return text or f"R{max(1, fallback_idx)}"


def _rule_order_sort_key(label: object) -> tuple[int, int, str]:
    text = str(label or "").strip().upper()
    if text == "RO":
        return (1, 0, text)
    m = _RULE_ORDER_RE.match(text)
    if m:
        return (0, int(m.group(1)), text)
    if text.isdigit():
        return (0, int(text), text)
    return (2, 0, text)


def _split_product_core(product: object) -> str:
    text = str(product or "").strip().replace("\\", "/")
    if not text:
        return ""
    text = Path(text).name
    m = _re.search(r"ML_TABLE_", text, flags=_re.I)
    if m:
        text = text[m.end():]
    ext = _re.search(r"\.(?:parquet|csv)(?=$|[\s,，、()[\]{}])", text, flags=_re.I)
    if ext:
        text = text[:ext.start()]
    return text.strip(" \t\r\n,，、()[]{}")


def _product_aliases(product: str) -> set[str]:
    """Soft-landing product matcher for registry/rulebook joins."""
    raw = str(product or "").strip()
    if not raw:
        return set()
    out = {raw.upper()}
    core = _split_product_core(raw)
    if core:
        out.add(core.upper())
    up = core.upper()
    if up.startswith("PRODUCT_A"):
        if up.endswith("0"):
            out.update({"PRODA", "PRODA0", "PRODUCT_A0"})
        elif up.endswith("1"):
            out.update({"PRODA", "PRODA1", "PRODUCT_A1"})
        else:
            out.update({"PRODA", "PRODA0", "PRODA1", "PRODUCT_A0", "PRODUCT_A1"})
    elif up.startswith("PRODUCT_B"):
        out.update({"PRODB", "PRODUCT_B"})
    elif up == "PRODA0":
        out.update({"PRODA", "PRODUCT_A0"})
    elif up == "PRODA1":
        out.update({"PRODA", "PRODUCT_A1"})
    elif up == "PRODA":
        out.update({"PRODA0", "PRODA1", "PRODUCT_A0", "PRODUCT_A1"})
    elif up == "PRODB":
        out.update({"PRODUCT_B"})
    return out


def _product_alias_keys(product: str) -> set[str]:
    return {str(alias or "").strip().casefold() for alias in _product_aliases(product) if str(alias or "").strip()}


def _product_value_matches(product: str, row_product: object, *, allow_common: bool = True) -> bool:
    """Case-insensitive product/alias match for rulebook rows."""
    if not str(product or "").strip():
        return True
    row_values = _product_cell_tokens(row_product)
    if not row_values:
        return allow_common
    product_keys = _product_alias_keys(product)
    return any(row_value.casefold() in product_keys for row_value in row_values)


def _step_matching_product_alias_keys(product: str) -> set[str]:
    raw = str(product or "").strip()
    if not raw:
        return set()
    core = _split_product_core(raw)
    aliases = {raw, core}
    if core:
        aliases.add(f"ML_TABLE_{core}")
    up = core.upper()
    if up == "PRODA0":
        aliases.add("PRODUCT_A0")
    elif up == "PRODA1":
        aliases.add("PRODUCT_A1")
    elif up == "PRODUCT_A0":
        aliases.add("PRODA0")
    elif up == "PRODUCT_A1":
        aliases.add("PRODA1")
    elif up == "PRODB":
        aliases.add("PRODUCT_B")
    elif up == "PRODUCT_B":
        aliases.add("PRODB")
    return {str(alias or "").strip().casefold() for alias in aliases if str(alias or "").strip()}


def _product_cell_tokens(row_product: object) -> list[str]:
    row_value = str(row_product or "").strip()
    if not row_value:
        return []
    return [part.strip() for part in _re.split(r"[,，、]", row_value) if part.strip()]


def _step_matching_product_matches(product: str, row_product: object, *, allow_common: bool = True) -> bool:
    if not str(product or "").strip():
        return True
    row_values = _product_cell_tokens(row_product)
    if not row_values:
        return allow_common
    product_keys = _step_matching_product_alias_keys(product)
    for row_value in row_values:
        row_keys = _step_matching_product_alias_keys(row_value)
        if row_keys and any(key in product_keys for key in row_keys):
            return True
    return False


def _step_desc_match_key(value: object) -> str:
    return _re.sub(r"\s+", " ", str(value or "").strip()).casefold()


_PRODUCT_FILE_EXTS = (".parquet", ".csv")


def _canonical_mltable_product_name(product: str, allow_bare: bool = False) -> str:
    """Return the canonical SplitTable product id for an ML_TABLE file/name."""
    raw = str(product or "").strip()
    if not raw:
        return ""
    if raw.casefold().startswith("ml_table_"):
        tail = raw[len("ML_TABLE_"):].strip()
    elif allow_bare:
        tail = raw
    else:
        return ""
    return f"ML_TABLE_{tail}".upper() if tail else ""


def _is_mltable_product_file(path: Path) -> bool:
    return (
        path.is_file()
        and path.suffix.lower() in _PRODUCT_FILE_EXTS
        and bool(_canonical_mltable_product_name(path.stem))
    )


def _lot_override_for(cfg: dict, product: str) -> dict:
    """Resolve lot_overrides by product name with case-insensitive ML_TABLE matching."""
    overrides = (cfg or {}).get("lot_overrides") or {}
    if not isinstance(overrides, dict):
        return {}
    keys = [str(product or "").strip()]
    canonical = _canonical_mltable_product_name(product, allow_bare=True)
    if canonical:
        keys.append(canonical)
    for key in keys:
        if key and isinstance(overrides.get(key), dict):
            return overrides.get(key) or {}
    folded = {k.casefold() for k in keys if k}
    for key, value in overrides.items():
        if str(key or "").casefold() in folded and isinstance(value, dict):
            return value
    return {}


def _build_col_rename_map(selected_cols: list, product: str) -> dict:
    """raw column name → display name. 매칭 없으면 원본 반환(맵에 키 없음)."""
    return {}


def _color_for_value(val, uniq_map, palette):
    """UI 와 동일하게 unique 값 별 색상 (RGB hex, openpyxl fgColor 포맷 - no #).
    palette 는 색 리스트, uniq_map 은 {value: index}.
    """
    if val is None or val == "":
        return None
    idx = uniq_map.get(val)
    if idx is None: return None
    return palette[idx % len(palette)]


def _detect_lot_wafer(lf, product: str = ""):
    """v8.4.4: product 별 source_config.json 의 lot_overrides 를 우선 참조.
    override 가 실제 schema 에 존재할 때만 사용 (소프트랜딩). 아니면 기본 감지.
    """
    if product:
        try:
            cfg = load_json(SOURCE_CFG, {"lot_overrides": {}}) if SOURCE_CFG.exists() else {}
            ov = _lot_override_for(cfg, product)
            schema_names_list = lf.collect_schema().names() if hasattr(lf, "collect_schema") else list(lf.schema.keys())
            root_col = _ci_resolve_in(ov.get("root_col") or "", schema_names_list) or None
            wf_col = _ci_resolve_in(ov.get("wf_col") or "", schema_names_list) or None
            if root_col or wf_col:
                # Fill missing with auto-detect
                auto_r, auto_w = find_lot_wafer_cols(schema_names_list)
                return (root_col or auto_r, wf_col or auto_w)
        except Exception:
            pass
    try:
        schema_names = lf.collect_schema().names()
    except Exception:
        schema_names = lf.schema
    return find_lot_wafer_cols(schema_names)


# _product_path 결과 메모이즈. 이 함수는 한 요청에서 2~5회 호출되고 1회가 ~3ms 다
# — 파일이 base_root 에 없으면 확장자마다 _find_ci_path(디렉터리 전체 스캔)를 돌린
# 뒤에야 db_base 로 넘어가기 때문이다. 실측에서 **캐시 히트 응답 시간의 42%** 가
# 이 헛도는 디렉터리 스캔이었고, 순수 파이썬이라 GIL 을 쥔 채 돌아 동시 검색에서
# 5배로 증폭됐다.
#
# 무효화는 TTL + 루트 시그니처 2중이다. TTL 만으로는 "제품 파일을 방금 넣었는데
# 404" 가 최대 TTL 만큼 남고, 루트 stat 만으로는 같은 디렉터리에 다른 파일이
# 생겨도 전량 무효화된다. 둘을 같이 보면 새 파일 투입은 루트 mtime 변화로 즉시
# 잡히고, 그 외에는 TTL 안에서 스캔을 건너뛴다.
_PRODUCT_PATH_CACHE: dict[str, tuple[float, tuple, Path]] = {}
_PRODUCT_PATH_CACHE_LOCK = threading.Lock()
_PRODUCT_PATH_TTL = 3.0
_PRODUCT_PATH_CACHE_MAX = 256


def _product_path_roots_sig() -> tuple:
    """base/db 루트의 (mtime, size) — 파일이 새로 들어오면 디렉터리 mtime 이 바뀐다."""
    sig = []
    for root in (_base_root(), _db_base()):
        try:
            st = root.stat()
            sig.append((str(root), st.st_mtime, st.st_size))
        except Exception:
            sig.append((str(root), 0.0, 0))
    return tuple(sig)


def _product_path(product: str):
    """Find product file. v8.4.3 — Base scope (ML_TABLE_PRODA/B etc.) 우선,
    이후 DB 루트(legacy) 로 폴백. ML 중심 설계로 전환.

    결과는 짧게 메모이즈한다(위 _PRODUCT_PATH_CACHE 주석 참고). 404 는 캐시하지
    않는다 — 파일을 넣자마자 다시 물었을 때 없다고 답하면 안 된다.
    """
    cache_key = str(product or "")
    now = time.monotonic()
    roots_sig = _product_path_roots_sig()
    with _PRODUCT_PATH_CACHE_LOCK:
        hit = _PRODUCT_PATH_CACHE.get(cache_key)
        if hit is not None and (now - hit[0]) < _PRODUCT_PATH_TTL and hit[1] == roots_sig:
            return hit[2]

    fp = _product_path_uncached(product)

    with _PRODUCT_PATH_CACHE_LOCK:
        if len(_PRODUCT_PATH_CACHE) >= _PRODUCT_PATH_CACHE_MAX:
            _PRODUCT_PATH_CACHE.clear()
        _PRODUCT_PATH_CACHE[cache_key] = (now, roots_sig, fp)
    return fp


def _clear_product_path_cache() -> None:
    """제품 파일 배치가 바뀌었을 때(업로드/삭제/루트 전환) 즉시 비운다."""
    with _PRODUCT_PATH_CACHE_LOCK:
        _PRODUCT_PATH_CACHE.clear()


def _product_path_uncached(product: str):
    raw = str(product or "").strip()
    canonical = _canonical_mltable_product_name(raw, allow_bare=True)
    names = []
    for name in (raw, canonical):
        if name and name not in names:
            names.append(name)
    base_root = _base_root()
    db_base = _db_base()
    for root in (base_root, db_base):
        if not root or not root.exists():
            continue
        for name in names:
            for ext in _PRODUCT_FILE_EXTS:
                fp = root / f"{name}{ext}"
                if fp.exists():
                    return fp
                ci = _find_ci_path(root, f"{name}{ext}")
                if ci is not None and ci.is_file():
                    return ci
        try:
            targets = {n.casefold() for n in names if n}
            for fp in sorted(root.iterdir(), key=lambda p: p.name.lower()):
                if fp.is_file() and fp.suffix.lower() in _PRODUCT_FILE_EXTS and fp.stem.casefold() in targets:
                    return fp
        except Exception:
            pass
    raise HTTPException(404, f"Product not found: {product}")


def _scan_product_base(product: str):
    """Scan the ML_TABLE file only, without FAB override joins."""
    product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    ram_lf = _product_ram_cache_lazyframe(product)
    if ram_lf is not None:
        return ram_lf
    fp = _product_path(product)
    if fp.suffix.lower() == ".csv":
        return _cast_cats_lazy(pl.scan_csv(str(fp), infer_schema_length=5000))
    return _cast_cats_lazy(_scan_parquet_compat(str(fp)))


def _scan_product_base_lookup_cache(
    product: str,
    root_lot_id: str = "",
    wafer_ids: str = "",
    runtime_profile: dict | None = None,
):
    """Use the ML_TABLE root lookup cache when a lot-scoped view is available."""
    if not str(root_lot_id or "").strip():
        return None
    fp = _product_path(product)
    _ml_table_lookup.record_root_access(fp, root_lot_id)
    ram_lf = _product_ram_cache_lazyframe(product)
    if ram_lf is not None:
        if runtime_profile is not None:
            runtime_profile["product_cache_hit"] = True
            runtime_profile["root_data_source"] = "product_ram"
        try:
            lot_col, wf_col = _detect_lot_wafer(ram_lf, product)
            return _filter_lot_wafer(ram_lf, lot_col, wf_col, root_lot_id, wafer_ids)
        except Exception as exc:
            logger.debug("Product RAM cache scope filter unavailable product=%s root=%s: %s", product, root_lot_id, exc)
            return ram_lf
    try:
        if fp.suffix.lower() != ".parquet":
            return None
        lf, status = _ml_table_lookup.scan_root_lot_cache(fp, root_lot_id, wafer_ids=wafer_ids, profile=runtime_profile)
        if runtime_profile is not None:
            runtime_profile["root_cache_status"] = status.get("status") or ""
            runtime_profile["root_cache_hit"] = lf is not None
        if lf is not None:
            return _cast_cats_lazy(lf)
    except Exception as exc:
        logger.debug("ML_TABLE lookup cache unavailable product=%s root=%s: %s", product, root_lot_id, exc)
    return None


def _env_bool(name: str, default: bool = False) -> bool:
    raw = os.environ.get(name)
    if raw is None or str(raw).strip() == "":
        return default
    return str(raw).strip().lower() in {"1", "true", "yes", "on", "enabled"}


def _truthy_value(value: Any) -> bool:
    if isinstance(value, bool):
        return value
    if isinstance(value, (int, float)):
        return value != 0
    if isinstance(value, str):
        return value.strip().lower() in {"1", "true", "yes", "on", "enabled"}
    return False


def _env_float(name: str, default: float) -> float:
    try:
        return float(os.environ.get(name, "") or default)
    except Exception:
        return default


def _split_view_raw_fallback_max_bytes() -> int:
    mb = max(0.0, _env_float("FLOW_SPLITTABLE_VIEW_RAW_FALLBACK_MAX_MB", _VIEW_RAW_FALLBACK_MAX_MB_DEFAULT))
    return int(mb * 1024 * 1024)


def _lookup_cache_public_meta(status: dict | None, queued: dict | None = None) -> dict:
    status = status or {}
    queued = queued or {}
    queued_status = str(queued.get("status") or "").strip()
    queued_flag = bool(queued.get("ok") or queued.get("queued") or queued_status in {"queued", "running"})
    meta = status.get("meta") or {}
    return {
        "status": queued_status if queued_flag else str(status.get("status") or ""),
        "has_cache": bool(status.get("has_cache")),
        "source_stale": bool(status.get("source_stale")),
        "job_status": status.get("job_status") or "",
        "queued": queued_flag,
        "root_lot_id_count": int(status.get("root_lot_id_count") or meta.get("root_lot_id_count") or 0),
        "candidate_index": bool(
            status.get("candidate_index")
            or (meta.get("candidate_index") or {}).get("has_index")
        ),
        # 산출물이 스스로 들고 있는 완료 시각/소요. 이벤트 로그가 잘리거나
        # (7일 창) 오프로드·스킵으로 완료 줄이 안 남아도 이 둘은 남는다.
        "built_at": str(meta.get("built_at") or ""),
        "build_seconds": float(meta.get("build_seconds") or 0.0),
    }


def _split_view_should_defer_raw_fallback(fp: Path) -> bool:
    if Path(fp).suffix.lower() != ".parquet":
        return False
    threshold = _split_view_raw_fallback_max_bytes()
    if threshold <= 0:
        return True
    try:
        source_size = int(Path(fp).stat().st_size)
    except Exception:
        source_size = 0
    return bool(source_size >= threshold)


def _lookup_cache_public_meta_for(product: str) -> dict:
    """제품의 lookup 캐시 상태만 (후보 목록 없이) 읽는다.

    `_root_lot_lookup_cache_candidates` 와 달리 candidate index 를 통째로
    읽지 않는다 — 상태 플래그만 필요할 때 쓴다.
    """
    try:
        fp = _product_path(product)
    except Exception:
        return {}
    if Path(fp).suffix.lower() != ".parquet":
        return {}
    try:
        return _lookup_cache_public_meta(_ml_table_lookup.cache_status(fp))
    except Exception:
        return {}


def _root_lot_lookup_cache_candidates(product: str, prefix: str = "", limit: int = 500) -> dict | None:
    try:
        fp = _product_path(product)
    except Exception:
        return None
    if fp.suffix.lower() != ".parquet":
        return None
    # 후보 목록은 stale-while-revalidate가 안전하다. stale 인덱스는 직전 ML_TABLE의
    # 완성 스냅샷이므로 새 lookup 빌드 중 빈 드롭다운만 보여주는 것보다 낫고,
    # 실제 SplitTable 데이터 조회는 여전히 별도의 fresh/stale 정책을 따른다.
    out = _ml_table_lookup.root_lot_candidates_from_lookup_cache(
        fp, prefix=prefix, limit=limit, allow_stale=True)
    out["source_fp"] = fp
    return out


def _product_ram_cache_available() -> bool:
    # 제품 전체 DataFrame RAM 상주는 검색 선행 계층(pivot)을 우회하지 못하면서
    # 메모리만 크게 점유하므로 자동/수동/env opt-in을 모두 폐기한다.
    return False


def _product_ram_cache_scheduler_enabled() -> bool:
    return False


def _product_ram_cache_refresh_minutes() -> int:
    raw = os.environ.get("FLOW_SPLITTABLE_PRODUCT_RAM_CACHE_REFRESH_MINUTES", "")
    if raw == "":
        raw = _match_cache_refresh_minutes()
    try:
        value = int(raw)
    except Exception:
        value = PRODUCT_RAM_CACHE_REFRESH_MINUTES_DEFAULT
    return max(PRODUCT_RAM_CACHE_REFRESH_MINUTES_MIN, min(PRODUCT_RAM_CACHE_REFRESH_MINUTES_MAX, value))


def _product_ram_cache_max_bytes() -> int:
    raw = os.environ.get("FLOW_SPLITTABLE_PRODUCT_RAM_CACHE_MAX_GB", "")
    if raw == "":
        # 톱니바퀴 설정(관리자 UI): 운영/개발 분리 — product_ram_gb(_dev)>0 이면 사용.
        try:
            from core import cache_settings
            _is_dev = _ml_table_lookup._root_ram_cache_use_dev()
            _gb_s = cache_settings.get_float_role("product_ram_gb", _is_dev)
            if _gb_s is not None and _gb_s > 0:
                raw = str(_gb_s)
        except Exception:
            pass
    try:
        gb = float(raw or PRODUCT_RAM_CACHE_MAX_GB_DEFAULT)
    except Exception:
        gb = PRODUCT_RAM_CACHE_MAX_GB_DEFAULT
    if gb <= 0:
        return 0
    budget = int(gb * 1024 * 1024 * 1024)
    try:
        from core import cache_budget
        # 명시값은 이 캐시의 희망 상한이며, 전체 캐시 안전 풀은 항상 적용한다.
        budget = cache_budget.capped("splittable_product_ram", budget)
    except Exception:
        pass
    return budget


def emergency_evict_view_cache(max_bytes: int) -> int:
    """메모리 워치독 긴급 축출 — 오래된 view payload 부터 최대 max_bytes 제거.

    다음 검색은 재계산으로 채워지므로 정확성 영향 없음. 반환: 회수 추정 바이트."""
    global _VIEW_CACHE_BYTES
    if max_bytes <= 0:
        return 0
    freed = 0
    with _VIEW_CACHE_LOCK:
        while _VIEW_CACHE and freed < max_bytes:
            _key, (_hard, _soft, _payload, approx) = _VIEW_CACHE.popitem(last=False)
            approx = int(approx or 0)
            freed += approx
            _VIEW_CACHE_BYTES = max(0, _VIEW_CACHE_BYTES - approx)
    return freed


def emergency_evict_product_ram(max_bytes: int) -> int:
    """메모리 워치독 긴급 축출 — 제품 RAM 캐시를 오래된 순으로 최대 max_bytes 제거.

    상태(_PRODUCT_RAM_CACHE_STATUS)는 남겨 스케줄러가 다음 주기에 재적재한다.
    반환: 회수 추정 바이트."""
    if max_bytes <= 0:
        return 0
    freed = 0
    with _PRODUCT_RAM_CACHE_LOCK:
        while _PRODUCT_RAM_CACHE and freed < max_bytes:
            product = next(iter(_PRODUCT_RAM_CACHE))
            entry = _PRODUCT_RAM_CACHE.pop(product, None) or {}
            try:
                freed += int(entry.get("estimated_bytes") or 0)
            except Exception:
                pass
    return freed


# ── 워치독이 못 보던 보조 캐시들 ────────────────────────────────────────────
# view payload / 제품 RAM / root RAM 세 tier 만 축출 대상이었다. 그런데 제품 RAM
# 은 `_product_ram_cache_available()` 이 False 로 굳어 항상 비어 있고, 검색 경로가
# 실제로 키우는 것은 아래 보조 캐시들이다 — 특히 `_CSV_ROWS_CACHE` 는 CSV 전체
# 행을 dict 로 32개까지 들고 있어 GB 단위가 된다. 축출이 `freed 0.0MB [nothing]`
# 만 반복하던 주된 이유가 이것이다.
_SCRATCH_CACHE_MAX_PROBE_NODES = 20000


def _approx_obj_bytes(obj, *, budget: int = _SCRATCH_CACHE_MAX_PROBE_NODES) -> int:
    """컨테이너를 제한된 노드 수까지만 훑는 근사 크기(bytes).

    정확할 필요는 없다 — 축출 tier 가 "얼마나 회수했는지" 를 과대평가하지만
    않으면 된다. budget 을 넘으면 그때까지 본 평균으로 나머지를 외삽한다.
    """
    import sys

    seen = 0
    total = 0

    def _walk(value) -> int:
        nonlocal seen, total
        if seen >= budget:
            return 0
        seen += 1
        try:
            size = sys.getsizeof(value)
        except Exception:
            return 64
        if isinstance(value, (str, bytes, bytearray, int, float, bool, type(None))):
            return size
        if isinstance(value, dict):
            n = len(value)
            if n and seen + n * 2 > budget:
                # 표본 몇 개로 항목당 평균을 잡고 외삽 — 큰 dict 를 통째로 안 훑는다.
                sample = 0
                taken = 0
                for k, v in value.items():
                    sample += _walk(k) + _walk(v)
                    taken += 1
                    if taken >= 16:
                        break
                return size + (sample // max(1, taken)) * n
            for k, v in value.items():
                size += _walk(k) + _walk(v)
            return size
        if isinstance(value, (list, tuple, set, frozenset)):
            n = len(value)
            if n and seen + n > budget:
                sample = 0
                taken = 0
                for item in value:
                    sample += _walk(item)
                    taken += 1
                    if taken >= 16:
                        break
                return size + (sample // max(1, taken)) * n
            for item in value:
                size += _walk(item)
            return size
        return size

    total = _walk(obj)
    return int(total)


def _scratch_cache_registry() -> list[tuple[str, dict, object]]:
    """(이름, 컨테이너, lock) — 축출해도 재계산으로 복구되는 보조 캐시들.

    순서는 회수 비용이 낮은 순이다. 전부 TTL/시그니처 기반이라 비워도 정확성에
    영향이 없고, 다음 조회가 다시 채운다.
    """
    return [
        ("csv_rows", _CSV_ROWS_CACHE, None),
        ("plan_root_index", _PLAN_ROOT_INDEX_CACHE, _PLAN_ROOT_INDEX_LOCK),
        ("ram_cache_lot_status", _RAM_CACHE_LOT_STATUS_CACHE, _RAM_CACHE_LOT_STATUS_LOCK),
        ("root_latest_step", _ROOT_LATEST_STEP_CACHE, _ROOT_LATEST_STEP_LOCK),
        ("step_order_ctx", _STEP_ORDER_CTX_CACHE, _STEP_ORDER_CTX_LOCK),
        ("latest_status_stats", _LATEST_STATUS_STATS_CACHE, _LATEST_STATUS_STATS_LOCK),
        ("lot_lookup", _LOT_LOOKUP_CACHE, None),
        ("view_product_sig", _VIEW_PRODUCT_SIG_CACHE, _VIEW_PRODUCT_SIG_LOCK),
        ("scan_schema", _SCAN_SCHEMA_CACHE, _SCAN_SCHEMA_CACHE_LOCK),
        ("schema_columns", _SCHEMA_COLUMNS_CACHE, None),
        ("plan_risk", _PLAN_RISK_CACHE, _PLAN_RISK_CACHE_LOCK),
        ("rglob", _RGLOB_CACHE, None),
        ("first_data_file", _FIRST_DATA_FILE_CACHE, None),
        ("db_roots", _DB_ROOTS_CACHE, None),
        ("latest_idx_fresh", _LATEST_IDX_FRESH_CACHE, _LATEST_IDX_FRESH_LOCK),
    ]


def emergency_evict_scratch_caches(max_bytes: int) -> int:
    """메모리 워치독 긴급 축출 — 검색 경로의 보조 캐시를 오래된 순으로 비운다.

    전부 재계산 가능한 파생 캐시다. 반환: 회수 추정 바이트."""
    if max_bytes <= 0:
        return 0
    freed = 0
    for _name, container, lock in _scratch_cache_registry():
        if freed >= max_bytes:
            break
        try:
            ctx = lock if lock is not None else contextlib.nullcontext()
            with ctx:
                # lock 없는 캐시는 다른 스레드가 동시에 넣고 뺀다 — 키 스냅샷을
                # 먼저 뜨고 pop 한다(iterator 중 크기 변경 예외 회피).
                for key in list(container.keys()):
                    if freed >= max_bytes:
                        break
                    value = container.pop(key, None)
                    if value is not None:
                        freed += _approx_obj_bytes(value)
        except Exception as exc:
            logger.warning("scratch cache evict failed (%s): %s", _name, exc)
    return freed


def scratch_cache_sizes() -> dict:
    """관리자 화면/진단용 — 보조 캐시가 지금 얼마나 들고 있는지."""
    out: dict = {}
    for name, container, _lock in _scratch_cache_registry():
        try:
            # 항목마다 새 budget 으로 잰다. 컨테이너를 통째로 넘기면 첫 항목이
            # budget 을 다 쓰고 나머지가 0 으로 잡혀 크게 과소평가된다.
            total = sum(_approx_obj_bytes(value) for value in list(container.values()))
            out[name] = {"entries": len(container), "bytes": int(total)}
        except Exception:
            out[name] = {"entries": 0, "bytes": 0}
    return out


def _product_ram_cache_products(product: str = "") -> list[str]:
    raw = str(product or "").strip()
    if raw:
        return [_canonical_mltable_product_name(raw, allow_bare=True) or raw]
    try:
        return [p.get("name") for p in list_products().get("products", []) if p.get("name")]
    except Exception:
        return []


def _product_ram_cache_source(product: str) -> tuple[str, Path]:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    return canonical, _product_path(canonical)


def _product_ram_cache_estimated_bytes(df: pl.DataFrame) -> int:
    try:
        return int(df.estimated_size())
    except Exception:
        try:
            return int(df.height) * max(1, len(df.columns)) * 16
        except Exception:
            return 0


def _product_ram_cache_source_memory_estimate(fp: Path) -> int:
    """collect 전에 계산하는 보수적 메모리 추정치.

    parquet 파일 크기는 압축 크기라 200MB 파일이 RAM에서 수 GB로 커질 수 있다.
    가능하면 row-group의 uncompressed byte 합을 쓰고 약간의 DataFrame overhead를
    더한다. metadata를 읽을 수 없을 때만 파일 크기로 폴백한다.
    """
    try:
        source_bytes = int(Path(fp).stat().st_size)
    except Exception:
        source_bytes = 0
    uncompressed = 0
    if Path(fp).suffix.lower() == ".parquet":
        try:
            import pyarrow.parquet as pq  # type: ignore

            meta = pq.ParquetFile(str(fp)).metadata
            for row_group_idx in range(meta.num_row_groups):
                row_group = meta.row_group(row_group_idx)
                for col_idx in range(row_group.num_columns):
                    uncompressed += int(row_group.column(col_idx).total_uncompressed_size or 0)
        except Exception:
            uncompressed = 0
    overhead = max(1.0, min(3.0, _env_float("FLOW_SPLITTABLE_PRODUCT_RAM_ESTIMATE_FACTOR", 1.25)))
    return int(max(source_bytes, uncompressed) * overhead)


def _product_ram_cache_total_bytes_locked(exclude_product: str = "") -> int:
    exclude = str(exclude_product or "").strip().upper()
    total = 0
    for product, entry in _PRODUCT_RAM_CACHE.items():
        if exclude and product.upper() == exclude:
            continue
        try:
            total += int(entry.get("estimated_bytes") or 0)
        except Exception:
            pass
    return total


def _read_product_ram_cache_frame(fp: Path) -> pl.DataFrame:
    if fp.suffix.lower() == ".csv":
        lf = pl.scan_csv(str(fp), infer_schema_length=5000)
    else:
        lf = _scan_parquet_compat(str(fp))
    return _cast_cats_lazy(lf).collect()


def _product_ram_cache_is_refreshing(product: str = "") -> bool:
    raw = str(product or "").strip()
    with _PRODUCT_RAM_CACHE_LOCK:
        if not raw:
            return bool(_PRODUCT_RAM_CACHE_REFRESHING)
        canonical = _canonical_mltable_product_name(raw, allow_bare=True) or raw
        return canonical in _PRODUCT_RAM_CACHE_REFRESHING or "*" in _PRODUCT_RAM_CACHE_REFRESHING


def _product_ram_cache_mark_status(product: str, updates: dict) -> None:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    if not canonical:
        return
    with _PRODUCT_RAM_CACHE_LOCK:
        cur = dict(_PRODUCT_RAM_CACHE_STATUS.get(canonical) or {})
        cur.update(updates)
        cur["product"] = canonical
        _PRODUCT_RAM_CACHE_STATUS[canonical] = cur


def _product_ram_cache_entry(product: str) -> dict | None:
    if not _product_ram_cache_available():
        return None
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    if not canonical:
        return None
    with _PRODUCT_RAM_CACHE_LOCK:
        return _PRODUCT_RAM_CACHE.get(canonical)


def _product_ram_cache_lazyframe(product: str):
    entry = _product_ram_cache_entry(product)
    if not entry:
        return None
    df = entry.get("df")
    if df is None:
        return None
    try:
        return _cast_cats_lazy(df.lazy())
    except Exception as exc:
        logger.debug("Product RAM cache lazyframe failed product=%s: %s", product, exc)
        return None


def _product_ram_cache_public_meta(product: str, *, include_detail: bool = False) -> dict:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    source_sig = ("", 0.0, 0)
    source_path = ""
    source_fp = None
    try:
        canonical, fp = _product_ram_cache_source(canonical)
        source_fp = fp
        source_path = str(fp)
        source_sig = _path_cache_sig(fp)
    except Exception:
        pass
    with _PRODUCT_RAM_CACHE_LOCK:
        entry = _PRODUCT_RAM_CACHE.get(canonical)
        status = dict(_PRODUCT_RAM_CACHE_STATUS.get(canonical) or {})
    stale = bool(entry and entry.get("source_sig") != source_sig)
    out = {
        "product": canonical,
        "enabled": _product_ram_cache_available(),
        "hit": bool(entry and _product_ram_cache_available()),
        "stale": stale,
        "refreshing": _product_ram_cache_is_refreshing(canonical),
        "loaded_at": entry.get("loaded_at", "") if entry else "",
        "row_count": int(entry.get("row_count") or 0) if entry else 0,
        "estimated_mb": round(float(entry.get("estimated_bytes") or 0) / (1024 * 1024), 3) if entry else 0.0,
        "source_mtime": source_sig[1] or None,
        "source_size": source_sig[2] or 0,
        "skipped": bool(status.get("skipped")) if not entry else False,
        "root_lot_cache": _ml_table_lookup.root_ram_cache_status(source_fp, include_detail=False),
    }
    if include_detail:
        out.update({
            "source_path": source_path,
            "source_sig": source_sig,
            "cache_source_sig": entry.get("source_sig") if entry else None,
            "error": status.get("error") or "",
            "skip_reason": status.get("skip_reason") or "",
            "last_refresh_at": status.get("last_refresh_at") or "",
            "root_lot_cache": _ml_table_lookup.root_ram_cache_status(source_fp, include_detail=True),
        })
    return out


def _product_ram_cache_response_meta(product: str) -> dict:
    meta = _product_ram_cache_public_meta(product, include_detail=False)
    return {
        "hit": meta.get("hit", False),
        "stale": meta.get("stale", False),
        "refreshing": meta.get("refreshing", False),
        "loaded_at": meta.get("loaded_at", ""),
        "row_count": meta.get("row_count", 0),
        "estimated_mb": meta.get("estimated_mb", 0.0),
        "source_mtime": meta.get("source_mtime"),
        "root_lot_cache": meta.get("root_lot_cache") or {},
    }


def _product_ram_cache_view_signature(product: str) -> tuple:
    if not _product_ram_cache_available():
        return ("product_ram_cache", "disabled")
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    with _PRODUCT_RAM_CACHE_LOCK:
        entry = _PRODUCT_RAM_CACHE.get(canonical)
        status = dict(_PRODUCT_RAM_CACHE_STATUS.get(canonical) or {})
    if entry:
        return (
            "product_ram_cache",
            canonical,
            entry.get("source_sig"),
            entry.get("loaded_epoch"),
            entry.get("estimated_bytes"),
        )
    return (
        "product_ram_cache",
        canonical,
        status.get("last_refresh_epoch"),
        status.get("skipped"),
        status.get("skip_reason"),
    )


def _refresh_product_ram_cache_products(products: list[str], force: bool = False) -> dict:
    products = [p for p in products if str(p or "").strip()]
    results: list[dict] = []
    if not _product_ram_cache_available():
        return {
            "ok": False,
            "products": [{"product": p, "ok": False, "skipped": True, "reason": "disabled"} for p in products],
            "interval_minutes": _product_ram_cache_refresh_minutes(),
        }
    max_bytes = _product_ram_cache_max_bytes()
    with _PRODUCT_RAM_CACHE_BUILD_LOCK:
        for raw_product in products:
            # 사용자 요청이 진행 중이면 다음 제품 로드를 미룬다 (백그라운드 양보).
            from core import request_priority
            request_priority.yield_to_users(max_wait_sec=15.0)
            canonical = _canonical_mltable_product_name(raw_product, allow_bare=True) or str(raw_product or "").strip()
            result = {"product": canonical, "ok": False, "skipped": False, "row_count": 0}
            try:
                canonical, fp = _product_ram_cache_source(canonical)
                result["product"] = canonical
                result["source_path"] = str(fp)
                source_sig = _path_cache_sig(fp)
                with _PRODUCT_RAM_CACHE_LOCK:
                    current = _PRODUCT_RAM_CACHE.get(canonical)
                    if current and not force and current.get("source_sig") == source_sig:
                        result.update({
                            "ok": True,
                            "skipped": True,
                            "reason": "fresh",
                            "row_count": int(current.get("row_count") or 0),
                            "estimated_mb": round(float(current.get("estimated_bytes") or 0) / (1024 * 1024), 3),
                        })
                        results.append(result)
                        continue
                    used_bytes = _product_ram_cache_total_bytes_locked(exclude_product=canonical)
                estimated_source_bytes = _product_ram_cache_source_memory_estimate(fp)
                remaining_budget = max(0, max_bytes - used_bytes) if max_bytes else 0
                if max_bytes and estimated_source_bytes > remaining_budget:
                    reason = "memory_budget_precheck"
                    _product_ram_cache_mark_status(canonical, {
                        "skipped": True,
                        "skip_reason": reason,
                        "error": "",
                        "last_refresh_at": datetime.datetime.now().isoformat(timespec="seconds"),
                        "last_refresh_epoch": time.time(),
                    })
                    result.update({
                        "skipped": True,
                        "reason": reason,
                        "estimated_mb": round(estimated_source_bytes / (1024 * 1024), 3),
                    })
                    results.append(result)
                    continue
                try:
                    from core.runtime_limits import process_memory_high, system_memory_snapshot
                    if process_memory_high():
                        reason = "process_memory_high"
                        _product_ram_cache_mark_status(canonical, {
                            "skipped": True,
                            "skip_reason": reason,
                            "error": "",
                            "last_refresh_at": datetime.datetime.now().isoformat(timespec="seconds"),
                            "last_refresh_epoch": time.time(),
                        })
                        result.update({"skipped": True, "reason": reason})
                        results.append(result)
                        continue
                    memory = system_memory_snapshot()
                    available_gb = float(memory.get("system_memory_available_gb") or 0.0)
                    reserve_gb = float(memory.get("system_memory_min_available_gb") or 2.0)
                    headroom_bytes = int(max(0.0, available_gb - reserve_gb) * (1024 ** 3))
                    if headroom_bytes > 0 and estimated_source_bytes > headroom_bytes:
                        reason = "host_memory_headroom_precheck"
                        _product_ram_cache_mark_status(canonical, {
                            "skipped": True,
                            "skip_reason": reason,
                            "error": "",
                            "last_refresh_at": datetime.datetime.now().isoformat(timespec="seconds"),
                            "last_refresh_epoch": time.time(),
                        })
                        result.update({
                            "skipped": True,
                            "reason": reason,
                            "estimated_mb": round(estimated_source_bytes / (1024 * 1024), 3),
                        })
                        results.append(result)
                        continue
                except Exception:
                    pass
                with _PRODUCT_RAM_CACHE_LOCK:
                    _PRODUCT_RAM_CACHE_REFRESHING.add(canonical)
                    # stale 확정(여기 도달 = fresh 아님) — 새 프레임을 읽기 전에
                    # 옛 프레임을 먼저 놓는다. 예전에는 old df + new df 가 로드
                    # 구간 동안 공존해 제품당 2배 스파이크가 났다 (예산 계산은
                    # exclude_product 로 old 를 빼고 보는데 실제로는 남아 있었음).
                    # 해제 후 로드 실패 시 다음 주기까지 캐시 미스 → 디스크 스캔
                    # 폴백이라 정확성 영향 없음.
                    _PRODUCT_RAM_CACHE.pop(canonical, None)
                try:
                    gc.collect()
                except Exception:
                    pass
                df = None
                try:
                    df = _read_product_ram_cache_frame(fp)
                    estimated_bytes = _product_ram_cache_estimated_bytes(df)
                    with _PRODUCT_RAM_CACHE_LOCK:
                        used_bytes = _product_ram_cache_total_bytes_locked(exclude_product=canonical)
                    if max_bytes and used_bytes + estimated_bytes > max_bytes:
                        reason = "memory_budget"
                        _product_ram_cache_mark_status(canonical, {
                            "skipped": True,
                            "skip_reason": reason,
                            "error": "",
                            "last_refresh_at": datetime.datetime.now().isoformat(timespec="seconds"),
                            "last_refresh_epoch": time.time(),
                        })
                        result.update({
                            "skipped": True,
                            "reason": reason,
                            "estimated_mb": round(estimated_bytes / (1024 * 1024), 3),
                        })
                        results.append(result)
                        continue
                    now = time.time()
                    loaded_at = datetime.datetime.fromtimestamp(now).isoformat(timespec="seconds")
                    entry = {
                        "version": PRODUCT_RAM_CACHE_VERSION,
                        "product": canonical,
                        "source_path": str(fp),
                        "source_sig": source_sig,
                        "df": df,
                        "row_count": int(df.height),
                        "estimated_bytes": int(estimated_bytes),
                        "loaded_at": loaded_at,
                        "loaded_epoch": now,
                        "error": "",
                        "skipped": False,
                    }
                    with _PRODUCT_RAM_CACHE_LOCK:
                        _PRODUCT_RAM_CACHE[canonical] = entry
                        _PRODUCT_RAM_CACHE_STATUS[canonical] = {
                            "product": canonical,
                            "skipped": False,
                            "skip_reason": "",
                            "error": "",
                            "last_refresh_at": loaded_at,
                            "last_refresh_epoch": now,
                        }
                    result.update({
                        "ok": True,
                        "row_count": int(df.height),
                        "estimated_mb": round(estimated_bytes / (1024 * 1024), 3),
                        "loaded_at": loaded_at,
                    })
                    df = None
                    _LOT_LOOKUP_CACHE.clear()
                    _invalidate_root_lot_pool()
                    _clear_split_view_cache_product(canonical)
                finally:
                    with _PRODUCT_RAM_CACHE_LOCK:
                        _PRODUCT_RAM_CACHE_REFRESHING.discard(canonical)
                    if df is not None:
                        try:
                            del df
                        except Exception:
                            pass
                results.append(result)
            except Exception as e:
                with _PRODUCT_RAM_CACHE_LOCK:
                    _PRODUCT_RAM_CACHE_REFRESHING.discard(canonical)
                reason = f"{type(e).__name__}: {e}"
                logger.warning("SplitTable product RAM cache refresh failed (product=%s) %s", canonical, reason, exc_info=True)
                _product_ram_cache_mark_status(canonical, {
                    "skipped": False,
                    "skip_reason": "",
                    "error": reason,
                    "last_refresh_at": datetime.datetime.now().isoformat(timespec="seconds"),
                    "last_refresh_epoch": time.time(),
                })
                result["reason"] = reason
                results.append(result)
            try:
                gc.collect()
            except Exception:
                pass
        # 제품 목록에서 빠진(삭제/개명) 항목은 갱신 대상이 다시 오지 않아 영구
        # 잔존한다 — 전체 갱신일 때만 정리 (단일 제품 호출이 남을 지우면 안 됨).
        if len(products) > 1:
            keep = {str(r.get("product") or "") for r in results}
            with _PRODUCT_RAM_CACHE_LOCK:
                for stale_key in [k for k in _PRODUCT_RAM_CACHE if k not in keep]:
                    _PRODUCT_RAM_CACHE.pop(stale_key, None)
                    _PRODUCT_RAM_CACHE_STATUS.pop(stale_key, None)
    return {
        "ok": any(r.get("ok") for r in results),
        "products": results,
        "interval_minutes": _product_ram_cache_refresh_minutes(),
        "max_gb": round(_product_ram_cache_max_bytes() / (1024 ** 3), 3) if _product_ram_cache_max_bytes() else 0,
    }


def refresh_product_ram_cache(product: str = "", force: bool = False) -> dict:
    return _refresh_product_ram_cache_products(_product_ram_cache_products(product), force=force)


def _product_ram_cache_job_status() -> dict:
    with _PRODUCT_RAM_CACHE_JOB_LOCK:
        out = dict(_PRODUCT_RAM_CACHE_JOB_STATE)
        out["products"] = [dict(r) for r in (_PRODUCT_RAM_CACHE_JOB_STATE.get("products") or [])]
        out["order"] = list(_PRODUCT_RAM_CACHE_JOB_STATE.get("order") or [])
    return out


def _product_ram_cache_job_update(**updates) -> None:
    with _PRODUCT_RAM_CACHE_JOB_LOCK:
        _PRODUCT_RAM_CACHE_JOB_STATE.update(updates)


def _begin_product_ram_cache_job(products: list[str], force: bool, reason: str) -> tuple[bool, dict]:
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _PRODUCT_RAM_CACHE_JOB_LOCK:
        if _PRODUCT_RAM_CACHE_JOB_STATE.get("running"):
            status = dict(_PRODUCT_RAM_CACHE_JOB_STATE)
            status["products"] = [dict(r) for r in (_PRODUCT_RAM_CACHE_JOB_STATE.get("products") or [])]
            status["order"] = list(_PRODUCT_RAM_CACHE_JOB_STATE.get("order") or [])
            return False, status
        _PRODUCT_RAM_CACHE_JOB_STATE.clear()
        _PRODUCT_RAM_CACHE_JOB_STATE.update({
            "running": True,
            "queued": True,
            "force": bool(force),
            "reason": reason or "manual",
            "started_at": now,
            "finished_at": "",
            "current_product": "",
            "total": len(products),
            "done": 0,
            "ok_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "last_error": "",
            "products": [],
            "order": list(products),
        })
        status = dict(_PRODUCT_RAM_CACHE_JOB_STATE)
        status["products"] = []
        status["order"] = list(products)
        return True, status


def _append_product_ram_cache_job_rows(rows: list[dict]) -> None:
    if not rows:
        return
    with _PRODUCT_RAM_CACHE_JOB_LOCK:
        current = [dict(r) for r in (_PRODUCT_RAM_CACHE_JOB_STATE.get("products") or [])]
        current.extend(dict(r) for r in rows)
        _PRODUCT_RAM_CACHE_JOB_STATE["products"] = current[-500:]
        _PRODUCT_RAM_CACHE_JOB_STATE["done"] = int(_PRODUCT_RAM_CACHE_JOB_STATE.get("done") or 0) + len(rows)
        _PRODUCT_RAM_CACHE_JOB_STATE["ok_count"] = int(_PRODUCT_RAM_CACHE_JOB_STATE.get("ok_count") or 0) + len([r for r in rows if r.get("ok")])
        _PRODUCT_RAM_CACHE_JOB_STATE["failed_count"] = int(_PRODUCT_RAM_CACHE_JOB_STATE.get("failed_count") or 0) + len([r for r in rows if not r.get("ok") and not r.get("skipped")])
        _PRODUCT_RAM_CACHE_JOB_STATE["skipped_count"] = int(_PRODUCT_RAM_CACHE_JOB_STATE.get("skipped_count") or 0) + len([r for r in rows if r.get("skipped")])
        for row in reversed(rows):
            if row.get("reason"):
                _PRODUCT_RAM_CACHE_JOB_STATE["last_error"] = str(row.get("reason") or "")
                break


def _run_started_product_ram_cache_job(products: list[str], force: bool, reason: str = "manual") -> dict:
    try:
        for product in products:
            if _PRODUCT_RAM_CACHE_STOP.is_set():
                break
            _product_ram_cache_job_update(current_product=product, queued=False)
            result = _refresh_product_ram_cache_products([product], force=force)
            _append_product_ram_cache_job_rows(result.get("products") or [])
    finally:
        _product_ram_cache_job_update(
            running=False,
            queued=False,
            current_product="",
            finished_at=datetime.datetime.now().isoformat(timespec="seconds"),
        )
    status = _product_ram_cache_job_status()
    return {
        "ok": bool(status.get("ok_count")),
        "queued": False,
        "products": status.get("products") or [],
        "interval_minutes": _product_ram_cache_refresh_minutes(),
        "job": status,
        "reason": reason,
    }


def enqueue_product_ram_cache_refresh(product: str = "", force: bool = True, reason: str = "manual") -> dict:
    """제품 원본 RAM 캐시 갱신을 서버 스캔 게이트에 넣고 즉시 반환한다.

    job state 는 큐에서 꺼내 실제로 시작할 때 잡는다 — 대기 중 running 표시로
    다른 스캔이 '이미 실행 중'으로 오판하는 것을 막는다."""
    products = _product_ram_cache_products(product)
    label = f"제품 원본 RAM 캐시 갱신 ({product or '전체 제품'})"

    def _start() -> dict:
        started, status = _begin_product_ram_cache_job(products, force=force, reason=reason)
        if not started:
            return {
                "ok": True,
                "skipped": True,
                "job": status,
                "detail": "SplitTable product RAM cache refresh is already running.",
            }
        return _run_started_product_ram_cache_job(products, force, reason)

    out = _submit_scan("product_ram", label, _start, product=product,
                       source="scheduler" if reason == "scheduler" else "manual",
                       dedupe_key=f"product_ram:{reason}:{product}")
    return {
        **out,
        "products": [{"product": p, "queued": True} for p in products],
        "interval_minutes": _product_ram_cache_refresh_minutes(),
        "job": _product_ram_cache_job_status(),
    }


def _product_ram_cache_loop() -> None:
    global _PRODUCT_RAM_CACHE_NEXT_TICK_AT
    while not _PRODUCT_RAM_CACHE_STOP.is_set():
        try:
            # 예약 갱신도 스캔 게이트를 통과한다 — 예전엔 job state 조차 잡지 않고
            # 바로 적재를 시작해서, 수동 스캔 2/3 단계와 그대로 겹칠 수 있었다.
            if _product_ram_cache_products(""):
                enqueue_product_ram_cache_refresh(product="", force=False, reason="scheduler")
        except Exception as e:
            logger.warning("SplitTable product RAM cache scheduler tick failed: %s", e)
        wait_s = max(60.0, _product_ram_cache_refresh_minutes() * 60.0)
        _PRODUCT_RAM_CACHE_NEXT_TICK_AT = datetime.datetime.fromtimestamp(
            time.time() + wait_s).isoformat(timespec="seconds")
        while wait_s > 0 and not _PRODUCT_RAM_CACHE_STOP.is_set():
            step = min(wait_s, 60.0)
            _PRODUCT_RAM_CACHE_STOP.wait(step)
            wait_s -= step


def start_product_ram_cache_scheduler() -> bool:
    global _PRODUCT_RAM_CACHE_THREAD, _PRODUCT_RAM_CACHE_STARTED
    if _PRODUCT_RAM_CACHE_STARTED:
        return False
    # Product rotation owns the complete persistent-cache sweep.  Keeping the
    # legacy product-RAM timer alive can insert an unrelated product between
    # two stages of the active product and makes the queue look cache-kind
    # driven again.
    if _auto_product_cache_enabled():
        logger.info("SplitTable product RAM timer retired; product rotation owns refresh")
        return False
    if not _product_ram_cache_scheduler_enabled():
        logger.info("SplitTable product RAM cache scheduler disabled")
        return False
    _PRODUCT_RAM_CACHE_STOP.clear()
    _PRODUCT_RAM_CACHE_THREAD = threading.Thread(
        target=_product_ram_cache_loop,
        name="splittable-product-ram-cache",
        daemon=True,
    )
    _PRODUCT_RAM_CACHE_THREAD.start()
    _PRODUCT_RAM_CACHE_STARTED = True
    logger.info("SplitTable product RAM cache scheduler started (interval=%sm)", _product_ram_cache_refresh_minutes())
    return True


# ── KNOB view payload 프리워밍 ────────────────────────────────────────────
# SplitTable 에서 압도적으로 많이 보는 것이 KNOB prefix 이고, 우선 lot 은 이미
# 관리자가 등록해 둔다. 그 조합(우선 lot × prefix=KNOB)의 view payload 를 미리
# 계산해 두면 가장 흔한 검색이 **cold 계산 자체를 건너뛰고** 캐시 HIT 로 끝난다.
# HIT 는 cold 레인에 줄서지 않으므로(self-gated 경로) 동시 사용자 대기도 함께 준다.
# 계산을 빠르게 하는 게 아니라 없애는 접근이라 다중 조회에서 효과가 가장 크다.
_KNOB_PREWARM_STOP = threading.Event()
_KNOB_PREWARM_THREAD = None
_KNOB_PREWARM_STARTED = False
_KNOB_PREWARM_LAST: dict = {"at": "", "warmed": 0, "skipped": 0, "failed": 0, "reason": ""}


def _knob_prewarm_enabled() -> bool:
    return str(os.environ.get("FLOW_SPLITTABLE_KNOB_PREWARM", "1")).strip().lower() not in {"0", "false", "no", "off"}


def _knob_prewarm_interval_sec() -> float:
    return max(60.0, min(24 * 3600.0, _env_float("FLOW_SPLITTABLE_KNOB_PREWARM_INTERVAL_SEC", 1800.0)))


def _knob_prewarm_max_lots() -> int:
    try:
        n = int(float(os.environ.get("FLOW_SPLITTABLE_KNOB_PREWARM_MAX_LOTS", "") or 50))
    except Exception:
        n = 50
    return max(1, min(500, n))


def _knob_prewarm_targets() -> list[tuple[str, str]]:
    """(product, root_lot_id) 목록 — 캐싱 활성화된 우선 lot 만, 상한 적용."""
    out: list[tuple[str, str]] = []
    try:
        data = _load_priority_lots_file()
    except Exception:
        return out
    limit = _knob_prewarm_max_lots()
    for product, lots in (data or {}).items():
        if not isinstance(lots, list):
            continue
        for entry in lots:
            if not isinstance(entry, dict):
                continue
            if entry.get("cache_enabled") is False:
                continue
            lot = str(entry.get("lot_id") or "").strip()
            if not lot:
                continue
            out.append((str(product or "").strip(), lot))
            if len(out) >= limit:
                return out
    return out


def _knob_prewarm_once() -> dict:
    warmed = skipped = failed = 0
    reason = ""
    targets = _knob_prewarm_targets()
    if not targets:
        reason = "우선 lot 없음"
    for product, root in targets:
        if _KNOB_PREWARM_STOP.is_set():
            reason = "중지 요청"
            break
        try:
            from core.runtime_limits import process_memory_high
            if process_memory_high():
                skipped += len(targets) - (warmed + skipped + failed)
                reason = "메모리 압박으로 중단"
                break
        except Exception:
            pass
        try:
            key = _split_view_cache_key(product, root, "", "KNOB", "", "all", "all", "", "")
            hard, soft = _split_view_cache_dep_signature(product)
            freshness, cached = _split_view_cache_get(key, hard, soft)
            if cached is not None and freshness == "fresh":
                skipped += 1
                continue
            params = {
                "product": product, "root_lot_id": root, "wafer_ids": "", "prefix": "KNOB",
                "custom_name": "", "view_mode": "all", "history_mode": "all",
                "fab_lot_id": "", "custom_cols": "",
            }
            result = _view_revalidate_execute(key, params) or {}
            hard, soft = _split_view_cache_dep_signature(product)
            _freshness, warmed_payload = _split_view_cache_get(key, hard, soft)
            if warmed_payload is not None:
                warmed += 1
            elif result.get("deferred") or result.get("queued"):
                skipped += 1
                reason = "개발 worker 큐에서 대기 중"
            else:
                failed += 1
        except Exception as exc:
            failed += 1
            logger.debug("KNOB prewarm 실패 (%s/%s): %s", product, root, exc)
    result = {
        "at": datetime.datetime.now().isoformat(timespec="seconds"),
        "warmed": warmed, "skipped": skipped, "failed": failed,
        "targets": len(targets), "reason": reason,
    }
    _KNOB_PREWARM_LAST.update(result)
    if warmed or failed:
        try:
            from core import cache_event_log
            cache_event_log.record(
                "cache_op", "knob_prewarm", ok=(failed == 0),
                detail={"warmed": warmed, "skipped": skipped, "failed": failed,
                        "targets": len(targets), "reason": reason})
        except Exception:
            pass
    return result


def _knob_prewarm_loop() -> None:
    # 기동 직후는 캐시/스케줄러가 자리를 잡을 시간을 준다.
    if _KNOB_PREWARM_STOP.wait(_env_float("FLOW_SPLITTABLE_KNOB_PREWARM_START_DELAY_SEC", 180.0)):
        return
    while not _KNOB_PREWARM_STOP.is_set():
        try:
            _knob_prewarm_once()
        except Exception as exc:
            logger.warning("KNOB prewarm loop 오류: %s", exc)
        if _KNOB_PREWARM_STOP.wait(_knob_prewarm_interval_sec()):
            return


def start_knob_prewarmer() -> bool:
    global _KNOB_PREWARM_THREAD, _KNOB_PREWARM_STARTED
    if _KNOB_PREWARM_STARTED:
        return False
    if not _knob_prewarm_enabled():
        logger.info("SplitTable KNOB prewarmer disabled (FLOW_SPLITTABLE_KNOB_PREWARM=0)")
        return False
    _KNOB_PREWARM_STOP.clear()
    _KNOB_PREWARM_THREAD = threading.Thread(
        target=_knob_prewarm_loop, name="splittable-knob-prewarm", daemon=True)
    _KNOB_PREWARM_THREAD.start()
    _KNOB_PREWARM_STARTED = True
    logger.info("SplitTable KNOB prewarmer started (interval=%.0fs, max %d lots)",
                _knob_prewarm_interval_sec(), _knob_prewarm_max_lots())
    return True


def _strip_non_authoritative_fab_fields(lf, product: str):
    """Hide FAB-only identifiers from ML tables unless they came from FAB source.

    `fab_lot_id` is an operational FAB identifier. If FAB override/source is off or
    failed, SplitTable should not surface a stale ML-side copy because users assume
    it came from live/real FAB lineage.  Do not synthesize it from ML_TABLE LOT_ID.
    """
    if not product or not str(product).casefold().startswith("ml_table_"):
        return lf
    try:
        names = lf.collect_schema().names()
    except Exception:
        return lf
    drop_cols = [n for n in names if n.casefold() == "fab_lot_id"]
    return lf.drop(drop_cols) if drop_cols else lf


def _view_identity_cols(schema_names, lot_col: str, wf_col: str, fab_lot_col: str = "") -> set[str]:
    """행이 아니라 헤더로 쓰이는 식별 컬럼(root lot / wafer / fab lot).

    fab lot 컬럼은 wafer 헤더 그룹으로 이미 나가는데 데이터 컬럼 목록에도 남아 있어서
    prefix=FAB 조회에 `fab_lot_id` 행이 섞여 나왔다. 헤더에서 keep 하는 fab 컬럼과
    같은 규칙으로 걸러 두 곳이 어긋나지 않게 한다.
    """
    names = list(schema_names or [])
    out = {c for c in (lot_col, wf_col) if c}
    fab_ident = "fab_lot_id" if "fab_lot_id" in names else (
        _ci_resolve_in(fab_lot_col, names)
        or _pick_first_present_ci(_FAB_COL_CANDIDATES, names)
        or ""
    )
    if fab_ident:
        out.add(fab_ident)
    return out


def _view_data_columns(schema_names, lot_col: str, wf_col: str, fab_lot_col: str = "") -> list:
    """SplitTable 행/컬럼 선택기의 후보 = 식별 컬럼을 뺀 나머지."""
    identity = _view_identity_cols(schema_names, lot_col, wf_col, fab_lot_col)
    return [c for c in (schema_names or []) if c not in identity]


def _known_column_prefix_tokens() -> set[str]:
    """등록된 prefix(톱니바퀴) + 내장 overlay prefix 집합. 모두 대문자."""
    out = set(DEFAULT_PREFIXES)
    out.update(getattr(_ml_table_lookup, "CANDIDATE_COLUMN_PREFIXES", ()))
    out.add(CUSTOM_TAG_PREFIX)
    out.add(MANAGEMENT_ROW_PREFIX)
    try:
        out.update(_load_prefixes())
    except Exception:
        pass
    return {p for p in out if p}


def _has_known_prefixed_columns(all_data_cols) -> bool:
    """원천 컬럼이 prefix 체계(KNOB_/VM_/ET_ ...)를 쓰고 있는지."""
    tokens = _known_column_prefix_tokens()
    if not tokens:
        return False
    for col in all_data_cols or []:
        head = str(col).split("_", 1)
        if len(head) == 2 and head[0].upper() in tokens:
            return True
    return False


def _select_columns(all_data_cols, custom_name: str, prefix: str, max_fallback: int = 50,
                    custom_cols: str = ""):
    """Multi-prefix ("KNOB,MASK") or ALL or custom-name/custom-cols based column selection.

    v8.8.16: CUSTOM 모드는 사용자가 저장한 columns 를 **그대로** 반환한다.
      - 기존: `all_data_cols` 에 없으면 걸러내어 → 값이 null 인 컬럼이 LOT 뷰에서 사라지는 문제.
      - 변경: custom 에 저장된 column 명을 있는 그대로 반환. view_split 이 null row 를
              자연스럽게 생성 (컬럼이 실제 df 에 없으면 모든 셀이 None, 컬럼명은 유지).
      - 빈 리스트면 기존 폴백 (상위 max_fallback) 유지.
    v8.8.33: `custom_cols` 쉼표 구분 문자열 지원 — 저장된 set 없이도 체크만 한 컬럼을 전송해
             즉시 view 에 반영. custom_name 보다 우선 (ad-hoc 입력 우선).
    v10.0.8: prefix 를 지정했는데 매칭 컬럼이 하나도 없으면 예전에는 상위 N개를 그대로
             돌려줘서 **다른 prefix 컬럼이 섞여 나왔다** (ET 선택인데 KNOB/VM 행이 보임).
             원천이 prefix 체계를 쓰고 있으면 빈 결과가 정답이므로 폴백하지 않는다.
             prefix 가 아예 없는(=bare 컬럼) 제품에서만 기존 상위 N개 폴백을 유지한다.
    """
    # ad-hoc custom_cols 우선
    if custom_cols:
        return _clean_custom_columns(custom_cols.split(","))
    if custom_name:
        try:
            cfp, _clean_name = _custom_file_path_for_name(custom_name)
        except HTTPException:
            return []
        data = load_json(cfp, {})
        cleaned = _sanitize_custom_record(data, cfp, persist=True)
        if cleaned:
            return list(cleaned.get("columns") or [])
        return all_data_cols[:max_fallback]
    if prefix.upper() == "ALL":
        return all_data_cols[: max_fallback * 4]
    pref_list = [p.strip().upper() + "_" for p in prefix.split(",") if p.strip()]
    if pref_list:
        sel = [c for c in all_data_cols if any(c.upper().startswith(p) for p in pref_list)]
        if sel:
            return sel
        if _has_known_prefixed_columns(all_data_cols):
            return []
    return all_data_cols[:max_fallback]


# ── Custom tag columns: runtime-only SplitTable overlay ───────────────
def _custom_tags_path() -> Path:
    return PLAN_DIR / "custom_tags.json"


def _load_custom_tags_data() -> dict:
    # cached — _clean_overlay_store_data 는 입력을 수정하지 않고 새 컨테이너를 만든다.
    data = load_json_cached(_custom_tags_path(), {"columns": [], "values": {}})
    cleaned, changed = _clean_overlay_store_data(data, allow_management=True)
    if changed:
        _save_custom_tags_data(cleaned)
    return cleaned


def _save_custom_tags_data(data: dict) -> None:
    save_json(_custom_tags_path(), {
        "columns": list(data.get("columns") or []),
        "values": dict(data.get("values") or {}),
    }, indent=2)


def _tag_column_id(name: str) -> str:
    raw = str(name or "").strip()
    if raw.upper().startswith(f"{CUSTOM_TAG_PREFIX}_"):
        raw = raw[len(CUSTOM_TAG_PREFIX) + 1:].strip()
    token = "".join(c for c in raw if c.isalnum() or c in "_-. ")[:72].strip().replace(" ", "_")
    token = "_".join(part for part in token.split("_") if part)
    if not token:
        raise HTTPException(400, "tag name required")
    return f"{CUSTOM_TAG_PREFIX}_{token}"


def _tag_value_key(product: str, root_lot_id: str, wafer_id: str, column: str) -> str:
    return "|".join([str(product or ""), str(root_lot_id or ""), str(wafer_id or ""), str(column or "")])


# TAG 행의 module 은 엔지니어가 직접 적는 자유 텍스트다. Vehicle_matching 의 module
# 열과 달리 원천이 없으므로 빈 값이 정상이며, 빈 값이면 화면에서도 그냥 비워 둔다.
def _clean_tag_module(value: Any) -> str:
    return str(value or "").strip()[:48]


def _ensure_custom_tag_column(
    data: dict, *, product: str, column: str, label: str, actor: str, now: str, module: Any = None
) -> dict:
    cols = data.setdefault("columns", [])
    product_key = str(product or "").strip()
    column_key = str(column or "").strip()
    existing = next((c for c in cols if c.get("product") == product_key and c.get("column") == column_key), None)
    if existing:
        existing["label"] = str(label or existing.get("label") or column_key).strip() or column_key
        existing["username"] = actor or existing.get("username", "")
        existing["updated"] = now
        # module 은 명시적으로 넘어왔을 때만 건드린다 — 값 저장 경로에서 지워지면 안 된다.
        if module is not None:
            existing["module"] = _clean_tag_module(module)
        return existing
    entry = {
        "product": product_key,
        "column": column_key,
        "label": str(label or column_key).strip() or column_key,
        "module": _clean_tag_module(module),
        "username": actor,
        "created": now,
        "updated": now,
    }
    cols.append(entry)
    return entry


def _custom_tag_columns_for_product(product: str) -> list[dict]:
    product_key = str(product or "").strip()
    data = _load_custom_tags_data()
    out = []
    seen = set()
    for raw in data.get("columns") or []:
        if not isinstance(raw, dict):
            continue
        if raw.get("product") != product_key:
            continue
        column = str(raw.get("column") or "").strip()
        if not column or column in seen:
            continue
        seen.add(column)
        label = str(raw.get("label") or column).strip() or column
        out.append({**raw, "column": column, "label": label, "module": _clean_tag_module(raw.get("module"))})
    return out


def _custom_tag_label_map(product: str) -> dict[str, str]:
    return {c["column"]: c.get("label") or c["column"] for c in _custom_tag_columns_for_product(product)}


def _custom_tag_values_for_root(product: str, root_lot_id: str) -> dict[str, str]:
    data = _load_custom_tags_data()
    prefix = f"{product}|{root_lot_id}|"
    out: dict[str, str] = {}
    for key, raw in (data.get("values") or {}).items():
        if not str(key).startswith(prefix):
            continue
        parts = str(key).split("|", 3)
        if len(parts) != 4:
            continue
        value = raw.get("value") if isinstance(raw, dict) else raw
        if value is None:
            continue
        out["|".join(parts[1:])] = str(value)
    return out


def _custom_tag_column_values(product: str, column: str, limit: int = 200) -> list[str]:
    data = _load_custom_tags_data()
    out: list[str] = []
    seen: set[str] = set()
    suffix = f"|{column}"
    prefix = f"{product}|"
    for key, raw in (data.get("values") or {}).items():
        if not str(key).startswith(prefix) or not str(key).endswith(suffix):
            continue
        value = raw.get("value") if isinstance(raw, dict) else raw
        if value is None:
            continue
        s = str(value).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
        if len(out) >= limit:
            break
    return out


# ── Management rows: runtime-only SplitTable row overlay ──────────────
def _management_rows_path() -> Path:
    return PLAN_DIR / "management_rows.json"


def _load_management_rows_data() -> dict:
    # cached — 위 _load_custom_tags_data 와 동일한 이유.
    data = load_json_cached(_management_rows_path(), {"columns": [], "values": {}})
    cleaned, changed = _clean_overlay_store_data(data, allow_management=True)
    if changed:
        _save_management_rows_data(cleaned)
    return cleaned


def _save_management_rows_data(data: dict) -> None:
    save_json(_management_rows_path(), {
        "columns": list(data.get("columns") or []),
        "values": dict(data.get("values") or {}),
    }, indent=2)


def _management_row_id(name: str) -> str:
    raw = str(name or "").strip()
    if raw.upper().startswith(f"{MANAGEMENT_ROW_PREFIX}_"):
        raw = raw[len(MANAGEMENT_ROW_PREFIX) + 1:].strip()
    token = safe_id(raw, max_len=72).strip().replace(" ", "_")
    token = "_".join(part for part in token.split("_") if part)
    if not token:
        raise HTTPException(400, "management row name required")
    return f"{MANAGEMENT_ROW_PREFIX}_{token}"


def _management_row_value_key(product: str, root_lot_id: str, wafer_id: str, column: str) -> str:
    return "|".join([str(product or ""), str(root_lot_id or ""), str(wafer_id or ""), str(column or "")])


def _ensure_management_row_column(data: dict, *, product: str, column: str, label: str, actor: str, now: str) -> dict:
    cols = data.setdefault("columns", [])
    product_key = str(product or "").strip()
    column_key = str(column or "").strip()
    existing = next((c for c in cols if c.get("product") == product_key and c.get("column") == column_key), None)
    if existing:
        existing["label"] = str(label or existing.get("label") or column_key).strip() or column_key
        existing["username"] = actor or existing.get("username", "")
        existing["updated"] = now
        return existing
    entry = {
        "product": product_key,
        "column": column_key,
        "label": str(label or column_key).strip() or column_key,
        "username": actor,
        "created": now,
        "updated": now,
    }
    cols.append(entry)
    return entry


def _management_row_columns_for_product(product: str) -> list[dict]:
    product_key = str(product or "").strip()
    data = _load_management_rows_data()
    out = []
    seen = set()
    for raw in data.get("columns") or []:
        if not isinstance(raw, dict):
            continue
        if raw.get("product") != product_key:
            continue
        column = str(raw.get("column") or "").strip()
        if not column or column in seen:
            continue
        seen.add(column)
        label = str(raw.get("label") or column).strip() or column
        out.append({**raw, "column": column, "label": label})
    return out


def _management_row_label_map(product: str) -> dict[str, str]:
    return {c["column"]: c.get("label") or c["column"] for c in _management_row_columns_for_product(product)}


def _management_row_values_for_root(product: str, root_lot_id: str) -> dict[str, str]:
    data = _load_management_rows_data()
    prefix = f"{product}|{root_lot_id}|"
    out: dict[str, str] = {}
    for key, raw in (data.get("values") or {}).items():
        if not str(key).startswith(prefix):
            continue
        parts = str(key).split("|", 3)
        if len(parts) != 4:
            continue
        value = raw.get("value") if isinstance(raw, dict) else raw
        if value is None:
            continue
        out["|".join(parts[1:])] = str(value)
    return out


def _management_row_column_values(product: str, column: str, limit: int = 200) -> list[str]:
    data = _load_management_rows_data()
    out: list[str] = []
    seen: set[str] = set()
    suffix = f"|{column}"
    prefix = f"{product}|"
    for key, raw in (data.get("values") or {}).items():
        if not str(key).startswith(prefix) or not str(key).endswith(suffix):
            continue
        value = raw.get("value") if isinstance(raw, dict) else raw
        if value is None:
            continue
        s = str(value).strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
        if len(out) >= limit:
            break
    return out


# ── Notes (v8.4.9-b): 검색된 wafer 태그 + 파라미터 메모 ───────────────
# 스키마: {data_root}/splittable/notes.json
#   { "entries": [
#       { "id": "n_xxxxxx",
#         "scope": "wafer" | "param",
#         "key":  "{product}__{root_lot_id}__W{wafer_id}"
#               | "{product}__{root_lot_id}__W{wafer_id}__{param_name}",
#         "text": "...",
#         "username": "hol",
#         "created_at": "2026-04-21T10:00:00" }
#     ] }
# 작성자 또는 admin 만 삭제 가능. 수정은 지원하지 않음 (메모 히스토리 유지).
def _load_notes() -> list:
    data = load_json(NOTES_FILE, {"entries": []})
    if isinstance(data, dict):
        return data.get("entries", [])
    return data if isinstance(data, list) else []


def _save_notes(entries: list) -> None:
    save_json(NOTES_FILE, {"entries": entries})


def _new_note_id() -> str:
    import secrets as _secrets
    return "n_" + _secrets.token_hex(5)


def _notes_key_wafer(product: str, root_lot_id: str, wafer_id) -> str:
    return f"{product}__{root_lot_id}__W{wafer_id}"


def _notes_key_param(product: str, root_lot_id: str, wafer_id, param: str) -> str:
    return f"{product}__{root_lot_id}__W{wafer_id}__{param}"


def _notes_key_lot(product: str, root_lot_id: str) -> str:
    """v8.7.8: LOT 단위 노트 (해당 root_lot_id 전역). param 태그와 달리 lot 에 묶임."""
    return f"{product}__LOT__{root_lot_id}"


def _notes_key_param_global(product: str, param: str) -> str:
    """v8.7.8: parameter 전역 태그 — product 내 모든 LOT 에서 동일 parameter 에 노출."""
    return f"{product}__PARAM__{param}"


def _notes_lot_prefix(product: str, root_lot_id: str) -> str:
    return f"{product}__{root_lot_id}__"


def _notes_product_param_prefix(product: str) -> str:
    return f"{product}__PARAM__"


def _notes_product_lot_prefix(product: str) -> str:
    return f"{product}__LOT__"


class NoteSaveReq(BaseModel):
    scope: str                 # "wafer" | "param" | "lot" | "param_global"
    product: str = ""
    root_lot_id: str = ""
    wafer_id: str = ""
    param: str = ""            # scope == "param" / "param_global" 일 때
    text: str
    username: str = ""
    images: list[dict] = Field(default_factory=list)


class NoteDeleteReq(BaseModel):
    id: str
    username: str = ""


class NoteCommentReq(BaseModel):
    note_id: str
    text: str = ""
    username: str = ""
    images: list[dict] = Field(default_factory=list)


def _clean_note_text(text: str) -> str:
    return (text or "").replace("\u200b", "").strip()


def _normalize_note_image(raw) -> dict | None:
    if not isinstance(raw, dict):
        return None
    url = (
        raw.get("url")
        or raw.get("downloadUrl")
        or raw.get("fileUrl")
        or ((raw.get("attachment") or {}).get("downloadUrl") if isinstance(raw.get("attachment"), dict) else "")
        or ((raw.get("file") or {}).get("fileUrl") if isinstance(raw.get("file"), dict) else "")
    )
    url = str(url or "").strip().split("?", 1)[0]
    if url.startswith("api/informs/files/"):
        url = "/" + url
    elif url.startswith("files/"):
        url = "/api/informs/" + url
    if not url.startswith("/api/informs/files/"):
        return None
    filename = (
        raw.get("filename")
        or raw.get("name")
        or raw.get("displayName")
        or Path(url).name
        or "image"
    )
    try:
        size = int(raw.get("size") or raw.get("bytes") or 0)
    except Exception:
        size = 0
    return {"filename": Path(str(filename)).name or "image", "url": url, "size": max(0, size)}


def _normalize_note_images(images) -> list[dict]:
    out = []
    seen = set()
    for raw in images or []:
        item = _normalize_note_image(raw)
        if not item:
            continue
        key = item["url"]
        if key in seen:
            continue
        seen.add(key)
        out.append(item)
    return out[:12]


def _normalize_note_entry(entry: dict) -> dict:
    e = dict(entry or {})
    e["text"] = _clean_note_text(str(e.get("text") or ""))
    e["images"] = _normalize_note_images(e.get("images") or [])
    comments = []
    for raw in e.get("comments") or []:
        if not isinstance(raw, dict):
            continue
        c = dict(raw)
        c["text"] = _clean_note_text(str(c.get("text") or ""))
        c["images"] = _normalize_note_images(c.get("images") or [])
        comments.append(c)
    e["comments"] = comments
    return e


def _note_scope_parts(entry: dict) -> tuple[str, str, str]:
    key = str(entry.get("key") or "")
    scope = entry.get("scope")
    parts = key.split("__")
    if scope == "lot" and len(parts) >= 3:
        return parts[0], parts[2], ""
    if scope in ("wafer", "param") and len(parts) >= 3:
        return parts[0], parts[1], str(parts[2]).replace("W", "", 1)
    return "", "", ""


def _append_splittable_note_knowledge(entry: dict, *, actor: str, text: str) -> None:
    try:
        from core import knowledge_impact
        product, root_lot_id, wafer_id = _note_scope_parts(entry)
        param = ""
        key = str(entry.get("key") or "")
        parts = key.split("__")
        if entry.get("scope") in {"param", "param_global"}:
            param = parts[-1] if parts else ""
        knowledge_impact.append_candidates_from_text(
            text,
            source_type="split_note",
            source_id=entry.get("id") or "",
            actor=actor,
            context={
                "product": product,
                "root_lot_id": root_lot_id,
                "wafer_id": wafer_id,
                "item_id": param,
                "knob_name": param if str(param).upper().startswith(("KNOB_", "MASK_")) else "",
                "source_refs": [{"type": "split_note", "id": entry.get("id") or "", "label": param or root_lot_id}],
            },
            allowed_event_types={"split_impact"},
            status="candidate",
            title_prefix="SplitTable",
        )
    except Exception:
        return


def _append_splittable_plan_knowledge(*, product: str, cell_key: str, old: Any, new: Any, actor: str, changed_at: str, conflicting: bool = False) -> None:
    try:
        from core import knowledge_impact
        parts = str(cell_key or "").split("|")
        root = parts[0] if len(parts) > 0 else ""
        wafer = parts[1] if len(parts) > 1 else ""
        col = parts[2] if len(parts) > 2 else ""
        if not col:
            return
        knowledge_impact.safe_append_domain_event(
            event_type="split_impact",
            source_type="split_note",
            source_id=f"{product}:{cell_key}",
            title="SplitTable plan impact candidate",
            summary=f"SplitTable plan changed {product} {cell_key}: {old} -> {new}",
            actor=actor,
            payload={
                "product": product,
                "root_lot_id": root,
                "wafer_id": wafer,
                "item_id": col,
                "knob_name": col if str(col).upper().startswith(("KNOB_", "MASK_")) else "",
                "split_value": "" if new is None else str(new),
                "previous_split_value": "" if old is None else str(old),
                "effect_direction": "unknown",
                "effect_confidence": "candidate",
                "status": "candidate",
                "changed_at": changed_at,
                "conflicting_evidence": bool(conflicting),
                "source_refs": [{"type": "split_plan", "id": cell_key, "label": f"{old} -> {new}"}],
            },
        )
    except Exception:
        return


def _split_plan_cell_key(cell_key: str) -> tuple[str, str, str]:
    parts = str(cell_key or "").split("|", 2)
    root = parts[0] if len(parts) > 0 else ""
    wafer = parts[1] if len(parts) > 1 else ""
    column = parts[2] if len(parts) > 2 else ""
    return root, wafer, column


def _plan_actual_mismatch(plan: Any, actual: Any) -> bool:
    plan_text = _clean_str(plan)
    actual_text = _clean_str(actual)
    return bool(plan_text and actual_text and plan_text != actual_text)


def _plan_mismatch_alert_key(cell_key: str, plan: Any, actual: Any) -> str:
    raw = json.dumps(
        {"cell": str(cell_key or ""), "plan": _clean_str(plan), "actual": _clean_str(actual)},
        ensure_ascii=False,
        sort_keys=True,
        default=str,
    )
    return hashlib.sha1(raw.encode("utf-8")).hexdigest()[:24]


def _actual_value_for_plan_cell(product: str, cell_key: str) -> str:
    root, wafer, column = _split_plan_cell_key(cell_key)
    if not root or not wafer or not column:
        return ""
    try:
        lf = _scan_product(product, root_lot_id=root, wafer_ids=wafer)
        lot_col, wf_col = _detect_lot_wafer(lf, product)
        lf = _filter_lot_wafer(lf, lot_col, wf_col, root, wafer)
        names = lf.collect_schema().names()
        actual_col = column if column in names else (_ci_resolve_in(column, names) or "")
        if not actual_col:
            return ""
        df = (
            lf.select(pl.col(actual_col).cast(_STR, strict=False).alias("actual"))
            .drop_nulls()
            .head(1)
            .collect()
        )
        if df.height == 0:
            return ""
        return _clean_str(df.item(0, 0))
    except Exception:
        return ""


def _product_mismatch_group_members(product: str) -> list[str]:
    """제품명과 이름이 같은(대소문자 무시) 관리자 그룹의 멤버 목록.

    SplitTable 제품은 ML_TABLE_<PROD> 형태라 prefix 를 뗀 이름도 같이 매칭한다.
    """
    raw = str(product or "").strip()
    if not raw:
        return []
    names = {raw.lower()}
    if raw.upper().startswith("ML_TABLE_"):
        tail = raw[len("ML_TABLE_"):].strip()
        if tail:
            names.add(tail.lower())
    try:
        from routers.groups import _load as _load_groups
        members: list[str] = []
        for g in _load_groups():
            if not isinstance(g, dict):
                continue
            if str(g.get("name") or "").strip().lower() not in names:
                continue
            for m in g.get("members") or []:
                username = str(m or "").strip()
                if username and username not in members:
                    members.append(username)
        return members
    except Exception:
        return []


def _send_plan_mismatch_mail(product: str, items: list[dict], targets: list[str], actor: str = "flow") -> None:
    """plan/actual 불일치 확인 요청 메일 — 비밀번호 찾기 메일과 같은 방식(from_addr 발신, HTML 본문)."""
    if not items or not targets:
        return
    try:
        import html as _html
        from core.mail import load_mail_cfg, send_mail
        try:
            sender = (load_mail_cfg().get("from_addr") or "").strip()
        except Exception:
            sender = ""
        rows = []
        for it in items[:50]:
            where = _html.escape(str(it.get("root") or ""))
            wafer = str(it.get("wafer") or "").strip()
            if wafer:
                where += f" WF{_html.escape(wafer)}"
            when = str(it.get("updated") or "").strip()[:16].replace("T", " ")
            rows.append(
                "<li style='margin:2px 0'>"
                f"<b>{where}</b> · {_html.escape(str(it.get('column') or ''))} — "
                f"<b>{_html.escape(str(it.get('owner') or ''))}</b> 님이 plan "
                f"<b>{_html.escape(str(it.get('plan') or ''))}</b> 적용"
                + (f" ({_html.escape(when)})" if when else "")
                + f" → 실제로는 <b>{_html.escape(str(it.get('actual') or ''))}</b> 로 진행됨"
                "</li>"
            )
        content = (
            "<div style='font-family:Arial,sans-serif;font-size:14px;line-height:1.6'>"
            f"<p><b>{_html.escape(str(product))}</b> SplitTable 에서 계획(plan)과 실제 진행이 다른 항목이 확인되었습니다.</p>"
            f"<ul style='padding-left:18px'>{''.join(rows)}</ul>"
            "<p>실제 진행 내용이 맞는지 확인해 주세요.</p>"
            "<p style='color:#666;font-size:12px'>본 메일은 동일 불일치 건에 대해 1회만 발송됩니다.</p>"
            "</div>"
        )
        send_mail(
            sender_username=sender or (actor or "flow"),
            receiver_usernames=list(targets),
            extra_emails=[],
            title=f"[flow] plan/actual 불일치 확인 요청 - {product}",
            content=content,
        )
    except Exception:
        logger.debug("mismatch mail send failed product=%s", product, exc_info=True)


def _notify_plan_actual_mismatches_once(product: str, mismatches: list[dict], actor: str = "flow") -> int:
    if not mismatches:
        return 0
    try:
        from core.notify import emit_event
        data = _load_plan_data(product)
        plans = data.get("plans") if isinstance(data.get("plans"), dict) else {}
        alerts = data.get("mismatch_alerts") if isinstance(data.get("mismatch_alerts"), dict) else {}
        mail_enabled = False
        try:
            _cfg = load_json(SOURCE_CFG, {}) or {}
            mail_enabled = bool(_cfg.get("mismatch_mail_enabled"))
        except Exception:
            mail_enabled = False
        # 수신 대상 = 계획 작성자 + 제품명과 동일한 이름의 그룹 멤버.
        group_recipients = _product_mismatch_group_members(product)
        sent = 0
        mail_items: list[dict] = []
        mail_targets: list[str] = []
        for mm in mismatches[:100]:
            cell_key = str(mm.get("key") or mm.get("cell") or "")
            if not cell_key:
                continue
            plan = mm.get("plan")
            actual = mm.get("actual")
            if not _plan_actual_mismatch(plan, actual):
                continue
            plan_info = plans.get(cell_key) if isinstance(plans.get(cell_key), dict) else {}
            owner = str(mm.get("plan_user") or plan_info.get("user") or "").strip()
            targets: list[str] = []
            if owner:
                targets.append(owner)
            for name in group_recipients:
                if name not in targets:
                    targets.append(name)
            if not targets:
                continue
            alert_key = _plan_mismatch_alert_key(cell_key, plan, actual)
            root, wafer, column = _split_plan_cell_key(cell_key)
            payload = {
                "product": product,
                "cell": cell_key,
                "root_lot_id": root,
                "wafer_id": wafer,
                "column": column,
                "plan": _clean_str(plan),
                "actual": _clean_str(actual),
                "plan_updated": plan_info.get("updated") or mm.get("plan_updated") or "",
            }
            alert_body = (
                f"! {product}/{root}"
                + (f" WF{wafer}" if wafer else "")
                + f" {column}: [plan] {payload['plan']} → [actual] {payload['actual']}"
            )
            new_targets: list[str] = []
            for target in targets:
                # 작성자는 기존 key 형식 유지(중복 재알람 방지), 팀 수신자는 사용자별 key.
                target_alert_key = alert_key if target == owner else f"{alert_key}|u:{target}"
                if target_alert_key in alerts:
                    continue
                ok = emit_event(
                    "my_plan_actual_mismatch",
                    actor=actor or "flow",
                    target_user=target,
                    title="[plan/actual 불일치]",
                    body=alert_body,
                    payload=payload,
                )
                if not ok:
                    continue
                alerts[target_alert_key] = {
                    "time": datetime.datetime.now().isoformat(),
                    "target_user": target,
                    **payload,
                }
                sent += 1
                new_targets.append(target)
            if new_targets:
                mail_items.append({
                    "owner": owner or "(작성자 미상)",
                    "root": root,
                    "wafer": wafer,
                    "column": column,
                    "plan": payload["plan"],
                    "actual": payload["actual"],
                    "updated": payload["plan_updated"],
                })
                for t in new_targets:
                    if t not in mail_targets:
                        mail_targets.append(t)
        if sent:
            if len(alerts) > 2000:
                for old_key in list(alerts.keys())[: len(alerts) - 2000]:
                    alerts.pop(old_key, None)
            data["mismatch_alerts"] = alerts
            save_json(_plan_history_path(product), data)
            # 톱니바퀴 설정에서 켠 경우에만 메일 발송 — 장부 저장 뒤라 동일 건은 1회만 나간다.
            if mail_enabled:
                _send_plan_mismatch_mail(product, mail_items, mail_targets, actor=actor)
        return sent
    except Exception:
        return 0


def _mismatch_notify_pending_key(product: str, mismatch: dict) -> tuple:
    return (
        str(product or "").strip(),
        str(mismatch.get("key") or mismatch.get("cell") or ""),
        _clean_str(mismatch.get("plan")),
        _clean_str(mismatch.get("actual")),
    )


def _mismatch_notify_worker() -> None:
    while True:
        _MISMATCH_NOTIFY_WAKE.wait(_MISMATCH_NOTIFY_DEBOUNCE_SEC)
        _MISMATCH_NOTIFY_WAKE.clear()
        with _MISMATCH_NOTIFY_LOCK:
            items = list(_MISMATCH_NOTIFY_PENDING.values())
            _MISMATCH_NOTIFY_PENDING.clear()
        if not items:
            with _MISMATCH_NOTIFY_LOCK:
                if not _MISMATCH_NOTIFY_PENDING:
                    return
            continue
        grouped: dict[tuple[str, str], list[dict]] = {}
        for item in items:
            grouped.setdefault((item["product"], item["actor"]), []).append(dict(item["mismatch"]))
        for (product, actor), batch in grouped.items():
            try:
                _notify_plan_actual_mismatches_once(product, batch, actor=actor)
            except Exception:
                logger.debug("background mismatch notification failed product=%s", product, exc_info=True)


def _enqueue_plan_actual_mismatches(product: str, mismatches: list[dict], actor: str = "flow") -> None:
    if not mismatches:
        return
    global _MISMATCH_NOTIFY_THREAD
    with _MISMATCH_NOTIFY_LOCK:
        for mm in mismatches[:100]:
            key = _mismatch_notify_pending_key(product, mm)
            if not key[1]:
                continue
            _MISMATCH_NOTIFY_PENDING[key] = {
                "product": str(product or "").strip(),
                "actor": actor or "flow",
                "mismatch": dict(mm),
            }
            _MISMATCH_NOTIFY_PENDING.move_to_end(key)
            while len(_MISMATCH_NOTIFY_PENDING) > _MISMATCH_NOTIFY_PENDING_MAX:
                _MISMATCH_NOTIFY_PENDING.popitem(last=False)
        if _MISMATCH_NOTIFY_THREAD is None or not _MISMATCH_NOTIFY_THREAD.is_alive():
            _MISMATCH_NOTIFY_THREAD = threading.Thread(
                target=_mismatch_notify_worker,
                name="splittable-mismatch-notify",
                daemon=True,
            )
            _MISMATCH_NOTIFY_THREAD.start()
    _MISMATCH_NOTIFY_WAKE.set()


def _drain_plan_actual_mismatch_notifications_for_tests(timeout: float = 2.0) -> None:
    deadline = time.monotonic() + max(0.1, float(timeout or 0.0))
    while time.monotonic() < deadline:
        with _MISMATCH_NOTIFY_LOCK:
            thread = _MISMATCH_NOTIFY_THREAD
            pending = bool(_MISMATCH_NOTIFY_PENDING)
        if not pending and (thread is None or not thread.is_alive()):
            return
        if thread is not None:
            thread.join(timeout=0.05)
        else:
            time.sleep(0.05)


def _notify_tracker_owner_for_note(entry: dict, actor: str) -> None:
    try:
        from core.notify import emit_event
        from core.mail import send_mail
        product, root_lot_id, wafer_id = _note_scope_parts(entry)
        if not product or not root_lot_id:
            return
        tracker_items = load_json(TRACKER_ISSUES_FILE, [])
        for issue in tracker_items or []:
            base_target = str(issue.get("username") or "").strip()
            if not base_target or base_target == actor:
                continue
            matched = False
            for row in issue.get("lots") or []:
                row_product = str(row.get("product") or issue.get("product") or "")
                row_root = str(row.get("root_lot_id") or "")
                row_wafer = _normalize_wafer_id(row.get("wafer_id"))
                if row_product and row_product not in (product, product.replace("ML_TABLE_", "")):
                    continue
                if not _root_lot_matches(row_root, root_lot_id):
                    continue
                if wafer_id and row_wafer and row_wafer != _normalize_wafer_id(wafer_id):
                    continue
                matched = True
                break
            if not matched:
                continue
            title = f"FLOW 알림 - {issue.get('title') or issue.get('id') or 'SplitTable note'}"
            body = f"{actor} 님이 SplitTable 노트를 추가했습니다. lot={root_lot_id}" + (f" wf={wafer_id}" if wafer_id else "")
            emit_event(
                "my_tracker_lot_note",
                actor=actor,
                target_user=base_target,
                title=title,
                body=body,
                payload={"issue_id": issue.get("id"), "product": product, "root_lot_id": root_lot_id, "wafer_id": wafer_id, "note_id": entry.get("id")},
            )
            mail_watch = issue.get("mail_watch") if isinstance(issue.get("mail_watch"), dict) else {}
            if mail_watch.get("enabled"):
                send_mail(
                    sender_username=actor or "flow",
                    receiver_usernames=[base_target],
                    extra_emails=[],
                    title=title,
                    content=body,
                )
    except Exception:
        pass


@router.get("/notes")
def list_notes(product: str = Query(""), root_lot_id: str = Query(""), username: str = Query("")):
    """필터:
      - product+root_lot_id → (wafer + param + lot) for that lot
        PLUS param_global for the product (전역 태그는 모든 LOT 에서 공통 노출)
      - product only → product 전역 (param_global + lot 전체)
      - 없으면 전체
    """
    entries = _load_notes()
    if product and root_lot_id:
        lot_pfx = _notes_lot_prefix(product, root_lot_id)
        lot_key = _notes_key_lot(product, root_lot_id)
        pg_pfx = _notes_product_param_prefix(product)
        def _match(e):
            k = str(e.get("key", ""))
            sc = e.get("scope")
            if sc == "wafer" and k.startswith(lot_pfx):
                return True
            if sc == "param" and k.startswith(lot_pfx):
                return True
            if sc == "lot" and k == lot_key:
                return True
            if sc == "param_global" and k.startswith(pg_pfx):
                return True
            return False
        entries = [e for e in entries if _match(e)]
    elif product:
        pg_pfx = _notes_product_param_prefix(product)
        lot_pfx = _notes_product_lot_prefix(product)
        entries = [e for e in entries
                   if str(e.get("key", "")).startswith(pg_pfx) or str(e.get("key", "")).startswith(lot_pfx)]
    entries = [_normalize_note_entry(e) for e in entries]
    entries.sort(key=lambda e: e.get("created_at", ""), reverse=True)
    return {"notes": entries, "total": len(entries)}


@router.post("/notes/save")
def save_note(req: NoteSaveReq, request: Request):
    from core.auth import current_user as _cu
    me = _cu(request)
    username = me.get("username") or req.username or "anonymous"
    scope = (req.scope or "").strip()
    if scope not in ("wafer", "param", "lot", "param_global"):
        raise HTTPException(400, "scope must be 'wafer'|'param'|'lot'|'param_global'")
    images = _normalize_note_images(req.images)
    text = _clean_note_text(req.text)
    if not text and not images:
        raise HTTPException(400, "empty text")
    if len(text) > 2000:
        raise HTTPException(400, "text too long (max 2000 chars)")
    if not req.product:
        raise HTTPException(400, "product required")
    if scope == "wafer":
        if not req.root_lot_id or not str(req.wafer_id or "").strip():
            raise HTTPException(400, "root_lot_id/wafer_id required for wafer scope")
        key = _notes_key_wafer(req.product, req.root_lot_id, req.wafer_id)
    elif scope == "param":
        if not req.root_lot_id or not str(req.wafer_id or "").strip() or not req.param:
            raise HTTPException(400, "root_lot_id/wafer_id/param required for param scope")
        key = _notes_key_param(req.product, req.root_lot_id, req.wafer_id, req.param)
    elif scope == "lot":
        if not req.root_lot_id:
            raise HTTPException(400, "root_lot_id required for lot scope")
        key = _notes_key_lot(req.product, req.root_lot_id)
    else:  # param_global
        if not req.param:
            raise HTTPException(400, "param required for param_global scope")
        key = _notes_key_param_global(req.product, req.param)
    entry = {
        "id": _new_note_id(),
        "scope": scope,
        "key": key,
        "text": text,
        "images": images,
        "comments": [],
        "username": username,
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    entries = _load_notes()
    entries.append(entry)
    _save_notes(entries)
    _append_splittable_note_knowledge(entry, actor=username, text=text)
    _notify_tracker_owner_for_note(entry, username)
    return {"ok": True, "entry": entry}


@router.post("/notes/comment")
def add_note_comment(req: NoteCommentReq, request: Request):
    from core.auth import current_user as _cu
    me = _cu(request)
    username = me.get("username") or req.username or "anonymous"
    text = _clean_note_text(req.text)
    images = _normalize_note_images(req.images)
    if not text and not images:
        raise HTTPException(400, "empty text")
    entries = _load_notes()
    target = next((e for e in entries if e.get("id") == req.note_id), None)
    if not target:
        raise HTTPException(404, "note not found")
    comment = {
        "id": "c_" + datetime.datetime.now().strftime("%y%m%d%H%M%S%f"),
        "text": text,
        "images": images,
        "username": username,
        "created_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    target.setdefault("comments", []).append(comment)
    _save_notes(entries)
    _append_splittable_note_knowledge(target, actor=username, text=text)
    return {"ok": True, "comment": comment}


@router.post("/notes/delete")
def delete_note(req: NoteDeleteReq, request: Request):
    from core.auth import current_user as _cu
    me = _cu(request)
    username = me.get("username") or ""
    role = me.get("role") or ""
    entries = _load_notes()
    target = next((e for e in entries if e.get("id") == req.id), None)
    if not target:
        raise HTTPException(404, "note not found")
    if role != "admin" and target.get("username") != username:
        raise HTTPException(403, "only author or admin can delete")
    entries = [e for e in entries if e.get("id") != req.id]
    _save_notes(entries)
    return {"ok": True}


def _normalize_wafer_id(raw, *, max_wafer: int = SPLITTABLE_MAX_WAFER_ID) -> str:
    text = str(raw or "").strip()
    if not text:
        return ""
    core = _re.sub(r"^(?:#|WAFER|WF|W)\s*", "", text, flags=_re.I).strip()
    if not _re.fullmatch(r"\d+", core):
        return ""
    try:
        n = int(core)
    except Exception:
        return ""
    return str(n) if 1 <= n <= max_wafer else ""


def _wafer_filter_set(raw: str) -> set[str]:
    out = set()
    for part in (raw or "").split(","):
        s = str(part).strip()
        if not s:
            continue
        norm = _normalize_wafer_id(s)
        if norm:
            out.add(norm)
    return {v for v in out if v not in ("", "None", "null")}


def _wafer_matches(wafer_value, wafer_set: set[str]) -> bool:
    if not wafer_set:
        return True
    s = _normalize_wafer_id(wafer_value)
    return bool(s and s in wafer_set)


def _scope_label(has_wafer: bool) -> str:
    return "wafer" if has_wafer else "lot"


def _root_lot_matches(candidate, root_lot_id: str) -> bool:
    cand = str(candidate or "").strip()
    root = str(root_lot_id or "").strip()
    if not cand or not root:
        return False
    if cand == root:
        return True
    # Legacy tracker/inform entries sometimes stored only the old 5-char root.
    if len(cand) <= 5 and root.startswith(cand):
        return True
    if len(root) <= 5 and cand.startswith(root):
        return True
    return False


def _lot_or_fab_matches_root(value, root_lot_id: str) -> bool:
    text = str(value or "").strip()
    root = str(root_lot_id or "").strip()
    if not text or not root:
        return False
    return text == root or text.startswith(root)


def _load_operational_history(product: str, root_lot_id: str, wafer_ids: str,
                              username: str, role: str) -> list[dict]:
    if not root_lot_id:
        return []
    wafer_set = _wafer_filter_set(wafer_ids)
    out: list[dict] = []
    try:
        from routers.groups import filter_by_visibility
    except Exception:
        def filter_by_visibility(items, username, role, key="group_ids"):
            return []

    tracker_items = filter_by_visibility(load_json(TRACKER_ISSUES_FILE, []), username, role, key="group_ids")
    for issue in tracker_items or []:
        matched_rows = []
        for row in (issue.get("lots") or []):
            rid = str(row.get("root_lot_id") or "").strip()
            lot_value = str(row.get("lot_id") or "").strip()
            if not (_root_lot_matches(rid, root_lot_id) or _lot_or_fab_matches_root(lot_value, root_lot_id)):
                continue
            wafer_val = str(row.get("wafer_id") or "").strip()
            if wafer_val and not _wafer_matches(wafer_val, wafer_set):
                continue
            if not wafer_val and wafer_set:
                continue
            matched_rows.append(row)
        if not matched_rows:
            continue
        for row in matched_rows:
            out.append({
                "source": "tracker",
                "scope": _scope_label(bool(str(row.get("wafer_id") or "").strip())),
                "time": issue.get("updated_at") or issue.get("created") or issue.get("timestamp") or "",
                "author": issue.get("username") or "",
                "title": issue.get("title") or "(untitled issue)",
                "detail": row.get("comment") or "",
                "status": issue.get("status") or "",
                "category": issue.get("category") or "",
                "root_lot_id": root_lot_id,
                "wafer_id": str(row.get("wafer_id") or ""),
                "lot_id": row.get("lot_id") or "",
                "ref_id": issue.get("id") or "",
            })
        for cm in (issue.get("comments") or []):
            for row in matched_rows:
                out.append({
                    "source": "tracker_comment",
                    "scope": _scope_label(bool(str(row.get("wafer_id") or "").strip())),
                    "time": cm.get("created_at") or "",
                    "author": cm.get("username") or "",
                    "title": issue.get("title") or "(issue comment)",
                    "detail": cm.get("text") or "",
                    "status": issue.get("status") or "",
                    "category": issue.get("category") or "",
                    "root_lot_id": root_lot_id,
                    "wafer_id": str(row.get("wafer_id") or ""),
                    "lot_id": row.get("lot_id") or "",
                    "ref_id": issue.get("id") or "",
                })

    inform_items = filter_by_visibility(load_json(INFORMS_FILE, []), username, role, key="group_ids")
    for inf in inform_items or []:
        inf_root = str(inf.get("root_lot_id") or "").strip()
        inf_lot = str(inf.get("lot_id") or "").strip()
        inf_fab = str(inf.get("fab_lot_id_at_save") or inf.get("lot_id") or "").strip()
        fab_parts = [p.strip() for p in inf_fab.split(",") if p.strip()]
        if not (
            _root_lot_matches(inf_root, root_lot_id)
            or _lot_or_fab_matches_root(inf_lot, root_lot_id)
            or any(_lot_or_fab_matches_root(part, root_lot_id) for part in fab_parts)
        ):
            continue
        inf_wafer = str(inf.get("wafer_id") or "").strip()
        if inf_wafer and not _wafer_matches(inf_wafer, wafer_set):
            continue
        if not inf_wafer and wafer_set:
            continue
        out.append({
            "source": "inform",
            "scope": _scope_label(bool(inf_wafer)),
            "time": inf.get("created_at") or "",
            "author": inf.get("author") or "",
            "title": f"{inf.get('module') or 'INFO'} · {inf.get('reason') or ''}".strip(" ·"),
            "detail": inf.get("text") or "",
            "status": inf.get("flow_status") or ("completed" if inf.get("checked") else "received"),
            "category": "inform",
            "root_lot_id": root_lot_id,
            "wafer_id": inf_wafer,
            "lot_id": inf.get("lot_id") or "",
            "ref_id": inf.get("id") or "",
        })
    out.sort(key=lambda x: x.get("time") or "", reverse=True)
    return out[:300]


def _issue_comment_count(issue: dict) -> int:
    total = 0
    for cm in issue.get("comments") or []:
        if not isinstance(cm, dict):
            continue
        total += 1
        total += len([r for r in (cm.get("replies") or []) if isinstance(r, dict)])
    return total


def _product_matches_issue(product: str, issue: dict, row: dict) -> bool:
    aliases = _product_aliases(product)
    if not aliases:
        return True
    values = [
        issue.get("product"),
        row.get("product"),
        row.get("monitor_prod"),
        row.get("prod"),
    ]
    candidates = {str(v or "").strip().upper() for v in values if str(v or "").strip()}
    if not candidates:
        return True
    return bool(candidates & aliases)


def _related_tracker_issues(product: str, root_lot_id: str,
                            username: str = "", role: str = "user",
                            limit: int = 8) -> list[dict]:
    root = str(root_lot_id or "").strip()
    if not root:
        return []
    try:
        from routers.groups import filter_by_visibility
    except Exception:
        def filter_by_visibility(items, username, role, key="group_ids"):
            return []
    try:
        tracker_items = filter_by_visibility(load_json(TRACKER_ISSUES_FILE, []), username, role, key="group_ids")
    except Exception:
        tracker_items = []
    out: list[dict] = []
    for issue in tracker_items or []:
        matched_lots = []
        matched_wafers = []
        for row in (issue.get("lots") or []):
            rid = str(row.get("root_lot_id") or "").strip()
            lot_value = str(row.get("lot_id") or "").strip()
            if not (_root_lot_matches(rid, root) or _lot_or_fab_matches_root(lot_value, root)):
                continue
            if not _product_matches_issue(product, issue, row):
                continue
            matched_lots.append(lot_value or rid or root)
            wafer = str(row.get("wafer_id") or "").strip()
            if wafer:
                matched_wafers.append(wafer)
        if not matched_lots:
            continue
        out.append({
            "id": issue.get("id") or "",
            "title": issue.get("title") or "(untitled issue)",
            "status": issue.get("status") or "",
            "category": issue.get("category") or "",
            "priority": issue.get("priority") or "",
            "username": issue.get("username") or "",
            "updated_at": issue.get("updated_at") or issue.get("created") or issue.get("timestamp") or "",
            "matched_lots": sorted({v for v in matched_lots if v}),
            "matched_wafers": sorted({v for v in matched_wafers if v}, key=_natural_param_key),
            "comment_count": _issue_comment_count(issue),
        })
    out.sort(key=lambda x: x.get("updated_at") or "", reverse=True)
    return out[:max(1, min(20, int(limit or 8)))]


# ── Products / schema ──
# v8.8.3: SplitTable 의 "제품" = 오직 ML_TABLE_* 파일로 한정.
#   - 기존에는 DB hive 테이블(FAB/INLINE/ET/EDS)과 레거시 루트 파일도 노출되어
#     실제 검색 가능한 테이블셋이 혼탁했다.
#   - 신규 요청: "검색되는 테이블셋 = ML_TABLE_~~" prefix 로 시작하는 단일 파일만.
#   - DB 하위 제품 폴더는 /fab-roots / /ml-table-match 가 따로 노출 → 오버라이드용 소스.
@router.get("/products")
def list_products():
    """Base/DB root 직하의 ML_TABLE_* 단일 파일만 노출. 다른 소스는 fab_source 자동 매칭 전용.
    Source 가시성(enabled) 토글은 여전히 이 리스트 기준."""
    products = []
    roots: list[tuple[str, str, Path]] = []
    for label, resolver in (("Base", _base_root), ("DB", _db_base)):
        try:
            root = resolver()
        except Exception:
            continue
        try:
            root_key = str(root.resolve())
        except Exception:
            root_key = str(root)
        if not root or any(existing_key == root_key for _, existing_key, _ in roots):
            continue
        roots.append((label, root_key, root))
    for label, _root_key, root in roots:
        try:
            if not root.exists():
                continue
            for f in sorted(root.iterdir(), key=lambda p: p.name.lower()):
                if not _is_mltable_product_file(f):
                    continue
                products.append({"name": _canonical_mltable_product_name(f.stem), "file": f.name, "size": f.stat().st_size,
                                 "root": label, "type": f.suffix.lower().lstrip("."), "source_type": "base_file"})
        except Exception:
            pass
    # dedup 은 불필요하지만 안정성을 위해 이름 기준 중복 제거.
    seen = set()
    dedup = []
    for p in products:
        n = p.get("name") or ""
        if n in seen:
            continue
        seen.add(n)
        dedup.append(p)
    dedup.sort(key=lambda p: (p.get("name") or ""))
    return {"products": dedup}


# v8.8.5: 사내 실데이터 구조 대응.
#   - base_root == db_root (동일 폴더).
#   - 상위 DB 폴더 이름이 `1.RAWDATA_DB*` prefix (예: `1.RAWDATA_DB`, `1.RAWDATA_DB_FAB`, `1.RAWDATA_DB_INLINE`).
#   - 제품 폴더 안은 hive 파티션: `PRODA/date=YYYYMMDD/part_*.parquet`.
#   - 동시에 Base 단일 파일 `ML_TABLE_<PROD>.parquet` 도 같은 폴더 레벨에 있음.
# v8.8.18: `1.RAWDATA_DB` 는 exact match — `_INLINE`/`_FAB` 등 suffix 붙은 변형은
#   별도 폴더로 취급 (override 소스로 자동 매칭하지 않음). 명시적 legacy 짧은 이름은 유지.
#   사용자가 직접 lot_overrides[product].fab_source 로 `1.RAWDATA_DB_INLINE/<PROD>` 를
#   지정하면 그 경로는 존중.
_RAWDATA_EXACT = "1.RAWDATA_DB"
_RAWDATA_FAB = "1.RAWDATA_DB_FAB"
_LEGACY_SHORT_ROOTS = {"FAB", "INLINE", "ET", "EDS"}

def _is_db_root_dir(p) -> bool:
    if not p.is_dir():
        return False
    n = p.name
    up = n.upper()
    if n == _RAWDATA_EXACT or up == _RAWDATA_FAB.upper():
        return True
    if up.startswith(_RAWDATA_EXACT.upper() + "_"):
        return True
    if up in _LEGACY_SHORT_ROOTS:
        return True
    return False


def _rank_db_root_name(name: str) -> tuple[int, str]:
    up = str(name or "").upper()
    if up == _RAWDATA_EXACT.upper():
        return (0, up)
    if up == _RAWDATA_FAB.upper():
        return (1, up)
    if up.startswith(_RAWDATA_EXACT.upper() + "_"):
        return (2, up)
    if "FAB" in up:
        return (3, up)
    if "INLINE" in up:
        return (4, up)
    if "ET" in up:
        return (5, up)
    if "EDS" in up:
        return (6, up)
    return (7, up)


# v8.8.22: case-insensitive 제품 폴더 lookup.
#   ML_TABLE_PRODA → DB/1.RAWDATA_DB/ProdA/ · proda/ · PRODA/ 모두 동일하게 매칭.
#   exact match 우선, 없으면 casefold 동등 비교.
def _find_ci_child(parent, name: str):
    """parent 아래에서 name 과 case-insensitive 동등한 디렉토리를 반환 (없으면 None)."""
    if not name or not parent or not parent.exists():
        return None
    try:
        exact = parent / name
        if exact.is_dir():
            return exact
    except Exception:
        pass
    try:
        target = name.casefold()
        for child in parent.iterdir():
            if child.is_dir() and child.name.casefold() == target:
                return child
    except Exception:
        pass
    return None


def _find_ci_path(root, rel: str):
    """root 아래의 쉼표 없는 상대경로 rel 을 case-insensitive 하게 찾아 반환.
    rel 이 '1.RAWDATA_DB/ProdA' 같이 슬래시 포함 시 각 세그먼트별로 CI 매칭 시도.
    파일이 아닌 경우에도 마지막 세그먼트가 .parquet/.csv 일 수 있어 is_file 도 허용.
    """
    if not rel or not root or not root.exists():
        return None
    # exact first
    try:
        exact = root / rel
        if exact.exists():
            return exact
    except Exception:
        pass
    parts = [p for p in rel.replace("\\", "/").split("/") if p]
    cur = root
    for i, seg in enumerate(parts):
        is_last = (i == len(parts) - 1)
        try:
            nxt = cur / seg
            if nxt.exists():
                cur = nxt
                continue
        except Exception:
            pass
        target = seg.casefold()
        found = None
        try:
            for child in cur.iterdir():
                if child.name.casefold() == target:
                    found = child
                    break
        except Exception:
            return None
        if found is None:
            return None
        cur = found
    return cur

def _list_db_roots():
    """사내/레거시 공통 DB 상위 폴더 후보 스캔. 반환 순서 = 우선순위.
    - 자동 연결은 `1.RAWDATA_DB` → `1.RAWDATA_DB_FAB` → 기타 `1.RAWDATA_DB_*` 순 우선.
    - 같은 우선군 안에서는 이름 오름차순.

    v8.8.17: db_root 자체가 `1.RAWDATA_DB*` 또는 그 안에 `1.RAWDATA_DB*/` 가
      없을 때도 작동하도록 확장.
        1) db_base 가 바로 `1.RAWDATA_DB*` 디렉토리면 → [db_base]
        2) db_base 아래에 `1.RAWDATA_DB*` 자식이 있으면 → 그 자식들 (기존 동작)
        3) 위 둘 다 아니고 db_base 바로 아래에 제품 폴더(parquet 포함) 가 있으면
           → [db_base] 자체를 rawdata 루트로 취급 (사용자가 rawdata 하위를 직접 지정한 경우).
    """
    db_base = _db_base()
    if not db_base.exists():
        return []
    try:
        cache_key = str(db_base.resolve())
    except Exception:
        cache_key = str(db_base)
    now = time.monotonic()
    cached = _DB_ROOTS_CACHE.get(cache_key)
    if cached and now - cached[0] < _DISCOVERY_CACHE_TTL_SEC:
        return list(cached[1])
    # Case 1: children match — legacy `Fab/` 아래의 `1.RAWDATA_DB_*` 구조를 우선 존중.
    cands = [p for p in db_base.iterdir() if _is_db_root_dir(p)]
    if cands:
        cands.sort(key=lambda p: _rank_db_root_name(p.name))
        _DB_ROOTS_CACHE[cache_key] = (now, list(cands))
        return cands
    # Case 2: db_base itself is a direct rawdata root (or a legacy short root with no rawdata children).
    if _is_db_root_dir(db_base):
        out = [db_base]
        _DB_ROOTS_CACHE[cache_key] = (now, out)
        return out
    # Case 3: db_base has no 1.RAWDATA_DB* children, but has product-like subfolders
    # (any subfolder that contains at least one parquet, possibly under hive date=* part).
    try:
        has_product = False
        for sub in db_base.iterdir():
            if not sub.is_dir():
                continue
            # Peek: is there any parquet under this subfolder (any depth ≤ 3)?
            for depth in range(3):
                pattern = "/".join(["*"] * depth) + ("/" if depth else "") + "*.parquet"
                # fall back to simple rglob
            found = _first_data_file_ci(sub, (".parquet",)) is not None
            if found:
                has_product = True
                break
        if has_product:
            out = [db_base]
            _DB_ROOTS_CACHE[cache_key] = (now, out)
            return out
    except Exception:
        pass
    _DB_ROOTS_CACHE[cache_key] = (now, [])
    return []


@router.get("/fab-roots")
def list_fab_roots():
    """v8.7.8/v8.8.5: DB 최상위 폴더 목록. `1.RAWDATA_DB*` 접두 폴더 + 레거시 FAB/INLINE/ET/EDS 짧은 이름 모두 인식.
    Returns: {roots: [{name, products: [...], total_size}], ...}
    """
    out = []
    for root_dir in _list_db_roots():
        products = []
        total_size = 0
        try:
            for prod_dir in sorted(root_dir.iterdir()):
                if not prod_dir.is_dir():
                    continue
                # 리스트 화면은 "제품으로 볼 수 있는가"만 필요하다. 실데이터에서
                # 전체 rglob+sort 는 수만 파티션을 훑으므로 첫 파일만 확인한다.
                f = _first_data_file_ci(prod_dir, (".parquet", ".csv"))
                has_data = f is not None
                if has_data:
                    try: total_size += f.stat().st_size
                    except Exception: pass
                if has_data:
                    products.append(prod_dir.name)
        except Exception:
            continue
        if products:
            out.append({"name": root_dir.name, "products": products, "total_size": total_size})
    return {"roots": out}


@router.get("/ml-table-match")
def ml_table_match(product: str = Query(...), detail: bool = False):
    """v8.7.8/v8.8.5: ML_TABLE_<PROD> 에서 PROD 추출 → `1.RAWDATA_DB*` / 레거시 짧은 이름 상위폴더 내 <PROD>/ 매칭.
    Ex) product=ML_TABLE_PRODA → {"matches": [{"root":"1.RAWDATA_DB_FAB","product":"PRODA","path":"1.RAWDATA_DB_FAB/PRODA"}, ...]}
    v8.8.3: 자동으로 선택된 fab_source (_auto_derive_fab_source) 와 현재 override 상태도 같이 반환.
    """
    pro = ""
    p = (product or "").strip()
    if p.casefold().startswith("ml_table_"):
        pro = p[len("ML_TABLE_"):].strip()
    elif "_" in p:
        pro = p.rsplit("_", 1)[-1]
    else:
        pro = p
    matches = []
    if pro:
        for root_dir in _list_db_roots():
            # v8.8.22: case-insensitive — ProdA/proda/PRODA 모두 같은 제품으로 매칭.
            sub = _find_ci_child(root_dir, pro)
            if sub is not None:
                matches.append({
                    "root": root_dir.name,
                    "product": sub.name,  # 실제 폴더 이름 (대소문자 반영)
                    "path": f"{root_dir.name}/{sub.name}",
                })
    auto_path = _auto_derive_fab_source(p)
    manual_ov = {}
    try:
        cfg = load_json(SOURCE_CFG, {}) or {}
        manual_ov = _lot_override_for(cfg, p)
    except Exception:
        pass
    manual_fs = _normalize_fab_source_path((manual_ov.get("fab_source") or "").strip())
    effective = manual_fs or auto_path
    # Default to the light resolver.  The full resolver scans FAB parquet just
    # to populate diagnostics, which made product switching feel slow.
    override_meta = _resolve_override_meta(p, include_diagnostics=False) if detail else _resolve_override_meta_light(p)
    return {
        "product": p,
        "derived_product": pro,
        "matches": matches,
        "auto_path": auto_path,
        "manual_override": bool(manual_fs),
        "effective_fab_source": effective,
        "override": override_meta,
        "match_cache": _match_cache_response_meta(p),
    }


@router.get("/override-link-preview")
def override_link_preview(
    product: str = Query(...),
    fab_root: str = Query(""),
    fab_source: str = Query(""),
    limit: int = Query(5, ge=1, le=20),
):
    """Preview a manual FAB link before persisting it.

    UI flow:
      1. select DB top folder (`fab_root`) or a full `fab_source`
      2. inspect detected columns / recommended fields
      3. preview most recent fab_lot_id values
      4. save into source-config only after confirmation
    """
    p = (product or "").strip()
    if not p:
        raise HTTPException(400, "product required")

    derived = ""
    if p.casefold().startswith("ml_table_"):
        derived = p[len("ML_TABLE_"):].strip()
    elif "_" in p:
        derived = p.rsplit("_", 1)[-1]
    else:
        derived = p

    selected_root = ""
    source = _normalize_fab_source_path(fab_source)
    if fab_root and not source:
        selected_root = str(fab_root or "").strip()
        root_dir = next((r for r in _list_db_roots() if r.name.casefold() == selected_root.casefold()), None)
        if root_dir is None:
            raise HTTPException(404, f"DB top folder not found: {fab_root}")
        prod_dir = _find_ci_child(root_dir, derived) if derived else None
        if prod_dir is None:
            return {
                "product": p,
                "derived_product": derived,
                "fab_root": root_dir.name,
                "fab_source": "",
                "matched_product_dir": "",
                "columns": [],
                "latest_fab_lot_ids": [],
                "recommended": {},
                "error": f"{root_dir.name} 아래에서 제품 폴더 '{derived}' 를 찾지 못했습니다.",
            }
        source = f"{root_dir.name}/{prod_dir.name}"
    elif source:
        selected_root = source.split("/", 1)[0]

    if not source:
        return {
            "product": p,
            "derived_product": derived,
            "fab_root": selected_root,
            "fab_source": "",
            "matched_product_dir": "",
            "columns": [],
            "latest_fab_lot_ids": [],
            "recommended": {},
            "error": "fab_root 또는 fab_source 가 필요합니다.",
        }

    raw_lf = _scan_fab_source_raw(source)
    fab_lf = _scan_fab_source(source)
    if fab_lf is None:
        return {
            "product": p,
            "derived_product": derived,
            "fab_root": selected_root,
            "fab_source": source,
            "matched_product_dir": source.split("/", 1)[1] if "/" in source else "",
            "columns": [],
            "raw_columns": [],
            "column_aliases": {},
            "schema_mode": "unknown",
            "latest_fab_lot_ids": [],
            "recommended": {},
            "error": f"소스를 읽지 못했습니다: {source}",
        }

    try:
        main_names = _scan_parquet_compat(str(_product_path(p))).collect_schema().names()
    except Exception:
        main_names = []
    fab_lf, fab_names = _ci_align_fab_to_main(fab_lf, main_names)
    if not fab_names:
        try:
            fab_names = fab_lf.collect_schema().names()
        except Exception:
            fab_names = []
    try:
        raw_names = raw_lf.collect_schema().names() if raw_lf is not None else []
    except Exception:
        raw_names = []
    column_aliases = _detect_source_column_aliases(raw_names, fab_names)
    schema_mode = "adapted" if column_aliases else "raw"

    root_col, wf_col = find_lot_wafer_cols(fab_names)
    fab_col = _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_names) or ""
    ts_col = _pick_ts_col(fab_names) or ""
    join_keys = _default_override_join_keys(main_names, fab_names)

    latest_fab_lot_ids: list[str] = []
    if fab_col and fab_col in fab_names:
        try:
            q = fab_lf
            if ts_col and ts_col in fab_names:
                q = q.sort(ts_col, descending=True, nulls_last=True)
            latest = (
                q.select([pl.col(fab_col).cast(_STR, strict=False)])
                 .filter(pl.col(fab_col).is_not_null() & (pl.col(fab_col).cast(_STR, strict=False) != ""))
                 .unique(maintain_order=True)
                 .head(limit)
                 .collect()
            )
            latest_fab_lot_ids = [str(v) for v in latest[fab_col].to_list() if v not in (None, "")]
        except Exception:
            latest_fab_lot_ids = []

    recommended_override_cols = []
    for c in list(_DEFAULT_OVERRIDE_COLS) + ([fab_col] if fab_col else []):
        actual = _resolve_source_col_name(c, fab_names)
        if actual and actual not in recommended_override_cols and actual not in join_keys:
            recommended_override_cols.append(actual)
    recommended_override_cols = [
        _prefer_raw_schema_name(c, raw_names, fab_names) for c in recommended_override_cols
    ]
    join_keys_preview = [_prefer_raw_schema_name(k, raw_names, fab_names) for k in join_keys]

    return {
        "product": p,
        "derived_product": derived,
        "fab_root": selected_root,
        "fab_source": source,
        "matched_product_dir": source.split("/", 1)[1] if "/" in source else "",
        "columns": fab_names,
        "raw_columns": raw_names or fab_names,
        "column_aliases": column_aliases,
        "schema_mode": schema_mode,
        "latest_fab_lot_ids": latest_fab_lot_ids,
        "recommended": {
            "root_col": _prefer_raw_schema_name(root_col or "", raw_names, fab_names),
            "wf_col": _prefer_raw_schema_name(wf_col or "", raw_names, fab_names),
            "fab_col": _prefer_raw_schema_name(fab_col, raw_names, fab_names),
            "ts_col": _prefer_raw_schema_name(ts_col, raw_names, fab_names),
            "join_keys": join_keys_preview,
            "override_cols": recommended_override_cols,
        },
        "recommended_runtime": {
            "root_col": root_col or "",
            "wf_col": wf_col or "",
            "fab_col": fab_col,
            "ts_col": ts_col,
            "join_keys": join_keys,
            "override_cols": [
                _resolve_source_col_name(c, fab_names) for c in recommended_override_cols
                if _resolve_source_col_name(c, fab_names)
            ],
        },
        "error": None,
    }


# v8.8.26: override 조인이 왜 실패했는지 진단용 — main vs fab 스키마/샘플/조인 결과를
#   한 번의 호출로 끝까지 보여줘 FE/운영자가 root cause 를 즉시 파악할 수 있게.
@router.get("/override-debug")
def override_debug(product: str = Query(...)):
    """진단 엔드포인트. override 조인이 비어있게 나올 때 어디서 문제가 났는지
    한 번에 확인하기 위한 용도. 반환:
      - meta: _resolve_override_meta (fab_source / join_keys / override_cols_*)
      - main_schema / main_schema_types (첫 30개)
      - fab_raw_schema / fab_raw_types (CI align 전, 첫 30개)
      - fab_aligned_schema (CI align 후, 첫 30개)
      - join_keys_resolved (main/fab 양쪽에 존재하는 것)
      - main_sample / fab_sample (join_keys + override_cols 각 3행)
      - main_lot_nonnull (main 의 root_lot_id 계열 컬럼 non-null 카운트)
      - join_probe_row_count (슬라이스 조인 결과 행 수)
    """
    out: dict = {"product": product, "error": None}
    try:
        fp = _product_path(product)
        if fp.suffix.lower() == ".csv":
            main_lf = _cast_cats_lazy(pl.scan_csv(str(fp), infer_schema_length=5000))
        else:
            main_lf = _cast_cats_lazy(_scan_parquet_compat(str(fp)))
        main_schema = main_lf.collect_schema()
        main_names = main_schema.names()
        out["main_schema"] = main_names[:30]
        out["main_schema_types"] = [str(main_schema[n]) for n in main_names[:30]]
    except Exception as e:
        out["error"] = f"main 스키마 조회 실패: {type(e).__name__}: {e}"
        return out

    meta = _resolve_override_meta(product)
    out["meta"] = meta
    fab_source = (meta.get("fab_source") or "").strip()
    if not fab_source:
        out["note"] = "fab_source 비어있음 → override off."
        return out

    fab_lf_raw = _scan_fab_source(fab_source)
    if fab_lf_raw is None:
        out["error"] = "_scan_fab_source 가 None 반환."
        return out
    try:
        raw_schema = fab_lf_raw.collect_schema()
        raw_names = raw_schema.names()
        out["fab_raw_schema"] = raw_names[:30]
        out["fab_raw_types"] = [str(raw_schema[n]) for n in raw_names[:30]]
    except Exception as e:
        out["error"] = f"fab raw 스키마 조회 실패: {type(e).__name__}: {e}"
        return out

    fab_lf_aligned, aligned_names = _ci_align_fab_to_main(fab_lf_raw, main_names)
    try:
        aligned_names = fab_lf_aligned.collect_schema().names()
    except Exception as e:
        out["align_error"] = f"{type(e).__name__}: {e}"
    out["fab_aligned_schema"] = aligned_names[:30]

    join_keys = list(meta.get("join_keys") or [])
    join_keys_resolved = [k for k in join_keys
                          if k in main_names and k in aligned_names]
    out["join_keys_resolved"] = join_keys_resolved

    override_cols = [c for c in (meta.get("override_cols_present") or [])
                     if c not in join_keys_resolved]
    out["override_cols_effective"] = override_cols

    # 샘플 행 — 에러나도 반환값은 유지.
    try:
        keep_main = [c for c in join_keys_resolved if c in main_names]
        if keep_main:
            ms = main_lf.select([pl.col(c).cast(_STR, strict=False) for c in keep_main]) \
                        .head(3).collect()
            out["main_sample"] = ms.to_dicts()
        else:
            out["main_sample"] = []
    except Exception as e:
        out["main_sample_error"] = f"{type(e).__name__}: {e}"
    try:
        keep_fab = list(dict.fromkeys(join_keys_resolved + override_cols[:5]))
        keep_fab = [c for c in keep_fab if c in aligned_names]
        if keep_fab:
            fs = fab_lf_aligned.select([pl.col(c).cast(_STR, strict=False) for c in keep_fab]) \
                               .head(3).collect()
            out["fab_sample"] = fs.to_dicts()
        else:
            out["fab_sample"] = []
    except Exception as e:
        out["fab_sample_error"] = f"{type(e).__name__}: {e}"

    # main lot 계열 컬럼의 non-null 카운트 (root_lot_id / lot_id CI).
    try:
        lot_candidates = []
        for n in main_names:
            if n.casefold() in ("root_lot_id", "lot_id"):
                lot_candidates.append(n)
        nonnull = {}
        if lot_candidates:
            row = main_lf.select(
                [pl.col(c).cast(_STR, strict=False).is_not_null().sum().alias(c)
                 for c in lot_candidates]
            ).collect()
            for c in lot_candidates:
                try:
                    nonnull[c] = int(row[c][0])
                except Exception:
                    nonnull[c] = None
        out["main_lot_nonnull"] = nonnull
    except Exception as e:
        out["main_lot_nonnull_error"] = f"{type(e).__name__}: {e}"

    # probe join: 작은 슬라이스로 실제 조인 결과가 나오는지 확인.
    try:
        if join_keys_resolved and override_cols:
            probe = _scan_product(product).select(
                join_keys_resolved + override_cols[:3]
            ).head(20).collect()
            out["join_probe_row_count"] = int(probe.height)
            out["join_probe_sample"] = probe.head(3).to_dicts()
        else:
            out["join_probe_row_count"] = 0
            out["join_probe_note"] = "join_keys_resolved 또는 override_cols 가 비어있음."
    except Exception as e:
        out["join_probe_error"] = f"{type(e).__name__}: {e}"

    return out



@router.get("/schema")
def get_schema(product: str = Query(...), root_lot_id: str = Query(""),
               fab_lot_id: str = Query(""), wafer_ids: str = Query("")):
    """v8.8.23: 오버라이드 조인을 포함한 실제 view 컬럼과 동일한 스키마를 반환.
       기존에는 ML_TABLE 원본 parquet 컬럼만 반환 → CUSTOM 선택 pool 에 root_lot_id 등
       오버라이드 컬럼이 들어가지 못해 검색/필터 드롭다운에서 누락. `_scan_product` 로
       post-join LazyFrame 스키마를 계산하고, `override_cols` (실제 join 성공한 오버라이드 컬럼)
       을 별도 필드로도 내려 FE 가 '오버라이드 제공' 뱃지를 표시할 수 있게 한다.
    """
    root_lot_id = root_lot_id if isinstance(root_lot_id, str) else ""
    fab_lot_id = fab_lot_id if isinstance(fab_lot_id, str) else ""
    wafer_ids = wafer_ids if isinstance(wafer_ids, str) else ""
    cols = []
    candidate_cache = ""
    knob_columns: list[str] = []
    try:
        # 제품 전체 스키마/KNOB 목록은 lookup 빌드가 이미 저장해 둔다. 제품 선택
        # 때 원천 공유 parquet의 metadata조차 다시 열지 않고 작은 JSON 두 개로 끝낸다.
        if not (root_lot_id or fab_lot_id or wafer_ids):
            fp = _product_path(product)
            status = _ml_table_lookup.cache_status(fp)
            schema_map = (status.get("meta") or {}).get("schema") or {}
            index = _ml_table_lookup.read_candidate_index(fp)
            if (
                status.get("status") == "fresh"
                and isinstance(schema_map, dict) and schema_map
                and int(index.get("version") or 0) == _ml_table_lookup.CANDIDATE_INDEX_VERSION
            ):
                cols = [{"name": name, "dtype": str(dtype)} for name, dtype in schema_map.items()]
                knob_columns = list((index.get("columns_by_prefix") or {}).get("KNOB") or [])
                candidate_cache = "lookup_index"
    except Exception:
        cols = []
    try:
        if cols:
            pass
        elif root_lot_id or fab_lot_id or wafer_ids:
            lf = _scan_product(product, root_lot_id=root_lot_id,
                               fab_lot_id=fab_lot_id, wafer_ids=wafer_ids)
            schema = lf.collect_schema()
            cols = [{"name": n, "dtype": str(d)} for n, d in schema.items()]
        else:
            lf = _scan_product_base(product)
            schema = lf.collect_schema()
            cols = [{"name": n, "dtype": str(d)} for n, d in schema.items()]
    except Exception:
        # fallback — 조인 실패해도 원본 컬럼은 반환.
        fp = _product_path(product)
        if fp.suffix.lower() == ".csv":
            lf = pl.scan_csv(str(fp), infer_schema_length=5000)
        else:
            lf = _scan_parquet_compat(str(fp))
        cols = [{"name": n, "dtype": str(d)} for n, d in lf.schema.items()]
    existing_cols = {str(c.get("name") or "") for c in cols}
    for tag_col in _custom_tag_columns_for_product(product):
        column = tag_col.get("column")
        if column and column not in existing_cols:
            cols.append({"name": column, "dtype": "custom_tag", "label": tag_col.get("label") or column})
            existing_cols.add(column)
    for mgmt_col in _management_row_columns_for_product(product):
        column = mgmt_col.get("column")
        if column and column not in existing_cols:
            cols.append({"name": column, "dtype": "management_row", "label": mgmt_col.get("label") or column})
            existing_cols.add(column)
    # 오버라이드에서 실제로 join 된 컬럼 목록 (FE 가 검색 pool 에서 '숨김 해제' 할 기준).
    override_cols_present: list = []
    try:
        meta = _resolve_override_meta_light(product)
        if meta.get("enabled"):
            override_cols_present = list(meta.get("override_cols_present") or meta.get("override_cols") or [])
    except Exception:
        pass
    return {
        "columns": cols,
        "total": len(cols),
        "override_cols_present": override_cols_present,
        "knob_columns": knob_columns,
        "candidate_cache": candidate_cache,
    }


# ── v4.1 Base-scope feature join (adapter-engineer slice) ─────────────────
_ET_FILE = "features_et_wafer.parquet"
_INLINE_FILE = "features_inline_agg.parquet"
_UNIQUES_FILE = "_uniques.json"
_JOIN_KEYS = ["lot_id", "wafer_id", "product"]


def _read_et_and_inline():
    """Read both wide-feature parquets from Base root (lazy→collect).

    Returns (et_df, inline_df). Raises HTTPException(404) if a file is missing.
    """
    base = _base_root()
    et_fp = base / _ET_FILE
    inl_fp = base / _INLINE_FILE
    missing = [f.name for f in (et_fp, inl_fp) if not f.is_file()]
    if missing:
        raise HTTPException(
            404,
            f"Base feature file(s) not found under {base}: {', '.join(missing)}",
        )
    try:
        from core.utils import filter_valid_wafer_ids_df
        et = filter_valid_wafer_ids_df(pl.read_parquet(str(et_fp)))
        inl = filter_valid_wafer_ids_df(pl.read_parquet(str(inl_fp)))
    except Exception as e:
        raise HTTPException(500, f"Failed to read Base parquet: {e}")
    return et, inl


def _join_features(et: pl.DataFrame, inl: pl.DataFrame) -> pl.DataFrame:
    """ET-left-join INLINE on (lot_id, wafer_id, product).

    Default per Q005 — ET has 750 rows (wafer coverage), INLINE has 50.
    Left join keeps the ET row count and nulls out inline-side columns for
    wafers without INLINE aggregation.
    """
    # Sanity: all join keys must exist on both sides
    keys = [k for k in _JOIN_KEYS if k in et.columns and k in inl.columns]
    if len(keys) < 2:
        raise HTTPException(
            500,
            f"Insufficient common join keys (need subset of {_JOIN_KEYS}, "
            f"found {keys}). ET cols: {et.columns[:5]}… INLINE cols: {inl.columns[:5]}…",
        )
    return et.join(inl, on=keys, how="left")


def _long_pivot_source(source: str) -> str:
    src = str(source or "").strip().lower()
    if src not in {"fab", "inline", "et"}:
        raise HTTPException(400, "source must be fab|inline|et")
    return src


def _long_pivot_product(product: str) -> str:
    prod = str(product or "").strip()
    if prod.upper().startswith("ML_TABLE_"):
        prod = prod[len("ML_TABLE_"):]
    return prod


def _long_pivot_key(source: str, product: str) -> str:
    return f"{_long_pivot_source(source)}:{_long_pivot_product(product).upper()}"


def _long_pivot_cache_path(source: str, product: str) -> Path:
    name = f"{_long_pivot_source(source)}_{safe_id(_long_pivot_product(product) or 'product')}.parquet"
    return _LONG_PIVOT_CACHE_DIR / name


def _long_pivot_meta_path(source: str, product: str) -> Path:
    return _long_pivot_cache_path(source, product).with_suffix(".json")


def _long_pivot_source_dir(source: str, product: str) -> Path:
    src = _long_pivot_source(source)
    folder = {
        "fab": "1.RAWDATA_DB_FAB",
        "inline": "1.RAWDATA_DB_INLINE",
        "et": "1.RAWDATA_DB_ET",
    }[src]
    return _db_base() / folder / _long_pivot_product(product)


def _long_pivot_source_signature(source: str, product: str) -> dict:
    root = _long_pivot_source_dir(source, product)
    count = 0
    total_size = 0
    max_mtime = 0.0
    try:
        files = sorted(root.rglob("*.parquet")) if root.is_dir() else []
    except Exception:
        files = []
    for fp in files:
        try:
            st = fp.stat()
        except Exception:
            continue
        count += 1
        total_size += int(st.st_size)
        max_mtime = max(max_mtime, float(st.st_mtime))
    return {
        "root": str(root),
        "file_count": count,
        "total_size": total_size,
        "max_mtime": max_mtime,
    }


def _read_long_pivot_meta(source: str, product: str) -> dict:
    fp = _long_pivot_meta_path(source, product)
    try:
        data = json.loads(fp.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _write_long_pivot_meta(source: str, product: str, meta: dict) -> None:
    fp = _long_pivot_meta_path(source, product)
    fp.parent.mkdir(parents=True, exist_ok=True)
    tmp = fp.with_suffix(fp.suffix + ".tmp")
    tmp.write_text(json.dumps(meta, ensure_ascii=False, indent=2), encoding="utf-8")
    tmp.replace(fp)


def _long_pivot_job_status(source: str, product: str) -> str:
    target = _long_pivot_key(source, product)
    with _LONG_PIVOT_JOB_LOCK:
        if _LONG_PIVOT_JOB_STATE.get("running") and _LONG_PIVOT_JOB_STATE.get("current") == target:
            return "running"
        queued = {_long_pivot_key(src, prod) for src, prod, _force in _LONG_PIVOT_QUEUE}
    return "queued" if target in queued else ""


def _long_pivot_cache_status(source: str, product: str) -> dict:
    src = _long_pivot_source(source)
    prod = _long_pivot_product(product)
    cache_fp = _long_pivot_cache_path(src, prod)
    meta = _read_long_pivot_meta(src, prod)
    source_sig = _long_pivot_source_signature(src, prod)
    has_cache = bool(cache_fp.is_file() and meta.get("version") == LONG_PIVOT_CACHE_VERSION)
    stale = bool(has_cache and meta.get("source_signature") != source_sig)
    status = "fresh" if has_cache and not stale else ("stale" if has_cache else "missing")
    job = _long_pivot_job_status(src, prod)
    if job and status != "fresh":
        status = job
    return {
        "ok": True,
        "source": src,
        "product": prod,
        "status": status,
        "has_cache": has_cache,
        "source_stale": stale,
        "source_exists": int(source_sig.get("file_count") or 0) > 0,
        "source_signature": source_sig,
        "cache_path": str(cache_fp),
        "meta_path": str(_long_pivot_meta_path(src, prod)),
        "job_status": job,
        "meta": meta,
    }


def _long_pivot_cache_public(status: dict | None, queued: dict | None = None) -> dict:
    status = status or {}
    queued = queued or {}
    queued_status = str(queued.get("status") or "").strip()
    queued_flag = bool(queued.get("queued") or queued_status in {"queued", "running"})
    meta = status.get("meta") or {}
    return {
        "status": queued_status if queued_flag else str(status.get("status") or ""),
        "hit": str(status.get("status") or "") == "fresh",
        "queued": queued_flag,
        "has_cache": bool(status.get("has_cache")),
        "source_stale": bool(status.get("source_stale")),
        "source_exists": bool(status.get("source_exists")),
        "row_count": int(meta.get("row_count") or 0),
        "built_at": meta.get("built_at") or "",
    }


def _scan_long_pivot_source(source: str, product: str):
    from core.long_pivot import scan_long_fab, scan_long_inline, scan_long_et

    src = _long_pivot_source(source)
    prod = _long_pivot_product(product)
    db_root = _db_base()
    if src == "fab":
        return scan_long_fab(prod, db_root)
    if src == "inline":
        return scan_long_inline(prod, db_root)
    return scan_long_et(prod, db_root)


def _long_pivot_function(source: str):
    from core.long_pivot import pivot_fab_wide, pivot_inline_wafer, pivot_et_wafer

    src = _long_pivot_source(source)
    if src == "fab":
        return pivot_fab_wide
    if src == "inline":
        return pivot_inline_wafer
    return pivot_et_wafer


def _build_long_pivot_cache(source: str, product: str, *, force: bool = False) -> dict:
    src = _long_pivot_source(source)
    prod = _long_pivot_product(product)
    status = _long_pivot_cache_status(src, prod)
    if status.get("status") == "fresh" and not force:
        return {"ok": True, "skipped": True, "reason": "fresh", "pivot_cache": _long_pivot_cache_public(status)}
    try:
        from core.runtime_limits import process_memory_high
        if process_memory_high():
            return {"ok": False, "skipped": True, "reason": "process_memory_high", "pivot_cache": _long_pivot_cache_public(status)}
    except Exception:
        pass
    lf = _scan_long_pivot_source(src, prod)
    if lf is None:
        return {"ok": False, "skipped": True, "reason": "source_missing", "pivot_cache": _long_pivot_cache_public(status)}
    pivot = _long_pivot_function(src)
    cache_fp = _long_pivot_cache_path(src, prod)
    tmp = cache_fp.with_suffix(cache_fp.suffix + ".tmp")
    cache_fp.parent.mkdir(parents=True, exist_ok=True)
    started = time.monotonic()
    wide = None
    try:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
        wide = pivot(lf)
        wide.write_parquet(str(tmp))
        tmp.replace(cache_fp)
        meta = {
            "version": LONG_PIVOT_CACHE_VERSION,
            "source": src,
            "product": prod,
            "source_signature": _long_pivot_source_signature(src, prod),
            "row_count": int(wide.height),
            "total_cols": len(wide.columns),
            "schema": {col: str(wide.schema[col]) for col in wide.columns},
            "built_at": datetime.datetime.now().isoformat(timespec="seconds"),
            "build_seconds": round(time.monotonic() - started, 3),
        }
        _write_long_pivot_meta(src, prod, meta)
        return {"ok": True, "cache_path": str(cache_fp), "meta": meta}
    finally:
        if wide is not None:
            try:
                del wide
            except Exception:
                pass
        try:
            gc.collect()
        except Exception:
            pass


def _long_pivot_worker_loop() -> None:
    while True:
        with _LONG_PIVOT_JOB_LOCK:
            if not _LONG_PIVOT_QUEUE:
                _LONG_PIVOT_JOB_STATE.update({"running": False, "queued": False, "current": ""})
                return
            source, product, force = _LONG_PIVOT_QUEUE.popleft()
            key = _long_pivot_key(source, product)
            _LONG_PIVOT_JOB_STATE.update({
                "running": True,
                "queued": bool(_LONG_PIVOT_QUEUE),
                "current": key,
                "started_at": datetime.datetime.now().isoformat(timespec="seconds"),
                "last_error": "",
            })
        try:
            result = _build_long_pivot_cache(source, product, force=force)
            with _LONG_PIVOT_JOB_LOCK:
                _LONG_PIVOT_JOB_STATE["last_source"] = key
                if result.get("reason") and not result.get("ok"):
                    _LONG_PIVOT_JOB_STATE["last_error"] = str(result.get("reason") or "")
        except Exception as exc:
            logger.warning("SplitTable long pivot cache build failed source=%s product=%s: %s", source, product, exc, exc_info=True)
            with _LONG_PIVOT_JOB_LOCK:
                _LONG_PIVOT_JOB_STATE["last_error"] = f"{type(exc).__name__}: {exc}"
                _LONG_PIVOT_JOB_STATE["last_source"] = key
        finally:
            with _LONG_PIVOT_JOB_LOCK:
                _LONG_PIVOT_JOB_STATE["finished_at"] = datetime.datetime.now().isoformat(timespec="seconds")


def enqueue_long_pivot_cache(source: str, product: str, *, force: bool = False) -> dict:
    global _LONG_PIVOT_JOB_THREAD
    src = _long_pivot_source(source)
    prod = _long_pivot_product(product)
    target = _long_pivot_key(src, prod)
    with _LONG_PIVOT_JOB_LOCK:
        current = str(_LONG_PIVOT_JOB_STATE.get("current") or "")
        queued = {_long_pivot_key(q_src, q_prod) for q_src, q_prod, _force in _LONG_PIVOT_QUEUE}
        if target != current and target not in queued:
            _LONG_PIVOT_QUEUE.append((src, prod, bool(force)))
        _LONG_PIVOT_JOB_STATE["queued"] = bool(_LONG_PIVOT_QUEUE)
        if _LONG_PIVOT_JOB_THREAD is None or not _LONG_PIVOT_JOB_THREAD.is_alive():
            _LONG_PIVOT_JOB_THREAD = threading.Thread(target=_long_pivot_worker_loop, name="splittable-long-pivot-cache", daemon=True)
            _LONG_PIVOT_JOB_THREAD.start()
        state = dict(_LONG_PIVOT_JOB_STATE)
    status = "running" if state.get("running") and state.get("current") == target else "queued"
    return {"ok": True, "queued": True, "status": status, "job": state}


def _long_pivot_inline_resource_guard() -> tuple[str, dict]:
    try:
        from core import runtime_limits
        if runtime_limits.process_memory_high():
            return "process_memory_high", runtime_limits.process_memory_snapshot()
        cpu = runtime_limits.process_cpu_snapshot()
        if bool(cpu.get("process_cpu_over_limit")):
            return "process_cpu_high", cpu
    except Exception:
        return "", {}
    return "", {}


# FAB/INLINE/ET datalake 진단 엔드포인트.
#   FAB 는 wafer 단위 공정이력이고, INLINE/ET 는 item/value 계측 long format 이다.
#   FAB preview 는 canonical 공정이력 컬럼을 보여주고, INLINE/ET 는 wide pivot sample 을 보여준다.
@router.get("/long-items")
def long_items(source: str = Query(..., description="fab|inline|et"),
               product: str = Query(..., description="PRODA 등 (ML_TABLE_ prefix 없이)")):
    """INLINE/ET item_id 레지스트리. FAB 는 공정이력이라 item_id 목록이 없을 수 있다."""
    from core.long_pivot import scan_long_fab, scan_long_inline, scan_long_et, list_items
    prod = product.replace("ML_TABLE_", "").strip()
    db_root = _db_base()
    lf = None
    if source == "fab":
        lf = scan_long_fab(prod, db_root)
    elif source == "inline":
        lf = scan_long_inline(prod, db_root)
    elif source == "et":
        lf = scan_long_et(prod, db_root)
    else:
        raise HTTPException(400, "source must be fab|inline|et")
    if lf is None:
        return {"source": source, "product": prod, "items": [],
                "note": f"hive 경로가 없음: {db_root} 에 1.RAWDATA_DB_{source.upper()}/{prod}/ 확인"}
    items = list_items(lf)
    note = "FAB 는 wafer 단위 공정이력이라 item_id 레지스트리가 비어 있을 수 있습니다." if source == "fab" and not items else ""
    return {"source": source, "product": prod, "items": items, "note": note}


@router.get("/long-wide-preview")
def long_wide_preview(source: str = Query(..., description="fab|inline|et"),
                      product: str = Query(...),
                      limit: int = Query(20)):
    """FAB 공정이력 또는 INLINE/ET pivot 결과 상위 N 행 미리보기."""
    src = _long_pivot_source(source)
    prod = _long_pivot_product(product)
    try:
        limit = max(1, min(500, int(limit or 20)))
    except Exception:
        limit = 20
    status = _long_pivot_cache_status(src, prod)
    if status.get("status") == "fresh":
        wide = pl.scan_parquet(status["cache_path"]).head(limit).collect()
        return {
            "source": src,
            "product": prod,
            "columns": wide.columns,
            "rows": wide.to_dicts(),
            "total_preview": wide.height,
            "pivot_cache": _long_pivot_cache_public(status),
        }
    guard_reason, guard_snapshot = _long_pivot_inline_resource_guard()
    if guard_reason:
        queued = enqueue_long_pivot_cache(src, prod, force=False) if status.get("source_exists") else {}
        return {
            "source": src,
            "product": prod,
            "columns": [],
            "rows": [],
            "total_preview": 0,
            "note": "Pivot cache is preparing in the background.",
            "pivot_cache": _long_pivot_cache_public(status, queued),
            "resource_guard": {"reason": guard_reason, **guard_snapshot},
        }
    lf = _scan_long_pivot_source(src, prod)
    if lf is None:
        return {
            "source": src,
            "product": prod,
            "rows": [],
            "columns": [],
            "note": "원천 hive 경로 미존재",
            "pivot_cache": _long_pivot_cache_public(_long_pivot_cache_status(src, prod)),
        }
    queued = enqueue_long_pivot_cache(src, prod, force=False)
    pivot = _long_pivot_function(src)
    wide = None
    try:
        wide = pivot(lf)
        preview = wide.head(limit)
        return {
            "source": src,
            "product": prod,
            "columns": preview.columns,
            "rows": preview.to_dicts(),
            "total_preview": preview.height,
            "note": "Pivot cache is preparing in the background.",
            "pivot_cache": _long_pivot_cache_public(status, queued),
        }
    finally:
        if wide is not None:
            try:
                del wide
            except Exception:
                pass
        try:
            gc.collect()
        except Exception:
            pass


@router.get("/features", deprecated=True)
def get_features_deprecated(rows: int = Query(50), cols: int = Query(40)):
    """v8.4.3 deprecated — ET+INLINE join 기반 features 는 ML_TABLE_PROD* 로 통합.
    임시로 빈 응답 유지 (기존 프론트 호환). 다음 frontend 릴리즈에서 호출 제거.
    """
    return {
        "join": "deprecated",
        "join_keys": [],
        "total_rows": 0, "total_cols": 0,
        "columns": [], "all_columns": [], "dtypes": {}, "sample": [],
        "deprecated": True,
        "replacement": "Use /api/splittable/view with product=ML_TABLE_PRODA|ML_TABLE_PRODB",
    }


def _get_features_legacy_stub(rows: int = 50, cols: int = 40):
    """Return the wide feature table from ET ⋈ INLINE (ET left join).

    Query params:
      - rows: sample rows to serialize (default 50, max 500)
      - cols: sample columns to serialize (default 40, max 200).
              `all_columns` is always full schema regardless of cols trim.

    Response shape (short):
      {
        "join": "et_left_inline",
        "join_keys": ["lot_id","wafer_id","product"],
        "total_rows": <int>,
        "total_cols": <int>,
        "et_rows":  <int>, "et_cols":  <int>,
        "inline_rows": <int>, "inline_cols": <int>,
        "columns":  [<first `cols` column names>],
        "all_columns": [<full list>],
        "dtypes":   {name: dtype_str, ...},
        "sample":   [ {col: val, ...}, ... ]   # first `rows` rows
      }
    """
    rows = max(1, min(500, int(rows)))
    cols = max(1, min(200, int(cols)))

    et, inl = _read_et_and_inline()
    joined = _join_features(et, inl)

    all_cols = list(joined.columns)
    schema = {n: str(d) for n, d in joined.schema.items()}
    show_cols = all_cols[:cols]
    sample = joined.head(rows).select(show_cols)

    # polars → JSON-safe rows (None passes through)
    data = sample.to_dicts()
    # Cast any non-JSON-friendly scalars to str as a defensive measure
    for r in data:
        for k, v in list(r.items()):
            if v is None or isinstance(v, (int, float, str, bool)):
                continue
            r[k] = str(v)

    return {
        "join": "et_left_inline",
        "join_keys": [k for k in _JOIN_KEYS if k in et.columns and k in inl.columns],
        "total_rows": joined.height,
        "total_cols": len(all_cols),
        "et_rows": et.height,
        "et_cols": et.width,
        "inline_rows": inl.height,
        "inline_cols": inl.width,
        "columns": show_cols,
        "all_columns": all_cols,
        "dtypes": schema,
        "sample": data,
        "base_root": str(_base_root()),
    }


@router.get("/uniques")
def get_uniques():
    """Proxy `<db_root>/_uniques.json` verbatim for feature-select catalogs.

    Returns the parsed JSON body + a small meta header. If the file is missing
    we return `{"uniques": {}, "exists": False, ...}` rather than 404 so the
    frontend can display a graceful empty state.
    """
    base = _base_root()
    fp = base / _UNIQUES_FILE
    if not fp.is_file():
        return {
            "exists": False,
            "path": str(fp),
            "uniques": {},
            "size": 0,
        }
    try:
        with open(fp, "r", encoding="utf-8") as f:
            parsed = json.load(f)
    except Exception as e:
        raise HTTPException(500, f"_uniques.json parse error: {e}")
    return {
        "exists": True,
        "path": str(fp),
        "size": fp.stat().st_size,
        "top_keys": list(parsed.keys()) if isinstance(parsed, dict) else None,
        "uniques": parsed,
    }


# ── Source visibility config (admin) ──
SOURCE_CFG = PLAN_DIR / "source_config.json"


# ── 쿼리 병렬 워커 수 관리 ──────────────────────────────────────────
# 운영 SplitTable Polars 풀은 기본 4코어, 1~4 범위에서 저장 가능하다.
# 개발은 1코어 고정이다. DuckDB 전역 설정은 건드리지 않아 파일탐색기 SQL과
# Flow-i가 SplitTable 튜닝에 함께 제한되지 않게 한다. Polars 풀은 시작 시 고정돼
# 운영값 변경은 서버 재시작 후 적용된다.

def _normalize_query_workers(value: Any) -> int:
    """운영 설정을 1~4 및 실제 CPU 범위로 클램프. 개발은 항상 1."""
    if _ml_table_lookup._root_ram_cache_use_dev():
        return 1
    try:
        n = int(value)
    except Exception:
        n = 4
    if n <= 0:
        n = 4
    from core.runtime_limits import effective_cpu_count
    max_cpu = max(1, min(4, int(effective_cpu_count())))
    return max(1, min(n, max_cpu))


def _apply_query_workers(workers: int) -> None:
    """Polars 풀은 런타임 재크기 불가. 저장만 하고 재시작 적용임을 명시한다.

    과거처럼 FLOW_DUCKDB_THREADS 를 바꾸지 않는다. 그 값은 파일탐색기 SQL 등
    SplitTable 밖의 DuckDB 사용자까지 제한해 개발 서버 우선순위와 충돌했다.
    """
    return None


def _query_workers_key() -> str:
    """운영 저장 키. 개발은 파일에 값을 쓰지 않는 고정 정책이다."""
    return "fixed_dev_1" if _ml_table_lookup._root_ram_cache_use_dev() else "query_workers"


def _current_query_workers() -> int:
    """운영 저장값(기본 4). 개발 서버는 저장 내용과 무관하게 1."""
    if _ml_table_lookup._root_ram_cache_use_dev():
        return 1
    try:
        cfg = load_json(SOURCE_CFG, {})
        return _normalize_query_workers(cfg.get("query_workers", 4))
    except Exception:
        return _normalize_query_workers(4)


def _current_query_workers_status() -> dict:
    """현재 쿼리 병렬화 상태를 반환."""
    from core.runtime_limits import effective_cpu_count
    configured = _current_query_workers()
    cpu_count = int(effective_cpu_count())
    is_dev = _ml_table_lookup._root_ram_cache_use_dev()
    desired = 1 if is_dev else configured
    try:
        effective = max(1, int(os.environ.get("POLARS_MAX_THREADS", desired)))
    except Exception:
        effective = desired
    # essential 세마포어 동시성 — 동시에 몇 명이 조회할 수 있는지
    essential_concurrency = int(os.environ.get(
        "FLOW_ESSENTIAL_REQUEST_CONCURRENCY", "0")) or None
    return {
        "configured": configured,
        "effective": effective,
        "auto_value": 1 if is_dev else min(4, max(1, cpu_count)),
        "cpu_count": cpu_count,
        "cpu_budget": desired,
        # 이 서버가 개발(dev) 몫인지와 저장에 쓰는 키 — 화면이 "지금 어느 서버를
        # 튜닝 중인지"를 명시해, 개발 값을 바꾼다는 게 운영에 영향 없음을 보인다.
        "is_dev": is_dev,
        "config_key": _query_workers_key(),
        # Polars 는 시작 시 역할 기반(api=4/worker=3/standalone=auto)으로 1회 고정 —
        # query_workers 변경에 영향받지 않고 재시작해야 바뀐다.
        "polars_threads": os.environ.get("POLARS_MAX_THREADS", ""),
        "polars_runtime_fixed": True,
        "fixed": is_dev,
        "restart_required": effective != desired,
        "desired": desired,
        "duckdb_threads": os.environ.get("FLOW_DUCKDB_THREADS", ""),
        # 예열 워커는 검색 부하와 분리해 1코어 고정(env 로만 조정).
        "ram_cache_workers": _ml_table_lookup._root_ram_cache_load_workers(),
        "essential_concurrency": essential_concurrency,
    }


@router.get("/query-workers")
def get_query_workers():
    """현재 쿼리 병렬 워커 설정과 상태를 반환."""
    return _current_query_workers_status()


@router.post("/query-workers/save")
def save_query_workers(req: dict, _perm=Depends(require_page_manager("splittable"))):
    """쿼리 병렬 워커 수를 저장하고 즉시 반영.

    운영 서버만 1~4 값을 저장한다. 개발 서버는 1코어 고정이며 운영 설정을
    건드리지 않는다. Polars 풀 특성상 운영 저장값은 재시작 후 적용된다."""
    if _ml_table_lookup._root_ram_cache_use_dev():
        # 개발은 정책상 1코어 고정. 공유 설정 파일의 운영값을 건드리지 않는다.
        return _current_query_workers_status()
    raw = req.get("query_workers", 4)
    workers = _normalize_query_workers(raw)
    cur = load_json(SOURCE_CFG, {"enabled": [], "lot_overrides": {}})
    cur[_query_workers_key()] = workers
    save_json(SOURCE_CFG, cur)
    return _current_query_workers_status()


# 실제 Polars 크기는 router import 전 core.runtime_limits 에서 적용된다.

@router.get("/source-config")
def get_source_config():
    cfg = load_json(SOURCE_CFG, {"enabled": []})
    cfg.setdefault("enabled", [])
    cfg.setdefault("lot_overrides", {})  # v8.4.4: product-scoped {root_col, fab_col, fab_source, ts_col, join_keys}
    cfg.setdefault("root_lot_cache", _ml_table_lookup.root_ram_cache_settings())
    # 지정 팀 수신자 폐기 — 수신 대상은 계획 작성자 + 제품 동명 그룹 멤버로 고정.
    cfg.pop("mismatch_alert_recipients", None)
    cfg.setdefault("mismatch_mail_enabled", False)  # plan/actual 불일치 메일 발송 (기본 off)
    cfg.setdefault("query_workers", 4)  # 운영 기본 4, 개발은 별도 정책으로 1 고정
    # v8.8.21: 응답 단에서도 root:~~ 남은 값은 표시 안 되게 정리.
    _migrate_legacy_root_prefix(cfg)
    return cfg

class SourceConfigReq(BaseModel):
    enabled: List[str] = []
    lot_overrides: dict = {}  # v8.4.4
    root_lot_cache: dict | None = None
    # plan/actual 불일치 알람을 메일로도 발송할지 여부 (기본 off).
    mismatch_mail_enabled: bool | None = None
    # 쿼리 병렬 워커 수 (Polars/DuckDB/RAM 캐시 예열). 0 = 자동(호스트 CPU 비례).
    query_workers: int | None = None


def _normalize_fab_source_path(v: str) -> str:
    s = str(v or "").strip().replace("\\", "/")
    if not s:
        return ""
    while s.startswith("./"):
        s = s[2:]
    if s.lower().startswith("db/"):
        s = s[3:]
    elif s.lower().startswith("base/"):
        s = s[5:]
    if s.startswith("/"):
        s = s.lstrip("/")
    return s

def _migrate_legacy_root_prefix(cfg: dict) -> dict:
    """Normalize stored fab_source values to db-relative paths."""
    try:
        lo = cfg.get("lot_overrides") or {}
        for _p, _ov in list(lo.items()):
            if not isinstance(_ov, dict):
                continue
            fs = str(_ov.get("fab_source") or "").strip()
            if fs.startswith("root:"):
                _ov["fab_source"] = ""
            else:
                _ov["fab_source"] = _normalize_fab_source_path(fs)
    except Exception:
        pass
    return cfg


def _normalize_root_lot_cache_settings(raw: dict | None, prev: dict | None = None) -> dict:
    data = raw if isinstance(raw, dict) else {}
    step_raw = data.get("step_ids")
    if isinstance(step_raw, str):
        step_parts = step_raw.replace("\n", ",").split(",")
    elif isinstance(step_raw, (list, tuple, set)):
        step_parts = list(step_raw)
    else:
        step_parts = []
    step_ids = []
    seen = set()
    for item in step_parts:
        step = str(item or "").strip().upper()
        if not step or step in seen:
            continue
        seen.add(step)
        step_ids.append(step)

    def _num(key: str, default: int) -> int:
        try:
            value = int(data.get(key))
        except Exception:
            value = default
        return max(0, min(ROOT_LOT_CACHE_LIMIT_MAX, value))

    out = {
        "step_ids": step_ids,
        "searched_limit": _num("searched_limit", 1000),
        "target_roots": _num("target_roots", 1000),
    }
    # AZ prefix 우선 적재 설정 — 요청에 없으면 기존 저장값 보존(전체 교체 시 소실 방지).
    prefix_raw = data.get("priority_root_prefix")
    if prefix_raw is None and isinstance(prev, dict):
        prefix_raw = prev.get("priority_root_prefix")
    if prefix_raw is not None:
        out["priority_root_prefix"] = str(prefix_raw or "").strip().upper()
    return out


@router.post("/source-config/save")
def save_source_config(req: SourceConfigReq, _perm=Depends(require_page_manager("splittable"))):
    cur = load_json(SOURCE_CFG, {"enabled": [], "lot_overrides": {}})
    cur["enabled"] = req.enabled
    if req.lot_overrides:
        cur.setdefault("lot_overrides", {}).update(req.lot_overrides)
    if req.root_lot_cache is not None:
        cur["root_lot_cache"] = _normalize_root_lot_cache_settings(req.root_lot_cache, cur.get("root_lot_cache"))
    # 지정 팀 수신자 폐기 — 저장 파일에서도 제거.
    cur.pop("mismatch_alert_recipients", None)
    if req.mismatch_mail_enabled is not None:
        cur["mismatch_mail_enabled"] = bool(req.mismatch_mail_enabled)
    if req.query_workers is not None and not _ml_table_lookup._root_ram_cache_use_dev():
        cur["query_workers"] = _normalize_query_workers(req.query_workers)
    # v8.8.21: legacy root:~~ 삭제.
    _migrate_legacy_root_prefix(cur)
    save_json(SOURCE_CFG, cur)
    return {"ok": True}


# ── Prefixes ──
@router.get("/prefixes")
def get_prefixes():
    return {"prefixes": _load_prefixes()}


# ── KNOB metadata (v8.4.7) ───────────────────────────────────────────
# Reverse-lookup helper used by SplitTable UI:
#   ppid_knob.csv:      feature_name, rule_order, step_desc, operator, value, category
#                        value = SplitTable cell value such as PPID_01_2
#   Vehicle_matching.csv: product, step_id, step_desc (preferred)
#   step_matching.csv:    product, step_id, function_step (legacy fallback)
# For each KNOB feature_name, we keep product-common ppid_knob CSV rule rows in
# rule_order, expand each step_desc through the current product's matching
# step_ids, and produce both a structured `groups` payload and a label:
#   GATE_PATTERN (AA200030/AA200040/AA200050) + PC_ETCH (AA200100/AA200110)
def _load_csv_rows(fp: Path) -> list[dict]:
    if not fp.is_file():
        return []
    try:
        st = fp.stat()
        key = str(fp.resolve())
        cached = _CSV_ROWS_CACHE.get(key)
        if cached and cached[0] == st.st_mtime and cached[1] == st.st_size:
            return [dict(row) for row in cached[2]]
        with open(fp, "r", encoding="utf-8-sig") as f:
            reader = csv_mod.DictReader(f)
            rows = []
            for row in reader:
                clean = {}
                for col_key, value in (row or {}).items():
                    if col_key is None:
                        continue
                    clean[str(col_key).lstrip("\ufeff").strip()] = value
                rows.append(clean)
        # 무한 성장 방지 — 파일 단위 키가 계속 늘 수 있어 오래된 것부터 정리.
        while len(_CSV_ROWS_CACHE) >= _CSV_ROWS_CACHE_MAX:
            _CSV_ROWS_CACHE.pop(next(iter(_CSV_ROWS_CACHE)), None)
        _CSV_ROWS_CACHE[key] = (st.st_mtime, st.st_size, [dict(row) for row in rows])
        return rows
    except Exception:
        return []


def _canonical_product_name(product: str) -> str:
    raw = str(product or "").strip()
    return _split_product_core(raw) or raw


def _mltable_schema_columns(product: str, prefix: str = "") -> list[str]:
    core = _canonical_product_name(product)
    if not core:
        return []
    names = [f"ML_TABLE_{core}.parquet"]
    for alias in sorted(_product_aliases(core)):
        if alias.startswith("ML_TABLE_"):
            names.append(f"{alias}.parquet")
        else:
            names.append(f"ML_TABLE_{alias}.parquet")
    seen_names = []
    for name in names:
        if name not in seen_names:
            seen_names.append(name)
    pref = str(prefix or "").strip().upper()
    for name in seen_names:
        fp = _base_root() / name
        if not fp.is_file():
            continue
        try:
            st = fp.stat()
            key = str(fp.resolve())
            cached = _SCHEMA_COLUMNS_CACHE.get(key)
            if cached and cached[0] == st.st_mtime and cached[1] == st.st_size:
                cols = list(cached[2])
            else:
                cols = _scan_parquet_compat(str(fp)).collect_schema().names()
                while len(_SCHEMA_COLUMNS_CACHE) >= _SCHEMA_COLUMNS_CACHE_MAX:
                    _SCHEMA_COLUMNS_CACHE.pop(next(iter(_SCHEMA_COLUMNS_CACHE)), None)
                _SCHEMA_COLUMNS_CACHE[key] = (st.st_mtime, st.st_size, list(cols))
        except Exception:
            continue
        if pref:
            return [c for c in cols if str(c).upper().startswith(pref + "_")]
        return list(cols)
    return []


def _stage_major(text: str):
    tail = str(text or "").strip()
    if "_" in tail and tail.split("_", 1)[0].upper() in {"KNOB", "INLINE", "VM"}:
        tail = tail.split("_", 1)[1].strip()
    m = _re.match(r"^\s*(\d+(?:\.\d+)?)", tail)
    if not m:
        return None
    try:
        return int(float(m.group(1)))
    except Exception:
        return None


def _dedup_list(values: list[str]) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    for value in values or []:
        s = str(value or "").strip()
        if not s or s in seen:
            continue
        seen.add(s)
        out.append(s)
    return out


def _first_row_value(row: dict, *cols: str) -> str:
    for col in cols:
        if not col:
            continue
        value = str((row or {}).get(col) or "").strip()
        if value:
            return value
    return ""


def _row_step_desc(row: dict, schema: dict) -> str:
    return _first_row_value(
        row,
        schema.get("step_desc_col", "step_desc"),
        schema.get("func_step_col", "function_step"),
        "step_desc",
        "function_step",
        "func_step",
    )


def _knob_step_matching_path(base: Path | None = None) -> Path:
    return _rulebook_path_for_base("step_matching", base)


def _load_knob_step_matching_rows(base: Path | None = None) -> list[dict]:
    return _load_csv_rows(_knob_step_matching_path(base))


# module 의 유일한 원천은 Vehicle_matching.csv 의 module 열이다.
# step_desc 로 공정 구간을 추측하던 폴백(classify_process_area)은 제거했다 —
# 파일에 module 열이 없는 환경에서도 값이 만들어져 SplitTable 좌측 module 열이
# 근거 없이 나타났다. 열이 없으면 빈 값이고, 그러면 프런트가 열 자체를 안 붙인다.
def _vehicle_module_of(row: dict, sm: dict) -> str:
    for col in (sm.get("module_col", "module"), "module", "area"):
        if not col:
            continue
        value = str(row.get(col) or "").strip()
        if value:
            return value
    return ""


def _product_step_map_by_desc(product: str, base: Path | None = None) -> dict[str, list[dict]]:
    matching = _load_knob_step_matching_rows(base)
    sm = _sch("step_matching")
    step_map: dict[str, list[dict]] = {}
    p_col = sm.get("product_col", "product")
    has_product_col = any(p_col in r or "product" in r for r in matching)
    for r in matching:
        row_prod = r.get(p_col)
        if row_prod is None and p_col != "product":
            row_prod = r.get("product")
        if not _step_matching_product_matches(product, row_prod, allow_common=not has_product_col):
            continue
        step_desc = _row_step_desc(r, sm)
        step_desc_key = _step_desc_match_key(step_desc)
        step_id = (r.get(sm.get("step_id_col", "step_id")) or r.get("raw_step_id") or "").strip()
        if not step_desc_key or not step_id:
            continue
        item = {
            "step_desc": step_desc,
            "step_id": step_id,
            "module": _vehicle_module_of(r, sm),
        }
        bucket = step_map.setdefault(step_desc_key, [])
        if not any(str(x.get("step_id") or "").strip().casefold() == step_id.casefold() for x in bucket):
            bucket.append(item)
    return step_map


def _stage_steps_by_major(product: str) -> dict[int, list[dict]]:
    matching = _load_knob_step_matching_rows()
    sm = _sch("step_matching")
    exact_has_numeric = False
    p_col = sm.get("product_col", "product")
    has_product_col = any(p_col in r or "product" in r for r in matching)
    for r in matching:
        row_prod = r.get(p_col)
        if row_prod is None and p_col != "product":
            row_prod = r.get("product")
        if not _step_matching_product_matches(product, row_prod, allow_common=not has_product_col):
            continue
        if _stage_major(_row_step_desc(r, sm)) is not None:
            exact_has_numeric = True
            break
    out: dict[int, list[dict]] = {}
    seen: dict[int, set[tuple[str, str]]] = {}
    for r in matching:
        row_prod = r.get(p_col)
        if row_prod is None and p_col != "product":
            row_prod = r.get("product")
        row_prod = str(row_prod or "").strip()
        if exact_has_numeric:
            if not _step_matching_product_matches(product, row_prod, allow_common=False):
                continue
        elif not _step_matching_product_matches(product, row_prod, allow_common=not has_product_col):
            continue
        fs = _row_step_desc(r, sm)
        sid = (r.get(sm.get("step_id_col", "step_id")) or r.get("raw_step_id") or "").strip()
        major = _stage_major(fs)
        if major is None or not fs:
            continue
        item = {
            "func_step": fs,
            "step_id": sid,
            "module": _vehicle_module_of(r, sm),
            # 표시용 module 과 달리 이건 **매칭 전용 추측값**이다. step_desc 로 어느
            # 공정 구간인지 짐작해 _stage_steps_for_tail 의 module 우선 매칭 단계를
            # 살려둔다(M1 이 M2_OVL_M1 에 붙는 걸 막는 tier). 화면의 module 열은
            # 이 값을 절대 쓰지 않는다 — Vehicle_matching 의 module 열만 근거다.
            "area_guess": classify_process_area(fs) or "",
            "step_class": str(r.get("step_class") or "").strip(),
        }
        key = (item["func_step"], item["step_id"])
        bucket_seen = seen.setdefault(major, set())
        if key in bucket_seen:
            continue
        bucket_seen.add(key)
        out.setdefault(major, []).append(item)
    return out


def _stage_token(text: str) -> str:
    tail = str(text or "").strip()
    if "_" in tail and tail.split("_", 1)[0].upper() in {"KNOB", "INLINE", "VM"}:
        tail = tail.split("_", 1)[1].strip()
    tail = _re.sub(r"^\s*\d+(?:\.\d+)?[A-Za-z]?\s*", "", tail).strip()
    return tail


def _norm_stage_text(text: str) -> str:
    return _re.sub(r"[^A-Z0-9]+", "", str(text or "").upper())


def _stage_aliases(token: str) -> list[str]:
    key = _norm_stage_text(token)
    aliases = {
        "WELL": ["WELL", "NWELL", "PWELL"],
        "VTN": ["VT", "VTN", "VTP", "WELL"],
        "GATEOX": ["GATEOX", "GATE_OX", "GATE", "HKMG"],
        "PC": ["PC", "POLYCONTACT", "GATE"],
        "SDEPI": ["SDEPI", "SD_EPI", "EPI"],
        "SILICIDE": ["SILICIDE", "SILI"],
        "CONTACT": ["CONTACT", "CT", "MOL"],
        "M0": ["MOL", "M0", "V0"],
        "VIA0": ["VIA0", "V0", "MOL"],
        "M1": ["BEOLM1", "M1"],
        "VIA1": ["VIA1", "BEOLM2", "M2"],
        "M2": ["BEOLM2", "M2"],
        "VIA2": ["VIA2", "BEOLM3", "M3"],
        "M3": ["BEOLM3", "M3"],
        "VIA3": ["VIA3", "BEOLM4", "M4"],
        "M4": ["BEOLM4", "M4"],
        "PAD": ["PAD", "PASSIVATION"],
        "PASSIVATION": ["PASSIVATION", "PAD"],
        "ETESTPREP": ["ETEST", "ET", "SORT"],
        "RELIABILITY": ["RELIABILITY", "REL"],
        "SORT": ["SORT", "ET"],
    }
    raw = [key]
    raw.extend(aliases.get(key, []))
    return _dedup_list([_norm_stage_text(x) for x in raw])


def _stage_steps_for_tail(tail: str, steps_by_major: dict[int, list[dict]]) -> list[dict]:
    token = _stage_token(tail)
    aliases = [a for a in _stage_aliases(token) if a]
    all_steps = [item for bucket in steps_by_major.values() for item in bucket]
    def _collect(match_fn):
        hits: list[dict] = []
        seen: set[tuple[str, str]] = set()
        for item in all_steps:
            if not match_fn(item):
                continue
            key = (item.get("func_step", ""), item.get("step_id", ""))
            if key in seen:
                continue
            seen.add(key)
            hits.append(item)
        return hits

    major = _stage_major(tail)
    stage_hits = _collect(lambda item:
        str(item.get("step_class") or "").strip().lower() == "stage"
        and _stage_major(item.get("func_step", "")) == major
        and _norm_stage_text(_stage_token(item.get("func_step", ""))) == _norm_stage_text(token)
    )
    if stage_hits:
        return stage_hits

    # Prefer module-level matches. This keeps e.g. M1 from matching M2_OVL_M1.
    # Vehicle_matching 의 module 열이 없으면 step_desc 추측값(area_guess)으로 본다 —
    # 이 tier 는 매칭 품질용이라 열 유무와 무관하게 유지되어야 한다.
    def _item_area(item: dict) -> str:
        return _norm_stage_text(item.get("module", "") or item.get("area_guess", ""))

    module_hits = _collect(lambda item: any(
        alias == _item_area(item) or (alias and alias in _item_area(item))
        for alias in aliases
    ))
    if module_hits:
        return module_hits

    def _func_match(item):
        body = _norm_stage_text(_stage_token(item.get("func_step", "")))
        return any(alias and body.startswith(alias) for alias in aliases)

    func_hits = _collect(_func_match)
    if func_hits:
        return func_hits
    if major is not None and major <= 8:
        return list(steps_by_major.get(major, []))
    return []


def _inferred_stage_meta(product: str, prefix: str) -> dict[str, dict]:
    pref = str(prefix or "").strip().upper()
    cols = _mltable_schema_columns(product, pref)
    if not cols:
        return {}
    steps_by_major = _stage_steps_by_major(product)
    out: dict[str, dict] = {}
    for full in cols:
        _, _, tail = str(full).partition("_")
        tail = tail.strip()
        if not tail:
            continue
        major = _stage_major(tail)
        steps = _stage_steps_for_tail(tail, steps_by_major)
        step_ids = _dedup_list([x.get("step_id", "") for x in steps])
        function_steps = _dedup_list([x.get("func_step", "") for x in steps])
        # INLINE 은 module 로 묶지 않는다 (_build_inline_meta 와 같은 규칙).
        modules = [] if pref == "INLINE" else _dedup_list([x.get("module", "") for x in steps])
        if pref == "KNOB":
            groups = [{
                "func_step": tail,
                "rule_order": major or 0,
                "ppid": "",
                "operator": "",
                "category": modules[0] if len(modules) == 1 else "",
                "step_ids": step_ids,
                "modules": modules,
                "module": modules[0] if len(modules) == 1 else "",
                "inferred": True,
            }]
            meta = {
                "groups": groups,
                "label": f"{tail} ({'/'.join(step_ids)})" if step_ids else tail,
                "modules": modules,
                "inferred": True,
            }
        else:
            group = {
                "function_step": tail,
                "step_id": step_ids[0] if len(step_ids) == 1 else "",
                "step_ids": step_ids,
                "function_steps": function_steps,
                "modules": modules,
                "module": modules[0] if len(modules) == 1 else "",
                "inferred": True,
            }
            if pref == "INLINE":
                group.update({"item_id": tail, "item_desc": tail})
                meta = {
                    "item_id": tail,
                    "item_desc": tail,
                    "step_id": step_ids[0] if len(step_ids) == 1 else "",
                    "step_ids": step_ids,
                    "function_step": tail,
                    "function_steps": function_steps,
                    "groups": [group],
                    "label": tail,
                    "sub": "/".join(step_ids) if step_ids else tail,
                    "inferred": True,
                }
            else:
                group.update({"feature_name": tail, "step_desc": tail})
                meta = {
                    "step_desc": tail,
                    "step_id": step_ids[0] if len(step_ids) == 1 else "",
                    "step_ids": step_ids,
                    "function_step": tail,
                    "function_steps": function_steps,
                    "groups": [group],
                    "label": tail,
                    "sub": "/".join(step_ids) if step_ids else tail,
                    "inferred": True,
                }
        out.setdefault(tail, meta)
        out.setdefault(str(full), meta)
    return out


def _build_knob_meta(product: str = "") -> dict:
    base = _base_root()
    ppid_knob_fp = base / "ppid_knob.csv"
    knob_rules = _load_csv_rows(ppid_knob_fp if ppid_knob_fp.is_file() else base / "knob_ppid.csv")
    # v8.8.10: 역할→컬럼명 매핑 soft-landing. 사내 CSV 의 컬럼 이름이 달라도 schema 만 바꾸면 됨.
    km = _sch("knob_ppid")

    # step_desc → [{step_id,module}, ...] (ordered, dedup)
    step_map = _product_step_map_by_desc(product, base)

    # feature_name → CSV rule row groups (sorted by rule_order)
    feats: dict[str, list[dict]] = {}
    for r in knob_rules:
        # ppid_knob.csv is product-common. Legacy product columns are ignored;
        # product scoping belongs only to Vehicle_matching.csv.
        fname = (r.get(km.get("feature_col", "feature_name")) or "").strip()
        step_desc = _row_step_desc(r, km)
        step_desc_key = _step_desc_match_key(step_desc)
        value = _first_row_value(
            r,
            km.get("value_col", "value"),
            km.get("ppid_col", "ppid"),
            "value",
            "ppid",
            "category",
        )
        if not fname or not step_desc_key:
            continue
        order_label = _rule_order_label(r.get(km.get("rule_order_col", "rule_order")), len(feats.get(fname, [])) + 1)
        feats.setdefault(fname, []).append({
            "func_step": step_desc,
            "step_desc": step_desc,
            "rule_order": order_label,
            "rule_order_sort": _rule_order_sort_key(order_label),
            "ppid": value,
            "value": value,
            "operator": (r.get(km.get("operator_col", "operator")) or "").strip(),
            "category": (r.get(km.get("category_col", "category")) or "").strip(),
            "step_ids": [str(x.get("step_id") or "").strip() for x in step_map.get(step_desc_key, []) if str(x.get("step_id") or "").strip()],
            "modules": [str(x.get("module") or "").strip() for x in step_map.get(step_desc_key, []) if str(x.get("module") or "").strip()],
        })

    # Sort each feature's groups by rule_order + build a human label
    out: dict[str, dict] = {}
    for fname, groups in feats.items():
        groups.sort(key=lambda g: g.get("rule_order_sort") or _rule_order_sort_key(g.get("rule_order")))
        parts: list[str] = []
        feat_modules: list[str] = []
        for i, g in enumerate(groups):
            sids = g["step_ids"]
            mods = []
            for mod in (g.get("modules") or []):
                mod = str(mod or "").strip()
                if mod and mod not in mods:
                    mods.append(mod)
                if mod and mod not in feat_modules:
                    feat_modules.append(mod)
            g["module"] = mods[0] if len(mods) == 1 else ""
            g["modules"] = mods
            if len(sids) == 0:
                seg = g["step_desc"]
            elif len(sids) == 1:
                seg = f"{g['step_desc']} ({sids[0]})"
            else:
                seg = f"{g['step_desc']} ({'/'.join(sids)})"
            parts.append(seg)
            if i < len(groups) - 1:
                parts.append(" + ")
        out[fname] = {
            "groups": groups,
            "label": "".join(parts),
            "modules": feat_modules,
        }
        # SplitTable rows carry the physical ML_TABLE column name
        # (for example `KNOB_5.0 PC`), while ppid_knob.csv stores only the
        # feature tail (`5.0 PC`).  Expose both keys so row click / tooltip /
        # display-rename paths all resolve to the same explicit ppid_knob rule.
        out[f"KNOB_{fname}"] = out[fname]
    for key, meta in _inferred_stage_meta(product, "KNOB").items():
        out.setdefault(key, meta)
    return out


# ── SplitTable 공정 순서(step order) 컨텍스트 ──
#   ① 표시 순서: KNOB 행을 feature 번호(00.0) 자연 정렬이 아니라 ppid_knob.csv 에
#      split 별로 지정된 step_desc → Vehicle_matching.csv step_id 의 공정 순서
#      (파일 행 순서)로 정렬한다. 한 step_desc 가 여러 step_id 로 확장되면
#      마지막 step_id 를 대표로 쓴다. 매핑이 없는 행은 기존 자연 정렬로 뒤에 붙는다.
#   ② 진행 셰이딩: 검색 root 의 latest-lot 캐시 step_id 보다 뒤(미진행) 공정에
#      해당하는 행 목록을 payload.step_progress 로 내려 FE 가 회색으로 칠한다.
_STEP_ORDER_CTX_CACHE: OrderedDict = OrderedDict()
_STEP_ORDER_CTX_TTL_SEC = 60.0
_STEP_ORDER_CTX_MAX = 32
_STEP_ORDER_CTX_LOCK = threading.Lock()
_STEP_ID_PREFIX_NUM_RE = _re.compile(r"^([A-Za-z]+)(\d+)")


def _split_step_order_context(product: str) -> dict:
    """product 의 공정 순서 컨텍스트.

    seq_rank:     step_id(upper) → Vehicle_matching 행 순서 rank
    prefix_steps: 영문 프리픽스 → [(step 번호, rank)] (근사 매칭용, 번호 오름차순)
    param_rank:   KNOB feature 명(물리 컬럼명/bare 명, upper) → rank
    param_step:   같은 키 → 대표 step_id
    """
    key = str(product or "").strip().upper()
    now = time.monotonic()
    with _STEP_ORDER_CTX_LOCK:
        hit = _STEP_ORDER_CTX_CACHE.get(key)
        if hit and now - hit[0] <= _STEP_ORDER_CTX_TTL_SEC:
            return hit[1]
    ctx = {"seq_rank": {}, "prefix_steps": {}, "param_rank": {}, "param_step": {}}
    try:
        matching = _load_knob_step_matching_rows()
        sm = _sch("step_matching")
        p_col = sm.get("product_col", "product")
        has_product_col = any(p_col in r or "product" in r for r in matching)
        seq_rank: dict[str, int] = {}
        prefix_steps: dict[str, list[tuple[int, int]]] = {}
        for r in matching:
            row_prod = r.get(p_col)
            if row_prod is None and p_col != "product":
                row_prod = r.get("product")
            if not _step_matching_product_matches(product, row_prod, allow_common=not has_product_col):
                continue
            sid = str(r.get(sm.get("step_id_col", "step_id")) or r.get("raw_step_id") or "").strip().upper()
            if not sid or sid in seq_rank:
                continue
            rank = len(seq_rank)
            seq_rank[sid] = rank
            m = _STEP_ID_PREFIX_NUM_RE.match(sid)
            if m:
                prefix_steps.setdefault(m.group(1).upper(), []).append((int(m.group(2)), rank))
        for steps in prefix_steps.values():
            steps.sort()
        param_rank: dict[str, int] = {}
        param_step: dict[str, str] = {}
        for fname, meta in (_build_knob_meta(product) or {}).items():
            groups = meta.get("groups") if isinstance(meta, dict) else None
            if not groups:
                continue
            sid = ""
            for g in groups:
                sids = [str(s or "").strip() for s in (g.get("step_ids") or []) if str(s or "").strip()]
                if sids:
                    # 한 step_desc 가 복수 step_id 로 확장되면 마지막(뒤의) step_id 를 대표로.
                    sid = sids[-1].upper()
                    break
            rank = seq_rank.get(sid) if sid else None
            if rank is None:
                continue
            param_rank[str(fname).strip().upper()] = rank
            param_step[str(fname).strip().upper()] = sid
        ctx = {"seq_rank": seq_rank, "prefix_steps": prefix_steps,
               "param_rank": param_rank, "param_step": param_step}
    except Exception as e:
        logger.warning("split step-order context 실패 (product=%s): %s", product, e)
    with _STEP_ORDER_CTX_LOCK:
        _STEP_ORDER_CTX_CACHE[key] = (now, ctx)
        while len(_STEP_ORDER_CTX_CACHE) > _STEP_ORDER_CTX_MAX:
            _STEP_ORDER_CTX_CACHE.popitem(last=False)
    return ctx


def _step_order_sort_key(col_name: str, display: str, param_rank: dict):
    """공정 순서 rank 우선, 같은 rank/매핑 없음은 기존 자연 정렬 유지.

    TAG_* 는 엔지니어가 직접 만든 주석 열이라 공정 순서와 무관하다. 표 맨 위에
    고정해 어떤 prefix 를 같이 보든 항상 먼저 눈에 들어오게 한다 (TAG 끼리는
    기존 자연 정렬 그대로).
    """
    tag_pfx = f"{CUSTOM_TAG_PREFIX}_"
    if (str(col_name or "").strip().upper().startswith(tag_pfx)
            or str(display or "").strip().upper().startswith(tag_pfx)):
        return (-1, 0, _natural_param_key(display or col_name))
    rank = param_rank.get(str(col_name or "").strip().upper())
    if rank is None:
        rank = param_rank.get(str(display or "").strip().upper())
    nat = _natural_param_key(display or col_name)
    if rank is None:
        return (1, 0, nat)
    return (0, rank, nat)


def _step_rank_for_progress(ctx: dict, step_id: str):
    """현재 진행 step_id 의 공정 순서 rank. 정확 매칭이 없으면(중간 step 등)
    같은 영문 프리픽스에서 step 번호가 작거나 같은 main step 중 가장 뒤의 rank."""
    sid = str(step_id or "").strip().upper()
    if not sid:
        return None
    rank = (ctx.get("seq_rank") or {}).get(sid)
    if rank is not None:
        return rank
    m = _STEP_ID_PREFIX_NUM_RE.match(sid)
    if not m:
        return None
    best = None
    for num, cand_rank in (ctx.get("prefix_steps") or {}).get(m.group(1).upper(), []):
        if num <= int(m.group(2)):
            best = cand_rank
        else:
            break
    return best


_ROOT_LATEST_STEP_CACHE: OrderedDict = OrderedDict()
_ROOT_LATEST_STEP_TTL_SEC = 180.0
_ROOT_LATEST_STEP_MAX = 64
_ROOT_LATEST_STEP_LOCK = threading.Lock()


def _root_latest_step_state(product: str, root_lot_id: str) -> dict:
    """Latest step for the root and independently for each physical wafer.

    per-root 파티션 우선, 없으면 monolithic 1회 필터 스캔(메모리 압박 시 생략).
    """
    root = str(root_lot_id or "").strip().upper()
    if not root:
        return {"step_id": "", "by_wafer": {}}
    key = (str(product or "").strip().upper(), root)
    now = time.monotonic()
    with _ROOT_LATEST_STEP_LOCK:
        hit = _ROOT_LATEST_STEP_CACHE.get(key)
        if hit and now - hit[0] <= _ROOT_LATEST_STEP_TTL_SEC:
            cached = hit[1]
            if isinstance(cached, dict):
                return {"step_id": str(cached.get("step_id") or ""),
                        "by_wafer": dict(cached.get("by_wafer") or {})}
            return {"step_id": str(cached or ""), "by_wafer": {}}
    state = {"step_id": "", "by_wafer": {}}
    try:
        lf = _latest_lot_index_partition_lf(product, root)
        streaming = False
        if lf is None:
            from core.runtime_limits import process_memory_high
            if not process_memory_high():
                lf = _latest_lot_step_cache_lf(product)
                streaming = lf is not None
        if lf is not None:
            names = lf.collect_schema().names()
            want = [c for c in ("root_lot_id", "wafer_id", "step_id", "tkout_time") if c in names]
            if "step_id" in want:
                q = lf.select(want)
                if "root_lot_id" in want:
                    q = q.filter(
                        pl.col("root_lot_id").cast(_STR, strict=False).str.to_uppercase() == root)
                if streaming:
                    from core.parquet_perf import collect_streaming
                    df = collect_streaming(q)
                else:
                    df = q.collect()
                if df.height:
                    if "tkout_time" in df.columns:
                        df = df.sort("tkout_time", descending=True, nulls_last=True)
                    rows = df.iter_rows(named=True)
                    first_sid = ""
                    by_wafer: dict[str, str] = {}
                    for row in rows:
                        sid = str(row.get("step_id") or "").strip().upper()
                        if not sid:
                            continue
                        if not first_sid:
                            first_sid = sid
                        wafer = _normalize_wafer_id(row.get("wafer_id")) if "wafer_id" in df.columns else ""
                        if wafer and wafer not in by_wafer:
                            by_wafer[wafer] = sid
                    state = {"step_id": first_sid, "by_wafer": by_wafer}
    except Exception as e:
        logger.warning("root latest step 조회 실패 (%s/%s): %s", product, root, e)
    with _ROOT_LATEST_STEP_LOCK:
        _ROOT_LATEST_STEP_CACHE[key] = (now, state)
        while len(_ROOT_LATEST_STEP_CACHE) > _ROOT_LATEST_STEP_MAX:
            _ROOT_LATEST_STEP_CACHE.popitem(last=False)
    return {"step_id": state["step_id"], "by_wafer": dict(state["by_wafer"])}


def _root_latest_step_id(product: str, root_lot_id: str) -> str:
    """Backward-compatible root-level latest step accessor."""
    return str(_root_latest_step_state(product, root_lot_id).get("step_id") or "")


def _split_step_progress(product: str, root_lot_id: str, selected: list[str],
                         wafer_keys: list | None = None,
                         fab_present: bool | None = None) -> dict:
    """Progress shading metadata, unique by product/root/wafer.

    ``fab_present=False`` is only supplied after an authoritative FAB lookup
    completed with zero matches.  That is different from a cache/source read
    failure: a confirmed not-yet-in-FAB root has no current step, so every
    displayed process is not reached and must be grey.
    """
    out = {"step_id": "", "not_reached": [], "by_wafer": {}, "fab_missing": False}
    if fab_present is False:
        all_selected = list(dict.fromkeys(str(col) for col in (selected or []) if str(col)))
        out["fab_missing"] = True
        out["not_reached"] = all_selected
        for raw in wafer_keys or []:
            wafer = _normalize_wafer_id(raw)
            if wafer and wafer not in out["by_wafer"]:
                out["by_wafer"][wafer] = {
                    "step_id": "",
                    "not_reached": list(all_selected),
                }
        return out
    try:
        ctx = _split_step_order_context(product)
        param_rank = ctx.get("param_rank") or {}
        if not param_rank or not str(root_lot_id or "").strip():
            return out
        latest = _root_latest_step_state(product, root_lot_id)
        cur_sid = str(latest.get("step_id") or "")
        out["step_id"] = cur_sid
        cur_rank = _step_rank_for_progress(ctx, cur_sid)
        if cur_rank is not None:
            for col in selected or []:
                rank = param_rank.get(str(col or "").strip().upper())
                if rank is not None and rank > cur_rank:
                    out["not_reached"].append(col)

        latest_by_wafer = latest.get("by_wafer") or {}
        requested_wafers: list[str] = []
        for raw in wafer_keys or latest_by_wafer.keys():
            wafer = _normalize_wafer_id(raw)
            if wafer and wafer not in requested_wafers:
                requested_wafers.append(wafer)
        for wafer in requested_wafers:
            wafer_sid = str(latest_by_wafer.get(wafer) or "")
            wafer_rank = _step_rank_for_progress(ctx, wafer_sid)
            not_reached = []
            if wafer_rank is not None:
                for col in selected or []:
                    rank = param_rank.get(str(col or "").strip().upper())
                    if rank is not None and rank > wafer_rank:
                        not_reached.append(col)
            out["by_wafer"][wafer] = {
                "step_id": wafer_sid,
                "not_reached": not_reached,
            }
    except Exception as e:
        logger.warning("split step-progress 계산 실패 (%s/%s): %s", product, root_lot_id, e)
    return out


# v8.7.5/v8.8.10: INLINE / VM_ prefix 매칭 메타 — schema 매핑 기반.
def _build_inline_meta(product: str = "") -> dict:
    """inline_matching.csv (product, step_id, item_id, optional map table)."""
    base = _base_root()
    rows = _load_csv_rows(base / "inline_matching.csv")
    im = _sch("inline_matching")
    # module 은 inline_matching.csv 에 없다. step_id 로 Vehicle_matching 을 눌러
    # 채우던 보강은 뺐다 — INLINE 은 module 로 따로 묶지 않고 '—' 로 둔다.
    grouped: dict[str, list[dict]] = {}
    p_col = im.get("product_col", "product")
    has_product_col = any(p_col in r or "product" in r for r in rows)
    for r in rows:
        row_prod = r.get(p_col)
        if row_prod is None and p_col != "product":
            row_prod = r.get("product")
        if not _step_matching_product_matches(product, row_prod, allow_common=not has_product_col):
            continue
        iid = (r.get(im.get("item_id_col", "item_id")) or "").strip()
        sid = (r.get(im.get("step_id_col", "step_id")) or "").strip()
        process_id = (r.get(im.get("process_id_col", "process_id")) or "").strip()
        desc = (r.get(im.get("item_desc_col", "item_desc")) or "").strip()
        matching_table = (r.get(im.get("matching_table_col", "matching_table")) or "").strip()
        func_step = (r.get("function_step") or "").strip()
        if not iid or not sid:
            continue
        grouped.setdefault(iid, []).append({
            "step_id": sid,
            "process_id": process_id,
            "item_id": iid,
            "item_desc": desc,
            "matching_table": matching_table,
            "function_step": func_step,
            "module": "",
        })
    out: dict[str, dict] = {}
    for iid, items in grouped.items():
        dedup = []
        seen = set()
        for item in items:
            key = (item.get("function_step", ""), item.get("step_id", ""), item.get("item_desc", ""), item.get("matching_table", ""))
            if key in seen:
                continue
            seen.add(key)
            dedup.append(item)
        step_ids = [x["step_id"] for x in dedup if x.get("step_id")]
        item_desc = next((x.get("item_desc") for x in dedup if x.get("item_desc")), "") or iid
        function_steps = [x["function_step"] for x in dedup if x.get("function_step")]
        process_ids = [x["process_id"] for x in dedup if x.get("process_id")]
        matching_tables = _dedup_list([x.get("matching_table", "") for x in dedup if x.get("matching_table")])
        modules = _dedup_list([x.get("module", "") for x in dedup if x.get("module")])
        out[iid] = {
            "item_id": iid,
            "item_desc": item_desc,
            "modules": modules,
            "module": modules[0] if len(modules) == 1 else "",
            "process_id": process_ids[0] if len(process_ids) == 1 else "",
            "process_ids": process_ids,
            "matching_table": matching_tables[0] if len(matching_tables) == 1 else "",
            "matching_tables": matching_tables,
            "step_id": step_ids[0] if len(step_ids) == 1 else "",
            "step_ids": step_ids,
            "function_step": function_steps[0] if len(function_steps) == 1 else "",
            "function_steps": function_steps,
            "groups": dedup,
            "label": item_desc,
            "sub": "/".join(step_ids) if step_ids else iid,
        }
    return out


def _build_vm_meta(product: str = "") -> dict:
    """vm_matching.csv has step_desc + item_id; step_id comes from Vehicle_matching.csv."""
    base = _base_root()
    rows = _load_csv_rows(base / "vm_matching.csv")
    vm = _sch("vm_matching")
    step_map = _product_step_map_by_desc(product, base)
    grouped: dict[str, list[dict]] = {}
    for r in rows:
        step_desc = _row_step_desc(r, vm)
        item_id = _first_row_value(
            r,
            vm.get("item_id_col", "item_id"),
            "item_id",
            vm.get("feature_col", "feature_name"),
            "feature_name",
        )
        if not step_desc or not item_id:
            continue
        steps = step_map.get(_step_desc_match_key(step_desc), [])
        if not steps:
            continue
        name = f"{step_desc}_{item_id}"
        step_ids = _dedup_list([str(x.get("step_id") or "").strip() for x in steps])
        modules = _dedup_list([str(x.get("module") or "").strip() for x in steps])
        grouped.setdefault(name, []).append({
            "feature_name": name,
            "item_id": item_id,
            "step_desc": step_desc,
            "step_id": step_ids[0] if len(step_ids) == 1 else "",
            "step_ids": step_ids,
            "function_step": step_desc,
            "function_steps": [step_desc],
            "modules": modules,
            "module": modules[0] if len(modules) == 1 else "",
        })
    out: dict[str, dict] = {}
    for fname, items in grouped.items():
        dedup = []
        seen = set()
        for item in items:
            key = (item.get("step_desc", ""), item.get("item_id", ""), tuple(item.get("step_ids") or []))
            if key in seen:
                continue
            seen.add(key)
            dedup.append(item)
        step_ids = _dedup_list([sid for x in dedup for sid in (x.get("step_ids") or [])])
        step_desc = next((x.get("step_desc") for x in dedup if x.get("step_desc")), "") or fname
        item_id = next((x.get("item_id") for x in dedup if x.get("item_id")), "")
        function_steps = _dedup_list([x["function_step"] for x in dedup if x.get("function_step")])
        modules = _dedup_list([mod for x in dedup for mod in (x.get("modules") or [])])
        out[fname] = {
            "feature_name": fname,
            "item_id": item_id,
            "step_desc": step_desc,
            "step_id": step_ids[0] if len(step_ids) == 1 else "",
            "step_ids": step_ids,
            "function_step": function_steps[0] if len(function_steps) == 1 else "",
            "function_steps": function_steps,
            "modules": modules,
            "module": modules[0] if len(modules) == 1 else "",
            "groups": dedup,
            "label": fname,
            "sub": "/".join(step_ids),
        }
    return out


# ── 적용 공정 정보(step label) 표기 ──────────────────────────────────────
#   화면의 "적용 공정 정보" 체크박스와 같은 규약을 서버에서 재현한다.
#     KNOB        → rule_order 별 step_id (서로 다른 step_desc 조건만 `&`)
#     INLINE / VM → `step_id | item_id`
#   표시할 공정이 하나도 없는 매칭 행은 화면에서 감추므로 내보내기에서도 뺀다.
#   TAG/관리 행처럼 매칭 규칙이 없는 행은 원래 이름 그대로 남는다.
#   frontend My_SplitTable.jsx 의 matchStepLines / knobStepLines 와 한 쌍이다.
def _step_label_match_kind(param: str) -> str:
    u = str(param or "").strip().upper()
    if u.startswith("KNOB_") or u == "KNOB":
        return "knob_ppid"
    if u.startswith("INLINE_") or u == "INLINE":
        return "inline_matching"
    if u.startswith("VM_") or u == "VM":
        return "vm_matching"
    return ""


def _step_label_meta_lookup(meta_map: dict, param: str, prefix: str) -> dict:
    if not param or not isinstance(meta_map, dict):
        return {}
    full = str(param or "").strip()
    tail = _re.sub(rf"^{prefix}_", "", full, flags=_re.I).strip()
    if isinstance(meta_map.get(full), dict):
        return meta_map[full]
    if isinstance(meta_map.get(tail), dict):
        return meta_map[tail]
    for k, v in meta_map.items():
        if str(k or "").strip().casefold() in {full.casefold(), tail.casefold()} and isinstance(v, dict):
            return v
    return {}


def _step_label_group_ids(group: dict) -> list[str]:
    ids = group.get("step_ids") if isinstance(group, dict) else None
    if isinstance(ids, list):
        return [str(v or "").strip() for v in ids if str(v or "").strip()]
    sid = str((group or {}).get("step_id") or "").strip()
    return [sid] if sid else []


def _step_label_knob_lines(groups, exclude_not_null: bool = True) -> list[str]:
    by_order: dict[str, list[dict]] = {}
    order_seq: list[str] = []
    for idx, g in enumerate(groups if isinstance(groups, list) else []):
        if not isinstance(g, dict):
            continue
        order = str(g.get("rule_order") or f"R{idx + 1}").strip() or f"R{idx + 1}"
        if order not in by_order:
            by_order[order] = []
            order_seq.append(order)
        by_order[order].append(g)
    lines: list[str] = []
    seen: set[str] = set()
    for order in order_seq:
        by_desc: dict[str, dict] = {}
        for g in by_order[order]:
            operator = _re.sub(r"[\s-]+", "_", str(g.get("operator") or "").strip().lower())
            if operator == "is_null":
                continue
            if exclude_not_null and operator == "not_null":
                continue
            desc = str(g.get("step_desc") or g.get("func_step") or "").strip()
            if not desc:
                continue
            bucket = by_desc.setdefault(desc.casefold(), {"step_ids": [], "seen": set()})
            for sid in _step_label_group_ids(g):
                if sid.casefold() not in bucket["seen"]:
                    bucket["seen"].add(sid.casefold())
                    bucket["step_ids"].append(sid)
        desc_groups = [item for item in by_desc.values() if item["step_ids"]]
        if not desc_groups:
            continue
        line = "\n&\n".join("\n".join(item["step_ids"]) for item in desc_groups)
        if line not in seen:
            seen.add(line)
            lines.append(line)
    return lines


def _step_label_item_lines(meta: dict) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    fallback_item = str((meta or {}).get("item_id") or "").strip()

    def push(sid, item_id):
        s = str(sid or "").strip()
        if not s:
            return
        i = str(item_id or fallback_item or "").strip()
        line = f"{s} | {i}" if i else s
        if line not in seen:
            seen.add(line)
            out.append(line)

    groups = (meta or {}).get("groups")
    for g in groups if isinstance(groups, list) else []:
        if not isinstance(g, dict):
            continue
        ids = _step_label_group_ids(g)
        if ids:
            for sid in ids:
                push(sid, g.get("item_id"))
        else:
            push(g.get("step_id"), g.get("item_id"))
    if not out:
        for sid in ((meta or {}).get("step_ids") or []):
            push(sid, fallback_item)
    return out


def _step_label_metas(product: str) -> dict:
    out = {"knob": {}, "inline": {}, "vm": {}}
    for key, fn in (("knob", _build_knob_meta), ("inline", _build_inline_meta), ("vm", _build_vm_meta)):
        try:
            meta = fn(product)
            out[key] = meta if isinstance(meta, dict) else {}
        except Exception:
            out[key] = {}
    return out


def _step_label_lines_for_param(param: str, metas: dict, exclude_not_null: bool = True) -> tuple[str, list[str]]:
    kind = _step_label_match_kind(param)
    if not kind:
        return "", []
    if kind == "knob_ppid":
        meta = _step_label_meta_lookup(metas.get("knob") or {}, param, "KNOB")
        return kind, _step_label_knob_lines((meta or {}).get("groups") or [], exclude_not_null)
    if kind == "inline_matching":
        return kind, _step_label_item_lines(_step_label_meta_lookup(metas.get("inline") or {}, param, "INLINE"))
    return kind, _step_label_item_lines(_step_label_meta_lookup(metas.get("vm") or {}, param, "VM"))


def _apply_step_label_columns(product: str, selected: list[str], col_rename: dict,
                              exclude_not_null: bool = True) -> tuple[list[str], dict]:
    """내보내기 항목 라벨을 적용 공정 표기로 바꾸고, 표시할 공정이 없는 매칭 행은 뺀다."""
    metas = _step_label_metas(product)
    kept: list[str] = []
    rename = dict(col_rename or {})
    for col in selected or []:
        kind, lines = _step_label_lines_for_param(col, metas, exclude_not_null)
        if not kind:
            kept.append(col)
            continue
        if not lines:
            continue
        rename[col] = "\n".join(lines)
        kept.append(col)
    return kept, rename


# ── 통합(병합) 표시 대상 ────────────────────────────────────────────────
# 옆칸과 같은 값을 하나로 묶는 표시는 split 조건 열에만 의미가 있다.
# INLINE/VM 은 wafer 별 실측값이고 TAG/관리 행은 자유 입력이라, 값이 우연히
# 같다고 묶으면 wafer 별 값이 몇 개인지 읽을 수 없게 된다.
MERGE_VIEW_PREFIXES = ("KNOB", "FAB", "MASK")


def _merge_view_allowed_param(param: str) -> bool:
    u = str(param or "").strip().upper()
    return any(u == p or u.startswith(f"{p}_") for p in MERGE_VIEW_PREFIXES)


def _virtual_columns_for_prefix(product: str, prefix: str) -> list[str]:
    pref = str(prefix or "").strip().upper()
    if not pref:
        return []
    out: list[str] = []
    seen: set[str] = set()

    def _push(name: str, pref_name: str):
        raw = str(name or "").strip()
        if not raw:
            return
        full = raw if raw.upper().startswith(pref_name + "_") else f"{pref_name}_{raw}"
        if full not in seen:
            seen.add(full)
            out.append(full)

    try:
        if pref == "KNOB":
            for key in (_build_knob_meta(product) or {}).keys():
                _push(key, "KNOB")
        elif pref == "INLINE":
            for key in (_build_inline_meta(product) or {}).keys():
                _push(key, "INLINE")
        elif pref == "VM":
            for key in (_build_vm_meta(product) or {}).keys():
                _push(key, "VM")
    except Exception:
        return out
    return out


@router.get("/inline-meta")
def inline_meta(product: str = Query("")):
    """v8.7.5/v8.8.15: INLINE prefix 항목 매칭 메타. product 필터 추가."""
    return {"items": _build_inline_meta(product)}


@router.get("/vm-meta")
def vm_meta(product: str = Query("")):
    """v8.7.5/v8.8.7: VM_ prefix 항목 매칭 메타. product 필터 추가."""
    return {"items": _build_vm_meta(product)}


@router.post("/infer-step-mapping")
def infer_step_mapping(request: Request, product: str = Query(...), kind: str = Query("inline")):
    """v8.8.33: FAB 공정이력을 활용해 INLINE / VM 의 step_id 자동 추론.
    보안: admin 또는 page_manager('splittable') 만 실행 가능 (rulebook CSV 쓰기 보호)."""
    me = current_user(request)
    if not is_page_manager(me, "splittable"):
        raise HTTPException(403, "admin or splittable page manager only")
    # 전략: INLINE 의 (lot_id, wafer_id, item_id, tkout_time/time) 에 대해 FAB 에서
    #   같은 (lot_id, wafer_id) 의 step_id 중 INLINE 측정 직전의 step_id 매칭.
    # 결과를 inline_matching.csv (or vm_matching.csv) 에 upsert. 수동 편집분은 보존.
    import polars as pl
    if not product:
        raise HTTPException(400, "product required")
    if kind not in ("inline", "vm"):
        raise HTTPException(400, "kind must be inline|vm")
    db_root = PATHS.db_root
    fab_root = db_root / "1.RAWDATA_DB_FAB" / product
    src_root = db_root / ("1.RAWDATA_DB_INLINE" if kind == "inline" else "1.RAWDATA_DB_VM") / product
    if not fab_root.is_dir():
        raise HTTPException(404, f"FAB folder not found: {fab_root}")
    if not src_root.is_dir():
        raise HTTPException(404, f"{kind.upper()} folder not found: {src_root}")
    fab_files = _rglob_files_ci(fab_root, (".parquet",))[-30:]
    src_files = _rglob_files_ci(src_root, (".parquet",))[-30:]
    if not fab_files or not src_files:
        raise HTTPException(404, "no parquet files")
    try:
        fab_lf = _scan_parquet_compat([str(f) for f in fab_files], hive_partitioning=True)
        src_lf = _scan_parquet_compat([str(f) for f in src_files], hive_partitioning=True)
    except Exception as e:
        raise HTTPException(500, f"scan error: {e}")
    fab_schema = fab_lf.collect_schema().names()
    src_schema = src_lf.collect_schema().names()
    if "step_id" not in fab_schema:
        raise HTTPException(400, "FAB has no step_id column")
    if "item_id" not in src_schema:
        raise HTTPException(400, f"{kind.upper()} has no item_id column")
    fab_time_col = "time" if "time" in fab_schema else ("tkout_time" if "tkout_time" in fab_schema else "tkin_time")
    src_time_col = "time" if "time" in src_schema else ("tkout_time" if "tkout_time" in src_schema else "tkin_time")
    if fab_time_col not in fab_schema:
        raise HTTPException(400, "FAB has no time/tkout_time/tkin_time column")
    if src_time_col not in src_schema:
        raise HTTPException(400, f"{kind.upper()} has no time/tkout_time/tkin_time column")
    fab_exprs = [pl.col(c) for c in ("lot_id", "wafer_id", "step_id") if c in fab_schema]
    fab_exprs.append(pl.col(fab_time_col).alias("time"))
    src_exprs = [pl.col(c) for c in ("item_id", "lot_id", "wafer_id") if c in src_schema]
    src_exprs.append(pl.col(src_time_col).alias("time"))
    fab_df = fab_lf.select(fab_exprs).collect()
    src_df = src_lf.select(src_exprs).collect()
    if fab_df.is_empty() or src_df.is_empty():
        raise HTTPException(404, "no rows after select")
    for label, df_name in (("FAB", "fab_df"), (kind.upper(), "src_df")):
        df = fab_df if df_name == "fab_df" else src_df
        if df.schema.get("time") != pl.Datetime:
            try:
                df = df.with_columns(pl.col("time").str.strptime(pl.Datetime, strict=False))
            except Exception:
                pass
            if df_name == "fab_df":
                fab_df = df
            else:
                src_df = df
    # item_id 별로 최빈 step_id.
    # 단순화: FAB 의 (lot_id, wafer_id) 그룹 내 max(time, step_id) 를 각 INLINE row 와 join_asof.
    try:
        fab_sorted = fab_df.sort(["lot_id", "wafer_id", "time"])
        src_sorted = src_df.sort(["lot_id", "wafer_id", "time"])
        joined = src_sorted.join_asof(
            fab_sorted, on="time", by=["lot_id", "wafer_id"], strategy="backward",
        )
    except Exception as e:
        raise HTTPException(500, f"join_asof failed: {e}")
    if "step_id" not in joined.columns:
        raise HTTPException(500, "step_id missing after join")
    joined = joined.filter(pl.col("step_id").is_not_null())
    if joined.is_empty():
        raise HTTPException(404, "no matched rows")
    # item_id 별로 가장 많이 붙은 step_id 선정.
    counts = (
        joined.group_by(["item_id", "step_id"])
              .agg(pl.len().alias("n"))
              .sort("n", descending=True)
    )
    winners: dict[str, str] = {}
    for r in counts.to_dicts():
        iid = r.get("item_id")
        if iid and iid not in winners:
            winners[str(iid)] = str(r.get("step_id") or "")
    # CSV upsert.
    base = _base_root()
    csv_name = "inline_matching.csv" if kind == "inline" else "vm_matching.csv"
    csv_fp = base / csv_name
    rulebook_meta = _RULEBOOK_FILES["inline_matching" if kind == "inline" else "vm_matching"]
    existing = _load_csv_rows(csv_fp)
    sid_to_step_desc = {}
    if kind == "vm":
        for steps in _product_step_map_by_desc(product, base).values():
            for step in steps:
                sid = str(step.get("step_id") or "").strip()
                desc = str(step.get("step_desc") or "").strip()
                if sid and desc:
                    sid_to_step_desc.setdefault(sid.casefold(), desc)
    added = []
    for iid, sid in winners.items():
        iid = str(iid or "").strip()
        sid = str(sid or "").strip()
        if not iid:
            continue
        if kind == "inline":
            if (product, iid) in {(str(r.get("product") or "").strip(), str(r.get("item_id") or "").strip()) for r in existing}:
                continue
            existing.append({"product": product, "item_id": iid, "step_id": sid, "item_desc": ""})
            added.append((iid, sid))
        else:
            step_desc = sid_to_step_desc.get(sid.casefold(), "")
            if not step_desc:
                continue
            existing_key = (step_desc.casefold(), iid.casefold())
            if existing_key in {
                (str(r.get("step_desc") or r.get("function_step") or "").strip().casefold(),
                 str(r.get("item_id") or r.get("feature_name") or "").strip().casefold())
                for r in existing
            }:
                continue
            existing.append({"step_desc": step_desc, "item_id": iid})
            added.append((iid, step_desc))
    if not added:
        return {"ok": True, "added": 0, "total": len(winners), "note": "모두 기존에 등록됨"}
    try:
        final_rows, dedupe_rows = _matching_cache.dedupe_rows(
            existing,
            key_cols=[k for k in rulebook_meta.get("cols", []) if k],
            required_cols=rulebook_meta.get("required", []),
            strict_required=True,
        )
    except ValueError as e:
        raise HTTPException(400, f"validation failed: {e}")

    # write back — header = union of all keys
    import csv as _csv
    all_keys: list = []
    for r in final_rows:
        for k in r.keys():
            if k not in all_keys:
                all_keys.append(k)
    try:
        csv_fp.parent.mkdir(parents=True, exist_ok=True)
        with open(csv_fp, "w", encoding="utf-8", newline="") as f:
            w = _csv.DictWriter(f, fieldnames=all_keys)
            w.writeheader()
            for r in final_rows:
                w.writerow({k: r.get(k, "") for k in all_keys})
        cache_result = _matching_cache.refresh_matching_csv(csv_fp)
        if not cache_result.get("ok", False):
            logger.warning("infer_step_mapping cache refresh failed: %s", cache_result)
    except Exception as e:
        raise HTTPException(500, f"csv write failed: {e}")
    return {
        "ok": True,
        "added": len(added),
        "deduped_rows": dedupe_rows,
        "total": len(winners),
        "cache_rows": cache_result.get("rows"),
        "csv": str(csv_fp.name),
        "sample_added": added[:10],
    }


# v8.8.10: Rulebook "컬럼 역할 → 실제 컬럼명" 매핑 저장소 (soft-landing).
#   사내 CSV 의 컬럼 이름이 기본값과 다를 때 admin 이 여기서 매핑만 바꾸면 _build_knob_meta /
#   _build_vm_meta / _build_inline_meta 가 그대로 동작. rulebook 파일 자체는 손대지 않음.
#
# v9.5.71: 이 파일에 있던 스키마 상수·로더·파일명 정리 함수의 **복사본을 제거**하고
#   `app_v2.modules.splittable.rulebook_repository` 한 곳으로 모았다. 같은 파일
#   (`{data_root}/splittable/rulebook_schema.json`)을 가리키는 정의가 두 벌이라
#   한쪽만 바꾸면 조용히 갈라졌고, 테스트에서도 한쪽만 격리돼 실제 운영 파일을
#   읽는 사고가 있었다. 이름은 기존 호출부 호환을 위해 얇은 위임으로 남긴다.
from app_v2.modules.splittable import rulebook_repository as _rulebook_repository
# pivot 빌드 청크 설정(캐시관리 톱니바퀴)에서 기본값/상한을 읽는다. 빌더 자체는
# 무거운 호출부에서 지연 import 하지만, 설정 payload 는 상수만 보므로 모듈 별칭으로
# 둔다 (cache_builder 는 routers 를 import 하지 않아 순환이 없다).
from app_v2.modules.splittable import cache_builder as _pivot_cache_builder

RULEBOOK_SCHEMA_FILE = _rulebook_repository.RULEBOOK_SCHEMA_FILE
_DEFAULT_RULEBOOK_SCHEMA = rulebook_repo.get_default_schema()


def _load_rulebook_schema() -> dict:
    return rulebook_repo.load_schema()


def _save_rulebook_schema(schema: dict) -> None:
    rulebook_repo.save_schema(schema)


def _sch(kind: str) -> dict:
    return rulebook_repo.get_sch(kind)


def _clean_rulebook_filename(value: object, default: str) -> str:
    return rulebook_repo.clean_rulebook_filename(value, default)


@router.get("/rulebook/schema")
def get_rulebook_schema():
    """현재 역할→컬럼명 매핑 + 기본값 같이 반환. FE 에서 diff 표시 가능."""
    return {"schema": rulebook_repo.load_schema(), "defaults": rulebook_repo.get_default_schema()}


class RulebookSchemaReq(BaseModel):
    kind: str
    mapping: dict
    username: str = ""


@router.post("/rulebook/schema/save")
def save_rulebook_schema(
    req: RulebookSchemaReq,
    request: Request,
    _perm=Depends(require_page_manager("splittable")),
):
    me = current_user(request)
    if req.kind not in rulebook_repo.get_default_schema():
        raise HTTPException(400, f"unknown rulebook: {req.kind}")
    cur = rulebook_repo.load_schema()
    defm = rulebook_repo.get_default_schema()[req.kind]
    new_map = {}
    for role, _dfl in defm.items():
        v = (req.mapping or {}).get(role, _dfl)
        if role == "file_name":
            v = rulebook_repo.clean_rulebook_filename(v, _dfl)
        else:
            v = str(v or "").strip() or _dfl
        new_map[role] = v
    cur[req.kind] = new_map
    rulebook_repo.save_schema(cur)
    _audit_user(req.username or (me.get("username") if isinstance(me, dict) else ""),
                "splittable:rulebook_schema_save",
                detail=f"kind={req.kind} mapping={new_map}")
    return {"ok": True, "kind": req.kind, "mapping": new_map}


# v8.8.7: Rulebook (knob_ppid.csv + Vehicle_matching.csv/step_matching.csv) admin 인라인 편집 CRUD.
#   admin 만 수정 가능. 저장 시 row 정규화 + 빈 행 제거 + 원자적 교체.
#   스키마는 _build_knob_meta 가 읽는 컬럼과 동일해야 함.
_RULEBOOK_FILES = {
    "knob_ppid": {
        "filename": "ppid_knob.csv",
        "legacy_filename": "knob_ppid.csv",
        "cols": ["feature_name", "rule_order", "step_desc", "operator", "value", "category"],
        "required": ["feature_name", "step_desc"],
    },
    "step_matching": {
        "filename": "Vehicle_matching.csv",
        "legacy_filename": "step_matching.csv",
        "cols": ["product", "step_id", "step_desc"],
        "required": ["product", "step_id", "step_desc"],
    },
    # v8.8.9: INLINE / VM 매칭도 동일 CRUD 로 관리.
    #   inline_matching.csv: (product, step_id, item_id, item_desc) — INLINE_<item_id> 측정 메타.
    "inline_matching": {
        "filename": "inline_matching.csv",
        "cols": ["product", "step_id", "item_id", "item_desc", "matching_table"],
        "required": ["product", "step_id", "item_id"],
    },
    #   vm_matching.csv: (step_desc, item_id) — VM_<step_desc>_<item_id>, step_id 는 Vehicle_matching.csv 에서 확장.
    "vm_matching": {
        "filename": "vm_matching.csv",
        "cols": ["step_desc", "item_id"],
        "required": ["step_desc", "item_id"],
    },
}


def _normalize_rulebook_rows(kind: str, rows: list[dict]) -> list[dict]:
    out = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        r = dict(row)
        if kind in {"knob_ppid", "step_matching", "vm_matching"} and not str(r.get("step_desc") or "").strip():
            r["step_desc"] = _row_step_desc(r, _sch(kind))
        if kind == "knob_ppid" and not str(r.get("value") or "").strip():
            r["value"] = _first_row_value(r, _sch(kind).get("value_col", "value"), "ppid", "category")
        if kind == "vm_matching" and not str(r.get("item_id") or "").strip():
            r["item_id"] = _first_row_value(r, _sch(kind).get("item_id_col", "item_id"), "feature_name")
        out.append(r)
    return out


def _rulebook_path_for_base(kind: str, base: Path | None = None) -> Path:
    meta = _RULEBOOK_FILES.get(kind)
    if not meta:
        raise HTTPException(400, f"unknown rulebook: {kind}")
    root = base or _base_root()
    configured = _clean_rulebook_filename(_sch(kind).get("file_name"), meta["filename"])
    primary = root / configured
    if configured != meta["filename"] or primary.exists() or not meta.get("legacy_filename"):
        return primary
    legacy = root / str(meta.get("legacy_filename") or "")
    return legacy if legacy.exists() else primary


def _rulebook_path(kind: str) -> Path:
    return _rulebook_path_for_base(kind)


def _rulebook_row_matches_product(kind: str, row: dict, product: str, *, allow_common: bool = True) -> bool:
    p_col = _sch(kind).get("product_col", "product")
    row_product = (row or {}).get(p_col)
    if row_product is None and p_col != "product":
        row_product = (row or {}).get("product")
    if kind in {"step_matching", "inline_matching"}:
        return _step_matching_product_matches(product, row_product, allow_common=allow_common)
    return _product_value_matches(product, row_product, allow_common=allow_common)


@router.get("/rulebook")
def get_rulebook(kind: str = Query("knob_ppid"), product: str = Query("")):
    """v8.8.7: rulebook CSV 를 JSON 으로 반환.

    KNOB and VM item rows are product-common. Step/INLINE matching rows remain product-scoped.
    """
    return rulebook_service.get_rulebook(kind, product)


class RulebookSaveReq(BaseModel):
    kind: str               # "knob_ppid" | "step_matching" | "inline_matching" | "vm_matching"
    rows: List[dict]        # 전체 대체 (혹은 product 스코프 대체)
    product: str = ""       # 주어지면 해당 제품 rows 만 대체, 빈값이면 파일 전체 대체
    username: str = ""


@router.post("/rulebook/save")
def save_rulebook(req: RulebookSaveReq, request: Request, _perm=Depends(require_page_manager("splittable"))):
    """Admin 또는 splittable page manager 전용. product 스코프면 해당 제품 행만 교체."""
    me = current_user(request)
    username = req.username or (me.get("username") if isinstance(me, dict) else "")
    return rulebook_service.save_rulebook(req.kind, req.rows, req.product, username)


@router.get("/knob-meta")
def knob_meta(product: str = Query("")):
    """v8.4.7: KNOB feature_name → step_desc(step_id) 역산 맵.

    응답 스키마:
      {
        "features": {
          "KNOB_GATE_PPID": {
            "groups": [
              {"step_desc":"GATE_PATTERN","step_ids":["AA200030","AA200040","AA200050"],
               "value":"PP_GATE_01","operator":"+","rule_order":"R1","category":"gate"},
              {"step_desc":"PC_ETCH","step_ids":["AA200100","AA200110"],
               "value":"PP_PC_01","operator":"","rule_order":"R2","category":"gate"}
            ],
            "label": "GATE_PATTERN (AA200030/AA200040/AA200050) + PC_ETCH (AA200100/AA200110)"
          },
          ...
        }
      }
    ppid_knob.csv는 product 없는 공용 룰북으로 읽고, product별 step_id 확장만 Vehicle_matching.csv에서 적용한다.
    """
    return {"features": _build_knob_meta(product)}


class PrefixSaveReq(BaseModel):
    prefixes: List[str]


@router.post("/prefixes/save")
def save_prefixes(req: PrefixSaveReq, _perm=Depends(require_page_manager("splittable"))):
    save_json(PREFIX_CFG, req.prefixes)
    return {"ok": True}


# ── Cell decimal precision (v8.1.1) ──
# Per-prefix decimal places for numeric cell display. Only INLINE/VM default;
# any prefix key can be added here. Admin-configurable.
PRECISION_CFG = PLAN_DIR / "precision_config.json"
DEFAULT_PRECISION = {"INLINE": 2, "VM": 2}


@router.get("/precision")
def get_precision():
    return {"precision": load_json(PRECISION_CFG, DEFAULT_PRECISION)}


class PrecisionReq(BaseModel):
    precision: dict   # {"INLINE": 2, "VM": 3, ...}


@router.post("/precision/save")
def save_precision(req: PrecisionReq, _perm=Depends(require_page_manager("splittable"))):
    # Sanitize: ensure int 0..10 per prefix
    out = {}
    for k, v in (req.precision or {}).items():
        if not isinstance(k, str) or not k.strip():
            continue
        try:
            n = int(v)
        except Exception:
            continue
        n = max(0, min(10, n))
        out[k.strip().upper()] = n
    save_json(PRECISION_CFG, out)
    return {"ok": True, "precision": out}


# ── v8.8.6: Paste sets (팀 공용 — 인폼·SplitTable paste 공유) ──────────────
# Schema: [{id, name, product, columns:[...], rows:[[...]], username, created, updated}]
#   - CUSTOM 탭에서 paste 세트를 직접 columns 로 취급 → as-is 뷰 (SplitTable custom 과 별개 보관).
#   - FE 는 로컬스토리지 대신 이 엔드포인트에서 읽고 씀. 로컬 폴백은 FE 가 알아서.
def _load_paste_sets() -> list:
    data = load_json(PASTE_SETS_FILE, [])
    return data if isinstance(data, list) else []

def _save_paste_sets(items: list) -> None:
    save_json(PASTE_SETS_FILE, items, indent=2)


class PasteSetSaveReq(BaseModel):
    name: str
    product: str = ""
    columns: List[str]
    rows: List[List] = []
    username: str = ""


@router.get("/paste-sets")
def list_paste_sets(product: str = Query("")):
    """팀 공용 paste 세트 목록. product 가 주어지면 해당 product 또는 빈 product(공용) 만 반환."""
    items = _load_paste_sets()
    if product:
        items = [s for s in items if not s.get("product") or s.get("product") == product]
    # recent first
    items = sorted(items, key=lambda s: s.get("updated", s.get("created", "")), reverse=True)
    return {"sets": items}


@router.post("/paste-sets/save")
def save_paste_set(
    req: PasteSetSaveReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    import secrets as _secrets
    nm = (req.name or "").strip()
    if not nm:
        raise HTTPException(400, "name required")
    cols = [str(c) for c in (req.columns or []) if c]
    if not cols:
        raise HTTPException(400, "columns required")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    items = _load_paste_sets()
    # upsert by (name, product) — 같은 이름·제품이면 덮어쓰기.
    existing = next((s for s in items if s.get("name") == nm and s.get("product", "") == (req.product or "")), None)
    if existing:
        existing.update({
            "columns": cols, "rows": req.rows or [], "username": actor or existing.get("username", ""),
            "updated": now,
        })
    else:
        items.append({
            "id": "ps_" + _secrets.token_hex(5),
            "name": nm, "product": req.product or "",
            "columns": cols, "rows": req.rows or [],
            "username": actor,
            "created": now, "updated": now,
        })
    _save_paste_sets(items)
    invalidate_splittable_sets_cache(req.product or "")
    return {"ok": True, "count": len(items)}


class PasteSetDeleteReq(BaseModel):
    id: str = ""
    name: str = ""
    product: str = ""
    username: str = ""


@router.post("/paste-sets/delete")
def delete_paste_set(
    req: PasteSetDeleteReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    items = _load_paste_sets()
    before = len(items)
    if req.id:
        items = [s for s in items if s.get("id") != req.id]
    elif req.name:
        items = [s for s in items if not (s.get("name") == req.name and s.get("product", "") == (req.product or ""))]
    else:
        raise HTTPException(400, "id or name required")
    if len(items) == before:
        raise HTTPException(404, "paste set not found")
    _save_paste_sets(items)
    invalidate_splittable_sets_cache(req.product or "")
    return {"ok": True, "removed": before - len(items)}


@router.post("/paste-sets/to-custom")
def paste_set_to_custom(
    req: PasteSetDeleteReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    """paste 세트의 columns 를 CUSTOM 커스텀 뷰로 승격.
    CUSTOM 탭에서 바로 선택 가능하게 `custom_<safe_name>.json` 생성."""
    items = _load_paste_sets()
    src = None
    if req.id:
        src = next((s for s in items if s.get("id") == req.id), None)
    elif req.name:
        src = next((s for s in items if s.get("name") == req.name and s.get("product", "") == (req.product or "")), None)
    if not src:
        raise HTTPException(404, "paste set not found")
    actor = req.username or src.get("username", "")
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    fp, name = _custom_file_path_for_name(src.get("name") or "paste_custom")
    columns = _clean_custom_columns(src.get("columns") or [])
    if not columns:
        raise HTTPException(400, "custom columns required")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    existing = load_json(fp, None) if fp.exists() else None
    save_json(fp, {
        "name": name, "username": actor,
        "columns": columns,
        "created": (existing or {}).get("created", now),
        "updated": now,
        "version": int((existing or {}).get("version", 0)) + 1,
        "source": "paste-set", "paste_id": src.get("id", ""),
    })
    invalidate_splittable_sets_cache(req.product or "")
    return {"ok": True, "custom_name": name}


# ── Customs ──
@router.get("/customs")
def list_customs():
    customs = []
    for f in sorted(PLAN_DIR.glob("custom_*.json")):
        c = _sanitize_custom_record(load_json(f, None), f, persist=True)
        if c:
            c["_file"] = f.name
            customs.append(c)
    return {"customs": customs}


class CustomSaveReq(BaseModel):
    name: str
    username: str
    columns: List[Any]
    # v8.6.1: 낙관적 잠금 — 동일 name 의 기존 커스텀이 있으면 expected_version 일치 시에만 덮어쓴다.
    # 신규(처음 저장)면 0 또는 None.
    expected_version: int | None = None


@router.post("/customs/save")
def save_custom(
    req: CustomSaveReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    actor = req.username
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    fp, name = _custom_file_path_for_name(req.name)
    columns = _clean_custom_columns(req.columns)
    if not columns:
        raise HTTPException(400, "custom columns required")
    now = datetime.datetime.now().isoformat()
    existing = load_json(fp, None) if fp.exists() else None
    if existing:
        cur_v = int(existing.get("version", 1))
        # 클라가 보낸 expected_version 이 None 이면 강제 덮어쓰기 (legacy).
        # 정수면 일치해야 함. 불일치 → conflict 응답.
        if req.expected_version is not None and int(req.expected_version) != cur_v:
            return {
                "ok": False, "conflict": True,
                "server_version": cur_v, "current": existing,
                "detail": "Version conflict — another user has saved this custom view.",
            }
        new_v = cur_v + 1
        created = existing.get("created", now)
    else:
        new_v = 1
        created = now
    save_json(fp, {
        "name": name, "username": actor, "columns": columns,
        "created": created, "updated": now, "version": new_v,
    })
    invalidate_splittable_sets_cache()
    return {"ok": True, "version": new_v}


class CustomDeleteReq(BaseModel):
    name: str
    username: str


@router.post("/customs/delete")
def delete_custom(
    req: CustomDeleteReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    fp = PLAN_DIR / f"custom_{safe_id(req.name)}.json"
    if not fp.exists():
        raise HTTPException(404)
    fp.unlink(missing_ok=True)
    invalidate_splittable_sets_cache()
    return {"ok": True}


class CustomTagColumnReq(BaseModel):
    product: str
    name: str
    username: str = ""
    # None = 건드리지 않음, "" = 비우기. 생성 시에는 빈 module 로 시작한다.
    module: Optional[str] = None


class CustomTagModuleReq(BaseModel):
    product: str
    column: str
    module: str = ""
    username: str = ""


class CustomTagColumnDeleteReq(BaseModel):
    product: str
    column: str = ""
    name: str = ""
    username: str = ""


class CustomTagValuesReq(BaseModel):
    product: str
    values: dict
    username: str = ""
    root_lot_id: str = ""


class ManagementRowColumnReq(BaseModel):
    product: str
    name: str
    username: str = ""


class ManagementRowValuesReq(BaseModel):
    product: str
    values: dict
    username: str = ""
    root_lot_id: str = ""


@router.get("/custom-tags")
def list_custom_tags(product: str = Query("")):
    columns = _custom_tag_columns_for_product(product) if product else []
    return {"columns": columns, "count": len(columns)}


@router.post("/custom-tags/columns/save")
def save_custom_tag_column(req: CustomTagColumnReq, request: Request = None):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    column = _tag_column_id(req.name)
    label = str(req.name or "").strip()
    if label.upper().startswith(f"{CUSTOM_TAG_PREFIX}_"):
        label = label[len(CUSTOM_TAG_PREFIX) + 1:].strip()
    now = datetime.datetime.now().isoformat(timespec="seconds")
    data = _load_custom_tags_data()
    entry = _ensure_custom_tag_column(
        data,
        product=product,
        column=column,
        label=label or column,
        actor=actor,
        now=now,
        module=req.module,
    )
    _save_custom_tags_data(data)
    return {
        "ok": True,
        "column": entry["column"],
        "label": entry["label"],
        "module": _clean_tag_module(entry.get("module")),
        "columns": _custom_tag_columns_for_product(product),
    }


# TAG 행의 module 은 이미 있는 열에만 붙인다 — 오타로 새 열이 생기면 안 된다.
@router.post("/custom-tags/columns/module")
def save_custom_tag_module(req: CustomTagModuleReq, request: Request = None):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    column = _tag_column_id(req.column)
    module = _clean_tag_module(req.module)
    data = _load_custom_tags_data()
    entry = next(
        (
            c
            for c in (data.get("columns") or [])
            if isinstance(c, dict)
            and str(c.get("product") or "").strip() == product
            and str(c.get("column") or "").strip() == column
        ),
        None,
    )
    if entry is None:
        raise HTTPException(404, f"tag column not found: {column}")
    entry["module"] = module
    entry["updated"] = datetime.datetime.now().isoformat(timespec="seconds")
    if actor:
        entry["username"] = actor
    _save_custom_tags_data(data)
    return {"ok": True, "column": column, "module": module, "columns": _custom_tag_columns_for_product(product)}


@router.post("/custom-tags/delete")
@router.post("/custom-tags/columns/delete")
def delete_custom_tag_column(
    req: CustomTagColumnDeleteReq,
    request: Request = None,
    _perm=Depends(require_page_manager("splittable")),
):
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    raw_column = str(req.column or req.name or "").strip()
    if not raw_column:
        raise HTTPException(400, "tag column required")
    column = _tag_column_id(raw_column)
    data = _load_custom_tags_data()

    columns = data.get("columns") if isinstance(data.get("columns"), list) else []
    kept_columns = []
    deleted_columns = 0
    for entry in columns:
        if (
            isinstance(entry, dict)
            and str(entry.get("product") or "").strip() == product
            and str(entry.get("column") or "").strip() == column
        ):
            deleted_columns += 1
            continue
        kept_columns.append(entry)

    values = data.get("values") if isinstance(data.get("values"), dict) else {}
    kept_values = {}
    deleted_values = 0
    for key, value in values.items():
        parts = str(key).split("|", 3)
        if len(parts) == 4 and parts[0] == product and parts[3] == column:
            deleted_values += 1
            continue
        kept_values[key] = value

    data["columns"] = kept_columns
    data["values"] = kept_values
    _save_custom_tags_data(data)
    actor = req.username or ""
    if not actor and isinstance(_perm, dict):
        actor = _perm.get("username") or ""
    _audit_user(actor, "splittable:custom_tag_delete", detail=f"product={product} column={column}")
    return {
        "ok": True,
        "column": column,
        "deleted_columns": deleted_columns,
        "deleted_values": deleted_values,
        "columns": _custom_tag_columns_for_product(product),
    }


@router.post("/custom-tags/values")
def save_custom_tag_values(req: CustomTagValuesReq, request: Request = None):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    data = _load_custom_tags_data()
    values = data.setdefault("values", {})
    saved = 0
    deleted = 0
    rejected: list[str] = []
    for cell_key, raw_value in (req.values or {}).items():
        parts = str(cell_key or "").split("|", 2)
        if len(parts) != 3:
            rejected.append(str(cell_key))
            continue
        root_lot_id, wafer_id, column = [p.strip() for p in parts]
        if not root_lot_id or not wafer_id or not column.upper().startswith(f"{CUSTOM_TAG_PREFIX}_"):
            rejected.append(str(cell_key))
            continue
        _ensure_custom_tag_column(
            data,
            product=product,
            column=column,
            label=column[len(CUSTOM_TAG_PREFIX) + 1:] or column,
            actor=actor,
            now=now,
        )
        store_key = _tag_value_key(product, root_lot_id, wafer_id, column)
        value = "" if raw_value is None else str(raw_value).strip()
        if value:
            values[store_key] = {"value": value, "username": actor, "updated": now}
            saved += 1
        elif store_key in values:
            values.pop(store_key, None)
            deleted += 1
    _save_custom_tags_data(data)
    return {"ok": True, "saved": saved, "deleted": deleted, "rejected": rejected}


@router.get("/management-rows")
def list_management_rows(product: str = Query("")):
    columns = _management_row_columns_for_product(product) if product else []
    return {"columns": columns, "count": len(columns)}


@router.post("/management-rows/columns/save")
def save_management_row_column(req: ManagementRowColumnReq, request: Request = None):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    column = _management_row_id(req.name)
    label = str(req.name or "").strip()
    if label.upper().startswith(f"{MANAGEMENT_ROW_PREFIX}_"):
        label = label[len(MANAGEMENT_ROW_PREFIX) + 1:].strip()
    now = datetime.datetime.now().isoformat(timespec="seconds")
    data = _load_management_rows_data()
    entry = _ensure_management_row_column(
        data,
        product=product,
        column=column,
        label=label or column,
        actor=actor,
        now=now,
    )
    _save_management_rows_data(data)
    return {"ok": True, "column": entry["column"], "label": entry["label"], "columns": _management_row_columns_for_product(product)}


@router.post("/management-rows/values")
def save_management_row_values(req: ManagementRowValuesReq, request: Request = None):
    actor = req.username or ""
    if request is not None:
        me = current_user(request)
        actor = me.get("username") or actor
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product required")
    now = datetime.datetime.now().isoformat(timespec="seconds")
    data = _load_management_rows_data()
    values = data.setdefault("values", {})
    saved = 0
    deleted = 0
    rejected: list[str] = []
    for cell_key, raw_value in (req.values or {}).items():
        parts = str(cell_key or "").split("|", 2)
        if len(parts) != 3:
            rejected.append(str(cell_key))
            continue
        root_lot_id, wafer_id, column = [p.strip() for p in parts]
        if not root_lot_id or not wafer_id or not column.upper().startswith(f"{MANAGEMENT_ROW_PREFIX}_"):
            rejected.append(str(cell_key))
            continue
        _ensure_management_row_column(
            data,
            product=product,
            column=column,
            label=column[len(MANAGEMENT_ROW_PREFIX) + 1:] or column,
            actor=actor,
            now=now,
        )
        store_key = _management_row_value_key(product, root_lot_id, wafer_id, column)
        value = "" if raw_value is None else str(raw_value).strip()
        if value:
            values[store_key] = {"value": value, "username": actor, "updated": now}
            saved += 1
        elif store_key in values:
            values.pop(store_key, None)
            deleted += 1
    _save_management_rows_data(data)
    return {"ok": True, "saved": saved, "deleted": deleted, "rejected": rejected}


def _resolve_fab_source_target(fab_source: str):
    """Resolve a db-relative fab_source to an existing file or directory."""
    fab_source = _normalize_fab_source_path(fab_source)
    if not fab_source:
        return None, fab_source
    if fab_source.startswith("root:"):
        return None, fab_source
    aliases = [fab_source]
    parts = [p for p in fab_source.split("/") if p]
    if parts:
        head = parts[0].casefold()
        tail = "/".join(parts[1:])
        if head == _RAWDATA_FAB.casefold():
            aliases.append(_RAWDATA_EXACT + (f"/{tail}" if tail else ""))
        elif head == _RAWDATA_EXACT.casefold():
            aliases.append(_RAWDATA_FAB + (f"/{tail}" if tail else ""))
    db_base = _db_base()
    base_root = _base_root()
    fp = None
    matched = fab_source
    for root in (db_base, base_root):
        if not root or not root.exists():
            continue
        for rel in aliases:
            # v8.8.22: CI 경로 매칭 — fab_source 내 제품 폴더 대소문자 무시.
            # v9.0.6: 1.RAWDATA_DB_FAB/<PROD> 와 1.RAWDATA_DB/<PROD> 는 둘 다 FAB
            # history 로 취급한다. 운영 환경은 exact 이름만 쓰는 경우가 있다.
            cand = _find_ci_path(root, rel)
            if cand is not None and cand.exists():
                fp = cand
                matched = rel
                break
            for ext in (".parquet", ".csv"):
                cand2 = _find_ci_path(root, f"{rel}{ext}")
                if cand2 is not None and cand2.exists():
                    fp = cand2
                    matched = rel
                    break
            if fp:
                break
        if fp:
            break
    return fp, matched


def _rglob_files_ci(root: Path, suffixes: tuple[str, ...]) -> list[Path]:
    suffix_set = {s.casefold() for s in suffixes}
    try:
        cache_key = (str(root.resolve()), tuple(sorted(suffix_set)))
    except Exception:
        cache_key = (str(root), tuple(sorted(suffix_set)))
    now = time.monotonic()
    cached = _RGLOB_CACHE.get(cache_key)
    if cached and now - cached[0] < _DISCOVERY_CACHE_TTL_SEC:
        return list(cached[1])
    try:
        out = sorted(
            [p for p in root.rglob("*") if p.is_file() and p.suffix.casefold() in suffix_set],
            key=lambda p: str(p).casefold(),
        )
        _RGLOB_CACHE[cache_key] = (now, out)
        return list(out)
    except Exception:
        return []


def _first_data_file_ci(root: Path, suffixes: tuple[str, ...]) -> Path | None:
    suffix_set = {s.casefold() for s in suffixes}
    try:
        cache_key = (str(root.resolve()), tuple(sorted(suffix_set)))
    except Exception:
        cache_key = (str(root), tuple(sorted(suffix_set)))
    now = time.monotonic()
    cached = _FIRST_DATA_FILE_CACHE.get(cache_key)
    if cached and now - cached[0] < _DISCOVERY_CACHE_TTL_SEC:
        return cached[1]
    try:
        for p in root.rglob("*"):
            if p.is_file() and p.suffix.casefold() in suffix_set:
                _FIRST_DATA_FILE_CACHE[cache_key] = (now, p)
                return p
    except Exception:
        pass
    _FIRST_DATA_FILE_CACHE[cache_key] = (now, None)
    return None


def _canon_file_key(path) -> str:
    """Normalized path string for cross-source file identity (sig ↔ scan)."""
    try:
        return str(Path(path).resolve()).casefold()
    except Exception:
        return str(path).casefold()


def _scan_fab_source_raw(fab_source: str, only_files: set[str] | None = None):
    """Scan a fab_source without applying the long-format compatibility adapter.

    only_files: 증분 fab_lot_index 빌드용 — `_canon_file_key` 로 정규화한 경로
    집합에 든 파일만 스캔한다 (None = 전체)."""
    fp, fab_source = _resolve_fab_source_target(fab_source)
    if not fp:
        return None
    try:
        if fp.is_dir():
            parquets = _rglob_files_ci(fp, (".parquet",))
            csvs = _rglob_files_ci(fp, (".csv",))
            if only_files is not None:
                parquets = [p for p in parquets if _canon_file_key(p) in only_files]
                csvs = [p for p in csvs if _canon_file_key(p) in only_files]
            if not parquets and not csvs:
                return None
            frames = []
            # v8.8.5: 사내 `PRODA/date=YYYYMMDD/part_*.parquet` hive 레이아웃 대응.
            # hive_partitioning 을 켜서 경로의 `date=...` 를 컬럼으로 노출 → ts_col 자동 추론 시
            # `date` 후보가 적중해 "가장 최신 date 의 fab_col" join 이 자동으로 동작.
            if parquets:
                try:
                    frames.append(_cast_cats_lazy(_scan_parquet_compat(
                        [str(p) for p in parquets], hive_partitioning=True)))
                except TypeError:
                    # polars 구버전 — 파라미터 미지원 시 폴백 (경로 기반 파티션 컬럼 없음).
                    frames.append(_cast_cats_lazy(_scan_parquet_compat([str(p) for p in parquets])))
            # source discovery/signature는 CSV를 FAB 데이터 파일로 인정했지만 디렉터리
            # scan은 parquet만 읽었다. 운영 FAB가 CSV 파티션이면 파일 수는 잡히는데
            # 모든 배치가 None이 되어 "읽을 수 있는 FAB 행 없음"으로 끝났다.
            frames.extend(_cast_cats_lazy(pl.scan_csv(str(p), infer_schema_length=5000)) for p in csvs)
            if len(frames) == 1:
                return frames[0]
            return _cast_cats_lazy(pl.concat(frames, how="diagonal_relaxed"))
        if only_files is not None and _canon_file_key(fp) not in only_files:
            return None
        if fp.suffix.lower() == ".csv":
            return _cast_cats_lazy(pl.scan_csv(str(fp), infer_schema_length=5000))
        return _cast_cats_lazy(_scan_parquet_compat(str(fp)))
    except Exception as e:
        logger.warning("FAB source scan failed (source=%s target=%s) %s: %s",
                       fab_source, fp, type(e).__name__, e)
        return None


def _scan_fab_source(fab_source: str, only_files: set[str] | None = None):
    """v8.8.0: fab_source 가 가리키는 DB 경로를 LazyFrame 으로 스캔.
    - "FAB/PRODA" / "1.RAWDATA_DB/PRODA" 같은 디렉토리면 그 아래 모든 *.parquet 을 union 으로 스캔.
    - 단일 .parquet/.csv 파일이면 그 파일을 스캔.
    v8.8.21: "root:<name>" legacy prefix 는 제품 scope 를 넘어서므로 더 이상 지원하지 않음.
      저장된 값이 있어도 무시 → 호출측이 _auto_derive_fab_source 로 자동 매칭하도록 None 반환.
    실패 시 None 반환 (조용히 폴백).
    """
    lf_raw = _scan_fab_source_raw(fab_source, only_files=only_files)
    if lf_raw is None:
        return None
    # FAB canonical adapter:
    #   - 정식 FAB 는 wafer 단위 공정이력(root_lot_id/lot_id/wafer_id/step_id/tkin_time/tkout_time/eqp_id...).
    #   - 구 demo alias(eqp/chamber/time)가 섞여 있으면 runtime schema 에서만 정규화한다.
    #   - 아주 오래된 item/value FAB demo data 만 기존 최신 row adapter 로 축약한다.
    try:
        raw_names = lf_raw.collect_schema().names()
        from core.long_pivot import normalize_fab_history
        lf_raw = normalize_fab_history(lf_raw)
        names = lf_raw.collect_schema().names()
    except Exception:
        return lf_raw
    process_markers = {"eqp_id", "chamber_id", "ppid", "reticle_id", "tkout_time", "tkin_time"}
    legacy_process_aliases = {"eqp", "chamber", "ppid", "reticle_id", "tkout_time", "tkin_time"}
    raw_has_process_history = bool((process_markers | legacy_process_aliases) & set(raw_names))
    if "item_id" in names and "value" in names and "lot_id" in names and not raw_has_process_history:
        logger.info("_scan_fab_source: long-format 감지 → fab_lot_id adapter 적용 (source=%s)", fab_source)
        keep = [c for c in ("product", "root_lot_id", "lot_id", "wafer_id", "time") if c in names]
        lf_adapt = lf_raw.select(keep)
        if "time" in keep:
            lf_adapt = lf_adapt.sort("time", descending=True, nulls_last=True)
        renames = {}
        if "lot_id" in keep:
            renames["lot_id"] = "fab_lot_id"
        if "time" in keep:
            renames["time"] = "tkout_time"
        if renames:
            lf_adapt = lf_adapt.rename(renames)
        key_cols = [c for c in ("root_lot_id", "wafer_id") if c in keep]
        if key_cols:
            lf_adapt = lf_adapt.unique(subset=key_cols, keep="first", maintain_order=True)
        return lf_adapt
    return lf_raw


def _foreground_global_fab_scan_enabled() -> bool:
    return str(os.environ.get("FLOW_SPLITTABLE_FOREGROUND_GLOBAL_FAB_SCAN", "")).strip().lower() in {
        "1", "true", "yes", "on", "enabled"
    }


def _global_fab_source_paths(preferred_source: str = "", include_all: bool = True) -> list[str]:
    """Return db-relative FAB product folders to use for lot-id matching.

    SplitTable renders one ML_TABLE product, but fab_lot_id lineage can be
    present in a different FAB product folder.  Build the matching table from
    the whole FAB DB root, keeping the product-derived source first when it
    exists so current behavior remains the common fast path.
    """
    out: list[str] = []
    seen: set[str] = set()

    def add(rel: str):
        rel = _normalize_fab_source_path(rel)
        if not rel:
            return
        key = rel.casefold()
        if key in seen:
            return
        seen.add(key)
        out.append(rel)

    add(preferred_source)
    if not include_all:
        return out
    db_base = _db_base()
    try:
        db_base_resolved = db_base.resolve()
    except Exception:
        db_base_resolved = None

    for root_dir in _list_db_roots():
        up = root_dir.name.upper()
        is_fab_root = (
            up == _RAWDATA_FAB.upper()
            or up == _RAWDATA_EXACT.upper()
            or "FAB" in up
        )
        try:
            root_is_db_base = db_base_resolved is not None and root_dir.resolve() == db_base_resolved
        except Exception:
            root_is_db_base = False
        if not is_fab_root and not root_is_db_base:
            continue
        try:
            children = sorted(
                [p for p in root_dir.iterdir() if p.is_dir() and not p.name.startswith((".", "_", "__"))],
                key=lambda p: p.name.lower(),
            )
        except Exception:
            continue
        for child in children:
            if _first_data_file_ci(child, (".parquet", ".csv")) is None:
                continue
            if root_is_db_base:
                add(child.name)
            else:
                add(f"{root_dir.name}/{child.name}")
    return out


def _fab_source_product(source: str) -> str:
    """FAB source 경로에서 제품명을 뽑는다 (`1.RAWDATA_DB_FAB/PRODA` → `PRODA`).

    FAB 는 `<DB루트>/<제품>/date=.../*.parquet` 구조이고, 이 앱의 제품 결속
    규칙도 "FAB DB root 바로 아래 제품 폴더명"이다(`lot_progress_cache.metadata`).
    루트 자체가 source 인 형태(제품명만 오는 경우)도 마지막 세그먼트가 제품이다.
    """
    rel = _normalize_fab_source_path(source)
    if not rel:
        return ""
    return str(rel).replace("\\", "/").rstrip("/").rsplit("/", 1)[-1].strip()


def _normalized_fab_product_expr(expr):
    """Canonical, case-insensitive FAB product value used by cache/dashboard."""
    return (
        expr.cast(_STR, strict=False)
        .fill_null("")
        .str.strip_chars()
        .str.replace(r"(?i)^ML_TABLE_", "")
        .str.to_uppercase()
    )


def _scan_global_fab_sources(preferred_source: str = "", include_all: bool = True,
                             only_files: set[str] | None = None,
                             tag_source_product: bool = False):
    """Scan all FAB DB product folders as one LazyFrame for matching.

    only_files: 증분 fab_lot_index 빌드용 파일 부분집합 (None = 전체).

    tag_source_product: 각 행에 **그 행이 온 FAB 제품 폴더**를 실어 준다
    (`__cache_src_product`). 매칭은 일부러 FAB 전체를 훑기 때문에(랏 lineage 가
    다른 제품 폴더에 있을 수 있다) 결과만 보면 어느 제품 물량인지 알 수 없다 —
    대시보드 WIP 이 제품마다 똑같은 수를 내던 원인이다. 기본은 끔: 다른 호출자
    (fab_lot_index·view join)의 스키마를 바꾸지 않는다.
    """
    frames = []
    used_sources: list[str] = []
    for source in _global_fab_source_paths(preferred_source, include_all=include_all):
        lf = _scan_fab_source(source, only_files=only_files)
        if lf is None:
            continue
        if tag_source_product:
            lf = lf.with_columns(
                _normalized_fab_product_expr(pl.lit(_fab_source_product(source)))
                .alias(MATCH_CACHE_SRC_PRODUCT_COL)
            )
        frames.append(lf)
        used_sources.append(source)
    if not frames:
        return None, used_sources
    if len(frames) == 1:
        return frames[0], used_sources
    try:
        return _cast_cats_lazy(pl.concat(frames, how="diagonal_relaxed")), used_sources
    except Exception as e:
        logger.warning("_scan_global_fab_sources concat 실패 %s: %s", type(e).__name__, e)
        return frames[0], used_sources[:1]


# v8.8.3/v8.8.5: ML_TABLE_<PROD> → DB 상위폴더 자동 매칭.
#   `_list_db_roots()` 에 위임 — 사내 `1.RAWDATA_DB*` 접두 폴더도 인식 (FAB 힌트 우선).
# v8.8.17: root_dir 이 db_base 자체일 때(Case 1/3) 는 제품명만 반환 —
#   `_scan_fab_source` 에서 `db_base / fab_source` 로 해석하므로 prefix 중복 방지.
def _auto_derive_fab_source(product: str) -> str:
    """Return a fab_source path like "1.RAWDATA_DB_FAB/PRODA" (or legacy "FAB/PRODA") if auto-matchable, else "".
    ML_TABLE_ prefix 가 아니면 "" 반환 (오버라이드 off)."""
    p = _canonical_mltable_product_name(product)
    if not p:
        return ""
    pro = p[len("ML_TABLE_"):].strip()
    if not pro:
        return ""
    db_base = _db_base()
    roots = _list_db_roots()
    roots.sort(key=lambda r: _rank_db_root_name(r.name))
    for root_dir in roots:
        up = root_dir.name.upper()
        if up not in (_RAWDATA_EXACT.upper(), _RAWDATA_FAB.upper()) and not up.startswith(_RAWDATA_EXACT.upper() + "_"):
            continue
        # v8.8.22: CI 매칭 — 폴더가 ProdA/proda/PRODA 중 무엇이든 인식.
        cand = _find_ci_child(root_dir, pro)
        if cand is not None:
            actual = cand.name
            try:
                if root_dir.resolve() == db_base.resolve():
                    return actual
            except Exception:
                pass
            return f"{root_dir.name}/{actual}"
    return ""


# v8.8.3/v8.8.5/v9.0.4: ts_col / fab_col 자동 추론.
#   - 사용자가 기대하는 실사용 우선순위: tkout_time > time 계열 > date.
#   - date 는 hive 파티션 키(`date=YYYYMMDD`) 전용 마지막 fallback.
_TS_COL_CANDIDATES = ("tkout_time", "time", "out_ts", "ts", "timestamp", "created_at", "log_ts", "event_ts", "update_ts")
_FAB_COL_CANDIDATES = ("fab_lot_id", "lot_id", "fab_lotid", "fab_lot")
_RAW_TO_RUNTIME_ALIAS_CANDIDATES = {
    "lot_id": "fab_lot_id",
    "time": "tkout_time",
    "eqp": "eqp_id",
    "chamber": "chamber_id",
}


# v8.8.22: case-insensitive 컬럼 정렬.
#   ML_TABLE 은 대문자(ROOT_LOT_ID/WAFER_ID), hive 원천은 소문자(root_lot_id/wafer_id) 로
#   다르게 찍히는 경우가 있음. casefold 같으면 같은 컬럼으로 취급해야 join/override 가 동작.
#   → fab_lf 의 컬럼을 main_lf 쪽 casing 으로 rename 하여 이후 로직이 그대로 exact 매칭되게.
# v8.8.26: 충돌 가드 단순화 + rename 후 실제 스키마 재확인 (rename 이 lazy 상 silently 실패하는 사례 방지).
def _ci_align_fab_to_main(fab_lf, main_names):
    """Rename fab_lf columns to match main_names casing when casefold is equal.

    규칙:
      - fab 의 컬럼 fn (casefold=key) 이 main 의 target 과 casefold 일치하고 casing 만 다르면
        rename[fn] = target.
      - target 이 이미 fab 에 (별도의 distinct 컬럼으로) 존재하면 rename 을 skip (clobber 방지).
      - target 이 이번 rename 맵의 다른 항목에 의해 이미 소비됐으면 skip.
      - rename 후 실제 schema 를 재조회해 실패 여부 확인 — 실패 시 경고 로깅.

    Returns (aligned_lf, new_fab_names_list).
    """
    if fab_lf is None:
        return fab_lf, []
    try:
        fab_names = fab_lf.collect_schema().names()
    except Exception as e:
        logger.warning("_ci_align_fab_to_main: fab schema 조회 실패 %s: %s", type(e).__name__, e)
        return fab_lf, []
    main_ci = {n.casefold(): n for n in main_names}
    fab_set = set(fab_names)
    rename: dict = {}
    used_targets: set = set()
    for fn in fab_names:
        key = fn.casefold()
        target = main_ci.get(key)
        if not target or target == fn:
            continue
        # 단순화된 충돌 가드: target 이 fab 에 별개 컬럼으로 존재하면 rename 불가 (clobber).
        if target in fab_set:
            continue
        if target in used_targets:
            continue
        rename[fn] = target
        used_targets.add(target)
    if rename:
        try:
            fab_lf = fab_lf.rename(rename)
        except Exception as e:
            logger.warning("_ci_align_fab_to_main: rename 실패 %s: %s (rename=%s)",
                           type(e).__name__, e, rename)
            # rename 실패 시 원본 이름 유지
            return fab_lf, list(fab_names)
        # rename 이 적용됐는지 실제 스키마로 재확인.
        try:
            post = fab_lf.collect_schema().names()
            missing = [t for t in rename.values() if t not in post]
            if missing:
                logger.warning("_ci_align_fab_to_main: rename 후 target 누락 %s (post=%s...)",
                               missing, post[:20])
            return fab_lf, post
        except Exception as e:
            logger.warning("_ci_align_fab_to_main: post-schema 조회 실패 %s: %s",
                           type(e).__name__, e)
    new_names = [rename.get(n, n) for n in fab_names]
    return fab_lf, new_names


def _ci_resolve_in(name: str, pool):
    """Return the actual column name from pool matching `name` case-insensitively (exact first)."""
    if not name:
        return ""
    resolved = resolve_column(list(pool), name)
    return resolved.matched if resolved else ""


def _default_override_join_keys(main_names, fab_names):
    """Prefer root_lot_id + wafer_id by default; fall back only when necessary."""
    main_ci = {str(n).casefold(): n for n in (main_names or [])}
    fab_ci = {str(n).casefold(): n for n in (fab_names or [])}
    preferred = []
    for cand in ("root_lot_id", "wafer_id"):
        key = cand.casefold()
        if key in main_ci and key in fab_ci:
            preferred.append(main_ci[key])
    if preferred:
        return preferred
    fallback = []
    for cand in ("lot_id", "wafer_id"):
        key = cand.casefold()
        if key in main_ci and key in fab_ci:
            fallback.append(main_ci[key])
    return fallback


def _join_key_expr(col_name: str):
    """Normalize join key values so main/fab joins are case-insensitive and trim-safe."""
    return (
        pl.col(col_name)
        .cast(_STR, strict=False)
        .str.strip_chars()
        .str.to_uppercase()
    )


def _scope_match_cache_to_main_keys(main_lf, fab_lf, join_keys: list[str],
                                    join_tmp_keys: list[str]):
    """Keep FAB history belonging to the target ML_TABLE product.

    FAB는 제품 폴더 전체를 훑어야 한다. 랏이 공정 중 다른 제품 폴더로 이동할 수
    있기 때문이다. 하지만 그 결과를 target ML_TABLE의 root/wafer key로 제한하지
    않으면 모든 제품 cache가 전 FAB 행을 똑같이 갖게 된다. normalized key semi-join은
    폴더 이동 이력은 보존하면서 다른 제품의 랏만 제거한다.
    """
    if not join_keys or len(join_keys) != len(join_tmp_keys):
        return fab_lf
    main_names = set(main_lf.collect_schema().names())
    if any(key not in main_names for key in join_keys):
        return fab_lf
    main_keys = main_lf.select([
        _join_key_expr(key).alias(tmp) for key, tmp in zip(join_keys, join_tmp_keys)
    ])
    for tmp in join_tmp_keys:
        main_keys = main_keys.filter(pl.col(tmp).is_not_null() & (pl.col(tmp) != ""))
    main_keys = main_keys.unique(subset=join_tmp_keys)
    return fab_lf.join(main_keys, on=join_tmp_keys, how="semi")


def _contains_literal_ci_expr(col_name: str, needle: str):
    """Case-insensitive literal contains for LazyFrame autocomplete filters."""
    return (
        pl.col(col_name)
        .cast(_STR, strict=False)
        .str.to_uppercase()
        .str.contains(str(needle or "").strip().upper(), literal=True)
    )


def _apply_fab_scope_filters(fab_lf, fab_names, ov: dict, root_lot_id: str = "",
                             fab_lot_id: str = "", wafer_ids: str = "",
                             fab_col: str = ""):
    """Limit FAB source rows before latest-row picking and join."""
    root_scope = str(root_lot_id or "").strip()
    fab_scope = str(fab_lot_id or "").strip()
    wafer_scope = str(wafer_ids or "").strip()
    if root_scope:
        root_col = _resolve_source_col_name((ov.get("root_col") or "").strip(), fab_names) \
                   or _ci_resolve_in("root_lot_id", fab_names)
        if root_col:
            fab_lf = fab_lf.filter(_join_key_expr(root_col) == root_scope.upper())
    if fab_scope:
        target_fab_col = fab_col if fab_col in fab_names else _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_names)
        if target_fab_col:
            fab_lf = fab_lf.filter(_join_key_expr(target_fab_col) == fab_scope.upper())
    if wafer_scope:
        wf_col = _resolve_source_col_name((ov.get("wf_col") or ov.get("wafer_col") or "").strip(), fab_names) \
                 or _pick_first_present_ci(("wafer_id", "wafer"), fab_names)
        if wf_col:
            wf_list = [w.strip() for w in wafer_scope.split(",") if w.strip()]
            try:
                wf_ints = [int(w) for w in wf_list]
                wf_strs = set()
                for n in wf_ints:
                    wf_strs.update([str(n), f"{n:02d}", f"W{n}", f"W{n:02d}"])
                fab_lf = fab_lf.filter(
                    pl.col(wf_col).cast(_STR, strict=False).is_in(list(wf_strs))
                    | pl.col(wf_col).cast(pl.Int64, strict=False).is_in(wf_ints)
                )
            except ValueError:
                fab_lf = fab_lf.filter(pl.col(wf_col).cast(_STR, strict=False).is_in(wf_list))
    return fab_lf


# v8.8.16: hive 원천에서 끌어와 ML_TABLE 값을 덮어쓸 기본 컬럼 집합.
#   사내 `1.RAWDATA_DB*/<PROD>/date=*/*.parquet` 레이아웃에서 이 이름이 있으면 소스값으로 교체.
#   fab_col(보통 fab_lot_id) 는 레거시 단일 필드와 병합되어 override_cols 에 합류.
_DEFAULT_OVERRIDE_COLS = (
    "root_lot_id", "lot_id", "wafer_id", "line_id", "process_id", "step_id",
    "tkin_time", "tkout_time", "eqp_id", "chamber_id", "reticle_id", "ppid",
    # lot_type: FAB 가 주는 랏 구분(양산/엔지니어링/모니터 등). 대시보드에서 이
    # 축으로 물량을 나눠 보기 위해 매칭 캐시에 함께 싣는다. 이 목록은 FAB 스키마에
    # 실제로 있는 열만 통과하므로(`c in fab_names`), 없는 환경에서는 조용히 빠진다.
    "lot_type",
)


def _match_cache_refresh_minutes() -> int:
    data = load_json(PATHS.data_root / "settings.json", {})
    raw = data.get("splittable_match_refresh_minutes", MATCH_CACHE_REFRESH_MINUTES_DEFAULT) if isinstance(data, dict) else MATCH_CACHE_REFRESH_MINUTES_DEFAULT
    try:
        value = int(raw)
    except Exception:
        value = MATCH_CACHE_REFRESH_MINUTES_DEFAULT
    return max(MATCH_CACHE_REFRESH_MINUTES_MIN, min(MATCH_CACHE_REFRESH_MINUTES_MAX, value))


def _latest_lot_step_cache_path() -> Path:
    return _db_base() / "cache" / LATEST_LOT_STEP_CACHE_FILE


def _legacy_latest_lot_step_cache_path() -> Path:
    return _db_base() / "cache" / LEGACY_LATEST_LOT_STEP_CACHE_FILE


def _cleanup_legacy_latest_lot_step_cache() -> None:
    try:
        _legacy_latest_lot_step_cache_path().unlink(missing_ok=True)
    except Exception:
        pass


def _empty_latest_lot_step_frame() -> pl.DataFrame:
    return pl.DataFrame({col: [] for col in LATEST_LOT_STEP_CACHE_COLUMNS})


def _match_cache_state() -> dict:
    data = load_json(MATCH_CACHE_STATE_FILE, {}) if MATCH_CACHE_STATE_FILE.is_file() else {}
    return data if isinstance(data, dict) else {}


def _match_cache_global_fresh(now: float | None = None) -> dict:
    now = time.time() if now is None else float(now)
    state = _match_cache_state()
    last = 0.0
    try:
        last = float(state.get("last_refresh_epoch") or 0)
    except Exception:
        last = 0.0
    interval_s = _match_cache_refresh_minutes() * 60
    cache_fp = _latest_lot_step_cache_path()
    fresh = bool(last and cache_fp.is_file() and (now - last) < interval_s)
    return {
        "fresh": fresh,
        "last_refresh_epoch": last or None,
        "last_refresh_at": state.get("last_refresh_at") or "",
        "age_seconds": max(0, int(now - last)) if last else None,
        "next_refresh_at": datetime.datetime.fromtimestamp(last + interval_s).isoformat(timespec="seconds") if last else "",
        "cache_path": str(cache_fp),
        "cache_exists": cache_fp.is_file(),
        "interval_minutes": _match_cache_refresh_minutes(),
    }


def _mark_match_cache_refreshed(export_result: dict) -> None:
    now = time.time()
    state = {
        "last_refresh_epoch": now,
        "last_refresh_at": datetime.datetime.fromtimestamp(now).isoformat(timespec="seconds"),
        "cache_path": export_result.get("path") or str(_latest_lot_step_cache_path()),
        "row_count": int(export_result.get("row_count") or 0),
        "products": export_result.get("products") or [],
        "updated_at": export_result.get("cache_updated_at") or "",
    }
    save_json(MATCH_CACHE_STATE_FILE, state)


def _float_env_clamped(name: str, default: float, lo: float, hi: float) -> float:
    try:
        value = float(os.environ.get(name, "") or default)
    except Exception:
        value = default
    return max(lo, min(hi, value))


def _match_cache_product_pause_seconds() -> float:
    return _float_env_clamped("FLOW_SPLITTABLE_MATCH_CACHE_PRODUCT_PAUSE_SEC", 5.0, 0.0, 300.0)


def _match_cache_memory_wait_seconds() -> float:
    return _float_env_clamped("FLOW_SPLITTABLE_MATCH_CACHE_MEMORY_WAIT_SEC", 60.0, 5.0, 600.0)


def _match_cache_memory_max_wait_seconds() -> float:
    """매칭캐시 빌드 전 메모리 가드 대기의 최대 한도(초). 초과하면 무한 대기 대신
    경고 후 진행한다(빌드는 스트리밍이라 메모리 제한적, run_heavy 에 2차 가드도 있음).
    0 = 무제한(구동작). 기본 600초 — 개발서버에서 통합 스캔이 영원히 멈추는 것 방지."""
    return _float_env_clamped("FLOW_SPLITTABLE_MATCH_CACHE_MEMORY_MAX_WAIT_SEC", 600.0, 0.0, 21600.0)


def _match_cache_stream_enabled() -> bool:
    """FAB 매칭캐시 빌드를 root_lot_id 배치 스트리밍으로 처리할지 여부(기본 ON).
    글로벌 sort+unique 의 peak RAM 폭발을 막는다. 0/false 로 끄면 legacy 경로."""
    raw = str(os.environ.get("FLOW_MATCH_CACHE_STREAM", "1") or "1").strip().lower()
    return raw not in ("0", "false", "no", "off")


MATCH_CACHE_STREAM_BATCH_ROOTS_DEFAULT = 300


def _match_cache_stream_batch_roots() -> int:
    """root_lot_id 배치 스트리밍에서 한 번에 처리할 root(정규화 join key) 값 개수.
    작을수록 peak RAM 이 낮고 대신 FAB 원천을 여러 번 재스캔해 느려진다.
    개발서버(메모리 작음) OOM 방지를 위해 기본값은 보수적으로 잡는다.

    우선순위: env(FLOW_MATCH_CACHE_STREAM_BATCH_ROOTS) > 캐시관리 예산설정 톱니바퀴
    (match_cache_batch_roots[_dev], 운영/개발 분리) > 기본값 300."""
    env_raw = os.environ.get("FLOW_MATCH_CACHE_STREAM_BATCH_ROOTS", "")
    if str(env_raw).strip():
        try:
            return max(1, min(100000, int(env_raw)))
        except Exception:
            pass
    try:
        from core import cache_settings
        is_dev = bool(_ml_table_lookup._root_ram_cache_use_dev())
        v = cache_settings.get_int_role("match_cache_batch_roots", is_dev, None)
        if v is not None:
            return max(1, min(100000, int(v)))
    except Exception:
        pass
    return MATCH_CACHE_STREAM_BATCH_ROOTS_DEFAULT


def _match_cache_stream_log_gap_seconds() -> float:
    """배치 진행 로그 최소 간격(초). 배치가 빠를 때 이벤트 로그 폭주를 막는다."""
    return _float_env_clamped("FLOW_MATCH_CACHE_STREAM_LOG_GAP_SEC", 1.5, 0.0, 30.0)


def _match_cache_products(product: str = "") -> list[str]:
    raw = str(product or "").strip()
    if raw:
        return [raw]
    try:
        products = [p.get("name") for p in list_products().get("products", [])]
    except Exception:
        products = []
    return [p for p in products if p]


class MatchCacheCancelled(RuntimeError):
    """관리자가 이 제품의 캐싱을 중단했다.

    부분 산출은 버리고(promote/meta 기록 안 함) 다음 제품으로 넘어간다. 중단한
    제품은 재개 지점이 남지 않으므로 다음 스캔에서 처음부터 다시 빌드된다."""


def _match_cache_product_key(product: str) -> str:
    name = str(product or "").strip().upper()
    return name[len("ML_TABLE_"):] if name.startswith("ML_TABLE_") else name


def _match_cache_cancel_file() -> Path:
    """중단 신호 파일. 빌드가 개발 워커(별도 프로세스·서버)에서 돌 수도 있어
    프로세스 메모리만으로는 신호가 닿지 않는다 — 공유 data_root 를 거친다."""
    return MATCH_CACHE_DIR / "cancel.json"


def request_match_cache_cancel(product: str, by: str = "") -> dict:
    """이 제품의 진행 중 캐싱을 중단 요청한다."""
    key = _match_cache_product_key(product)
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _MATCH_CACHE_CANCEL_LOCK:
        _MATCH_CACHE_CANCEL.update({"product": key, "by": str(by or ""), "at": now})
    try:
        MATCH_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        save_json(_match_cache_cancel_file(), {"product": key, "by": str(by or ""), "at": now})
    except Exception as e:
        logger.warning("match cache cancel signal write failed: %s", e)
    return {"product": key, "by": str(by or ""), "at": now}


def clear_match_cache_cancel(product: str = "") -> None:
    """중단 신호 해제. product 를 주면 그 제품 신호일 때만 지운다 — 다음 제품을
    시작하면서 남의 중단 요청까지 지우지 않도록."""
    key = _match_cache_product_key(product)
    with _MATCH_CACHE_CANCEL_LOCK:
        if not key or _MATCH_CACHE_CANCEL.get("product") == key:
            _MATCH_CACHE_CANCEL.update({"product": "", "by": "", "at": ""})
    try:
        fp = _match_cache_cancel_file()
        if fp.is_file():
            current = load_json(fp, {}) or {}
            if not key or _match_cache_product_key(current.get("product") or "") == key:
                fp.unlink(missing_ok=True)
    except Exception:
        pass


def _match_cache_cancel_target() -> dict:
    """지금 중단 대상으로 지정된 제품. 프로세스 상태가 우선, 없으면 신호 파일."""
    with _MATCH_CACHE_CANCEL_LOCK:
        if _MATCH_CACHE_CANCEL.get("product"):
            return dict(_MATCH_CACHE_CANCEL)
    try:
        fp = _match_cache_cancel_file()
        if fp.is_file():
            raw = load_json(fp, {}) or {}
            key = _match_cache_product_key(raw.get("product") or "")
            if key:
                return {"product": key, "by": str(raw.get("by") or ""), "at": str(raw.get("at") or "")}
    except Exception:
        pass
    return {"product": "", "by": "", "at": ""}


def _match_cache_cancelled(product: str) -> bool:
    key = _match_cache_product_key(product)
    return bool(key) and _match_cache_cancel_target().get("product") == key


def _match_cache_job_status() -> dict:
    with _MATCH_CACHE_JOB_LOCK:
        out = dict(_MATCH_CACHE_JOB_STATE)
        out["products"] = [dict(r) for r in (_MATCH_CACHE_JOB_STATE.get("products") or [])]
        out["order"] = list(_MATCH_CACHE_JOB_STATE.get("order") or [])
    out["cancel"] = _match_cache_cancel_target()
    return out


def _match_cache_job_update(**updates) -> None:
    with _MATCH_CACHE_JOB_LOCK:
        _MATCH_CACHE_JOB_STATE.update(updates)


def _match_cache_job_append_products(rows: list[dict]) -> None:
    if not rows:
        return
    with _MATCH_CACHE_JOB_LOCK:
        current = [dict(r) for r in (_MATCH_CACHE_JOB_STATE.get("products") or [])]
        current.extend(dict(r) for r in rows)
        _MATCH_CACHE_JOB_STATE["products"] = current[-500:]
        _MATCH_CACHE_JOB_STATE["done"] = int(_MATCH_CACHE_JOB_STATE.get("done") or 0) + len(rows)
        _MATCH_CACHE_JOB_STATE["ok_count"] = int(_MATCH_CACHE_JOB_STATE.get("ok_count") or 0) + len([r for r in rows if r.get("ok")])
        # 관리자 중단은 실패가 아니다 — 별도로 센다.
        _MATCH_CACHE_JOB_STATE["failed_count"] = int(_MATCH_CACHE_JOB_STATE.get("failed_count") or 0) + len([r for r in rows if not r.get("ok") and not r.get("skipped") and not r.get("cancelled")])
        _MATCH_CACHE_JOB_STATE["cancelled_count"] = int(_MATCH_CACHE_JOB_STATE.get("cancelled_count") or 0) + len([r for r in rows if r.get("cancelled")])
        _MATCH_CACHE_JOB_STATE["skipped_count"] = int(_MATCH_CACHE_JOB_STATE.get("skipped_count") or 0) + len([r for r in rows if r.get("skipped")])
        for row in reversed(rows):
            if row.get("reason"):
                _MATCH_CACHE_JOB_STATE["last_error"] = str(row.get("reason") or "")
                break


def _begin_match_cache_job(products: list[str], force: bool, reason: str) -> tuple[bool, dict]:
    now = datetime.datetime.now().isoformat(timespec="seconds")
    with _MATCH_CACHE_JOB_LOCK:
        if _MATCH_CACHE_JOB_STATE.get("running"):
            status = dict(_MATCH_CACHE_JOB_STATE)
            status["products"] = [dict(r) for r in (_MATCH_CACHE_JOB_STATE.get("products") or [])]
            status["order"] = list(_MATCH_CACHE_JOB_STATE.get("order") or [])
            return False, status
        _MATCH_CACHE_JOB_STATE.clear()
        _MATCH_CACHE_JOB_STATE.update({
            "running": True,
            "queued": True,
            "force": bool(force),
            "reason": reason or "manual",
            "started_at": now,
            "finished_at": "",
            "current_product": "",
            "total": len(products),
            "done": 0,
            "ok_count": 0,
            "failed_count": 0,
            "skipped_count": 0,
            "cancelled_count": 0,
            "paused": False,
            "last_error": "",
            "products": [],
            "order": list(products),
        })
        status = dict(_MATCH_CACHE_JOB_STATE)
        status["products"] = []
        status["order"] = list(products)
        return True, status


def _wait_for_match_cache_memory() -> bool:
    try:
        from core.runtime_limits import process_memory_high, process_memory_snapshot
    except Exception:
        return True
    wait_s = _match_cache_memory_wait_seconds()
    max_wait = _match_cache_memory_max_wait_seconds()
    waited = 0.0
    while not _MATCH_CACHE_STOP.is_set():
        try:
            high = process_memory_high()
            snap = process_memory_snapshot()
        except Exception:
            return True
        if not high:
            _match_cache_job_update(paused=False, memory=snap)
            return True
        # 무한 대기 방지 — 메모리 가드가 max_wait 넘게 계속 high 면 경고 후 진행한다.
        # (스트리밍 빌드라 메모리 제한적이고, run_heavy 로컬 실행에 2차 메모리 admission
        #  가드가 또 있다.) 개발서버에서 통합 스캔이 stage 1 에서 영원히 멈추던 문제 해결.
        if max_wait > 0 and waited >= max_wait:
            logger.warning(
                "match cache memory guard high for %.0fs — 대기 한도 초과, 진행함 "
                "(snap=%s)", waited, snap,
            )
            _match_cache_job_update(paused=False, memory=snap)
            return True
        _match_cache_job_update(paused=True, memory=snap)
        _MATCH_CACHE_STOP.wait(wait_s)
        waited += wait_s
    return False


def _write_match_cache_lazyframe(q, tmp: Path) -> int:
    """Write cache output with the lowest available peak-memory path."""
    try:
        tmp.unlink(missing_ok=True)
    except Exception:
        pass
    try:
        q.sink_parquet(str(tmp))
        try:
            return int(pl.scan_parquet(str(tmp)).select(pl.len().alias("row_count")).collect().item(0, 0))
        except Exception:
            try:
                return int(pl.read_parquet(str(tmp)).height)
            except Exception:
                return 0
    except Exception:
        try:
            tmp.unlink(missing_ok=True)
        except Exception:
            pass
    df = None
    try:
        try:
            from core.parquet_perf import collect_streaming
            df = collect_streaming(q)
        except Exception:
            df = q.collect()
        df.write_parquet(tmp)
        return int(df.height)
    finally:
        try:
            del df
        except Exception:
            pass
        try:
            gc.collect()
        except Exception:
            pass


def _current_fab_override(product: str) -> tuple[str, dict, str]:
    ml_product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    cfg = load_json(SOURCE_CFG, {}) if SOURCE_CFG.exists() else {}
    ov = _lot_override_for(cfg, ml_product)
    fab_source = _normalize_fab_source_path((ov.get("fab_source") or "").strip())
    if fab_source.startswith("root:"):
        fab_source = ""
    if not fab_source:
        fab_source = _auto_derive_fab_source(ml_product)
    return ml_product, ov, fab_source


def _match_cache_path(product: str) -> Path:
    name = safe_id(_canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip() or "product")
    return MATCH_CACHE_DIR / f"{name}.parquet"


def _match_cache_meta_path(product: str) -> Path:
    name = safe_id(_canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip() or "product")
    return MATCH_CACHE_DIR / f"{name}.json"


def _match_cache_config_key(product: str, ov: dict, fab_source: str) -> str:
    keys = ("root_col", "wf_col", "wafer_col", "fab_col", "ts_col", "join_keys", "override_cols")
    clean_ov = {k: ov.get(k) for k in keys if k in ov}
    payload = {
        "version": MATCH_CACHE_VERSION,
        "product": _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip(),
        "fab_source": _normalize_fab_source_path(fab_source),
        "fab_sources": _global_fab_source_paths(fab_source),
        "db_root": str(_db_base()),
        "base_root": str(_base_root()),
        "override": clean_ov,
    }
    try:
        return json.dumps(payload, sort_keys=True, ensure_ascii=False, default=str)
    except Exception:
        return str(payload)


def _match_cache_current(product: str) -> dict | None:
    ml_product, ov, fab_source = _current_fab_override(product)
    if not ml_product:
        return None
    if not fab_source and not _global_fab_source_paths(""):
        return None
    fp = _match_cache_path(ml_product)
    meta_fp = _match_cache_meta_path(ml_product)
    if not fp.is_file() or not meta_fp.is_file():
        return None
    meta = load_json(meta_fp, {})
    if not isinstance(meta, dict):
        return None
    if meta.get("version") != MATCH_CACHE_VERSION:
        return None
    if meta.get("config_key") != _match_cache_config_key(ml_product, ov, fab_source):
        return None
    try:
        lf = _cast_cats_lazy(_scan_parquet_compat(str(fp)))
    except Exception as e:
        logger.warning("SplitTable match cache scan failed (product=%s) %s: %s",
                       ml_product, type(e).__name__, e)
        return None
    return {"product": ml_product, "ov": ov, "fab_source": fab_source, "path": fp, "meta": meta, "lf": lf}


def _match_cache_response_meta(product: str) -> dict:
    """Small response payload for UI badges and Agent trace tables."""
    status = _latest_lot_step_cache_status(product)
    if status.get("cache_exists") and int(status.get("product_row_count") or status.get("row_count") or 0) > 0:
        return {
            "hit": True,
            "source": "lot_progress_latest_cache",
            "product": _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip(),
            "fab_source": "lot_progress_latest_lot_by_root_wafer",
            "path": status.get("cache_path") or str(_latest_lot_step_cache_path()),
            "built_at": status.get("updated_at") or status.get("latest_updated_at") or "",
            "row_count": int(status.get("product_row_count") or status.get("row_count") or 0),
            "join_keys": ["root_lot_id", "wafer_id"],
            "override_cols": ["lot_id", "fab_lot_id"],
            "fab_col": "lot_id",
            "ts_col": "tkout_time",
            "dedup_keys": ["product", "root_lot_id", "wafer_id"],
        }
    return {"hit": False, "source": "lot_progress_latest_cache"}


def _ensure_match_cache_current(product: str, *, force: bool = False) -> dict | None:
    """Ensure the product FAB projection is persisted before falling back to raw scan."""
    current = _match_cache_current(product)
    if current:
        return current
    ml_product, ov, fab_source = _current_fab_override(product)
    if not ml_product:
        return None
    if not fab_source and not _global_fab_source_paths(""):
        return None
    config_key = _match_cache_config_key(ml_product, ov, fab_source)
    missed = _MATCH_CACHE_AUTO_BUILD_MISS.get(ml_product)
    now = time.time()
    if (
        not force
        and missed
        and missed[1] == config_key
        and now - missed[0] < _MATCH_CACHE_AUTO_BUILD_MISS_TTL_SEC
    ):
        return None
    try:
        result = _refresh_match_cache_products([ml_product], force=force)
        if not result.get("ok"):
            _MATCH_CACHE_AUTO_BUILD_MISS[ml_product] = (now, config_key)
            return None
        _MATCH_CACHE_AUTO_BUILD_MISS.pop(ml_product, None)
        return _match_cache_current(ml_product)
    except Exception as e:
        logger.warning("SplitTable match cache auto-build failed (product=%s) %s: %s",
                       ml_product, type(e).__name__, e, exc_info=True)
        _MATCH_CACHE_AUTO_BUILD_MISS[ml_product] = (now, config_key)
        return None


def _latest_cache_product_values(product: str) -> set[str]:
    raw = str(product or "").strip()
    if not raw:
        return set()
    canonical = _canonical_mltable_product_name(raw, allow_bare=True) or raw
    values = {raw.upper(), canonical.upper()}
    if canonical.upper().startswith("ML_TABLE_"):
        bare = canonical[len("ML_TABLE_"):].strip()
        if bare:
            values.add(bare.upper())
    else:
        values.add(f"ML_TABLE_{canonical}".upper())
    return values


def _latest_lot_step_cache_lf(product: str = "", root_lot_id: str = ""):
    # Per-root fast path (SplitTable pivot-cache 방식): root 검색이면 해당 root
    # 파티션만 읽는다. 파티션이 stale/miss 면 monolithic 풀스캔으로 폴백.
    if str(root_lot_id or "").strip():
        part_lf = _latest_lot_index_partition_lf(product, root_lot_id)
        if part_lf is not None:
            return part_lf
    fp = _latest_lot_step_cache_path()
    if not fp.is_file():
        return None
    try:
        lf = _cast_cats_lazy(_scan_parquet_compat(str(fp)))
        names = lf.collect_schema().names()
    except Exception as e:
        logger.warning("SplitTable latest lot-step cache scan failed (%s) %s: %s",
                       fp, type(e).__name__, e)
        return None
    if LATEST_LOT_STEP_CACHE_FORMAT_COLUMN not in names:
        logger.info("SplitTable ignores legacy latest cache without format version: %s", fp)
        return None
    lf = lf.filter(
        pl.col(LATEST_LOT_STEP_CACHE_FORMAT_COLUMN).cast(pl.Int64, strict=False)
        == LATEST_LOT_STEP_CACHE_FORMAT_VERSION
    )
    if product and "product" in names:
        values = _latest_cache_product_values(product)
        if values:
            lf = lf.filter(pl.col("product").cast(_STR, strict=False).str.to_uppercase().is_in(sorted(values)))
    return lf


# ── Per-root latest-lot cache partitions ─────────────────────────────────────
# The canonical latest-lot cache (lot_progress_latest_lot_by_root_wafer.parquet)
# is a single monolithic file, so every root-scoped lookup (the fab identity
# join in _scan_product, fab-lot snapshots, history scope) re-scanned the whole
# file with a cast+upper filter that defeats parquet predicate pushdown. The
# per-root partition layout is owned by core.latest_lot_partitions and is
# written by BOTH monolithic exporters at write time, so a root search normally
# reads a fresh partition directly. The freshness check + enqueue below remain
# as self-heal only (crash mid-write, files produced by older code): on any
# miss/stale/error the caller falls back to the monolithic scan while a
# rebuild is scheduled.
_LATEST_IDX_ROOT_COL = _latest_lot_partitions.ROOT_KEY_COL
_LATEST_IDX_DIR_NAME = _latest_lot_partitions.PARTITION_DIR_NAME
_LATEST_IDX_META_FILE = _latest_lot_partitions.META_FILE
_LATEST_IDX_BUILD_LOCK = threading.Lock()
_LATEST_IDX_BUILD_STATE: dict = {"inprogress": False, "last": 0.0}
_LATEST_IDX_BUILD_COOLDOWN_SEC = 60.0
_LATEST_IDX_FRESH_TTL_SEC = 2.0
_LATEST_IDX_FRESH_LOCK = threading.Lock()
_LATEST_IDX_FRESH_CACHE: dict[str, tuple[float, bool]] = {}


def _latest_lot_index_enabled() -> bool:
    return _env_bool("FLOW_SPLITTABLE_LATEST_LOT_INDEX", True)


def _latest_lot_index_dir() -> Path:
    return _latest_lot_partitions.partitions_dir(_latest_lot_step_cache_path())


def _latest_lot_index_meta_path() -> Path:
    return _latest_lot_partitions.meta_path(_latest_lot_step_cache_path())


def _latest_lot_index_source_sig() -> list:
    """(path, mtime, size) of the monolithic file — the partition staleness key."""
    return _latest_lot_partitions.source_signature(_latest_lot_step_cache_path())


def _latest_lot_index_fresh() -> bool:
    """True → 파티션 세트가 현재 monolithic 파일과 정확히 일치.

    stale/miss 면 백그라운드 재빌드를 예약하고 False 를 반환한다 (호출측은
    monolithic 폴백 — 오늘과 동일한 경로라 정확성 저하 없음). 판정은 짧은 TTL
    로 캐시해 요청마다 meta 재읽기/재-stat 을 피한다. TTL 캐시는 monolithic
    경로를 키로 쓴다 — DB 루트가 런타임에 재지정되면(관리자 설정/테스트 sandbox)
    이전 루트의 fresh 판정이 새 루트로 새어 빈 파티션 응답을 내면 안 된다."""
    if not _latest_lot_index_enabled():
        return False
    mono_fp = _latest_lot_step_cache_path()
    cache_key = str(mono_fp)
    now = time.monotonic()
    with _LATEST_IDX_FRESH_LOCK:
        cached = _LATEST_IDX_FRESH_CACHE.get(cache_key)
        if cached is not None and now - cached[0] < _LATEST_IDX_FRESH_TTL_SEC:
            return cached[1]
    fresh = False
    try:
        meta = load_json(_latest_lot_index_meta_path(), {}) or {}
        fresh = bool(meta) and meta.get("source_sig") == _latest_lot_index_source_sig()
    except Exception:
        fresh = False
    if not fresh and mono_fp.is_file():
        _enqueue_latest_lot_index_build(reason="stale")
    with _LATEST_IDX_FRESH_LOCK:
        _LATEST_IDX_FRESH_CACHE[cache_key] = (now, fresh)
        while len(_LATEST_IDX_FRESH_CACHE) > 8:
            _LATEST_IDX_FRESH_CACHE.pop(next(iter(_LATEST_IDX_FRESH_CACHE)))
    return fresh


def _latest_lot_index_partition_lf(product: str, root_lot_id: str):
    """Return the one-root LazyFrame from the partitioned latest-lot cache, or
    None to signal fallback to the monolithic scan."""
    root = str(root_lot_id or "").strip().upper()
    if not root or not _latest_lot_index_fresh():
        return None
    try:
        part = _latest_lot_index_dir() / f"{_LATEST_IDX_ROOT_COL}={root}"
        if not part.is_dir():
            # 파티션 세트가 fresh 인데 이 root 파티션이 없다 → monolithic 에도
            # 이 root 는 없다. 풀스캔 폴백은 같은 결과를 느리게 낼 뿐이므로
            # 빈 프레임으로 즉시 응답한다. 단, 특수문자 root 는 파티션 디렉터리명
            # 인코딩이 다를 수 있으므로 단정하지 않고 monolithic 폴백.
            if _re.fullmatch(r"[A-Z0-9_\-.]+", root):
                return _empty_latest_lot_step_frame().lazy()
            return None
        files = sorted(part.glob("*.parquet"))
        if not files:
            return None
        lf = _scan_parquet_compat([str(p) for p in files])
        names = lf.collect_schema().names()
        if LATEST_LOT_STEP_CACHE_FORMAT_COLUMN not in names:
            return None
        lf = lf.filter(
            pl.col(LATEST_LOT_STEP_CACHE_FORMAT_COLUMN).cast(pl.Int64, strict=False)
            == LATEST_LOT_STEP_CACHE_FORMAT_VERSION
        )
        if _LATEST_IDX_ROOT_COL in names:
            lf = lf.drop(_LATEST_IDX_ROOT_COL)
        lf = _cast_cats_lazy(lf)
        if product and "product" in names:
            values = _latest_cache_product_values(product)
            if values:
                lf = lf.filter(pl.col("product").cast(_STR, strict=False).str.to_uppercase().is_in(sorted(values)))
        return lf
    except Exception:
        logger.debug("latest_lot_index scan failed root=%s", root, exc_info=True)
        return None


def _build_latest_lot_index(reason: str = "reader_self_heal") -> bool:
    """Re-partition the monolithic latest-lot cache by normalized root key."""
    ok = _latest_lot_partitions.sync_partitions(
        _latest_lot_step_cache_path(), reason=reason)
    if ok:
        with _LATEST_IDX_FRESH_LOCK:
            _LATEST_IDX_FRESH_CACHE.clear()
    return ok


def _enqueue_latest_lot_index_build(reason: str = "") -> bool:
    """Single-flight, cooldown-guarded background rebuild of the root partitions.

    Self-heal only — the exporters write the partitions synchronously, so this
    fires just for crash-truncated layouts or files written by older code."""
    if not _latest_lot_index_enabled():
        return False
    now = time.time()
    with _LATEST_IDX_BUILD_LOCK:
        if _LATEST_IDX_BUILD_STATE.get("inprogress"):
            return False
        if now - float(_LATEST_IDX_BUILD_STATE.get("last") or 0.0) < _LATEST_IDX_BUILD_COOLDOWN_SEC:
            return False
        _LATEST_IDX_BUILD_STATE["inprogress"] = True

    def _run():
        try:
            _build_latest_lot_index(reason=reason or "reader_self_heal")
        except Exception as exc:
            logger.warning("latest_lot_index build failed (%s): %s", reason, exc)
        finally:
            with _LATEST_IDX_BUILD_LOCK:
                _LATEST_IDX_BUILD_STATE["inprogress"] = False
                _LATEST_IDX_BUILD_STATE["last"] = time.time()

    threading.Thread(target=_run, daemon=True, name="splittable-latestidx").start()
    logger.info("latest_lot_index build queued (%s)", reason)
    return True


def _filter_latest_lot_step_cache(lf, *, root_lot_id: str = "", fab_lot_id: str = "",
                                  wafer_ids: str = ""):
    try:
        names = lf.collect_schema().names()
    except Exception:
        names = []
    root_scope = str(root_lot_id or "").strip()
    fab_scope = str(fab_lot_id or "").strip()
    wafer_scope = str(wafer_ids or "").strip()
    if root_scope and "root_lot_id" in names:
        lf = lf.filter(_join_key_expr("root_lot_id") == root_scope.upper())
    if fab_scope:
        fab_filters = []
        if "lot_id" in names:
            fab_filters.append(_join_key_expr("lot_id") == fab_scope.upper())
        if fab_filters:
            expr = fab_filters[0]
            for item in fab_filters[1:]:
                expr = expr | item
            lf = lf.filter(expr)
    if wafer_scope and "wafer_id" in names:
        wf_list = [w.strip() for w in wafer_scope.split(",") if w.strip()]
        try:
            wf_ints = [int(w) for w in wf_list]
            wf_strs = set()
            for n in wf_ints:
                wf_strs.update([str(n), f"{n:02d}", f"W{n}", f"W{n:02d}"])
            lf = lf.filter(
                pl.col("wafer_id").cast(_STR, strict=False).is_in(list(wf_strs))
                | pl.col("wafer_id").cast(pl.Int64, strict=False).is_in(wf_ints)
            )
        except ValueError:
            lf = lf.filter(pl.col("wafer_id").cast(_STR, strict=False).is_in(wf_list))
    return lf


def _latest_lot_step_cache_source(product: str, current: dict | None = None) -> str:
    return "lot_progress_latest_cache"


def _fab_history_scope_from_latest_cache(product: str, root_lot_id: str = "", fab_lot_id: str = "",
                                         prefix: str = "", limit: int = 500) -> dict | None:
    cache_lf = _latest_lot_step_cache_lf(product, root_lot_id=root_lot_id)
    if cache_lf is None:
        return None
    source = _latest_lot_step_cache_source(product)
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return None
    if not {"root_lot_id", "lot_id"}.issubset(set(names)):
        return None
    q = _filter_latest_lot_step_cache(
        cache_lf,
        root_lot_id=root_lot_id,
        fab_lot_id=fab_lot_id,
    ).select([
        pl.col("root_lot_id").cast(_STR, strict=False).alias("root"),
        pl.col("lot_id").cast(_STR, strict=False).alias("fab"),
        *([pl.col("wafer_id").cast(_STR, strict=False).alias("wafer")] if "wafer_id" in names else []),
    ]).filter(pl.col("root").is_not_null() & pl.col("fab").is_not_null())
    root_scope = str(root_lot_id or "").strip()
    fab_scope = str(fab_lot_id or "").strip()
    if fab_scope:
        q = q.filter(_join_key_expr("fab") == fab_scope.upper())
    elif str(prefix or "").strip():
        q = q.filter(_contains_literal_ci_expr("fab", prefix))
    try:
        fabs = _limited_unique_values(
            q,
            "fab",
            prefix="",
            limit=limit,
            preview_only=not bool(root_scope or fab_scope or str(prefix or "").strip()),
        )
        roots: list[str] = [root_scope] if root_scope else []
        wafers: list[str] = []
        if fab_scope and fabs:
            meta_cols = [pl.col("root")]
            if "wafer" in q.collect_schema().names():
                meta_cols.append(pl.col("wafer"))
            meta_df = q.select(meta_cols).unique().collect()
            roots = sorted({s for s in (_clean_str(v) for v in meta_df["root"].to_list()) if s})
            if "wafer" in meta_df.columns:
                wafers = sorted({s for s in (_clean_str(v) for v in meta_df["wafer"].to_list()) if s}, key=_wafer_sort_key)
    except Exception as e:
        logger.warning("_fab_history_scope_from_latest_cache 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return None
    return {
        "candidates": fabs,
        "root_ids": roots,
        "wafer_ids": wafers,
        "source": source,
        "cache": True,
        "query_ok": True,
    }


def _fab_history_root_candidates_from_latest_cache(product: str, prefix: str = "", limit: int = 500) -> dict | None:
    cache_lf = _latest_lot_step_cache_lf(product)
    if cache_lf is None:
        return None
    source = _latest_lot_step_cache_source(product)
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return None
    if "root_lot_id" not in names:
        return None
    try:
        values = _limited_unique_values(cache_lf, "root_lot_id", prefix=prefix, limit=limit)
    except Exception as e:
        logger.warning("_fab_history_root_candidates_from_latest_cache 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return None
    return {"candidates": values, "source": source, "cache": True}


def _fab_lot_snapshot_from_latest_cache(product: str, root_lot_id: str, wafer_id: str = "") -> str:
    root = str(root_lot_id or "").strip()
    if not root:
        return ""
    cache_lf = _latest_lot_step_cache_lf(product, root_lot_id=root)
    if cache_lf is None:
        return ""
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return ""
    if "lot_id" not in names:
        return ""
    q = (
        _filter_latest_lot_step_cache(cache_lf, root_lot_id=root, wafer_ids=str(wafer_id or ""))
        .select([
            pl.col("lot_id").cast(_STR, strict=False).alias("fab"),
            *([pl.col("tkout_time").cast(_STR, strict=False).alias("ts")] if "tkout_time" in names else []),
        ])
        .filter(pl.col("fab").is_not_null() & (pl.col("fab") != ""))
    )
    if "ts" in q.collect_schema().names():
        q = q.sort("ts", descending=True, nulls_last=True)
    else:
        q = q.sort("fab")
    try:
        df = q.head(1).collect()
    except Exception as e:
        logger.warning("_fab_lot_snapshot_from_latest_cache 실패 (product=%s root=%s wafer=%s) %s: %s",
                       product, root_lot_id, wafer_id, type(e).__name__, e)
        return ""
    if df.is_empty():
        return ""
    return _clean_str(df.item(0, 0))


def _latest_cache_src_product_expr(names: list[str], fallback_product: str):
    """latest cache 의 `src_product` 열 — **그 행이 온 FAB 제품 폴더**.

    `product`(ML_TABLE 제품)와 **의미가 다르고 둘 다 필요하다**:

    - `product` — "이 ML_TABLE 제품 화면에서 도달 가능한 랏". 매칭은 일부러 FAB
      전체를 훑으므로, 자기 FAB 폴더가 없는 제품도 다른 폴더에서 lot lineage 를
      찾아온다. SplitTable 의 fab lot 라벨이 이 의미에 의존한다.
    - `src_product` — "이 랏이 실제로 속한 FAB 제품 폴더". 물량을 세는 대시보드
      WIP 은 이걸 봐야 한다. 예전에는 이 구분이 없어 모든 제품이 같은 랏을 자기
      것으로 세었고, 제품마다 똑같은 wafer 수가 나왔다(2026-07-31).

    출처 열이 없는 옛 캐시는 `product` 와 같은 값으로 채운다 — 읽는 쪽이 항상
    `src_product` 를 쓸 수 있게 하되 동작은 예전과 같아진다.
    """
    fallback = _normalized_fab_product_expr(pl.lit(fallback_product))
    if MATCH_CACHE_SRC_PRODUCT_COL not in names:
        return fallback.alias("src_product")
    src = _normalized_fab_product_expr(pl.col(MATCH_CACHE_SRC_PRODUCT_COL))
    return (
        pl.when(src == "").then(fallback).otherwise(src).alias("src_product")
    )


def export_latest_lot_step_cache(products: list[str] | None = None, *, update_state: bool = False) -> dict:
    """Export product match caches into the canonical latest lot/step parquet."""
    raw_products = [p for p in (products or _match_cache_products("")) if p]
    cache_updated_at = datetime.datetime.now().isoformat(timespec="seconds")
    frames = []
    exported_products: list[str] = []
    skipped: list[dict] = []
    for raw_product in raw_products:
        current = _match_cache_current(raw_product)
        if not current:
            skipped.append({"product": raw_product, "reason": "match_cache_missing"})
            continue
        lf = current.get("lf")
        product = _normalize_latest_cache_product(current.get("product") or raw_product)
        try:
            names = lf.collect_schema().names()
        except Exception as e:
            skipped.append({"product": product, "reason": f"schema_failed: {type(e).__name__}"})
            continue
        if MATCH_CACHE_ROOT_COL not in names or MATCH_CACHE_FAB_COL not in names:
            skipped.append({"product": product, "reason": "required_columns_missing"})
            continue
        lot_type_col = _ci_resolve_in("lot_type", names)
        exprs = [
            pl.lit(LATEST_LOT_STEP_CACHE_FORMAT_VERSION).cast(pl.Int16).alias(LATEST_LOT_STEP_CACHE_FORMAT_COLUMN),
            pl.lit(LATEST_LOT_STEP_CACHE_SOURCE).alias(LATEST_LOT_STEP_CACHE_SOURCE_COLUMN),
            # product 는 예전 그대로 ML_TABLE 제품 — SplitTable 의 교차 폴더
            # lineage 조회가 이 의미에 의존한다. 실제 소속은 src_product 로 따로.
            pl.lit(product).alias("product"),
            _latest_cache_src_product_expr(names, product),
            pl.col(MATCH_CACHE_ROOT_COL).cast(_STR, strict=False).alias("root_lot_id"),
            (
                pl.col(MATCH_CACHE_WAFER_COL).cast(_STR, strict=False).alias("wafer_id")
                if MATCH_CACHE_WAFER_COL in names else pl.lit("").alias("wafer_id")
            ),
            pl.col(MATCH_CACHE_FAB_COL).cast(_STR, strict=False).alias("lot_id"),
            (
                pl.col("step_id").cast(_STR, strict=False).alias("step_id")
                if "step_id" in names else pl.lit("").alias("step_id")
            ),
            pl.lit("").alias("function_step"),
            # 매칭 캐시는 (root, wafer) 별 FAB 최신 행을 들고 있고 아래에서
            # tkout_time 내림차순 unique 로 한 번 더 최신만 남기므로, 여기서
            # 그대로 실으면 "가장 최신 FAB 행의 lot_type" 이 된다.
            (
                # 앞뒤 공백은 여기서 없앤다 — 대시보드 lot_type 필터가 원본 문자열을
                # 그대로 구분해서, "BUILD" 와 "BUILD " 가 눈에 똑같은 항목 두 개로
                # 나오고 한쪽을 고르면 나머지 wafer 가 사라졌다.
                pl.col(lot_type_col).cast(_STR, strict=False).fill_null("")
                  .str.strip_chars().alias("lot_type")
                if lot_type_col else pl.lit("").alias("lot_type")
            ),
            (
                pl.col(MATCH_CACHE_TS_COL).cast(_STR, strict=False).alias("tkout_time")
                if MATCH_CACHE_TS_COL in names else pl.lit("").alias("tkout_time")
            ),
            pl.lit(cache_updated_at).alias("update_time"),
        ]
        frames.append(lf.select(exprs))
        exported_products.append(product)
    fp = _latest_lot_step_cache_path()
    fp.parent.mkdir(parents=True, exist_ok=True)
    _cleanup_legacy_latest_lot_step_cache()
    if frames:
        q = pl.concat(frames)
        q = q.filter(
            pl.col("product").is_not_null()
            & (pl.col("product") != "")
            & pl.col("root_lot_id").is_not_null()
            & (pl.col("root_lot_id") != "")
            & pl.col("lot_id").is_not_null()
            & (pl.col("lot_id") != "")
        )
        q = (
            q.sort("tkout_time", descending=True, nulls_last=True)
             .unique(subset=["product", "root_lot_id", "wafer_id"], keep="first", maintain_order=True)
             .sort(["product", "root_lot_id", "wafer_id"])
        )
        try:
            from core.parquet_perf import collect_streaming
            df = collect_streaming(q)
        except Exception:
            df = q.collect()
        function_steps = []
        step_meta_cache: dict[tuple[str, str], str] = {}
        try:
            from core.lot_step import lookup_step_meta
        except Exception:
            lookup_step_meta = None
        for product_value, step_value in df.select(["product", "step_id"]).iter_rows():
            product_text = str(product_value or "")
            step_text = str(step_value or "").strip()
            key = (product_text, step_text)
            if key not in step_meta_cache:
                meta = lookup_step_meta(product=product_text, step_id=step_text) if lookup_step_meta and step_text else {}
                step_meta_cache[key] = str((meta or {}).get("function_step") or (meta or {}).get("func_step") or "")
            function_steps.append(step_meta_cache[key])
        df = df.with_columns(pl.Series("function_step", function_steps)).select(LATEST_LOT_STEP_CACHE_COLUMNS)
    else:
        df = _empty_latest_lot_step_frame()
    tmp = fp.with_suffix(fp.suffix + ".tmp")
    try:
        tmp.unlink(missing_ok=True)
    except Exception:
        pass
    df.write_parquet(tmp)
    tmp.replace(fp)
    # per-root 파티션을 같은 쓰기 시점에 동기화 — df 가 손에 있으므로 read-back
    # 없이 즉시 파티션이 fresh 가 된다. 실패해도 reader 의 monolithic 폴백 +
    # self-heal 재빌드가 있으므로 export 는 성공으로 처리한다.
    try:
        _latest_lot_partitions.sync_partitions(fp, df=df, reason="match_cache_export")
    except Exception as e:
        logger.warning("latest-lot per-root partition sync failed %s: %s",
                       type(e).__name__, e)
    with _LATEST_IDX_FRESH_LOCK:
        _LATEST_IDX_FRESH_CACHE.clear()
    result = {
        "ok": True,
        "path": str(fp),
        "row_count": int(df.height),
        "products": exported_products,
        "skipped": skipped,
        "cache_updated_at": cache_updated_at,
    }
    if update_state:
        _mark_match_cache_refreshed(result)
    return result


# status 의 parquet 파생 수치(전체/제품별 row 수, product 목록, max update_time)는
# monolithic 파일 시그니처가 같으면 불변이다. /view 가 캐시 미스마다 이 함수를
# 호출해 monolithic 파일을 4회 full-collect 하던 것이 검색 지연의 고정비용이었다
# — (sig, product) 키로 메모이즈해 파일이 재기록될 때만 재계산한다.
_LATEST_STATUS_STATS_LOCK = threading.Lock()
_LATEST_STATUS_STATS_CACHE: dict[str, tuple[tuple, dict]] = {}
_LATEST_STATUS_STATS_MAX = 64


def _latest_lot_step_cache_parquet_stats(product: str, fp: Path) -> dict:
    key = str(product or "").strip().upper()
    sig = _path_cache_sig(fp)
    with _LATEST_STATUS_STATS_LOCK:
        cached = _LATEST_STATUS_STATS_CACHE.get(key)
        if cached is not None and cached[0] == sig:
            return dict(cached[1])
    lf = _latest_lot_step_cache_lf("")
    if lf is None:
        raise RuntimeError("latest cache is not readable")
    names = lf.collect_schema().names()
    format_version = 0
    cache_source = ""
    if LATEST_LOT_STEP_CACHE_FORMAT_COLUMN in names:
        try:
            format_version = int(
                lf.select(
                    pl.col(LATEST_LOT_STEP_CACHE_FORMAT_COLUMN)
                    .cast(pl.Int64, strict=False)
                    .max()
                ).collect().item(0, 0) or 0
            )
        except Exception:
            format_version = 0
    if LATEST_LOT_STEP_CACHE_SOURCE_COLUMN in names:
        try:
            cache_source = str(
                lf.select(pl.col(LATEST_LOT_STEP_CACHE_SOURCE_COLUMN).drop_nulls().first())
                .collect().item(0, 0) or ""
            )
        except Exception:
            cache_source = ""
    total_df = lf.select(pl.len().alias("row_count")).collect()
    row_count = int(total_df.item(0, 0) or 0)
    products: list[str] = []
    if "product" in names:
        prod_df = (
            lf.select(pl.col("product").cast(_STR, strict=False).alias("product"))
            .filter(pl.col("product").is_not_null() & (pl.col("product") != ""))
            .unique()
            .sort("product")
            .head(500)
            .collect()
        )
        products = [str(v) for v in prod_df["product"].to_list() if str(v or "").strip()]
    product_row_count = row_count
    if str(product or "").strip():
        product_row_count = 0
        if "product" in names:
            product_lf = _latest_lot_step_cache_lf(product)
            if product_lf is not None:
                product_row_count = int(product_lf.select(pl.len().alias("row_count")).collect().item(0, 0) or 0)
    updated_at = ""
    if "update_time" in names:
        try:
            value = lf.select(pl.col("update_time").cast(_STR, strict=False).max().alias("updated_at")).collect().item(0, 0)
            if value:
                updated_at = str(value)
        except Exception:
            pass
    stats = {
        "format_version": format_version,
        "expected_format_version": LATEST_LOT_STEP_CACHE_FORMAT_VERSION,
        "format_current": format_version == LATEST_LOT_STEP_CACHE_FORMAT_VERSION,
        "cache_source": cache_source,
        "row_count": row_count,
        "product_row_count": product_row_count,
        "products": products,
        "updated_at": updated_at,
    }
    with _LATEST_STATUS_STATS_LOCK:
        _LATEST_STATUS_STATS_CACHE[key] = (sig, dict(stats))
        while len(_LATEST_STATUS_STATS_CACHE) > _LATEST_STATUS_STATS_MAX:
            _LATEST_STATUS_STATS_CACHE.pop(next(iter(_LATEST_STATUS_STATS_CACHE)))
    return stats


def _latest_lot_step_cache_status(product: str = "") -> dict:
    """Return a non-throwing status summary for the canonical FAB match cache."""
    fp = _latest_lot_step_cache_path()
    freshness = _match_cache_global_fresh()
    state = _match_cache_state()
    base = {
        "ok": True,
        "cache_path": str(fp),
        "cache_exists": fp.is_file(),
        "row_count": 0,
        "product_row_count": 0,
        "products": [],
        "updated_at": state.get("updated_at") or state.get("last_refresh_at") or "",
        "latest_updated_at": state.get("updated_at") or "",
        "last_refresh_at": state.get("last_refresh_at") or "",
        "interval_minutes": _match_cache_refresh_minutes(),
        "latest_cache": freshness,
        "format_version": 0,
        "expected_format_version": LATEST_LOT_STEP_CACHE_FORMAT_VERSION,
        "format_current": False,
        "cache_source": "",
    }
    if not fp.is_file():
        return base
    try:
        stats = _latest_lot_step_cache_parquet_stats(product, fp)
        updated_at = stats.get("updated_at") or base["updated_at"]
        if not updated_at:
            try:
                updated_at = datetime.datetime.fromtimestamp(fp.stat().st_mtime).isoformat(timespec="seconds")
            except Exception:
                updated_at = ""
        return {
            **base,
            "format_version": int(stats.get("format_version") or 0),
            "expected_format_version": LATEST_LOT_STEP_CACHE_FORMAT_VERSION,
            "format_current": bool(stats.get("format_current")),
            "cache_source": str(stats.get("cache_source") or ""),
            "row_count": int(stats.get("row_count") or 0),
            "product_row_count": int(stats.get("product_row_count") or 0),
            "products": list(stats.get("products") or []),
            "updated_at": updated_at,
            "latest_updated_at": updated_at,
        }
    except Exception as e:
        logger.warning("SplitTable latest lot-step cache status failed (%s) %s: %s", fp, type(e).__name__, e)
        return {**base, "ok": False, "error": f"{type(e).__name__}: {e}"}


def _resolve_match_cache_columns(ov: dict, main_names_list: list[str], fab_schema_names: list[str]) -> dict:
    main_names = set(main_names_list)
    fab_names = set(fab_schema_names)
    join_keys = ov.get("join_keys") or []
    if isinstance(join_keys, str):
        join_keys = [k.strip() for k in join_keys.split(",") if k.strip()]
    if join_keys:
        mapped = []
        for k in join_keys:
            actual = _ci_resolve_in(k, main_names_list) or _resolve_source_col_name(k, fab_schema_names)
            if actual:
                mapped.append(actual)
        join_keys = mapped
    if not join_keys:
        join_keys = _default_override_join_keys(main_names_list, fab_schema_names)
    join_keys = [k for k in join_keys if k in main_names and k in fab_names]

    root_col = _resolve_source_col_name((ov.get("root_col") or "").strip(), fab_schema_names) \
               or _pick_first_present_ci(("root_lot_id",), fab_schema_names)
    wafer_col = _resolve_source_col_name((ov.get("wf_col") or ov.get("wafer_col") or "").strip(), fab_schema_names) \
                or _pick_first_present_ci(("wafer_id", "wafer"), fab_schema_names)
    fc_raw = (ov.get("fab_col") or "").strip()
    fab_col = (_resolve_source_col_name(fc_raw, fab_schema_names) if fc_raw else "") \
              or _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_schema_names) \
              or "fab_lot_id"
    tc_raw = (ov.get("ts_col") or "").strip()
    ts_col = (_resolve_source_col_name(tc_raw, fab_schema_names) if tc_raw else "") \
             or _pick_ts_col(fab_schema_names)

    raw_oc = ov.get("override_cols")
    if isinstance(raw_oc, str):
        raw_oc = [c.strip() for c in raw_oc.split(",") if c.strip()]
    if not raw_oc:
        raw_oc = list(_DEFAULT_OVERRIDE_COLS)
    if fab_col and fab_col not in raw_oc:
        raw_oc = list(raw_oc) + [fab_col]
    resolved_oc = []
    for c in raw_oc:
        actual = _resolve_source_col_name(c, fab_schema_names)
        resolved_oc.append(actual or c)
    override_cols = [c for c in dict.fromkeys(resolved_oc)
                     if c in fab_names and c not in join_keys]
    return {
        "join_keys": join_keys,
        "root_col": root_col,
        "wafer_col": wafer_col,
        "fab_col": fab_col,
        "ts_col": ts_col,
        "override_cols": override_cols,
    }


def _join_fab_projection_into_main(lf, main_names: set[str], fab_proj, join_keys: list[str],
                                   override_cols: list[str], *, fab_has_join_tmp: bool = False):
    join_aliases = [(k, f"__join_key_{i}") for i, k in enumerate(join_keys)]
    join_tmp_keys = [tmp for _, tmp in join_aliases]
    if not fab_has_join_tmp:
        fab_proj = fab_proj.with_columns([_join_key_expr(k).alias(tmp) for k, tmp in join_aliases])
    lf = lf.with_columns([_join_key_expr(k).alias(tmp) for k, tmp in join_aliases])
    backup_cols: list = []
    for c in override_cols:
        if c in main_names:
            bk = f"__main_bk_{c}"
            lf = lf.with_columns(pl.col(c).alias(bk))
            backup_cols.append((c, bk))
            lf = lf.drop(c)
    lf = lf.join(fab_proj, on=join_tmp_keys, how="left").drop(join_tmp_keys)
    for c, bk in backup_cols:
        if c.casefold() == "fab_lot_id":
            # FAB lot ids should come from the FAB DB connection table.
            lf = lf.drop(bk)
        else:
            lf = lf.with_columns(pl.coalesce([pl.col(c), pl.col(bk)]).alias(c)).drop(bk)
    joined_lot_col = next((c for c in override_cols if str(c).casefold() == "lot_id"), "")
    joined_fab_col = next((c for c in override_cols if str(c).casefold() == "fab_lot_id"), "")
    if joined_lot_col and not joined_fab_col:
        # Raw FAB now uses lot_id as the fab-lot key. Keep raw schema clean, but
        # expose the legacy view label so SplitTable grouping/export still works.
        lf = lf.with_columns(pl.col(joined_lot_col).cast(_STR, strict=False).alias("fab_lot_id"))
    return lf


def _latest_lot_progress_projection(product: str, main_names_list: list[str],
                                     root_lot_id: str = "", fab_lot_id: str = "",
                                     wafer_ids: str = "") -> dict | None:
    """Use the canonical LOT progress cache as SplitTable's lot identity source."""
    cache_lf = _latest_lot_step_cache_lf(product, root_lot_id=root_lot_id)
    if cache_lf is None:
        return None
    main_names = set(main_names_list)
    # New products do not always use the literal root_lot_id/wafer_id names.
    # Resolve configured aliases first, then use the same schema detection as the
    # SplitTable renderer.  Previously LOT_ID (or a configured root column) made
    # this projection return None, so the FAB cache existed but every wafer header
    # was rendered with an unassigned "-" lot label.
    configured_root = ""
    configured_wafer = ""
    try:
        cfg = load_json_cached(SOURCE_CFG, {}) or {}
        ov = _lot_override_for(cfg, product)
        configured_root = _ci_resolve_in(str(ov.get("root_col") or ""), main_names_list)
        configured_wafer = _ci_resolve_in(
            str(ov.get("wf_col") or ov.get("wafer_col") or ""), main_names_list)
    except Exception:
        pass
    detected_root, detected_wafer = find_lot_wafer_cols(main_names_list)
    root_key = (
        configured_root
        or _ci_resolve_in("root_lot_id", main_names_list)
        or detected_root
    )
    wafer_key = (
        configured_wafer
        or _ci_resolve_in("wafer_id", main_names_list)
        or _pick_first_present_ci(("wafer_id", "wf_id", "wafer"), main_names_list)
        or detected_wafer
    )
    if not root_key or root_key not in main_names:
        return None
    join_keys = [root_key]
    if wafer_key and wafer_key in main_names:
        join_keys.append(wafer_key)
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return None
    if not {"root_lot_id", "lot_id"}.issubset(set(names)):
        return None
    q = _filter_latest_lot_step_cache(
        cache_lf,
        root_lot_id=root_lot_id,
        fab_lot_id=fab_lot_id,
        wafer_ids=wafer_ids,
    )
    join_aliases = [(k, f"__join_key_{i}") for i, k in enumerate(join_keys)]
    exprs = []
    for source_col, (_main_key, tmp) in zip(["root_lot_id", "wafer_id"], join_aliases):
        if source_col not in names:
            return None
        exprs.append(_join_key_expr(source_col).alias(tmp))
    exprs.extend([
        pl.col("lot_id").cast(_STR, strict=False).alias("lot_id"),
        pl.col("lot_id").cast(_STR, strict=False).alias("fab_lot_id"),
        (
            pl.col("tkout_time").cast(_STR, strict=False).alias(MATCH_CACHE_TS_COL)
            if "tkout_time" in names else pl.lit("").alias(MATCH_CACHE_TS_COL)
        ),
    ])
    try:
        proj = (
            q.select(exprs)
            .filter(pl.col("lot_id").is_not_null() & (pl.col("lot_id") != ""))
            .sort(MATCH_CACHE_TS_COL, descending=True, nulls_last=True)
            .unique(subset=[tmp for _k, tmp in join_aliases], keep="first", maintain_order=True)
            .select([tmp for _k, tmp in join_aliases] + ["lot_id", "fab_lot_id"])
        )
    except Exception as e:
        logger.warning("latest LOT progress projection failed (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return None
    return {
        "lf": proj,
        "join_keys": join_keys,
        "override_cols": ["lot_id", "fab_lot_id"],
        "meta": {
            "source": "lot_progress_latest_cache",
            "path": str(_latest_lot_step_cache_path()),
        },
    }


def _filter_match_cache_scope(cache_lf, root_lot_id: str = "", fab_lot_id: str = "",
                              wafer_ids: str = ""):
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        names = []
    root_scope = str(root_lot_id or "").strip()
    fab_scope = str(fab_lot_id or "").strip()
    wafer_scope = str(wafer_ids or "").strip()
    if root_scope and MATCH_CACHE_ROOT_COL in names:
        cache_lf = cache_lf.filter(_join_key_expr(MATCH_CACHE_ROOT_COL) == root_scope.upper())
    if fab_scope and MATCH_CACHE_FAB_COL in names:
        cache_lf = cache_lf.filter(_join_key_expr(MATCH_CACHE_FAB_COL) == fab_scope.upper())
    if wafer_scope and MATCH_CACHE_WAFER_COL in names:
        wf_list = [w.strip() for w in wafer_scope.split(",") if w.strip()]
        try:
            wf_ints = [int(w) for w in wf_list]
            wf_strs = set()
            for n in wf_ints:
                wf_strs.update([str(n), f"{n:02d}", f"W{n}", f"W{n:02d}"])
            cache_lf = cache_lf.filter(
                pl.col(MATCH_CACHE_WAFER_COL).cast(_STR, strict=False).is_in(list(wf_strs))
                | pl.col(MATCH_CACHE_WAFER_COL).cast(pl.Int64, strict=False).is_in(wf_ints)
            )
        except ValueError:
            cache_lf = cache_lf.filter(pl.col(MATCH_CACHE_WAFER_COL).cast(_STR, strict=False).is_in(wf_list))
    return cache_lf


def _cached_fab_projection(product: str, ov: dict, fab_source: str, main_names_list: list[str],
                           root_lot_id: str = "", fab_lot_id: str = "", wafer_ids: str = "") -> dict | None:
    current = _match_cache_current(product)
    if not current:
        return None
    meta = current["meta"]
    if meta.get("fab_source") != _normalize_fab_source_path(fab_source):
        return None
    join_keys = [k for k in (meta.get("join_keys") or []) if k in main_names_list]
    join_tmp_keys = list(meta.get("join_tmp_keys") or [])
    if not join_keys or len(join_keys) != len(join_tmp_keys):
        return None
    cache_lf = _filter_match_cache_scope(current["lf"], root_lot_id=root_lot_id,
                                         fab_lot_id=fab_lot_id, wafer_ids=wafer_ids)
    try:
        cache_names = cache_lf.collect_schema().names()
    except Exception:
        return None
    override_cols = [c for c in (meta.get("override_cols") or []) if c in cache_names]
    if not override_cols:
        return None
    keep = list(dict.fromkeys(join_tmp_keys + override_cols + ([MATCH_CACHE_TS_COL] if MATCH_CACHE_TS_COL in cache_names else [])))
    fab_proj = cache_lf.select(keep)
    if MATCH_CACHE_TS_COL in keep:
        fab_proj = fab_proj.sort(MATCH_CACHE_TS_COL, descending=True, nulls_last=True)
        fab_proj = fab_proj.unique(subset=join_tmp_keys, keep="first", maintain_order=True)
    else:
        fab_proj = fab_proj.unique(subset=join_tmp_keys, keep="last")
    return {
        "lf": fab_proj.select(list(dict.fromkeys(join_tmp_keys + override_cols))),
        "join_keys": join_keys,
        "join_tmp_keys": join_tmp_keys,
        "override_cols": override_cols,
        "meta": meta,
    }


def _fab_history_scope_from_cache(product: str, root_lot_id: str = "", fab_lot_id: str = "",
                                  prefix: str = "", limit: int = 500) -> dict | None:
    latest = _fab_history_scope_from_latest_cache(
        product,
        root_lot_id=root_lot_id,
        fab_lot_id=fab_lot_id,
        prefix=prefix,
        limit=limit,
    )
    if latest is not None:
        return latest
    current = _match_cache_current(product)
    if not current:
        return None
    cache_lf = current["lf"]
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return None
    if MATCH_CACHE_ROOT_COL not in names or MATCH_CACHE_FAB_COL not in names:
        return None
    root_scope = str(root_lot_id or "").strip()
    fab_scope = str(fab_lot_id or "").strip()
    q = cache_lf.select([
        pl.col(MATCH_CACHE_ROOT_COL).cast(_STR, strict=False).alias("root"),
        pl.col(MATCH_CACHE_FAB_COL).cast(_STR, strict=False).alias("fab"),
        *([pl.col(MATCH_CACHE_WAFER_COL).cast(_STR, strict=False).alias("wafer")] if MATCH_CACHE_WAFER_COL in names else []),
    ]).filter(pl.col("root").is_not_null() & pl.col("fab").is_not_null())
    if root_scope:
        q = q.filter(_join_key_expr("root") == root_scope.upper())
    if fab_scope:
        q = q.filter(_join_key_expr("fab") == fab_scope.upper())
    elif str(prefix or "").strip():
        q = q.filter(_contains_literal_ci_expr("fab", prefix))
    try:
        fabs = _limited_unique_values(q, "fab", prefix="", limit=limit,
                                      preview_only=not bool(root_scope or fab_scope or str(prefix or "").strip()))
        roots: list[str] = [root_scope] if root_scope else []
        wafers: list[str] = []
        if fab_scope and fabs:
            meta_cols = [pl.col("root")]
            if "wafer" in q.collect_schema().names():
                meta_cols.append(pl.col("wafer"))
            meta_df = q.select(meta_cols).unique().collect()
            roots = sorted({s for s in (_clean_str(v) for v in meta_df["root"].to_list()) if s})
            if "wafer" in meta_df.columns:
                wafers = sorted({s for s in (_clean_str(v) for v in meta_df["wafer"].to_list()) if s}, key=_wafer_sort_key)
    except Exception as e:
        logger.warning("_fab_history_scope_from_cache 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return None
    return {
        "candidates": fabs,
        "root_ids": roots,
        "wafer_ids": wafers,
        "source": current.get("fab_source", ""),
        "cache": True,
        "query_ok": True,
    }


def _fab_history_root_candidates_from_cache(product: str, prefix: str = "", limit: int = 500) -> dict | None:
    latest = _fab_history_root_candidates_from_latest_cache(product, prefix=prefix, limit=limit)
    if latest is not None:
        return latest
    current = _match_cache_current(product)
    if not current:
        return None
    cache_lf = current["lf"]
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return None
    if MATCH_CACHE_ROOT_COL not in names:
        return None
    try:
        values = _limited_unique_values(cache_lf, MATCH_CACHE_ROOT_COL, prefix=prefix, limit=limit)
    except Exception as e:
        logger.warning("_fab_history_root_candidates_from_cache 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return None
    return {"candidates": values, "source": current.get("fab_source", ""), "cache": True}


def _fab_lot_snapshot_from_cache(product: str, root_lot_id: str, wafer_id: str = "") -> str:
    latest = _fab_lot_snapshot_from_latest_cache(product, root_lot_id, wafer_id)
    if latest:
        return latest
    current = _match_cache_current(product)
    root = str(root_lot_id or "").strip()
    if not current or not root:
        return ""
    cache_lf = _filter_match_cache_scope(current["lf"], root_lot_id=root, wafer_ids=str(wafer_id or ""))
    try:
        names = cache_lf.collect_schema().names()
    except Exception:
        return ""
    if MATCH_CACHE_FAB_COL not in names:
        return ""
    q = (
        cache_lf
        .select([
            pl.col(MATCH_CACHE_FAB_COL).cast(_STR, strict=False).alias("fab"),
            *([pl.col(MATCH_CACHE_TS_COL).cast(_STR, strict=False).alias("ts")] if MATCH_CACHE_TS_COL in names else []),
        ])
        .filter(pl.col("fab").is_not_null() & (pl.col("fab") != ""))
    )
    if "ts" in q.collect_schema().names():
        q = q.sort("ts", descending=True, nulls_last=True)
    else:
        q = q.sort("fab")
    try:
        df = q.head(1).collect()
    except Exception as e:
        logger.warning("_fab_lot_snapshot_from_cache 실패 (product=%s root=%s wafer=%s) %s: %s",
                       product, root_lot_id, wafer_id, type(e).__name__, e)
        return ""
    if df.is_empty():
        return ""
    return _clean_str(df.item(0, 0))


def _build_match_cache_streamed(
    q_base,
    tmp: Path,
    *,
    unique_subset: list[str],
    ts_present: bool,
    batch_col: str,
    roots_per_batch: int,
    product: str = "",
) -> int:
    """root_lot_id(=dedup subset 컬럼) 단위로 FAB 매칭캐시를 나눠 빌드한다.

    글로벌 ``q.sort(ts).unique()`` 는 파이프라인 브레이커라 전체 FAB 를 메모리에
    버퍼링해 peak RAM 이 폭발한다(수십 GB). 여기서는 dedup subset 에 속한 한 컬럼
    (``batch_col``, 정규화된 join key = root_lot_id)의 값으로 배치를 나눈다.
    dedup subset 의 한 컬럼으로 나누면 서로 다른 배치는 dedup 그룹이 절대 겹치지
    않으므로, 각 배치를 개별적으로 sort+unique 한 뒤 **마지막에 sort/unique 없이
    스트리밍 병합만** 하면 글로벌 sort+unique 와 결과가 동일하다.

    peak RAM 은 한 배치(roots_per_batch 개의 root) 크기로 제한된다. 대신 FAB 원천을
    배치 수만큼 재스캔하므로 느리다(속도↔메모리 트레이드오프, 메모리 우선).
    배치 사이에서 프레임을 해제(del+gc)하고 메모리 가드를 확인한다.
    """
    def _dedup(part_q):
        if ts_present:
            part_q = part_q.sort(MATCH_CACHE_TS_COL, descending=True, nulls_last=True)
            return part_q.unique(subset=unique_subset, keep="first", maintain_order=True)
        return part_q.unique(subset=unique_subset, keep="last")

    # 캐시 이벤트 로그(사용자가 보는 화면)로 진행 상황을 내보내는 헬퍼.
    # warmup/cache_op 와 같은 category 라 수동스캔 '전체' 필터에 실시간 표시된다.
    def _emit(event: str, *, ok: bool = True, detail: dict | None = None):
        try:
            from core.cache_event_log import record as _rec
            _rec("cache_op", event, ok=ok, detail=detail or {}, product=product)
        except Exception:
            pass

    def _rss_gb() -> float:
        try:
            from core.runtime_limits import process_memory_snapshot
            snap = process_memory_snapshot()
            return round(float(snap.get("process_rss_gb")
                               or snap.get("process_memory_effective_gb") or 0.0), 2)
        except Exception:
            return 0.0

    def _fmt_dur(sec: float) -> str:
        sec = int(max(0, sec))
        if sec < 60:
            return f"{sec}초"
        if sec < 3600:
            return f"{sec // 60}분 {sec % 60}초"
        return f"{sec // 3600}시간 {(sec % 3600) // 60}분"

    partdir = tmp.parent / (tmp.name + ".parts")
    try:
        if partdir.exists():
            for f in partdir.glob("*.parquet"):
                try:
                    f.unlink(missing_ok=True)
                except Exception:
                    pass
            try:
                partdir.rmdir()
            except Exception:
                pass
    except Exception:
        pass

    # 글로벌 FAB(q_base = _scan_global_fab_sources 전체 concat)를 배치마다 재스캔하면
    # 배치 수만큼 전체 FAB 를 다시 읽어 매우 느리고, Windows 워킹셋이 반복 mmap 으로
    # 계속 부풀어 RSS 가 꾸준히 오른다. 따라서 필요한 컬럼만 투영된 q_base 를 **딱 한 번**
    # 임시 parquet 로 흘려쓴다(정렬 없음 → 완전 스트리밍, 저메모리). 이후 배치/중복제거는
    # 이 작은 로컬 파일만 읽으므로 글로벌 재스캔이 사라진다(속도·RSS 동시 개선).
    base_tmp = tmp.parent / (tmp.name + ".base.parquet")
    try:
        base_tmp.unlink(missing_ok=True)
    except Exception:
        pass
    _emit(f"[매칭] {product}: 원본 FAB 1회 스캔·투영 중… (배치 재스캔 제거)",
          detail={"phase": "base_scan", "rss_gb": _rss_gb()})
    _write_match_cache_lazyframe(q_base, base_tmp)
    base_lf = pl.scan_parquet(str(base_tmp))

    # 배치 키 값 목록 — 작은 투영본에서 한 컬럼만 collect (글로벌 재스캔 아님).
    try:
        vals_df = base_lf.select(pl.col(batch_col)).unique().collect(streaming=True)
    except TypeError:
        vals_df = base_lf.select(pl.col(batch_col)).unique().collect()
    vals = sorted({str(v) for v in vals_df.get_column(batch_col).drop_nulls().to_list()})

    if not vals:
        # 빈 결과: 스키마만 있는 빈 parquet 생성 (legacy 경로와 동일 산출).
        try:
            return _write_match_cache_lazyframe(_dedup(base_lf), tmp)
        finally:
            try:
                base_tmp.unlink(missing_ok=True)
            except Exception:
                pass

    partdir.mkdir(parents=True, exist_ok=True)
    part_paths: list[Path] = []
    total_roots = len(vals)
    total_batches = (total_roots + roots_per_batch - 1) // roots_per_batch
    started = time.time()
    rows_done = 0
    last_emit = 0.0
    emit_min_gap = _match_cache_stream_log_gap_seconds()
    _emit(
        f"[매칭] {product}: 스트리밍 빌드 시작 — 총 {total_roots:,} 랏 → {total_batches:,} 배치"
        f" (배치당 {roots_per_batch:,} 랏)",
        detail={"phase": "start", "roots": total_roots, "batches": total_batches,
                "roots_per_batch": roots_per_batch, "rss_gb": _rss_gb(),
                "progress": _match_progress(0, total_roots),
                "stage": _stage("match", "start")},
    )
    try:
        for idx in range(0, total_roots, roots_per_batch):
            if _MATCH_CACHE_STOP.is_set():
                raise RuntimeError("match cache build stopped")
            if _match_cache_cancelled(product):
                # 배치 경계에서만 끊는다 — 쓰는 중인 part 파일을 남기지 않는다.
                raise MatchCacheCancelled(product)
            _wait_for_match_cache_memory()
            batch_no = idx // roots_per_batch + 1
            b_start = time.time()
            chunk = vals[idx:idx + roots_per_batch]
            part_q = _dedup(base_lf.filter(pl.col(batch_col).is_in(chunk)))
            part_path = partdir / f"part_{batch_no:05d}.parquet"
            part_rows = _write_match_cache_lazyframe(part_q, part_path)
            part_paths.append(part_path)
            rows_done += int(part_rows or 0)
            b_dur = time.time() - b_start
            elapsed = time.time() - started
            roots_done = min(idx + len(chunk), total_roots)
            pct = roots_done * 100 // total_roots if total_roots else 100
            # ETA: 지금까지 배치당 평균 소요 × 남은 배치.
            avg = elapsed / batch_no if batch_no else 0.0
            eta = avg * (total_batches - batch_no)
            rss = _rss_gb()
            try:
                mem_lots = _ml_table_lookup.root_ram_cache_lot_count()
            except Exception:
                mem_lots = 0
            try:
                _match_cache_job_update(
                    stream_product=product, stream_batch=batch_no,
                    stream_batch_total=total_batches, stream_roots=total_roots,
                    stream_roots_done=roots_done, stream_rows=rows_done,
                    stream_rss_gb=rss, stream_eta_sec=round(eta, 1),
                    stream_mem_lots=mem_lots,
                )
            except Exception:
                pass
            # 첫·마지막 배치는 항상, 그 외엔 최소 간격 스로틀 — 작은 배치의 로그 폭주 방지.
            now = time.time()
            if batch_no == 1 or batch_no == total_batches or (now - last_emit) >= emit_min_gap:
                last_emit = now
                _emit(
                    f"[매칭] {product}: 랏 {roots_done:,}/{total_roots:,} ({pct}%)"
                    f" · 배치 {batch_no:,}/{total_batches:,} · {b_dur:.1f}s"
                    f" · 누적 {rows_done:,}행 · 메모리 {mem_lots}랏 상주 · RSS {rss}GB"
                    f" · 남은 ~{_fmt_dur(eta)}",
                    detail={"phase": "batch", "batch": batch_no, "batches": total_batches,
                            "roots_done": roots_done, "roots_total": total_roots,
                            "batch_sec": round(b_dur, 2), "elapsed_sec": round(elapsed, 1),
                            "eta_sec": round(eta, 1), "rows": rows_done,
                            "mem_lots": mem_lots, "rss_gb": rss,
                            "progress": _match_progress(roots_done, total_roots)},
                )
            try:
                del part_q
                gc.collect()
            except Exception:
                pass

        # 배치끼리 batch_col 값이 배타적 → sort/unique 없이 스트리밍 병합만 하면 된다.
        merged = pl.scan_parquet([str(p) for p in part_paths])
        row_count = _write_match_cache_lazyframe(merged, tmp)
        _emit(
            f"[매칭] {product}: 완료 — {total_roots:,}랏 · {row_count:,}행 · {total_batches:,}배치"
            f" · 총 {_fmt_dur(time.time() - started)} · RSS {_rss_gb()}GB",
            detail={"phase": "done", "roots": total_roots, "rows": int(row_count),
                    "batches": total_batches, "total_sec": round(time.time() - started, 1),
                    "rss_gb": _rss_gb(),
                    "progress": _match_progress(total_roots, total_roots, state="done"),
                    "stage": _stage("match", "done")},
        )
    finally:
        for p in part_paths:
            try:
                p.unlink(missing_ok=True)
            except Exception:
                pass
        try:
            partdir.rmdir()
        except Exception:
            pass
        try:
            base_tmp.unlink(missing_ok=True)
        except Exception:
            pass
    return row_count


def _refresh_match_cache_products(products: list[str], force: bool = False) -> dict:
    """Build persisted FAB root/fab/wafer connection tables for known products."""
    products = [p for p in products if p]
    results: list[dict] = []
    with _MATCH_CACHE_BUILD_LOCK:
        MATCH_CACHE_DIR.mkdir(parents=True, exist_ok=True)
        canceled_scan = False
        for raw_product in products:
            # 제품 경계 = 안전한 취소 지점. 이미 만든 제품 캐시는 그대로 두고
            # 남은 제품을 건너뛴 뒤 다음 큐 작업으로 넘어간다.
            if _scan_cancel_requested():
                if not canceled_scan:
                    canceled_scan = True
                    logger.info("match cache: 스캔 큐 중단 요청 — 남은 제품 건너뜀")
                results.append({"product": raw_product, "ok": False, "skipped": True,
                                "row_count": 0, "reason": "scan_canceled"})
                continue
            ml_product, ov, fab_source = _current_fab_override(raw_product)
            result = {"product": ml_product or raw_product, "ok": False, "skipped": False, "row_count": 0, "fab_source": fab_source}
            try:
                if not ml_product:
                    result["reason"] = "FAB source not matched"
                    results.append(result)
                    continue
                if not fab_source and not _global_fab_source_paths(""):
                    result["reason"] = "FAB source not matched"
                    results.append(result)
                    continue
                config_key = _match_cache_config_key(ml_product, ov, fab_source)
                fp = _match_cache_path(ml_product)
                meta_fp = _match_cache_meta_path(ml_product)
                old_meta = load_json(meta_fp, {}) if meta_fp.is_file() else {}
                if not force and fp.is_file() and isinstance(old_meta, dict) and old_meta.get("config_key") == config_key:
                    age_s = time.time() - float(old_meta.get("built_epoch") or 0)
                    if age_s < _match_cache_refresh_minutes() * 60:
                        result.update({"ok": True, "skipped": True, "row_count": int(old_meta.get("row_count") or 0)})
                        results.append(result)
                        continue

                main_lf = _scan_product_base(ml_product)
                main_names_list = main_lf.collect_schema().names()
                fab_lf, fab_sources = _scan_global_fab_sources(fab_source, tag_source_product=True)
                if fab_lf is None:
                    result["reason"] = "FAB source scan failed"
                    result["fab_sources"] = fab_sources
                    results.append(result)
                    continue
                fab_lf, fab_schema_names = _ci_align_fab_to_main(fab_lf, main_names_list)
                try:
                    fab_schema_names = fab_lf.collect_schema().names()
                except Exception:
                    pass
                result["fab_sources"] = fab_sources
                cols = _resolve_match_cache_columns(ov, main_names_list, fab_schema_names)
                join_keys = cols["join_keys"]
                override_cols = cols["override_cols"]
                if not join_keys or not override_cols:
                    result["reason"] = "join keys or override columns missing"
                    result["join_keys"] = join_keys
                    result["override_cols"] = override_cols
                    results.append(result)
                    continue

                wanted = list(dict.fromkeys(
                    join_keys
                    + override_cols
                    + ([cols["ts_col"]] if cols["ts_col"] else [])
                    + ([cols["root_col"]] if cols["root_col"] else [])
                    + ([cols["wafer_col"]] if cols["wafer_col"] else [])
                    + ([cols["fab_col"]] if cols["fab_col"] else [])
                ))
                # 출처 제품 열은 override_cols 해석 대상이 아니므로(FAB 원본 열이
                # 아니다) 여기서 명시적으로 살려 둔다.
                if MATCH_CACHE_SRC_PRODUCT_COL in fab_schema_names:
                    wanted = list(wanted) + [MATCH_CACHE_SRC_PRODUCT_COL]
                wanted = [c for c in wanted if c in fab_schema_names]
                q = fab_lf.select(wanted)
                join_tmp_keys = [f"__join_key_{i}" for i, _ in enumerate(join_keys)]
                exprs = [_join_key_expr(k).alias(tmp) for k, tmp in zip(join_keys, join_tmp_keys)]
                if cols["root_col"] and cols["root_col"] in fab_schema_names:
                    exprs.append(pl.col(cols["root_col"]).cast(_STR, strict=False).alias(MATCH_CACHE_ROOT_COL))
                if cols["wafer_col"] and cols["wafer_col"] in fab_schema_names:
                    exprs.append(pl.col(cols["wafer_col"]).cast(_STR, strict=False).alias(MATCH_CACHE_WAFER_COL))
                if cols["fab_col"] and cols["fab_col"] in fab_schema_names:
                    exprs.append(pl.col(cols["fab_col"]).cast(_STR, strict=False).alias(MATCH_CACHE_FAB_COL))
                if cols["ts_col"] and cols["ts_col"] in fab_schema_names:
                    exprs.append(pl.col(cols["ts_col"]).cast(_STR, strict=False).alias(MATCH_CACHE_TS_COL))
                q = q.with_columns(exprs)
                # 전 FAB 폴더를 읽되 target ML_TABLE의 랏/wafer key에 해당하는 행만
                # 남긴다. 이 제한이 없으면 제품마다 같은 전 FAB cache가 만들어진다.
                q = _scope_match_cache_to_main_keys(main_lf, q, join_keys, join_tmp_keys)
                keep = list(dict.fromkeys(
                    join_tmp_keys
                    + [MATCH_CACHE_ROOT_COL, MATCH_CACHE_WAFER_COL, MATCH_CACHE_FAB_COL,
                       MATCH_CACHE_TS_COL, MATCH_CACHE_SRC_PRODUCT_COL]
                    + override_cols
                ))
                q_names = q.collect_schema().names()
                keep = [c for c in keep if c in q_names]
                q = q.select(keep)
                for k in join_tmp_keys:
                    q = q.filter(pl.col(k).is_not_null() & (pl.col(k) != ""))
                # The persisted cache is the authoritative SplitTable
                # root/wafer -> FAB lot mapping.  Keep exactly one FAB row per
                # root_lot_id + wafer_id join key, chosen by latest tkout/time,
                # so SplitTable and Inform snapshots read the same lot_id basis.
                unique_subset = [c for c in join_tmp_keys if c in keep]
                if not unique_subset:
                    unique_subset = [c for c in (MATCH_CACHE_ROOT_COL, MATCH_CACHE_WAFER_COL) if c in keep]
                if not unique_subset:
                    unique_subset = [c for c in (MATCH_CACHE_FAB_COL,) if c in keep]
                ts_present = MATCH_CACHE_TS_COL in keep
                tmp = fp.with_suffix(fp.suffix + ".tmp")

                # root_lot_id 배치 스트리밍: dedup subset 에 속한 정규화 join key 한 컬럼
                # (가능하면 root_lot_id) 으로 배치를 나눠 peak RAM 을 한 배치 크기로 제한.
                # batch_col 은 반드시 join_tmp_keys(=null 필터 완료) 중 하나여야 행 손실이 없다.
                subset_join_keys = [c for c in unique_subset if c in join_tmp_keys]
                batch_col = ""
                if subset_join_keys:
                    root_src = cols.get("root_col") or ""
                    if root_src:
                        for jk, tmpk in zip(join_keys, join_tmp_keys):
                            if tmpk in subset_join_keys and str(jk).casefold() == str(root_src).casefold():
                                batch_col = tmpk
                                break
                    if not batch_col:
                        batch_col = subset_join_keys[0]

                if _match_cache_stream_enabled() and batch_col:
                    row_count = _build_match_cache_streamed(
                        q, tmp,
                        unique_subset=unique_subset,
                        ts_present=ts_present,
                        batch_col=batch_col,
                        roots_per_batch=_match_cache_stream_batch_roots(),
                        product=ml_product,
                    )
                else:
                    # legacy: 글로벌 sort+unique (streaming off 또는 배치 키 없음).
                    if ts_present:
                        q = q.sort(MATCH_CACHE_TS_COL, descending=True, nulls_last=True)
                        q = q.unique(subset=unique_subset, keep="first", maintain_order=True)
                    else:
                        q = q.unique(subset=unique_subset, keep="last")
                    row_count = _write_match_cache_lazyframe(q, tmp)
                tmp.replace(fp)
                meta = {
                    "version": MATCH_CACHE_VERSION,
                    "product": ml_product,
                    "fab_source": _normalize_fab_source_path(fab_source),
                    "fab_sources": fab_sources,
                    "config_key": config_key,
                    "built_at": datetime.datetime.now().isoformat(timespec="seconds"),
                    "built_epoch": time.time(),
                    "row_count": int(row_count),
                    "join_keys": join_keys,
                    "join_tmp_keys": join_tmp_keys,
                    "dedup_keys": unique_subset,
                    "override_cols": override_cols,
                    "root_col": cols["root_col"],
                    "wafer_col": cols["wafer_col"],
                    "fab_col": cols["fab_col"],
                    "ts_col": cols["ts_col"],
                }
                save_json(meta_fp, meta)
                _LOT_LOOKUP_CACHE.clear()
                # root 후보 풀은 ML_TABLE lookup candidate index가 authoritative다.
                # FAB match/pivot 캐시 갱신은 root 집합을 바꾸지 않으므로 이미 준비된
                # 신규 제품 후보 풀을 지우지 않는다.
                result.update({"ok": True, "row_count": int(row_count), "join_keys": join_keys,
                               "override_cols": override_cols, "fab_sources": fab_sources})
            except MatchCacheCancelled:
                # 중단: 기존 캐시 파일과 meta 는 건드리지 않고 임시 산출만 지운다.
                # meta 를 안 쓰므로 다음 스캔은 이 제품을 처음부터 다시 빌드한다.
                try:
                    _match_cache_path(ml_product).with_suffix(
                        _match_cache_path(ml_product).suffix + ".tmp").unlink(missing_ok=True)
                except Exception:
                    pass
                cancel = _match_cache_cancel_target()
                result.update({
                    "ok": False,
                    "skipped": False,
                    "cancelled": True,
                    "reason": f"관리자 중단 ({cancel.get('by') or '-'}) — 다음 스캔에서 처음부터 다시 빌드합니다",
                })
                logger.warning("SplitTable match cache build cancelled by admin (product=%s)", ml_product)
            except Exception as e:
                logger.warning("SplitTable match cache build failed (product=%s) %s: %s",
                               raw_product, type(e).__name__, e, exc_info=True)
                result["reason"] = f"{type(e).__name__}: {e}"
            results.append(result)
            try:
                gc.collect()
            except Exception:
                pass
    return {"ok": any(r.get("ok") for r in results), "products": results, "interval_minutes": _match_cache_refresh_minutes()}


def _canonical_product_set(products: list[str]) -> set[str]:
    out = set()
    for p in products or []:
        text = _canonical_mltable_product_name(p, allow_bare=True) or str(p or "").strip()
        if text:
            out.add(text.upper())
    return out


def _match_cache_products_cover_all(products: list[str]) -> bool:
    expected = _canonical_product_set(_match_cache_products(""))
    got = _canonical_product_set(products)
    return bool(expected) and expected.issubset(got)


def refresh_match_cache(product: str = "", force: bool = False, max_products: int | None = None) -> dict:
    """Build persisted FAB root/fab/wafer connection tables for SplitTable.

    Callers that warm the whole cache can pass max_products to pace large
    sweeps. Product-specific calls keep the historical synchronous behavior.
    """
    products = _match_cache_products(product)
    if max_products is not None:
        try:
            n = max(1, int(max_products))
            products = products[:n]
        except Exception:
            pass
    result = _refresh_match_cache_products(products, force=force)
    try:
        export_products = _match_cache_products("") or products
        export = export_latest_lot_step_cache(products=export_products, update_state=_match_cache_products_cover_all(products))
        result["latest_cache"] = export
    except Exception as e:
        logger.warning("SplitTable unified latest cache export failed: %s", e, exc_info=True)
        result["latest_cache"] = {"ok": False, "reason": f"{type(e).__name__}: {e}"}
    return result


def _match_cache_progress(job_id: str, message: str, *, ok: bool = True) -> None:
    """FAB 매칭 캐시 진행 신호 — 이벤트 로그 한 줄 + job heartbeat.

    로그가 곧 진행 신호다: cache_event_log.record 가 detail.job_id 로 작업의
    updated_ts 를 갱신하고, reap_stale_jobs() 는 그 값으로 정지를 판정한다.
    조용히 오래 도는 구간이 있으면 그 작업은 기본 5분 뒤 '응답 없음'으로 실패
    처리되고 뒤 단계는 시작조차 못 한다.

    stage 는 일부러 붙이지 않는다 — 단계 시작/종료 이벤트와 섞이면 화면의
    단계별 이력이 진행 로그로 오염된다.
    """
    try:
        from core.cache_event_log import record as _rec, heartbeat as _beat
    except Exception:
        return
    try:
        _rec("scan", message, ok=ok, detail={"job_id": job_id} if job_id else None)
        if job_id:
            _beat(job_id, message)
    except Exception:
        logger.debug("match cache progress log failed", exc_info=True)


def _match_cache_followups(force: bool, *, job_id: str = "", built: list[str] | None = None,
                           refresh_plan_risk: bool = False) -> dict:
    """FAB 매칭 캐시 뒤에 이어지는 파생 캐시 갱신.

    제품 루프가 끝난 뒤에도 남는 무거운 작업이다 — 통합 latest-lot 캐시 export 는
    전 제품 매칭 캐시를 concat/sort/unique 후 전량 collect 하고 행마다 step 메타를
    붙인다. 예전에는 이 구간이 로그도 heartbeat 도 없이 돌아서, 화면상 'FAB 100%'
    인 채로 멈춘 것처럼 보이고 5분 뒤엔 작업 전체가 '응답 없음'으로 실패한다.
    각 단계를 진행 신호로 남긴다.

    built: 이번에 실제로 빌드한 제품 목록. export 는 항상 전 제품을 대상으로 하되,
    '전역 매칭 캐시가 갱신됐다'는 상태 표시는 전 제품을 덮었을 때만 남긴다 —
    제품 하나만 수동 갱신한 것을 전체 갱신으로 기록하면 다음 주기가 건너뛰어진다.
    """
    out: dict[str, Any] = {}
    if refresh_plan_risk and not _MATCH_CACHE_STOP.is_set():
        _match_cache_progress(job_id, "[FAB 후처리] plan risk 캐시 갱신 중…")
        try:
            refresh_plan_risk_cache(force=False)
            out["plan_risk"] = {"ok": True}
        except Exception as e:
            logger.warning("SplitTable plan risk cache refresh after match cache failed: %s", e)
            out["plan_risk"] = {"ok": False, "reason": f"{type(e).__name__}: {e}"}
            _match_cache_progress(job_id, f"[FAB 후처리] plan risk 캐시 실패: {e}", ok=False)
    if not _MATCH_CACHE_STOP.is_set():
        _match_cache_progress(job_id, "[FAB 후처리] 통합 latest-lot 캐시 export 중… "
                                      "(전 제품 매칭 캐시 병합 — 데이터가 크면 수 분 걸립니다)")
        try:
            products = _match_cache_products("")
            export = export_latest_lot_step_cache(
                products=products or list(built or []),
                update_state=_match_cache_products_cover_all(list(built or products)))
            out["latest_cache"] = export
            _match_cache_progress(
                job_id, f"[FAB 후처리] 통합 latest-lot 캐시 완료 — "
                        f"{int(export.get('row_count') or 0):,}행 / "
                        f"{len(export.get('products') or [])}개 제품")
        except Exception as e:
            logger.warning("SplitTable unified latest cache export after match cache failed: %s",
                           e, exc_info=True)
            out["latest_cache"] = {"ok": False, "reason": f"{type(e).__name__}: {e}"}
            _match_cache_progress(job_id, f"[FAB 후처리] 통합 latest-lot 캐시 실패: {e}", ok=False)
    if not _MATCH_CACHE_STOP.is_set():
        _match_cache_progress(job_id, "[FAB 후처리] LOT 진행 캐시 갱신 중…")
        try:
            from core.lot_progress_cache import refresh_lot_progress_cache
            # 진행 문구를 그대로 이벤트 로그 + heartbeat 로 흘린다 — 이 구간이 수십 분
            # 걸리는데 예전에는 '갱신 중…' 한 줄만 남아 멈춘 것과 구분이 안 됐다.
            state = refresh_lot_progress_cache(
                force=force,
                required_products=list(built or []),
                progress=lambda msg: _match_cache_progress(
                    job_id, f"[FAB 후처리] LOT 진행 캐시 — {msg}"))
            out["lot_progress"] = {"ok": True}
            _match_cache_progress(
                job_id, "[FAB 후처리] LOT 진행 캐시 완료 — "
                        f"{int((state or {}).get('count') or 0):,}건")
        except Exception as e:
            logger.warning("LOT progress cache refresh after SplitTable match cache failed: %s", e)
            out["lot_progress"] = {"ok": False, "reason": f"{type(e).__name__}: {e}"}
            _match_cache_progress(job_id, f"[FAB 후처리] LOT 진행 캐시 실패: {e}", ok=False)
    return out


def _run_started_match_cache_job(products: list[str], force: bool, reason: str = "manual",
                                 refresh_plan_risk: bool = False, job_id: str = "",
                                 followups: bool = True) -> dict:
    """제품별 FAB 매칭 캐시 빌드. followups=False 면 파생 캐시 갱신은 호출자가 맡는다.

    통합 스캔은 followups=False 로 부른다 — 파생 캐시(통합 latest-lot export 등)를
    이 단계 안에서 돌리면 제품이 100% 끝난 뒤에도 stage 가 안 끝나, 화면에서는
    'FAB 100% 인데 다음 단계로 안 넘어감' 으로 보인다. 제품 원본 캐시는 파생 캐시에
    의존하지 않으므로 스캔은 바로 다음 단계로 넘어가고, 파생 캐시는 Root lot 단계
    직전(실제로 필요한 지점)에 별도 진행 표시와 함께 돈다.
    """
    pause_s = _match_cache_product_pause_seconds()
    total = len(products)
    try:
        if not products:
            return {"ok": True, "queued": False, "products": [], "job": _match_cache_job_status()}
        for idx, raw_product in enumerate(products):
            if _MATCH_CACHE_STOP.is_set():
                break
            if not _wait_for_match_cache_memory():
                break
            _match_cache_job_update(current_product=raw_product, paused=False)
            try:
                from core import worker_dispatch as _wd
                automatic = reason not in {"manual", "unified_scan"}

                result = _wd.run_heavy(
                    "splittable_match_cache_refresh",
                    {"product": raw_product, "force": bool(force)},
                    lambda: _refresh_match_cache_products([raw_product], force=force),
                    label=f"match_cache:{raw_product}",
                    local_idle_only=(reason != "unified_scan"),
                    local_fallback=not automatic,
                    durable=automatic,
                    priority="maintenance" if automatic else "normal",
                    dedupe_key=f"match_cache:{raw_product}",
                    timeout_sec=6 * 3600.0 if automatic else None,
                ) or {"ok": False, "products": []}
            except Exception as e:
                logger.warning("SplitTable match cache queued build failed (product=%s) %s: %s",
                               raw_product, type(e).__name__, e, exc_info=True)
                result = {
                    "ok": False,
                    "products": [{
                        "product": raw_product,
                        "ok": False,
                        "skipped": False,
                        "row_count": 0,
                        "reason": f"{type(e).__name__}: {e}",
                    }],
                    "interval_minutes": _match_cache_refresh_minutes(),
                }
            _match_cache_job_append_products(result.get("products") or [])
            # 이 제품에 걸린 중단 신호는 여기서 푼다 — 안 풀면 다음 제품까지
            # 연쇄로 끊기거나, 다음 스캔이 시작하자마자 다시 중단된다.
            cancelled = any(r.get("cancelled") for r in (result.get("products") or []))
            if cancelled or _match_cache_cancelled(raw_product):
                clear_match_cache_cancel(raw_product)
            # 제품 하나가 수 분 걸릴 수 있다 — 완료마다 진행 신호를 남겨야 작업이
            # '응답 없음'으로 오판되지 않는다.
            _match_cache_progress(
                job_id,
                f"[FAB 매칭] {idx + 1}/{total} {raw_product} "
                + ("관리자 중단 — 다음 제품으로 넘어갑니다 (이 제품은 다음 스캔에서 처음부터)"
                   if cancelled else "처리 완료"),
                ok=not cancelled,
            )
            if idx < len(products) - 1 and pause_s > 0:
                _MATCH_CACHE_STOP.wait(pause_s)
        if followups:
            _match_cache_followups(force, job_id=job_id, built=products,
                                   refresh_plan_risk=refresh_plan_risk)
    finally:
        _match_cache_job_update(
            running=False,
            queued=False,
            current_product="",
            paused=False,
            finished_at=datetime.datetime.now().isoformat(timespec="seconds"),
        )
    status = _match_cache_job_status()
    return {
        "ok": bool(status.get("ok_count")),
        "queued": False,
        "products": status.get("products") or [],
        "interval_minutes": _match_cache_refresh_minutes(),
        "job": status,
        "reason": reason,
    }


def enqueue_match_cache_refresh(product: str = "", force: bool = True, reason: str = "manual",
                                refresh_plan_risk: bool = False) -> dict:
    """Queue a paced match-cache refresh and return immediately.

    서버 스캔 게이트를 통과하므로 다른 스캔(통합 스캔·전체 셋업·예약)과 절대
    겹치지 않는다. job state(_begin_match_cache_job)는 큐에서 꺼내 **실제로
    시작할 때** 잡는다 — 대기 중에 running 으로 잡아두면 다른 스캔이 '이미
    실행 중'으로 오판하고, 화면에는 진행 없는 running 이 오래 떠 있게 된다."""
    products = _match_cache_products(product)
    label = f"FAB 매칭 캐시 갱신 ({product or '전체 제품'})"

    def _start() -> dict:
        started, status = _begin_match_cache_job(products, force=force, reason=reason)
        if not started:
            return {
                "ok": True,
                "skipped": True,
                "job": status,
                "detail": "SplitTable match cache refresh is already running.",
            }
        return _run_started_match_cache_job(products, force, reason,
                                            refresh_plan_risk=refresh_plan_risk)

    out = _submit_scan("match_cache", label, _start, product=product,
                       source="scheduler" if reason == "scheduler" else "manual",
                       dedupe_key=f"match_cache:{reason}:{product}")
    return {
        **out,
        "products": [{"product": p, "queued": True} for p in products],
        "interval_minutes": _match_cache_refresh_minutes(),
        "job": _match_cache_job_status(),
    }


def _seconds_until_next_match_cache_tick() -> float:
    return max(60.0, _match_cache_refresh_minutes() * 60.0)


def _match_cache_loop() -> None:
    global _MATCH_CACHE_NEXT_TICK_AT
    while not _MATCH_CACHE_STOP.is_set():
        try:
            from core.background_owner import is_owner
            if not is_owner():
                _MATCH_CACHE_STOP.wait(5.0)
                continue
            freshness = _match_cache_global_fresh()
            if freshness.get("fresh"):
                logger.info("SplitTable match cache scheduler skipped; latest cache fresh until %s",
                            freshness.get("next_refresh_at") or "")
            else:
                # 예약 갱신도 수동 스캔과 같은 게이트를 통과한다 — 예전엔 통합
                # 스캔이 2/3 단계에 들어간 사이 이 tick 이 FAB 전체 스캔을 새로
                # 시작해 한 서버에서 두 스캔이 겹쳤다.
                enqueue_match_cache_refresh(product="", force=False, reason="scheduler",
                                            refresh_plan_risk=True)
        except Exception as e:
            logger.warning("SplitTable match cache scheduler tick failed: %s", e)
        wait_s = _seconds_until_next_match_cache_tick()
        _MATCH_CACHE_NEXT_TICK_AT = datetime.datetime.fromtimestamp(
            time.time() + wait_s).isoformat(timespec="seconds")
        while wait_s > 0 and not _MATCH_CACHE_STOP.is_set():
            step = min(wait_s, 60.0)
            _MATCH_CACHE_STOP.wait(step)
            wait_s -= step


def start_match_cache_scheduler() -> bool:
    global _MATCH_CACHE_THREAD, _MATCH_CACHE_STARTED
    if _MATCH_CACHE_STARTED:
        return False
    if _auto_product_cache_enabled():
        logger.info("SplitTable match-cache timer retired; product rotation owns refresh")
        return False
    try:
        from core.runtime_limits import splittable_match_cache_enabled
        if not splittable_match_cache_enabled():
            logger.info("SplitTable match cache scheduler disabled")
            return False
    except Exception:
        pass
    _MATCH_CACHE_STOP.clear()
    _MATCH_CACHE_THREAD = threading.Thread(target=_match_cache_loop, name="splittable-match-cache", daemon=True)
    _MATCH_CACHE_THREAD.start()
    _MATCH_CACHE_STARTED = True
    logger.info("SplitTable match cache scheduler started (interval=%sm)", _match_cache_refresh_minutes())
    return True


class MatchCacheRefreshReq(BaseModel):
    product: str = ""
    force: bool = True


class MatchCacheStopReq(BaseModel):
    # 비우면 지금 돌고 있는 제품을 중단한다.
    product: str = ""


class ProductRamCacheRefreshReq(BaseModel):
    product: str = ""
    force: bool = True


class RootLotRamCacheRefreshReq(BaseModel):
    product: str = ""
    force: bool = True


@router.get("/match-cache/status")
def match_cache_status(request: Request, product: str = Query("")):
    me = current_user(request)
    if me.get("role") != "admin":
        raise HTTPException(403, "admin only")
    try:
        from core.runtime_limits import splittable_match_cache_enabled
        enabled = splittable_match_cache_enabled()
    except Exception:
        enabled = True
    products = [product] if str(product or "").strip() else [p.get("name") for p in list_products().get("products", [])]
    rows = []
    for prod in [p for p in products if p]:
        current = _match_cache_current(prod)
        if not current:
            continue
        meta = current.get("meta") or {}
        rows.append({
            "product": current.get("product") or prod,
            "fab_source": current.get("fab_source") or "",
            "path": str(current.get("path") or ""),
            "built_at": meta.get("built_at", ""),
            "row_count": int(meta.get("row_count") or 0),
            "join_keys": meta.get("join_keys") or [],
        })
    return {
        "ok": True,
        "enabled": enabled,
        "interval_minutes": _match_cache_refresh_minutes(),
        "products": rows,
        "latest_cache": _match_cache_global_fresh(),
        "job": _match_cache_job_status(),
    }


@router.post("/match-cache/refresh")
def refresh_match_cache_now(req: MatchCacheRefreshReq, request: Request, _a=Depends(require_page_manager("splittable"))):
    return enqueue_match_cache_refresh(product=req.product or "", force=bool(req.force), reason="manual")


@router.post("/match-cache/stop")
def stop_match_cache_product(req: MatchCacheStopReq, request: Request,
                             _a=Depends(require_page_manager("splittable"))):
    """진행 중인 제품 하나의 캐싱을 중단하고 다음 제품으로 넘긴다.

    중단한 제품은 부분 산출을 버리므로 **다음 스캔에서 처음부터** 다시 빌드된다
    (이어받기 없음). 기존에 완성돼 있던 캐시 파일은 그대로 남는다."""
    me = current_user(request)
    job = _match_cache_job_status()
    target = str(req.product or "").strip() or str(job.get("current_product") or "")
    if not target:
        raise HTTPException(409, "중단할 진행 중인 캐싱 작업이 없습니다.")
    if not job.get("running"):
        raise HTTPException(409, "실행 중인 FAB 매칭 캐시 작업이 없습니다.")
    cancel = request_match_cache_cancel(target, by=me.get("username") or "admin")
    try:
        from core.cache_event_log import record as _record
        _record("scan", f"[FAB 매칭] {target} 캐싱 중단 요청 — {me.get('username') or 'admin'}", ok=False)
    except Exception:
        pass
    return {"ok": True, "cancel": cancel, "job": _match_cache_job_status()}


@router.get("/product-cache/status")
def product_ram_cache_status(request: Request, product: str = Query("")):
    me = current_user(request)
    include_detail = is_page_manager(me, "splittable")
    products = _product_ram_cache_products(product)
    rows = [_product_ram_cache_public_meta(prod, include_detail=include_detail) for prod in products]
    return {
        "ok": True,
        "enabled": _product_ram_cache_available(),
        "scheduler_enabled": _product_ram_cache_scheduler_enabled(),
        "interval_minutes": _product_ram_cache_refresh_minutes(),
        "max_gb": round(_product_ram_cache_max_bytes() / (1024 ** 3), 3) if _product_ram_cache_max_bytes() else 0,
        "products": rows,
        "job": _product_ram_cache_job_status() if include_detail else {
            "running": _product_ram_cache_job_status().get("running", False),
            "queued": _product_ram_cache_job_status().get("queued", False),
        },
    }


@router.post("/product-cache/refresh")
def refresh_product_ram_cache_now(req: ProductRamCacheRefreshReq, request: Request):
    me = current_user(request)
    if not is_page_manager(me, "splittable"):
        raise HTTPException(403, "Admin or page manager (splittable) only")
    return {
        "ok": False,
        "disabled": True,
        "product": str(req.product or "").strip(),
        "detail": "제품 전체 RAM 예열은 폐기되어 수동 적재도 실행하지 않습니다.",
    }


@router.get("/root-lot-cache/status")
def root_lot_ram_cache_status(request: Request, product: str = Query("")):
    me = current_user(request)
    include_detail = is_page_manager(me, "splittable")
    source_fp = None
    if str(product or "").strip():
        try:
            source_fp = _product_path(product)
        except Exception:
            source_fp = None
    out = {
        "ok": True,
        "settings": _ml_table_lookup.root_ram_cache_settings(),
        "cache": _ml_table_lookup.root_ram_cache_status(source_fp, include_detail=include_detail),
    }
    # 관리자에게만 최근 검색 타이밍 breakdown 을 노출한다.
    if include_detail:
        out["recent_searches"] = recent_search_timings(limit=30)
    return out


@router.get("/search-timings")
def get_search_timings(request: Request, hours: float = Query(24.0), limit: int = Query(200),
                       origin: str = Query("")):
    """검색 타이밍 기간 조회 + 집계 (관리자 전용).

    공유 JSONL 에 누적된 기록을 읽어 wait(줄서기) / compute(계산) 분포를 낸다.
    slow_wait_pct 가 지속적으로 높으면 cold 레인 슬롯을 늘릴 근거가 되고,
    wait 은 낮은데 compute 가 크면 레인이 아니라 캐시/계산 쪽 문제다.

    origin 은 서버 라벨(예: "운영", "개발(worker)") — 로그가 두 서버 공유라
    기본값(빈 값)은 합산이다. 이 서버만 보려면 origin 을 지정한다."""
    if not is_page_manager(current_user(request), "splittable"):
        raise HTTPException(403, "관리자 전용")
    # "__self__" = 이 서버 — 프런트가 라벨을 미리 알 필요 없이 자기 서버만 볼 수 있게.
    if str(origin or "").strip() == "__self__":
        origin = _search_timing_log.this_origin()
    out = _search_timing_log.query(hours=hours, limit=limit, origin=origin)
    out["ok"] = True
    out["cold_lane"] = _view_cold_lane_stats()
    return out


@router.post("/root-lot-cache/refresh")
def refresh_root_lot_ram_cache_now(req: RootLotRamCacheRefreshReq, _perm=Depends(require_page_manager("splittable"))):
    """관리자: root lot RAM 캐시 적재를 스캔 게이트에 넣는다(비동기).

    예전엔 요청 스레드에서 그대로 적재해 응답이 수십 분 걸릴 수 있었고, 진행
    중인 다른 스캔과도 겹쳤다. 결과는 캐시 이벤트 로그와 `/root-lot-cache/status`
    로 확인한다."""
    return {
        "ok": False,
        "disabled": True,
        "product": str(req.product or "").strip(),
        "detail": "Root lot RAM 예열은 폐기되어 수동 적재도 실행하지 않습니다.",
    }


class ScanQueueCancelReq(BaseModel):
    task_id: str = ""


@router.post("/scan-queue/cancel")
def cancel_scan_queue_task(req: ScanQueueCancelReq, request: Request,
                           _a=Depends(require_page_manager("splittable"))):
    """진행 중이거나 대기 중인 캐시 작업을 중단한다.

    대기 중이면 큐에서 바로 빼고, 진행 중이면 중단을 요청한다 — **즉시 멈추지
    않는다.** 현재 제품/배치가 끝나는 안전한 지점에서 접고 다음 큐 작업으로
    넘어간다(스레드를 강제 종료하면 parquet 파티션이 깨진다).

    이미 완성된 제품 캐시는 그대로 남고 즉시 계속 사용한다. FAB 랏 인덱스는
    완료 배치를 staging에 보존해 다음 실행에서 이어받는다.
    """
    from core import scan_gate

    me = current_user(request)
    out = scan_gate.cancel(str(req.task_id or ""), by=me.get("username") or "admin")
    if not out.get("ok"):
        raise HTTPException(409, out.get("detail") or "중단할 작업을 찾을 수 없습니다.")
    # 자동 순환 작업이 아직 대기 중일 때는 scan_gate 에서 제거하는 것만으로는
    # 스케줄러의 queued_product 가 남는다. 그러면 다음 poll 에서 취소한 작업이
    # 다시 현재/다음으로 보이거나 재등록된다. 같은 취소 응답으로 커서도 정리한다.
    if out.get("state") == "removed" and str(out.get("source") or "") == "scheduler":
        _auto_product_cache_on_cancelled(
            str(out.get("product") or ""), str(out.get("id") or "")
        )
    # FAB 매칭 단계가 돌고 있으면 제품 단위 중단 신호도 함께 세운다 — 그 단계는
    # 자체 취소 경로(_MATCH_CACHE_CANCEL)로 현재 배치를 접는 게 가장 빠르다.
    try:
        job = _match_cache_job_status()
        if job.get("running") and str(job.get("current_product") or "").strip():
            request_match_cache_cancel(str(job.get("current_product")),
                                       by=me.get("username") or "admin")
    except Exception:
        logger.debug("scan queue cancel: match cache cancel passthrough failed", exc_info=True)
    return {**out, "queue": _scan_gate_snapshot()}


class LotListCacheRefreshReq(BaseModel):
    product: str = ""


@router.get("/lot-pool/status")
def lot_list_cache_status(request: Request):
    """제품별 root_lot_id 풀 캐시 상태 (관리자 전용).

    드롭다운이 느리다는 신고가 오면 여기부터 본다 — `entries` 에 제품이 없거나
    `hit_ram`/`hit_disk` 대비 `miss` 가 크면 소스 시그니처가 계속 흔들리고
    있다는 뜻이다(스케줄러가 캐시를 자주 다시 쓰는 경우)."""
    if not is_page_manager(current_user(request), "splittable"):
        raise HTTPException(403, "관리자 전용")
    return {**_lot_list_cache.stats(), "prewarm": dict(_CANDIDATE_PREWARM_LAST)}


@router.post("/lot-pool/refresh")
def refresh_lot_list_cache(req: LotListCacheRefreshReq,
                           _perm=Depends(require_page_manager("splittable"))):
    """관리자: root_lot_id 풀 캐시를 비우고 즉시 다시 만든다.

    product 를 비우면 전체를 버리기만 한다(다음 조회에서 각자 다시 빌드).
    """
    product = str(req.product or "").strip()
    dropped = _invalidate_root_lot_pool(product)
    rebuilt = 0
    fab_rebuilt = 0
    if product:
        try:
            pool = _root_lot_pool(product)
            rebuilt = len(pool.get("values") or [])
            fab_pool = _fab_lot_pool(product)
            fab_rebuilt = len(fab_pool.get("values") or [])
        except HTTPException:
            raise
        except Exception as e:
            logger.warning("/lot-pool/refresh 재빌드 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
    return {"ok": True, "product": product, "dropped": dropped,
            "root_lot_id_count": rebuilt, "lot_id_count": fab_rebuilt}


class RootLotRamCacheEvictReq(BaseModel):
    source_path: str = ""
    root_lot_id: str = ""


@router.post("/root-lot-cache/evict")
def evict_root_lot_ram_cache_entry(req: RootLotRamCacheEvictReq, _perm=Depends(require_page_manager("splittable"))):
    """관리자: 개별 root lot 캐시 항목 제거."""
    return _ml_table_lookup.evict_root_ram_cache_entry(source_path=req.source_path, root_lot_id=req.root_lot_id)


# ── 통합 수동 스캔 (FAB + product + root lot 일괄) ──────────────────────

class UnifiedScanReq(BaseModel):
    product: str = ""
    force: bool = True


_UNIFIED_SCAN_LOCK = threading.Lock()
_UNIFIED_SCAN_BUSY = False
_UNIFIED_SCAN_JOB_ID = ""
_UNIFIED_SCAN_THREAD: threading.Thread | None = None

# 대기 중 진행 로그 간격 — 대기도 '진행'으로 보이게 해 정지와 구분한다.
_SCAN_WAIT_LOG_GAP_SEC = 15.0


def _scan_stage_timeout_sec() -> float:
    return max(60.0, min(6 * 3600.0, _env_float("FLOW_UNIFIED_SCAN_STAGE_TIMEOUT_SEC", 3600.0)))


def _submit_scan(kind: str, label: str, run, *, product: str = "",
                 source: str = "manual", dedupe_key: str = "") -> dict:
    """스캔 작업을 서버 단위 게이트(core.scan_gate)에 넘긴다.

    한 서버에서 스캔은 항상 하나만 돈다 — 이미 다른 스캔이 진행 중이면 거절하지
    않고 대기열에 넣고, 앞 작업이 성공하든 실패하든 끝나면 이어서 실행한다.
    게이트를 못 쓰는 예외 상황(모듈 로드 실패)에서는 예전처럼 전용 스레드로
    떨어뜨려, 스캔 자체가 불가능해지지는 않게 한다."""
    try:
        from core import scan_gate
    except Exception:
        logger.warning("scan gate unavailable — running %s without serialization", kind, exc_info=True)
        threading.Thread(target=run, name=f"splittable-{kind}", daemon=True).start()
        return {"ok": True, "queued": True, "ahead": 0, "detail": "스캔 시작됨(직렬화 게이트 없음)"}
    return scan_gate.submit(kind, label, run, product=product, source=source,
                            dedupe_key=dedupe_key)


def _reap_dead_unified_scan() -> bool:
    """스캔 스레드가 죽었는데 busy 플래그만 남은 상태를 정리한다.

    이 방어가 없으면 스레드가 예기치 않게 사라졌을 때 이후 모든 스캔 요청이
    '이미 실행 중'으로 거부되고 화면은 영원히 '스캔 중...'에 머문다.
    반환: 정리했으면 True."""
    global _UNIFIED_SCAN_BUSY, _UNIFIED_SCAN_JOB_ID, _UNIFIED_SCAN_THREAD
    with _UNIFIED_SCAN_LOCK:
        if not _UNIFIED_SCAN_BUSY:
            return False
        thread = _UNIFIED_SCAN_THREAD
        if thread is not None and thread.is_alive():
            return False
        job_id = _UNIFIED_SCAN_JOB_ID
        _UNIFIED_SCAN_BUSY = False
        _UNIFIED_SCAN_THREAD = None
    logger.warning("unified scan thread vanished — busy flag cleared (job=%s)", job_id)
    try:
        from core.cache_event_log import finish_job, record
        if job_id:
            finish_job(job_id, ok=False, detail={"error": "scan_thread_vanished"})
        record("scan", "[작업 실패] 스캔 스레드가 종료되어 작업을 중단 처리했습니다 "
                       "(서버 재시작/강제 종료 등). 다시 실행할 수 있습니다.", ok=False)
    except Exception:
        logger.debug("unified scan reap logging failed", exc_info=True)
    return True


def _wait_for_cache_job(status_fn, *, stage: str, stage_label: str = "", job_id: str = "") -> dict:
    """이미 실행 중인 캐시 stage가 끝날 때까지 통합 스캔 스레드에서 대기.

    대기 동안 주기적으로 진행 로그를 남긴다 — 예전엔 최대 1시간을 아무 로그 없이
    대기해 화면이 멈춘 것처럼 보였다."""
    from core.cache_event_log import record as _rec, heartbeat as _beat
    timeout = _scan_stage_timeout_sec()
    label = stage_label or stage
    started = time.monotonic()
    deadline = started + timeout
    next_log = started + _SCAN_WAIT_LOG_GAP_SEC
    _rec("scan", f"[대기] {label}: 다른 요청이 이미 실행 중 — 끝날 때까지 기다립니다"
                 f" (최대 {_fmt_dur_ko(timeout)})",
         detail={"job_id": job_id, "stage": stage, "phase": "waiting"} if job_id else {"stage": stage})
    while True:
        now = time.monotonic()
        status = status_fn()
        if not status.get("running"):
            return {
                "ok": bool(status.get("ok_count")),
                "queued": False,
                "products": status.get("products") or [],
                "job": status,
                "joined_existing": True,
            }
        if now >= deadline:
            break
        if now >= next_log:
            next_log = now + _SCAN_WAIT_LOG_GAP_SEC
            _beat(job_id, f"{label} 대기 중")
            done = int(status.get("done") or 0)
            total = int(status.get("total") or 0)
            progress = f" · {done}/{total} 진행" if total else ""
            _rec("scan",
                 f"[대기] {label}: 실행 중인 작업 대기 {_fmt_dur_ko(now - started)} 경과{progress}"
                 f" (현재: {status.get('current_product') or '-'})",
                 detail={"job_id": job_id, "stage": stage, "phase": "waiting"} if job_id else {"stage": stage})
        time.sleep(1.0)
    _rec("scan", f"[실패] {label}: 대기 시간 초과({_fmt_dur_ko(timeout)}) — 앞선 작업이 끝나지 않았습니다."
                 f" env FLOW_UNIFIED_SCAN_STAGE_TIMEOUT_SEC 로 대기 한도를 조절할 수 있습니다.",
         ok=False, detail={"job_id": job_id, "stage": stage, "phase": "failed"} if job_id else {"stage": stage})
    return {"ok": False, "error": f"{stage}_timeout", "joined_existing": True,
            "detail": f"대기 시간 초과({int(timeout)}초)"}


def _scan_lookup_targets(product: str) -> list[Path]:
    """수동 스캔이 랏(lookup) 캐시를 보장해야 할 ML_TABLE 파일 목록."""
    if str(product or "").strip():
        fp = _ml_table_lookup.resolve_ml_table_file(product=product)
        return [fp] if fp else []
    try:
        return list(_ml_table_lookup.discover_ml_table_files() or [])
    except Exception:
        logger.debug("scan lookup target discovery failed", exc_info=True)
        return []


def _wait_for_lookup_builds(targets: dict[str, Path], *, job_id: str = "",
                            note: str = "랏캐시 빌드") -> set[str]:
    """targets(파일명 → 경로)의 lookup 캐시가 fresh 가 될 때까지 기다린다.

    대기 중에도 진행 로그와 heartbeat 를 남긴다 — 조용한 대기는 화면에서 정지와
    구분되지 않고, heartbeat 가 끊기면 job tracker 가 정지로 보고 실패 처리한다.
    반환: 제한 시간 안에 끝나지 못한 파일명 집합(비어 있으면 전부 완료).

    **기다리는 동안 캐시 슬롯을 반납한다** (`scan_gate.lend`). 이 함수는 스캔
    게이트 작업 안에서 불리고, 기다리는 대상인 빌드 스레드는 이제 같은 슬롯을
    잡아야 한다 — 붙들고 기다리면 서로를 기다리는 교착이 된다. 대기 중인 이
    스레드는 자원을 쓰지 않으므로 반납이 안전하고, 빌드가 끝나면 다시 잡는다."""
    from core.cache_event_log import record as _rec, heartbeat as _beat
    pending = set(targets)
    if not pending:
        return set()
    total = len(pending)
    timeout = _scan_stage_timeout_sec()
    started = time.monotonic()
    deadline = started + timeout
    next_log = started + _SCAN_WAIT_LOG_GAP_SEC
    _detail = {"job_id": job_id, "stage": "root_lot_ram", "phase": "waiting"}

    def _wait_loop() -> set[str]:
        nonlocal next_log
        while pending:
            now = time.monotonic()
            for name in sorted(pending):
                fp = targets.get(name)
                try:
                    if fp is not None and _ml_table_lookup.cache_status(fp).get("status") == "fresh":
                        pending.discard(name)
                except Exception:
                    logger.debug("lookup cache status check failed file=%s", name, exc_info=True)
            if not pending or now >= deadline:
                break
            if now >= next_log:
                next_log = now + _SCAN_WAIT_LOG_GAP_SEC
                _beat(job_id, f"{note} 대기 중")
                queue = _ml_table_lookup.build_queue_snapshot()
                current = Path(str(queue.get("current") or "")).name or "-"
                last_error = str(queue.get("last_error") or "")
                _rec("scan",
                     f"[대기] {note} {total - len(pending)}/{total} 완료 · "
                     f"{_fmt_dur_ko(now - started)} 경과 · 빌드 중: {current}"
                     + (f" · 최근 오류: {last_error}" if last_error else "")
                     + " · 남은 제품: " + ", ".join(sorted(pending)[:3])
                     + ("…" if len(pending) > 3 else ""),
                     detail=_detail)
            time.sleep(1.0)
        return pending

    try:
        from core import scan_gate as _gate
        _lend = _gate.lend()
    except Exception:
        logger.debug("scan gate lend unavailable during lookup build wait", exc_info=True)
        import contextlib as _ctx
        _lend = _ctx.nullcontext()
    with _lend:
        return _wait_loop()


def _ensure_scan_lookup_caches(product: str = "", *, job_id: str = "") -> dict:
    """랏(lookup) 디스크 캐시를 대상 제품에 대해 빌드하고 완료까지 기다린다.

    랏캐시는 프로세스 로컬 RAM 이 아니라 운영/개발 두 서버가 함께 읽는 공유
    파티션 산출물이다. 그래서 root RAM 캐시가 꺼진 서버(개발/워커 역할)에서도
    이 단계는 돌아야 한다 — 예전엔 RAM 이 비활성이면 3/3 단계 전체를 건너뛰어,
    개발서버에서 수동 스캔을 아무리 돌려도 랏캐시가 만들어지지 않았다.

    fresh 한 캐시는 건드리지 않는다. 화면의 수동 스캔은 항상 force=True 로
    오지만, 랏캐시에서 force 는 예전에도 '무조건 재빌드'가 아니라 자동 빌드
    크기 상한을 넘긴다는 뜻이었다 — 매 스캔마다 전 제품을 다시 빌드하면
    몇 시간짜리 작업이 된다."""
    from core.cache_event_log import record as _rec
    _detail = {"job_id": job_id, "stage": "root_lot_ram", "phase": "waiting"}
    files = _scan_lookup_targets(product)
    if not files:
        return {"ok": False, "total": 0, "built": 0,
                "detail": "랏캐시 대상 ML_TABLE_*.parquet 파일을 찾지 못했습니다"}
    targets: dict[str, Path] = {}
    for fp in files:
        try:
            fresh = _ml_table_lookup.cache_status(fp).get("status") == "fresh"
        except Exception:
            fresh = False
        if not fresh:
            targets[Path(fp).name] = Path(fp)
    if not targets:
        return {"ok": True, "total": len(files), "built": 0, "skipped": True,
                "reason": "fresh", "detail": "랏캐시가 이미 최신입니다"}
    for fp in targets.values():
        # immediate=True: 관리자가 직접 요청한 빌드는 idle 창을 기다리지 않는다.
        # 진행 화면이 2.5초마다 폴링하는 동안 서버는 절대 idle 이 되지 않아,
        # idle 대기가 곧 무한 정지였다.
        _ml_table_lookup.enqueue_build(fp, immediate=True)
    _rec("scan",
         f"[대기] 랏캐시 빌드 — {len(targets)}/{len(files)}개 제품 빌드 큐 등록"
         f" (최대 {_fmt_dur_ko(_scan_stage_timeout_sec())})",
         detail=_detail)
    pending = _wait_for_lookup_builds(targets, job_id=job_id)
    built = len(targets) - len(pending)
    if pending:
        _rec("scan",
             f"[실패] 랏캐시 빌드 대기 시간 초과 — 미완료 {len(pending)}/{len(targets)}개: "
             + ", ".join(sorted(pending)[:5]),
             ok=False,
             detail={"job_id": job_id, "stage": "root_lot_ram", "phase": "failed"})
        return {"ok": False, "total": len(files), "built": built,
                "pending_files": sorted(pending),
                "detail": f"랏캐시 빌드 대기 시간 초과 — 미완료 {len(pending)}개"}
    _rec("scan", f"[대기] 랏캐시 빌드 {built}/{len(targets)} 완료", detail=_detail)
    return {"ok": True, "total": len(files), "built": built}


def _wait_for_root_lookup_caches(result: dict, *, product: str, job_id: str = "") -> dict:
    """root RAM 예열이 요청한 lookup 파티션 빌드를 기다린 뒤 한 번 재시도.

    빌드는 수 분~수십 분 걸릴 수 있으므로 대기 진행을 주기적으로 로그에 남긴다."""
    from core.cache_event_log import record as _rec, heartbeat as _beat
    pending = {
        str(row.get("file") or "")
        for row in (result.get("products") or [])
        if row.get("build_pending") and row.get("file")
    }
    if not pending:
        return result
    total = len(pending)
    timeout = _scan_stage_timeout_sec()
    started = time.monotonic()
    deadline = started + timeout
    next_log = started + _SCAN_WAIT_LOG_GAP_SEC
    _detail = {"job_id": job_id, "stage": "root_lot_ram", "phase": "waiting"}
    _rec("scan", f"[대기] 랏캐시(원본 lookup) 빌드 대기 — {total}개 제품 "
                 f"(최대 {_fmt_dur_ko(timeout)}). 빌드 진행은 [랏캐시빌드] 로그에 표시됩니다.",
         detail=_detail)
    while pending:
        now = time.monotonic()
        ready: set[str] = set()
        for file_name in pending:
            fp = _ml_table_lookup.resolve_ml_table_file(file=file_name)
            if fp is not None and _ml_table_lookup.cache_status(fp).get("status") == "fresh":
                ready.add(file_name)
        pending.difference_update(ready)
        if not pending:
            break
        if now >= deadline:
            break
        if now >= next_log:
            next_log = now + _SCAN_WAIT_LOG_GAP_SEC
            _beat(job_id, "랏캐시 빌드 대기 중")
            _rec("scan",
                 f"[대기] 랏캐시 빌드 {total - len(pending)}/{total} 완료 · "
                 f"{_fmt_dur_ko(now - started)} 경과 · 남은 제품: "
                 + ", ".join(sorted(pending)[:3]) + ("…" if len(pending) > 3 else ""),
                 detail=_detail)
        time.sleep(1.0)
    if pending:
        _rec("scan",
             f"[실패] 랏캐시 빌드 대기 시간 초과({_fmt_dur_ko(timeout)}) — "
             f"미완료 {len(pending)}/{total}개: " + ", ".join(sorted(pending)[:5]),
             ok=False, detail={"job_id": job_id, "stage": "root_lot_ram", "phase": "failed"})
        out = dict(result)
        out.update(ok=False, error="root_lookup_build_timeout", pending_files=sorted(pending),
                   detail=f"랏캐시 빌드 대기 시간 초과 — 미완료 {len(pending)}개")
        return out
    _rec("scan", f"[대기] 랏캐시 빌드 {total}/{total} 완료 — RAM 적재를 이어서 진행합니다",
         detail=_detail)
    # lookup build 완료 직후 운영 프로세스 RAM에 실제 root frame을 올린다.
    return _ml_table_lookup.refresh_root_lot_ram_cache(product=product, force=False, load_now=True)


_MANUAL_LATEST_REFRESH_LOCK = threading.Lock()
_MANUAL_LATEST_REFRESH_RUNNING = False


def _enqueue_manual_lot_progress_refresh(products: list[str]) -> bool:
    """Queue one normal-priority refresh of the shared WIP/latest-lot cache."""
    global _MANUAL_LATEST_REFRESH_RUNNING
    with _MANUAL_LATEST_REFRESH_LOCK:
        if _MANUAL_LATEST_REFRESH_RUNNING:
            return False
        _MANUAL_LATEST_REFRESH_RUNNING = True

    def _run() -> None:
        global _MANUAL_LATEST_REFRESH_RUNNING
        label = ", ".join(products or []) or "전체 제품"
        started_ts = time.time()
        # pivot 빌드와 같은 이유로 실행 자체를 이벤트 로그에 남긴다 — 예전에는
        # "큐 등록 완료"만 남고 실제 갱신은 서버 터미널에만 찍혔다.
        for _prod in (products or [""]):
            _cache_build_emit(_prod, f"[WIP latest-lot] 갱신 시작 — {label}",
                              detail={"products": list(products or []),
                                      "stage": _stage("latest_lot", "start")})
        ok = False
        detail: dict = {}
        try:
            from core import lot_progress_cache as _lpc
            from core import worker_dispatch as _wd

            def _local_refresh() -> dict:
                state = _lpc.refresh_lot_progress_cache(
                    force=True, required_products=list(products or []))
                return {"ok": bool((state or {}).get("generated_at")),
                        "count": int((state or {}).get("count") or 0)}

            res = _wd.run_heavy(
                "splittable_lot_progress_cache_refresh",
                {"force": True, "required_products": list(products or [])},
                _local_refresh,
                label="WIP latest lot cache",
                local_idle_only=False,
                local_fallback=True,
                durable=False,
                priority="normal",
                dedupe_key="lot_progress_refresh",
            )
            ok = bool((res or {}).get("ok"))
            detail = {"count": int((res or {}).get("count") or 0)}
        except Exception as exc:
            logger.warning("manual WIP/latest-lot cache refresh failed", exc_info=True)
            detail = {"error": str(exc)}
        finally:
            elapsed = round(time.time() - started_ts, 1)
            for _prod in (products or [""]):
                _cache_build_emit(
                    _prod,
                    (f"[WIP latest-lot] 갱신 완료 — {detail.get('count', 0)}건 · {elapsed}s"
                     if ok else f"[WIP latest-lot] 갱신 실패 — {detail.get('error', '결과 없음')} ({elapsed}s)"),
                    ok=ok, detail={**detail, "elapsed_sec": elapsed,
                                   "stage": _stage("latest_lot", "done" if ok else "fail")},
                )
            with _MANUAL_LATEST_REFRESH_LOCK:
                _MANUAL_LATEST_REFRESH_RUNNING = False

    threading.Thread(target=_run, daemon=True, name="manual-lot-progress-refresh").start()
    return True


def _refresh_dashboard_latest_v4(products: list[str], *, force: bool,
                                 job_id: str = "") -> dict:
    """Build and atomically publish the dashboard's canonical format-v4 cache.

    The generic LOT-progress scanner owns ``lot_wf_current.parquet`` only.  The
    dashboard canonical file needs the SplitTable match projection (including
    its embedded format/source columns), so rebuilding only the scanner cache
    can never upgrade an old canonical file to v4.  Build all requested match
    inputs first and publish only after every requested product succeeded; a
    failed refresh therefore leaves the previous canonical generation intact.
    """
    targets = [str(value or "").strip() for value in products if str(value or "").strip()]
    if not targets:
        return {"ok": False, "error": "no_products", "latest_cache": {}}
    if job_id:
        from core.cache_event_log import heartbeat
        heartbeat(job_id, "대시보드 format v4 입력 캐시 빌드 중")
    match = _refresh_match_cache_products(targets, force=force)
    rows = list(match.get("products") or [])
    failed = [row for row in rows if not row.get("ok")]
    if failed or len(rows) < len(targets):
        return {
            "ok": False,
            "error": "match_cache_incomplete",
            "failed": failed,
            "match_cache": match,
            "latest_cache": {},
        }

    # Export from the complete catalog so a product-specific refresh does not
    # replace the shared dashboard file with a one-product subset.  The writer
    # uses .tmp + replace, preserving the old v4 file throughout the build.
    export_products = _match_cache_products("") or targets
    export = export_latest_lot_step_cache(
        products=export_products,
        update_state=_match_cache_products_cover_all(targets),
    )
    exported = _canonical_product_set(list(export.get("products") or []))
    expected = _canonical_product_set(targets)
    ok = bool(export.get("ok") and expected and expected.issubset(exported))
    if job_id:
        from core.cache_event_log import heartbeat
        heartbeat(
            job_id,
            (f"대시보드 format v{LATEST_LOT_STEP_CACHE_FORMAT_VERSION} 완료 — "
             f"{int(export.get('row_count') or 0):,}행") if ok
            else f"대시보드 format v{LATEST_LOT_STEP_CACHE_FORMAT_VERSION} 생성 실패",
        )
    return {
        "ok": ok,
        "error": "" if ok else "canonical_export_incomplete",
        "match_cache": match,
        "latest_cache": export,
    }


def _enqueue_required_split_caches(product: str, force: bool, job_id: str = "", *,
                                   owns_job: bool = True,
                                   local_only: bool = False) -> dict:
    """Build the four required SplitTable caches and keep the scan task cancellable.

    local_only=True 면 네 단계 모두 이 서버에서 직접 빌드한다 (개발 워커 오프로드
    금지). 관리자가 "수동 캐싱" 을 누른 서버가 결과와 진행률을 보는 서버이므로,
    그 서버에서 바로 돌아야 진행 표시와 중단이 모두 같은 곳에서 동작한다.

    The previous implementation returned as soon as it had spawned independent
    daemon queues.  The UI therefore lost its cancellable scan task while the
    expensive work was still running.  Keep this orchestration task alive until
    each real artifact finishes; a cancel request stops before the next product
    or stage while the current safe build batch is allowed to finish.
    """
    global _UNIFIED_SCAN_BUSY
    from core.cache_event_log import heartbeat, record, stage_started, stage_finished

    requested = str(product or "").strip()
    products = _match_cache_products(requested)
    if requested and not products:
        products = [requested]
    results: dict = {
        "ok": True, "product": requested, "mode": "product_pipeline",
        "response_cache": {"queued": False, "reason": "query_specific_read_through"},
        "products": [],
    }
    cancelled = False
    try:
        stage_timeout = max(60.0, float(os.environ.get(
            "FLOW_CACHE_PIPELINE_STAGE_TIMEOUT_SEC", "21600") or 21600.0))
    except Exception:
        stage_timeout = 21600.0

    class _PipelineCancelled(RuntimeError):
        pass

    def _cancel_point() -> None:
        if _scan_cancel_requested():
            raise _PipelineCancelled("관리자 중단 요청")

    def _wait_for(label: str, ready, *, timeout: float | None = None) -> bool:
        from core import scan_gate as _scan_gate
        started = time.monotonic()
        last_beat = 0.0
        limit = stage_timeout if timeout is None else max(1.0, float(timeout))
        # The actual lookup/pivot/FAB builder runs on another thread and uses
        # the same single-cache slot.  Lend that slot while merely waiting or
        # the orchestrator and its child builder would deadlock each other.
        with _scan_gate.lend():
            while time.monotonic() - started < limit:
                _cancel_point()
                if ready():
                    return True
                now = time.monotonic()
                if job_id and now - last_beat >= 2.0:
                    heartbeat(job_id, label)
                    last_beat = now
                time.sleep(0.25)
        return False

    def _release_stage_memory(stage_id: str) -> dict:
        """단계 사이에서 작업 메모리를 OS 로 돌려준다.

        lookup/pivot/latest/fab 네 단계는 각각 수 GB 를 쓰고 끝난다. gc 만으로는
        allocator arena 에 남아 RSS 가 단계마다 누적됐고, 마지막에는 워치독이
        비울 캐시조차 없는 상태로 임계값에 붙어 있었다(`freed 0.0MB [nothing]`).
        단계 경계는 참조가 확실히 끊긴 지점이라 여기서 반환하는 게 가장 안전하다.
        """
        try:
            from core import memory_trim

            return memory_trim.trim(reason=f"cache_stage:{stage_id}")
        except Exception:
            try:
                gc.collect()
            except Exception:
                pass
            return {}

    def _stage_run(stage_id: str, label: str, fn) -> bool:
        _cancel_point()
        if job_id:
            stage_started(job_id, stage_id)
        try:
            ok = bool(fn())
            released = int((_release_stage_memory(stage_id) or {}).get("released_bytes") or 0)
            if released > 0:
                logger.info("[cache pipeline] %s 단계 후 %.0fMB OS 반환",
                            stage_id, released / (1024 * 1024))
            if job_id:
                stage_finished(job_id, stage_id, ok=ok,
                               detail={"completed": ok, "released_mb": round(released / (1024 * 1024), 1)})
            record("scan", f"[제품별 캐싱] {label} " + ("완료" if ok else "일부 실패"),
                   ok=ok, product=requested,
                   detail={"job_id": job_id, "stage": stage_id,
                           "phase": "finished" if ok else "failed"})
            if not ok:
                results["ok"] = False
            return ok
        except _PipelineCancelled:
            # 중단도 단계가 끝난 것이다 — 여기서 안 돌려주면 취소한 빌드의 작업
            # 메모리가 그대로 남는다.
            _release_stage_memory(stage_id)
            if job_id:
                stage_finished(job_id, stage_id, ok=False, detail={"cancelled": True})
            raise
        except Exception as exc:
            results["ok"] = False
            _release_stage_memory(stage_id)
            if job_id:
                stage_finished(job_id, stage_id, ok=False, detail={"error": str(exc)})
            record("scan", f"[수동 캐싱] {label} 큐 등록 실패: {exc}", ok=False,
                   product=requested,
                   detail={"job_id": job_id, "stage": stage_id, "phase": "failed"})
            return False

    try:
        rows = []
        for prod in products:
            try:
                rows.append({"product": prod, "source": str(_product_path(prod))})
            except Exception as exc:
                rows.append({"product": prod, "error": str(exc)})
                results["ok"] = False

        def _build_lookup():
            ok = True
            for row in rows:
                _cancel_point()
                if row.get("source"):
                    source = Path(row["source"]).resolve()
                    target = str(source)
                    row["lookup"] = _ml_table_lookup.enqueue_build(
                        source, immediate=True, local_only=local_only)

                    def _lookup_finished() -> bool:
                        # Wait for the request itself, not just the published
                        # generation.  The previous generation deliberately
                        # remains ready while rebuilding, so readiness alone
                        # can otherwise advance this product to the next kind
                        # while its lookup worker is still active.
                        snapshot = _ml_table_lookup.build_queue_snapshot()
                        current = str(snapshot.get("current") or "")
                        queued = {str(value) for value in (snapshot.get("queued") or [])}
                        active = bool(snapshot.get("running") and current == target)
                        return (
                            not active
                            and target not in queued
                            and _ml_table_lookup.cache_status(source).get("status") == "fresh"
                        )
                    finished = _wait_for(
                        f"랏 lookup · {row['product']}",
                        _lookup_finished,
                    )
                    row["lookup_ready"] = bool(finished)
                    ok = ok and bool(finished)
            return ok

        _stage_run("lookup_build", "랏 lookup", _build_lookup)

        def _build_pivot():
            ok = True
            for row in rows:
                _cancel_point()
                if row.get("source"):
                    row["pivot_queued"] = _enqueue_pivot_cache_build(
                        row["product"], reason="manual_queue", immediate=True,
                        local_only=local_only)
                    finished = _wait_for(
                        f"SplitTable pivot · {row['product']}",
                        lambda prod=row["product"]: _pivot_cache_build_state(prod) != "building",
                    )
                    try:
                        ready = finished and not _pivot_cache_needs_build(
                            row["product"], Path(row["source"]))
                    except Exception:
                        ready = False
                    row["pivot_ready"] = bool(ready)
                    ok = ok and bool(ready)
            return ok

        _stage_run("pivot_build", "SplitTable pivot", _build_pivot)

        def _build_latest():
            results["latest_queued"] = bool(_enqueue_manual_lot_progress_refresh(products))
            scanner_done = _wait_for(
                "WIP latest-lot", lambda: not _MANUAL_LATEST_REFRESH_RUNNING)
            if not scanner_done:
                return False
            canonical = _refresh_dashboard_latest_v4(
                products, force=force, job_id=job_id)
            results["dashboard_latest_v4"] = canonical
            return bool(canonical.get("ok"))

        _stage_run("latest_lot", "WIP latest-lot", _build_latest)

        def _build_fab():
            include_all = _foreground_global_fab_scan_enabled()
            ok = True
            for row in rows:
                _cancel_point()
                if not row.get("source"):
                    continue
                _ml_product, _ov, fab_source = _current_fab_override(row["product"])
                row["fab_index_queued"] = _enqueue_fab_lot_index_build(
                    row["product"], fab_source, include_all=include_all,
                    reason="manual_queue", immediate=True, local_only=local_only)
                finished = _wait_for(
                    f"FAB latest · {row['product']}",
                    lambda prod=row["product"]: _pivotless_fab_build_done(prod),
                )
                meta = _fab_lot_index_read_meta(row["product"])
                ready = bool(finished and meta.get("built_at") and meta.get("root_col"))
                row["fab_index_ready"] = ready
                ok = ok and ready
            return ok

        def _pivotless_fab_build_done(prod: str) -> bool:
            canonical = (_canonical_mltable_product_name(prod, allow_bare=True)
                         or str(prod or "").strip().upper())
            with _FAB_IDX_BUILD_LOCK:
                return canonical not in _FAB_IDX_BUILD_INPROGRESS

        _stage_run("fab_index", "FAB latest 인덱스", _build_fab)
        results["products"] = rows
        results["queued_products"] = sum(1 for row in rows if row.get("source"))
        results["detail"] = (
            f"제품별 캐시 파이프라인을 처리했습니다: {results['queued_products']}개 제품 · "
            "lookup → SplitTable → WIP latest → FAB latest."
        )
        record("scan", f"[제품별 캐싱] 처리 완료 — {results['detail']}",
               ok=bool(results["ok"]), product=requested,
               detail={"job_id": job_id, "mode": "product_pipeline"})
        return results
    except _PipelineCancelled as exc:
        cancelled = True
        results["ok"] = False
        results["cancelled"] = True
        results["detail"] = str(exc)
        record("scan", f"[제품별 캐싱] 중단 — {exc}. 현재 안전 배치 이후 다음 단계는 시작하지 않습니다.",
               ok=False, product=requested,
               detail={"job_id": job_id, "cancelled": True})
        return results
    finally:
        if owns_job and job_id:
            try:
                from core.cache_event_log import finish_job
                finish_job(job_id, ok=bool(results.get("ok")),
                           detail={"product": requested, "mode": "product_pipeline",
                                   "cancelled": cancelled})
            except Exception:
                logger.debug("manual cache enqueue job finish failed", exc_info=True)
        if owns_job:
            with _UNIFIED_SCAN_LOCK:
                _UNIFIED_SCAN_BUSY = False


def _run_unified_scan(product: str, force: bool, job_id: str = "", *,
                      owns_job: bool = True, local_only: bool = False) -> dict:
    """FAB 매칭 캐시 → 제품 원본 RAM 캐시 → Root lot RAM 캐시를 순서대로 갱신.

    owns_job=False 면 job 종료/busy 해제를 호출자가 맡는다(전체 셋업이 이 함수를
    Phase B 로 감싸 쓰므로, 감싼 쪽이 끝나기 전에 작업이 '완료'로 닫히면 안 된다)."""
    # 구형 match/product-RAM/root-RAM 직렬 스캔 대신 필수 공유 디스크
    # 산출물만 일반 큐에 넣는다. 아래 본문은 과거 작업 이력 호환용으로 보존한다.
    return _enqueue_required_split_caches(product, force, job_id, owns_job=owns_job,
                                          local_only=local_only)

    global _UNIFIED_SCAN_BUSY
    results: dict = {"ok": True, "product": product}
    _label = product or "전체 제품"
    try:
        from core.cache_event_log import record as _log_event
    except Exception:
        _log_event = None

    def _log(cat, msg, *, ok=True, detail=None, stage="", phase=""):
        """통합 스캔 이벤트를 작업/단계와 함께 남긴다.

        캐시 이벤트 로그는 일반 캐시 이벤트도 함께 보관하므로, 수동 스캔의
        각 단계가 어느 작업에서 나온 것인지 구분할 수 있는 식별자를 넣는다.
        프런트는 이 값을 사용해 FAB / 제품 원본 / Root lot 이력을 별도 열에
        표시한다.
        """
        if _log_event:
            try:
                event_detail = dict(detail or {})
                if job_id:
                    event_detail["job_id"] = job_id
                if stage:
                    event_detail["stage"] = stage
                if phase:
                    event_detail["phase"] = phase
                _log_event(cat, msg, ok=ok, product=product, detail=event_detail or None)
            except Exception:
                pass

    # 단계별 결과 요약 — 마지막에 '무엇이 되고 무엇이 실패했는지'를 한 줄로 남긴다.
    stage_outcomes: dict[str, str] = {}

    def _err_text(payload: dict) -> str:
        """실패 사유를 사람이 읽을 수 있는 한 줄로.

        주의: 최상위 `reason` 은 트리거 사유("unified_scan")라 실패 원인이 아니다 —
        여기서 쓰면 화면에 'unified_scan' 이 실패 사유처럼 뜬다. 실제 원인은
        top-level error/detail 또는 제품별 결과의 reason 에 있다."""
        payload = payload or {}
        for key in ("detail", "error", "last_error"):
            value = str(payload.get(key) or "").strip()
            if value:
                return value
        rows = payload.get("products") or []
        bad = [row for row in rows if isinstance(row, dict) and not row.get("ok")]
        if bad:
            parts = [f"{row.get('product') or row.get('file') or '?'}: "
                     f"{row.get('reason') or row.get('error') or '실패'}"
                     for row in bad[:3]]
            more = f" 외 {len(bad) - 3}개" if len(bad) > 3 else ""
            return "; ".join(parts) + more
        if not rows:
            return ("처리 대상 제품이 없습니다 — 제품명이 올바른지, ML_TABLE 원본 파일이 "
                    "존재하는지 확인하세요")
        return "원인 미상(서버 터미널 로그 확인 필요)"

    try:
        # 이 함수 자체가 이미 background thread다. 각 stage가 다시 별도 thread를
        # 띄우면 FAB scan + 전체 product collect + root warmup이 동시에 실행되어
        # 개발서버 OOM의 직접 원인이 된다. 여기서는 stage를 실제 완료 순서로 직렬화.
        # 1. FAB match cache
        if job_id:
            from core.cache_event_log import stage_started
            stage_started(job_id, "match_cache")
        _log("scan", f"[수동 스캔] 1/3 FAB 매칭 캐시 갱신 시작 ({_label})",
             stage="match_cache", phase="started")
        try:
            match_products = _match_cache_products(product)
            started, _status = _begin_match_cache_job(
                match_products, force=force, reason="unified_scan"
            )
            if started:
                # followups=False — 파생 캐시(통합 latest-lot export 등)는 3/3 직전에
                # 돈다. 여기서 돌리면 제품이 100% 끝나도 이 단계가 안 끝나 화면이
                # 'FAB 100% 인데 다음 단계로 안 넘어감' 상태로 보인다.
                results["match_cache"] = _run_started_match_cache_job(
                    match_products,
                    force,
                    reason="unified_scan",
                    refresh_plan_risk=False,
                    job_id=job_id,
                    followups=False,
                )
            else:
                results["match_cache"] = _wait_for_cache_job(
                    _match_cache_job_status, stage="match_cache",
                    stage_label="FAB 매칭 캐시", job_id=job_id,
                )
            mc = results["match_cache"]
            mc_ok = bool(mc.get("ok"))
            stage_outcomes["1/3 FAB 매칭"] = "완료" if mc_ok else "실패"
            if job_id:
                from core.cache_event_log import stage_finished
                stage_finished(job_id, "match_cache", ok=mc_ok,
                               detail={"products": len(mc.get("products") or []),
                                       "error": "" if mc_ok else _err_text(mc)})
            _log("scan",
                 f"[수동 스캔] 1/3 FAB 매칭 캐시 "
                 + (f"완료 — {len(mc.get('products') or [])}개 제품" if mc_ok
                    else f"실패: {_err_text(mc)}")
                 + f" ({_label})",
                 ok=mc_ok,
                 detail={"products": len(mc.get("products") or [])},
                 stage="match_cache", phase="finished" if mc_ok else "failed")
        except Exception as e:
            results["match_cache"] = {"ok": False, "error": str(e)}
            stage_outcomes["1/3 FAB 매칭"] = "실패"
            if job_id:
                from core.cache_event_log import stage_finished
                stage_finished(job_id, "match_cache", ok=False, detail={"error": str(e)})
            _log("scan", f"[수동 스캔] 1/3 FAB 매칭 캐시 실패 ({_label}): {e}", ok=False,
                 stage="match_cache", phase="failed")

        # 2. Product RAM cache
        #    비활성(설정/서버 역할에서 꺼짐)이면 job 을 시작하지도 않고 즉시 다음
        #    단계로 넘어간다. 예전엔 꺼져 있어도 전 제품을 한 바퀴 돌며 "disabled"
        #    행만 쌓았고, 다른 제품 RAM 작업이 돌고 있으면 _wait_for_cache_job 이
        #    최대 1시간을 기다렸다 — 개발서버처럼 항상 비활성인 곳에서 3/3 랏캐시가
        #    시작조차 못 하던 원인이다.
        if job_id:
            from core.cache_event_log import stage_started
            stage_started(job_id, "product_ram")
        pc_disabled = not _product_ram_cache_available()
        if pc_disabled:
            results["product_cache"] = {"ok": True, "skipped": True, "reason": "disabled",
                                        "products": []}
            stage_outcomes["2/3 제품 원본"] = "건너뜀"
            if job_id:
                from core.cache_event_log import stage_skipped
                stage_skipped(job_id, "product_ram", {"disabled": True, "reason": "설정에서 꺼짐"})
            _log("scan",
                 "[수동 스캔] 2/3 제품 원본 RAM 캐시 건너뜀(비활성 — 설정에서 꺼짐. "
                 f"랏캐시와는 무관하며 랏캐시는 계속 갱신됩니다) ({_label})",
                 detail={"disabled": True}, stage="product_ram", phase="finished")
        else:
            _log("scan", f"[수동 스캔] 2/3 제품 원본 RAM 캐시 갱신 시작 ({_label})",
                 stage="product_ram", phase="started")
            try:
                product_ram_products = _product_ram_cache_products(product)
                started, _status = _begin_product_ram_cache_job(
                    product_ram_products, force=force, reason="unified_scan"
                )
                if started:
                    results["product_cache"] = _run_started_product_ram_cache_job(
                        product_ram_products, force, reason="unified_scan"
                    )
                else:
                    results["product_cache"] = _wait_for_cache_job(
                        _product_ram_cache_job_status, stage="product_ram_cache",
                        stage_label="제품 원본 RAM 캐시", job_id=job_id,
                    )
                pc = results["product_cache"]
                pc_ok = bool(pc.get("ok"))
                stage_outcomes["2/3 제품 원본"] = "완료" if pc_ok else "실패"
                if job_id:
                    from core.cache_event_log import stage_finished
                    stage_finished(job_id, "product_ram", ok=pc_ok,
                                   detail={"products": len(pc.get("products") or []),
                                           "error": "" if pc_ok else _err_text(pc)})
                _log("scan",
                     f"[수동 스캔] 2/3 제품 원본 RAM 캐시 "
                     + (f"완료 — {len(pc.get('products') or [])}개 제품" if pc_ok else
                        f"실패: {_err_text(pc)}")
                     + f" ({_label})",
                     ok=pc_ok,
                     detail={"products": len(pc.get("products") or [])},
                     stage="product_ram", phase="finished" if pc_ok else "failed")
            except Exception as e:
                results["product_cache"] = {"ok": False, "error": str(e)}
                stage_outcomes["2/3 제품 원본"] = "실패"
                if job_id:
                    from core.cache_event_log import stage_finished
                    stage_finished(job_id, "product_ram", ok=False, detail={"error": str(e)})
                _log("scan", f"[수동 스캔] 2/3 제품 원본 RAM 캐시 실패 ({_label}): {e}", ok=False,
                     stage="product_ram", phase="failed")

        # 3. 랏(lookup) 디스크 캐시 → Root lot RAM 적재
        if job_id:
            from core.cache_event_log import stage_started
            stage_started(job_id, "root_lot_ram")
        _log("scan", f"[수동 스캔] 3/3 랏(lookup) 캐시 빌드 + Root lot RAM 적재 시작 ({_label})",
             stage="root_lot_ram", phase="started")
        try:
            # FAB 파생 캐시 — Root lot 예열이 통합 latest-lot 캐시에서 예열 후보
            # root 를 고르므로 여기서(실제로 필요한 지점에서) 먼저 갱신한다.
            # 1/3 안에 있던 것을 옮긴 것이며, 진행 로그가 함께 나간다.
            if not _MATCH_CACHE_STOP.is_set():
                results["match_followups"] = _match_cache_followups(
                    force, job_id=job_id, built=_match_cache_products(product))
            # 랏캐시(디스크 파티션)는 운영/개발이 함께 읽는 공유 산출물이라 서버
            # 역할과 무관하게 항상 만든다. RAM 적재만 조회를 서빙하는 운영 서버
            # 전용이다 — 예전엔 root RAM 이 비활성이면 이 단계 전체를 건너뛰어,
            # 개발(worker) 서버에서 수동 스캔을 돌려도 랏캐시가 만들어지지 않았다.
            lookup = _ensure_scan_lookup_caches(product, job_id=job_id)
            results["lookup_cache"] = lookup
            lookup_ok = bool(lookup.get("ok"))
            # 통합 스캔의 기본 목적은 공유 디스크 lookup 갱신이다. pivot이 먼저
            # 읽히는 현재 검색 경로에서 root RAM 자동 예열은 중복이므로 명시적으로
            # scheduler를 opt-in한 배포에서만 함께 적재한다. 개별 수동 적재 API는
            # 그대로 유지해 캐시관리/진단 계약을 깨지 않는다.
            try:
                from core.runtime_limits import splittable_root_lot_ram_cache_scheduler_enabled
                ram_auto_enabled = bool(splittable_root_lot_ram_cache_scheduler_enabled())
            except Exception:
                ram_auto_enabled = False
            ram_disabled = (
                not _ml_table_lookup.root_ram_cache_available() or not ram_auto_enabled
            )
            if ram_disabled:
                results["root_lot_cache"] = {
                    "ok": lookup_ok, "skipped": True, "reason": "ram_disabled",
                    "products": [], "warmed_products": 0, "build_pending": 0,
                    "lookup": lookup,
                }
            else:
                results["root_lot_cache"] = _ml_table_lookup.refresh_root_lot_ram_cache(
                    product=product, force=force, load_now=True)
                results["root_lot_cache"] = _wait_for_root_lookup_caches(
                    results["root_lot_cache"], product=product, job_id=job_id)
            rc = results["root_lot_cache"]
            pending = int(rc.get("build_pending") or 0)
            warmed = int(rc.get("warmed_products") or 0)
            # 원본 lookup 캐시가 아직 빌드 중이면 예열은 '대기'다(실패 아님) — 빌드 완료
            # 후 자동 적재된다. 대기 개수를 표시해 조용한 미적재로 오인하지 않게 한다.
            note = (f" · 빌드 대기 {pending}개(원본 lookup 캐시 빌드 완료 후 자동 적재)"
                    if pending else "")
            lookup_note = ("랏캐시 최신(빌드 불필요)" if lookup.get("skipped")
                           else f"랏캐시 {lookup.get('built') or 0}/{lookup.get('total') or 0} 빌드")
            rc_ok = lookup_ok and (ram_disabled or bool(rc.get("ok")))
            if not lookup_ok:
                phase_label = f"실패: {lookup.get('detail') or _err_text(lookup)}"
            elif ram_disabled:
                phase_label = (f"{lookup_note} 완료 · RAM 적재는 건너뜀"
                               "(이 서버 역할/설정에서 꺼짐 — 조회를 서빙하는 운영 서버 전용)")
            elif rc_ok:
                phase_label = f"{lookup_note} 완료 · {warmed}개 제품 RAM 예열{note}"
            else:
                phase_label = f"RAM 예열 실패: {_err_text(rc)}"
            stage_outcomes["3/3 랏캐시"] = "완료" if rc_ok else "실패"
            if job_id:
                from core.cache_event_log import stage_finished
                stage_finished(job_id, "root_lot_ram", ok=rc_ok,
                               detail={"products": len(rc.get("products") or []),
                                       "max_gb": rc.get("max_gb"),
                                       "ram_disabled": ram_disabled,
                                       "lookup_built": lookup.get("built") or 0,
                                       "lookup_total": lookup.get("total") or 0,
                                       "build_pending": pending, "warmed": warmed,
                                       "error": "" if rc_ok else
                                       (lookup.get("detail") or _err_text(rc))})
            _log("scan", f"[수동 스캔] 3/3 랏(lookup) 캐시 {phase_label} ({_label})",
                 ok=rc_ok,
                 detail={"products": len(rc.get("products") or []),
                         "max_gb": rc.get("max_gb"), "ram_disabled": ram_disabled,
                         "lookup_built": lookup.get("built") or 0,
                         "build_pending": pending, "warmed": warmed},
                 stage="root_lot_ram", phase="finished" if rc_ok else "failed")
        except Exception as e:
            results["root_lot_cache"] = {"ok": False, "error": str(e)}
            stage_outcomes["3/3 랏캐시"] = "실패"
            if job_id:
                from core.cache_event_log import stage_finished
                stage_finished(job_id, "root_lot_ram", ok=False, detail={"error": str(e)})
            _log("scan", f"[수동 스캔] 3/3 랏(lookup) 캐시 실패 ({_label}): {e}", ok=False,
                 stage="root_lot_ram", phase="failed")

        # 전체 성공 판정 — 예전엔 any() 라 3개 중 2개가 실패해도 '완료(ok=True)'로
        # 보였다. 하나라도 실패하면 실패로 본다(건너뜀은 실패가 아니다).
        failed = [name for name, outcome in stage_outcomes.items() if outcome == "실패"]
        results["ok"] = not failed
        results["stages"] = dict(stage_outcomes)
        summary = " · ".join(f"{name} {outcome}" for name, outcome in stage_outcomes.items())
        _log("scan",
             (f"[수동 스캔] 전체 완료 ({_label}) — {summary}" if not failed
              else f"[수동 스캔] 실패 ({_label}) — {summary} · 실패 단계: {', '.join(failed)}"),
             ok=results["ok"], detail={"stages": dict(stage_outcomes)})
    finally:
        if owns_job:
            if job_id:
                try:
                    from core.cache_event_log import finish_job
                    finish_job(job_id, ok=bool(results.get("ok")),
                               detail={"product": product, "stages": results.get("stages") or {}})
                except Exception:
                    logger.debug("unified scan job tracker finish failed", exc_info=True)
            with _UNIFIED_SCAN_LOCK:
                _UNIFIED_SCAN_BUSY = False
    return results


_PRODUCT_CACHE_PIPELINE_STAGES = [
    ("lookup_build", "root별 ML_TABLE lookup"),
    ("pivot_build", "root별 SplitTable pivot"),
    ("latest_lot", "WIP latest-lot"),
    ("fab_index", "root별 FAB latest 인덱스"),
]


def _submit_product_cache_scan(product: str, *, force: bool, source: str,
                               on_started=None, on_finished=None,
                               local_only: bool = False) -> dict:
    """Queue one product's four cache stages as a single serial pipeline.

    Manual and scheduled work share this entry point so a server never starts
    independent per-cache schedules.  The scan gate serializes pipelines and
    the pipeline itself waits for each cache kind before advancing to the next.

    local_only=True 는 관리자가 누른 수동 캐싱 전용이다 — 네 단계를 이 서버에서
    직접 돌린다. 자동(스케줄러) 캐싱은 종전대로 개발 워커로 오프로드할 수 있다.
    """
    product = str(product or "").strip()
    _reap_dead_unified_scan()
    scheduled = str(source or "").strip().lower() == "scheduler"
    label = (
        f"자동 제품 캐싱 ({product})"
        if scheduled else f"SplitTable 통합 캐시 스캔 ({product or '전체 제품'})"
    )

    def _start() -> dict:
        global _UNIFIED_SCAN_BUSY, _UNIFIED_SCAN_JOB_ID, _UNIFIED_SCAN_THREAD
        from core.cache_event_log import start_job
        job_id = start_job(
            "scheduled_product_cache" if scheduled else "unified_scan",
            label,
            list(_PRODUCT_CACHE_PIPELINE_STAGES),
            product=product,
        )
        with _UNIFIED_SCAN_LOCK:
            _UNIFIED_SCAN_BUSY = True
        _UNIFIED_SCAN_JOB_ID = job_id
        _UNIFIED_SCAN_THREAD = threading.current_thread()
        if on_started:
            try:
                on_started(product, job_id)
            except Exception:
                logger.debug("product cache on_started callback failed", exc_info=True)
        result: dict = {"ok": False, "error": "pipeline_not_started"}
        try:
            result = _run_unified_scan(product, force, job_id, local_only=local_only)
            return result
        finally:
            # _run_unified_scan 도 finally 에서 풀지만, 여기서 한 번 더 확실히 푼다 —
            # busy 가 남으면 scan-status 가 영원히 running 이라 화면이 안 끝난다.
            with _UNIFIED_SCAN_LOCK:
                _UNIFIED_SCAN_BUSY = False
            if on_finished:
                try:
                    on_finished(product, result)
                except Exception:
                    logger.debug("product cache on_finished callback failed", exc_info=True)

    out = _submit_scan(
        "scheduled_product_cache" if scheduled else "unified_scan",
        label,
        _start,
        product=product,
        source=source or "manual",
        dedupe_key=f"product_cache_pipeline:{product}",
    )
    return {**out, "product": product, "label": label}


@router.post("/ram-cache/unified-scan")
def unified_scan(req: UnifiedScanReq, _perm=Depends(require_page_manager("splittable"))):
    """관리자: 선택 제품의 필수 공유 캐시 작업을 1회 일반 큐에 등록.

    스캔은 서버당 하나만 돈다. 다른 스캔(수동/전체 셋업/예약)이 진행 중이면
    거절하지 않고 대기열에 넣고, 앞 작업이 끝나면(실패해도) 이어서 실행한다.

    작업(job) 은 **큐에서 꺼내 실제로 시작하는 순간** 만든다 — 대기 중에 미리
    만들면 heartbeat 이 없어 작업 추적기가 정지로 보고 실패 처리한다."""
    product = str(req.product or "").strip()
    force = bool(req.force)
    # 수동 캐싱은 누른 서버에서 바로 돈다 — 개발 워커로 넘기지 않는다. 진행률과
    # 중단 버튼이 모두 이 서버의 scan gate 를 보고 있으므로, 실작업만 다른 서버로
    # 가면 화면은 "등록됨" 인데 아무 진행도 안 보이고 중단도 먹지 않는다.
    out = _submit_product_cache_scan(product, force=force, source="manual",
                                     local_only=True)
    return {
        **out,
        "product": product,
        "detail": (f"수동 캐싱 — 이 서버에서 제품별 4단계 캐시를 순서대로 처리합니다. "
                   f"({product or '전체 제품'}) " + str(out.get("detail") or "")).strip(),
    }


# ── 전체 셋업 (초기 1회, 운영 로컬·고자원 빠른 캐싱) ───────────────────────
#   개발 워커로 오프로드하지 않고 운영 서버에서 직접, 큰 메모리로 전 제품
#   캐시(랏 lookup → 매칭 → 제품RAM → 예열)를 한 번에 빌드한다. 관리자 전용.
#
#   workers 기본값은 1 이다 (2026-07-28). "한 서버에서 캐싱 작업은 한 번에 한 제품
#   한 가지"라는 운영 규칙에 이 버튼만 예외로 남겨 두면, 전체 셋업 중에는 여전히
#   여러 제품의 랏캐시가 동시에 돌아 RSS 피크가 겹친다. 초기 반입처럼 속도가
#   우선일 때만 FLOW_FULL_SETUP_WORKERS 로 올린다(예: 5).
#
#   **청크·배치 크기도 기본으로 건드리지 않는다 (2026-07-31).** 예전에는 속도를
#   위해 랏캐시 청크를 4→32, 매칭 배치를 300→2000 으로 키워서 넘겼다. 그런데 이
#   두 값이 곧 peak RAM 이다 — 랏캐시 peak ≈ "청크 개수만큼의 root wide 데이터",
#   매칭 peak ≈ "배치 1개 크기". 8배·6.7배로 키우면 peak 도 그만큼 커진다.
#   메모리 가드는 **청크/배치 사이에서만** 확인하므로 한 단위 안에서 벌어지는
#   초과는 못 막는다 — 그래서 28GB 호스트에서 상한을 넘겨 캐시가 축출됐다.
#   전체 셋업은 "제품별로 순차 전체 캐싱"이면 충분하다는 게 요구사항이므로,
#   평소와 같은 메모리 프로파일로 돌리고 속도가 필요할 때만 env 로 올린다.

def _full_setup_config() -> dict:
    """전체 셋업 실행 파라미터. 0 = '평소 설정을 그대로 쓴다(오버라이드 없음)'."""
    return {
        # 한 서버에서는 제품/캐시 파이프라인을 반드시 하나씩 실행한다. 과거
        # FLOW_FULL_SETUP_WORKERS 값이 남아 있어도 전체 셋업만 병렬로 새는 예외를
        # 만들지 않는다.
        "workers": 1,
        "memory_gb": _float_env_clamped("FLOW_FULL_SETUP_MEMORY_GB", 20.0, 2.0, 512.0),
        "lookup_chunk": int(_float_env_clamped("FLOW_FULL_SETUP_LOOKUP_CHUNK", 0.0, 0.0, 500.0)),
        "match_batch": int(_float_env_clamped("FLOW_FULL_SETUP_MATCH_BATCH_ROOTS", 0.0, 0.0, 100000.0)),
    }


def _full_setup_env_apply(cfg: dict) -> dict:
    """전체 셋업 동안만 적용할 env 오버라이드. 원래 값 dict 반환(복원용)."""
    overrides = {
        "FLOW_WORKER_OFFLOAD": "0",                                  # 개발 오프로드 금지 → 운영 로컬 처리
        "FLOW_PROCESS_MEMORY_LIMIT_GB": str(cfg["memory_gb"]),       # 메모리 가드 상한 고정
    }
    # 청크/배치는 명시 지정(>0)일 때만 덮어쓴다. env 가 캐시관리 ⚙ 설정보다
    # 우선순위가 높으므로, 무조건 넣으면 관리자가 ⚙ 에서 낮춰둔 값이 전체 셋업
    # 동안 조용히 무시된다 — 메모리를 지키려고 낮춘 설정이 정작 가장 무거운
    # 작업에서만 안 먹는 셈이었다.
    if int(cfg.get("lookup_chunk") or 0) > 0:
        overrides["FLOW_ML_TABLE_LOOKUP_CACHE_BUILD_CHUNK_SIZE"] = str(cfg["lookup_chunk"])
    if int(cfg.get("match_batch") or 0) > 0:
        overrides["FLOW_MATCH_CACHE_STREAM_BATCH_ROOTS"] = str(cfg["match_batch"])
    saved = {k: os.environ.get(k) for k in overrides}
    os.environ.update({k: str(v) for k, v in overrides.items()})
    return saved


def _full_setup_env_restore(saved: dict) -> None:
    for k, old in (saved or {}).items():
        if old is None:
            os.environ.pop(k, None)
        else:
            os.environ[k] = old


def _proc_rss_gb() -> float:
    try:
        from core.runtime_limits import process_memory_snapshot
        snap = process_memory_snapshot()
        return round(float(snap.get("process_rss_gb")
                           or snap.get("process_memory_effective_gb") or 0.0), 2)
    except Exception:
        return 0.0


def _fmt_dur_ko(sec: float) -> str:
    sec = int(max(0, sec))
    if sec < 60:
        return f"{sec}초"
    if sec < 3600:
        return f"{sec // 60}분 {sec % 60}초"
    return f"{sec // 3600}시간 {(sec % 3600) // 60}분"


def _full_setup_build_lookups_parallel(cfg: dict, job_id: str = "") -> dict:
    """전 제품의 랏(lookup) 캐시를 운영 로컬에서 N개 병렬로 직접 빌드.

    큐/오프로드/로컬-heavy-게이트(Semaphore 1)를 우회하고 build_lookup_cache 를
    직접 호출한다(파일별 빌드락으로 동일 파일 중복만 방지, 다른 제품끼리는 병렬).
    각 빌드는 청크 스트리밍이라 제품당 메모리가 제한된다."""
    import concurrent.futures as _cf
    from core.cache_event_log import record as _rec, heartbeat as _beat
    files = _ml_table_lookup._discover_ml_table_files() or []
    total = len(files)
    ok_count = 0
    failed_products: list[str] = []
    if not total:
        _rec("cache_op", "[전체셋업] 랏캐시 빌드 대상 제품이 없습니다 (ML_TABLE_*.parquet 미검출)",
             ok=False, product="", detail={"job_id": job_id} if job_id else None)
        return {"ok": False, "total": 0, "built": 0, "detail": "빌드 대상 제품 없음"}
    _rec("cache_op",
         f"[전체셋업] 랏캐시 빌드 시작 — {total:,}개 제품 · "
         f"{'순차(제품 1개씩)' if cfg['workers'] <= 1 else str(cfg['workers']) + '병렬'}(운영 로컬)"
         f" · 청크 {_ml_table_lookup._lookup_cache_build_chunk_size()}"
         f" · 메모리 상한 {cfg['memory_gb']}GB"
         f" · 이미 빌드된 제품은 건너뜀(재실행 시 이어서 진행)",
         product="")
    done = [0]
    lock = threading.Lock()
    started = time.time()

    canceled = [False]

    def _one(fp) -> bool:
        prod = Path(fp).stem
        # 제품 경계가 안전한 취소 지점이다 — 지금까지 만든 제품 캐시는 그대로
        # 남고, 남은 제품은 손대지 않은 채 다음 큐 작업으로 넘어간다.
        if _scan_cancel_requested():
            if not canceled[0]:
                canceled[0] = True
                _rec("cache_op",
                     f"[전체셋업] 중단 요청 — 남은 제품({total - done[0]:,}개)을 건너뜁니다. "
                     f"이미 빌드된 제품은 그대로 남습니다(다시 실행하면 이어서 진행).",
                     ok=False, product="", detail={"job_id": job_id} if job_id else None)
            return False
        okp = True
        err = ""
        try:
            # force=False: 이미 빌드된(fresh) 캐시는 건너뛴다 → 전체 셋업이 중간에
            # 끊겨 재실행해도 처음부터 다시 하지 않고 남은 것만 이어서 빌드(수렴 보장).
            _ml_table_lookup.build_lookup_cache(Path(fp), force=False)
        except Exception as exc:
            okp = False
            err = f"{type(exc).__name__}: {exc}"
            logger.warning("full setup lookup build failed source=%s: %s", fp, exc)
        with lock:
            done[0] += 1
            n = done[0]
            if not okp:
                failed_products.append(prod)
        elapsed = time.time() - started
        eta = (elapsed / n) * (total - n) if n else 0.0
        try:
            _beat(job_id, f"랏캐시 빌드 {n}/{total}")
            _rec("cache_op",
                 f"[전체셋업] 랏캐시 {n:,}/{total:,} — {prod}"
                 + ("" if okp else f" · 빌드 실패: {err}")
                 + f" · RSS {_proc_rss_gb()}GB · 남은 ~{_fmt_dur_ko(eta)}",
                 ok=okp, product=prod, detail={"job_id": job_id} if job_id else None)
        except Exception:
            pass
        return okp

    with _cf.ThreadPoolExecutor(max_workers=int(cfg["workers"]), thread_name_prefix="full-setup") as ex:
        for res in ex.map(_one, files):
            if res:
                ok_count += 1
    all_ok = ok_count == total
    _rec("cache_op",
         (f"[전체셋업] 랏캐시 완료 — {ok_count:,}/{total:,} 성공 · 총 {_fmt_dur_ko(time.time() - started)}"
          if all_ok else
          f"[전체셋업] 랏캐시 일부 실패 — {ok_count:,}/{total:,} 성공, {len(failed_products)}개 실패: "
          + ", ".join(sorted(failed_products)[:5]) + ("…" if len(failed_products) > 5 else "")),
         ok=all_ok, product="", detail={"job_id": job_id} if job_id else None)
    return {"ok": ok_count > 0, "all_ok": all_ok, "total": total, "built": ok_count,
            "failed": sorted(failed_products),
            "detail": "" if all_ok else f"{len(failed_products)}개 제품 랏캐시 빌드 실패"}


def _run_full_setup_scan(job_id: str = "") -> dict:
    global _UNIFIED_SCAN_BUSY
    cfg = _full_setup_config()
    saved = _full_setup_env_apply(cfg)
    try:
        from core import cache_budget
        cache_budget.invalidate()
    except Exception:
        pass
    from core.cache_event_log import record as _rec
    started = time.time()
    ok = True
    try:
        _rec("cache_op",
             f"[전체셋업] 시작 — 운영 로컬 · "
             f"{'순차(제품 1개씩)' if cfg['workers'] <= 1 else str(cfg['workers']) + '병렬'}"
             f" · 메모리 {cfg['memory_gb']}GB"
             f" · 랏캐시 청크 {_ml_table_lookup._lookup_cache_build_chunk_size()}"
             f" · 매칭 배치 {_match_cache_stream_batch_roots()} root"
             f" (랏캐시 → 매칭 → 제품RAM → 예열)",
             product="", detail={"job_id": job_id} if job_id else None)
        # Phase A: 랏(lookup) 캐시 전 제품 로컬 빌드 (가장 무겁다).
        #   기본은 제품 1개씩 순차 — FLOW_FULL_SETUP_WORKERS 로 병렬도를 올린다.
        if job_id:
            from core.cache_event_log import stage_started, stage_finished
            stage_started(job_id, "lookup_build")
        phase_a = _full_setup_build_lookups_parallel(cfg, job_id)
        if job_id:
            from core.cache_event_log import stage_finished
            stage_finished(job_id, "lookup_build", ok=bool(phase_a.get("all_ok")),
                           detail={"built": phase_a.get("built"), "total": phase_a.get("total"),
                                   "error": phase_a.get("detail") or ""})
        # Phase B: 매칭 캐시 → 제품 원본 RAM → root 예열 (기존 통합 스캔 재사용).
        #   랏캐시가 이미 빌드돼 있어 예열이 skip 없이 즉시 적재된다.
        #   _run_unified_scan 이 finally 에서 job 종료 + busy 해제까지 처리한다.
        #   force=False: 이미 최신인 매칭/제품RAM/예열은 건너뛰어 재실행 시 이어서 진행.
        phase_b = _run_unified_scan("", False, job_id, owns_job=False)
        ok = bool(phase_b.get("ok")) and bool(phase_a.get("all_ok"))
        _rec("cache_op",
             (f"[전체셋업] 전체 완료 · 총 {_fmt_dur_ko(time.time() - started)} · RSS {_proc_rss_gb()}GB"
              if ok else
              f"[전체셋업] 완료(일부 실패) · 랏캐시 {phase_a.get('built')}/{phase_a.get('total')}"
              f" · 후속 단계: {phase_b.get('stages') or {}}"
              f" · 총 {_fmt_dur_ko(time.time() - started)}"),
             ok=ok, product="")
    except Exception as exc:
        ok = False
        logger.warning("full setup scan failed: %s", exc, exc_info=True)
        try:
            _rec("cache_op", f"[전체셋업] 실패: {exc}", ok=False, product="")
            if job_id:
                from core.cache_event_log import finish_job
                finish_job(job_id, ok=False, detail={"error": str(exc)})
        except Exception:
            pass
    finally:
        _full_setup_env_restore(saved)
        try:
            from core import cache_budget
            cache_budget.invalidate()
        except Exception:
            pass
        if job_id:
            try:
                from core.cache_event_log import finish_job
                finish_job(job_id, ok=ok, detail={"kind": "full_setup"})
            except Exception:
                logger.debug("full setup job tracker finish failed", exc_info=True)
        with _UNIFIED_SCAN_LOCK:
            _UNIFIED_SCAN_BUSY = False
    return {"ok": ok}


@router.post("/ram-cache/full-setup")
def full_setup_scan(_admin=Depends(require_admin)):
    """호환용 버튼: 전 제품 필수 공유 캐시를 일반 큐에 한 번 등록한다."""
    _reap_dead_unified_scan()
    cfg = _full_setup_config()
    label = "전체 셋업 (운영 로컬 · 빠른 캐싱)"

    def _start() -> dict:
        global _UNIFIED_SCAN_BUSY, _UNIFIED_SCAN_JOB_ID, _UNIFIED_SCAN_THREAD
        from core.cache_event_log import start_job
        job_id = start_job(
            "full_setup",
            label,
            [
                ("lookup_build", "root별 ML_TABLE lookup"),
                ("pivot_build", "root별 SplitTable pivot"),
                ("latest_lot", "WIP latest-lot"),
                ("fab_index", "root별 FAB latest 인덱스"),
            ],
            product="",
        )
        with _UNIFIED_SCAN_LOCK:
            _UNIFIED_SCAN_BUSY = True
        _UNIFIED_SCAN_JOB_ID = job_id
        _UNIFIED_SCAN_THREAD = threading.current_thread()
        try:
            return _enqueue_required_split_caches("", True, job_id)
        finally:
            with _UNIFIED_SCAN_LOCK:
                _UNIFIED_SCAN_BUSY = False

    out = _submit_scan("full_setup", label, _start, dedupe_key="full_setup")
    return {
        **out,
        "config": cfg,
        "effective": {
            "lookup_chunk": _ml_table_lookup._lookup_cache_build_chunk_size(),
            "match_batch_roots": _match_cache_stream_batch_roots(),
        },
        "detail": ("전체 제품 필수 디스크 캐시를 일반 작업 큐에 1회 등록합니다. "
                   "제품/Root RAM 예열과 구형 FAB 매칭 캐시는 실행하지 않습니다. "
                   "개발 워커가 없으면 운영 서버가 같은 작업을 이어받습니다. "
                   + str(out.get("detail") or "")).strip(),
    }


def _scan_gate_snapshot() -> dict:
    """서버 스캔 게이트 현황(실행 1건 + 대기열). 게이트가 없으면 빈 스냅샷."""
    try:
        from core import scan_gate
        return scan_gate.snapshot()
    except Exception:
        logger.debug("scan gate snapshot failed", exc_info=True)
        return {"running": False, "busy": False, "current": None, "pending": [],
                "depth": 0, "last": None, "worker_alive": False}


def _cache_queue_snapshot() -> dict:
    """관리 화면에 노출할 캐시 관련 실행/대기 큐의 작은 스냅샷."""
    try:
        from core import worker_dispatch
        worker = worker_dispatch.queue_snapshot(limit=50)
    except Exception:
        worker = {"depth": 0, "queued": [], "running": []}
    try:
        lookup_raw = _ml_table_lookup.build_queue_snapshot()
        lookup = {
            "running": bool(lookup_raw.get("running")),
            "current": Path(str(lookup_raw.get("current") or "")).name,
            "queued": [Path(str(value)).name for value in (lookup_raw.get("queued") or [])],
            "last_error": str(lookup_raw.get("last_error") or ""),
        }
    except Exception:
        lookup = {"running": False, "current": "", "queued": [], "last_error": ""}
    try:
        root_prefetch = _ml_table_lookup.root_ram_prefetch_snapshot(limit=50)
    except Exception:
        root_prefetch = {"running": False, "depth": 0, "queued": []}
    match = _match_cache_job_status()
    product = _product_ram_cache_job_status()
    try:
        root_status = _ml_table_lookup.root_ram_cache_status(include_detail=False)
    except Exception:
        root_status = {}
    auto_schedule = _auto_product_cache_schedule_snapshot()
    return {
        "worker": worker,
        "lookup_build": lookup,
        "root_prefetch": root_prefetch,
        "match_cache": {
            "running": bool(match.get("running")),
            "current": str(match.get("current_product") or ""),
            "order": list(match.get("order") or []),
            "done": int(match.get("done") or 0),
            "total": int(match.get("total") or 0),
            # 관리자 중단 버튼 노출/표시용 — 요청이 걸린 제품과 남은 중단 건수.
            "cancel_product": str((match.get("cancel") or {}).get("product") or ""),
            "cancelled_count": int(match.get("cancelled_count") or 0),
            "remaining": max(0, int(match.get("total") or 0) - int(match.get("done") or 0)),
            "queued": max(
                0,
                int(match.get("total") or 0)
                - int(match.get("done") or 0)
                - (1 if match.get("running") and match.get("current_product") else 0),
            ),
        },
        "product_ram": {
            "running": bool(product.get("running")),
            "current": str(product.get("current_product") or ""),
            "order": list(product.get("order") or []),
            "done": int(product.get("done") or 0),
            "total": int(product.get("total") or 0),
            "remaining": max(0, int(product.get("total") or 0) - int(product.get("done") or 0)),
            "queued": max(
                0,
                int(product.get("total") or 0)
                - int(product.get("done") or 0)
                - (1 if product.get("running") and product.get("current_product") else 0),
            ),
        },
        "root_lot_ram": {
            "running": bool(root_status.get("running")),
            "current": str(root_status.get("current_product") or ""),
            "order": list(root_status.get("order") or []),
            "done": int(root_status.get("done") or 0),
            "total": len(root_status.get("order") or []),
        },
        # One schedule owns all four required cache kinds.  Each tick processes
        # one product serially, then advances the product cursor.
        "schedules": [auto_schedule],
        "auto_product_cache": auto_schedule,
    }


_PRODUCT_CACHE_STATUS_LOCK = threading.Lock()
_PRODUCT_CACHE_STATUS_SNAPSHOT: dict = {"at": 0.0, "value": None, "refreshing": False}


def _product_cache_status_snapshot(*, nonblocking: bool = False) -> dict:
    """Recent lifecycle plus real four-artifact readiness for every product.

    Directory fingerprints are more expensive than the event reducer, so the
    admin page's frequent poll reuses a short snapshot.  Ten seconds is still
    immediate for operations while avoiding repeated directory walks.
    """
    from core.cache_event_log import product_status
    now = time.monotonic()
    with _PRODUCT_CACHE_STATUS_LOCK:
        cached = _PRODUCT_CACHE_STATUS_SNAPSHOT.get("value")
        if cached is not None and now - float(_PRODUCT_CACHE_STATUS_SNAPSHOT.get("at") or 0.0) < 10.0:
            return cached
        refreshing = bool(_PRODUCT_CACHE_STATUS_SNAPSHOT.get("refreshing"))
        if nonblocking and not refreshing:
            _PRODUCT_CACHE_STATUS_SNAPSHOT["refreshing"] = True

    if nonblocking:
        if not refreshing:
            def _refresh_status() -> None:
                try:
                    _product_cache_status_snapshot(nonblocking=False)
                finally:
                    with _PRODUCT_CACHE_STATUS_LOCK:
                        _PRODUCT_CACHE_STATUS_SNAPSHOT["refreshing"] = False

            threading.Thread(
                target=_refresh_status,
                name="cache-product-status-refresh",
                daemon=True,
            ).start()
        if cached is not None:
            return cached
        # Cold page load must not wait for every product's lookup/pivot/FAB
        # directory walk.  Return the cheap lifecycle projection immediately;
        # the background refresh publishes artifact-accurate readiness for the
        # next poll.
        status = product_status(products=_match_cache_products(""))
        status["artifact_status_pending"] = True
        return status
    status = product_status(products=_match_cache_products(""))
    for row in status.get("products") or []:
        try:
            actual = _required_split_cache_status(row.get("product") or "")
        except Exception:
            continue
        by_kind = {item.get("kind"): item for item in (actual.get("kinds") or [])}
        for item in row.get("kinds") or []:
            artifact = by_kind.get(item.get("kind")) or {}
            lifecycle_state = str(item.get("state") or "")
            item["ready"] = bool(artifact.get("ready"))
            item["refreshing"] = bool(item["ready"] and lifecycle_state == "running")
            item["artifact_state"] = artifact.get("state") or "missing"
            item["artifact_done"] = int(artifact.get("done") or 0)
            item["built_ts"] = float(artifact.get("built_ts") or 0.0)
            # 산출물이 실제로 준비돼 있으면 그 캐시는 성공이다. 이벤트 로그는
            # 7일 창이고, fresh 스킵·워커 오프로드로 완료 줄 자체가 안 남는
            # 경로가 있어서 "4개 다 성공인데 3/4 준비"처럼 요약과 상세가
            # 어긋났다. 준비 여부는 산출물, 시각/소요는 로그 → 산출물 순.
            if item["ready"]:
                # A rebuild is a new unpublished generation.  Keep serving and
                # reporting the previous published artifact until the builder
                # atomically promotes its replacement; progress is exposed via
                # `refreshing`/the progress panel instead of downgrading 4/4.
                if item.get("state") != "ok":
                    item["state"] = "ok"
                built_ts = item["built_ts"]
                if built_ts and built_ts > float(item.get("success_ts") or 0.0):
                    item["success_ts"] = built_ts
                    item["last_ts"] = max(float(item.get("last_ts") or 0.0), built_ts)
                    build_seconds = float(artifact.get("build_seconds") or 0.0)
                    started_ts = float(item.get("started_ts") or 0.0)
                    if build_seconds > 0:
                        item["duration_sec"] = round(build_seconds, 1)
                    elif started_ts and built_ts >= started_ts:
                        item["duration_sec"] = round(built_ts - started_ts, 1)
                artifact_message = str(artifact.get("message") or "")
                if artifact_message and (
                    not item.get("message")
                    or built_ts >= float(item.get("success_ts") or 0.0)
                ):
                    item["message"] = artifact_message
                elif not item.get("message"):
                    item["message"] = "실제 캐시 산출물 준비됨"
            elif not item["ready"] and item.get("state") == "ok":
                # 반대 방향의 어긋남: 빌드 로그는 성공인데 산출물이 없다.
                item["state"] = "stale"
                item["message"] = "빌드 기록은 성공이지만 산출물이 준비되지 않았습니다"
        row["ready_count"] = int(actual.get("ready_count") or 0)
        row["total"] = int(actual.get("total") or 4)
        # 개별 kind 상태를 산출물 기준으로 고쳤으니 제품 요약도 다시 만든다.
        kind_rows = row.get("kinds") or []
        successes = [float(k.get("success_ts") or 0.0) for k in kind_rows if k.get("success_ts")]
        row["last_success_ts"] = max(successes) if successes else 0.0
        row["last_failure_ts"] = max(
            [float(k.get("failed_ts") or 0.0) for k in kind_rows if k.get("state") == "failed"] or [0.0])
        row["failed_kinds"] = [k.get("label") for k in kind_rows if k.get("state") == "failed"]
        row["running_kinds"] = [k.get("label") for k in kind_rows if k.get("state") == "running"]
        states = {item.get("state") for item in kind_rows}
        if "running" in states:
            row["state"] = "running"
        elif "failed" in states:
            row["state"] = "failed"
        elif row["ready_count"] == row["total"]:
            row["state"] = "ok"
        else:
            row["state"] = "partial"
        # '전체 성공 시각' = 마지막으로 완성된 캐시의 시각. 4종이 다 준비됐을
        # 때만 의미가 있다.
        row["all_success_ts"] = (
            max(successes) if successes and row["state"] == "ok"
            and len(successes) == len(kind_rows) else 0.0
        )
    status["ok_count"] = sum(1 for row in status.get("products") or [] if row.get("state") == "ok")
    status["failed_count"] = sum(1 for row in status.get("products") or [] if row.get("state") == "failed")
    status["running_count"] = sum(1 for row in status.get("products") or [] if row.get("state") == "running")
    with _PRODUCT_CACHE_STATUS_LOCK:
        _PRODUCT_CACHE_STATUS_SNAPSHOT.update({"at": now, "value": status})
    return status


@router.get("/ram-cache/scan-status")
def unified_scan_status():
    """통합 수동 스캔 진행 여부 — 프런트가 진행 로그를 폴링할 동안 종료 감지용.

    반환 전 죽은 스레드/정지 작업을 정리한다 — 화면이 끝나지 않는 '스캔 중'에
    갇히지 않도록, running=False 는 반드시 언젠가 온다."""
    global _UNIFIED_SCAN_BUSY
    reaped = _reap_dead_unified_scan()
    from core.cache_event_log import get_jobs
    jobs = get_jobs(recent=1)          # 내부에서 정지 작업(stale)을 실패로 확정
    # 작업 추적기가 정지로 판정했으면 busy 플래그도 함께 푼다.
    if not reaped and any(job.get("id") == _UNIFIED_SCAN_JOB_ID and job.get("status") != "running"
                          for job in jobs):
        with _UNIFIED_SCAN_LOCK:
            if _UNIFIED_SCAN_BUSY and not (_UNIFIED_SCAN_THREAD and _UNIFIED_SCAN_THREAD.is_alive()):
                _UNIFIED_SCAN_BUSY = False
    with _UNIFIED_SCAN_LOCK:
        unified_running = bool(_UNIFIED_SCAN_BUSY)
    # 게이트에 대기 중인 작업까지 running 으로 본다 — 대기 중에 running=False 를
    # 주면 프런트가 폴링을 멈추고 "완료"로 표시해 버린다(실제로는 곧 시작한다).
    gate = _scan_gate_snapshot()
    last = next((job for job in jobs if job.get("id") == _UNIFIED_SCAN_JOB_ID), None)
    return {
        "ok": True,
        "running": unified_running or bool(gate.get("busy")),
        "scan_queue": gate,
        "job_id": _UNIFIED_SCAN_JOB_ID,
        # 마지막 작업의 최종 상태 — 프런트가 '완료/실패'를 명확히 표시하는 데 쓴다.
        "last_status": str((last or {}).get("status") or ""),
        "last_stages": [
            {"id": s.get("id"), "label": s.get("label"), "status": s.get("status"),
             "error": str((s.get("detail") or {}).get("error") or "")}
            for s in ((last or {}).get("stages") or [])
        ],
        "jobs": jobs,
        "queues": _cache_queue_snapshot(),
    }


# ── 캐시 예산 조절 (톱니바퀴) ─────────────────────────────────────────────
def _cache_budget_settings_payload() -> dict:
    """저장된 예산 설정 + 실제 적용(effective) 값 + 호스트 정보."""
    from core import cache_settings, cache_budget
    saved = cache_settings.read()

    def _gb(n):
        return round(float(n or 0) / (1024 ** 3), 2)

    host_gb = 0.0
    try:
        from core.runtime_limits import system_memory_snapshot
        host_gb = round(float(system_memory_snapshot().get("system_memory_total_gb") or 0.0), 1)
    except Exception:
        pass
    env_pins = {
        "pool_fraction": "FLOW_CACHE_TOTAL_BUDGET_FRACTION" in os.environ,
        "dev_factor": "FLOW_DEV_CACHE_BUDGET_FACTOR" in os.environ,
        "root_ram_gb": "FLOW_SPLITTABLE_ROOT_LOT_RAM_CACHE_MAX_GB" in os.environ,
        "view_mb": "FLOW_SPLITTABLE_VIEW_CACHE_MAX_MB" in os.environ,
        "product_ram_gb": "FLOW_SPLITTABLE_PRODUCT_RAM_CACHE_MAX_GB" in os.environ,
        "product_ram_enabled": "FLOW_DISABLE_SPLITTABLE_PRODUCT_RAM_CACHE" in os.environ,
        "match_cache_batch_roots": "FLOW_MATCH_CACHE_STREAM_BATCH_ROOTS" in os.environ,
        "view_cold_concurrency": "FLOW_SPLITTABLE_VIEW_COLD_CONCURRENCY" in os.environ,
        "view_cold_concurrency_dev": True,
        "lookup_build_chunk_roots": "FLOW_ML_TABLE_LOOKUP_CACHE_BUILD_CHUNK_SIZE" in os.environ,
        "pivot_build_chunk_roots": "FLOW_PIVOT_CACHE_CHUNK_SIZE" in os.environ,
        "cache_speed_level": (
            "FLOW_ML_TABLE_LOOKUP_CACHE_BUILD_CHUNK_SIZE" in os.environ
            or "FLOW_PIVOT_CACHE_CHUNK_SIZE" in os.environ
        ),
        "auto_product_cache_enabled": "FLOW_SPLITTABLE_AUTO_PRODUCT_CACHE_ENABLED" in os.environ,
        "auto_product_cache_interval_minutes": "FLOW_SPLITTABLE_AUTO_PRODUCT_CACHE_INTERVAL_MINUTES" in os.environ,
    }
    return {
        "ok": True,
        "saved": saved,
        "is_dev": _ml_table_lookup._root_ram_cache_use_dev(),
        "host_total_gb": host_gb,
        "env_pins": env_pins,   # env 로 고정된 항목은 UI 편집 무시됨
        "effective": {
            "pool_fraction": round(cache_budget._pool_fraction(), 3),
            "dev_factor": round(cache_budget.worker_budget_factor(), 3),
            "pool_gb": _gb(cache_budget.pool_bytes()),
            "root_ram_gb": _gb(_ml_table_lookup._root_ram_cache_max_bytes()),
            "view_mb": round(_view_cache_max_bytes() / (1024 ** 2), 1),
            "product_ram_enabled": _product_ram_cache_available(),
            "product_ram_gb": _gb(_product_ram_cache_max_bytes()),
            "match_cache_batch_roots": _match_cache_stream_batch_roots(),
            "view_cold_concurrency": _view_cold_lane_concurrency(),
            "lookup_build_chunk_roots": _ml_table_lookup._lookup_cache_build_chunk_size(),
            "pivot_build_chunk_roots": _pivot_cache_builder.pivot_build_chunk_roots(),
            "cache_speed_level": cache_settings.cache_speed_level(
                _ml_table_lookup._root_ram_cache_use_dev()),
            "auto_product_cache_enabled": _auto_product_cache_enabled(),
            "auto_product_cache_interval_minutes": _auto_product_cache_interval_minutes(),
        },
        "defaults": {
            "pool_fraction": 0.45,
            "dev_factor": 0.35,
            "product_ram_enabled": False,
            "view_mb": None,
            "match_cache_batch_roots": MATCH_CACHE_STREAM_BATCH_ROOTS_DEFAULT,
            "view_cold_concurrency": 3,
            "view_cold_concurrency_dev": 1,
            "lookup_build_chunk_roots": _ml_table_lookup.lookup_cache_build_chunk_default(),
            "pivot_build_chunk_roots": _pivot_cache_builder.pivot_build_chunk_default(),
            "cache_speed_level": 1,
            "auto_product_cache_enabled": False,
            "auto_product_cache_interval_minutes": _AUTO_PRODUCT_CACHE_DEFAULT_INTERVAL_MINUTES,
        },
        "auto_schedule": _auto_product_cache_schedule_snapshot(),
        # 검색 동시성 레인 현황 — 톱니바퀴에서 현재 몇 개가 돌고 있는지 함께 보여준다.
        "cold_lane": _view_cold_lane_stats(),
    }


@router.get("/cache-budget/settings")
def get_cache_budget_settings(request: Request):
    if not is_page_manager(current_user(request), "splittable"):
        raise HTTPException(403, "관리자 전용")
    return _cache_budget_settings_payload()


class CacheBudgetSaveReq(BaseModel):
    pool_fraction: float | None = None
    pool_fraction_dev: float | None = None
    dev_factor: float | None = None
    root_ram_gb: float | None = None
    root_ram_gb_dev: float | None = None
    view_mb: float | None = None
    view_mb_dev: float | None = None
    product_ram_enabled: bool | None = None
    product_ram_enabled_dev: bool | None = None
    product_ram_gb: float | None = None
    product_ram_gb_dev: float | None = None
    match_cache_batch_roots: int | None = None
    match_cache_batch_roots_dev: int | None = None
    view_cold_concurrency: int | None = None
    view_cold_concurrency_dev: int | None = None
    lookup_build_chunk_roots: int | None = None
    lookup_build_chunk_roots_dev: int | None = None
    pivot_build_chunk_roots: int | None = None
    pivot_build_chunk_roots_dev: int | None = None
    cache_speed_level: int | None = None
    cache_speed_level_dev: int | None = None
    auto_product_cache_enabled: bool | None = None
    auto_product_cache_enabled_dev: bool | None = None
    auto_product_cache_interval_minutes: int | None = None
    auto_product_cache_interval_minutes_dev: int | None = None


@router.post("/cache-budget/settings/save")
def save_cache_budget_settings(req: CacheBudgetSaveReq, _perm=Depends(require_page_manager("splittable"))):
    """캐시 예산 조절값 저장. 빈 값(None)은 미변경, 0/음수 GB 는 '자동'(키 삭제).
    운영/개발 분리: <key> = 운영, <key>_dev = 개발서버 전용(미설정 시 운영값 폴백)."""
    from core import cache_settings
    partial: dict = {}
    fields = req.model_dump(exclude_unset=True)

    def _put_gb(key):
        if key in fields:
            v = fields[key]
            partial[key] = None if (v is None or float(v) <= 0) else max(0.1, min(64.0, float(v)))

    if "pool_fraction" in fields:
        partial["pool_fraction"] = None if fields["pool_fraction"] is None else max(0.1, min(0.8, float(fields["pool_fraction"])))
    if "pool_fraction_dev" in fields:
        v = fields["pool_fraction_dev"]
        partial["pool_fraction_dev"] = None if (v is None or v == "") else max(0.1, min(0.8, float(v)))
    if "dev_factor" in fields:
        partial["dev_factor"] = None if fields["dev_factor"] is None else max(0.05, min(1.0, float(fields["dev_factor"])))
    _put_gb("root_ram_gb"); _put_gb("root_ram_gb_dev")
    _put_gb("product_ram_gb"); _put_gb("product_ram_gb_dev")
    if "view_mb" in fields:
        v = fields["view_mb"]
        partial["view_mb"] = None if (v is None or float(v) <= 0) else max(64.0, min(8192.0, float(v)))
    if "view_mb_dev" in fields:
        v = fields["view_mb_dev"]
        partial["view_mb_dev"] = None if (v is None or float(v) <= 0) else max(64.0, min(8192.0, float(v)))

    def _put_int(key, lo, hi):
        if key in fields:
            v = fields[key]
            partial[key] = None if (v is None or v == "" or float(v) <= 0) else int(max(lo, min(hi, float(v))))
    _put_int("match_cache_batch_roots", 1, 100000)
    _put_int("match_cache_batch_roots_dev", 1, 100000)
    # 검색 cold 레인 슬롯 — 0/빈값이면 키 삭제(코어수 기반 자동값 복귀).
    _put_int("view_cold_concurrency", _VIEW_COLD_CONCURRENCY_MIN, _VIEW_COLD_CONCURRENCY_MAX)
    # 개발 서버 슬롯은 정책상 1로 고정한다. 예전 저장값도 다음 저장 때 정리한다.
    partial["view_cold_concurrency_dev"] = 1
    # 랏캐시 빌드 청크 — 0/빈값이면 키 삭제(역할별 기본값 복귀).
    _put_int("lookup_build_chunk_roots",
             _ml_table_lookup.LOOKUP_CACHE_BUILD_CHUNK_MIN, _ml_table_lookup.LOOKUP_CACHE_BUILD_CHUNK_MAX)
    _put_int("lookup_build_chunk_roots_dev",
             _ml_table_lookup.LOOKUP_CACHE_BUILD_CHUNK_MIN, _ml_table_lookup.LOOKUP_CACHE_BUILD_CHUNK_MAX)
    # pivot 빌드 청크 — 0/빈값이면 키 삭제(역할별 기본값 운영 3 / 개발 1 복귀).
    _put_int("pivot_build_chunk_roots",
             _pivot_cache_builder.PIVOT_BUILD_CHUNK_MIN, _pivot_cache_builder.PIVOT_BUILD_CHUNK_MAX)
    _put_int("pivot_build_chunk_roots_dev",
             _pivot_cache_builder.PIVOT_BUILD_CHUNK_MIN, _pivot_cache_builder.PIVOT_BUILD_CHUNK_MAX)
    _put_int("cache_speed_level", cache_settings.CACHE_SPEED_MIN, cache_settings.CACHE_SPEED_MAX)
    _put_int("cache_speed_level_dev", cache_settings.CACHE_SPEED_MIN, cache_settings.CACHE_SPEED_MAX)
    _put_int("auto_product_cache_interval_minutes", 1, 1440)
    _put_int("auto_product_cache_interval_minutes_dev", 1, 1440)
    if "product_ram_enabled" in fields:
        partial["product_ram_enabled"] = None if fields["product_ram_enabled"] is None else bool(fields["product_ram_enabled"])
    if "product_ram_enabled_dev" in fields:
        partial["product_ram_enabled_dev"] = None if fields["product_ram_enabled_dev"] is None else bool(fields["product_ram_enabled_dev"])
    if "auto_product_cache_enabled" in fields:
        partial["auto_product_cache_enabled"] = None if fields["auto_product_cache_enabled"] is None else bool(fields["auto_product_cache_enabled"])
    if "auto_product_cache_enabled_dev" in fields:
        partial["auto_product_cache_enabled_dev"] = None if fields["auto_product_cache_enabled_dev"] is None else bool(fields["auto_product_cache_enabled_dev"])
    cache_settings.save(partial)
    try:
        from core import cache_budget
        cache_budget.invalidate()   # 풀 메모 무효화 → 즉시 반영
    except Exception:
        pass
    if any(key.startswith("auto_product_cache_") for key in partial):
        # Recompute next product/time immediately instead of waiting for the old
        # interval to expire after an operator saves the schedule.
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            _AUTO_PRODUCT_CACHE_STATE.update({
                "next_product": "", "next_at": "", "next_at_ts": 0.0,
                "delayed": False, "delayed_reason": "",
            })
        _SEARCH_CACHE_MAINT_WAKE.set()
    return _cache_budget_settings_payload()


# ── RAM Cache 우선 Lot 등록 / 전체 목록 / 제품별 예산 ──────────────────
_RAM_CACHE_PRIORITY_FILE = PLAN_DIR / "ram_cache_priority_lots.json"


def _load_priority_lots_file() -> dict:
    """우선 lot 등록 JSON 전체를 읽는다. 파일 없으면 빈 dict."""
    try:
        return load_json(_RAM_CACHE_PRIORITY_FILE, {})
    except Exception:
        return {}


def _save_priority_lots_file(data: dict) -> None:
    save_json(_RAM_CACHE_PRIORITY_FILE, data)


@router.get("/ram-cache/priority-lots")
def get_ram_cache_priority_lots(product: str = Query("")):
    """우선 lot 등록 목록 조회."""
    product = str(product or "").strip()
    if not product:
        raise HTTPException(400, "product is required")
    data = _load_priority_lots_file()
    # 대소문자 무시 매칭
    product_upper = product.upper()
    lots = []
    for key, val in data.items():
        if str(key).strip().upper() == product_upper and isinstance(val, list):
            lots = val
            break
    # legacy note → comment 하위호환 (comment 빈 항목만)
    for entry in lots:
        if isinstance(entry, dict) and not entry.get("comment") and entry.get("note"):
            entry["comment"] = entry["note"]
    return {"ok": True, "product": product, "lots": lots}


class RamCachePriorityLotItem(BaseModel):
    lot_id: str
    purpose: str = ""
    note: str = ""      # legacy — comment 로 대체, 하위호환 위해 유지
    comment: str = ""   # 엔지니어 코멘트 (랏 운영 관리용)
    cache_enabled: bool = True  # False = 목록에 유지하되 캐싱에서 제외


class RamCachePriorityLotsSaveReq(BaseModel):
    product: str
    lots: List[RamCachePriorityLotItem] = []


@router.post("/ram-cache/priority-lots/save")
def save_ram_cache_priority_lots(
    req: RamCachePriorityLotsSaveReq,
    request: Request,
    _perm=Depends(require_page_manager("splittable")),
):
    """우선 lot 등록 목록 저장 (product 전체 교체)."""
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product is required")
    me = current_user(request)
    user_id = str(me.get("user_id") or me.get("name") or "unknown") if isinstance(me, dict) else "unknown"
    now_iso = datetime.datetime.now(datetime.timezone.utc).replace(microsecond=0).isoformat()
    entries: list[dict] = []
    for item in req.lots:
        lot_id = str(item.lot_id or "").strip().upper()
        if not lot_id:
            continue
        root_lot_id = lot_id[:5].upper()
        entries.append({
            "lot_id": lot_id,
            "root_lot_id": root_lot_id,
            "purpose": str(item.purpose or "").strip(),
            "note": str(item.note or "").strip(),
            "comment": str(item.comment or "").strip(),
            "cache_enabled": bool(item.cache_enabled),
            "created_at": now_iso,
            "created_by": user_id,
        })
    data = _load_priority_lots_file()
    data[product] = entries
    _save_priority_lots_file(data)
    return {"ok": True, "product": product, "lots": entries}


@router.get("/ram-cache/contents")
def get_ram_cache_contents(product: str = Query("")):
    """현재 RAM 캐시에 올라간 root lot 전체 목록 (전체 목록 서브탭용)."""
    product = str(product or "").strip()
    if not product:
        raise HTTPException(400, "product is required")
    # 소스 파일 해석
    source_fp = None
    try:
        source_fp = _product_path(product)
    except Exception:
        pass
    source_path = str(Path(source_fp).resolve()) if source_fp else ""
    # 우선 lot 등록 root 목록 로드 (is_priority 판정용)
    priority_data = _load_priority_lots_file()
    priority_roots: set[str] = set()
    product_upper = product.upper()
    for key, val in priority_data.items():
        if str(key).strip().upper() == product_upper and isinstance(val, list):
            for entry in val:
                if isinstance(entry, dict):
                    root = str(entry.get("root_lot_id") or "").strip().upper()
                    if root:
                        priority_roots.add(root)
            break
    # RAM 캐시 OrderedDict 에서 항목 수집
    with _ml_table_lookup._ROOT_RAM_CACHE_LOCK:
        entries = []
        product_bytes = 0
        total_bytes = 0
        for key, entry in _ml_table_lookup._ROOT_RAM_CACHE.items():
            eb = _ml_table_lookup.root_ram_entry_bytes(entry)
            total_bytes += eb
            if source_path and key[0] != source_path:
                continue
            product_bytes += eb
            root_lot_id = str(entry.get("root_lot_id") or key[1] or "").strip().upper()
            entries.append({
                "root_lot_id": root_lot_id,
                "row_count": int(entry.get("row_count") or 0),
                "estimated_mb": round(float(eb) / (1024 * 1024), 3),
                "loaded_at": entry.get("loaded_at") or "",
                "access_count": int(entry.get("access_count") or 0),
                "cache_group": str(entry.get("cache_group") or "other"),
                "cache_sources": list(entry.get("cache_sources") or []),
                "is_priority": root_lot_id in priority_roots,
                # 개별 축출(evict) 호출에 필요 — 미스 속도 측정용 "축출→재검색" 흐름.
                "source_path": str(key[0] or ""),
            })
    max_bytes = _ml_table_lookup._root_ram_cache_max_bytes()
    return {
        "ok": True, "product": product, "entries": entries,
        "product_mb": round(product_bytes / (1024 * 1024), 1),
        "total_mb": round(total_bytes / (1024 * 1024), 1),
        "max_mb": round(max_bytes / (1024 * 1024), 1) if max_bytes else 0,
        "total_roots": len(entries),
    }


@router.get("/ram-cache/overview")
def get_ram_cache_overview(request: Request):
    """제품별 RAM 캐시 현황 요약 — 캐시 관리 페이지 상단 제품별 분해용.

    RAM 캐시 키(소스 경로)를 제품명으로 되돌려 제품 단위로 집계한다.
    캐시에 아직 안 올라간 제품도 /products 목록 기준으로 0 값으로 포함한다."""
    # 제품 목록 (source 파일 경로 → 제품명 매핑 준비)
    known: dict[str, str] = {}  # resolved source path → product name
    file_names: dict[str, str] = {}  # product name → source file name (예열 결과 매칭용)
    product_rows: dict[str, dict] = {}
    try:
        for p in list_products().get("products", []):
            name = str(p.get("name") or "")
            if not name:
                continue
            product_rows.setdefault(name, {
                "product": name, "roots": 0, "mb": 0.0,
                "priority_total": 0, "priority_cached": 0,
            })
            try:
                fp = _product_path(name)
                if fp:
                    known[str(Path(fp).resolve())] = name
                    file_names[name] = Path(fp).name
            except Exception:
                pass
    except Exception:
        pass
    # 우선 lot 등록 현황 (제품별 등록 수)
    priority_data = _load_priority_lots_file()
    priority_roots_by_product: dict[str, set[str]] = {}
    for key, val in priority_data.items():
        if not isinstance(val, list):
            continue
        prod_name = str(key).strip()
        roots = {
            str(e.get("root_lot_id") or "").strip().upper()
            for e in val if isinstance(e, dict) and e.get("cache_enabled", True)
        }
        roots.discard("")
        priority_roots_by_product[prod_name] = roots
        row = product_rows.setdefault(prod_name, {
            "product": prod_name, "roots": 0, "mb": 0.0,
            "priority_total": 0, "priority_cached": 0,
        })
        row["priority_total"] = len(roots)
    # RAM 캐시 집계
    total_bytes = 0
    with _ml_table_lookup._ROOT_RAM_CACHE_LOCK:
        for key, entry in _ml_table_lookup._ROOT_RAM_CACHE.items():
            eb = _ml_table_lookup.root_ram_entry_bytes(entry)
            total_bytes += eb
            source_path = str(key[0] or "")
            prod_name = known.get(source_path)
            if not prod_name:
                # 미등록 소스 — 파일 stem 으로 제품명 유추
                stem = Path(source_path).stem if source_path else ""
                prod_name = _canonical_mltable_product_name(stem, allow_bare=True) or (stem or "(unknown)")
            row = product_rows.setdefault(prod_name, {
                "product": prod_name, "roots": 0, "mb": 0.0,
                "priority_total": 0, "priority_cached": 0,
            })
            row["roots"] += 1
            row["mb"] += eb / (1024 * 1024)
            root_lot_id = str(entry.get("root_lot_id") or key[1] or "").strip().upper()
            if root_lot_id in priority_roots_by_product.get(prod_name, set()):
                row["priority_cached"] += 1
    # 제품별 예산 병합 — 운영/개발 구분. 실제 캐싱 로직(_root_ram_cache_product_budget)과
    # 동일하게: 개발서버에서 max_roots_dev 가 없으면 운영 max_roots 가 아니라 개발
    # 기본 target 을 적용한다(표시값이 실제 적재량과 일치하도록).
    cfg = load_json(SOURCE_CFG, {})
    budgets = cfg.get("ram_cache_product_budgets") or {}
    _use_dev = _ml_table_lookup._root_ram_cache_use_dev()
    # 기본 target — 서버 역할 반영.
    default_max_roots = int(_ml_table_lookup._root_ram_cache_default_target_roots())
    for prod_name, row in product_rows.items():
        b = budgets.get(prod_name) if isinstance(budgets, dict) else None
        custom_val = None  # 명시 설정된 값 (없으면 기본 target)
        if isinstance(b, dict):
            try:
                if _use_dev:
                    if b.get("max_roots_dev") is not None:
                        custom_val = int(b.get("max_roots_dev"))
                    # dev 예산 미설정 → 기본 target 폴백 (custom 아님)
                else:
                    if b.get("max_roots") is not None:
                        custom_val = int(b.get("max_roots"))
            except Exception:
                custom_val = None
        elif b is not None and not _use_dev:
            # 스칼라 예산 = 운영 전용. 개발서버는 기본 target.
            try:
                custom_val = int(b)
            except Exception:
                custom_val = None
        row["max_roots"] = custom_val if custom_val else default_max_roots
        row["max_roots_custom"] = bool(custom_val)
        row["mb"] = round(row["mb"], 1)
    # 마지막 예열(warmup) 사이클 결과를 제품 행에 붙인다 — "왜 이 제품은 0인가"
    # (lookup 빌드 대기 / 자원 가드 / 예산 가득 / 사용자 요청 중 중단)를 화면이
    # 그대로 읽어 줄 수 있게. 이게 없어서 0 을 보고 원인을 알 수 없었다.
    warm = _ml_table_lookup.root_ram_warmup_overview()
    warm_products = warm.get("products") or {}
    for prod_name, row in product_rows.items():
        wrow = warm_products.get(file_names.get(prod_name) or f"{prod_name}.parquet") or {}
        row["warm"] = {
            "cached_roots": int(wrow.get("cached_roots") or 0),
            "target_roots": int(wrow.get("target_roots") or 0),
            "missing_roots": int(wrow.get("missing_roots") or 0),
            "resource_skipped_roots": int(wrow.get("resource_skipped_roots") or 0),
            "budget_skipped_roots": int(wrow.get("budget_skipped_roots") or 0),
            "skip_reason": str(wrow.get("last_skip_reason") or ""),
            "cache_status": str(wrow.get("cache_status") or ""),
            # 랏캐시에는 root 가 있는데 예열 후보가 0이던 상황을 화면에서 가른다.
            "available_roots": int(wrow.get("available_roots") or 0),
            "index_target_roots": int(wrow.get("index_target_roots") or 0),
            # 우선적재 미등록 제품의 적재 순서 기준 (0 = 미적용 = 우선 lot 등록됨).
            "step_threshold": int(wrow.get("step_threshold") or 0),
            "step_threshold_roots": int(wrow.get("step_threshold_roots") or 0),
            "build_pending": bool(wrow.get("build_pending")),
            "incomplete": bool(wrow.get("incomplete")),
            "seen": bool(wrow),
        } if wrow else None
    max_bytes = _ml_table_lookup._root_ram_cache_max_bytes()
    products = sorted(product_rows.values(), key=lambda r: (-r["mb"], r["product"]))
    return {
        "ok": True,
        "products": products,
        "total_mb": round(total_bytes / (1024 * 1024), 1),
        "max_mb": round(max_bytes / (1024 * 1024), 1) if max_bytes else 0,
        "default_max_roots": default_max_roots,
        "is_dev": _use_dev,
        # 예열 상태 요약 — 이 서버(운영/개발)에서 예열이 실제로 돌고 있는지,
        # 설정한 예산이 그대로 적용됐는지. 제품별 현황 위에 한 줄로 표시한다.
        "warmup": {
            "server_role": warm.get("server_role") or "",
            "scheduler_started": bool(warm.get("scheduler_started")),
            "disabled_reason": warm.get("disabled_reason") or "",
            "last_refresh_at": warm.get("last_refresh_at") or "",
            "last_resource_guard_reason": warm.get("last_resource_guard_reason") or "",
            "last_cycle_incomplete": bool(warm.get("last_cycle_incomplete")),
            "interval_minutes": warm.get("interval_minutes"),
            "retry_minutes": warm.get("retry_minutes"),
            "budget_gb": warm.get("budget_gb"),
            "budget_setting_gb": warm.get("budget_setting_gb"),
            "budget_capped": bool(warm.get("budget_capped")),
            "load_workers": warm.get("load_workers"),
        },
        # 캐시 관리 화면의 관리자 블록 노출 기준 — 이 페이지의 관리 기능은 모두
        # splittable page manager 권한으로 보호돼 있어 그 판정을 그대로 내려준다.
        # (일반 유저에겐 주요 Lot / 전체 캐시만 보인다)
        "can_manage": is_page_manager(current_user(request), "splittable"),
    }


@router.get("/memory/overview")
def get_memory_overview():
    """프로세스 내 전 캐시 메모리 종합 현황 — 캐시 관리 페이지용.

    root RAM 캐시(관리 페이지에 원래 보이던 것)만이 아니라 filebrowser
    preview 메모리 캐시, reformatize 캐시, view payload 캐시까지 합산해
    "표시 2GB vs 실제 RSS 8.5GB" 괴리를 없앤다. 메타데이터만 반환.
    """
    def _mb(n: int | float) -> float:
        return round(float(n or 0) / (1024 * 1024), 1)

    caches: list[dict] = []

    # 1) SplitTable root RAM 캐시 (ml_table_lookup)
    with _ml_table_lookup._ROOT_RAM_CACHE_LOCK:
        root_bytes = sum(
            _ml_table_lookup.root_ram_entry_bytes(e)
            for e in _ml_table_lookup._ROOT_RAM_CACHE.values()
        )
        root_entries = len(_ml_table_lookup._ROOT_RAM_CACHE)
    caches.append({
        "key": "splittable_root_ram",
        "label": "SplitTable root RAM 캐시",
        "entries": root_entries,
        "mb": _mb(root_bytes),
        "budget_mb": _mb(_ml_table_lookup._root_ram_cache_max_bytes()),
    })

    # 1-b) SplitTable 제품 전체 RAM 캐시 (product cache) — root 캐시와 별개로
    #      제품 단위 DataFrame 을 통째로 상주시키는 층. overview 미집계 항목이었다.
    with _PRODUCT_RAM_CACHE_LOCK:
        product_bytes = sum(
            (_product_ram_cache_estimated_bytes(e.get("df"))
             if e.get("df") is not None else int(e.get("estimated_bytes") or 0))
            for e in _PRODUCT_RAM_CACHE.values()
        )
        product_entries = len(_PRODUCT_RAM_CACHE)
    caches.append({
        "key": "splittable_product_ram",
        "label": "SplitTable 제품 RAM 캐시",
        "entries": product_entries,
        "mb": _mb(product_bytes),
        "budget_mb": _mb(_product_ram_cache_max_bytes()),
    })

    # 2) SplitTable view payload 캐시
    with _VIEW_CACHE_LOCK:
        view_bytes = int(_VIEW_CACHE_BYTES)
        view_entries = len(_VIEW_CACHE)
    caches.append({
        "key": "splittable_view_payload",
        "label": "SplitTable 조회결과(payload) 캐시",
        "entries": view_entries,
        "mb": _mb(view_bytes),
        "budget_mb": _mb(_view_cache_max_bytes()),
    })

    # 3) 파일탐색기 preview 메모리 캐시
    try:
        from core import filebrowser_cache as _fb_cache
        fb = _fb_cache.memory_cache_stats()
        caches.append({
            "key": "filebrowser_preview",
            "label": "파일탐색기 preview 캐시",
            "entries": int(fb.get("entries") or 0),
            "mb": _mb(fb.get("bytes")),
            "budget_mb": _mb(fb.get("budget_bytes")),
        })
    except Exception as e:
        logger.debug("memory overview: filebrowser stats unavailable: %s", e)

    # 4) reformatize (ET Index) 캐시 — wide 결과 + raw ET
    try:
        from routers import reformatize as _reformatize
        rf = _reformatize.cache_stats()
        for sub_key, label in (("wide", "ET Index 결과(wide) 캐시"), ("raw", "ET raw 데이터 캐시")):
            sub = rf.get(sub_key) or {}
            caches.append({
                "key": f"reformatize_{sub_key}",
                "label": label,
                "entries": int(sub.get("entries") or 0),
                "mb": _mb(sub.get("bytes")),
                "budget_mb": _mb(sub.get("budget_bytes")),
            })
    except Exception as e:
        logger.debug("memory overview: reformatize stats unavailable: %s", e)

    # 5) 검색 경로 보조 캐시 + lookup candidate index memo — 어느 tier 도 안 보던
    #    구간이라 "캐시 합계는 작은데 RSS 는 높다" 의 상당 부분이 여기였다.
    try:
        scratch = scratch_cache_sizes()
        scratch_bytes = sum(int((row or {}).get("bytes") or 0) for row in scratch.values())
        caches.append({
            "key": "splittable_scratch",
            "label": "SplitTable 보조 캐시(CSV 행·인덱스·스키마)",
            "entries": sum(int((row or {}).get("entries") or 0) for row in scratch.values()),
            "mb": _mb(scratch_bytes),
            "budget_mb": 0,
            "detail": scratch,
        })
    except Exception as e:
        logger.debug("memory overview: scratch cache stats unavailable: %s", e)
    try:
        memo = _ml_table_lookup.candidate_index_memo_stats()
        caches.append({
            "key": "lookup_candidate_index",
            "label": "lookup candidate index 파싱 사본",
            "entries": int(memo.get("entries") or 0),
            "mb": float(memo.get("used_mb") or 0.0),
            "budget_mb": float(memo.get("budget_mb") or 0.0),
        })
    except Exception as e:
        logger.debug("memory overview: candidate index memo stats unavailable: %s", e)

    caches_total_mb = round(sum(float(c.get("mb") or 0) for c in caches), 1)
    budget_total_mb = round(sum(float(c.get("budget_mb") or 0) for c in caches), 1)

    # 프로세스/호스트 메모리 — 캐시 합계와 실제 RSS 의 괴리를 함께 보여준다.
    process = {}
    try:
        from core.runtime_limits import process_memory_snapshot
        snap = process_memory_snapshot()
        process = {
            "rss_gb": snap.get("process_rss_gb"),
            "effective_gb": snap.get("process_memory_effective_gb"),
            "effective_kind": snap.get("process_memory_effective_kind"),
            "container_working_set_gb": snap.get("container_memory_working_set_gb"),
            "container_current_gb": snap.get("container_memory_current_gb"),
            "limit_gb": snap.get("process_memory_limit_gb"),
            "system_total_gb": snap.get("system_memory_total_gb"),
            "system_available_gb": snap.get("system_memory_available_gb"),
        }
    except Exception as e:
        logger.debug("memory overview: process snapshot unavailable: %s", e)

    # 메모리 워치독 + 캐시 풀 상한 — OOM 방어 상태를 같은 화면에서 확인.
    watchdog = {}
    try:
        from core import memory_watchdog as _memory_watchdog
        watchdog = _memory_watchdog.status()
    except Exception as e:
        logger.debug("memory overview: watchdog status unavailable: %s", e)
    budget_pool = {}
    try:
        from core import cache_budget as _cache_budget
        budget_pool = _cache_budget.overview()
    except Exception as e:
        logger.debug("memory overview: cache budget overview unavailable: %s", e)
    proxy = {}
    try:
        from core import upstream_proxy as _upstream_proxy
        proxy = _upstream_proxy.status()
    except Exception as e:
        logger.debug("memory overview: upstream proxy status unavailable: %s", e)

    return {
        "ok": True,
        "caches": caches,
        "caches_total_mb": caches_total_mb,
        "caches_budget_total_mb": budget_total_mb,
        "process": process,
        "watchdog": watchdog,
        "cache_budget_pool": budget_pool,
        "upstream_proxy": proxy,
    }


@router.get("/cache-event-log")
def get_cache_event_log(
    request: Request,
    limit: int = Query(100, ge=1, le=500),
    category: str = Query(""),
):
    """관리자 전용 — 캐시 성공/실패 이벤트 로그 + peak RAM."""
    user = current_user(request)
    if not is_page_manager(user, "splittable"):
        raise HTTPException(403, "관리자 전용")
    from core.cache_event_log import (get_events, get_jobs, milestones, peak_ram_info,
                                      progress_snapshot)
    return {
        "ok": True,
        "events": get_events(limit=limit, category=category),
        "peak_ram": peak_ram_info(),
        "jobs": get_jobs(recent=3),
        # 지금 돌고 있는 빌드의 "몇 랏 중 몇 랏". 끝난 작업은 빠진다.
        # category 필터와 무관하게 항상 전체 기준으로 준다.
        "progress": progress_snapshot(),
        # 제품 × 캐싱 작업별 최근 성공/실패 — 진행률보다 이쪽이 평소 보는 화면이다.
        "product_status": _product_cache_status_snapshot(nonblocking=True),
        # 제품 단위 시작/완료/실패만 뽑은 로그 — 청크 진행에 묻히지 않게.
        "milestones": milestones(),
        "queues": _cache_queue_snapshot(),
        # 서버 스캔 게이트(실행 1건 + 대기열) — 수동 스캔 중이 아닐 때도 화면에서
        # "예약 스캔이 대기 중"을 볼 수 있어야 한다.
        "scan_queue": _scan_gate_snapshot(),
    }


def _artifact_built_ts(value) -> float:
    """ISO 문자열 → epoch 초. tz 가 없으면 로컬 시각으로 본다. 실패하면 0."""
    raw = str(value or "").strip()
    if not raw:
        return 0.0
    try:
        return datetime.datetime.fromisoformat(raw.replace("Z", "+00:00")).timestamp()
    except Exception:
        return 0.0


def _path_mtime(path: Path) -> float:
    try:
        return float(path.stat().st_mtime)
    except Exception:
        return 0.0


def _required_split_cache_status(product: str) -> dict:
    """Return the four persistent SplitTable caches managed as one pipeline.

    This is deliberately an artifact check, not a reconstruction from recent
    event-log messages.  A seven-day-old successful event does not prove that
    its files still exist, and an existing cache must not be reported missing
    merely because its build event aged out of the log window.
    """
    canonical = (_canonical_mltable_product_name(product, allow_bare=True)
                 or str(product or "").strip().upper())
    if not canonical:
        return {"product": "", "kinds": [], "ready_count": 0, "total": 4,
                "all_ready": False, "missing_labels": []}

    lookup = _lookup_cache_public_meta_for(canonical)
    lookup_roots = int(lookup.get("root_lot_id_count") or 0)
    lookup_building = str(lookup.get("job_status") or lookup.get("status") or "") in {
        "queued", "running", "building",
    }
    lookup_ready = bool(
        lookup.get("has_cache")
        and (lookup.get("status") == "fresh" or lookup_building)
    )

    lookup_built_ts = _artifact_built_ts(lookup.get("built_at"))
    lookup_build_seconds = float(lookup.get("build_seconds") or 0.0)

    pivot_dir = _pivot_cache_path(canonical, "_probe").parent
    pivot_files = 0
    try:
        pivot_files = sum(1 for path in pivot_dir.glob("*.parquet") if path.is_file())
    except Exception:
        pivot_files = 0
    pivot_running = _pivot_cache_build_state(canonical) == "building"
    try:
        pivot_complete = not _pivot_cache_needs_build(canonical, _product_path(canonical))
    except Exception:
        pivot_complete = False
    # The builder's root fingerprint is the authoritative membership check.
    # File count alone can pass when one stale root remains while a current root
    # is missing, so it is only retained as a compact progress number.
    pivot_ready = bool(pivot_files > 0 and (pivot_complete or pivot_running))

    latest = _latest_lot_step_cache_status(canonical)
    latest_ready = bool(
        latest.get("cache_exists")
        and latest.get("format_current")
        and int(latest.get("product_row_count") or 0) > 0
    )

    fab_meta = _fab_lot_index_read_meta(canonical)
    fab_dir = _fab_lot_index_dir(canonical)
    fab_roots = 0
    try:
        fab_roots = sum(
            1 for path in fab_dir.iterdir()
            if path.is_dir() and path.name.startswith(f"{_FAB_IDX_ROOT_COL}=")
        )
    except Exception:
        fab_roots = 0
    with _FAB_IDX_BUILD_LOCK:
        fab_running = canonical in _FAB_IDX_BUILD_INPROGRESS
    fab_ready = bool(fab_meta.get("built_at") and fab_meta.get("root_col") and fab_roots > 0)

    # 각 산출물이 스스로 들고 있는 완료 시각. 이벤트 로그(7일 창)가 아니라
    # 파일 쪽이 근거라, 로그가 잘렸거나 워커 오프로드·fresh 스킵으로 완료 줄이
    # 안 남은 캐시도 "언제 끝났는지"를 화면에 그대로 줄 수 있다.
    kinds = [
        {"kind": "lookup", "label": "랏 lookup", "ready": lookup_ready,
         "state": "building" if lookup_building else ("ready" if lookup_ready else (lookup.get("status") or "missing")),
         "done": lookup_roots, "total": lookup_roots,
         "built_ts": lookup_built_ts, "build_seconds": lookup_build_seconds,
         "message": (f"lookup 빌드 완료 — {lookup_roots:,} roots"
                     + (f" · {lookup_build_seconds:.1f}s" if lookup_build_seconds > 0 else ""))},
        {"kind": "pivot", "label": "SplitTable pivot", "ready": pivot_ready,
         "state": "building" if pivot_running else ("ready" if pivot_ready else "missing"),
         "done": pivot_files, "total": lookup_roots,
         "built_ts": _path_mtime(pivot_dir / ".root_fingerprints.json"), "build_seconds": 0.0,
         "message": f"Pivot 빌드 완료 — {pivot_files:,} roots"},
        {"kind": "latest_lot", "label": "WIP latest-lot", "ready": latest_ready,
         "state": "ready" if latest_ready else ("building" if _MANUAL_LATEST_REFRESH_RUNNING else "missing"),
         "done": int(latest.get("product_row_count") or 0), "total": 0,
         "built_ts": _artifact_built_ts(latest.get("updated_at")), "build_seconds": 0.0,
         "message": f"WIP latest-lot 빌드 완료 — {int(latest.get('product_row_count') or 0):,} rows"},
        {"kind": "fab_index", "label": "FAB latest 인덱스", "ready": fab_ready,
         "state": "building" if fab_running else ("ready" if fab_ready else "missing"),
         "done": fab_roots, "total": 0,
         "built_ts": _artifact_built_ts(fab_meta.get("built_at")), "build_seconds": 0.0,
         "message": f"FAB latest 인덱스 빌드 완료 — {fab_roots:,} roots"},
    ]
    ready_count = sum(1 for row in kinds if row["ready"])
    return {
        "product": canonical,
        "kinds": kinds,
        "ready_count": ready_count,
        "total": len(kinds),
        "all_ready": ready_count == len(kinds),
        "missing_labels": [row["label"] for row in kinds if not row["ready"]],
    }


@router.get("/cache/required-status")
def required_split_cache_status(request: Request, product: str = Query(...)):
    """Selected product's four-cache readiness, without internal column names."""
    user = current_user(request)
    if not is_page_manager(user, "splittable"):
        raise HTTPException(403, "관리자 전용")
    return {"ok": True, **_required_split_cache_status(product)}


# ── /ram-cache/lot-status 메모리 안전장치 ────────────────────────────────
# 주요 lot 위치 조회는 per-root 파티션 미스 시 monolithic latest-lot 캐시
# 풀스캔으로 떨어진다. 예전에는 root 마다 풀스캔을 반복하고 latest_main 을
# 위해 전체 정렬까지 해서, 캐시 관리 페이지 진입만으로 RSS 가 수 GB 튀어
# 프로세스가 죽었다(OOM). 지금은 ①폴백 스캔을 제품당 1회로 배치, ②streaming
# collect 로 상한 고정, ③메모리 압박 시 스캔 스킵, ④결과 TTL 캐시 + 스캔
# 직렬화로 재진입/연속 클릭 시 재스캔을 막는다.
_RAM_CACHE_LOT_STATUS_TTL_SEC = 180.0
_RAM_CACHE_LOT_STATUS_LOCK = threading.Lock()
_RAM_CACHE_LOT_STATUS_COMPUTE_LOCK = threading.Lock()
_RAM_CACHE_LOT_STATUS_CACHE: OrderedDict = OrderedDict()
_RAM_CACHE_LOT_STATUS_COLS = ("root_lot_id", "lot_id", "step_id", "function_step", "tkout_time")


def _ram_cache_lot_status_compute(product: str, roots: list[str]) -> dict:
    product_upper = product.upper()
    # Vehicle_matching step_id → step_desc 맵 (main step 목록이기도 하다)
    step_desc_map: dict[str, str] = {}
    # 근사 매칭용: 영문 프리픽스별 (숫자, step_id, desc) 목록 — 정확 매칭이 없으면
    # 같은 프리픽스에서 step_id 보다 작은 것 중 가장 가까운 main step 의 desc 를 쓴다.
    prefix_steps: dict[str, list[tuple[int, str, str]]] = {}
    _STEP_SPLIT_RE = _re.compile(r"^([A-Z]+)(\d+)")
    try:
        from core import fab_reference
        vm_rows = fab_reference._read_rows(fab_reference.VEHICLE_MATCHING_FILE)
        bare = product_upper[len("ML_TABLE_"):] if product_upper.startswith("ML_TABLE_") else product_upper
        scoped = [r for r in vm_rows if str(r.get("product") or "").upper() in (product_upper, bare)] or vm_rows
        for r in scoped:
            sid = str(r.get("step_id") or "").strip().upper()
            desc = str(r.get("step_desc") or "").strip()
            if sid and desc and sid not in step_desc_map:
                step_desc_map[sid] = desc
                m = _STEP_SPLIT_RE.match(sid)
                if m:
                    prefix_steps.setdefault(m.group(1), []).append((int(m.group(2)), sid, desc))
        for steps in prefix_steps.values():
            steps.sort()
    except Exception:
        pass

    def _resolve_step_desc(step_id: str) -> tuple[str, bool]:
        """(desc, approx). 정확 매칭 우선, 없으면 같은 영문 프리픽스에서
        step_id 보다 작은 main step 중 가장 가까운 것의 desc (approx=True)."""
        sid = str(step_id or "").strip().upper()
        if not sid:
            return "", False
        exact = step_desc_map.get(sid)
        if exact:
            return exact, False
        m = _STEP_SPLIT_RE.match(sid)
        if not m:
            return "", False
        prefix, num = m.group(1), int(m.group(2))
        best = None
        for cand_num, _cand_sid, cand_desc in prefix_steps.get(prefix, []):
            if cand_num <= num:
                best = cand_desc
            else:
                break
        return (best or ""), bool(best)

    def _empty_status() -> dict:
        return {"step_id": "", "step_desc": "", "step_desc_approx": False,
                "fab_lot_id": "", "tkout_time": "", "wafer_count": 0}

    def _status_from_df(df) -> dict:
        status = _empty_status()
        if df is None or not df.height:
            return status
        status["wafer_count"] = int(df.height)
        if "tkout_time" in df.columns:
            df = df.sort("tkout_time", descending=True, nulls_last=True)
        top = df.row(0, named=True)
        step_id = str(top.get("step_id") or "").strip()
        status["step_id"] = step_id
        status["fab_lot_id"] = str(top.get("lot_id") or "").strip()
        tk = top.get("tkout_time")
        status["tkout_time"] = str(tk) if tk is not None else ""
        # step_desc 는 vehicle_matching 기준: 정확 매칭 → 같은 영문
        # 프리픽스의 이전 main step 근사 → (둘 다 없으면) function_step.
        desc, approx = _resolve_step_desc(step_id)
        if not desc:
            desc = str(top.get("function_step") or "").strip()
            approx = False
        status["step_desc"] = desc
        status["step_desc_approx"] = approx
        return status

    statuses: dict[str, dict] = {root: _empty_status() for root in roots}
    pending: list[str] = []  # per-root 파티션이 없어 monolithic 폴백이 필요한 root
    for root in roots:
        try:
            part_lf = _latest_lot_index_partition_lf(product, root)
            if part_lf is None:
                pending.append(root)
                continue
            names = part_lf.collect_schema().names()
            want = [c for c in _RAM_CACHE_LOT_STATUS_COLS if c in names]
            if "root_lot_id" in names:
                part_lf = part_lf.filter(
                    pl.col("root_lot_id").cast(_STR, strict=False).str.to_uppercase() == root)
            statuses[root] = _status_from_df(part_lf.select(want).collect())
        except Exception as e:
            logger.warning("ram-cache lot-status failed for %s/%s: %s", product, root, e)

    # 메모리 압박이면 monolithic 풀스캔(폴백·latest_main)을 건너뛴다 — 이
    # 페이지가 프로세스를 죽이면 안 된다. 파티션에서 얻은 위치만 노출.
    skipped_reason = ""
    try:
        from core.runtime_limits import process_memory_high
        if process_memory_high():
            skipped_reason = "process_memory_high"
    except Exception:
        pass

    if pending and not skipped_reason:
        # 폴백 root 들은 monolithic 캐시를 '한 번만' 스캔해 일괄 조회한다.
        try:
            lf = _latest_lot_step_cache_lf(product)
            if lf is not None:
                names = lf.collect_schema().names()
                if "root_lot_id" in names:
                    want = [c for c in _RAM_CACHE_LOT_STATUS_COLS if c in names]
                    from core.parquet_perf import collect_streaming
                    batch = collect_streaming(
                        lf.filter(
                            pl.col("root_lot_id").cast(_STR, strict=False)
                            .str.to_uppercase().is_in(sorted(pending))
                        ).select(want)
                    )
                    for root in pending:
                        sub = batch.filter(
                            pl.col("root_lot_id").cast(_STR, strict=False).str.to_uppercase() == root)
                        statuses[root] = _status_from_df(sub)
        except Exception as e:
            logger.warning("ram-cache lot-status batch scan failed for %s: %s", product, e)

    # 참고 헤더용: 이 제품에서 가장 최근 진행된 main step (vehicle_matching 에
    # 등재된 step 중 tkout_time 최신). 정확 매칭이 없으면 최신 행의 근사 desc.
    latest_main: dict = {}
    if not skipped_reason:
        try:
            lf = _latest_lot_step_cache_lf(product)
            if lf is not None:
                names = lf.collect_schema().names()
                want = [c for c in _RAM_CACHE_LOT_STATUS_COLS if c in names]
                lf2 = lf.select(want)
                if "tkout_time" in want:
                    lf2 = lf2.sort("tkout_time", descending=True, nulls_last=True)
                from core.parquet_perf import collect_streaming
                df = collect_streaming(lf2.head(2000))
                fallback = None
                for row in df.iter_rows(named=True):
                    sid = str(row.get("step_id") or "").strip().upper()
                    if not sid:
                        continue
                    entry = {
                        "step_id": sid,
                        "root_lot_id": str(row.get("root_lot_id") or "").strip(),
                        "fab_lot_id": str(row.get("lot_id") or "").strip(),
                        "tkout_time": str(row.get("tkout_time") or ""),
                    }
                    if sid in step_desc_map:
                        latest_main = {**entry, "step_desc": step_desc_map[sid], "approx": False}
                        break
                    if fallback is None:
                        desc, approx = _resolve_step_desc(sid)
                        if desc:
                            fallback = {**entry, "step_desc": desc, "approx": approx}
                if not latest_main and fallback:
                    latest_main = fallback
        except Exception as e:
            logger.warning("ram-cache lot-status latest_main failed for %s: %s", product, e)
    return {"ok": True, "product": product, "statuses": statuses,
            "latest_main_step": latest_main, "skipped_reason": skipped_reason}


@router.get("/ram-cache/lot-status")
def get_ram_cache_lot_status(product: str = Query("")):
    """주요 lot 의 현재 위치(최신 step) — 캐시 관리 페이지 주요 Lot 표의
    step_id/step_desc 자동 컬럼용. 등록된 lot 의 root 별로 latest-lot 캐시에서
    tkout_time 최신 행을 뽑는다."""
    product = str(product or "").strip()
    if not product:
        raise HTTPException(400, "product is required")
    data = _load_priority_lots_file()
    product_upper = product.upper()
    roots: list[str] = []
    for key, val in data.items():
        if str(key).strip().upper() == product_upper and isinstance(val, list):
            for entry in val:
                if isinstance(entry, dict):
                    root = str(entry.get("root_lot_id") or "").strip().upper()
                    if root and root not in roots:
                        roots.append(root)
            break
    try:
        source_sig = repr(_latest_lot_index_source_sig())
    except Exception:
        source_sig = ""
    cache_key = "|".join([product_upper, ",".join(sorted(roots)), source_sig])
    with _RAM_CACHE_LOT_STATUS_LOCK:
        hit = _RAM_CACHE_LOT_STATUS_CACHE.get(cache_key)
        if hit and time.monotonic() - hit[0] < _RAM_CACHE_LOT_STATUS_TTL_SEC:
            return hit[1]
    # 무거운 스캔은 한 번에 하나만 — 페이지 재진입/제품 연속 클릭이 몰려도
    # monolithic 스캔이 겹쳐 메모리가 배로 튀지 않게 직렬화한다.
    with _RAM_CACHE_LOT_STATUS_COMPUTE_LOCK:
        with _RAM_CACHE_LOT_STATUS_LOCK:
            hit = _RAM_CACHE_LOT_STATUS_CACHE.get(cache_key)
            if hit and time.monotonic() - hit[0] < _RAM_CACHE_LOT_STATUS_TTL_SEC:
                return hit[1]
        out = _ram_cache_lot_status_compute(product, roots)
        # 메모리 압박으로 스킵한 응답은 캐시하지 않는다 — 압박이 풀리면 재시도.
        if not out.get("skipped_reason"):
            with _RAM_CACHE_LOT_STATUS_LOCK:
                _RAM_CACHE_LOT_STATUS_CACHE[cache_key] = (time.monotonic(), out)
                while len(_RAM_CACHE_LOT_STATUS_CACHE) > 16:
                    _RAM_CACHE_LOT_STATUS_CACHE.popitem(last=False)
        return out


@router.get("/ram-cache/knob-allocation")
def get_ram_cache_knob_allocation(product: str = Query(""), lot_id: str = Query("")):
    """lot 의 KNOB 할당 비율 — 각 KNOB 항목에서 split 값별 wafer 수/%.
    SplitTable /view (payload 캐시 활용) 를 내부 호출해 계산한다."""
    product = str(product or "").strip()
    root = str(lot_id or "").strip().upper()[:5]
    if not product or not root:
        raise HTTPException(400, "product and lot_id are required")
    try:
        view = view_split(
            product=product, root_lot_id=root, wafer_ids="", prefix="KNOB",
            custom_name="", view_mode="all", history_mode="all", fab_lot_id="",
            custom_cols="", include_related=False, cache_first=True, request=None,
        )
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(500, f"view load failed: {e}")
    wafer_keys = view.get("wafer_keys") or []
    total_wafers = len(wafer_keys)
    out_rows = []
    for r in view.get("rows") or []:
        cells = r.get("_cells") or {}
        first = next(iter(cells.values()), {})
        if first.get("is_custom_tag") or first.get("is_management_row"):
            continue
        tally: dict[str, int] = {}
        for cell in cells.values():
            v = cell.get("actual")
            if v is None or str(v).strip() == "":
                continue
            sv = str(v).strip()
            tally[sv] = tally.get(sv, 0) + 1
        if not tally:
            continue
        assigned = sum(tally.values())
        values = [
            {"value": val, "count": cnt, "pct": round(cnt / assigned * 100, 1)}
            for val, cnt in sorted(tally.items(), key=lambda kv: (-kv[1], kv[0]))
        ][:15]
        out_rows.append({
            "param": r.get("_param") or "",
            "display": r.get("_display") or r.get("_param") or "",
            "assigned": assigned,
            "unassigned": max(0, total_wafers - assigned),
            "split_count": len(tally),
            "values": values,
        })
        if len(out_rows) >= 500:
            break
    # prefix(KNOB 등)가 최우선, 그 다음 split(값 2개 이상) 행 우선, 그 안에서 이름 순
    out_rows.sort(key=lambda row: (
        _natural_param_key(row["param"])[0],
        0 if row["split_count"] >= 2 else 1,
        _natural_param_key(row["param"])[1:]
    ))
    return {
        "ok": True, "product": product, "root_lot_id": root,
        "total_wafers": total_wafers, "rows": out_rows,
        "split_rows": sum(1 for row in out_rows if row["split_count"] >= 2),
    }


@router.get("/ram-cache/product-budgets")
def get_ram_cache_product_budgets():
    """제품별 RAM 캐시 예산(max_roots) + 적재순서 step 임계값 조회 — 운영/개발 구분."""
    cfg = load_json(SOURCE_CFG, {})
    budgets = cfg.get("ram_cache_product_budgets") or {}
    products: dict[str, dict] = {}
    if isinstance(budgets, dict):
        for prod_key, prod_val in budgets.items():
            if isinstance(prod_val, dict):
                products[prod_key] = {
                    "max_roots": int(prod_val.get("max_roots") or 1000),
                    "max_roots_dev": int(prod_val["max_roots_dev"]) if prod_val.get("max_roots_dev") is not None else None,
                    "step_threshold": int(prod_val["step_threshold"]) if prod_val.get("step_threshold") is not None else None,
                }
            else:
                try:
                    products[prod_key] = {"max_roots": int(prod_val), "max_roots_dev": None, "step_threshold": None}
                except Exception:
                    products[prod_key] = {"max_roots": 1000, "max_roots_dev": None, "step_threshold": None}
    default_max_roots = int(
        _ml_table_lookup._root_ram_cache_target_roots()
    )
    _is_dev = _ml_table_lookup._root_ram_cache_use_dev()
    return {
        "ok": True,
        "products": products,
        "default_max_roots": default_max_roots,
        "default_max_roots_dev": int(_ml_table_lookup._root_ram_cache_target_roots_dev()),
        "is_dev": _is_dev,
        # 우선적재(priority) 미등록 제품의 랏캐시 적재 순서 기준 — step_id 숫자 임계값.
        "default_step_threshold": int(_ml_table_lookup._root_ram_cache_step_threshold("")),
    }


class RamCacheProductBudgetSaveReq(BaseModel):
    product: str
    max_roots: int = 1000
    max_roots_dev: int | None = None
    step_threshold: int | None = None


@router.post("/ram-cache/product-budgets/save")
def save_ram_cache_product_budget(
    req: RamCacheProductBudgetSaveReq,
    _perm=Depends(require_page_manager("splittable")),
):
    """제품별 RAM 캐시 예산(max_roots) + 적재순서 step 임계값 저장 — 운영/개발 구분.
    source_config.json 의 ram_cache_product_budgets 아래에 기록한다."""
    product = str(req.product or "").strip()
    if not product:
        raise HTTPException(400, "product is required")
    max_roots = max(0, min(ROOT_LOT_CACHE_LIMIT_MAX, int(req.max_roots or 1000)))
    entry: dict = {"max_roots": max_roots}
    if req.max_roots_dev is not None:
        entry["max_roots_dev"] = max(0, min(ROOT_LOT_CACHE_LIMIT_MAX, int(req.max_roots_dev)))
    if req.step_threshold is not None:
        entry["step_threshold"] = max(0, int(req.step_threshold))
    cfg = load_json(SOURCE_CFG, {})
    budgets = cfg.setdefault("ram_cache_product_budgets", {})
    if not isinstance(budgets, dict):
        budgets = {}
        cfg["ram_cache_product_budgets"] = budgets
    # 기존 값 보존 — 화면이 일부 필드만 보내도 나머지 설정이 지워지지 않게.
    existing = budgets.get(product)
    if isinstance(existing, dict):
        if req.max_roots_dev is None and existing.get("max_roots_dev") is not None:
            entry["max_roots_dev"] = existing["max_roots_dev"]
        if req.step_threshold is None and existing.get("step_threshold") is not None:
            entry["step_threshold"] = existing["step_threshold"]
    budgets[product] = entry
    save_json(SOURCE_CFG, cfg)
    return {"ok": True, "product": product, "max_roots": max_roots,
            "max_roots_dev": entry.get("max_roots_dev"),
            "step_threshold": entry.get("step_threshold")}


def _resolve_override_meta(product: str, include_diagnostics: bool = True) -> dict:
    """v8.8.5: view / ml-table-match 양쪽에서 공용. 현재 product 에 대해 적용된 오버라이드 설정 요약.

    Returns (모든 필드 optional, 에러 시 error 로 이유 표기):
      {
        "enabled": bool,              # 조인 실제 수행 여부
        "manual_override": bool,      # SOURCE_CFG 에 명시된 fab_source 사용 여부
        "fab_source": str,            # 사용된 fab_source 경로 (e.g. "1.RAWDATA_DB_FAB/PRODA")
        "fab_col": str,               # 실제 join 하는 fab 컬럼 이름
        "ts_col": str,                # 최신도 판정에 쓰는 ts 컬럼 (빈 문자열이면 레거시 keep=last)
        "join_keys": [str],
        "scanned_files": [str],       # fab_source 아래 발견된 parquet 들 (최대 20)
        "scanned_count": int,         # 실제 파일 개수
        "row_count": int,             # fab_source LazyFrame 전체 row 수 (scanned)
        "sample_fab_values": [str],   # head(5) 의 fab_col 값 — "어디서 읽어옴?" 답변용
        "error": str | None,
      }
    """
    meta = {
        "enabled": False, "manual_override": False,
        "fab_source": "", "fab_col": "", "ts_col": "",
        "join_keys": [], "scanned_files": [], "scanned_count": 0,
        "row_count": 0, "sample_fab_values": [], "error": None,
        "raw_columns": [], "runtime_columns": [], "column_aliases": {}, "schema_mode": "unknown",
        # v8.8.16: hive 원천에서 끌어오기로 한 override 컬럼 목록 + 실제 스키마에 존재하는 것만.
        "override_cols": [], "override_cols_present": [], "override_cols_missing": [],
    }
    try:
        product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
        cfg = load_json(SOURCE_CFG, {}) if SOURCE_CFG.exists() else {}
        ov = _lot_override_for(cfg, product)
        manual = (ov.get("fab_source") or "").strip()
        # v8.8.21: root:~~ 는 deprecated — 저장된 값이 남아있어도 무시하고 auto-derive 로 재매칭.
        if manual.startswith("root:"):
            manual = ""
        fab_source = manual or _auto_derive_fab_source(product)
        meta["manual_override"] = bool(manual)
        meta["fab_source"] = fab_source
        # v8.8.19: 진단 정보 — 어떤 data_root/DB 에서 어떤 후보를 탐색했는지 노출.
        db_base = _db_base()
        base_root = _base_root()
        meta["db_root"] = str(db_base)
        meta["base_root"] = str(base_root)
        meta["db_root_exists"] = bool(db_base.exists())
        meta["searched_db_roots"] = [p.name for p in _list_db_roots()]

        if not fab_source:
            if product.casefold().startswith("ml_table_"):
                pro = product[len("ML_TABLE_"):].strip()
                # 실제로 탐색한 후보 경로를 모두 리스트업
                tried = []
                for root_dir in _list_db_roots():
                    tried.append(f"{root_dir.name}/{pro}")
                if not _list_db_roots():
                    tried.append(f"(db_root 비어있거나 '1.RAWDATA_DB' 하위 제품 폴더 없음: {db_base})")
                meta["error"] = (
                    f"자동 매칭 실패: product='{product}' → pro='{pro}'. "
                    f"db_root='{db_base}'. "
                    f"후보 탐색: {tried if tried else '(없음)'}. "
                    f"권장 해결: data_root/DB 아래 '1.RAWDATA_DB/{pro}/' 가 존재하거나, "
                    f"수동으로 lot_overrides.{product}.fab_source 를 지정."
                )
                meta["tried_candidates"] = tried
            else:
                meta["error"] = "ML_TABLE_ prefix 아님 — 오버라이드 off."
            return meta

        # locate fab_source folder/file to list scanned files.  The resolver also
        # treats 1.RAWDATA_DB_FAB/<PROD> and 1.RAWDATA_DB/<PROD> as equivalent
        # FAB-history roots for production soft landing.
        fp, resolved_fab_source = _resolve_fab_source_target(fab_source)
        tried = []
        for root in (db_base, base_root):
            if not root or not root.exists():
                tried.append(f"{root} (not exist)" if root else "(None)")
                continue
            for rel in dict.fromkeys([fab_source, resolved_fab_source]):
                if rel:
                    tried.append(str(root / rel) + ("" if fp is not None else "  (not found)"))
        if fp is None:
            meta["tried_candidates"] = tried
            meta["error"] = (
                f"fab_source 경로를 찾을 수 없음: '{fab_source}'. "
                f"탐색 경로: {tried}. db_root='{db_base}' base_root='{base_root}'. "
                f"fab_source 는 데모/운영 모두 db_root 기준 상대경로만 사용하세요 "
                f"(예: '1.RAWDATA_DB_FAB/PRODA', not 'DB/1.RAWDATA_DB_FAB/PRODA')."
            )
            return meta
        if resolved_fab_source:
            meta["fab_source"] = resolved_fab_source
            fab_source = resolved_fab_source
        if fp.is_dir():
            parquets = _rglob_files_ci(fp, (".parquet",))
            base_for_rel = fp.parent if fp.parent.exists() else fp
            rels = []
            for p in parquets:
                try:
                    rels.append(str(p.relative_to(_db_base())))
                except Exception:
                    try:
                        rels.append(str(p.relative_to(_base_root())))
                    except Exception:
                        rels.append(str(p))
            meta["scanned_count"] = len(parquets)
            meta["scanned_files"] = [r.replace("\\", "/") for r in rels[:20]]
        else:
            meta["scanned_count"] = 1
            try:
                meta["scanned_files"] = [str(fp.relative_to(_db_base())).replace("\\", "/")]
            except Exception:
                meta["scanned_files"] = [str(fp)]

        raw_lf = _scan_fab_source_raw(fab_source)
        fab_lf = _scan_fab_source(fab_source)
        if fab_lf is None:
            meta["error"] = f"스캔 실패 (parquet 없음 또는 읽기 불가): {fab_source}"
            return meta
        # v8.8.22: CI 정렬 — ML_TABLE 대문자 vs hive 소문자 컬럼 이름 차이를 흡수.
        try:
            main_fp = _product_path(product)
            if main_fp.suffix.lower() == ".csv":
                main_names_list = pl.scan_csv(str(main_fp), infer_schema_length=5000).collect_schema().names()
            else:
                main_names_list = _scan_parquet_compat(str(main_fp)).collect_schema().names()
        except Exception:
            main_names_list = []
        fab_lf, fab_schema_names = _ci_align_fab_to_main(fab_lf, main_names_list)
        fab_names = fab_schema_names  # list after rename
        main_names = main_names_list
        try:
            raw_names = raw_lf.collect_schema().names() if raw_lf is not None else []
        except Exception:
            raw_names = []
        meta["raw_columns"] = raw_names
        meta["runtime_columns"] = list(fab_names)
        meta["column_aliases"] = _detect_source_column_aliases(raw_names, fab_names)
        meta["schema_mode"] = "adapted" if meta["column_aliases"] else "raw"

        # join keys
        join_keys = ov.get("join_keys") or []
        if isinstance(join_keys, str):
            join_keys = [k.strip() for k in join_keys.split(",") if k.strip()]
        # 유저가 지정한 키도 CI 로 실제 컬럼명에 매핑.
        if join_keys:
            mapped = []
            for k in join_keys:
                actual = _ci_resolve_in(k, main_names) or _resolve_source_col_name(k, fab_names)
                if actual:
                    mapped.append(actual)
            join_keys = mapped
        if not join_keys:
            join_keys = _default_override_join_keys(main_names, fab_names)
        join_keys = [k for k in join_keys if k in fab_names]
        meta["join_keys"] = join_keys

        # fab_col / ts_col 추론 (v8.8.22: CI 매칭 — fab_lf 는 이미 main casing 으로 align 됨).
        fc_raw = (ov.get("fab_col") or "").strip()
        meta["fab_col"] = (_resolve_source_col_name(fc_raw, fab_names) if fc_raw else "") \
                         or _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_names) \
                         or "fab_lot_id"
        tc_raw = (ov.get("ts_col") or "").strip()
        meta["ts_col"] = (_resolve_source_col_name(tc_raw, fab_names) if tc_raw else "") \
                         or _pick_ts_col(fab_names) \
                         or ""

        # v8.8.16: override_cols — 기본 (_DEFAULT_OVERRIDE_COLS) + manual ov.override_cols + 레거시 fab_col 병합.
        raw_oc = ov.get("override_cols")
        if isinstance(raw_oc, str):
            raw_oc = [c.strip() for c in raw_oc.split(",") if c.strip()]
        if not raw_oc:
            raw_oc = list(_DEFAULT_OVERRIDE_COLS)
        # 레거시 fab_col 도 합류 (중복 제거).
        if meta["fab_col"] and meta["fab_col"] not in raw_oc:
            raw_oc = list(raw_oc) + [meta["fab_col"]]
        # v8.8.22: CI 매칭 — 사용자가 소문자로 적었어도 실제 스키마의 casing 으로 맵핑.
        resolved_oc = []
        for c in raw_oc:
            actual = _resolve_source_col_name(c, fab_names)
            resolved_oc.append(actual or c)
        meta["override_cols"] = list(resolved_oc)
        meta["override_cols_present"] = [c for c in resolved_oc if c in fab_names]
        meta["override_cols_missing"] = [c for c in resolved_oc if c not in fab_names]

        if meta["fab_col"] not in fab_names:
            meta["error"] = f"fab_col '{meta['fab_col']}' 이 소스 스키마에 없음. 소스 컬럼: {fab_names[:20]}"
            return meta
        if not join_keys:
            meta["error"] = f"공통 join key 없음. 소스 컬럼: {fab_names[:20]}"
            return meta

        # row count + sample
        if include_diagnostics:
            try:
                rc = fab_lf.select(pl.len()).collect()
                meta["row_count"] = int(rc.item()) if rc.height > 0 else 0
            except Exception as e:
                meta["row_count"] = -1
        try:
            sample_cols = [c for c in (join_keys + [meta["fab_col"]] + ([meta["ts_col"]] if meta["ts_col"] else [])) if c in fab_names]
            sample = fab_lf.select(sample_cols)
            if include_diagnostics and meta["ts_col"] and meta["ts_col"] in fab_names:
                sample = sample.sort(meta["ts_col"], descending=True, nulls_last=True)
            vals = sample.head(5).collect()
            if meta["fab_col"] in vals.columns:
                meta["sample_fab_values"] = [
                    ("" if v is None else str(v)) for v in vals[meta["fab_col"]].to_list()
                ]
        except Exception as e:
            pass
        meta["enabled"] = True
    except Exception as e:
        meta["error"] = f"resolve 중 예외: {type(e).__name__}: {e}"
    return meta


def _resolve_override_meta_light(product: str) -> dict:
    """Cheap view badge metadata; avoid rescanning FAB source after /view already did."""
    meta = {
        "enabled": False, "manual_override": False,
        "fab_source": "", "fab_col": "fab_lot_id", "ts_col": "",
        "root_col": "", "wf_col": "", "join_keys": [], "override_cols": [],
        "override_cols_present": [],
        "scanned_count": 0, "row_count": 0, "sample_fab_values": [],
        "raw_columns": [], "runtime_columns": [], "column_aliases": {},
        "error": None,
    }
    try:
        product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
        cfg = load_json(SOURCE_CFG, {}) if SOURCE_CFG.exists() else {}
        ov = _lot_override_for(cfg, product)
        manual = _normalize_fab_source_path((ov.get("fab_source") or "").strip())
        if manual.startswith("root:"):
            manual = ""
        fab_source = manual or _auto_derive_fab_source(product)
        meta["manual_override"] = bool(manual)
        meta["fab_source"] = fab_source
        meta["enabled"] = bool(fab_source)
        meta["root_col"] = (ov.get("root_col") or "").strip()
        meta["wf_col"] = (ov.get("wf_col") or ov.get("wafer_col") or "").strip()
        meta["fab_col"] = (ov.get("fab_col") or "fab_lot_id").strip() or "fab_lot_id"
        meta["ts_col"] = (ov.get("ts_col") or "").strip()
        join_keys = ov.get("join_keys") or []
        if isinstance(join_keys, str):
            join_keys = [k.strip() for k in join_keys.split(",") if k.strip()]
        meta["join_keys"] = list(join_keys)
        raw_oc = ov.get("override_cols")
        if isinstance(raw_oc, str):
            raw_oc = [c.strip() for c in raw_oc.split(",") if c.strip()]
        meta["override_cols"] = list(raw_oc or _DEFAULT_OVERRIDE_COLS)
        meta["override_cols_present"] = list(meta["override_cols"])
        if not fab_source and product.casefold().startswith("ml_table_"):
            meta["error"] = "FAB source not matched"
    except Exception as e:
        meta["error"] = f"{type(e).__name__}: {e}"
    return meta

def _pick_first_present(candidates, available_names):
    av = set(available_names)
    for c in candidates:
        if c in av:
            return c
    return ""


def _pick_first_present_ci(candidates, available_names):
    """v8.8.22: case-insensitive 버전. 실제 스키마의 정확한 casing 을 반환."""
    ci = {n.casefold(): n for n in available_names}
    for c in candidates:
        actual = ci.get(c.casefold())
        if actual:
            return actual
    return ""


def _pick_ts_col(available_names):
    """Pick the most-likely time column from a FAB source."""
    primary = _pick_first_present_ci(_TS_COL_CANDIDATES, available_names)
    if primary:
        return primary
    for name in (available_names or []):
        low = str(name).casefold()
        if "time" in low or "timestamp" in low or low.endswith("_ts") or low.startswith("ts_"):
            return name
    return _pick_first_present_ci(("date",), available_names)


def _resolve_source_col_name(name: str, available_names):
    """Resolve user-facing raw/runtime column names against runtime source schema."""
    actual = _ci_resolve_in(name, available_names)
    if actual:
        return actual
    folded = str(name or "").strip().casefold()
    if not folded:
        return ""
    ci = {str(n).casefold(): n for n in (available_names or [])}
    for raw_name, runtime_name in _RAW_TO_RUNTIME_ALIAS_CANDIDATES.items():
        if folded == raw_name.casefold():
            actual = ci.get(runtime_name.casefold())
            if actual:
                return actual
        if folded == runtime_name.casefold():
            actual = ci.get(raw_name.casefold())
            if actual:
                return actual
    return ""


def _detect_source_column_aliases(raw_names, runtime_names):
    """Return raw->runtime aliases introduced by source adaptation."""
    raw_ci = {str(n).casefold(): n for n in (raw_names or [])}
    runtime_ci = {str(n).casefold(): n for n in (runtime_names or [])}
    out = {}
    for raw_name, runtime_name in _RAW_TO_RUNTIME_ALIAS_CANDIDATES.items():
        raw_actual = raw_ci.get(raw_name.casefold())
        runtime_actual = runtime_ci.get(runtime_name.casefold())
        if raw_actual and runtime_actual and runtime_name.casefold() not in raw_ci:
            out[raw_actual] = runtime_actual
    return out


def _prefer_raw_schema_name(name: str, raw_names, runtime_names):
    """Map runtime alias names back to physical raw schema names for UI display."""
    actual_raw = _ci_resolve_in(name, raw_names)
    if actual_raw:
        return actual_raw
    aliases = _detect_source_column_aliases(raw_names, runtime_names)
    runtime_to_raw = {str(v).casefold(): k for k, v in aliases.items()}
    return runtime_to_raw.get(str(name or "").strip().casefold(), name)


def _fab_source_context(product: str) -> dict:
    """Return the active FAB history source and resolved key columns for a product."""
    p = (product or "").strip()
    if not p:
        return {}
    ml_product = _canonical_mltable_product_name(p, allow_bare=True)
    try:
        cfg = load_json(SOURCE_CFG, {}) if SOURCE_CFG.exists() else {}
        ov = _lot_override_for(cfg, ml_product)
        fab_source = (ov.get("fab_source") or "").strip()
        if fab_source.startswith("root:"):
            fab_source = ""
        if not fab_source:
            fab_source = _auto_derive_fab_source(ml_product)
        include_all = _foreground_global_fab_scan_enabled()
        if not fab_source and not _global_fab_source_paths("", include_all=include_all):
            return {}
        _, resolved_fab_source = _resolve_fab_source_target(fab_source) if fab_source else (None, "")
        if resolved_fab_source:
            fab_source = resolved_fab_source
        fab_lf, fab_sources = _scan_global_fab_sources(fab_source, include_all=include_all)
        if fab_lf is None:
            return {}
        try:
            main_fp = _product_path(ml_product)
            if main_fp.suffix.lower() == ".csv":
                main_names = pl.scan_csv(str(main_fp), infer_schema_length=5000).collect_schema().names()
            else:
                main_names = _scan_parquet_compat(str(main_fp)).collect_schema().names()
        except Exception:
            main_names = []
        fab_lf, fab_names = _ci_align_fab_to_main(fab_lf, main_names)
        try:
            fab_names = fab_lf.collect_schema().names()
        except Exception:
            pass
        root_col = _resolve_source_col_name((ov.get("root_col") or "").strip(), fab_names) \
                   or _pick_first_present_ci(("root_lot_id",), fab_names)
        wafer_col = _resolve_source_col_name((ov.get("wf_col") or ov.get("wafer_col") or "").strip(), fab_names) \
                    or _pick_first_present_ci(("wafer_id", "wafer"), fab_names)
        fab_col = _resolve_source_col_name((ov.get("fab_col") or "").strip(), fab_names) \
                  or _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_names)
        ts_col = _resolve_source_col_name((ov.get("ts_col") or "").strip(), fab_names) \
                 or _pick_ts_col(fab_names)
        if not root_col or not fab_col:
            return {}
        return {
            "lf": fab_lf,
            "source": fab_source,
            "sources": fab_sources,
            "root_col": root_col,
            "wafer_col": wafer_col,
            "fab_col": fab_col,
            "ts_col": ts_col,
            "columns": fab_names,
        }
    except Exception as e:
        logger.warning("_fab_source_context 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return {}


def _clean_str(v) -> str:
    s = "" if v is None else str(v).strip()
    return "" if s in ("", "None", "null") else s


def _wafer_sort_key(v: str):
    s = str(v or "").strip()
    try:
        return (0, int(s.upper().lstrip("W")))
    except Exception:
        return (1, s.upper())


def _merge_wafer_scope(user_wafer_ids: str, source_wafers: list[str]) -> str:
    """Intersect user wafer filter with FAB-source wafer scope when both exist."""
    source = [_clean_str(w) for w in (source_wafers or [])]
    source = [w for w in source if w]
    if not source:
        return user_wafer_ids or ""
    user = [w.strip() for w in str(user_wafer_ids or "").split(",") if w.strip()]
    if not user:
        return ",".join(sorted(dict.fromkeys(source), key=_wafer_sort_key))

    def norm(w):
        s = str(w or "").strip().upper()
        try:
            return str(int(s.lstrip("W")))
        except Exception:
            return s

    user_norm = {norm(w) for w in user}
    kept = [w for w in source if norm(w) in user_norm]
    if not kept:
        return "__NO_WAFER_MATCH__"
    return ",".join(sorted(dict.fromkeys(kept), key=_wafer_sort_key))


def _fab_history_scope(product: str, root_lot_id: str = "", fab_lot_id: str = "",
                       prefix: str = "", limit: int = 500,
                       prefer_raw_latest: bool = False) -> dict:
    """Query FAB history as current SplitTable lot identity.

    Candidate LOT_ID values must follow the same contract as ML_TABLE/SplitTable:
    pick the latest FAB row per root_lot_id + wafer_id first, then expose the
    resulting fab_lot_id/lot_id set.
    """
    root_lot_id = root_lot_id if isinstance(root_lot_id, str) else ""
    fab_lot_id = fab_lot_id if isinstance(fab_lot_id, str) else ""
    prefix = prefix if isinstance(prefix, str) else ""
    try:
        limit = int(limit)
    except Exception:
        limit = 500
    cache_key = (
        "fab_history_scope",
        _lot_lookup_cache_sig(product),
        str(product or "").strip(),
        root_lot_id.strip(),
        fab_lot_id.strip(),
        prefix.strip(),
        limit,
        bool(prefer_raw_latest),
    )
    cached = _lot_lookup_cache_get(cache_key)
    if cached is not None:
        return cached

    def finish(payload: dict) -> dict:
        return _lot_lookup_cache_set(cache_key, payload)

    if not prefer_raw_latest:
        cached_scope = _fab_history_scope_from_cache(
            product, root_lot_id=root_lot_id, fab_lot_id=fab_lot_id,
            prefix=prefix, limit=limit,
        )
        if cached_scope is not None:
            has_explicit_scope = bool(root_lot_id.strip() or fab_lot_id.strip())
            if cached_scope.get("candidates") or not has_explicit_scope:
                return finish(cached_scope)

    ctx = _fab_source_context(product)
    if not ctx:
        return finish({"candidates": [], "root_ids": [], "wafer_ids": [], "source": "", "query_ok": False})
    root_col = ctx["root_col"]
    fab_col = ctx["fab_col"]
    wafer_col = ctx.get("wafer_col") or ""
    ts_col = ctx.get("ts_col") or ""
    select_exprs = [
        pl.col(root_col).cast(_STR, strict=False).alias("root"),
        pl.col(fab_col).cast(_STR, strict=False).alias("fab"),
    ]
    if wafer_col:
        select_exprs.append(pl.col(wafer_col).cast(_STR, strict=False).alias("wafer"))
    if ts_col:
        select_exprs.append(pl.col(ts_col).cast(_STR, strict=False).alias("ts"))
    q = ctx["lf"].select(select_exprs)
    q = q.filter(pl.col("root").is_not_null() & pl.col("fab").is_not_null())
    root_scope = (root_lot_id or "").strip()
    fab_scope = (fab_lot_id or "").strip()
    if root_scope:
        q = q.filter(_join_key_expr("root") == root_scope.upper())
    if fab_scope:
        q = q.filter(_join_key_expr("fab") == fab_scope.upper())
    latest_subset = ["root"] + (["wafer"] if wafer_col else [])
    if ts_col:
        q = q.sort("ts", descending=True, nulls_last=True)
        q = q.unique(subset=latest_subset, keep="first", maintain_order=True)
    else:
        q = q.unique(subset=latest_subset, keep="last", maintain_order=True)
    if not fab_scope and prefix.strip():
        q = q.filter(_contains_literal_ci_expr("fab", prefix))
    try:
        fabs = _limited_unique_values(
            q, "fab", prefix="", limit=limit,
            preview_only=not bool(root_scope or fab_scope or prefix.strip()),
        )
        roots: list[str] = [root_scope] if root_scope else []
        wafers: list[str] = []
        # Exact fab lookup is used by /view to infer the root and wafer scope.
        # Keep that metadata precise, but avoid collecting it for broad previews.
        if fab_scope and fabs:
            meta_cols = [pl.col("root")]
            if wafer_col:
                meta_cols.append(pl.col("wafer"))
            meta_df = q.select(meta_cols).unique().collect()
            roots = sorted({s for s in (_clean_str(v) for v in meta_df["root"].to_list()) if s})
            if "wafer" in meta_df.columns:
                wafers = sorted({s for s in (_clean_str(v) for v in meta_df["wafer"].to_list()) if s}, key=_wafer_sort_key)
    except Exception as e:
        logger.warning("_fab_history_scope 조회 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return finish({"candidates": [], "root_ids": [], "wafer_ids": [],
                       "source": ctx.get("source", ""), "query_ok": False})
    if not fabs:
        return finish({"candidates": [], "root_ids": [], "wafer_ids": [],
                       "source": ctx.get("source", ""), "query_ok": True})
    return finish({
        "candidates": fabs,
        "root_ids": roots,
        "wafer_ids": wafers,
        "source": ctx.get("source", ""),
        "query_ok": True,
    })


def _fab_history_root_candidates(product: str, prefix: str = "", limit: int = 500) -> dict:
    """Return root_lot_id candidates from the configured FAB DB source.

    SplitTable's editable source is ML_TABLE_*, but operators choose lots from
    the live FAB history.  Use the configured fab_source first so the dropdown
    follows the same DB path that /view and fab_lot_id matching use.
    """
    try:
        limit = max(1, int(limit or 500))
    except Exception:
        limit = 500
    cache_key = (
        "fab_history_root_candidates",
        _lot_lookup_cache_sig(product),
        str(product or "").strip(),
        str(prefix or "").strip(),
        limit,
    )
    cached = _lot_lookup_cache_get(cache_key)
    if cached is not None:
        return cached

    def finish(payload: dict) -> dict:
        return _lot_lookup_cache_set(cache_key, payload)

    cached_roots = _fab_history_root_candidates_from_cache(product, prefix=prefix, limit=limit)
    if cached_roots is not None:
        return finish(cached_roots)

    ctx = _fab_source_context(product)
    if not ctx:
        return finish({"candidates": [], "source": ""})
    root_col = ctx.get("root_col") or ""
    if not root_col:
        return finish({"candidates": [], "source": ctx.get("source", "")})
    try:
        values = _limited_unique_values(ctx["lf"], root_col, prefix=prefix, limit=limit)
    except Exception as e:
        logger.warning("_fab_history_root_candidates 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        return finish({"candidates": [], "source": ctx.get("source", "")})
    return finish({"candidates": values, "source": ctx.get("source", "")})


def _merge_candidate_values(*groups, limit: int = 500) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    try:
        limit = max(1, int(limit or 500))
    except Exception:
        limit = 500
    for group in groups:
        for value in group or []:
            text = _clean_str(value)
            if not text:
                continue
            key = text.upper()
            if key in seen:
                continue
            seen.add(key)
            out.append(text)
            if len(out) >= limit:
                return out
    return out


def _plan_product_name(product: str) -> str:
    raw = str(product or "").strip()
    canonical = _canonical_mltable_product_name(raw, allow_bare=True)
    return canonical or safe_id(raw or "product")


def _plan_history_path(product: str) -> Path:
    return PLAN_DIR / f"{_plan_product_name(product)}.json"


def _plan_alias_paths(product: str) -> list[Path]:
    """Plan store aliases kept for older callers that used bare product names."""
    canonical = _plan_history_path(product)
    out = [canonical]
    raw = str(product or "").strip()
    if raw:
        legacy = PLAN_DIR / f"{safe_id(raw)}.json"
        if legacy != canonical:
            out.insert(0, legacy)
    return out


def _load_plan_data(product: str) -> dict:
    merged = {"plans": {}, "history": [], "mismatch_alerts": {}}
    seen_history: set[str] = set()
    for fp in _plan_alias_paths(product):
        # cached — merged 로 복사만 하고 원본은 건드리지 않는다. cold 검색마다 plan
        # 전체를 재파싱하던 비용(GIL 점유 + 공유드라이브 read)을 없앤다.
        data = load_json_cached(fp, {}) if fp.exists() else {}
        if not isinstance(data, dict):
            continue
        plans = data.get("plans")
        if isinstance(plans, dict):
            merged["plans"].update(plans)
        mismatch_alerts = data.get("mismatch_alerts")
        if isinstance(mismatch_alerts, dict):
            merged["mismatch_alerts"].update(mismatch_alerts)
        hist = data.get("history")
        if isinstance(hist, list):
            for row in hist:
                if not isinstance(row, dict):
                    continue
                try:
                    key = json.dumps(row, sort_keys=True, ensure_ascii=False, default=str)
                except Exception:
                    key = str(row)
                if key in seen_history:
                    continue
                seen_history.add(key)
                merged["history"].append(row)
    return merged


# ── plan 변경 이력 아카이브 ────────────────────────────────────────────────────
# 제품 JSON 안의 history 는 저장할 때마다 최근 1000건으로 잘린다(`[-1000:]`).
# "누가 언제 무엇을 어떻게 바꿨는지"를 나중에 되짚으려면 잘리면 안 되므로, 같은
# 엔트리를 append-only JSONL 로 한 벌 더 남긴다. JSON 쪽 창은 기존 소비자(plan
# risk payload 등) 호환을 위해 그대로 둔다.
PLAN_HISTORY_LOG_SUBDIR = "plan_history_log"
_PLAN_HISTORY_LOG_LOCK = threading.Lock()
_PLAN_HISTORY_FACET_COLUMN_CAP = 500


def _plan_history_log_path(product: str) -> Path:
    # PLAN_DIR 은 테스트가 monkeypatch 하는 지점이라 **호출 시점에** 읽는다.
    # 모듈 상수로 굳히면 PLAN_DIR 만 갈아끼운 테스트가 실제 데이터 루트에 쓴다.
    return PLAN_DIR / PLAN_HISTORY_LOG_SUBDIR / f"{_plan_product_name(product)}.jsonl"


def _plan_history_entry_key(entry: dict) -> str:
    return "".join(
        str(entry.get(k) if entry.get(k) is not None else "")
        for k in ("time", "user", "action", "cell", "old", "new")
    )


PLAN_REASON_MAX_LEN = 500


def _clean_plan_reason(reason: Any) -> str:
    """변경 사유. 선택 입력이라 비어 있는 게 정상이고, 길이만 잘라 둔다
    (이력 파일이 자유 텍스트로 무한정 커지지 않게)."""
    text = str(reason or "").strip()
    return text[:PLAN_REASON_MAX_LEN]


def _new_history_batch_id() -> str:
    """한 번의 저장/삭제 요청을 묶는 식별자 — 화면에서 '이 사람이 이때 한 작업'
    단위로 접어 볼 수 있게 한다."""
    return os.urandom(6).hex()


def _archive_plan_history(product: str, entries: list[dict],
                          prior_history: list[dict] | None = None) -> None:
    """새 이력을 아카이브에 덧붙인다. 아카이브가 아직 없으면 기존 JSON 이력으로 seed
    해서, 아카이브 도입 이전 기록도 한 화면에서 이어 보이게 한다."""
    if not entries:
        return
    try:
        from core.utils import jsonl_append
        path = _plan_history_log_path(product)
        with _PLAN_HISTORY_LOG_LOCK:
            if not path.exists():
                path.parent.mkdir(parents=True, exist_ok=True)
                for old_row in (prior_history or []):
                    if isinstance(old_row, dict):
                        jsonl_append(path, old_row, add_timestamp=False, max_lines=None)
            for row in entries:
                jsonl_append(path, row, add_timestamp=False, max_lines=None)
    except Exception as exc:
        logger.warning(f"plan history archive append failed for {product}: {exc}")


def _plan_history_entries(product: str) -> list[dict]:
    """JSON 창 + 아카이브를 합친 전체 이력 (시간 오름차순)."""
    inline = [h for h in (_load_plan_data(product).get("history") or [])
              if isinstance(h, dict)]
    try:
        from core.utils import jsonl_read
        archived = [h for h in jsonl_read(_plan_history_log_path(product), limit=0)
                    if isinstance(h, dict)]
    except Exception:
        archived = []
    if not archived:
        return inline
    seen = {_plan_history_entry_key(h) for h in archived}
    merged = archived + [h for h in inline if _plan_history_entry_key(h) not in seen]
    merged.sort(key=lambda h: str(h.get("time") or ""))
    return merged


def _history_cell_parts(entry: dict) -> tuple[str, str, str]:
    """엔트리에서 (root, wafer, column). 새 엔트리는 필드로 갖고 있고, 예전
    엔트리는 cell key 를 쪼개서 얻는다."""
    root, wafer, column = _split_plan_cell_key(entry.get("cell") or "")
    return (str(entry.get("root_lot_id") or root or ""),
            str(entry.get("wafer_id") or wafer or ""),
            str(entry.get("column") or column or ""))


def _history_matches_root(entry: dict, root_lot_id: str) -> bool:
    if not root_lot_id:
        return True
    return (entry.get("root_lot_id") == root_lot_id
            or str(entry.get("cell") or "").startswith(root_lot_id + "|"))


def _history_matches_filters(entry: dict, *, user: str = "", action: str = "",
                             column: str = "", wafer_id: str = "", q: str = "",
                             since: str = "", until: str = "",
                             has_reason: bool = False) -> bool:
    _root, wafer, col = _history_cell_parts(entry)
    if has_reason and not str(entry.get("reason") or "").strip():
        return False
    if user and str(entry.get("user") or "") != user:
        return False
    if action and str(entry.get("action") or "") != action:
        return False
    if column and column.lower() not in col.lower():
        return False
    if wafer_id and wafer_id.lower() not in wafer.lower():
        return False
    when = str(entry.get("time") or "")
    # ISO 문자열은 사전순 비교가 곧 시간순 비교다. until 은 그 날 전체를 포함한다.
    if since and when[:len(since)] < since:
        return False
    if until and when[:10] > until[:10]:
        return False
    if q:
        needle = q.lower()
        haystack = " ".join(str(entry.get(k) or "") for k in
                            ("user", "cell", "old", "new", "action", "reason"))
        if needle not in haystack.lower():
            return False
    return True


def _plan_history_facets(entries: list[dict]) -> dict:
    """필터 드롭다운 소스 — 필터 적용 전 집합에서 뽑아야 선택지가 사라지지 않는다."""
    users: set[str] = set()
    actions: set[str] = set()
    columns: set[str] = set()
    for entry in entries:
        if entry.get("user"):
            users.add(str(entry["user"]))
        if entry.get("action"):
            actions.add(str(entry["action"]))
        _root, _wafer, col = _history_cell_parts(entry)
        if col:
            columns.add(col)
    return {
        "users": sorted(users),
        "actions": sorted(actions),
        "columns": sorted(columns)[:_PLAN_HISTORY_FACET_COLUMN_CAP],
    }


# ── root 별 plan 인덱스 ────────────────────────────────────────────────────────
# 저장 파일은 그대로 두고(포맷 변경 없음, 마이그레이션 없음) **읽는 방법만** 바꾼다.
#
# 문제였던 것: JSON 파싱은 이미 load_json_cached 로 캐시돼 있었는데(0.04ms),
# _load_plan_data 가 요청마다 제품의 plan 전체를 새 dict 로 복사하고 history 를
# json.dumps 로 dedup 했다(3.5ms, 90배). 그런데 /view 가 실제로 쓰는 건 "이 root
# 의 plan" 뿐이고 history 는 아예 안 본다.
#
# 그래서 파일 시그니처가 그대로면 root 별로 쪼갠 인덱스를 재사용한다. 인덱스를
# 만드는 비용(전체 1회 순회)은 파일이 바뀔 때만 든다.
_PLAN_ROOT_INDEX_CACHE: dict[str, tuple[tuple, dict[str, dict]]] = {}
_PLAN_ROOT_INDEX_LOCK = threading.Lock()
_PLAN_ROOT_INDEX_MAX = 64

# 반환값을 수정하면 다음 요청이 오염된다 — 읽기 전용 계약이다(load_json_cached 와
# 같은 규칙). 편집 경로는 계속 _load_plan_data 로 자기 사본을 받는다.
_EMPTY_READONLY: dict = {}


def _split_key_root(key: str) -> str:
    """`root|wafer|column` 에서 root 만. 형식이 아니면 빈 문자열."""
    text = str(key or "")
    idx = text.find("|")
    return text[:idx] if idx > 0 else ""


def _build_root_index(values: dict) -> dict[str, dict]:
    """`root|...` 키 dict → {root: {원본키: 값}}."""
    index: dict[str, dict] = {}
    for key, value in values.items():
        root = _split_key_root(key)
        if not root:
            continue
        index.setdefault(root, {})[key] = value
    return index


def _plan_entries_for_root(product: str, root_lot_id: str) -> dict:
    """이 root 의 plan 엔트리만 (읽기 전용).

    반환 dict 의 키는 원본과 동일한 `root|wafer|column` 이라, 호출측은 기존
    `plans.get(f"{root}|{wafer}|{col}")` 코드를 그대로 쓴다.
    """
    root = str(root_lot_id or "").strip()
    if not root:
        return _EMPTY_READONLY
    paths = _plan_alias_paths(product)
    sig = _plan_risk_cache_sig(paths)
    cache_key = str(product or "")
    with _PLAN_ROOT_INDEX_LOCK:
        hit = _PLAN_ROOT_INDEX_CACHE.get(cache_key)
        if hit is not None and hit[0] == sig:
            return hit[1].get(root) or _EMPTY_READONLY

    # alias 병합은 _load_plan_data 와 같은 순서(뒤가 이김)를 유지한다.
    merged_plans: dict = {}
    for fp in paths:
        data = load_json_cached(fp, {}) if fp.exists() else {}
        if not isinstance(data, dict):
            continue
        plans = data.get("plans")
        if isinstance(plans, dict):
            merged_plans.update(plans)
    index = _build_root_index(merged_plans)

    with _PLAN_ROOT_INDEX_LOCK:
        if len(_PLAN_ROOT_INDEX_CACHE) >= _PLAN_ROOT_INDEX_MAX:
            _PLAN_ROOT_INDEX_CACHE.clear()
        _PLAN_ROOT_INDEX_CACHE[cache_key] = (sig, index)
    return index.get(root) or _EMPTY_READONLY


def _plan_risk_cache_key(product: str, include_deleted: bool) -> tuple[str, bool]:
    fp = _plan_history_path(product)
    try:
        return (str(fp.resolve()), bool(include_deleted))
    except Exception:
        return (str(fp), bool(include_deleted))


def _plan_risk_cache_sig(fp: Path | list[Path]) -> tuple:
    if isinstance(fp, list):
        return tuple(_plan_risk_cache_sig(p) for p in fp)
    try:
        st = fp.stat()
        return (str(fp.resolve()), st.st_mtime, st.st_size)
    except Exception:
        return (str(fp), 0.0, 0)


def _empty_plan_risk_payload(cache: bool = False) -> dict:
    return {"final": [], "drift": [], "drift_count": 0, "total_cells": 0, "cache": cache}


def _copy_plan_risk_payload(payload: dict, root_lot_id: str = "") -> dict:
    root = str(root_lot_id or "").strip()
    if root:
        by_root = payload.get("_by_root") if isinstance(payload.get("_by_root"), dict) else {}
        scoped = by_root.get(root) or {"final": [], "drift": []}
        final_rows = [dict(r) for r in (scoped.get("final") or [])]
        drift_rows = [dict(r) for r in (scoped.get("drift") or [])]
    else:
        final_rows = [dict(r) for r in (payload.get("final") or [])]
        drift_rows = [dict(r) for r in (payload.get("drift") or [])]
    return {
        "final": final_rows,
        "drift": drift_rows,
        "drift_count": len(drift_rows),
        "total_cells": len(final_rows),
        "cache": bool(payload.get("cache")),
        "cache_built_at": payload.get("cache_built_at", ""),
    }


def _build_plan_risk_payload(hist: list, include_deleted: bool = False) -> dict:
    per_cell: dict[str, list] = {}
    for h in hist or []:
        if not isinstance(h, dict):
            continue
        ck = h.get("cell")
        if not ck:
            continue
        per_cell.setdefault(str(ck), []).append(h)

    final_rows = []
    drift_rows = []
    for ck, entries in per_cell.items():
        entries.sort(key=lambda x: x.get("time", ""))
        last = entries[-1]
        action = last.get("action") or "set"
        if action == "delete" and not include_deleted:
            continue
        sets = [e for e in entries if (e.get("action") or "set") == "set"]
        distinct_values = list({e.get("new") for e in sets if e.get("new") is not None})
        distinct_users = list({e.get("user") for e in sets if e.get("user")})
        set_count = len(sets)
        delete_count = sum(1 for e in entries if e.get("action") == "delete")
        drift_flags = []
        if set_count >= 2 and len(distinct_values) >= 2:
            drift_flags.append("multi_change")
        if len(distinct_users) >= 2:
            drift_flags.append("multi_user")
        if delete_count >= 1 and set_count >= 1:
            drift_flags.append("reinstated")
        parts = (ck or "").split("|")
        lot = parts[0] if len(parts) > 0 else ""
        wf = parts[1] if len(parts) > 1 else ""
        col = parts[2] if len(parts) > 2 else ""
        row = {
            "cell": ck,
            "root_lot_id": lot,
            "wafer_id": wf,
            "column": col,
            "final_value": last.get("new"),
            "final_action": action,
            "final_user": last.get("user"),
            "final_time": last.get("time"),
            "set_count": set_count,
            "delete_count": delete_count,
            "distinct_values": distinct_values,
            "distinct_users": distinct_users,
            "drift": drift_flags,
            # 지금 값이 왜 이렇게 됐는지 = 마지막 변경의 사유. 그 변경에 사유가
            # 없으면 이 셀에 마지막으로 남은 사유로 폴백해 최소한의 맥락은 남긴다.
            "final_reason": _clean_plan_reason(last.get("reason")),
            "last_reason": next((_clean_plan_reason(e.get("reason"))
                                 for e in reversed(entries)
                                 if str(e.get("reason") or "").strip()), ""),
            "reason_count": sum(1 for e in entries if str(e.get("reason") or "").strip()),
        }
        final_rows.append(row)
        if drift_flags:
            drift_rows.append(row)

    final_rows.sort(key=lambda r: r.get("final_time") or "", reverse=True)
    drift_rows.sort(key=lambda r: r.get("final_time") or "", reverse=True)
    by_root: dict[str, dict[str, list]] = {}
    for row in final_rows:
        root = str(row.get("root_lot_id") or "").strip()
        if not root:
            continue
        bucket = by_root.setdefault(root, {"final": [], "drift": []})
        bucket["final"].append(row)
        if row.get("drift"):
            bucket["drift"].append(row)

    return {
        "final": final_rows,
        "drift": drift_rows,
        "drift_count": len(drift_rows),
        "total_cells": len(final_rows),
        "_by_root": by_root,
    }


def _get_plan_risk_payload(product: str, include_deleted: bool = False, force: bool = False) -> dict:
    paths = _plan_alias_paths(product)
    if not any(fp.exists() for fp in paths):
        return _empty_plan_risk_payload(cache=True)
    # 아카이브도 sig 에 넣는다 — 최종 Log 의 변경 횟수/사용자 수를 1000건 창이
    # 아니라 전체 이력에서 세기 때문에, 아카이브만 갱신돼도 무효화돼야 한다.
    sig = _plan_risk_cache_sig(paths + [_plan_history_log_path(product)])
    key = _plan_risk_cache_key(product, include_deleted)
    with _PLAN_RISK_CACHE_LOCK:
        cached = _PLAN_RISK_CACHE.get(key)
        if cached and not force and cached.get("_sig") == sig:
            return cached
    hist = _plan_history_entries(product)
    payload = _build_plan_risk_payload(hist if isinstance(hist, list) else [], include_deleted=include_deleted)
    payload.update({
        "_sig": sig,
        "cache": True,
        "cache_built_at": datetime.datetime.now().isoformat(timespec="seconds"),
    })
    with _PLAN_RISK_CACHE_LOCK:
        if len(_PLAN_RISK_CACHE) >= _PLAN_RISK_CACHE_MAX:
            try:
                _PLAN_RISK_CACHE.pop(next(iter(_PLAN_RISK_CACHE)))
            except Exception:
                _PLAN_RISK_CACHE.clear()
        _PLAN_RISK_CACHE[key] = payload
    return payload


def _invalidate_plan_risk_cache(product: str) -> None:
    if not product:
        return
    keys = {
        _plan_risk_cache_key(product, False),
        _plan_risk_cache_key(product, True),
    }
    with _PLAN_RISK_CACHE_LOCK:
        for key in keys:
            _PLAN_RISK_CACHE.pop(key, None)


def refresh_plan_risk_cache(product: str = "", force: bool = False) -> dict:
    products = [product] if str(product or "").strip() else []
    if not products:
        try:
            products = [p.get("name") for p in list_products().get("products", []) if p.get("name")]
        except Exception:
            products = []
    results = []
    for raw_product in products:
        fp = _plan_history_path(raw_product)
        if not fp.exists():
            results.append({"product": raw_product, "ok": True, "skipped": True, "reason": "no plan history"})
            continue
        try:
            payload = _get_plan_risk_payload(raw_product, include_deleted=False, force=force)
            results.append({
                "product": raw_product,
                "ok": True,
                "skipped": False,
                "total_cells": int(payload.get("total_cells") or 0),
                "drift_count": int(payload.get("drift_count") or 0),
            })
        except Exception as e:
            logger.warning("plan risk cache refresh failed (product=%s) %s: %s",
                           raw_product, type(e).__name__, e)
            results.append({"product": raw_product, "ok": False, "reason": f"{type(e).__name__}: {e}"})
    return {"ok": all(r.get("ok") for r in results) if results else True, "products": results}


def _candidate_values_from_frame(rows, value_col: str = "v", limit: int = 500) -> list[str]:
    """Return clean string autocomplete values from a collected Polars frame."""
    values: list[str] = []
    seen: set[str] = set()
    try:
        limit = max(1, int(limit or 500))
    except Exception:
        limit = 500
    if rows is None or value_col not in rows.columns:
        return values
    for value in rows[value_col].to_list():
        text = _clean_str(value)
        if not text:
            continue
        key = text.upper()
        if key in seen:
            continue
        seen.add(key)
        values.append(text)
        if len(values) >= limit:
            break
    return values


def _limited_unique_values(lf, col: str, prefix: str = "", limit: int = 500,
                           preview_only: bool = True) -> list[str]:
    """Return bounded autocomplete values without scanning broad empty-prefix lists.

    Empty dropdowns only need a preview.  Once a user types, prefix filtering must
    search the full source so values outside the preview are still discoverable.
    """
    try:
        limit = max(1, int(limit or 500))
    except Exception:
        limit = 500
    prefix = prefix if isinstance(prefix, str) else ""
    q = (
        lf.select(pl.col(col).cast(_STR, strict=False).alias("v"))
        .filter(pl.col("v").is_not_null())
    )
    if prefix.strip():
        q = q.filter(_contains_literal_ci_expr("v", prefix))
        rows = q.unique().sort("v").head(limit).collect()
    elif not preview_only:
        rows = q.unique().sort("v").head(limit).collect()
    else:
        sample_limit = max(limit, min(limit * 20, 10000))
        rows = q.head(sample_limit).unique(maintain_order=True).head(limit).collect()
    values = _candidate_values_from_frame(rows, "v", limit)
    return sorted(values, key=lambda s: str(s).upper())


def _split_table_cache_dir(product: str) -> Path:
    """Pre-pivoted split_table 캐시 디렉터리 (root_lot_id 당 parquet 1개).

    캐시 디렉터리는 canonical ML_TABLE_* 대문자 이름으로 저장된다 — raw product
    문자열로 찾으면 대소문자 구분 FS(운영 Linux)에서 조용히 빗나간다.
    """
    from app_v2.modules.splittable.cache_builder import canonical_product_dir

    base = PATHS.db_cache_dir if hasattr(PATHS, "db_cache_dir") else Path("data/cache")
    return base / "split_table" / (canonical_product_dir(product) or str(product or ""))


def _main_table_candidates(product: str, col: str = "root_lot_id", prefix: str = "",
                           limit: int = 500, root_lot_id: str = "") -> dict:
    """Return candidates from the actual SplitTable render source.

    FAB history can contain operational roots that are not present in the
    current ML_TABLE. Those roots are useful for lineage, but they produce an
    empty SplitTable view. Autocomplete should therefore prefer values that can
    actually render in /view.
    """
    try:
        limit = max(1, int(limit or 500))
    except Exception:
        limit = 500
        
    cache_key = (
        "main_table_candidates",
        _lot_lookup_cache_sig(product),
        str(product or "").strip(),
        str(col or "").strip(),
        str(prefix or "").strip(),
        str(root_lot_id or "").strip(),
        limit,
    )
    cached = _lot_lookup_cache_get(cache_key)
    if cached is not None:
        return cached

    def finish(payload: dict) -> dict:
        return _lot_lookup_cache_set(cache_key, payload)

    try:
        lookup_meta = {}
        
        split_table_cache_dir = _split_table_cache_dir(product)
        catalog_path = split_table_cache_dir / "_lot_catalog.json"
        if catalog_path.exists():
            try:
                import json
                with open(catalog_path, "r", encoding="utf-8") as f:
                    catalog = json.load(f)
                
                target_col = next((k for k in catalog.keys() if k.casefold() == str(col or "").casefold()), None)
                if not target_col and str(col or "").casefold() in {c.casefold() for c in _FAB_COL_CANDIDATES}:
                    target_col = next((k for k in catalog.keys() if k.casefold() == "fab_lot_id"), None)
                    if not target_col:
                        target_col = next((k for k in catalog.keys() if k.casefold() == "lot_id"), None)
                
                if target_col and target_col in catalog:
                    cands = catalog[target_col]
                    if prefix.strip():
                        cands = [c for c in cands if prefix.strip().upper() in str(c).upper()]
                    if cands:
                        return finish({
                            "col": col,
                            "candidates": cands[:limit],
                            "prefix": prefix,
                            "root_scope": "",
                            "match_mode": "pivot_catalog_fast",
                            "source": "pivot_catalog",
                            "fab_source": "",
                            "lookup_cache": lookup_meta,
                            "strict": False,
                        })
            except Exception as e:
                logger.warning(f"Failed to read _lot_catalog.json for {product}: {e}")

        if str(col or "").casefold() == "root_lot_id":
            lookup = _root_lot_lookup_cache_candidates(product, prefix=prefix, limit=limit)
            if lookup is not None:
                lookup_meta = _lookup_cache_public_meta(lookup)
                candidates = lookup.get("candidates") or []
                if candidates:
                    return finish({
                        "candidates": candidates,
                        "source_col": "root_lot_id",
                        "root_ids": candidates,
                        "match_mode": "lookup_cache_roots",
                        "lookup_cache": lookup_meta,
                    })
                if lookup.get("has_cache") and not lookup.get("source_stale") and prefix.strip():
                    # prefix 검색에서 fresh 캐시가 빈 결과면 그대로 신뢰한다
                    # (여기서 raw 폴백하면 키 입력마다 원천 전체 스캔이 된다).
                    return finish({
                        "candidates": [],
                        "source_col": "root_lot_id",
                        "root_ids": [],
                        "match_mode": "lookup_cache_roots",
                        "lookup_cache": lookup_meta,
                    })
                # 빈 prefix(초기 목록)인데 fresh 캐시가 비어 있으면 캐시가 잘못
                # 빌드된 회귀일 수 있으므로 아래 bounded raw preview 로 재확인한다.
                source_fp = lookup.get("source_fp")
                if source_fp and _split_view_should_defer_raw_fallback(source_fp) and (
                    not lookup.get("has_cache") or lookup.get("source_stale")
                ):
                    queued = _ml_table_lookup.enqueue_build(source_fp)
                    lookup_meta = _lookup_cache_public_meta(lookup, queued)
            # lookup 캐시로 못 채웠을 때만 split_table 캐시 디렉터리를 본다.
            # 이건 **이미 캐시된 root 만** 열거하므로 완전한 목록이 아니다 —
            # 예전엔 이 분기가 맨 앞에 있어서 캐시된 4개만 나오고 나머지 26개
            # root 는 제품 선택 후 드롭다운에서 아예 안 보였다.
            split_table_cache_dir = _split_table_cache_dir(product)
            if split_table_cache_dir.exists():
                cands = sorted(fp.stem for fp in split_table_cache_dir.glob("*.parquet"))
                if prefix.strip():
                    cands = [c for c in cands if prefix.strip().upper() in c.upper()]
                if cands:
                    return finish({
                        "col": "root_lot_id",
                        "candidates": cands[:limit],
                        "prefix": prefix,
                        "root_scope": "",
                        "match_mode": "split_table_cache_fast",
                        "source": "split_table_cache",
                        "fab_source": "",
                        "lookup_cache": lookup_meta,
                        "strict": False,
                    })
        lf = _scan_product_base(product)
        schema_names = lf.collect_schema().names()
        lot_col, _ = _detect_lot_wafer(lf, product)
        target = ""
        if str(col or "").casefold() == "root_lot_id":
            target = lot_col or _ci_resolve_in("root_lot_id", schema_names)
        elif str(col or "").casefold() in {c.casefold() for c in _FAB_COL_CANDIDATES}:
            target = (
                _ci_resolve_in("fab_lot_id", schema_names)
                or _ci_resolve_in("lot_id", schema_names)
                or _pick_first_present_ci(_FAB_COL_CANDIDATES, schema_names)
            )
        else:
            target = _ci_resolve_in(col, schema_names)
        if not target or target not in schema_names:
            return finish({"candidates": [], "source_col": target or col, "root_ids": []})

        root_scope = _clean_str(root_lot_id)
        if root_scope:
            root_col = lot_col or _ci_resolve_in("root_lot_id", schema_names)
            if root_col and root_col in schema_names:
                lf = lf.filter(_join_key_expr(root_col) == root_scope.upper())

        # 빈 prefix 드롭다운(초기 root_lot_id 목록)은 미리보기 앞부분 N개면 충분하다.
        # 전체 컬럼을 unique + sort 하면 큰 원천에서 수 초가 걸려 즉시 뜨지 않으므로,
        # 앞부분만 샘플링해 즉시 응답한다. 사용자가 입력하면 _limited_unique_values 가
        # prefix 로 원천 전체를 서버에서 필터링하므로 미리보기 밖 값도 검색된다.
        # (root_scope 가 지정된 fab_lot_id 조회는 이미 좁혀진 집합이라 전체 스캔해도 빠르다.)
        preview_only = not bool(root_scope)
        values = _limited_unique_values(lf, target, prefix=prefix, limit=limit,
                                        preview_only=preview_only)
        payload = {"candidates": values, "source_col": target, "root_ids": values if str(col or "").casefold() == "root_lot_id" else []}
        if str(col or "").casefold() == "root_lot_id":
            payload["match_mode"] = "splittable_roots"
            if lookup_meta:
                payload["lookup_cache"] = lookup_meta
        return finish(payload)
    except Exception as e:
        logger.warning("_main_table_candidates 실패 (product=%s col=%s) %s: %s",
                       product, col, type(e).__name__, e)
        return finish({"candidates": [], "source_col": col, "root_ids": []})


def _scan_product(
    product: str,
    root_lot_id: str = "",
    fab_lot_id: str = "",
    wafer_ids: str = "",
    base_lf=None,
    runtime_profile: dict | None = None,
):
    """Scan ML_TABLE_<PROD>.parquet + hive override join.

    v8.8.26: 실패 경로마다 logger.warning 로 가시화 (이전 blanket except 제거).
      - CI align 이후 fab schema 를 **재조회** 해서 rename 이 실제로 적용됐는지 확인.
      - override_cols 가 join_keys 만 남으면 경고 후 raw lf 반환.
    """
    product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    lf = base_lf
    if lf is not None and runtime_profile is not None:
        runtime_profile["root_cache_hit"] = True
    if lf is None:
        lf = _scan_product_base_lookup_cache(
            product,
            root_lot_id=root_lot_id,
            wafer_ids=wafer_ids,
            runtime_profile=runtime_profile,
        )
    if lf is None:
        lf = _scan_product_base(product)

    # v8.8.3: 오버라이드 로직 근본 재정리.
    #   1) 매뉴얼 config(lot_overrides[product].fab_source) 가 있으면 그 값을 사용.
    #   2) 없으면 ML_TABLE_<PROD> → DB/<root>/<PROD> 자동 매칭 시도.
    #   3) ts_col / fab_col 도 매뉴얼 > 자동 추론 순.
    #   4) 조인은 항상 "ts_col 기준 최신 레코드만" join keys 별로 picking 후 left-join.
    try:
        product, ov, fab_source = _current_fab_override(product)
        include_all = _foreground_global_fab_scan_enabled()

        try:
            main_names_list = lf.collect_schema().names()
        except Exception as e:
            logger.warning("_scan_product: main schema 조회 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
            return lf
        if root_lot_id or wafer_ids:
            try:
                main_lot_col, main_wf_col = _detect_lot_wafer(lf, product)
                lf = _filter_lot_wafer(
                    lf, main_lot_col, main_wf_col,
                    root_lot_id=root_lot_id,
                    wafer_ids=wafer_ids,
                )
            except Exception as e:
                logger.warning("_scan_product: main scope filter 실패 (product=%s root=%s wafer=%s) %s: %s",
                               product, root_lot_id, wafer_ids, type(e).__name__, e)

        cached = _latest_lot_progress_projection(
            product, main_names_list,
            root_lot_id=root_lot_id,
            fab_lot_id=fab_lot_id,
            wafer_ids=wafer_ids,
        )
        if cached:
            return _join_fab_projection_into_main(
                lf, set(main_names_list), cached["lf"],
                cached["join_keys"], cached["override_cols"],
                fab_has_join_tmp=True,
            )

        if not fab_source and not _global_fab_source_paths("", include_all=include_all):
            return _strip_non_authoritative_fab_fields(lf, product)

        # Fast layer: read only the searched root's FAB partition from the
        # precomputed per-root index, bounding the latest-lot pick to O(one root)
        # instead of scanning the whole FAB source. Additive — a miss returns None
        # and we fall back to the full scan below while a build is scheduled.
        fab_lf = None
        fab_sources: list = []
        if str(root_lot_id or "").strip():
            fab_lf = _fab_lot_index_scan_root(
                product, root_lot_id, fab_source=fab_source, include_all=include_all)
            if fab_lf is not None:
                fab_sources = ["<fab_lot_index>"]
        if fab_lf is None:
            if str(root_lot_id or "").strip():
                enqueued = _enqueue_fab_lot_index_build(
                    product, fab_source, include_all=include_all, reason="scan_miss")
                canonical_product = (
                    _canonical_mltable_product_name(product, allow_bare=True)
                    or str(product or "").strip().upper()
                )
                with _FAB_IDX_BUILD_LOCK:
                    fab_index_running = canonical_product in _FAB_IDX_BUILD_INPROGRESS
                runtime_profile = runtime_profile if runtime_profile is not None else {}
                runtime_profile["fab_index_queued"] = bool(
                    enqueued or fab_index_running
                )
                runtime_profile["cache_incomplete"] = True
                # Never scan a multi-GB FAB tree in an interactive root request.
                # The main ML table is still immediately useful; the worker-built
                # per-root FAB index will refresh labels on the next UI poll.
                if not _env_bool("FLOW_SPLITTABLE_INTERACTIVE_FAB_RAW_FALLBACK", False):
                    return _strip_non_authoritative_fab_fields(lf, product)
            fab_lf, fab_sources = _scan_global_fab_sources(fab_source, include_all=include_all)
        if fab_lf is None:
            logger.warning("_scan_product: FAB source scan 실패 (product=%s fab_source=%s sources=%s)",
                           product, fab_source, fab_sources)
            return _strip_non_authoritative_fab_fields(lf, product)

        # v8.8.22: CI 정렬 — fab_lf 컬럼명을 main 쪽 casing 으로 rename.
        #   ex) ML_TABLE 의 ROOT_LOT_ID ↔ hive root_lot_id → join 성공.
        fab_lf, _ = _ci_align_fab_to_main(fab_lf, main_names_list)
        # v8.8.26: rename 이 silently 실패할 수 있으므로 schema 를 재조회 — 신뢰 가능한 true state.
        try:
            fab_schema_names = fab_lf.collect_schema().names()
        except Exception as e:
            logger.warning("_scan_product: fab post-align schema 조회 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
            return lf
        main_names = set(main_names_list)
        fab_names = set(fab_schema_names)

        join_keys = ov.get("join_keys") or []
        if isinstance(join_keys, str):
            join_keys = [k.strip() for k in join_keys.split(",") if k.strip()]
        if join_keys:
            mapped = []
            for k in join_keys:
                actual = _ci_resolve_in(k, main_names_list) or _resolve_source_col_name(k, fab_schema_names)
                if actual:
                    mapped.append(actual)
            join_keys = mapped
        if not join_keys:
            join_keys = _default_override_join_keys(main_names_list, fab_schema_names)
        join_keys = [k for k in join_keys if k in main_names and k in fab_names]
        if not join_keys:
            logger.warning(
                "_scan_product: 공통 join key 없음 (product=%s fab_source=%s main=%s fab=%s)",
                product, fab_source, main_names_list[:20], fab_schema_names[:20],
            )
            return lf

        fc_raw = (ov.get("fab_col") or "").strip()
        fab_col = (_resolve_source_col_name(fc_raw, fab_schema_names) if fc_raw else "") \
                  or _pick_first_present_ci(_FAB_COL_CANDIDATES, fab_schema_names)
        if not fab_col:
            fab_col = "fab_lot_id"
        tc_raw = (ov.get("ts_col") or "").strip()
        ts_col = (_resolve_source_col_name(tc_raw, fab_schema_names) if tc_raw else "") \
                 or _pick_ts_col(fab_schema_names)
        fab_lf = _apply_fab_scope_filters(
            fab_lf, fab_schema_names, ov,
            root_lot_id=root_lot_id,
            fab_lot_id=fab_lot_id,
            wafer_ids=wafer_ids,
            fab_col=fab_col,
        )

        raw_oc = ov.get("override_cols")
        if isinstance(raw_oc, str):
            raw_oc = [c.strip() for c in raw_oc.split(",") if c.strip()]
        if not raw_oc:
            raw_oc = list(_DEFAULT_OVERRIDE_COLS)
        if fab_col and fab_col not in raw_oc:
            raw_oc = list(raw_oc) + [fab_col]
        resolved_oc = []
        for c in raw_oc:
            actual = _resolve_source_col_name(c, fab_schema_names)
            resolved_oc.append(actual or c)
        override_cols = [c for c in dict.fromkeys(resolved_oc)
                         if c in fab_names and c not in join_keys]
        wanted = list(dict.fromkeys(join_keys + override_cols + ([ts_col] if ts_col else [])))
        wanted = [c for c in wanted if c in fab_names]
        if not override_cols:
            logger.warning(
                "_scan_product: override_cols 가 비어있음 — join 없이 raw lf 반환 "
                "(product=%s fab_source=%s raw_oc=%s fab_names=%s)",
                product, fab_source, raw_oc, fab_schema_names[:20],
            )
            return lf

        fab_proj = fab_lf.select(wanted)
        join_aliases = [(k, f"__join_key_{i}") for i, k in enumerate(join_keys)]
        fab_proj = fab_proj.with_columns([_join_key_expr(k).alias(tmp) for k, tmp in join_aliases])
        lf = lf.with_columns([_join_key_expr(k).alias(tmp) for k, tmp in join_aliases])
        join_tmp_keys = [tmp for _, tmp in join_aliases]
        fab_proj = fab_proj.select(list(dict.fromkeys(join_tmp_keys + override_cols + ([ts_col] if ts_col else []))))
        if ts_col and ts_col in fab_names:
            fab_proj = fab_proj.sort(ts_col, descending=True, nulls_last=True)
            fab_proj = fab_proj.unique(subset=join_tmp_keys, keep="first", maintain_order=True)
        else:
            fab_proj = fab_proj.unique(subset=join_tmp_keys, keep="last")
        return _join_fab_projection_into_main(
            lf, main_names, fab_proj, join_keys, override_cols,
            fab_has_join_tmp=True,
        )
    except Exception as e:
        # v8.8.26: blanket except 유지하되 반드시 로그를 남겨 진단 가능하게.
        logger.warning("_scan_product: 예상치 못한 예외 (product=%s) %s: %s",
                       product, type(e).__name__, e, exc_info=True)
        return _strip_non_authoritative_fab_fields(lf, product)

# ── 제품별 root_lot_id 풀 (드롭다운 즉시 표시) ────────────────────────────
# 프런트(My_SplitTable)는 제품을 고를 때 root_lot_id 전체 목록을 한 번 받아
# lotPoolRef 에 담고 이후 키 입력은 로컬 필터로 처리한다. 그 "한 번"이 느렸다:
# split_table 캐시 디렉터리 glob + FAB latest 캐시 unique 스캔을 매번 다시 했고,
# 그 위의 _LOT_LOOKUP_CACHE 는 TTL 60초 + 256개 FIFO 라 1분마다 콜드였다.
#
# 여기서는 소스 시그니처가 바뀔 때까지 유효한 풀을 만들어 core.lot_list_cache
# (RAM LRU + 디스크 JSON) 에 맡긴다. 목록 조립 자체는 아래 빌더가 그대로
# 기존 경로를 쓰므로 결과가 달라지지 않는다 — 다시 계산하는 횟수만 줄어든다.
_ROOT_LOT_POOL_MAX = 50000   # 프런트 ROOT_LOT_CACHE_LIMIT_MAX 와 동일
_FAB_LOT_POOL_MAX = 50000

def _root_lot_pool_sig(product: str) -> str:
    """Fingerprint only the authoritative ML_TABLE lookup candidate source.

    FAB latest/match/pivot caches do not change which roots the ML_TABLE can
    render.  Including their mtimes made the pool miss repeatedly while a cache
    job was steadily writing partitions — exactly when a newly added product
    needed fast candidates.  The ML_TABLE file and lookup meta already change
    atomically when the complete candidate index is rebuilt.
    """
    try:
        source_fp = _product_path(product)
        parts = (
            _path_cache_sig(source_fp),
            _path_cache_sig(_ml_table_lookup.meta_path_for(source_fp)),
            _path_cache_sig(_ml_table_lookup.candidate_index_path_for(source_fp)),
        )
    except Exception:
        parts = ((str(product or ""), 0.0, 0),)
    return hashlib.sha1(repr(parts).encode("utf-8", "ignore")).hexdigest()


def _fab_lot_pool_sig(product: str) -> str:
    """제품 전체 LOT_ID 후보에 실제로 영향을 주는 파생 캐시 지문."""
    try:
        source_fp = _product_path(product)
        parts = (
            _path_cache_sig(source_fp),
            _path_cache_sig(_ml_table_lookup.meta_path_for(source_fp)),
            _path_cache_sig(_ml_table_lookup.candidate_index_path_for(source_fp)),
            _path_cache_sig(_latest_lot_step_cache_path()),
            _path_cache_sig(_match_cache_path(product)),
        )
    except Exception:
        parts = ((str(product or ""), 0.0, 0),)
    return hashlib.sha1(repr(parts).encode("utf-8", "ignore")).hexdigest()


def _invalidate_root_lot_pool(product: str = "") -> int:
    """root_lot 풀 캐시를 버린다 (버린 항목 수 반환).

    소스가 바뀌면 시그니처가 달라져 어차피 miss 지만, 캐시 재빌드 직후처럼
    "지금 바로 비워졌음"을 보장해야 하는 지점에서 명시적으로 부른다.
    """
    key = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    _root_lot_provisional_drop(key)
    return _lot_list_cache.invalidate(key)


def _build_root_lot_pool(product: str) -> tuple[list[str], dict, bool]:
    """완성된 lookup 후보 인덱스만 읽어 제품 root 목록을 게시한다.

    이 함수는 사용자 요청 경로에서도 호출된다. 따라서 lookup 미스 때 ML_TABLE
    원천이나 FAB history를 동기 스캔하지 않고 빌드만 큐에 넣은 뒤 즉시 준비 중
    응답을 돌려준다. lookup 빌드 완료 콜백/사전 예열기가 같은 함수를 다시 불러
    완전한 목록을 RAM+디스크 캐시에 게시한다.
    """
    lookup = _root_lot_lookup_cache_candidates(
        product, prefix="", limit=_ROOT_LOT_POOL_MAX)
    if lookup is None:
        return [], {
            "match_mode": "lookup_cache_unavailable",
            "source": "mltable_lookup",
            "lookup_cache": {},
        }, False

    candidates = _merge_candidate_values(
        lookup.get("candidates") or [], limit=_ROOT_LOT_POOL_MAX)
    fresh = bool(lookup.get("has_cache") and not lookup.get("source_stale"))
    index_ready = bool(fresh and lookup.get("candidate_index"))
    total_count = int(lookup.get("root_lot_id_count") or len(candidates))
    truncated = bool(total_count > len(candidates))
    # ``complete``는 이 API가 약속한 최대 50,000개 풀이 준비됐다는 뜻이다.
    # 종전에는 len == 50,000이면 인덱스가 완성돼도 영구 미완성으로 판정해 같은
    # lookup 빌드를 계속 큐에 넣었다. 원천 전체 포함 여부는 truncated로 분리한다.
    complete = index_ready
    queued = {}
    if not complete:
        source_fp = lookup.get("source_fp")
        if source_fp:
            queued = _ml_table_lookup.enqueue_build(source_fp)
    lookup_meta = _lookup_cache_public_meta(lookup, queued)
    # candidate index 가 아직 없어도 lookup 파티션 디렉터리(root_lot_id=…)는 이미
    # 그 시점의 root 를 전부 들고 있다. 종전에는 그걸 읽어 놓고 `complete` 가
    # 아니라는 이유로 버려서, 인덱스가 나올 때까지 드롭다운이 비어 있었다 —
    # 목록이 늦게 뜬다는 체감의 직접 원인. 잠정 목록으로 바로 내려주고
    # (디스크/RAM 에 굳히지는 않는다) 완성본이 나오면 프런트가 교체한다.
    provisional = bool(not complete and candidates)
    if complete:
        match_mode = "lookup_cache_roots"
    elif provisional:
        match_mode = "lookup_cache_partial"
    else:
        match_mode = "lookup_cache_preparing"
    meta = {
        "match_mode": match_mode,
        "source": "mltable_lookup",
        "fab_source": "",
        "provisional": provisional,
        "truncated": truncated,
        "total_count": total_count,
        "lookup_cache": lookup_meta,
    }
    return candidates, meta, complete


# 잠정(빌드 중) root 목록 — TTL 짧게, 제품 수만큼만. 완성본은 `_lot_list_cache`
# 로 가고 여기는 절대 디스크에 안 남는다.
_ROOT_LOT_PROVISIONAL_TTL_SEC = 10.0
_ROOT_LOT_PROVISIONAL_MAX = 16
_ROOT_LOT_PROVISIONAL_LOCK = threading.Lock()
_ROOT_LOT_PROVISIONAL: OrderedDict[str, tuple[float, dict]] = OrderedDict()


def _root_lot_provisional_get(product: str) -> dict | None:
    now = time.monotonic()
    with _ROOT_LOT_PROVISIONAL_LOCK:
        hit = _ROOT_LOT_PROVISIONAL.get(product)
        if hit is None:
            return None
        if now - hit[0] > _ROOT_LOT_PROVISIONAL_TTL_SEC:
            _ROOT_LOT_PROVISIONAL.pop(product, None)
            return None
        _ROOT_LOT_PROVISIONAL.move_to_end(product)
        return dict(hit[1])


def _root_lot_provisional_put(product: str, payload: dict) -> None:
    with _ROOT_LOT_PROVISIONAL_LOCK:
        _ROOT_LOT_PROVISIONAL[product] = (time.monotonic(), dict(payload))
        _ROOT_LOT_PROVISIONAL.move_to_end(product)
        while len(_ROOT_LOT_PROVISIONAL) > _ROOT_LOT_PROVISIONAL_MAX:
            _ROOT_LOT_PROVISIONAL.popitem(last=False)


def _root_lot_provisional_drop(product: str = "") -> None:
    with _ROOT_LOT_PROVISIONAL_LOCK:
        if product:
            _ROOT_LOT_PROVISIONAL.pop(product, None)
        else:
            _ROOT_LOT_PROVISIONAL.clear()


def _root_lot_pool(product: str) -> dict:
    """캐시된 root_lot_id 풀. 미스면 빌드해서 채운다.

    반환: `{"values", "meta", "complete", "cached"}`. `complete` 가 False 면
    (lookup 캐시 빌드 중이거나 raw 미리보기 폴백) **캐시하지 않는다** — 부분
    목록을 굳혀두면 빌드가 끝나도 잘린 드롭다운이 남고, 프런트의 재폴링
    프로토콜(`lookup_cache.queued`)도 깨진다. 호출자는 기존 경로로 넘어간다.

    시그니처를 **빌드보다 먼저** 읽는 순서가 중요하다. split_table 캐시
    디렉터리는 백그라운드로 root 파일이 하나씩 늘어나므로, 목록을 먼저 읽고
    나중에 stat 하면 "더 많이 본 목록"을 "더 새로운 지문"으로 굳혀 다음
    조회가 계속 hit 한다. 지금 순서라면 빌드 중 스냅샷을 담더라도 그 뒤에
    추가된 파일이 디렉터리 mtime 을 밀어 다음 조회에서 miss → 재빌드다.
    """
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    sig = _root_lot_pool_sig(canonical)
    cached = _lot_list_cache.get(canonical, sig)
    if cached is not None and cached.get("values") and cached.get("complete"):
        return cached
    provisional = _root_lot_provisional_get(canonical)
    if provisional is not None:
        return provisional
    values, meta, complete = _build_root_lot_pool(canonical)
    if not values or not complete:
        out = {"values": values, "meta": meta, "complete": False, "cached": ""}
        if values:
            # 빌드가 끝날 때까지 프런트가 2초마다 재확인한다. 그 사이 매번
            # 파티션 디렉터리를 다시 훑지 않도록 짧게만 들고 있는다.
            _root_lot_provisional_put(canonical, out)
        return out
    _root_lot_provisional_drop(canonical)
    return _lot_list_cache.put(canonical, sig, values, meta=meta, complete=True)


def _build_fab_lot_pool(product: str) -> tuple[list[str], dict, bool]:
    """ML_TABLE + canonical latest cache의 제품 전체 LOT_ID 목록을 사전 계산한다.

    원천 FAB tree는 여기서 절대 스캔하지 않는다. lookup 빌드가 수집한 ML_TABLE
    식별자와 narrow latest/match cache만 읽으므로 백그라운드 예열에 안전하고,
    둘 중 하나라도 준비되지 않았으면 부분 결과를 디스크에 굳히지 않는다.
    """
    main_values: list[str] = []
    main_ready = False
    main_complete = False
    main_source_col = ""
    main_stale = False
    needs_candidate_rebuild = False
    try:
        fp = _product_path(product)
        status = _ml_table_lookup.cache_status(fp)
        main_stale = bool(status.get("has_cache") and status.get("status") != "fresh")
        meta_schema = (status.get("meta") or {}).get("schema") or {}
        schema_names = list(meta_schema) if isinstance(meta_schema, dict) else []
        fab_col = next((name for name in schema_names if str(name).casefold() == "fab_lot_id"), "")
        lot_col = next((name for name in schema_names if str(name).casefold() == "lot_id"), "")
        target = fab_col or lot_col
        if status.get("status") == "fresh" and not target:
            # fresh schema에 두 컬럼이 없다는 것 자체가 완전한 빈 ML 후보다.
            main_ready = True
            main_complete = True
        elif target:
            indexed = _ml_table_lookup.candidate_values_from_lookup_cache(
                fp, target, limit=_FAB_LOT_POOL_MAX + 1, allow_stale=True)
            main_values = list(indexed.get("values") or [])
            main_ready = bool(indexed.get("available"))
            main_complete = bool(indexed.get("available") and indexed.get("complete"))
            main_source_col = str(indexed.get("source_column") or target)
            needs_candidate_rebuild = not main_ready
        if main_stale or needs_candidate_rebuild:
            _ml_table_lookup.enqueue_build(fp)
    except Exception as exc:
        logger.debug("fab lot ML candidate index read failed (%s): %s", product, exc)

    history_values: list[str] = []
    history_ready = False
    history_complete = False
    history_source = ""
    try:
        history_lf = _latest_lot_step_cache_lf(product)
        history_col = "lot_id"
        if history_lf is None:
            current = _match_cache_current(product)
            history_lf = current.get("lf") if current else None
            history_col = MATCH_CACHE_FAB_COL
            history_source = str((current or {}).get("fab_source") or "match_cache")
        else:
            history_source = _latest_lot_step_cache_source(product)
        if history_lf is not None:
            names = history_lf.collect_schema().names()
            if history_col in names:
                rows = (
                    history_lf.select(pl.col(history_col).cast(_STR, strict=False).alias("v"))
                    .filter(pl.col("v").is_not_null() & (pl.col("v") != ""))
                    .unique().sort("v").head(_FAB_LOT_POOL_MAX + 1).collect()
                )
                history_values = _candidate_values_from_frame(
                    rows, "v", _FAB_LOT_POOL_MAX + 1)
                history_complete = len(history_values) <= _FAB_LOT_POOL_MAX
                history_values = history_values[:_FAB_LOT_POOL_MAX]
                history_ready = True
            else:
                history_ready = True
                history_complete = True
    except Exception as exc:
        logger.debug("fab lot latest candidate cache read failed (%s): %s", product, exc)

    merged_probe = _merge_candidate_values(
        main_values, history_values, limit=_FAB_LOT_POOL_MAX + 1)
    merged_truncated = len(merged_probe) > _FAB_LOT_POOL_MAX
    merged = merged_probe[:_FAB_LOT_POOL_MAX]
    exhaustive = bool(main_complete and history_complete and not merged_truncated)
    # 두 사전계산 소스를 정상적으로 읽었다면 50,000개 bounded pool은 캐시할 수
    # 있다. 원천 후보가 상한을 넘었다는 이유로 매 요청마다 다시 조립하지 않는다.
    ready = bool(main_ready and history_ready)
    return merged, {
        "match_mode": "cached_fab_lot_pool",
        "source": "mltable+latest_cache",
        "source_col": main_source_col,
        "fab_source": history_source,
        "truncated": not exhaustive,
        "exhaustive": exhaustive,
        "source_stale": main_stale,
    }, ready


def _fab_lot_pool(product: str) -> dict:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    sig = _fab_lot_pool_sig(canonical)
    cached = _lot_list_cache.get(canonical, sig, kind="fab")
    if cached is not None and cached.get("values") and cached.get("complete"):
        return cached
    values, meta, complete = _build_fab_lot_pool(canonical)
    if not values or not complete:
        return {"values": values, "meta": meta, "complete": False, "cached": ""}
    return _lot_list_cache.put(
        canonical, sig, values, kind="fab", meta=meta, complete=True)


def _refresh_root_lot_pool_after_lookup_build(fp: Path) -> None:
    """Publish complete candidate pools before the first user opens them."""
    product = Path(fp).stem
    try:
        _invalidate_root_lot_pool(product)
        pool = _root_lot_pool(product)
        fab_pool = _fab_lot_pool(product)
        logger.info("lot candidate pools warmed product=%s roots=%s lots=%s",
                    product, len(pool.get("values") or []),
                    len(fab_pool.get("values") or []))
    except Exception as exc:
        # Lookup cache itself remains valid; a request can retry pool creation.
        logger.warning("root lot candidate pool warm failed product=%s %s: %s",
                       product, type(exc).__name__, exc)


# 제품 선택 전에 root/LOT 후보 디스크 사본을 만들어 두는 경량 스케줄러. lookup
# 자체의 무거운 계산은 ml_table_lookup worker가 담당하고, 이 루프는 완성된 인덱스와
# narrow latest cache를 합쳐 게시만 한다.
_CANDIDATE_PREWARM_STOP = threading.Event()
_CANDIDATE_PREWARM_THREAD: threading.Thread | None = None
_CANDIDATE_PREWARM_STARTED = False
_CANDIDATE_PREWARM_LAST: dict = {"at": "", "products": 0, "roots": 0, "lots": 0, "failed": 0}


def _candidate_list_prewarm_once() -> dict:
    products = []
    try:
        products = [
            row.get("name") for row in (list_products().get("products") or [])
            if isinstance(row, dict) and row.get("name")
        ]
    except Exception:
        products = []
    roots = lots = failed = 0
    for product in products:
        if _CANDIDATE_PREWARM_STOP.is_set():
            break
        try:
            root_pool = _root_lot_pool(product)
            fab_pool = _fab_lot_pool(product)
            roots += len(root_pool.get("values") or [])
            lots += len(fab_pool.get("values") or [])
        except Exception as exc:
            failed += 1
            logger.debug("candidate list prewarm failed (%s): %s", product, exc)
    result = {
        "at": datetime.datetime.now().isoformat(timespec="seconds"),
        "products": len(products), "roots": roots, "lots": lots, "failed": failed,
    }
    _CANDIDATE_PREWARM_LAST.update(result)
    return result


def _candidate_list_prewarm_loop() -> None:
    start_delay = max(1.0, _env_float("FLOW_SPLITTABLE_LIST_PREWARM_START_DELAY_SEC", 30.0))
    interval = max(60.0, _env_float("FLOW_SPLITTABLE_LIST_PREWARM_INTERVAL_SEC", 600.0))
    if _CANDIDATE_PREWARM_STOP.wait(start_delay):
        return
    while not _CANDIDATE_PREWARM_STOP.is_set():
        try:
            _candidate_list_prewarm_once()
        except Exception as exc:
            logger.warning("candidate list prewarm loop 오류: %s", exc)
        if _CANDIDATE_PREWARM_STOP.wait(interval):
            return


def start_candidate_list_prewarmer() -> bool:
    global _CANDIDATE_PREWARM_THREAD, _CANDIDATE_PREWARM_STARTED
    if _CANDIDATE_PREWARM_STARTED:
        return False
    if str(os.environ.get("FLOW_SPLITTABLE_LIST_PREWARM", "1")).strip().lower() in {"0", "false", "no", "off"}:
        return False
    _CANDIDATE_PREWARM_STOP.clear()
    _CANDIDATE_PREWARM_THREAD = threading.Thread(
        target=_candidate_list_prewarm_loop,
        name="splittable-list-prewarm",
        daemon=True,
    )
    _CANDIDATE_PREWARM_THREAD.start()
    _CANDIDATE_PREWARM_STARTED = True
    logger.info("SplitTable candidate list prewarmer started")
    return True


@router.get("/lot-ids")
def get_lot_ids(product: str = Query(...), limit: int = Query(200)):
    # 여기는 의도적으로 root_lot 풀을 쓰지 않는다. /lot-ids 는 "렌더 가능한 root
    # 의 authoritative 폴백" 이고, 풀의 가장 빠른 소스인 split_table 캐시
    # 디렉터리는 **이미 pre-pivot 된 root 만** 담는 부분집합이라 여기에 쓰면
    # 드롭다운이 조용히 잘린다. 프런트도 /lot-candidates 가 빈손일 때만 이걸
    # 부르므로 — 즉 풀이 없을 때만 — 캐시로 얻을 게 없다.
    catalog_path = _split_table_cache_dir(product) / "_lot_catalog.json"
    if catalog_path.exists():
        try:
            import json
            with open(catalog_path, "r", encoding="utf-8") as f:
                catalog = json.load(f)
            target_col = next((k for k in catalog.keys() if k.casefold() == "root_lot_id"), None)
            if target_col and target_col in catalog:
                return {
                    "lot_col": "root_lot_id",
                    "lot_ids": catalog[target_col][:limit],
                    "fallback": "",
                    "fab_source": "pivot_catalog",
                    "lookup_cache": {},
                }
        except Exception as e:
            logger.warning(f"Failed to read _lot_catalog.json in /lot-ids for {product}: {e}")

    lookup = _root_lot_lookup_cache_candidates(product, prefix="", limit=limit)
    lookup_meta = _lookup_cache_public_meta(lookup) if lookup is not None else {}
    if lookup is not None and lookup.get("candidates"):
        lots_list = _merge_candidate_values(lookup.get("candidates") or [], limit=limit)
        fab_source = ""
        try:
            hist = _fab_history_root_candidates(product, limit=limit)
            fab_source = hist.get("source") or ""
            fab_roots = hist.get("candidates") or []
            if fab_roots:
                main_keys = {str(v).upper() for v in lots_list}
                lots_list = _merge_candidate_values(
                    lots_list,
                    [v for v in fab_roots if str(v).upper() in main_keys],
                    limit=limit,
                )
        except Exception as e:
            logger.warning("/lot-ids: FAB root 후보 조회 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
        return {
            "lot_col": "root_lot_id",
            "lot_ids": lots_list,
            "fallback": "",
            "fab_source": fab_source,
            "lookup_cache": lookup_meta,
        }
    if lookup is not None:
        source_fp = lookup.get("source_fp")
        if source_fp and _split_view_should_defer_raw_fallback(source_fp) and (
            not lookup.get("has_cache") or lookup.get("source_stale")
        ):
            queued = _ml_table_lookup.enqueue_build(source_fp)
            lookup_meta = _lookup_cache_public_meta(lookup, queued)
    lf = _scan_product(product)
    lot_col, _ = _detect_lot_wafer(lf)
    lots_list: list = []
    fallback_used = False
    try:
        # /lot-ids 는 렌더 가능한 root 의 authoritative 폴백 목록이라 완전성이 계약이다
        # (검색은 /lot-candidates?prefix= 가 담당). 초기 목록 즉시성은 주 경로인
        # _main_table_candidates(Tier A split_table 캐시)와 FE 재폴링이 담당하므로
        # 여기서는 전체 스캔을 유지한다. 이 경로는 lookup 캐시 미스에서만 도달한다.
        lots_list = _limited_unique_values(lf, lot_col, limit=limit, preview_only=False)
    except Exception as e:
        logger.warning("/lot-ids: main lf 조회 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
        lots_list = []
    fab_roots: list[str] = []
    fab_source = ""
    try:
        hist = _fab_history_root_candidates(product, limit=limit)
        fab_roots = hist.get("candidates") or []
        fab_source = hist.get("source") or ""
    except Exception as e:
        logger.warning("/lot-ids: FAB root 후보 조회 실패 (product=%s) %s: %s",
                       product, type(e).__name__, e)
    if fab_roots:
        # Keep the dropdown aligned with what /view can render.  If ML_TABLE has
        # roots, only append FAB roots that are also present there; otherwise a
        # user can pick a valid FAB history root and still get an empty table.
        if lots_list:
            main_keys = {str(v).upper() for v in lots_list}
            fab_roots = [v for v in fab_roots if str(v).upper() in main_keys]
            lots_list = _merge_candidate_values(lots_list, fab_roots, limit=limit)
        else:
            lots_list = _merge_candidate_values(fab_roots, limit=limit)
            fallback_used = True
    # v8.8.26: main 이 all-null 이거나 비어있으면 override fab_source 로 폴백.
    if not lots_list:
        try:
            meta = _resolve_override_meta(product, include_diagnostics=False)
            fab_source = (meta.get("fab_source") or "").strip()
            if fab_source and not meta.get("error"):
                fab_lf = _scan_fab_source(fab_source)
                if fab_lf is not None:
                    fab_names = fab_lf.collect_schema().names()
                    # CI 매칭으로 root_lot_id 를 찾는다.
                    target = next((n for n in fab_names
                                   if n.casefold() == "root_lot_id"), None)
                    if target:
                        lots_list = _limited_unique_values(fab_lf, target, limit=limit)
                        if lots_list:
                            fallback_used = True
                            lot_col = target
        except Exception as e:
            logger.warning("/lot-ids: override 폴백 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
    return {"lot_col": lot_col, "lot_ids": lots_list,
            "fallback": "fab_source" if fallback_used else "",
            "fab_source": fab_source,
            "lookup_cache": lookup_meta}


@router.get("/lot-candidates")
def get_lot_candidates(
    product: str = Query(...),
    col: str = Query("root_lot_id"),
    prefix: str = Query(""),
    limit: int = Query(30),
    source: str = Query("auto"),   # v8.8.19: auto|override|mltable
    root_lot_id: str = Query(""),  # v9.0.0 (Q1): fab_lot_id 드롭다운을 특정 root 로 제한
):
    """Autocomplete 후보 반환. col 은 'root_lot_id' 또는 'fab_lot_id'. prefix 가
    비어있으면 최신/정렬 상위 N개, 아니면 prefix 포함 매칭을 정렬 순 top N.

    v8.8.19: `source` 인자 추가.
    v9.0.0: `root_lot_id` 파라미터 추가 — fab_lot_id 후보를 해당 root (앞 5자) 로 제한.
      (예: root_lot_id=A0001 → A0001 로 시작하는 fab_lot_id 만 반환)
    """
    # v9.0.5: fab_lot_id 후보는 DB FAB 원천 이력의 정확한 root/fab 매칭만 허용.
    #   DB FAB 에 없으면 ML_TABLE LOT_ID, starts_with, 전체 후보 fallback 으로 회피하지 않는다.
    root_scope = (root_lot_id or "").strip()
    if col.casefold() == "root_lot_id":
        # root 후보는 lookup candidate index → 제품별 RAM/디스크 풀만 사용한다.
        # 캐시 미스에서 원천/FAB/lot-ids를 동기 스캔하면 드롭다운 하나가 수 초~수
        # 분을 점유하므로, 빌드만 큐에 넣고 즉시 준비 중 응답을 반환한다.
        pool = _root_lot_pool(product)
        # 잠정 목록(빌드 중 파티션 스냅샷)도 그대로 내려준다. 종전에는 완성본만
        # 내보내 인덱스가 나올 때까지 드롭다운이 비어 있었다. `complete=False`
        # 를 함께 알려 프런트가 완성본으로 교체할 때까지 재확인을 이어간다.
        if pool.get("values"):
            values = pool["values"]
            needle = str(prefix or "").strip().upper()
            if needle:
                values = [v for v in values if needle in str(v).upper()]
            return {
                "col": "root_lot_id",
                "candidates": values[:max(1, int(limit or 30))],
                "prefix": prefix,
                "root_scope": root_scope,
                "strict": False,
                "complete": bool(pool.get("complete")),
                "pool_cache": pool.get("cached") or "",
                **(pool.get("meta") or {}),
            }
        meta = pool.get("meta") or {}
        return {
            "col": "root_lot_id",
            "candidates": [],
            "prefix": prefix,
            "root_scope": root_scope,
            "complete": False,
            "provisional": False,
            "match_mode": meta.get("match_mode") or "lookup_cache_preparing",
            "source": meta.get("source") or "mltable_lookup",
            "fab_source": "",
            "lookup_cache": meta.get("lookup_cache") or {},
            "strict": False,
        }
    if col.casefold() in {c.casefold() for c in _FAB_COL_CANDIDATES}:
        # 신규 Inform과 root 미선택 SplitTable은 제품 전체 LOT_ID 풀을 쓴다.
        # RAM/디스크에 이미 게시돼 있으면 prefix 입력도 서버 스캔 없이 필터한다.
        if not root_scope:
            pool = _fab_lot_pool(product)
            if pool.get("complete") and pool.get("values"):
                values = pool["values"]
                needle = str(prefix or "").strip().upper()
                if needle:
                    values = [value for value in values if needle in str(value).upper()]
                return {
                    "col": col,
                    "candidates": values[:max(1, int(limit or 30))],
                    "prefix": prefix,
                    "root_scope": "",
                    "strict": False,
                    "pool_complete": True,
                    "pool_cache": pool.get("cached") or "",
                    **(pool.get("meta") or {}),
                }
        main = _main_table_candidates(product, col, prefix=prefix, limit=limit, root_lot_id=root_scope)
        hist = _fab_history_scope(
            product,
            root_lot_id=root_scope,
            prefix=prefix,
            limit=limit,
            prefer_raw_latest=bool(root_scope or str(prefix or "").strip()),
        )
        main_candidates = main.get("candidates") or []
        hist_candidates = hist.get("candidates") or []
        if root_scope and hist_candidates:
            # A root can legitimately span multiple operational fab_lot_id values.
            # Keep the FAB history set authoritative for scoped lookups; intersecting
            # with ML_TABLE lot_id/fab_lot_id collapses cases like A1003A.2/A1003A.3.
            merged = _merge_candidate_values(hist_candidates, limit=limit)
        else:
            # Unscoped Inform LOT_ID search must also surface operational FAB
            # history lots that are not present in the current ML_TABLE render.
            # Intersecting here made searches such as "A1003" show only
            # A1003A.1 while hiding related A1003A.2/A1003A.3 entries.
            merged = _merge_candidate_values(main_candidates, hist_candidates, limit=limit)
        if merged:
            return {
                "col": col,
                "candidates": merged,
                "prefix": prefix,
                "root_scope": root_scope,
                "match_mode": "splittable_fab_lots" if root_scope else "splittable_fab_lots_all",
                "source": "mltable",
                "fab_source": hist.get("source", ""),
                "strict": False,
            }
        return {
            "col": col,
            "candidates": hist.get("candidates") or [],
            "prefix": prefix,
            "root_scope": root_scope,
            "match_mode": "fab_history_root" if root_scope else "fab_history",
            "source": "fab_source_history",
            "fab_source": hist.get("source", ""),
            "strict": True,
        }
    use_override = False
    lf = None
    if source == "override" and product.casefold().startswith("ml_table_"):
        try:
            meta = _resolve_override_meta(product, include_diagnostics=False)
            fab_source = (meta.get("fab_source") or "").strip()
            if fab_source and not meta.get("error"):
                fab_lf = _scan_fab_source(fab_source)
                if fab_lf is not None:
                    lf = fab_lf
                    use_override = True
        except Exception:
            lf = None
        if lf is None:
            return {"col": col, "candidates": [], "source": "override",
                    "note": "override 비활성 또는 fab_source 없음"}
    if lf is None:
        lf = _scan_product(
            product,
            root_lot_id=root_scope if col.casefold() != "root_lot_id" else "",
        )

    schema_names = lf.collect_schema().names()
    # v8.8.26: CI 매칭 — FE 가 "ROOT_LOT_ID"(ML_TABLE casing) 로 요청해도 raw 소스의
    # "root_lot_id" 로 정확히 매핑 (이전에는 exact match 만 되어 override 경로에서 누락).
    if col not in schema_names:
        col_ci = next((n for n in schema_names if n.casefold() == col.casefold()), None)
        if col_ci:
            col = col_ci
        else:
            # fallback — root 이면 auto-detect lot col, fab 는 그대로
            if col.casefold() == "root_lot_id":
                lot_col, _ = _detect_lot_wafer(lf)
                col = lot_col or col
            if col not in schema_names:
                return {"col": col, "candidates": [], "available_cols": schema_names[:20],
                        "source": "override" if use_override else "mltable"}

    match_mode = "all"
    fallback_used = False

    # v9.0.1: root_scope + fab_lot_id 조회 시 데이터-중심 매칭.
    #   데이터에서 root_lot_id 와 fab_lot_id 의 앞 5자가 자연 일치하지 않는 케이스 (예:
    #   ML_TABLE root=A0015 → fab_lot=A0005B.1) 에서 단순 starts_with 가 0건을 반환하던 문제.
    #   1) main lf 에서 root_lot_id 컬럼을 CI 매칭으로 찾고, 같은 row 의 fab_lot_id 를 unique 추출.
    #   2) (1) 결과가 비면 → 기존 starts_with 폴백.
    #   3) (2) 도 비면 → root_scope 무시하고 전체 후보 반환 (sentinel: fallback_used=True).
    if root_scope and col.casefold() != "root_lot_id":
        root_col = next((n for n in schema_names if n.casefold() == "root_lot_id"), None)
        if root_col:
            try:
                q_join = (lf.filter(_join_key_expr(root_col) == root_scope.strip().upper())
                            .select(pl.col(col).cast(_STR, strict=False).alias("v"))
                            .drop_nulls().unique())
                if prefix.strip():
                    q_join = q_join.filter(_contains_literal_ci_expr("v", prefix))
                rows_join = q_join.sort("v").head(limit).collect()
                cand_join = [v for v in rows_join["v"].to_list()
                             if v and str(v).strip() not in ("", "None", "null")]
                if cand_join:
                    return {"col": col, "candidates": cand_join, "prefix": prefix,
                            "root_scope": root_scope, "match_mode": "root_join",
                            "source": "override" if use_override else "mltable"}
            except Exception as e:
                logger.warning("/lot-candidates: root_join 실패 (product=%s) %s: %s",
                               product, type(e).__name__, e)

    q = lf.select(pl.col(col).cast(_STR, strict=False).alias("v")).drop_nulls().unique()
    if prefix.strip():
        q = q.filter(_contains_literal_ci_expr("v", prefix))
    if root_scope and col.casefold() != "root_lot_id":
        # 폴백 1: starts_with 5자 prefix
        try:
            q_sw = q.filter(pl.col("v").str.starts_with(root_scope[:5]))
            rows_sw = q_sw.sort("v").head(limit).collect()
            if rows_sw.height > 0:
                match_mode = "starts_with"
                return {"col": col, "candidates": rows_sw["v"].to_list(), "prefix": prefix,
                        "root_scope": root_scope, "match_mode": match_mode,
                        "source": "override" if use_override else "mltable"}
        except Exception:
            pass
        fallback_used = True
        match_mode = "all_fallback"
    rows = q.sort("v").head(limit).collect()
    return {"col": col, "candidates": rows["v"].to_list(), "prefix": prefix,
            "root_scope": root_scope, "match_mode": match_mode,
            "root_scope_fallback": fallback_used,
            "source": "override" if use_override else "mltable"}


@router.get("/column-values")
def get_column_values(product: str = Query(...), col: str = Query(...), limit: int = Query(200)):
    """빈셀 dbl-click edit suggestion — col 값의 unique 리스트 (전체 데이터셋 범위) +
    해당 product 의 plan 에 등록된 값 union. null/빈값 제외.
    """
    out: list[str] = []
    seen: set[str] = set()
    indexed = None
    try:
        fp = _product_path(product)
        indexed = _ml_table_lookup.candidate_values_from_lookup_cache(
            fp, col, limit=limit)
        if indexed.get("available"):
            values = indexed.get("values") or []
        else:
            lf = _scan_product(product)
            schema_names = lf.collect_schema().names()
            values = []
            if col in schema_names:
                rows = (lf.select(pl.col(col).cast(_STR, strict=False).alias("v"))
                        .drop_nulls().unique().sort("v").head(limit).collect())
                values = rows["v"].to_list()
        for v in values:
            if v is None: continue
            s = str(v).strip()
            if not s or s in ("None", "null"): continue
            if s in seen: continue
            seen.add(s); out.append(s)
    except Exception:
        pass
    try:
        for v in _custom_tag_column_values(product, col, limit=limit):
            if v in seen:
                continue
            seen.add(v)
            out.append(v)
            if len(out) >= limit:
                break
    except Exception:
        pass
    try:
        for v in _management_row_column_values(product, col, limit=limit):
            if v in seen:
                continue
            seen.add(v)
            out.append(v)
            if len(out) >= limit:
                break
    except Exception:
        pass
    # Union with plan values stored under this column
    try:
        plans = _load_plan_data(product).get("plans", {})
        for ck, pv in plans.items():
            # ck format: root_lot_id|wafer_id|col_name
            parts = str(ck).split("|")
            if len(parts) >= 3 and parts[2] == col:
                v = pv.get("value") if isinstance(pv, dict) else pv
                if v is None: continue
                s = str(v).strip()
                if not s or s in ("None", "null"): continue
                if s in seen: continue
                seen.add(s); out.append(s)
    except Exception:
        pass
    return {
        "col": col, "values": out, "count": len(out),
        "candidate_cache": "lookup_index" if indexed and indexed.get("available") else "",
        "candidate_cache_complete": bool(indexed and indexed.get("complete")),
    }


def _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id: str, wafer_ids: str,
                      fab_lot_id: str = "", fab_lot_col: str = "fab_lot_id"):
    """Apply lot + (optional) wafer filter to LazyFrame. v8.4.3 — fab_lot_id
    경로 추가. root_lot_id / fab_lot_id 중 하나로 조회 가능.
    """
    root_scope = root_lot_id.strip()
    fab_scope = fab_lot_id.strip()
    schema_names = lf.collect_schema().names()
    if root_scope and lot_col and lot_col in schema_names:
        lf = lf.filter(_join_key_expr(lot_col) == root_scope.upper())
    if fab_scope and fab_lot_col in schema_names:
        lf = lf.filter(_join_key_expr(fab_lot_col) == fab_lot_id.strip().upper())
    if wafer_ids.strip() and wf_col:
        wf_list = [w.strip() for w in wafer_ids.split(",") if w.strip()]
        try:
            wf_ints = [int(w) for w in wf_list]
            # Build all possible formats: 1 → ["1", "01", "W01", "W1"]
            wf_strs = set()
            for n in wf_ints:
                wf_strs.update([str(n), f"{n:02d}", f"W{n}", f"W{n:02d}"])
            lf = lf.filter(
                pl.col(wf_col).cast(_STR, strict=False).is_in(list(wf_strs))
                | pl.col(wf_col).cast(pl.Int64, strict=False).is_in(wf_ints)
            )
        except ValueError:
            lf = lf.filter(pl.col(wf_col).cast(_STR, strict=False).is_in(wf_list))
    return lf


def _ml_product_name(product: str) -> str:
    p = str(product or "").strip()
    if not p:
        return ""
    return _canonical_mltable_product_name(p, allow_bare=True)


def resolve_fab_lot_snapshot(product: str, root_lot_id: str, wafer_id: str = "") -> str:
    """Return the fab_lot_id from the same coalesced SplitTable data users see."""
    ml_product = _ml_product_name(product)
    root = str(root_lot_id or "").strip()
    if not ml_product or not root:
        return ""
    try:
        cached = _fab_lot_snapshot_from_cache(ml_product, root, wafer_id)
        if cached:
            return cached
        lf = _scan_product(ml_product, root_lot_id=root, wafer_ids=str(wafer_id or ""))
        lot_col, wf_col = _detect_lot_wafer(lf, ml_product)
        if not lot_col:
            return ""
        names = lf.collect_schema().names()
        fab_col = "fab_lot_id" if "fab_lot_id" in names else ""
        if not fab_col:
            fab_col = _pick_first_present_ci(_FAB_COL_CANDIDATES, names) or ""
        if not fab_col:
            return ""
        lf = _filter_lot_wafer(lf, lot_col, wf_col, root, str(wafer_id or ""),
                               fab_lot_col=fab_col)
        df = (
            lf.select(pl.col(fab_col).cast(_STR, strict=False).alias("fab_lot_id"))
            .drop_nulls()
            .unique()
            .sort("fab_lot_id")
            .head(1)
            .collect()
        )
        if df.height == 0:
            return ""
        return str(df.item(0, 0) or "").strip()
    except Exception as e:
        logger.warning("resolve_fab_lot_snapshot 실패 (product=%s root=%s wafer=%s) %s: %s",
                       product, root_lot_id, wafer_id, type(e).__name__, e)
        return ""


def _resolve_fab_lot_for_cell(product: str, cell_key: str, root_lot_id: str = "") -> str:
    parts = str(cell_key or "").split("|")
    root = str(root_lot_id or (parts[0] if len(parts) >= 1 else "") or "").strip()
    wafer = str(parts[1] if len(parts) >= 2 else "").strip()
    return resolve_fab_lot_snapshot(product, root, wafer)


def _split_view_cache_key(product: str, root_lot_id: str, wafer_ids: str, prefix: str,
                          custom_name: str, view_mode: str, history_mode: str,
                          fab_lot_id: str, custom_cols: str) -> tuple:
    canonical_product = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    cleaned_custom_cols = ",".join(_clean_custom_columns(str(custom_cols or "").split(","))) if custom_cols else ""
    return (
        canonical_product,
        str(root_lot_id or "").strip(),
        str(wafer_ids or "").strip(),
        str(prefix or "").strip().upper(),
        str(custom_name or "").strip(),
        str(view_mode or "all").strip().lower() or "all",
        str(history_mode or "all").strip().lower() or "all",
        str(fab_lot_id or "").strip(),
        cleaned_custom_cols,
    )


def _split_view_cache_stats(hit: bool, key: tuple | None = None, *, stale: bool = False) -> dict:
    key_hash = ""
    if key is not None:
        try:
            raw = json.dumps(key, sort_keys=True, ensure_ascii=False, default=str)
            key_hash = hashlib.sha1(raw.encode("utf-8")).hexdigest()[:12]
        except Exception:
            key_hash = ""
    with _VIEW_CACHE_LOCK:
        size = len(_VIEW_CACHE)
    return {
        "hit": bool(hit),
        "payload_cache_hit": bool(hit),
        # stale=True → 캐시로 즉시 응답했고 백그라운드 재검증이 예약됨(SWR).
        "stale": bool(stale),
        "entries": size,
        "max_entries": _view_cache_max_entries(),
        "key": key_hash,
    }


def _split_view_data_source_label(src: dict, *, payload_cache_hit: bool) -> str:
    """검색이 실제로 어느 계층에서 데이터를 얻었는지 한 단어로 분류.

    payload_cache: 응답 전체 캐시 히트(가장 빠름) · pivot_cache: per-root pivot 캐시 ·
    product_ram/ram: 메모리 캐시 히트 · ram_load: 첫 검색으로 파티션을 메모리 적재 ·
    disk: 파티션 parquet 디스크 스캔(첫 검색) · raw/fallback: 캐시 없이 원본 경로.
    """
    if payload_cache_hit:
        return "payload_cache"
    ds = str(src.get("root_data_source") or "").strip()
    if ds:
        return ds
    if src.get("product_cache_hit"):
        return "product_ram"
    if src.get("root_cache_hit"):
        return "root_cache"
    return "raw"


def _request_lane_wait_ms(request: Request | None) -> float:
    """ResourceGuardMiddleware 가 남긴 레인 대기 시간(ms). 없으면 0."""
    try:
        return max(0.0, float(getattr(request.state, "lane_wait_ms", 0.0) or 0.0))
    except Exception:
        return 0.0


# ── 단계별 계측 ────────────────────────────────────────────────────────────────
# scan/collect/matrix/overlay 4개만 재던 시절엔 느린 검색의 절반 이상이 어느
# 단계인지 알 수 없었다(실측 17,862건에서 미계측 57%). 아래 _PHASE_KEYS 는 그
# 빈칸을 메우는 구간들이고, unaccounted_ms 가 남은 오차를 그대로 드러낸다 —
# 이 값이 크면 "아직 못 재고 있는 구간이 있다"는 뜻이지 최적화 대상이 없다는
# 뜻이 아니다.
#
# 주의: collect_ms 의 의미가 좁아졌다. 예전에는 컬럼 선택(파이썬)+polars collect
# 를 합친 값이었고 지금은 collect 만이다. 그 앞 파이썬 구간은 select_ms 다.
_PHASE_KEYS: tuple[str, ...] = (
    "prelude_ms",      # 진입~payload 캐시 판정 (인증/감사/캐시키/의존 시그니처)
    "fastpath_ms",     # pivot 캐시 탐색 + 스키마 확인 + KNOB 사이드카 판정
    "scan_ms",         # 파티션/RAM 소스 결정 + lazy join 구성
    "root_scan_ms",    #   그중 root RAM/디스크 파티션 확보
    "schema_ms",       # lot/wafer/fab 컬럼 해석 (스키마 + 오버라이드 설정)
    "fabscope_ms",     # fab_lot_id 히스토리 스코프 조회
    "select_ms",       # 컬럼 선택·rename·공정순 정렬 (순수 파이썬)
    "collect_ms",      # polars collect
    "emptyfallback_ms",  # 빈 결과일 때의 재해석(root-only / 붙여넣은 fab_lot) 시도
    "header_ms",       # 웨이퍼 헤더/그룹 조립 (df→파이썬 리스트)
    "overlay_ms",      # plan/tag/management 로드 + override/fab 후보
    "matrix_ms",       # 셀 매트릭스 조립
    "mismatch_ms",     # plan↔actual 불일치 알림 적재
    "progress_ms",     # step_progress (미진행 행 회색 처리)
    "payload_ms",      # 응답 dict 조립 (prefixes/precision/all_columns/meta)
    "cacheput_ms",     # payload 캐시 저장
    "finish_ms",       # 마무리 메타 (lookup_cache/related_issues/product_cache)
    "serialize_ms",    # orjson 직렬화 (HTTP 경로 전용)
)


_LAP_MARK = "_lap_mark"


def _lap(profile: dict | None, key: str | None) -> None:
    """구간 계측 — 직전 lap 이후 흐른 시간을 key 에 누적하고 마커를 옮긴다.

    구간마다 따로 t0 를 잡지 않고 마커 하나를 이어 달리는 이유는 **빈틈이 생기지
    않게** 하기 위해서다. 개별 타이머 방식은 타이머와 타이머 사이의 코드가 아무
    키에도 안 잡혀서 조용히 사라진다 — 기존 계측이 절반 이상을 놓친 원인이 정확히
    이거였다. 이 방식에서는 lap 을 부르지 않은 시간만 unaccounted_ms 로 남는다.

    key=None 은 "이 구간은 다른 데서 이미 셌으니 버린다"(레인/단일비행 대기).
    같은 key 를 여러 번 lap 해도 합산된다(폴백 경로 재실행).

    마커를 profile 안에 두므로 핸들러 밖의 헬퍼(_attach_split_view_runtime_fields)
    도 같은 흐름을 이어서 잴 수 있다.
    """
    if profile is None:
        return
    now = time.perf_counter()
    mark = profile.get(_LAP_MARK)
    if mark is not None and key is not None:
        profile[key] = float(profile.get(key) or 0.0) + (now - float(mark)) * 1000.0
    profile[_LAP_MARK] = now


def _split_view_runtime_profile(started: float, runtime_profile: dict | None, *, payload_cache_hit: bool) -> dict:
    src = runtime_profile or {}
    data_source = _split_view_data_source_label(src, payload_cache_hit=payload_cache_hit)
    total_ms = (time.perf_counter() - started) * 1000.0
    lane_wait_ms = float(src.get("lane_wait_ms") or 0.0)
    cold_lane_wait_ms = float(src.get("cold_lane_wait_ms") or 0.0)
    singleflight_wait_ms = round(float(src.get("singleflight_wait_ms") or 0.0), 3)
    phases = {key: round(float(src.get(key) or 0.0), 3) for key in _PHASE_KEYS}
    # root_scan_ms 는 scan_ms 안쪽 구간이라 합계에서 뺀다(이중 계산 방지).
    # serialize_ms 는 total_ms 밖에서 일어나므로 역시 제외한다.
    accounted = sum(
        v for k, v in phases.items() if k not in ("root_scan_ms", "serialize_ms")
    ) + cold_lane_wait_ms + singleflight_wait_ms
    unaccounted = max(0.0, total_ms - accounted)
    return {
        # 검색 상세 이력에서 실제 요청자를 보여 주기 위한 세션 정보. 응답에도 포함되지만
        # 비밀번호/토큰 같은 인증 정보는 넣지 않고 식별자와 역할만 보관한다.
        "username": str(src.get("username") or ""),
        "user_role": str(src.get("user_role") or ""),
        "is_user_search": bool(src.get("is_user_search")),
        # total_ms 는 핸들러 전체 = cold 레인 대기(핸들러 '안'에서 줄 선 시간)를
        # 포함한다. 순수 계산만 보려면 compute_ms 를 쓸 것 — total 만 보면 줄서기가
        # 계산 시간으로 오인된다.
        "total_ms": round(total_ms, 3),
        "compute_ms": round(max(0.0, total_ms - cold_lane_wait_ms), 3),
        # 핸들러 밖에서 줄 선 시간까지 포함한 체감 시간. total_ms 만 보면 "서버는
        # 빠른데 사용자는 느린" 상태의 원인이 안 보인다.
        "wall_ms": round(total_ms + lane_wait_ms, 3),
        # 줄 선 시간 합계 = 미들웨어 레인 + cold 계산 레인.
        "wait_ms": round(lane_wait_ms + cold_lane_wait_ms, 3),
        "lane_wait_ms": round(lane_wait_ms, 3),
        "cold_lane_wait_ms": round(cold_lane_wait_ms, 3),
        "root_cache_hit": bool(src.get("root_cache_hit")),
        "product_cache_hit": bool(src.get("product_cache_hit")),
        "payload_cache_hit": bool(payload_cache_hit),
        "data_source": data_source,
        # KNOB 전용 사이드카(좁은 pivot 파일)로 읽었는지 — 운영에서 이 경로가
        # 실제로 도는지 확인용. False 면 전체 pivot 파일로 폴백한 것.
        "knob_sidecar": bool(src.get("root_data_source_detail") == "knob_sidecar"),
        "singleflight_wait_ms": singleflight_wait_ms,
        # 단계별 breakdown — 전 구간 계측. serialize_ms 만 total_ms 밖이다
        # (직렬화는 이 프로필이 만들어진 뒤 HTTP 계층에서 일어난다).
        **phases,
        # 아직 어느 단계로도 설명되지 않은 시간. 0 에 가까워야 정상이고, 크면
        # 계측이 빠진 구간이 남아 있다는 신호다.
        "unaccounted_ms": round(unaccounted, 3),
        "root_cache_status": str(src.get("root_cache_status") or ""),
        "root_prefetch_queued": bool(src.get("root_prefetch_queued")),
        "fab_index_queued": bool(src.get("fab_index_queued")),
        "cache_incomplete": bool(src.get("cache_incomplete")),
    }


def _split_view_finish_payload(
    payload: dict,
    *,
    started: float,
    runtime_profile: dict | None,
    payload_cache_hit: bool,
    view_cache_key: tuple | None,
    view_stale: bool = False,
) -> dict:
    out = dict(payload)
    rp = _split_view_runtime_profile(started, runtime_profile, payload_cache_hit=payload_cache_hit)
    out["runtime_profile"] = rp
    out["view_cache"] = _split_view_cache_stats(payload_cache_hit, view_cache_key, stale=view_stale)
    _record_search_timing(out, rp)
    _view_compute_finish(view_cache_key)
    return out


# ── SplitTable 검색 타이밍 로그 (관리자 breakdown 용) ──────────────────────────
# 단계별 소요시간을 기록해 관리자 화면에서 "캐시 히트일 때 속도 / 첫 검색(DB 조회)
# 단계별 breakdown / 동시 요청 때문에 줄 선 시간" 을 보여준다.
# 보관은 core/search_timing_log.py 가 담당한다 — 인메모리 링버퍼(최근 조회용) +
# 공유 JSONL(며칠 치 관찰용, 운영/개발 서버 통합).
def _record_search_timing(payload: dict, rp: dict) -> None:
    try:
        # request=None 인 캐시 재검증/캐시 빌드/Inform·Flow-i 내부 조회는 제외한다.
        # 실제 브라우저의 /view HTTP 요청만 runtime_profile 에 이 표식을 갖는다.
        if not bool(rp.get("is_user_search")):
            return
        root = str(payload.get("root_lot_id") or "").strip()
        if not root:
            return
        rows = payload.get("rows")
        if not isinstance(rows, list):
            rows = payload.get("rows_compact")
        row_count = len(rows) if isinstance(rows, list) else 0
        entry = {
            "actor_type": "user_search",
            "at": datetime.datetime.now().isoformat(timespec="seconds"),
            "product": str(payload.get("product") or ""),
            "root_lot_id": root,
            "data_source": rp.get("data_source") or "",
            "total_ms": rp.get("total_ms") or 0.0,
            "compute_ms": rp.get("compute_ms") or 0.0,
            "wall_ms": rp.get("wall_ms") or 0.0,
            "wait_ms": rp.get("wait_ms") or 0.0,
            "lane_wait_ms": rp.get("lane_wait_ms") or 0.0,
            "cold_lane_wait_ms": rp.get("cold_lane_wait_ms") or 0.0,
            "singleflight_wait_ms": rp.get("singleflight_wait_ms") or 0.0,
            # 단계별 breakdown 전량 + 미설명 잔여.
            **{key: rp.get(key) or 0.0 for key in _PHASE_KEYS},
            "unaccounted_ms": rp.get("unaccounted_ms") or 0.0,
            "root_cache_hit": bool(rp.get("root_cache_hit")),
            "payload_cache_hit": bool(rp.get("payload_cache_hit")),
            "row_count": row_count,
            "col_count": int(payload.get("selected_count") or 0),
            "wafer_count": len(payload.get("headers") or []),
            "cache_status": rp.get("root_cache_status") or "",
            "username": str(rp.get("username") or ""),
            "user_role": str(rp.get("user_role") or ""),
        }
        # HTTP 경로면 직렬화까지 재고 나서 한 줄로 쓴다. 직렬화는 이 함수가
        # 끝난 뒤(view_split_http)에 일어나므로, 여기서 바로 쓰면 KNOB 처럼 큰
        # 응답의 직렬화 비용이 영영 기록되지 않는다.
        pending = getattr(_VIEW_TIMING_TLS, "pending", None)
        if isinstance(pending, list):
            pending.append(entry)
            return
        _search_timing_log.record(entry)
    except Exception:
        pass


def recent_search_timings(limit: int = 50) -> list[dict]:
    return _search_timing_log.recent(limit)


def _view_global_stat_sig() -> tuple:
    """product-독립 전역 파일들의 stat 시그니처를 (global_hard, global_soft) 로 반환.

    짧은 TTL 로 캐시 — 동시 다수 사용자가 매 요청 같은 전역 파일(config/rulebook/
    lot_progress 파생)을 재-stat 하던 공유드라이브 부하를 없앤다. 전역 hard(config/
    rulebook) 변경은 admin 행위라 ≤TTL 지연 허용, soft(lot_progress 파생) 변경은
    어차피 SWR 이 흡수하므로 지연 무해.
    """
    now = time.monotonic()
    with _VIEW_GLOBAL_SIG_LOCK:
        cached = _VIEW_GLOBAL_SIG_CACHE.get("v")
        if cached is not None and (now - cached[0]) < _VIEW_GLOBAL_SIG_TTL:
            return cached[1]
    hard_paths: list[Path] = [SOURCE_CFG, PREFIX_CFG, PRECISION_CFG, RULEBOOK_SCHEMA_FILE]
    for kind in _RULEBOOK_FILES:
        try:
            hard_paths.append(_rulebook_path(kind))
        except Exception:
            pass
    global_hard = tuple(_path_cache_sig(path) for path in hard_paths)
    soft_paths: list[Path] = [
        MATCH_CACHE_STATE_FILE,
        _latest_lot_step_cache_path(),
        PATHS.cache_dir / "lot_progress" / "lot_wf_current.json",
        PATHS.cache_dir / "lot_progress" / "lot_wf_current.parquet",
    ]
    global_soft = tuple(_path_cache_sig(path) for path in soft_paths)
    val = (global_hard, global_soft)
    with _VIEW_GLOBAL_SIG_LOCK:
        _VIEW_GLOBAL_SIG_CACHE["v"] = (now, val)
    return val


def _split_view_cache_dep_signature(product: str, custom_name: str = "", product_fp: Path | None = None) -> tuple:
    """View payload 캐시의 2-tier 의존 시그니처 (hard_sig, soft_sig) 반환.

    hard_sig — 즉시 무효화 대상. 소스 ML_TABLE(신규 lot 신호) + 사용자가 직접
      편집하는 입력(prefix/precision/rulebook/custom tag/management/plan/custom).
      per-product 편집 파일은 항상 fresh stat 하므로 편집·신규 lot 이 지연 없이 반영.
    soft_sig — 백그라운드 스케줄러가 주기적으로 재기록하는 파생 캐시(lot_progress
      최신 lot, match cache, product RAM cache). soft 만 달라졌을 때는 stale-while-
      revalidate 로 캐시를 즉시 서빙하고 백그라운드에서 갱신. (이전에는 이것들이
      hard 와 묶여 lot_progress 재기록마다 모든 검색이 캐시 miss → 풀 재계산.)

    HIT 경로 stat 비용 절감: product-독립 전역 파일은 _view_global_stat_sig 로 짧은
    TTL 캐시해 동시 요청 폭주 시 재-stat 를 제거한다.
    """
    # per-product/사용자편집 파일 — 항상 fresh stat (즉시 무효화 보장).
    fresh_paths: list[Path] = [
        product_fp or _product_path(product),
        _view_product_invalidation_path(product),
        _custom_tags_path(),
        _management_rows_path(),
    ]
    fresh_paths.extend(_plan_alias_paths(product))
    if str(custom_name or "").strip():
        try:
            custom_fp, _clean_name = _custom_file_path_for_name(custom_name)
            fresh_paths.append(custom_fp)
        except HTTPException:
            pass
    per_product_hard = _view_product_signature(fresh_paths)
    global_hard, global_soft = _view_global_stat_sig()
    hard_sig = (per_product_hard, global_hard)
    soft_sig = global_soft + (_product_ram_cache_view_signature(product),)
    return (hard_sig, soft_sig)


def _view_product_invalidation_path(product: str) -> Path:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    digest = hashlib.sha1(canonical.encode("utf-8")).hexdigest()[:16]
    return _base_root() / "cache" / "split_table_view_payload" / "invalidate" / f"{digest}.token"


def _touch_view_product_invalidation(product: str) -> None:
    fp = _view_product_invalidation_path(product)
    try:
        fp.parent.mkdir(parents=True, exist_ok=True)
        tmp = fp.with_name(f"{fp.name}.{os.getpid()}.{threading.get_ident()}.tmp")
        tmp.write_text(str(time.time_ns()), encoding="utf-8")
        os.replace(tmp, fp)
    except Exception:
        logger.debug("SplitTable product invalidation token write failed: %s", product, exc_info=True)


def _clear_split_view_cache_product(product: str) -> None:
    """한 제품의 RAM/disk view만 무효화하고 다른 사용자·제품의 HIT는 보존한다."""
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    if not canonical:
        return
    _touch_view_product_invalidation(canonical)
    global _VIEW_CACHE_BYTES
    with _VIEW_CACHE_LOCK:
        doomed = [key for key in _VIEW_CACHE if str(key[0] if key else "").upper() == canonical.upper()]
        for key in doomed:
            entry = _VIEW_CACHE.pop(key, None)
            if entry is not None:
                _VIEW_CACHE_BYTES = max(0, _VIEW_CACHE_BYTES - entry[3])
    with _VIEW_PRODUCT_SIG_LOCK:
        _VIEW_PRODUCT_SIG_CACHE.clear()
    with _VIEW_REVALIDATE_LOCK:
        for key in list(_VIEW_REVALIDATE_PENDING):
            if str(key[0] if key else "").upper() == canonical.upper():
                _VIEW_REVALIDATE_PENDING.pop(key, None)


def _clear_split_view_cache() -> None:
    global _VIEW_CACHE_BYTES
    with _VIEW_CACHE_LOCK:
        _VIEW_CACHE.clear()
        _VIEW_CACHE_BYTES = 0
    # 전역 시그니처 TTL 캐시도 함께 비운다 — 캐시 재빌드/명시적 무효화가 TTL(≤1s)
    # 지연 없이 즉시 반영되도록.
    with _VIEW_GLOBAL_SIG_LOCK:
        _VIEW_GLOBAL_SIG_CACHE.clear()
    with _VIEW_PRODUCT_SIG_LOCK:
        _VIEW_PRODUCT_SIG_CACHE.clear()
    # 제품 파일 탐색 결과도 함께 — 명시적 무효화는 "배치가 바뀌었을 수 있다"는
    # 신호이므로 TTL 을 기다리지 않는다.
    _clear_product_path_cache()
    # 대기 중이던 재검증도 함께 버린다 — 대상 엔트리가 방금 전부 비워졌으므로
    # 재계산해도 다음 stale hit 때 다시 등록될 뿐인 낭비다.
    with _VIEW_REVALIDATE_LOCK:
        _VIEW_REVALIDATE_PENDING.clear()


# lookup 캐시(hive 파티션) 재빌드가 끝나면 view payload 캐시를 비운다 — stale
# 파티션으로 렌더해 캐시된 payload 를 fresh 데이터로 재계산시키기 위함.
# root_lot 풀도 같이 버린다: 빌드 직후 root 집합이 달라져 있고, 시그니처만
# 믿으면 mtime 해상도 안에서 옛 목록이 잠깐 더 살아남을 수 있다.
try:
    _ml_table_lookup.register_build_complete_hook(
        lambda fp: _clear_split_view_cache_product(Path(fp).stem)
    )
    _ml_table_lookup.register_build_complete_hook(_refresh_root_lot_pool_after_lookup_build)
except Exception:
    logger.debug("ml_table_lookup build-complete hook 등록 실패", exc_info=True)


# ── Pre-pivoted root_lot cache: background build (single-flight per product) ──
# v9.1.x: plan 저장 후 백그라운드 작업 스레드 핸들 — 테스트가 join 으로 완료를 기다린다.
_PLAN_POST_SAVE_LAST_THREAD: threading.Thread | None = None

_PIVOT_BUILD_LOCK = threading.Lock()
_PIVOT_BUILD_INPROGRESS: set[str] = set()
_PIVOT_BUILD_LAST: dict[str, float] = {}
_PIVOT_BUILD_COOLDOWN_SEC = 300.0
_SEARCH_CACHE_MAINT_LOCK = threading.Lock()
_SEARCH_CACHE_MAINT_THREAD: threading.Thread | None = None
_SEARCH_CACHE_MAINT_WAKE = threading.Event()
_AUTO_PRODUCT_CACHE_STATE_LOCK = threading.RLock()
_AUTO_PRODUCT_CACHE_STATE: dict = {
    "started": False,
    "current_product": "",
    "queued_product": "",
    "queued_task_id": "",
    "next_product": "",
    "next_at": "",
    "next_at_ts": 0.0,
    "last_product": "",
    "last_at": "",
    "last_ok": None,
    "job_id": "",
    "cycle_first_product": "",
    "cycle_completed_products": 0,
    "last_duration_sec": 0.0,
    "delayed": False,
    "delayed_reason": "",
    "loaded": False,
}
_AUTO_PRODUCT_CACHE_RETRY_SEC = 15.0
_AUTO_PRODUCT_CACHE_DEFAULT_INTERVAL_MINUTES = 15


def _auto_product_cache_iso(epoch_seconds: float) -> str:
    return datetime.datetime.fromtimestamp(
        float(epoch_seconds), tz=datetime.timezone.utc
    ).astimezone().isoformat(timespec="seconds")


def _pivot_cache_path(product: str, root_lot_id: str) -> Path:
    # Resolve under the *active* base root (db_root) rather than a value frozen
    # at import time. In production `_base_root()/cache/split_table` is identical
    # to the builder's CACHE_DIR (db_cache_dir == db_root/cache); resolving it
    # live keeps reader/writer consistent if the DB root is re-pointed at runtime
    # (admin_settings takes effect without a restart) and lets tests that patch
    # `_base_root` sandbox the pivot cache instead of reading the global one.
    from app_v2.modules.splittable.cache_builder import canonical_product_dir
    canonical = canonical_product_dir(product) or str(product or "").strip()
    safe_root = str(root_lot_id).replace("/", "_").replace("\\", "_")
    return _base_root() / "cache" / "split_table" / canonical / f"{safe_root}.parquet"


def _pivot_cache_knob_path(product: str, root_lot_id: str) -> Path:
    """KNOB 전용 사이드카 경로 (하위 폴더 — lot-candidates 의 *.parquet 열거와 분리)."""
    from app_v2.modules.splittable.cache_builder import KNOB_SIDECAR_DIR
    main = _pivot_cache_path(product, root_lot_id)
    return main.parent / KNOB_SIDECAR_DIR / main.name


def _knob_only_request(prefix: str, custom_name: str, custom_cols: str) -> bool:
    """이 검색이 KNOB prefix '만' 보는가 (사이드카 사용 조건)."""
    if str(custom_name or "").strip() or str(custom_cols or "").strip():
        return False
    parts = [p.strip().upper() for p in str(prefix or "").split(",") if p.strip()]
    return parts == ["KNOB"]


def _merge_all_columns(frame_cols: list, full_cols: list | None, lot_col: str, wf_col: str) -> list:
    """FE 컬럼 선택기용 전체 컬럼 목록.

    KNOB 사이드카로 읽으면 프레임에는 KNOB 컬럼밖에 없다. 그대로 내보내면
    커스텀 세트 만들 때 INLINE/VM 등이 사라지므로, 전체 pivot 스키마를 기준으로
    복원하고 프레임에만 있는 항목(태그/관리 행 오버레이)을 뒤에 붙인다."""
    if not full_cols:
        return frame_cols
    skip = {c for c in (lot_col, wf_col) if c}
    out = [c for c in full_cols if c not in skip]
    seen = set(out)
    for c in frame_cols or []:
        if c not in seen:
            out.append(c)
            seen.add(c)
    return out


def _knob_sidecar_usable(sidecar: Path, main: Path) -> bool:
    """사이드카를 써도 결과가 같은가.

    ① 본 파일보다 오래되지 않았고 ② 렌더에 필요한 앵커 컬럼(lot/wafer/fab)이
    모두 들어 있어야 한다. 하나라도 어긋나면 전체 파일로 폴백한다 — 사이드카가
    결과를 바꾸는 경로를 만들지 않는다."""
    try:
        if sidecar.stat().st_mtime < main.stat().st_mtime - 1.0:
            return False
    except OSError:
        return False
    names = _cached_scan_schema(sidecar)[1]
    if not names:
        return False
    lower = {str(n).lower() for n in names}
    if not any(n in lower for n in ("wafer_id", "waferid", "wafer")):
        return False
    return any(n in lower for n in ("root_lot_id", "lot_id"))


def _pivot_cache_build_state(product: str) -> str:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    with _PIVOT_BUILD_LOCK:
        if canonical in _PIVOT_BUILD_INPROGRESS:
            return "building"
        if _PIVOT_BUILD_LAST.get(canonical):
            return "built"
    return ""


def _match_progress(done: int, total: int, *, state: str = "running") -> dict:
    """전체 랏 진행률 집계용 표준 블록 (매칭 캐시). 수집기가 없으면 빈 dict."""
    try:
        from core.cache_event_log import progress_detail
        return progress_detail("match", done, total, state=state)
    except Exception:
        return {}


def _stage(kind: str, phase: str) -> dict:
    """제품별 캐시 이력 집계용 수명주기 표식. 수집기가 없으면 빈 dict."""
    try:
        from core.cache_event_log import stage_detail
        return stage_detail(kind, phase)
    except Exception:
        return {}


def _cache_build_emit(product: str, event: str, *, ok: bool = True, detail: dict | None = None) -> None:
    """백그라운드 캐시 빌드(pivot·WIP latest-lot) 진행을 캐시 이벤트 로그로 내보낸다.

    **`logger.info` 는 화면에 안 보인다** — 캐시관리 화면은 `cache_event_log.record`
    만 읽는다. 수동 캐싱은 pivot 빌드를 큐에 넣고 "큐 등록 완료"만 남겼고, 실제
    빌드/실패/lease 스킵은 서버 터미널에만 찍혀 "캐시를 지웠는데 이벤트 로그에
    아무것도 안 뜬다"가 됐다 (FAB 랏인덱스는 `_fab_idx_emit` 로 이미 해결한 것과
    같은 문제). 같은 category(`cache_op`)라 '전체' 필터에 실시간으로 섞여 보인다.
    """
    try:
        from core.cache_event_log import record as _rec
        _rec("cache_op", event, ok=ok, detail=detail or {}, product=product)
    except Exception:
        pass


def _enqueue_pivot_cache_build(product: str, reason: str = "", *, immediate: bool = False,
                               local_only: bool = False) -> bool:
    """Rebuild the product's pre-pivoted root_lot cache in a daemon thread.
    Single-flight per product with a cooldown so view-triggered rebuilds cannot
    stampede; the daily 03:00 scheduler remains the full sweep.

    local_only=True 면 개발 워커로 오프로드하지 않고 이 서버에서 직접 빌드한다
    (관리자가 이 서버에서 누른 수동 캐싱)."""
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    if not canonical:
        return False
    now = time.time()
    with _PIVOT_BUILD_LOCK:
        if canonical in _PIVOT_BUILD_INPROGRESS:
            # 검색이 유발하는 스킵은 빌드 1건당 수십 번 발생하므로 로그에 남기지
            # 않는다. 관리자가 직접 요청한 경우(immediate)만 "왜 안 도는지" 를 남긴다.
            if immediate:
                _cache_build_emit(canonical, "[Pivot캐시] 이미 빌드 중 — 이번 요청은 건너뜁니다",
                            detail={"reason": reason})
            return False
        if (not immediate and
                now - _PIVOT_BUILD_LAST.get(canonical, 0.0) < _PIVOT_BUILD_COOLDOWN_SEC):
            return False
        _PIVOT_BUILD_INPROGRESS.add(canonical)

    try:
        source_fp = _product_path(canonical)
    except HTTPException:
        source_fp = None

    # 중단 신호(scan_gate)는 **스레드 로컬**이다 — 아래 데몬 스레드에서 인자 없이
    # cancel_requested() 를 부르면 항상 False 라, 관리자가 "중단" 을 눌러도 빌드가
    # 끝까지 돌았다. 호출한 스레드(스캔 큐 작업)에서 지금 id 를 붙잡아 빌더까지
    # 명시적으로 넘긴다. 큐 작업 밖에서 시작된 빌드는 id 가 없어 종전대로 무중단.
    try:
        from core import scan_gate as _scan_gate_capture
        cancel_task_id = _scan_gate_capture.current_task_id()
    except Exception:
        cancel_task_id = ""

    def _run():
        ok = False
        started_ts = time.time()
        _cache_build_emit(canonical, f"[Pivot캐시] 빌드 시작 — {reason or '요청'}",
                    detail={"reason": reason, "immediate": immediate})

        def _local_build() -> dict:
            # v9.1.x: 공유 flow-data lease — 개발/운영 서버가 같은 product 를 동시에
            # 빌드하지 않게 한다. lease 실패 = 다른 서버가 빌드 중 → 건너뛰고
            # cooldown 후 재시도 (그 사이 빌드 완료본을 그대로 사용).
            lease_name = f"splittable_pivot_{canonical}"
            lease_held = False
            try:
                from core import shared_lease as _shared_lease
                lease_held = _shared_lease.try_acquire(lease_name, ttl_sec=1800.0)
                if not lease_held:
                    holder = _shared_lease.holder(lease_name)
                    logger.info(f"pivot cache build skipped for {canonical} — 다른 서버가 빌드 중 (holder={holder})")
                    # lease 는 30분 TTL 이라 죽은 홀더가 남으면 그 시간 동안 빌드가
                    # 조용히 아무것도 안 한다 — 반드시 화면에서 보여야 한다.
                    _cache_build_emit(canonical,
                                f"[Pivot캐시] 빌드 건너뜀 — 다른 서버가 빌드 중 (holder={holder or '?'})",
                                ok=False, detail={"reason": reason, "lease": lease_name,
                                                  "stage": _stage("pivot", "fail")})
                    return {"ok": False}
                from app_v2.modules.splittable.cache_builder import build_pivoted_cache_for_product
                # 사용자 검색이 진행 중이면 빌드 시작을 미룬다 — 검색 collect 와 전역
                # polars 풀을 두고 경쟁하지 않도록 (빌더 내부 배치 사이 yield 와 별개로
                # 시작 시점 자체를 한 번 양보).
                from core import request_priority as _rp
                _rp.yield_to_users(max_wait_sec=60.0)

                from core import scan_gate as _shared_cancel

                def _should_cancel() -> bool:
                    if not cancel_task_id:
                        return False
                    try:
                        return bool(_shared_cancel.cancel_requested(cancel_task_id))
                    except Exception:
                        return False

                def _renew_lease() -> None:
                    # try_acquire 의 TTL 은 30 분인데 빌드는 그보다 오래 걸릴 수
                    # 있다. 갱신하지 않으면 lease 가 stale 이 되어 다른 서버가
                    # 같은 제품을 동시에 빌드한다.
                    if lease_held:
                        _shared_lease.renew(lease_name, ttl_sec=1800.0)

                with _HEAVY_BUILD_SEMAPHORE:
                    return {"ok": bool(build_pivoted_cache_for_product(
                        canonical, product_path=source_fp,
                        should_cancel=_should_cancel, on_chunk_done=_renew_lease))}
            except Exception as exc:
                logger.warning(f"pivot cache build failed for {canonical} ({reason}): {exc}")
                _cache_build_emit(canonical, f"[Pivot캐시] 빌드 실패 — {exc}", ok=False,
                            detail={"reason": reason, "error": str(exc),
                                    "stage": _stage("pivot", "fail")})
                return {"ok": False}
            finally:
                if lease_held:
                    try:
                        from core import shared_lease as _shared_lease
                        _shared_lease.release(lease_name)
                    except Exception:
                        pass

        try:
            if local_only:
                # 관리자가 이 서버에서 요청한 수동 캐싱 — 워커로 넘기지 않는다.
                # 오프로드된 빌드는 중단 신호가 워커까지 전파되지 않아 "중단을
                # 눌러도 안 멈춘다" 로 보인다. 로컬 실행은 should_cancel 이
                # 그대로 배선되므로 중단이 즉시 먹는다.
                res = _local_build()
            else:
                # v9.4.x: 개발서버(워커) 생존 시 빌드를 워커로 오프로드 — 결과는 공유
                # cache/split_table 파일로 돌아온다. 워커 다운/타임아웃/원격 실패면
                # 위 로컬 경로로 자동 폴백 (core.worker_dispatch.run_heavy).
                from core import worker_dispatch as _wd
                res = _wd.run_heavy(
                    "splittable_pivot_build",
                    {"product": canonical, "product_path": str(source_fp) if source_fp else ""},
                    _local_build,
                    label=f"pivot:{canonical}",
                    local_idle_only=not immediate,
                    local_fallback=bool(immediate),
                    durable=not immediate,
                    priority="normal" if immediate else "maintenance",
                    dedupe_key=f"pivot:{canonical}",
                    timeout_sec=6 * 3600.0,
                )
            ok = bool((res or {}).get("ok"))
        except Exception as exc:
            # run_heavy 자체가 터지면 예전에는 스레드만 조용히 죽어 화면에
            # 아무 흔적도 남지 않았다.
            logger.warning(f"pivot cache build dispatch failed for {canonical} ({reason}): {exc}")
            _cache_build_emit(canonical, f"[Pivot캐시] 빌드 실패 — {exc}", ok=False,
                        detail={"reason": reason, "error": str(exc),
                                "stage": _stage("pivot", "fail")})
        finally:
            with _PIVOT_BUILD_LOCK:
                _PIVOT_BUILD_INPROGRESS.discard(canonical)
                _PIVOT_BUILD_LAST[canonical] = time.time()
        elapsed = round(time.time() - started_ts, 1)
        if ok:
            _cache_build_emit(canonical, f"[Pivot캐시] 빌드 완료 — {elapsed}s",
                        detail={"reason": reason, "elapsed_sec": elapsed,
                                "stage": _stage("pivot", "done")})
            # 새 pivot 파일은 view payload cache 의존 시그니처에 잡히지 않으므로
            # 빌드 완료 시점에 명시적으로 비운다.
            _clear_split_view_cache_product(canonical)
        elif cancel_task_id and _scan_cancel_requested_for(cancel_task_id):
            # 관리자 중단은 실패가 아니다 — "생성된 캐시 없음" 으로 뭉뚱그리면
            # 중단을 눌렀는데 빌드가 깨진 것처럼 읽힌다. 이미 쓴 root 는 그대로
            # 서빙되고 남은 root 는 다음 빌드가 이어받는다.
            _cache_build_emit(canonical,
                        f"[Pivot캐시] 빌드 중단됨 — 관리자 요청 ({elapsed}s) · "
                        "이미 만들어진 캐시는 유지되고 남은 랏은 다음 빌드가 이어받습니다",
                        ok=False, detail={"reason": reason, "elapsed_sec": elapsed,
                                          "cancelled": True,
                                          "stage": _stage("pivot", "fail")})
        else:
            # 실패 사유는 위에서 이미 남겼을 수 있지만(lease/예외), 빌더가 조용히
            # False 만 돌려준 경우에도 반드시 한 줄은 남는다.
            _cache_build_emit(canonical, f"[Pivot캐시] 빌드 종료 — 생성된 캐시 없음 ({elapsed}s)",
                        ok=False, detail={"reason": reason, "elapsed_sec": elapsed,
                                          "stage": _stage("pivot", "fail")})

    threading.Thread(target=_run, daemon=True, name=f"splittable-pivot-{canonical}").start()
    logger.info(f"pivot cache build queued for {canonical} ({reason})")
    return True


def _pivot_cache_needs_build(product: str, source_fp: Path) -> bool:
    """Cheap completeness check; the worker performs the expensive fingerprint."""
    try:
        out_dir = _pivot_cache_path(product, "__probe__").parent
        fingerprint_fp = out_dir / ".root_fingerprints.json"
        if not fingerprint_fp.is_file():
            return True
        if source_fp.stat().st_mtime > fingerprint_fp.stat().st_mtime:
            return True
        expected = {
            str(value or "").strip()
            for value in (
                _ml_table_lookup.read_candidate_index(source_fp).get("root_lot_ids") or []
            )
            if str(value or "").strip()
        }
        if not expected:
            return False
        meta = json.loads(fingerprint_fp.read_text(encoding="utf-8"))
        built = {str(value or "").strip() for value in (meta.get("roots") or {})}
        if not expected.issubset(built):
            return True
        return sum(1 for _path in out_dir.glob("*.parquet")) < len(expected)
    except Exception:
        return True


def _auto_product_cache_role_key() -> str:
    return "dev" if _ml_table_lookup._root_ram_cache_use_dev() else "prod"


def _auto_product_cache_enabled() -> bool:
    """자동 제품 순환 캐싱 사용 여부. **기본은 꺼짐**(2026-08-04).

    예전 기본값은 켜짐이었다 — 반입 직후 아무도 켜지 않았는데 전 제품 순환이
    돌기 시작해, 관리자가 상황을 파악하기도 전에 무거운 빌드가 서버를 물고
    있었다. 자동 캐싱은 관리자가 톱니바퀴에서 명시적으로 켜는 기능으로 둔다.
    """
    raw = os.environ.get("FLOW_SPLITTABLE_AUTO_PRODUCT_CACHE_ENABLED")
    if raw is not None and str(raw).strip() != "":
        return str(raw).strip().lower() in {"1", "true", "yes", "on", "enabled"}
    try:
        from core import cache_settings
        return cache_settings.get_bool_role(
            "auto_product_cache_enabled",
            _ml_table_lookup._root_ram_cache_use_dev(),
            False,
        )
    except Exception:
        return False


def _auto_product_cache_interval_minutes() -> int:
    raw = os.environ.get("FLOW_SPLITTABLE_AUTO_PRODUCT_CACHE_INTERVAL_MINUTES")
    if raw is not None and str(raw).strip() != "":
        try:
            return int(max(1, min(1440, float(raw))))
        except Exception:
            pass
    try:
        from core import cache_settings
        value = cache_settings.get_int_role(
            "auto_product_cache_interval_minutes",
            _ml_table_lookup._root_ram_cache_use_dev(),
            _AUTO_PRODUCT_CACHE_DEFAULT_INTERVAL_MINUTES,
        )
        return int(max(1, min(1440, int(value))))
    except Exception:
        return _AUTO_PRODUCT_CACHE_DEFAULT_INTERVAL_MINUTES


def _auto_product_cache_state_path() -> Path:
    return PLAN_DIR / "auto_product_cache_schedule.json"


def _auto_product_cache_load_cursor() -> None:
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        if _AUTO_PRODUCT_CACHE_STATE.get("loaded"):
            return
        _AUTO_PRODUCT_CACHE_STATE["loaded"] = True
    try:
        saved = load_json(_auto_product_cache_state_path(), {})
        role_state = saved.get(_auto_product_cache_role_key()) or {}
    except Exception:
        role_state = {}
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        _AUTO_PRODUCT_CACHE_STATE["last_product"] = str(role_state.get("last_product") or "")
        _AUTO_PRODUCT_CACHE_STATE["last_at"] = str(role_state.get("last_at") or "")
        _AUTO_PRODUCT_CACHE_STATE["last_ok"] = role_state.get("last_ok")


def _auto_product_cache_save_cursor() -> None:
    try:
        saved = load_json(_auto_product_cache_state_path(), {})
        if not isinstance(saved, dict):
            saved = {}
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            saved[_auto_product_cache_role_key()] = {
                "last_product": _AUTO_PRODUCT_CACHE_STATE.get("last_product") or "",
                "last_at": _AUTO_PRODUCT_CACHE_STATE.get("last_at") or "",
                "last_ok": _AUTO_PRODUCT_CACHE_STATE.get("last_ok"),
            }
        save_json(_auto_product_cache_state_path(), saved)
    except Exception:
        logger.debug("auto product cache cursor save failed", exc_info=True)


def _next_auto_cache_product(products: list[str], last_product: str = "") -> str:
    order = list(dict.fromkeys(str(value or "").strip() for value in products if str(value or "").strip()))
    if not order:
        return ""
    last = str(last_product or "").strip().casefold()
    for index, product in enumerate(order):
        if product.casefold() == last:
            return order[(index + 1) % len(order)]
    return order[0]


def _auto_product_cache_set_next(delay_sec: float, *, delayed: bool = False,
                                 reason: str = "") -> None:
    when = time.time() + max(0.0, float(delay_sec))
    products = _match_cache_products("")
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        last = _AUTO_PRODUCT_CACHE_STATE.get("last_product") or ""
        _AUTO_PRODUCT_CACHE_STATE.update({
            "next_product": _next_auto_cache_product(products, last),
            "next_at_ts": when,
            "next_at": _auto_product_cache_iso(when),
            "delayed": bool(delayed),
            "delayed_reason": str(reason or ""),
        })


def _auto_product_cache_on_started(product: str, job_id: str) -> None:
    now = time.time()
    next_product = _next_auto_cache_product(_match_cache_products(""), product)
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        # The configured interval is the pause after a complete catalog sweep,
        # not a gap between products.  While a product is active, estimate the
        # next product from the latest observed whole-product duration.
        estimate = max(
            60.0,
            float(_AUTO_PRODUCT_CACHE_STATE.get("last_duration_sec") or 0.0)
            or _auto_product_cache_interval_minutes() * 60.0,
        )
        next_at = now + estimate
        _AUTO_PRODUCT_CACHE_STATE.update({
            "current_product": product,
            "current_started_ts": now,
            "queued_product": "",
            "queued_task_id": "",
            "next_product": next_product,
            # Earliest prediction while a product is running.  The status
            # snapshot keeps moving this floor forward if the current stage
            # runs longer, so the UI never advertises a time already passed.
            "next_at": _auto_product_cache_iso(next_at),
            "next_at_ts": next_at,
            "job_id": job_id,
            "delayed": False,
            "delayed_reason": "",
        })
    _SEARCH_CACHE_MAINT_WAKE.set()


def _auto_product_cache_advance_after_product(product: str, *, ok: bool) -> None:
    """Advance inside one continuous catalog sweep.

    Products inside a sweep start back-to-back.  The configured interval is
    applied only after the cursor wraps to the sweep's first product.
    """
    now = time.time()
    products = _match_cache_products("")
    next_product = _next_auto_cache_product(products, product)
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        first = str(_AUTO_PRODUCT_CACHE_STATE.get("cycle_first_product") or product)
        started = float(_AUTO_PRODUCT_CACHE_STATE.get("current_started_ts") or 0.0)
        duration = max(0.0, now - started) if started else 0.0
        completed = int(_AUTO_PRODUCT_CACHE_STATE.get("cycle_completed_products") or 0) + 1
        wrapped = (
            not next_product
            or next_product.casefold() == first.casefold()
            or len(products) <= 1
            or completed >= max(1, len(products))
        )
        delay = _auto_product_cache_interval_minutes() * 60.0 if wrapped else 0.0
        _AUTO_PRODUCT_CACHE_STATE.update({
            "current_product": "",
            "current_started_ts": 0.0,
            "queued_product": "",
            "queued_task_id": "",
            "last_product": product,
            "last_at": _auto_product_cache_iso(now),
            "last_ok": bool(ok),
            "job_id": "",
            "next_product": next_product,
            "next_at_ts": now + delay,
            "next_at": _auto_product_cache_iso(now + delay),
            "cycle_first_product": "" if wrapped else first,
            "cycle_completed_products": 0 if wrapped else completed,
            "last_duration_sec": duration or float(
                _AUTO_PRODUCT_CACHE_STATE.get("last_duration_sec") or 0.0),
            "delayed": False,
            "delayed_reason": "",
        })
    _auto_product_cache_save_cursor()
    _SEARCH_CACHE_MAINT_WAKE.set()


def _auto_product_cache_on_finished(product: str, result: dict) -> None:
    _auto_product_cache_advance_after_product(
        product, ok=bool((result or {}).get("ok")))


def _auto_product_cache_on_cancelled(product: str, task_id: str = "") -> None:
    """Keep the rotation cursor in sync when its pending queue item is removed.

    A cancelled automatic item must not remain as the UI's current/next item and
    must not remain stuck in the UI. Treat it as a visited cursor entry and
    continue the same catalog sweep with the following product; the configured
    interval is used only after the sweep wraps.
    """
    product = str(product or "").strip()
    task_id = str(task_id or "").strip()
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        queued_id = str(_AUTO_PRODUCT_CACHE_STATE.get("queued_task_id") or "")
        queued_product = str(_AUTO_PRODUCT_CACHE_STATE.get("queued_product") or "")
        if task_id and queued_id and task_id != queued_id:
            return
        if product and queued_product and product != queued_product:
            return
        target = product or queued_product
        _AUTO_PRODUCT_CACHE_STATE.update({
            "queued_product": "",
            "queued_task_id": "",
        })
    _auto_product_cache_advance_after_product(target, ok=False)


def _auto_product_cache_schedule_snapshot() -> dict:
    _auto_product_cache_load_cursor()
    enabled = _auto_product_cache_enabled()
    interval = _auto_product_cache_interval_minutes()
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        state = dict(_AUTO_PRODUCT_CACHE_STATE)
    queued_task_id = str(state.get("queued_task_id") or "")
    if queued_task_id:
        gate = _scan_gate_snapshot()
        gate_ids = {
            str(row.get("id") or "")
            for row in [gate.get("current"), *(gate.get("pending") or [])]
            if isinstance(row, dict)
        }
        # scan_gate가 5분 무진행 pending을 자동 제거했으면 순환 상태도 함께
        # 비운다. 그렇지 않으면 화면과 스케줄러만 영원히 queued로 남는다.
        if queued_task_id not in gate_ids:
            _auto_product_cache_on_cancelled(
                str(state.get("queued_product") or ""), queued_task_id,
            )
            with _AUTO_PRODUCT_CACHE_STATE_LOCK:
                state = dict(_AUTO_PRODUCT_CACHE_STATE)
    active_product = str(state.get("current_product") or state.get("queued_product") or "")
    if enabled and active_product:
        # Products in the same sweep are back-to-back.  Predict the next start
        # from the last whole-product duration and keep moving an expired
        # estimate forward while the current four-stage pipeline is still live.
        now = time.time()
        estimate = max(
            60.0,
            float(state.get("last_duration_sec") or 0.0) or interval * 60.0,
        )
        started = float(state.get("current_started_ts") or 0.0)
        floor_ts = (started + estimate) if started else (now + estimate)
        if floor_ts <= now:
            floor_ts = now + 60.0
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            _AUTO_PRODUCT_CACHE_STATE["next_product"] = _next_auto_cache_product(
                _match_cache_products(""), active_product)
            if float(_AUTO_PRODUCT_CACHE_STATE.get("next_at_ts") or 0.0) < floor_ts:
                _AUTO_PRODUCT_CACHE_STATE["next_at_ts"] = floor_ts
                _AUTO_PRODUCT_CACHE_STATE["next_at"] = _auto_product_cache_iso(floor_ts)
            state = dict(_AUTO_PRODUCT_CACHE_STATE)
    if enabled and not state.get("current_product") and not state.get("queued_product") and not state.get("next_product"):
        _auto_product_cache_set_next(0.0)
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            state = dict(_AUTO_PRODUCT_CACHE_STATE)
    return {
        "key": "product_rotation",
        "label": "제품별 필수 캐시 순환",
        "started": bool(_SEARCH_CACHE_MAINT_THREAD and _SEARCH_CACHE_MAINT_THREAD.is_alive()),
        "enabled": enabled,
        "interval_minutes": interval,
        "current_product": str(state.get("current_product") or ""),
        "queued_product": str(state.get("queued_product") or ""),
        "queued_task_id": str(state.get("queued_task_id") or ""),
        "next_product": str(state.get("next_product") or ""),
        "next_at": str(state.get("next_at") or ""),
        "last_product": str(state.get("last_product") or ""),
        "last_at": str(state.get("last_at") or ""),
        "last_ok": state.get("last_ok"),
        "job_id": str(state.get("job_id") or ""),
        "cycle_active": bool(state.get("cycle_first_product")),
        "cycle_first_product": str(state.get("cycle_first_product") or ""),
        "cycle_completed_products": int(state.get("cycle_completed_products") or 0),
        "cycle_total_products": len(_match_cache_products("")),
        "next_after_current": bool(active_product),
        "delayed": bool(state.get("delayed")),
        "delayed_reason": str(state.get("delayed_reason") or ""),
        "serial_policy": "one_product_one_cache_kind_per_server",
    }


def _split_search_cache_maintenance_loop() -> None:
    """Run one product's required cache pipeline at a time in rotation order."""
    _auto_product_cache_load_cursor()
    try:
        first_delay = float(os.environ.get("FLOW_SPLITTABLE_SEARCH_CACHE_FIRST_DELAY_SEC", "") or 20.0)
    except Exception:
        first_delay = 20.0
    _auto_product_cache_set_next(max(0.0, min(600.0, first_delay)))
    while True:
        _SEARCH_CACHE_MAINT_WAKE.clear()
        try:
            from core.background_owner import is_owner
            if not is_owner():
                _SEARCH_CACHE_MAINT_WAKE.wait(5.0)
                continue
        except Exception:
            _SEARCH_CACHE_MAINT_WAKE.wait(5.0)
            continue
        if not _auto_product_cache_enabled():
            with _AUTO_PRODUCT_CACHE_STATE_LOCK:
                _AUTO_PRODUCT_CACHE_STATE.update({
                    "next_product": "", "next_at": "", "next_at_ts": 0.0,
                    "delayed": False, "delayed_reason": "",
                })
            _SEARCH_CACHE_MAINT_WAKE.wait(30.0)
            continue
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            state = dict(_AUTO_PRODUCT_CACHE_STATE)
        if state.get("current_product") or state.get("queued_product"):
            _SEARCH_CACHE_MAINT_WAKE.wait(5.0)
            continue
        if not state.get("next_product"):
            _auto_product_cache_set_next(0.0)
            with _AUTO_PRODUCT_CACHE_STATE_LOCK:
                state = dict(_AUTO_PRODUCT_CACHE_STATE)
        next_at_ts = float(state.get("next_at_ts") or 0.0)
        now = time.time()
        if next_at_ts > now:
            _SEARCH_CACHE_MAINT_WAKE.wait(min(30.0, next_at_ts - now))
            continue
        product = str(state.get("next_product") or "")
        if not product:
            _auto_product_cache_set_next(60.0, delayed=True, reason="제품 목록 없음")
            _SEARCH_CACHE_MAINT_WAKE.wait(30.0)
            continue
        # 예약 시각이 오면 다른 작업이 실행 중이어도 실제 공용 FIFO 큐에 넣는다.
        # 예전처럼 메모리 상태로만 15초씩 미루면 화면의 "다음 작업"과 실제 큐가
        # 서로 달랐고, 취소할 task id도 없었다.
        with _AUTO_PRODUCT_CACHE_STATE_LOCK:
            following = _next_auto_cache_product(_match_cache_products(""), product)
            now = time.time()
            estimate = max(
                60.0,
                float(_AUTO_PRODUCT_CACHE_STATE.get("last_duration_sec") or 0.0)
                or _auto_product_cache_interval_minutes() * 60.0,
            )
            if not _AUTO_PRODUCT_CACHE_STATE.get("cycle_first_product"):
                _AUTO_PRODUCT_CACHE_STATE["cycle_first_product"] = product
                _AUTO_PRODUCT_CACHE_STATE["cycle_completed_products"] = 0
            _AUTO_PRODUCT_CACHE_STATE.update({
                "queued_product": product,
                "queued_task_id": "",
                "next_product": following,
                "next_at": _auto_product_cache_iso(now + estimate),
                "next_at_ts": now + estimate,
                "delayed": False,
                "delayed_reason": "",
            })
        out = _submit_product_cache_scan(
            product,
            force=False,
            source="scheduler",
            on_started=_auto_product_cache_on_started,
            on_finished=_auto_product_cache_on_finished,
        )
        accepted = bool(out.get("accepted"))
        if accepted:
            # 바로 시작한 경우 on_started 가 먼저 queued_product 를 비울 수 있다.
            # 그때 완료된 callback 상태 위에 오래된 task id를 다시 덮지 않는다.
            with _AUTO_PRODUCT_CACHE_STATE_LOCK:
                if (_AUTO_PRODUCT_CACHE_STATE.get("queued_product") == product
                        and not _AUTO_PRODUCT_CACHE_STATE.get("current_product")):
                    _AUTO_PRODUCT_CACHE_STATE["queued_task_id"] = str(out.get("id") or "")
        if not accepted:
            with _AUTO_PRODUCT_CACHE_STATE_LOCK:
                _AUTO_PRODUCT_CACHE_STATE.update({
                    "queued_product": "", "queued_task_id": "",
                })
            _auto_product_cache_set_next(
                _AUTO_PRODUCT_CACHE_RETRY_SEC,
                delayed=True,
                reason=str(out.get("detail") or "캐시 실행 큐 대기"),
            )
        _SEARCH_CACHE_MAINT_WAKE.wait(5.0)


def start_split_search_cache_maintainer() -> bool:
    """Start the single product-rotation cache scheduler for this server."""
    global _SEARCH_CACHE_MAINT_THREAD
    with _SEARCH_CACHE_MAINT_LOCK:
        if _SEARCH_CACHE_MAINT_THREAD is not None and _SEARCH_CACHE_MAINT_THREAD.is_alive():
            return False
        _SEARCH_CACHE_MAINT_THREAD = threading.Thread(
            target=_split_search_cache_maintenance_loop,
            daemon=True,
            name="splittable-search-cache-maintainer",
        )
        _SEARCH_CACHE_MAINT_THREAD.start()
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        _AUTO_PRODUCT_CACHE_STATE["started"] = True
    return True


try:
    _ml_table_lookup.register_build_complete_hook(
        lambda fp: _enqueue_pivot_cache_build(Path(fp).stem, reason="lookup_complete")
    )
except Exception:
    logger.debug("lookup-complete pivot hook registration failed", exc_info=True)


# ── Per-root FAB latest-lot index (additive fast layer for the fab join) ──────
# Profiling (5000-root / 20M-row FAB sandbox) pinned the dominant SplitTable
# cost to a single collect() in the fab override join: picking the latest FAB
# lot per (root_lot_id, wafer_id) scanned the WHOLE FAB source on every search.
# Two things defeat parquet pruning there: the source is not partitioned by root,
# and the root filter is wrapped in _join_key_expr (cast+upper), so predicate
# pushdown cannot use row-group stats. That collect was ~2.5s and grew with FAB
# size — it fired even on pivot-cache hits whenever the lot_progress projection
# cache did not cover the searched root.
#
# This layer precomputes, in the background, the global FAB source re-partitioned
# by a normalized root key. A search reads only the one root's partition (a few
# hundred rows) and the EXISTING align/scope/latest-pick/join logic then runs
# unchanged on that tiny frame. Purely additive: on any miss / staleness / error
# it returns None and the caller falls back to the original full-scan path while
# a (re)build is scheduled. Does not touch the SWR/signature/scan_root_lot_cache
# paths. Measured per-root read: ~13–40ms vs ~2130ms full-scan (~50–160x).
_FAB_IDX_ROOT_COL = "__fab_idx_root"
_FAB_IDX_META_FILE = "_meta.json"
_FAB_IDX_BUILD_LOCK = threading.Lock()


def _heavy_build_concurrency() -> int:
    # 운영 계약: 한 서버에서 캐시 산출은 한 제품의 한 종류만 실행한다. 배포
    # 환경에 과거 튜닝값이 남아 있어도 이 경로만 병렬로 빠져나가지 않게 고정한다.
    return 1


# 무거운 백그라운드 빌드(제품 pivot, fab index)는 전역 직렬화한다 — 스윕/뷰가
# 여러 product 를 동시에 enqueue 해도 빌드 피크(각 수백 MB)가 겹쳐 쌓이지
# 않게 한다. 배치/청크 단위 메모리 가드와 함께 12~20GB 호스트에서 총 상한을
# "빌드 1개 피크" 로 유지하는 장치.
_HEAVY_BUILD_SEMAPHORE = threading.Semaphore(_heavy_build_concurrency())
_FAB_IDX_BUILD_INPROGRESS: set[str] = set()
_FAB_IDX_BUILD_LAST: dict[str, float] = {}
_FAB_IDX_BUILD_COOLDOWN_SEC = 120.0
# central revalidator (startup service) — keeps every built index in line with
# the FAB sources without any staleness work on the search hot path
_FAB_IDX_SWEEP_THREAD_LOCK = threading.Lock()
_FAB_IDX_SWEEP_THREAD: threading.Thread | None = None
_FAB_IDX_SWEEP_WAKE = threading.Event()
_FAB_IDX_SWEEP_FIRST_DELAY_SEC = 10.0


def _fab_lot_index_enabled() -> bool:
    return _env_bool("FLOW_SPLITTABLE_FAB_LOT_INDEX", True)


def _fab_lot_index_dir(product: str) -> Path:
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    return _base_root() / "cache" / "fab_lot_index" / canonical


def _is_fab_lot_index_staging_product(product: str) -> bool:
    """True when a `<product>.build` staging path was mistaken for a product."""
    raw = str(product or "").strip().rstrip("/\\")
    return bool(raw) and raw.casefold().endswith(".build")


def _fab_lot_index_meta_path(product: str) -> Path:
    return _fab_lot_index_dir(product) / _FAB_IDX_META_FILE


def _fab_source_signature(fab_source: str, include_all: bool) -> list:
    """(path, mtime, size) for every FAB source file — the index staleness key."""
    sig: list = []
    seen_targets: set[str] = set()
    for source in _global_fab_source_paths(fab_source, include_all=include_all):
        # Keep signature discovery identical to the scanner. Directly joining
        # db_base/source loses case-insensitive matches on Linux and misses a
        # usable alias/source mounted under base_root.
        base, _resolved_source = _resolve_fab_source_target(source)
        if base is None:
            continue
        target_key = _canon_file_key(base)
        if target_key in seen_targets:
            continue
        seen_targets.add(target_key)
        try:
            if base.is_file():
                st = base.stat()
                sig.append([str(base), int(st.st_mtime), int(st.st_size)])
                continue
            for p in sorted(base.rglob("*")):
                if p.is_file() and p.suffix.lower() in (".parquet", ".csv"):
                    st = p.stat()
                    sig.append([str(p), int(st.st_mtime), int(st.st_size)])
        except Exception:
            continue
    return sig


def _fab_lot_index_read_meta(product: str) -> dict:
    try:
        return load_json(_fab_lot_index_meta_path(product), {}) or {}
    except Exception:
        return {}


def _fab_lot_index_partition_dir(product: str, root_lot_id: str) -> Path | None:
    root = str(root_lot_id or "").strip().upper()
    if not root:
        return None
    part = _fab_lot_index_dir(product) / f"{_FAB_IDX_ROOT_COL}={root}"
    return part if part.is_dir() else None


def _fab_lot_index_sweep_interval_sec() -> float:
    try:
        value = float(os.environ.get("FLOW_SPLITTABLE_FAB_LOT_INDEX_SWEEP_SEC", "") or 60.0)
    except Exception:
        value = 60.0
    return max(10.0, min(3600.0, value))


def _fab_lot_index_sweep_once() -> None:
    """Compare every built index against the live FAB source signature and
    enqueue rebuilds on drift. One signature walk is shared by all products
    that resolve to the same (fab_source, include_all) source set."""
    base = _base_root() / "cache" / "fab_lot_index"
    products: set[str] = set()
    try:
        # `<product>.build` 는 중단 시 이어받기용 임시 산출물이다. 제품으로
        # 다시 넣으면 다음 staging이 `.BUILD.BUILD`로 계속 증식한다.
        products.update(
            p.name for p in base.iterdir()
            if p.is_dir() and not _is_fab_lot_index_staging_product(p.name)
        )
    except Exception:
        pass
    # A sweep that only visits existing index directories can never create the
    # first index. Include configured products so a fresh deployment prepares
    # every root before the first operator searches it.
    try:
        cfg = load_json(SOURCE_CFG, {}) or {}
        products.update(
            _canonical_mltable_product_name(value, allow_bare=True)
            or str(value or "").strip().upper()
            for value in (cfg.get("enabled") or [])
            if str(value or "").strip()
        )
    except Exception:
        pass
    products.discard("")
    if not products:
        return
    include_all = _foreground_global_fab_scan_enabled()
    sig_memo: dict[tuple, list] = {}
    for product in sorted(products):
        try:
            ml_product, _ov, fab_source = _current_fab_override(product)
            if not ml_product:
                continue
            key = (fab_source, include_all)
            if key not in sig_memo:
                sig_memo[key] = _fab_source_signature(fab_source, include_all)
            meta = _fab_lot_index_read_meta(product)
            if not meta or meta.get("source_sig") != sig_memo[key]:
                _enqueue_fab_lot_index_build(
                    product, fab_source, include_all=include_all,
                    reason="sweep_missing_meta" if not meta else "sweep_stale",
                )
        except Exception:
            logger.debug("fab_lot_index sweep failed product=%s", product, exc_info=True)


def _fab_lot_index_sweep_loop() -> None:
    _FAB_IDX_SWEEP_WAKE.wait(_FAB_IDX_SWEEP_FIRST_DELAY_SEC)
    while True:
        _FAB_IDX_SWEEP_WAKE.clear()
        try:
            if _fab_lot_index_enabled():
                _fab_lot_index_sweep_once()
        except Exception:
            logger.debug("fab_lot_index sweep tick failed", exc_info=True)
        _FAB_IDX_SWEEP_WAKE.wait(_fab_lot_index_sweep_interval_sec())


def start_fab_lot_index_revalidator() -> bool:
    """Retired: FAB refresh is owned by the product-rotation scheduler.

    Keeping an independent FAB timer would allow a second cache kind/product to
    enter the queue while the scheduled product pipeline is running.  The sweep
    helper remains callable for diagnostics and tests, but startup no longer
    creates a per-cache scheduler thread.
    """
    logger.info("fab_lot_index independent scheduler retired; product rotation owns refresh")
    return False


def notify_fab_sources_changed(reason: str = "") -> None:
    """Move the next product-rotation check forward after FAB source ingest."""
    logger.info("fab sources changed (%s) — product cache rotation waked", reason or "-")
    with _AUTO_PRODUCT_CACHE_STATE_LOCK:
        if not (_AUTO_PRODUCT_CACHE_STATE.get("current_product") or
                _AUTO_PRODUCT_CACHE_STATE.get("queued_product")):
            _AUTO_PRODUCT_CACHE_STATE["next_at_ts"] = time.time()
            _AUTO_PRODUCT_CACHE_STATE["next_at"] = _auto_product_cache_iso(time.time())
    _SEARCH_CACHE_MAINT_WAKE.set()


def _fab_lot_index_scan_root(product: str, root_lot_id: str,
                             fab_source: str = "", include_all: bool = False):
    """Return a LazyFrame of the FAB source rows for one root (schema identical to
    _scan_global_fab_sources), or None to signal fallback to the full scan.
    Serve-immediately: freshness is maintained by the central revalidator sweep,
    so the search hot path does no staleness work at all."""
    if not _fab_lot_index_enabled():
        return None
    root = str(root_lot_id or "").strip()
    if not root:
        return None
    try:
        part = _fab_lot_index_partition_dir(product, root)
        if part is None:
            return None
        files = sorted(part.glob("*.parquet"))
        if not files:
            return None
        lf = _scan_parquet_compat([str(p) for p in files])
        names = lf.collect_schema().names()
        if _FAB_IDX_ROOT_COL in names:
            lf = lf.drop(_FAB_IDX_ROOT_COL)
        return _cast_cats_lazy(lf)
    except Exception:
        logger.debug("fab_lot_index scan failed product=%s root=%s", product, root, exc_info=True)
        return None


def _fab_source_sig_delta(old_sig, new_sig) -> set[str] | None:
    """ADDED file keys between two source signatures, or None when any file was
    removed/rewritten (→ 전체 재빌드 필요)."""
    try:
        old_map = {_canon_file_key(p): (int(m), int(s)) for p, m, s in (old_sig or [])}
        new_map = {_canon_file_key(p): (int(m), int(s)) for p, m, s in (new_sig or [])}
    except Exception:
        return None
    if not old_map:
        return None
    for key, sig in old_map.items():
        if new_map.get(key) != sig:
            return None
    return {k for k in new_map if k not in old_map}


def _build_fab_lot_index(product: str, fab_source: str, include_all: bool) -> bool:
    """Build a per-root FAB lot index: the latest FAB row per (root, wafer),
    partitioned by a normalized root key.

    Storing the *reduced* latest-per-(root,wafer) frame (rather than every raw
    FAB row) keeps the index tiny — a few rows per root instead of thousands —
    which makes the build I/O an order of magnitude cheaper (shorter cold window
    after a data refresh) and per-root reads near-instant. The reduction is by
    (root, wafer) keyed on the FAB timestamp, so it is equivalent to the fab
    join's own latest-pick for any join key ⊆ {root_lot_id, wafer_id} (the
    default identity join): reducing by (root,wafer)-latest and then re-picking
    latest per join key yields the same rows because the timestamp order is
    preserved. The downstream sort+unique in _scan_product still runs and stays
    correct on the tiny frame.

    FAB 원천은 보통 새 date 파티션 파일이 추가되는 append 형이다 — 기존 파일이
    그대로고 파일만 늘었으면 새 파일만 스캔해 기존 인덱스와 (root,wafer)-latest
    로 병합하고 영향받은 root 파티션만 교체한다(수 초). 파일이 지워졌거나
    재기록됐으면 전체 재빌드로 폴백한다."""
    if _is_fab_lot_index_staging_product(product):
        logger.warning("fab_lot_index staging path rejected as product: %s", product)
        return False
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip()
    if not canonical:
        return False

    job_id = ""
    try:
        from core.cache_event_log import start_job, stage_started
        job_id = start_job(
            "fab_index_background",
            f"FAB latest 인덱스 백그라운드 갱신 ({canonical})",
            [("fab_index", "FAB latest 인덱스")],
            product=canonical
        )
        stage_started(job_id, "fab_index")
    except Exception:
        pass

    try:
        live_sig = _fab_source_signature(fab_source, include_all)
        old_meta = _fab_lot_index_read_meta(canonical)
        if old_meta.get("source_sig") and _fab_lot_index_dir(canonical).is_dir():
            added = _fab_source_sig_delta(old_meta.get("source_sig"), live_sig)
            if added is not None:
                if not added:
                    # 파일 변화 없음 — 인덱스가 이미 최신. 이것도 알려야 한다.
                    # 아니면 스캔 큐에 뜬 작업이 "아무 것도 안 하고 끝난" 것처럼 보인다.
                    _fab_idx_emit(canonical, "[FAB랏인덱스] 최신 — FAB 원천 변화 없음(건너뜀)",
                                  detail={"stage": _stage("fab_index", "skip")})
                    return True
                if _build_fab_lot_index_incremental(
                        canonical, fab_source, include_all, added, live_sig, old_meta):
                    return True
                logger.info("fab_lot_index incremental 실패 — 전체 재빌드 (product=%s)", canonical)
        return _build_fab_lot_index_full(canonical, fab_source, include_all)
    finally:
        if job_id:
            try:
                from core.cache_event_log import stage_finished, finish_job
                stage_finished(job_id, "fab_index")
                finish_job(job_id)
            except Exception:
                pass

def _fab_latest_reduce_lf(fab_lf):
    """(reduced_lf, root_col) — 정규화 root 키 부여 + (root,wafer)-latest 축소.

    Reduce to the latest row per (root, wafer). Keep every FAB column so the
    read path is a drop-in replacement for _scan_global_fab_sources output.
    이미 _FAB_IDX_ROOT_COL 이 있는 프레임(파티션/병합 재축소)에도 그대로 동작."""
    try:
        names = fab_lf.collect_schema().names()
    except Exception:
        return None, ""
    root_col = _ci_resolve_in("root_lot_id", names) or _pick_first_present_ci(("root_lot_id",), names)
    if not root_col:
        return None, ""
    wf_col = _ci_resolve_in("wafer_id", names) or _pick_first_present_ci(("wafer_id", "wafer"), names)
    ts_col = _pick_ts_col(names)
    lf = (
        fab_lf
        .with_columns(_join_key_expr(root_col).alias(_FAB_IDX_ROOT_COL))
        .filter(pl.col(_FAB_IDX_ROOT_COL).is_not_null() & (pl.col(_FAB_IDX_ROOT_COL) != ""))
    )
    reduce_subset = [_FAB_IDX_ROOT_COL] + ([wf_col] if wf_col else [])
    try:
        if ts_col and ts_col in names:
            lf = lf.sort(ts_col, descending=True, nulls_last=True).unique(
                subset=reduce_subset, keep="first", maintain_order=True)
        else:
            lf = lf.unique(subset=reduce_subset, keep="last")
    except Exception:
        # If reduction is not expressible, fall back to storing raw per-root rows
        # (still correct — the fab join reduces at read time).
        logger.debug("fab_lot_index reduction skipped", exc_info=True)
    return lf, root_col


def _scan_cancel_requested() -> bool:
    """이 스레드가 실행 중인 스캔 큐 작업에 중단 요청이 왔는가.

    긴 루프의 **안전한 지점**(제품 경계·배치 경계 — 이미 메모리 가드나 사용자
    양보를 확인하는 자리)에서만 확인한다. 중간에서 끊으면 parquet 파티션이
    깨지므로 절대 강제 종료하지 않는다.
    """
    try:
        from core import scan_gate
        return bool(scan_gate.cancel_requested())
    except Exception:
        return False


def _scan_cancel_requested_for(task_id: str) -> bool:
    """**다른 스레드**에서 실행 중인 큐 작업의 중단 여부.

    인자 없는 `_scan_cancel_requested()` 는 스레드 로컬이라 데몬 빌드 스레드에서
    부르면 항상 False 다. 작업을 띄운 스레드에서 붙잡아 둔 id 로 물어볼 때 쓴다.
    """
    if not str(task_id or "").strip():
        return False
    try:
        from core import scan_gate
        return bool(scan_gate.cancel_requested(str(task_id)))
    except Exception:
        return False


def _fab_idx_log_gap_sec() -> float:
    return _float_env_clamped("FLOW_SPLITTABLE_FAB_INDEX_LOG_GAP_SEC", 1.5, 0.0, 60.0)


def _fab_idx_emit(product: str, event: str, *, ok: bool = True, detail: dict | None = None) -> None:
    """FAB 랏 인덱스 빌드 진행을 캐시 이벤트 로그로 내보낸다.

    **`logger.info` 는 화면에 안 보인다** — 캐시관리 화면(수동 스캔/스캔 큐)은
    `cache_event_log.record` 만 읽는다. 그래서 이 빌드는 스캔 큐에 이름만 뜨고
    "몇 분째 뭘 하는지" 가 전혀 안 보였다(매칭 캐시는 이미 배치별로 내보내고
    있었다). 같은 category(`cache_op`)라 '전체' 필터에 실시간으로 섞여 보인다.
    """
    try:
        from core.cache_event_log import record as _rec
        _rec("cache_op", event, ok=ok, detail=detail or {}, product=product)
    except Exception:
        pass


def _fab_idx_progress(done: int, total: int, *, state: str = "running") -> dict:
    """Standard progress payload used by the cache dashboard (batch X/Y)."""
    from core.cache_event_log import progress_detail
    return progress_detail("fab_index", done, total, state=state, unit="배치")


def _fab_index_build_batch_rows() -> int:
    # 수 GB FAB도 약 50만 행씩 축소해 총 원천 크기와 무관하게 피크를 낮춘다.
    # 캐시는 느려도 괜찮고 안정성이 우선이므로 기존 100만 행보다 보수적으로 시작.
    try:
        rows = float(os.environ.get("FLOW_SPLITTABLE_FAB_INDEX_BUILD_BATCH_ROWS", "") or 500_000)
    except Exception:
        rows = 500_000.0
    return int(max(100_000.0, min(100_000_000.0, rows)))


def _parquet_row_count(path: str) -> int | None:
    """Row count from the parquet footer only (no data read)."""
    try:
        if not str(path).lower().endswith(".parquet"):
            return None
        return int(pl.scan_parquet(str(path)).select(pl.len()).collect().item(0, 0))
    except Exception:
        return None


def _write_fab_index_partitions(product: str, lf) -> None:
    idx_dir = _fab_lot_index_dir(product)
    tmp_dir = idx_dir.with_name(idx_dir.name + ".tmp")
    if tmp_dir.exists():
        shutil.rmtree(tmp_dir, ignore_errors=True)
    tmp_dir.mkdir(parents=True, exist_ok=True)
    try:
        sink_target = pl.PartitionBy(
            tmp_dir, key=_FAB_IDX_ROOT_COL, include_key=True,
            approximate_bytes_per_file="auto",
        )
        lf.sink_parquet(sink_target, mkdir=True, maintain_order=False)
    except Exception:
        # Older polars / sink edge cases — fall back to an eager partitioned write.
        df = lf.collect()
        if tmp_dir.exists():
            shutil.rmtree(tmp_dir, ignore_errors=True)
        tmp_dir.mkdir(parents=True, exist_ok=True)
        if df.height:
            df.write_parquet(tmp_dir, partition_by=_FAB_IDX_ROOT_COL)
    # 디렉터리 전체를 지웠다가 바꾸면 그 짧은 틈에 모든 검색이 full FAB scan으로
    # 떨어진다. 완성된 root 파티션만 하나씩 교체하고 meta는 호출측이 마지막에
    # publish한다. 빌드 중에는 기존 인덱스가 계속 읽힌다.
    idx_dir.mkdir(parents=True, exist_ok=True)
    new_names = {p.name for p in tmp_dir.iterdir() if p.is_dir()}
    for old in list(idx_dir.iterdir()):
        if old.is_dir() and old.name.startswith(f"{_FAB_IDX_ROOT_COL}=") and old.name not in new_names:
            shutil.rmtree(old, ignore_errors=True)
    for child in list(tmp_dir.iterdir()):
        target = idx_dir / child.name
        if target.exists():
            if target.is_dir():
                shutil.rmtree(target, ignore_errors=True)
            else:
                target.unlink()
        child.replace(target)
    shutil.rmtree(tmp_dir, ignore_errors=True)


def _build_fab_lot_index_full(product: str, fab_source: str, include_all: bool) -> bool:
    """전체 재빌드 — FAB 파일을 row 수 기준 배치로 나눠 배치별로 latest
    축소 후 병합한다. (root,wafer)-latest 는 결합법칙이 성립하므로 결과는 단일
    패스 전역 sort 와 동일하고(동률 ts 는 앞선 파일 우선 — stable sort 와 일치),
    피크 메모리가 FAB 전체 크기가 아니라 배치 크기에 비례한다. row 수는
    parquet footer 만 읽어 세므로(데이터 미접근) 배치 산정은 저렴하고, 축소
    결과는 root×wafer 로 상한되는 작은 프레임이라 누적 병합 비용도 미미하다."""
    live_sig = _fab_source_signature(fab_source, include_all)
    batch_rows = _fab_index_build_batch_rows()
    batches: list[list[str] | None] = []
    cur: list[str] = []
    cur_rows = 0
    for path, _mtime, _size in live_sig:
        n = _parquet_row_count(path)
        cur.append(path)
        # row 수를 모르는 파일(CSV 등)은 보수적으로 배치를 끊는다.
        cur_rows += n if n is not None else batch_rows
        if cur_rows >= batch_rows:
            batches.append(cur)
            cur, cur_rows = [], 0
    if cur:
        batches.append(cur)
    if not batches:
        # 시그니처가 파일을 못 찾는 엣지 소스 — 필터 없는 전체 스캔 1배치.
        batches = [None]

    total_batches = len(batches)
    build_sig = hashlib.sha256(json.dumps(
        {"source_sig": live_sig, "batch_rows": batch_rows, "total": total_batches},
        ensure_ascii=False, sort_keys=True, default=str,
    ).encode("utf-8")).hexdigest()
    idx_parent = _fab_lot_index_meta_path(product).parent
    staging = idx_parent.with_name(idx_parent.name + ".build")
    manifest_fp = staging / "_resume.json"
    manifest = load_json(manifest_fp, {}) if manifest_fp.is_file() else {}
    if str(manifest.get("build_sig") or "") != build_sig:
        shutil.rmtree(staging, ignore_errors=True)
        manifest = {}
    staging.mkdir(parents=True, exist_ok=True)
    completed = {
        int(v) for v in (manifest.get("completed") or [])
        if str(v).isdigit() and (staging / f"{int(v):06d}.parquet").is_file()
    }
    used_all: list[str] = [str(v) for v in (manifest.get("used") or [])]
    root_col_meta = str(manifest.get("root_col") or "")
    reduced_rows = int(manifest.get("reduced_rows") or 0)
    batch_failures: list[str] = []
    started = time.time()
    last_log = 0.0
    _fab_idx_emit(
        product,
        f"[FAB랏인덱스] 전체 재빌드 시작 — FAB 파일 {len(live_sig):,}개 · "
        f"배치 {total_batches:,}개(배치당 최대 {batch_rows:,}행) · "
        f"이어받기 {len(completed):,}개 · RSS {_proc_rss_gb()}GB",
        detail={
            "stage": _stage("fab_index", "start"),
            "progress": _fab_idx_progress(len(completed), total_batches),
        },
    )
    for bi, batch in enumerate(batches, 1):
        part_fp = staging / f"{bi:06d}.parquet"
        if bi in completed and part_fp.is_file():
            continue
        # 배치 경계 = 안전한 취소 지점. 완료 배치는 staging에 남겨 다음 실행에서
        # 이어받고, 공개 인덱스는 meta publish 전까지 건드리지 않는다.
        if _scan_cancel_requested():
            _fab_idx_emit(product,
                          f"[FAB랏인덱스] 중단됨 — 배치 {bi - 1:,}/{total_batches:,} 에서 접습니다. "
                          f"완료 배치는 보존해 다음 스캔에서 이어받습니다.",
                          ok=False,
                          detail={"progress": _fab_idx_progress(
                              bi - 1, total_batches, state="cancelled")})
            return False
        only = {_canon_file_key(p) for p in batch} if batch is not None else None
        fab_lf, used = _scan_global_fab_sources(fab_source, include_all=include_all,
                                                only_files=only)
        if fab_lf is None:
            batch_failures.append(f"batch {bi}: source scan returned no frame")
            continue
        reduced, root_col = _fab_latest_reduce_lf(fab_lf)
        if reduced is None:
            try:
                cols = fab_lf.collect_schema().names()
            except Exception:
                cols = []
            batch_failures.append(
                f"batch {bi}: root_lot_id column missing (columns={cols[:20]})"
            )
            continue
        root_col_meta = root_col_meta or root_col
        try:
            batch_df = reduced.collect()
        except Exception as e:
            logger.warning("fab_lot_index 배치 축소 실패 (product=%s) %s: %s",
                           product, type(e).__name__, e)
            _fab_idx_emit(product,
                          f"[FAB랏인덱스] 배치 {bi}/{total_batches} 축소 실패: "
                          f"{type(e).__name__}: {e}", ok=False,
                          detail={"progress": _fab_idx_progress(
                              bi - 1, total_batches, state="failed")})
            return False
        used_all.extend(u for u in used if u not in used_all)
        tmp_part = part_fp.with_suffix(".parquet.tmp")
        batch_df.write_parquet(tmp_part)
        os.replace(tmp_part, part_fp)
        completed.add(bi)
        reduced_rows += batch_df.height
        manifest = {
            "build_sig": build_sig,
            "completed": sorted(completed),
            "used": used_all,
            "root_col": root_col_meta,
            "reduced_rows": reduced_rows,
        }
        save_json(manifest_fp, manifest)
        # 첫·마지막 배치는 항상, 나머지는 스로틀 — 배치가 많아도 로그가 화면을
        # 덮지 않으면서 "멈춘 게 아니라 진행 중"이 계속 보이게.
        now = time.time()
        if bi == 1 or bi == total_batches or (now - last_log) >= _fab_idx_log_gap_sec():
            last_log = now
            eta = ((now - started) / bi) * (total_batches - bi)
            _fab_idx_emit(
                product,
                f"[FAB랏인덱스] 배치 {bi:,}/{total_batches:,}"
                f" ({int(bi * 100 / max(1, total_batches))}%) · 누적 "
                f"{reduced_rows:,}행"
                f" · RSS {_proc_rss_gb()}GB · 남은 ~{_fmt_dur_ko(eta)}",
                detail={
                    "batch": bi,
                    "total_batches": total_batches,
                    "progress": _fab_idx_progress(bi, total_batches),
                },
            )
        batch_df = None
        if len(batches) > 1:
            # 배치 사이에 사용자 요청/메모리 회복에 양보 — 빌드가 길어도
            # interactive 검색과 RSS 안정에 영향을 주지 않게 한다.
            try:
                from core.runtime_limits import process_memory_high as _pmh
                from core import request_priority as _rp
                if _pmh():
                    time.sleep(0.5)
                _rp.yield_to_users(max_wait_sec=20.0)
            except Exception:
                pass
    part_files = [staging / f"{bi:06d}.parquet" for bi in range(1, total_batches + 1)]
    part_files = [p for p in part_files if p.is_file()]
    if not part_files or not root_col_meta:
        sources = _global_fab_source_paths(fab_source, include_all=include_all)
        reason = batch_failures[-1] if batch_failures else "FAB data files not found"
        _fab_idx_emit(
            product,
            "[FAB랏인덱스] 전체 재빌드 중단 — 읽을 수 있는 FAB 행이 없습니다"
            f" · {reason}",
            ok=False,
            detail={
                "fab_source": fab_source,
                "resolved_sources": sources[:20],
                "source_file_count": len(live_sig),
                "batch_failures": batch_failures[-20:],
                "stage": _stage("fab_index", "fail"),
                "progress": _fab_idx_progress(
                    len(completed), total_batches, state="failed"),
            },
        )
        return False
    _fab_idx_emit(product,
                  f"[FAB랏인덱스] root 파티션 기록 중 — 배치 축소 {reduced_rows:,}행")
    combined = pl.concat([pl.scan_parquet(str(p)) for p in part_files], how="diagonal_relaxed")
    final_lf, _ = _fab_latest_reduce_lf(combined)
    if final_lf is None:
        return False
    _write_fab_index_partitions(product, final_lf)
    meta = {
        "built_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "sources": used_all,
        "root_col": root_col_meta,
        "source_sig": live_sig,
    }
    try:
        save_json(_fab_lot_index_meta_path(product), meta)
    except Exception:
        logger.debug("fab_lot_index meta write failed product=%s", product, exc_info=True)
    shutil.rmtree(staging, ignore_errors=True)
    _fab_idx_emit(product,
                  f"[FAB랏인덱스] 전체 재빌드 완료 — 배치 축소 {reduced_rows:,}행 · "
                  f"총 {_fmt_dur_ko(time.time() - started)} · RSS {_proc_rss_gb()}GB",
                  detail={
                      "stage": _stage("fab_index", "done"),
                      "progress": _fab_idx_progress(
                          total_batches, total_batches, state="done"),
                  })
    return True


def _build_fab_lot_index_incremental(product: str, fab_source: str, include_all: bool,
                                     added_files: set[str], live_sig: list,
                                     old_meta: dict) -> bool:
    """Merge newly added FAB files into the existing index; rewrite only the
    affected root partitions. False → caller falls back to the full rebuild.

    타이 규칙: 같은 (root,wafer) 에 동일 timestamp 행이 기존 인덱스와 새 파일
    양쪽에 있으면 기존 행을 유지한다 — 전체 재빌드의 stable sort 에서 경로
    정렬상 앞서는(=기존) 파일이 이기는 것과 일치한다."""
    started = time.time()
    _fab_idx_emit(product,
                  f"[FAB랏인덱스] 증분 갱신 시작 — 새 FAB 파일 {len(added_files):,}개")
    try:
        idx_dir = _fab_lot_index_dir(product)
        fab_lf, used_sources = _scan_global_fab_sources(
            fab_source, include_all=include_all, only_files=added_files)
        if fab_lf is None:
            # 추가 파일이 이 product 의 소스 범위 밖(다른 소스 폴더)일 수 있다 —
            # 인덱스 내용 불변이므로 시그니처만 갱신한다.
            meta = dict(old_meta)
            meta["built_at"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
            meta["source_sig"] = live_sig
            save_json(_fab_lot_index_meta_path(product), meta)
            return True
        new_lf, root_col = _fab_latest_reduce_lf(fab_lf)
        if new_lf is None:
            return False
        new_df = new_lf.collect()
        roots = sorted({str(v) for v in new_df[_FAB_IDX_ROOT_COL].to_list() if str(v or "").strip()})
        if not roots:
            meta = dict(old_meta)
            meta["built_at"] = datetime.datetime.now(datetime.timezone.utc).isoformat()
            meta["source_sig"] = live_sig
            save_json(_fab_lot_index_meta_path(product), meta)
            return True
        # 리터럴 디렉터리명으로 안전하게 교체 가능한 root 만 증분 처리한다
        # (특수문자는 hive 인코딩과 어긋날 수 있음 → 전체 재빌드).
        if any(not _re.fullmatch(r"[A-Z0-9_\-.]+", r) for r in roots):
            return False
        # 새 파일이 기존 root 대부분을 건드리면 per-root 병합(파티션별 읽기+
        # 교체)이 한 번의 전체 sort 보다 비싸다 — 실측상 30% 를 넘으면 전체
        # 재빌드가 더 빠르므로 폴백한다.
        try:
            existing = sum(1 for p in idx_dir.iterdir() if p.is_dir())
        except Exception:
            existing = 0
        if existing and len(roots) > max(16, int(existing * 0.3)):
            logger.info("fab_lot_index incremental 포기 — 영향 root %d/%d (product=%s)",
                        len(roots), existing, product)
            _fab_idx_emit(product,
                          f"[FAB랏인덱스] 증분 포기 — 새 파일이 root {len(roots):,}/{existing:,}개를 "
                          f"건드려 전체 재빌드가 더 빠릅니다")
            return False

        frames = []
        for root in roots:
            part = idx_dir / f"{_FAB_IDX_ROOT_COL}={root}"
            if part.is_dir():
                files = sorted(part.glob("*.parquet"))
                if files:
                    frames.append(_scan_parquet_compat([str(p) for p in files]))
        frames.append(new_df.lazy())  # 기존 인덱스 행이 앞 — 타이에서 기존 우선
        merged_input = pl.concat(frames, how="diagonal_relaxed") if len(frames) > 1 else frames[0]
        merged, _ = _fab_latest_reduce_lf(merged_input)
        if merged is None:
            return False

        staging = idx_dir.with_name(idx_dir.name + ".delta")
        if staging.exists():
            shutil.rmtree(staging, ignore_errors=True)
        staging.mkdir(parents=True, exist_ok=True)
        try:
            try:
                sink_target = pl.PartitionBy(
                    staging, key=_FAB_IDX_ROOT_COL, include_key=True,
                    approximate_bytes_per_file="auto",
                )
                merged.sink_parquet(sink_target, mkdir=True, maintain_order=False)
            except Exception:
                merged_df = merged.collect()
                shutil.rmtree(staging, ignore_errors=True)
                staging.mkdir(parents=True, exist_ok=True)
                if merged_df.height:
                    merged_df.write_parquet(staging, partition_by=_FAB_IDX_ROOT_COL)
            written = {p.name for p in staging.iterdir() if p.is_dir()}
            expected = {f"{_FAB_IDX_ROOT_COL}={r}" for r in roots}
            if written != expected:
                return False
            for child in sorted(staging.iterdir()):
                if not child.is_dir():
                    continue
                target = idx_dir / child.name
                if target.exists():
                    shutil.rmtree(target, ignore_errors=True)
                child.replace(target)
        finally:
            shutil.rmtree(staging, ignore_errors=True)

        sources = list(dict.fromkeys(list(old_meta.get("sources") or []) + list(used_sources)))
        meta = {
            "built_at": datetime.datetime.now(datetime.timezone.utc).isoformat(),
            "sources": sources,
            "root_col": old_meta.get("root_col") or root_col,
            "source_sig": live_sig,
        }
        save_json(_fab_lot_index_meta_path(product), meta)
        logger.info("fab_lot_index incremental merge: %d file(s) → %d root(s) (product=%s)",
                    len(added_files), len(roots), product)
        _fab_idx_emit(product,
                      f"[FAB랏인덱스] 증분 갱신 완료 — 새 파일 {len(added_files):,}개 → "
                      f"root {len(roots):,}개 교체 · 총 {_fmt_dur_ko(time.time() - started)}")
        return True
    except Exception:
        logger.debug("fab_lot_index incremental build failed product=%s", product, exc_info=True)
        _fab_idx_emit(product, "[FAB랏인덱스] 증분 갱신 실패 — 전체 재빌드로 넘어갑니다", ok=False)
        return False


def _enqueue_fab_lot_index_build(product: str, fab_source: str = "",
                                 include_all: bool = False, reason: str = "", *,
                                 immediate: bool = False,
                                 local_only: bool = False) -> bool:
    """Single-flight, cooldown-guarded background (re)build of the fab lot index.

    local_only=True 면 개발 워커로 오프로드하지 않고 이 서버에서 직접 빌드한다."""
    if not _fab_lot_index_enabled():
        return False
    if _is_fab_lot_index_staging_product(product):
        logger.warning("fab_lot_index staging product enqueue rejected: %s", product)
        return False
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    if not canonical:
        return False
    now = time.time()
    with _FAB_IDX_BUILD_LOCK:
        if canonical in _FAB_IDX_BUILD_INPROGRESS:
            return False
        if (not immediate and
                now - _FAB_IDX_BUILD_LAST.get(canonical, 0.0) < _FAB_IDX_BUILD_COOLDOWN_SEC):
            return False
        _FAB_IDX_BUILD_INPROGRESS.add(canonical)

    def _run():
        ok = False

        def _local_build() -> dict:
            lease_name = f"splittable_fabidx_{canonical}"
            lease_held = False
            try:
                try:
                    from core import shared_lease as _shared_lease
                    lease_held = _shared_lease.try_acquire(lease_name, ttl_sec=1800.0)
                    if not lease_held:
                        logger.info(f"fab_lot_index build skipped for {canonical} — 다른 서버가 빌드 중")
                        return {"ok": False}
                except Exception:
                    lease_held = False  # lease infra optional; proceed without it
                # 사용자 검색이 진행 중이면 빌드 시작을 미룬다 (pivot 빌드와 동일한 이유).
                from core import request_priority as _rp
                _rp.yield_to_users(max_wait_sec=60.0)
                with _HEAVY_BUILD_SEMAPHORE:
                    return {"ok": bool(_build_fab_lot_index(canonical, fab_source, include_all))}
            except Exception as exc:
                logger.warning(f"fab_lot_index build failed for {canonical} ({reason}): {exc}")
                return {"ok": False}
            finally:
                if lease_held:
                    try:
                        from core import shared_lease as _shared_lease
                        _shared_lease.release(lease_name)
                    except Exception:
                        pass

        try:
            if local_only:
                # 이 서버에서 요청한 수동 캐싱 — 워커로 넘기지 않는다 (pivot 과 동일).
                res = _local_build()
            else:
                # v9.4.x: 워커 생존 시 오프로드, 아니면 로컬 폴백 (pivot 빌드와 동일).
                from core import worker_dispatch as _wd
                res = _wd.run_heavy(
                    "splittable_fab_lot_index_build",
                    {"product": canonical, "fab_source": fab_source, "include_all": include_all},
                    _local_build,
                    # 스캔 큐에 그대로 뜨는 문자열이다 — 'fabidx:PRODA' 로는
                    # 무슨 작업인지 알 수 없어 사람이 읽는 이름으로 둔다.
                    label=f"FAB 랏 인덱스 (root별 최신 FAB lot): {canonical}",
                    local_idle_only=not immediate,
                    local_fallback=bool(immediate),
                    durable=not immediate,
                    priority="normal" if immediate else "maintenance",
                    dedupe_key=f"fab_lot_index:{canonical}",
                    timeout_sec=6 * 3600.0,
                )
            ok = bool((res or {}).get("ok"))
        finally:
            with _FAB_IDX_BUILD_LOCK:
                _FAB_IDX_BUILD_INPROGRESS.discard(canonical)
                _FAB_IDX_BUILD_LAST[canonical] = time.time()
        if ok:
            # New fab labels are not captured by the view payload cache signature;
            # clear it so the next search recomputes with fresh joined lot ids.
            _clear_split_view_cache_product(canonical)

    threading.Thread(target=_run, daemon=True, name=f"splittable-fabidx-{canonical}").start()
    logger.info(f"fab_lot_index build queued for {canonical} ({reason})")
    return True


def _split_view_cache_get(key: tuple, hard_sig: tuple, soft_sig: tuple) -> tuple[str, dict | None]:
    """(freshness, payload) 반환. freshness ∈ {"miss","fresh","stale"}.

    - hard_sig 불일치(신규 lot / 사용자 편집) → ("miss", None) + 엔트리 폐기.
    - soft_sig 까지 일치 → ("fresh", payload).
    - soft_sig 만 불일치(백그라운드 파생 캐시 재기록) → ("stale", payload) —
      즉시 서빙하고 호출측이 백그라운드 재검증을 예약한다.
    """
    global _VIEW_CACHE_BYTES
    with _VIEW_CACHE_LOCK:
        cached = _VIEW_CACHE.get(key)
        if cached:
            cached_hard, cached_soft, payload, approx_bytes = cached
            if cached_hard != hard_sig:
                _VIEW_CACHE.pop(key, None)
                _VIEW_CACHE_BYTES = max(0, _VIEW_CACHE_BYTES - approx_bytes)
            else:
                _VIEW_CACHE.move_to_end(key)
                if cached_soft == soft_sig:
                    return "fresh", dict(payload)
                return "stale", dict(payload)
    # RAM LRU에서 밀려도 공유 디스크의 압축 payload를 읽어 cold parquet/FAB 계산을
    # 피한다. 디스크 엔트리도 동일 hard/soft 시그니처 계약을 사용한다.
    freshness, payload = _view_disk_cache_read(key, hard_sig, soft_sig)
    if payload is not None:
        _split_view_cache_put_memory(key, hard_sig, soft_sig, payload)
        return freshness, dict(payload)
    return "miss", None


def _split_view_cache_prepare_payload(payload: dict) -> dict:
    stored = dict(payload)
    stored.pop("related_issues", None)
    stored.pop("runtime_profile", None)
    stored.pop("view_cache", None)
    # 레거시 fat rows(`_cells`, 셀당 9키 dict ≈441B)는 캐시에 담지 않는다. HTTP 응답은
    # rows_compact(≈40B/셀)만 쓰고 fat rows 는 버려지므로, 저장하면 엔트리 하나가
    # 예산의 ~11배를 먹어 다른 사용자의 검색 결과를 계속 밀어냈다(= 여러 명이
    # 서로 다른 lot 을 검색하면 전부 cold 로 떨어지던 원인). 레거시 형태가 필요한
    # 내부 호출자는 _expand_view_rows() 로 그때 복원한다.
    if stored.get("rows_compact") is not None:
        stored.pop("rows", None)
    # v: lookup_cache 는 그대로 저장한다 — HIT 경로에서 _attach 가 매번
    # _ml_table_lookup.cache_status(meta 읽기 + partition dir glob) 를 재계산하던
    # 비용을 없앤다. 저장 시점의 배지값이 약간 stale 할 수 있으나, 캐시가 렌더된
    # 상태에서 빌드 진행 배지는 행동 가치가 없으므로 허용.
    return stored


def _split_view_cache_put_memory(key: tuple, hard_sig: tuple, soft_sig: tuple, stored: dict) -> None:
    global _VIEW_CACHE_BYTES
    approx_bytes = _estimate_view_payload_bytes(stored)
    budget = _view_cache_max_bytes()
    with _VIEW_CACHE_LOCK:
        old = _VIEW_CACHE.pop(key, None)
        if old is not None:
            _VIEW_CACHE_BYTES = max(0, _VIEW_CACHE_BYTES - old[3])
        _VIEW_CACHE[key] = (hard_sig, soft_sig, stored, approx_bytes)
        _VIEW_CACHE_BYTES += approx_bytes
        while _VIEW_CACHE and (
            len(_VIEW_CACHE) > _view_cache_max_entries() or _VIEW_CACHE_BYTES > budget
        ):
            if len(_VIEW_CACHE) == 1:
                break  # 방금 넣은 항목은 예산 초과라도 유지 (miss 반복 방지)
            _, evicted = _VIEW_CACHE.popitem(last=False)
            _VIEW_CACHE_BYTES = max(0, _VIEW_CACHE_BYTES - evicted[3])


def _split_view_cache_put(key: tuple, hard_sig: tuple, soft_sig: tuple, payload: dict) -> None:
    stored = _split_view_cache_prepare_payload(payload)
    _split_view_cache_put_memory(key, hard_sig, soft_sig, stored)
    _view_disk_cache_write(key, hard_sig, soft_sig, stored)


def _view_revalidate_cooldown_sec() -> float:
    try:
        v = float(os.environ.get("FLOW_SPLITTABLE_VIEW_REVALIDATE_COOLDOWN_SEC", "")
                  or _VIEW_REVALIDATE_COOLDOWN_SEC_DEFAULT)
    except Exception:
        v = _VIEW_REVALIDATE_COOLDOWN_SEC_DEFAULT
    return max(0.0, min(86400.0, v))


def _view_revalidate_delay_sec() -> float:
    try:
        v = float(os.environ.get("FLOW_SPLITTABLE_VIEW_REVALIDATE_DELAY_SEC", "")
                  or _VIEW_REVALIDATE_DELAY_SEC_DEFAULT)
    except Exception:
        v = _VIEW_REVALIDATE_DELAY_SEC_DEFAULT
    return max(0.0, min(600.0, v))


def _view_revalidate_execute(view_cache_key: tuple, params: dict) -> dict:
    """재검증 1건을 운영 서버에서 실행해 운영 RAM/disk view 캐시를 갱신한다."""
    from core import request_priority

    # 사용자 HTTP 요청이 진행 중이면 백그라운드 재검증 시작을 미룬다. 실제 사용자
    # cold 조회는 이 대기 경로를 거치지 않고 운영 서버의 전용 cold 레인에서 즉시 돈다.
    request_priority.yield_to_users(max_wait_sec=120.0, quiet_for_sec=3.0)
    _VIEW_REVALIDATE_TLS.force = True
    try:
        view_split_core(request=None, include_related=False, **params)
    finally:
        _VIEW_REVALIDATE_TLS.force = False
    return {"ok": True, "stored": "operating"}


def _view_revalidate_worker_loop() -> None:
    """전역 재검증 워커 — 큐를 한 건씩 처리한다.

    재계산은 워커(개발서버) 생존 시 오프로드되어 이 서버의 polars 풀을 쓰지
    않고, 로컬 폴백 시에도 동시 재계산은 항상 최대 1건: 20명이 몰려도
    백그라운드가 polars 풀에서 점유하는 collect 는 하나뿐이라 사용자 검색의
    CPU 경쟁이 상수로 묶인다."""
    while True:
        _VIEW_REVALIDATE_WAKE.wait(timeout=60.0)
        while True:
            now = time.time()
            key = None
            params: dict = {}
            wait_hint = 0.0
            with _VIEW_REVALIDATE_LOCK:
                if not _VIEW_REVALIDATE_PENDING:
                    _VIEW_REVALIDATE_WAKE.clear()
                    break
                head_key, (enq_ts, head_params) = next(iter(_VIEW_REVALIDATE_PENDING.items()))
                age = now - enq_ts
                burst_quiet = now - _VIEW_REVALIDATE_LAST_ENQUEUE_TS
                # 디바운스: 새 stale 검색이 계속 들어오는 동안(burst)은 처리를 미루고,
                # 잠잠해지면 처리한다. 상시 트래픽에서도 head 가 delay 이상 기다리면
                # 진행을 보장해 워커가 굶지 않는다.
                if burst_quiet >= _VIEW_REVALIDATE_BURST_QUIET_SEC or age >= _view_revalidate_delay_sec():
                    _VIEW_REVALIDATE_PENDING.pop(head_key, None)
                    _VIEW_REVALIDATE_INFLIGHT.add(head_key)
                    key = head_key
                    params = head_params
                else:
                    wait_hint = min(5.0, max(0.5, _VIEW_REVALIDATE_BURST_QUIET_SEC - burst_quiet))
            if key is None:
                time.sleep(wait_hint)
                continue
            try:
                _view_revalidate_execute(key, params)
            except Exception as exc:
                logger.debug("view revalidate 실패 (product=%s): %s", params.get("product"), exc)
            finally:
                with _VIEW_REVALIDATE_LOCK:
                    _VIEW_REVALIDATE_INFLIGHT.discard(key)
                    _VIEW_REVALIDATE_LAST[key] = time.time()
                    if len(_VIEW_REVALIDATE_LAST) > _VIEW_REVALIDATE_LAST_MAX:
                        overflow = len(_VIEW_REVALIDATE_LAST) - _VIEW_REVALIDATE_LAST_MAX
                        for old_key in sorted(_VIEW_REVALIDATE_LAST, key=_VIEW_REVALIDATE_LAST.get)[:overflow]:
                            _VIEW_REVALIDATE_LAST.pop(old_key, None)


def _ensure_view_revalidate_worker_locked() -> None:
    global _VIEW_REVALIDATE_THREAD
    t = _VIEW_REVALIDATE_THREAD
    if t is not None and t.is_alive():
        return
    t = threading.Thread(target=_view_revalidate_worker_loop, daemon=True,
                         name="splittable-view-revalidate")
    _VIEW_REVALIDATE_THREAD = t
    t.start()


def _enqueue_view_revalidate(view_cache_key: tuple, params: dict) -> bool:
    """Stale hit 시 백그라운드에서 view payload 를 재계산해 최신 lot 라벨로 갱신.

    key 단위 병합(중복 제거) + 쿨다운(기본 3h) — lot_progress 스케줄러가 파생
    캐시를 자주 재기록해도 같은 검색을 반복 재계산하지 않는다. 사용자 요청은 이미
    stale 캐시로 즉시 응답했으므로 이 갱신은 다음 조회를 fresh 로 만드는 용도다.
    실행은 전역 단일 워커가 담당한다 (_view_revalidate_worker_loop)."""
    global _VIEW_REVALIDATE_LAST_ENQUEUE_TS
    now = time.time()
    with _VIEW_REVALIDATE_LOCK:
        if view_cache_key in _VIEW_REVALIDATE_INFLIGHT or view_cache_key in _VIEW_REVALIDATE_PENDING:
            return False
        if now - _VIEW_REVALIDATE_LAST.get(view_cache_key, 0.0) < _view_revalidate_cooldown_sec():
            return False
        _VIEW_REVALIDATE_PENDING[view_cache_key] = (now, dict(params))
        _VIEW_REVALIDATE_LAST_ENQUEUE_TS = now
        while len(_VIEW_REVALIDATE_PENDING) > _VIEW_REVALIDATE_PENDING_MAX:
            # 큐 상한 초과 — 가장 오래된 항목을 버린다. 버려진 검색은 다음 stale hit
            # 때 다시 등록되므로 유실이 아니라 지연일 뿐이다.
            _VIEW_REVALIDATE_PENDING.popitem(last=False)
        _ensure_view_revalidate_worker_locked()
    _VIEW_REVALIDATE_WAKE.set()
    return True


def _split_view_request_user(request: Request | None) -> tuple[str, str]:
    if request is None:
        return "", "user"
    try:
        me = current_user(request)
        return me.get("username") or "", me.get("role") or "user"
    except Exception:
        return "", "user"


# 비동기 감사 로그: /view 는 요청마다 감사 로그를 남기는데, jsonl_append 는 공유
# 드라이브 파일 락 + open/write 라 동시 요청을 직렬화한다. 큐에 쌓고 단일 데몬
# 스레드가 비워 요청 지연 경로에서 제거한다(감사 유실 없이).
_AUDIT_QUEUE: deque = deque()
_AUDIT_QUEUE_WAKE = threading.Event()
_AUDIT_QUEUE_MAX = 10000
_AUDIT_WORKER_STARTED = False
_AUDIT_WORKER_LOCK = threading.Lock()


def _audit_worker_loop() -> None:
    while True:
        _AUDIT_QUEUE_WAKE.wait(timeout=5.0)
        _AUDIT_QUEUE_WAKE.clear()
        while _AUDIT_QUEUE:
            try:
                username, action, detail, tab = _AUDIT_QUEUE.popleft()
            except IndexError:
                break
            try:
                _audit_user(username, action, detail=detail, tab=tab)
            except Exception:
                pass


def _ensure_audit_worker() -> None:
    global _AUDIT_WORKER_STARTED
    if _AUDIT_WORKER_STARTED:
        return
    with _AUDIT_WORKER_LOCK:
        if _AUDIT_WORKER_STARTED:
            return
        threading.Thread(target=_audit_worker_loop, daemon=True, name="splittable-audit").start()
        _AUDIT_WORKER_STARTED = True


def _audit_enqueue(username: str, action: str, detail: str = "", tab: str = "") -> None:
    if len(_AUDIT_QUEUE) >= _AUDIT_QUEUE_MAX:
        return  # 과부하 시 드롭 — 요청 지연보다 우선.
    _ensure_audit_worker()
    _AUDIT_QUEUE.append((username, action, detail, tab))
    _AUDIT_QUEUE_WAKE.set()


def _audit_split_view_search(
    request: Request | None,
    product: str,
    root_lot_id: str,
    fab_lot_id: str,
    wafer_ids: str,
    prefix: str,
) -> None:
    if request is None or not (str(root_lot_id or "").strip() or str(fab_lot_id or "").strip()):
        return
    username, _role = _split_view_request_user(request)
    if not username:
        return
    detail = (
        f"product={str(product or '').strip()} "
        f"root_lot_id={str(root_lot_id or '').strip()} "
        f"fab_lot_id={str(fab_lot_id or '').strip()} "
        f"wafer_ids={str(wafer_ids or '').strip()} "
        f"prefix={str(prefix or '').strip()}"
    )
    _audit_enqueue(username, "splittable:view_search", detail=detail, tab="splittable")


def _split_view_lookup_cache_public(status: dict | None, queued: dict | None = None) -> dict:
    return _lookup_cache_public_meta(status, queued)


def _split_view_cache_preparing_payload(
    product: str,
    root_lot_id: str,
    wafer_ids: str,
    prefix: str,
    history_mode: str,
    status: dict | None,
    queued: dict | None,
    *,
    message: str,
    started: float,
    runtime_profile: dict,
    view_cache_key: tuple,
) -> dict:
    payload = {
        "product": product,
        "lot_col": "root_lot_id",
        "wf_col": "wafer_id",
        "headers": [],
        "rows": [],
        "header_groups": [],
        "wafer_fab_list": [],
        "prefixes": _load_prefixes(),
        "root_lot_id": str(root_lot_id or "").strip(),
        "prefix": str(prefix or "").strip(),
        "history_mode": history_mode,
        "mismatch_count": 0,
        "product_cache": _product_ram_cache_response_meta(product),
        "lookup_cache": _split_view_lookup_cache_public(status, queued),
        "msg": message,
    }
    return _split_view_finish_payload(
        payload,
        started=started,
        runtime_profile=runtime_profile,
        payload_cache_hit=False,
        view_cache_key=view_cache_key,
    )


def _split_view_large_root_cache_or_defer(
    product: str,
    root_lot_id: str,
    wafer_ids: str,
    fp: Path,
    *,
    started: float,
    runtime_profile: dict,
    view_cache_key: tuple,
    prefix: str,
    history_mode: str,
    force_defer_raw_fallback: bool = False,
) -> tuple[Any | None, dict | None]:
    root = str(root_lot_id or "").strip()
    if not root or fp.suffix.lower() != ".parquet":
        return None, None
    if _product_ram_cache_entry(product):
        return None, None
    if not force_defer_raw_fallback and not _split_view_should_defer_raw_fallback(fp):
        return None, None
    status = _ml_table_lookup.cache_status(fp)
    runtime_profile["root_cache_status"] = status.get("status") or ""
    if not status.get("has_cache"):
        queued = _ml_table_lookup.enqueue_build(fp)
        runtime_profile["_lookup_cache"] = _split_view_lookup_cache_public(status, queued)
        runtime_profile["root_cache_hit"] = False
        # 수백 MB ML_TABLE을 HTTP 요청에서 직접 collect하면 동시 검색 2~3건만으로
        # 수 GB peak가 겹친다. 큰 소스는 공유 lookup 빌드를 먼저 끝내고 프런트가
        # 짧게 재시도한다. 작은 개발 fixture만 기존 raw fallback을 허용한다.
        # 여기서 바로 반환하므로 호출측 scan_ms lap 을 못 탄다 — 직접 닫는다.
        _lap(runtime_profile, "scan_ms")
        return None, _split_view_cache_preparing_payload(
            product,
            root,
            wafer_ids,
            prefix,
            history_mode,
            status,
            queued,
            message="Root lot 인덱스를 준비 중입니다. 완료되면 자동 재검색됩니다.",
            started=started,
            runtime_profile=runtime_profile,
            view_cache_key=view_cache_key,
        )
    # 소스가 갱신돼 cache 가 stale 여도, 해당 root 의 hive 파티션이 있으면 즉시
    # 서빙하고 백그라운드 재빌드만 예약한다(allow_stale). 데이터 갱신 직후마다
    # 소스 전체를 재스캔(5~10초)하던 것을 파티션 인덱스 읽기로 대체 — 이 stale
    # 구간이 SplitTable 검색이 캐시가 있어도 느리던 주원인이었다.
    lf, status = _ml_table_lookup.scan_root_lot_cache(fp, root, wafer_ids=wafer_ids, allow_stale=True, profile=runtime_profile)
    runtime_profile["root_cache_status"] = status.get("status") or ""
    runtime_profile["root_cache_hit"] = lf is not None
    runtime_profile["_lookup_cache"] = _split_view_lookup_cache_public(status, {})
    if lf is None:
        _lap(runtime_profile, "scan_ms")
        return None, _split_view_cache_preparing_payload(
            product,
            root,
            wafer_ids,
            prefix,
            history_mode,
            status,
            {},
            message="No data",
            started=started,
            runtime_profile=runtime_profile,
            view_cache_key=view_cache_key,
        )
    return _cast_cats_lazy(lf), None


def _attach_split_view_runtime_fields(
    payload: dict,
    request: Request | None,
    *,
    include_related: bool = False,
    started: float | None = None,
    runtime_profile: dict | None = None,
    payload_cache_hit: bool = False,
    view_cache_key: tuple | None = None,
    view_stale: bool = False,
) -> dict:
    out = dict(payload)
    if "lookup_cache" not in out:
        status = None
        try:
            product = out.get("product") or ""
            root = str(out.get("root_lot_id") or "").strip()
            if product and root:
                fp = _product_path(product)
                if fp.suffix.lower() == ".parquet":
                    status = _ml_table_lookup.cache_status(fp)
        except Exception:
            status = None
        out["lookup_cache"] = _split_view_lookup_cache_public(status, None)
    if include_related:
        username, role = _split_view_request_user(request)
        out["related_issues"] = _related_tracker_issues(
            out.get("product") or "",
            out.get("root_lot_id") or "",
            username,
            role,
        )
    out["product_cache"] = _product_ram_cache_response_meta(out.get("product") or "")
    # 마무리 메타(lookup_cache 재조회/related_issues/product_cache) — 캐시 히트
    # 응답에서도 도는 구간이라 여기가 무거우면 "캐시 히트인데 느리다" 가 된다.
    _lap(runtime_profile, "finish_ms")
    if runtime_profile and runtime_profile.get("fab_index_queued"):
        out["background_cache"] = {
            "queued": True,
            "kind": "fab_root_index",
            "message": "FAB 랏 인덱스를 백그라운드에서 준비 중입니다. 표는 자동으로 갱신됩니다.",
        }
    if started is not None:
        out = _split_view_finish_payload(
            out,
            started=started,
            runtime_profile=runtime_profile,
            payload_cache_hit=payload_cache_hit,
            view_cache_key=view_cache_key,
            view_stale=view_stale,
        )
    return out


@router.get("/related-issues")
def related_issues_for_view(
    product: str = Query(...),
    root_lot_id: str = Query(""),
    request: Request = None,
):
    username, role = _split_view_request_user(request)
    return {
        "product": product,
        "root_lot_id": root_lot_id,
        "related_issues": _related_tracker_issues(product, root_lot_id, username, role),
    }


# ── View ──


try:
    import orjson as _orjson
except ImportError:
    _orjson = None


def _view_orjson_response(payload):
    """/view 전용 직렬화 우회. FastAPI 기본 경로는 dict 반환 시 jsonable_encoder 를
    payload 전체에 재귀 적용하는데, KNOB 처럼 행×웨이퍼 셀이 많은 응답(수만 셀)에서
    이 인코딩만 수 초가 걸린다. Response 객체를 직접 반환하면 그 경로를 건너뛴다.
    orjson 미설치·직렬화 실패 시 dict 를 그대로 돌려 기본 경로로 폴백한다.

    (응답, 직렬화 ms, 본문 바이트) 를 돌려준다 — 이 비용은 핸들러 total_ms 밖이라
    호출측이 타이밍 로그에 따로 실어야 보인다. 폴백(dict 반환) 시 바이트는 0."""
    if _orjson is None or not isinstance(payload, dict):
        return payload, 0.0, 0
    _t0 = time.perf_counter()
    try:
        body = _orjson.dumps(
            payload,
            default=str,
            option=_orjson.OPT_SERIALIZE_NUMPY | _orjson.OPT_NON_STR_KEYS,
        )
    except Exception:
        return payload, (time.perf_counter() - _t0) * 1000.0, 0
    return (
        Response(content=body, media_type="application/json"),
        (time.perf_counter() - _t0) * 1000.0,
        len(body),
    )


def _expand_view_rows(payload: dict) -> dict:
    """슬림 셀 포맷(cells_format v2) → 레거시 `_cells` 행으로 복원.

    v2 행은 actual 배열(a) + sparse plan(p) + sparse mismatch(m) + 행-상수
    플래그만 담는다. 셀마다 행-상수 플래그와 파생 가능한 key 를 반복하던 레거시
    `_cells` 는 KNOB 2000행×25웨이퍼에서 ≈10.9MB(셀당 실측 441B)였고, HTTP 응답은
    이미 v2 만 보낸다. 그래서 서버는 v2 만 만들고 캐시하며, 레거시 형태가 필요한
    내부 호출자(informs embed, 백그라운드 revalidate, knob-allocation, worker
    task)만 이 함수로 복원한다.

    셀 key 조립 규칙 `root_lot_id|wafer_keys[ci]|_param` 은 FE(My_SplitTable
    expandViewRows)와 동일한 계약이다 — 어긋나면 plan 저장 키가 틀어진다."""
    if not isinstance(payload, dict):
        return payload
    compact = payload.get("rows_compact")
    if compact is None:
        return payload
    existing = payload.get("rows")
    if isinstance(existing, list) and existing:
        return payload
    wafer_keys = payload.get("wafer_keys") or []
    n_keys = len(wafer_keys)
    root = str(payload.get("root_lot_id") or "")
    rows = []
    for r in compact:
        vals = r.get("a") or []
        plans_sparse = r.get("p") or {}
        mism = set(r.get("m") or [])
        can_plan = bool(r.get("can_plan"))
        is_tag = bool(r.get("tag"))
        is_mgmt = bool(r.get("mgmt"))
        param = r.get("_param")
        cells = {}
        for ci in range(len(vals)):
            key = str(ci)
            wf = wafer_keys[ci] if ci < n_keys else ci
            cells[key] = {
                "actual": vals[ci],
                "plan": plans_sparse.get(key),
                "key": f"{root}|{wf}|{param}",
                "can_plan": can_plan,
                "mismatch": ci in mism,
                "is_custom_tag": is_tag,
                "can_tag": is_tag,
                "is_management_row": is_mgmt,
                "can_management_edit": is_mgmt,
            }
        rows.append({"_param": param, "_display": r.get("_display"), "_cells": cells})
    out = dict(payload)
    out["rows"] = rows
    return out


@router.get("/view")
def view_split_http(product: str = Query(...), root_lot_id: str = Query(""),
                    wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
                    custom_name: str = Query(""), view_mode: str = Query("all"),
                    history_mode: str = Query("all"),
                    fab_lot_id: str = Query(""),
                    custom_cols: str = Query(""),
                    include_related: bool = Query(False),
                    cache_first: bool = Query(False),
                    request: Request = None):
    # HTTP 진입점 — 서버는 슬림 셀 포맷(v2)만 만들고 캐시한다. 레거시 rows(_cells)를
    # 기대하는 내부 호출자(재검증 스레드/informs embed/knob-allocation/테스트)는
    # view_split() 래퍼가 _expand_view_rows() 로 그때 복원해 준다.
    #
    # 타이밍 기록은 직렬화까지 끝난 뒤 flush 한다 — 큰 응답에서는 직렬화 자체가
    # 무시 못 할 비용이고, 핸들러 안에서만 재면 그 시간이 통째로 사라진다.
    pending: list[dict] = []
    _VIEW_TIMING_TLS.pending = pending
    try:
        payload = view_split_core(
            product=product, root_lot_id=root_lot_id, wafer_ids=wafer_ids,
            prefix=prefix, custom_name=custom_name, view_mode=view_mode,
            history_mode=history_mode, fab_lot_id=fab_lot_id,
            custom_cols=custom_cols, include_related=include_related,
            cache_first=cache_first, request=request,
        )
        compact = payload.pop("rows_compact", None)
        if compact is not None:
            payload["rows"] = compact
            payload["cells_format"] = "v2"
        response, serialize_ms, body_bytes = _view_orjson_response(payload)
        for entry in pending:
            entry["serialize_ms"] = round(serialize_ms, 3)
            entry["response_bytes"] = body_bytes
        return response
    finally:
        _VIEW_TIMING_TLS.pending = None
        for entry in pending:
            _search_timing_log.record(entry)


def view_split(**kwargs) -> dict:
    """레거시 계약(rows = `_cells` dict) 을 유지하는 내부 호출자용 래퍼.

    HTTP 경로는 view_split_core 를 직접 쓰고 슬림 포맷 그대로 내보낸다."""
    return _expand_view_rows(view_split_core(**kwargs))


def view_split_core(product: str = Query(...), root_lot_id: str = Query(""),
               wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
               custom_name: str = Query(""), view_mode: str = Query("all"),
               history_mode: str = Query("all"),
               fab_lot_id: str = Query(""),
               custom_cols: str = Query(""),
               include_related: bool = Query(False),
               cache_first: bool = Query(False),
               request: Request = None):
    # v8.8.33: custom_cols (쉼표 구분) 추가 — Save 없이 체크만 한 컬럼을 ad-hoc 으로 전달.
    # v9.0.3: 한 root_lot_id 아래 여러 fab_lot_id 가 정상이다. FAB 공정 진행 중
    #   fab_lot_id 가 바뀔 수 있으므로 앞 5자 일치 여부를 검증/경고 기준으로 쓰지 않는다.
    started = time.perf_counter()
    request_username, request_user_role = _split_view_request_user(request)
    runtime_profile = {
        "username": request_username,
        "user_role": request_user_role,
        "is_user_search": request is not None,
        "root_cache_hit": False,
        "product_cache_hit": False,
        "scan_ms": 0.0,
        "collect_ms": 0.0,
        "matrix_ms": 0.0,
        "overlay_ms": 0.0,
        "root_cache_status": "",
        "root_data_source": "",
        # 미들웨어 레인 대기(있으면). started 는 레인 통과 후에 찍히므로 이 값을
        # 더해야 사용자가 체감한 시간(wall_ms)이 된다.
        "lane_wait_ms": _request_lane_wait_ms(request),
        "cold_lane_wait_ms": 0.0,
    }
    runtime_profile[_LAP_MARK] = started
    _history_mode = (history_mode or "all").strip().lower() or "all"
    if _history_mode not in ("all", "final", "lot_all"):
        raise HTTPException(400, "history_mode must be one of: all, final, lot_all")
    cache_first_enabled = _truthy_value(cache_first)
    # 백그라운드 stale-revalidate 스레드의 재진입이면 캐시 서빙/감사로그를 건너뛰고
    # 순수 재계산 후 fresh 시그니처로 캐시를 덮어쓴다.
    force_recompute = bool(getattr(_VIEW_REVALIDATE_TLS, "force", False))
    if not force_recompute and not cache_first_enabled:
        _audit_split_view_search(request, product, root_lot_id, fab_lot_id, wafer_ids, prefix)
    _lot_warn = ""
    fp = _product_path(product)
    view_cache_key = _split_view_cache_key(
        product, root_lot_id, wafer_ids, prefix, custom_name,
        view_mode, _history_mode, fab_lot_id, custom_cols,
    )
    view_hard_sig, view_soft_sig = _split_view_cache_dep_signature(
        product, custom_name=custom_name, product_fp=fp)
    if not force_recompute:
        freshness, cached_view = _split_view_cache_get(view_cache_key, view_hard_sig, view_soft_sig)
        # 진입~여기 = 인증/감사/캐시키/의존 시그니처(공유드라이브 stat)/캐시 조회.
        # payload 캐시 히트 응답은 사실상 전부 이 구간이다.
        _lap(runtime_profile, "prelude_ms")
        if cached_view is not None:
            # 신규 lot 없음(hard 일치) → fresh/stale 모두 캐시 즉시 서빙. soft 만
            # 달라진 stale 이면 백그라운드에서 최신 lot 라벨로 재검증을 예약한다.
            if freshness == "stale":
                _enqueue_view_revalidate(view_cache_key, {
                    "product": product, "root_lot_id": root_lot_id, "wafer_ids": wafer_ids,
                    "prefix": prefix, "custom_name": custom_name, "view_mode": view_mode,
                    "history_mode": history_mode, "fab_lot_id": fab_lot_id,
                    "custom_cols": custom_cols,
                })
                _lap(runtime_profile, "prelude_ms")
            return _attach_split_view_runtime_fields(
                cached_view,
                request,
                include_related=include_related,
                started=started,
                runtime_profile=runtime_profile,
                payload_cache_hit=True,
                view_cache_key=view_cache_key,
                view_stale=(freshness == "stale"),
            )
    # force_recompute(재검증 워커) 경로는 위 lap 을 타지 않는다 — 여기서 한 번 더
    # 불러 진입 구간을 닫는다. 이미 닫혔으면 ~0 이 더해질 뿐이다.
    _lap(runtime_profile, "prelude_ms")
    if not root_lot_id.strip() and not fab_lot_id.strip():
        return _split_view_finish_payload(
            {"product": product, "lot_col": "root_lot_id", "wf_col": "wafer_id",
             "headers": [], "rows": [], "prefixes": _load_prefixes(),
             "product_cache": _product_ram_cache_response_meta(product),
             "msg": "Enter a Root Lot ID or Fab Lot ID to view"},
            started=started,
            runtime_profile=runtime_profile,
            payload_cache_hit=False,
            view_cache_key=view_cache_key,
        )
    # 같은 cold key가 동시에 들어오면 한 요청만 계산하고 나머지는 결과를 기다린다.
    # 수백 MB 원본/파티션 scan이 중복 실행되는 메모리 스파이크와 CPU 경합을 방지한다.
    compute_owner, compute_event = _view_compute_begin(view_cache_key)
    if not compute_owner:
        wait_started = time.perf_counter()
        completed = compute_event.wait(timeout=_view_compute_wait_seconds())
        runtime_profile["singleflight_wait_ms"] = (time.perf_counter() - wait_started) * 1000.0
        # 대기는 singleflight_wait_ms 에 이미 잡혔다 — 단계 합계에서 빼기 위해 버린다.
        _lap(runtime_profile, None)
        view_hard_sig, view_soft_sig = _split_view_cache_dep_signature(
            product, custom_name=custom_name, product_fp=fp)
        freshness, cached_view = _split_view_cache_get(view_cache_key, view_hard_sig, view_soft_sig)
        _lap(runtime_profile, "prelude_ms")
        if cached_view is not None:
            return _attach_split_view_runtime_fields(
                cached_view,
                request,
                include_related=include_related,
                started=started,
                runtime_profile=runtime_profile,
                payload_cache_hit=True,
                view_cache_key=view_cache_key,
                view_stale=(freshness == "stale"),
            )
        if completed:
            # 이전 owner가 빈 결과/오류로 끝났다면 한 요청만 다음 owner가 된다.
            compute_owner, compute_event = _view_compute_begin(view_cache_key)
        if not compute_owner:
            status = _ml_table_lookup.cache_status(fp) if root_lot_id.strip() and fp.suffix.lower() == ".parquet" else {}
            return _split_view_cache_preparing_payload(
                product,
                root_lot_id,
                wafer_ids,
                prefix,
                _history_mode,
                status,
                {"status": "running", "queued": True},
                message="같은 SplitTable 검색을 처리 중입니다. 완료되면 자동 재검색됩니다.",
                started=started,
                runtime_profile=runtime_profile,
                view_cache_key=view_cache_key,
            )
    # 여기부터가 실제 메모리/CPU 를 쓰는 cold 계산이다. 미들웨어가 아니라 이 지점에서
    # 직렬화하므로, 위에서 이미 반환된 캐시 HIT/단일비행 대기 요청은 슬롯을 기다리지
    # 않는다. 슬롯 확보 실패(레인 포화)는 429 대신 "준비 중" 페이로드로 응답해
    # FE 의 기존 자동 재조회 흐름을 그대로 태운다.
    if not _view_cold_lane_acquire(runtime_profile):
        _lap(runtime_profile, None)  # 레인 대기는 cold_lane_wait_ms 에 기록됨.
        _view_compute_finish(view_cache_key)
        return _split_view_cache_preparing_payload(
            product, root_lot_id, wafer_ids, prefix, _history_mode,
            {}, {"status": "running", "queued": True},
            message="검색 요청이 몰려 대기 중입니다. 잠시 후 자동으로 다시 조회합니다.",
            started=started,
            runtime_profile=runtime_profile,
            view_cache_key=view_cache_key,
        )
    _lap(runtime_profile, None)  # 레인 대기는 cold_lane_wait_ms 에 기록됨.
    pivot_base_lf = None
    knob_sidecar_all_columns = None
    try:
        if root_lot_id.strip() and not cache_first_enabled:
            fast_cache_path = _pivot_cache_path(product, root_lot_id.strip())
            if fast_cache_path.exists():
                try:
                    if fp and fast_cache_path.stat().st_mtime < fp.stat().st_mtime:
                        # 원본 ML_TABLE 이 pivot cache 보다 최신 — 즉시성은 유지하고
                        # 백그라운드 재빌드를 예약해 다음 조회부터 최신 데이터를 쓴다.
                        _enqueue_pivot_cache_build(product, reason="stale_pivot")
                except Exception:
                    pass
                # v9.2: native-orientation per-root cache (wafer rows × param
                # cols). Feed it straight into the normal renderer as base_lf so
                # column projection (prefix/custom) stays index-fast AND the
                # latest-lot join runs (lot_id/fab label). Legacy transposed
                # files (a "parameter" column, no wafer_id) are skipped + rebuilt.
                try:
                    cache_names = pl.scan_parquet(str(fast_cache_path)).collect_schema().names()
                except Exception:
                    cache_names = []
                is_legacy = ("parameter" in cache_names) and not any(
                    c.lower() == "wafer_id" for c in cache_names
                )
                if is_legacy:
                    _enqueue_pivot_cache_build(product, reason="legacy_pivot_format")
                elif cache_names:
                    # KNOB 만 보는 검색이면 KNOB 전용 사이드카(좁은 파일)를 읽는다.
                    # 없거나 앵커 컬럼이 빠져 있으면 그대로 전체 파일 — 결과 동일.
                    read_path = fast_cache_path
                    if _knob_only_request(prefix, custom_name, custom_cols):
                        knob_path = _pivot_cache_knob_path(product, root_lot_id.strip())
                        if _knob_sidecar_usable(knob_path, fast_cache_path):
                            read_path = knob_path
                            runtime_profile["root_data_source_detail"] = "knob_sidecar"
                            # 좁은 파일로 읽으면 프레임 스키마에 KNOB 밖 컬럼이 없다.
                            # all_columns(FE 컬럼 선택기 목록)는 전체 스키마여야 하므로
                            # 여기서 확보해 둔다 — 안 그러면 커스텀 세트에서 INLINE/VM 이 사라진다.
                            knob_sidecar_all_columns = list(cache_names)
                        else:
                            # 구버전 캐시엔 사이드카가 없다 — 다음 빌드에서 backfill.
                            _enqueue_pivot_cache_build(product, reason="knob_sidecar_missing")
                    pivot_base_lf = _cast_cats_lazy(_scan_parquet_compat(str(read_path)))
                    runtime_profile["root_cache_hit"] = True
            else:
                # pivot cache miss — 이번 요청은 아래 일반 경로로 처리하고,
                # 백그라운드에서 제품 전체 pivot cache 를 빌드해 다음 검색을 즉시화한다.
                _enqueue_pivot_cache_build(product, reason="cache_miss")
    except HTTPException:
        _view_compute_finish(view_cache_key)
        raise
    except Exception as e:
        logger.warning(f"Fast path failed: {e}")
    # pivot 캐시 탐색 = 경로 stat + 스키마 읽기 + KNOB 사이드카 판정. 넓은 pivot
    # 파일에서는 스키마 읽기만으로도 무시 못 할 시간이 든다.
    _lap(runtime_profile, "fastpath_ms")

    try:
        if pivot_base_lf is not None:
            base_lf = pivot_base_lf
            runtime_profile["root_data_source"] = "pivot_cache"
        else:
            base_lf, deferred_payload = _split_view_large_root_cache_or_defer(
                product,
                root_lot_id,
                wafer_ids,
                fp,
                started=started,
                runtime_profile=runtime_profile,
                view_cache_key=view_cache_key,
                prefix=prefix,
                history_mode=_history_mode,
                force_defer_raw_fallback=cache_first_enabled,
            )
            _lap(runtime_profile, "scan_ms")
            if deferred_payload is not None:
                return deferred_payload
        lf = _scan_product(
            product,
            root_lot_id=root_lot_id,
            fab_lot_id=fab_lot_id,
            wafer_ids=wafer_ids,
            base_lf=base_lf,
            runtime_profile=runtime_profile,
        )
        # _scan_product 은 base_lf(파티션/RAM) + latest-lot/fab override join 을
        # lazy 로 구성한다(실제 실행은 뒤 collect). 여기서는 그 구성 시간을 scan_ms 에
        # 합산 — DB-first 경로에서 join 준비 비용을 breakdown 에 노출한다.
        _lap(runtime_profile, "scan_ms")
        lot_col, wf_col = _detect_lot_wafer(lf, product)
        # v8.4.4/v8.8.3: fab_lot_col — 매뉴얼 override > 자동 추론 > "fab_lot_id".
        fab_lot_col = "fab_lot_id"
        try:
            schema_names = lf.collect_schema().names()
            _cfg = load_json_cached(SOURCE_CFG, {}) or {}
            _ov = _lot_override_for(_cfg, product)
            _fc = (_ov.get("fab_col") or "").strip()
            if _fc and _fc in schema_names:
                fab_lot_col = _fc
            elif "fab_lot_id" not in schema_names:
                # 자동 보강된 컬럼 이름 중 하나로 대체.
                for c in _FAB_COL_CANDIDATES:
                    if c in schema_names:
                        fab_lot_col = c
                        break
        except Exception:
            pass
        # lot/wafer/fab 컬럼 해석 — 스키마 확보 + 소스 오버라이드 설정 읽기.
        _lap(runtime_profile, "schema_ms")

        fab_scope = {}
        fab_filter_for_join = fab_lot_id
        forced_fab_scope_label = ""
        if fab_lot_id.strip():
            # v9.0.5: fab_lot_id 는 DB FAB 원천에서 정확히 매칭될 때만 유효하다.
            # v9.0.6: 다만 사내/데모 파일이 이미 ML_TABLE 안에 fab/lot 값을 가진 경우도
            # 있으므로 FAB history scope 가 없다고 즉시 종료하지 않고 coalesced /view
            # 데이터에서 한 번 더 필터한다.
            fab_scope = _fab_history_scope(product, root_lot_id=root_lot_id,
                                           fab_lot_id=fab_lot_id, limit=5000)
            src_wafers = fab_scope.get("wafer_ids") or []
            if src_wafers:
                if not root_lot_id.strip() and fab_scope.get("root_ids"):
                    root_lot_id = fab_scope["root_ids"][0]
                wafer_ids = _merge_wafer_scope(wafer_ids, src_wafers)
                fab_filter_for_join = ""
                forced_fab_scope_label = fab_lot_id.strip()
        # FAB 히스토리 스코프 조회 — fab_lot_id 를 넣은 검색에서만 돈다.
        _lap(runtime_profile, "fabscope_ms")

        joined_lf = lf
        lf = _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id, wafer_ids,
                               fab_lot_id=fab_filter_for_join, fab_lot_col=fab_lot_col)

        def _prepare_view_frame(view_lf):
            view_schema = view_lf.collect_schema().names()
            all_data = _view_data_columns(view_schema, lot_col, wf_col, fab_lot_col)
            tag_labels = _custom_tag_label_map(product)
            for tag_col in tag_labels:
                if tag_col not in all_data:
                    all_data.append(tag_col)
            management_labels = _management_row_label_map(product)
            if custom_name or custom_cols:
                for mgmt_col in management_labels:
                    if mgmt_col not in all_data:
                        all_data.append(mgmt_col)
            sel = _select_columns(all_data, custom_name, prefix,
                                  max_fallback=50, custom_cols=custom_cols)
            if not custom_name and not custom_cols:
                for raw_pref in [p.strip() for p in str(prefix or "").split(",") if p.strip()]:
                    for virt in _virtual_columns_for_prefix(product, raw_pref):
                        if virt not in sel:
                            sel.append(virt)
            rename = _build_col_rename_map(sel, product)
            rename.update({col: f"{CUSTOM_TAG_PREFIX}_{label}" for col, label in tag_labels.items()})
            rename.update({col: label for col, label in management_labels.items()})
            # ppid_knob step_desc → step_id 공정 순서 우선, 매핑 없는 행은 기존 자연 정렬.
            step_rank = _split_step_order_context(product).get("param_rank") or {}
            sel = sorted(sel, key=lambda c: _step_order_sort_key(c, rename.get(c, c), step_rank))
            keep_cols = []
            for c in (lot_col, wf_col):
                if c and c in view_schema and c not in keep_cols:
                    keep_cols.append(c)
            keep_fab_col = "fab_lot_id" if "fab_lot_id" in view_schema else None
            if not keep_fab_col:
                keep_fab_col = (
                    _ci_resolve_in(fab_lot_col, view_schema)
                    or _pick_first_present_ci(_FAB_COL_CANDIDATES, view_schema)
                    or None
                )
            if keep_fab_col and keep_fab_col in view_schema and keep_fab_col not in keep_cols:
                keep_cols.append(keep_fab_col)
            for c in sel:
                if c in view_schema and c not in keep_cols:
                    keep_cols.append(c)
            q = view_lf.select(keep_cols) if keep_cols else view_lf
            # 여기까지가 컬럼 선택·rename·공정순 정렬 — 전부 파이썬이고 컬럼 수에
            # 비례한다(멀티 prefix 검색에서 커지는 쪽). collect 와 분리해야 어느
            # 쪽이 문제인지 보인다.
            _lap(runtime_profile, "select_ms")
            df_out = q.head(SPLITTABLE_VIEW_MAX_WAFERS).collect()
            _lap(runtime_profile, "collect_ms")
            return df_out, all_data, sel, rename

        df, all_data_cols, selected, col_rename = _prepare_view_frame(lf)
        if df.height == 0 and root_lot_id.strip() and fab_lot_id.strip():
            # If the UI carries a stale Fab Lot while the operator searches a
            # valid root lot, do not let the stale secondary field hide the
            # renderable SplitTable rows. Root remains the primary scope.
            try:
                root_only_lf = _filter_lot_wafer(
                    joined_lf, lot_col, wf_col, root_lot_id, wafer_ids,
                    fab_lot_col=fab_lot_col,
                )
                root_only_df, all_data_cols, selected, col_rename = _prepare_view_frame(root_only_lf)
                if root_only_df.height > 0:
                    df = root_only_df
                    _lot_warn = "Fab Lot ID와 Root Lot ID 조합이 없어 Root Lot ID 기준으로 조회했습니다."
            except Exception as e:
                logger.warning("view_split root-only fallback 실패 (product=%s root=%s fab=%s) %s: %s",
                               product, root_lot_id, fab_lot_id, type(e).__name__, e)
        if df.height == 0:
            # Operators often paste the FAB lot value they found in File Browser
            # into the Root Lot field. Treat that as a fab_lot_id lookup before
            # declaring the SplitTable empty.
            root_input = root_lot_id.strip()
            if root_input and not fab_lot_id.strip():
                try:
                    pasted_fab_scope = _fab_history_scope(
                        product, fab_lot_id=root_input, limit=5000
                    )
                    pasted_wafers = pasted_fab_scope.get("wafer_ids") or []
                    pasted_roots = pasted_fab_scope.get("root_ids") or []
                    if pasted_wafers and pasted_roots:
                        fallback_root = pasted_roots[0]
                        fallback_wafers = _merge_wafer_scope(wafer_ids, pasted_wafers)
                        fallback_lf = _scan_product(
                            product, root_lot_id=fallback_root,
                            wafer_ids=fallback_wafers,
                            runtime_profile=runtime_profile,
                        )
                    else:
                        fallback_root = ""
                        fallback_lf = _scan_product(product, fab_lot_id=root_input,
                                                    wafer_ids=wafer_ids,
                                                    runtime_profile=runtime_profile)
                    fallback_names = fallback_lf.collect_schema().names()
                    fallback_fab_col = (
                        _ci_resolve_in(fab_lot_col, fallback_names)
                        or _pick_first_present_ci(_FAB_COL_CANDIDATES, fallback_names)
                    )
                    if pasted_wafers and pasted_roots:
                        fallback_df, all_data_cols, selected, col_rename = _prepare_view_frame(fallback_lf)
                        if fallback_df.height > 0:
                            df = fallback_df
                            fab_lot_id = root_input
                            root_lot_id = fallback_root
                            forced_fab_scope_label = root_input
                            _lot_warn = "입력한 Root Lot ID를 fab_lot_id로 해석해 조회했습니다."
                    elif fallback_fab_col:
                        fallback_lf = _filter_lot_wafer(
                            fallback_lf, lot_col, wf_col, "",
                            wafer_ids, fab_lot_id=root_input,
                            fab_lot_col=fallback_fab_col,
                        )
                        fallback_df, all_data_cols, selected, col_rename = _prepare_view_frame(fallback_lf)
                        if fallback_df.height > 0:
                            df = fallback_df
                            fab_lot_id = root_input
                            root_lot_id = ""
                            _lot_warn = "입력한 Root Lot ID를 fab_lot_id로 해석해 조회했습니다."
                except Exception as e:
                    logger.warning("view_split fab_lot fallback 실패 (product=%s input=%s) %s: %s",
                                   product, root_input, type(e).__name__, e)
        # 빈 결과 재해석 구간(위 두 폴백). 안쪽 _prepare_view_frame 은 자기 몫을
        # select/collect 로 이미 떼 갔으므로 여기 남는 건 폴백 자체의 비용이다.
        _lap(runtime_profile, "emptyfallback_ms")
        if df.height == 0:
            return _split_view_finish_payload(
                {"product": product, "lot_col": lot_col, "wf_col": wf_col,
                 "headers": [], "rows": [], "prefixes": _load_prefixes(),
                 "product_cache": _product_ram_cache_response_meta(product),
                 "msg": "No data"},
                started=started,
                runtime_profile=runtime_profile,
                payload_cache_hit=False,
                view_cache_key=view_cache_key,
            )
        if not root_lot_id.strip() and lot_col and lot_col in df.columns:
            roots = []
            for v in df[lot_col].cast(_STR, strict=False).to_list():
                s = str(v or "").strip()
                if s and s not in ("None", "null") and s not in roots:
                    roots.append(s)
            if roots:
                root_lot_id = sorted(roots)[0]

        # Wafer header list + fab_lot_id grouping (v8.4.4)
        fab_col = "fab_lot_id" if "fab_lot_id" in df.columns else None
        if wf_col and wf_col in df.columns:
            # Wafer IDs are physically 1..25. Some upstream DBs contain
            # placeholder values like 1000; do not expose or plan against them.
            wf_raw = [_normalize_wafer_id(v) for v in df[wf_col].to_list()]
            # Per-wafer fab_lot_id (first non-null occurrence per wafer)
            wf2fab: dict = {}
            if fab_col:
                fab_vals = [(None if v is None else str(v)) for v in df[fab_col].to_list()]
                for w, f in zip(wf_raw, fab_vals):
                    if not w: continue
                    if w not in wf2fab and f and f not in ("None", "null"):
                        wf2fab[w] = f
            if forced_fab_scope_label:
                wf2fab = {w: forced_fab_scope_label for w in dict.fromkeys(wf_raw) if w}
            # Sort: (fab_lot_id 그룹, wafer_id 숫자-aware) — fab_lot 미정이면 "~" 로 후순위.
            # v8.8.3: wafer_id 가 문자열일 때 "10" < "2" 오작동 → 숫자 가능하면 int 로 cast 해서 secondary 키.
            wf_uniq = [w for w in dict.fromkeys(wf_raw) if w]
            def _wf_sort_key(w):
                primary = wf2fab.get(w, "~")
                try:
                    n = int(w)
                    return (primary, 0, n)
                except (TypeError, ValueError):
                    s = str(w)
                    # 선행 'W' 제거 후 숫자 시도
                    if s.upper().startswith("W"):
                        try:
                            return (primary, 0, int(s[1:]))
                        except ValueError:
                            pass
                    return (primary, 1, s)
            wf_sorted = sorted(wf_uniq, key=_wf_sort_key)
            headers = [f"#{v}" for v in wf_sorted]
            wf_idx = {v: i for i, v in enumerate(wf_sorted)}
            # Build header_groups: consecutive same-fab_lot segments
            wafer_fab_list = [wf2fab.get(w, "") for w in wf_sorted]
            header_groups = []
            if fab_col:
                cur = None; span = 0
                for f in wafer_fab_list:
                    if f == cur:
                        span += 1
                    else:
                        # fab_lot_id 를 못 찾은 구간도 라벨은 남긴다 — 빈 문자열로
                        # 두면 프런트가 lot_id 헤더 행에 빈 칸을 그려서 "lot id 가
                        # 사라졌다" 로 보인다.
                        if span > 0: header_groups.append({"label": cur or "—", "span": span})
                        cur = f; span = 1
                if span > 0: header_groups.append({"label": cur or "—", "span": span})
        else:
            wf_raw = list(range(df.height))
            wf_sorted = list(range(df.height))
            headers = [f"#{i}" for i in wf_sorted]
            wf_idx = {i: i for i in wf_sorted}
            wafer_fab_list = []
            header_groups = []
        # 웨이퍼 헤더/그룹 조립 — df 컬럼을 파이썬 리스트로 꺼내 정렬·그룹핑한다.
        _lap(runtime_profile, "header_ms")

        # Load plans — 이 root 것만. 아래 조회는 전부 `root|wafer|col` 키라서
        # 제품 전체를 들고 있을 이유가 없었고, 그 전량 복사가 요청당 3.5ms 였다.
        # (편집 경로는 계속 _load_plan_data 로 전체 사본을 받는다.)
        plans = _plan_entries_for_root(product, root_lot_id)
        tag_labels = _custom_tag_label_map(product)
        tag_values = _custom_tag_values_for_root(product, root_lot_id)
        management_labels = _management_row_label_map(product)
        management_values = _management_row_values_for_root(product, root_lot_id)
        _lap(runtime_profile, "overlay_ms")

        # 슬림 셀 포맷(v2)을 곧바로 만든다. 예전에는 셀마다 9키 dict + f-string key 를
        # 만들고(KNOB 2000행×25웨이퍼 ≈ 5만 dict / 10만 f-string) 그 결과를 다시
        # 전량 순회해 compact 로 변환했다. 이 구간은 순수 파이썬이라 GIL 을 놓지
        # 않아 동시 검색이 여기서 그대로 직렬화됐다. 셀 key 는 plan/mismatch 가
        # 실제로 있는 셀에서만 조립한다(대부분 lot 은 plan 이 없거나 몇 개뿐).
        rows_compact = []
        row_mismatches = []  # rows_compact 와 같은 인덱스 — diff 필터 후 평탄화
        df_cols_set = set(df.columns)
        n_cols = len(wf_sorted)
        has_plans = bool(plans)
        for col_name in selected:
            is_tag_col = col_name in tag_labels
            is_management_row = col_name in management_labels
            vals = [None] * n_cols
            plan_vals = [None] * n_cols
            # v8.8.16: CUSTOM 에 저장된 컬럼이 현재 df 에 없더라도 빈 행으로 표시.
            #   (e.g. plan 전용 가상 컬럼, 다른 제품에서 저장된 컬럼 등). plan 값은 여전히 lookup.
            if is_tag_col:
                for ci, wf_key in enumerate(wf_sorted):
                    vals[ci] = tag_values.get(f"{root_lot_id}|{wf_key}|{col_name}")
            elif is_management_row:
                for ci, wf_key in enumerate(wf_sorted):
                    vals[ci] = management_values.get(f"{root_lot_id}|{wf_key}|{col_name}")
            elif col_name in df_cols_set:
                try:
                    col_data = df[col_name].to_list()
                    for i, val in enumerate(col_data):
                        key = wf_raw[i] if i < len(wf_raw) else None
                        idx = wf_idx.get(key)
                        if idx is not None:
                            vals[idx] = val
                            if has_plans:
                                pv = plans.get(f"{root_lot_id}|{key}|{col_name}", {}).get("value")
                                if pv is not None:
                                    plan_vals[idx] = pv
                except Exception:
                    pass
            elif has_plans:
                # 가상 컬럼 — plan 값만 확인.
                for ci, wf_key in enumerate(wf_sorted):
                    pv = plans.get(f"{root_lot_id}|{wf_key}|{col_name}", {}).get("value")
                    if pv is not None:
                        plan_vals[ci] = pv

            plans_sparse = {}
            mism = []
            cell_mismatches = []
            for ci in range(n_cols):
                actual = vals[ci]
                actual_str = None if actual is None else str(actual)
                if actual_str in ("None", "null"):
                    actual_str = None
                vals[ci] = actual_str
                plan = plan_vals[ci]
                if plan is None:
                    continue
                plans_sparse[str(ci)] = plan
                if plan and actual_str and str(plan) != actual_str:
                    mism.append(ci)
                    ck = f"{root_lot_id}|{wf_sorted[ci]}|{col_name}"
                    plan_info = plans.get(ck, {})
                    cell_mismatches.append({
                        "param": col_name, "key": ck,
                        "plan": plan, "actual": actual_str,
                        "plan_user": plan_info.get("user", ""),
                        "plan_updated": plan_info.get("updated", ""),
                    })

            # v8.8.14: _display — rule_order + step_desc를 포함한 렌더용 이름.
            #   없으면 원본과 동일. FE 는 _display 를 사용하고 prefix strip 후 표시.
            row_c = {"_param": col_name, "_display": col_rename.get(col_name, col_name), "a": vals}
            if plans_sparse:
                row_c["p"] = plans_sparse
            if mism:
                row_c["m"] = mism
            if n_cols:
                # 행-상수 플래그 — 셀이 하나도 없으면(웨이퍼 0개) 레거시와 동일하게 생략.
                col_upper = col_name.upper()
                if (not is_tag_col and not is_management_row) and any(
                        col_upper.startswith(p + "_") for p in PLAN_ALLOWED_PREFIXES):
                    row_c["can_plan"] = True
                if is_tag_col:
                    row_c["tag"] = True
                if is_management_row:
                    row_c["mgmt"] = True
            rows_compact.append(row_c)
            row_mismatches.append(cell_mismatches)

        if view_mode == "diff":
            keep = [i for i, r in enumerate(rows_compact)
                    if len(set(v for v in r["a"] if v is not None)) > 1]
            rows_compact = [rows_compact[i] for i in keep]
            row_mismatches = [row_mismatches[i] for i in keep]

        # Detect mismatches and send notifications to plan owners
        # (diff 모드에서는 레거시와 동일하게 화면에 남은 행만 대상)
        mismatches = [m for per_row in row_mismatches for m in per_row]
        _lap(runtime_profile, "matrix_ms")
        if not force_recompute:
            # 백그라운드 재검증은 동일 데이터를 재계산하는 것이므로 알림을 중복 발송하지 않는다.
            _enqueue_plan_actual_mismatches(product, mismatches, actor="flow")
        _lap(runtime_profile, "mismatch_ms")

        # v8.8.5: view 응답에 오버라이드 resolve 결과 동봉 — FE 상단 배지에 "어디서 읽어왔는지" 바로 표시.
        override_meta = _resolve_override_meta_light(product)
        # v9.0.5: FAB 후보는 DB FAB 원천의 정확한 root 매칭만 노출한다.
        #   DB FAB 에 없는 root 는 ML_TABLE LOT_ID / joined null fallback 을 쓰지 않는다.
        available_fab_lots = sorted(
            {str(v).strip() for v in wafer_fab_list if str(v or "").strip()},
            key=lambda s: s.upper(),
        )
        fab_presence_known = bool(available_fab_lots)
        if not available_fab_lots:
            hist_lots = _fab_history_scope(product, root_lot_id=root_lot_id, limit=1000)
            fab_presence_known = bool(hist_lots.get("query_ok"))
            if hist_lots.get("candidates"):
                available_fab_lots = hist_lots["candidates"]
        fab_present = True if available_fab_lots else (False if fab_presence_known else None)
        _lap(runtime_profile, "overlay_ms")
        # latest-lot 캐시 기준 현재 진행 step 이후(미진행) 행 — FE 회색 셰이딩용.
        # FAB 조회가 정상 완료됐는데 root 매칭이 0건이면 아직 FAB에 없는 lot이므로
        # 모든 표시 공정을 회색으로 보낸다. 조회 실패/소스 미설정은 미확인으로 둔다.
        step_progress = _split_step_progress(
            product, root_lot_id, selected, wf_sorted, fab_present=fab_present)
        _lap(runtime_profile, "progress_ms")
        payload = {
            "product": product, "lot_col": lot_col, "wf_col": wf_col,
            "step_progress": step_progress,
            "headers": headers, "rows_compact": rows_compact,
            "wafer_keys": [f"{k}" for k in wf_sorted],
            "header_groups": header_groups, "wafer_fab_list": wafer_fab_list,
            "row_labels": {"root_lot_id": "root_lot_id", "lot_id": "lot_id", "parameter": "항목"},
            "available_fab_lots": available_fab_lots,
            "prefixes": _load_prefixes(), "precision": load_json_cached(PRECISION_CFG, DEFAULT_PRECISION), "root_lot_id": root_lot_id,
            "all_columns": _merge_all_columns(all_data_cols, knob_sidecar_all_columns, lot_col, wf_col),
            "selected_count": len(selected),
            "prefix": prefix or (custom_name if custom_name else ""),
            "history_mode": _history_mode,
            "plan_allowed_prefixes": PLAN_ALLOWED_PREFIXES,
            "mismatch_count": len(mismatches),
            "override": override_meta,
            "match_cache": _match_cache_response_meta(product),
            "product_cache": _product_ram_cache_response_meta(product),
            "lookup_cache": runtime_profile.get("_lookup_cache") or _split_view_lookup_cache_public(None, None),
            "lot_warn": _lot_warn,
        }
        # 응답 dict 조립 — prefixes/precision 설정 로드, all_columns 병합, 캐시 메타.
        _lap(runtime_profile, "payload_ms")
        # FAB root index가 아직 없을 때의 부분 응답을 payload cache에 고정하지 않는다.
        # 인덱스 완료 후 UI polling 또는 다음 검색이 즉시 완전한 결과를 받게 한다.
        if not runtime_profile.get("cache_incomplete"):
            _split_view_cache_put(view_cache_key, view_hard_sig, view_soft_sig, payload)
        _lap(runtime_profile, "cacheput_ms")
        return _attach_split_view_runtime_fields(
            payload,
            request,
            include_related=include_related,
            started=started,
            runtime_profile=runtime_profile,
            payload_cache_hit=False,
            view_cache_key=view_cache_key,
        )
    except HTTPException:
        _view_compute_finish(view_cache_key)
        raise
    except Exception as e:
        _view_compute_finish(view_cache_key)
        raise HTTPException(400, f"View error: {str(e)}")


# ── Pre-pivoted cache: on-demand refresh ──
class PivotRefreshReq(BaseModel):
    product: str
    username: str = ""


@router.post("/cache/pivot/refresh")
def refresh_pivot_cache(req: PivotRefreshReq, _perm=Depends(require_page_manager("splittable"))):
    """수동 pivot cache 재빌드 트리거. 빌드는 백그라운드에서 돌고 완료 시
    view payload cache 를 비워 다음 조회부터 최신 데이터가 보인다."""
    queued = _enqueue_pivot_cache_build(req.product, reason="manual_refresh", immediate=True)
    return {
        "ok": True,
        "queued": queued,
        "state": _pivot_cache_build_state(req.product),
    }


@router.get("/cache/pivot/status")
def pivot_cache_status(product: str = Query(...), username: str = Query("")):
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    cache_dir = _pivot_cache_path(canonical, "_probe").parent
    files = 0
    latest_mtime = 0.0
    try:
        if cache_dir.exists():
            for fp_ in cache_dir.glob("*.parquet"):
                files += 1
                latest_mtime = max(latest_mtime, fp_.stat().st_mtime)
    except Exception:
        pass
    return {
        "product": canonical,
        "state": _pivot_cache_build_state(canonical),
        "files": files,
        "last_built": datetime.datetime.fromtimestamp(latest_mtime).isoformat(timespec="seconds") if latest_mtime else None,
    }


# ── Plans ──
class PlanReq(BaseModel):
    product: str
    plans: dict
    username: str = "unknown"
    root_lot_id: str = ""
    # 선택 입력 — 왜 바꿨는지. 비워 두는 게 기본이고, 남기면 이력에 그대로 붙는다.
    reason: str = ""


@router.post("/plan")
def save_plan(req: PlanReq, request: Request = None):
    if request is not None:
        try:
            me = current_user(request)
            req.username = me.get("username") or req.username or "unknown"
        except Exception:
            raise
    # Validate: only KNOB/MASK/FAB columns can have plans
    rejected = []
    for ck in list(req.plans.keys()):
        col_name = ck.split("|")[-1] if "|" in ck else ck
        col_upper = col_name.upper()
        if not any(col_upper.startswith(p + "_") for p in PLAN_ALLOWED_PREFIXES):
            rejected.append(col_name)
            del req.plans[ck]
    if rejected and not req.plans:
        raise HTTPException(400, f"Plan not allowed for: {', '.join(rejected)}. Only {'/'.join(PLAN_ALLOWED_PREFIXES)} columns.")

    pf = _plan_history_path(req.product)
    data = _load_plan_data(req.product)
    data.setdefault("history", [])
    now = datetime.datetime.now().isoformat()
    changed_entries = []
    # v8.8.33: my_plan_changed 이벤트 대상자 수집.
    #   같은 cell 에 과거 plan 이 있었으면 그 plan 을 만든 user 에게 "내 plan 이 변경됨" 알림.
    original_owners: dict[str, str] = {}
    for ck in req.plans.keys():
        prev_user = (data["plans"].get(ck) or {}).get("user")
        if prev_user:
            original_owners[ck] = prev_user
    prior_history = list(data["history"])
    batch_id = _new_history_batch_id()
    reason = _clean_plan_reason(req.reason)
    new_history: list[dict] = []
    for ck, val in req.plans.items():
        old = data["plans"].get(ck, {}).get("value")
        prev_owner = original_owners.get(ck) or ""
        data["plans"][ck] = {"value": val, "user": req.username, "updated": now}
        cell_root, cell_wafer, cell_column = _split_plan_cell_key(ck)
        entry = {
            "cell": ck, "old": old, "new": val, "user": req.username,
            "time": now, "action": "set", "root_lot_id": req.root_lot_id or cell_root,
            # 아래는 화면 필터·CSV 를 위해 풀어 둔 것 — cell key 파싱을 소비자마다
            # 반복하지 않게 한다. 예전 엔트리는 없으므로 읽는 쪽이 폴백을 갖는다.
            "wafer_id": cell_wafer, "column": cell_column,
            "batch": batch_id, "batch_size": len(req.plans),
            "prev_user": prev_owner, "reason": reason,
        }
        data["history"].append(entry)
        new_history.append(entry)
        changed_entries.append((ck, old, val))
    data["history"] = data["history"][-1000:]
    save_json(pf, data)
    _archive_plan_history(req.product, new_history, prior_history)
    _invalidate_plan_risk_cache(req.product)

    # v9.1.x: plan 저장은 여기서 즉시 완료 — actual 대조(셀당 파케이 스캔)·knowledge
    # 적재·알림은 백그라운드로 옮겨 저장 응답 지연을 없앤다 (가장 사용 빈도 높은 경로).
    product = req.product
    username = req.username
    root_lot_id_req = req.root_lot_id

    def _plan_post_save():
        try:
            for ck, old, val in changed_entries:
                _append_splittable_plan_knowledge(
                    product=product,
                    cell_key=ck,
                    old=old,
                    new=val,
                    actor=username,
                    changed_at=now,
                    conflicting=bool(old not in (None, "") and old != val),
                )
            save_mismatches = []
            for ck, _old, val in changed_entries:
                actual = _actual_value_for_plan_cell(product, ck)
                if not _plan_actual_mismatch(val, actual):
                    continue
                root, wafer, column = _split_plan_cell_key(ck)
                save_mismatches.append({
                    "key": ck,
                    "plan": val,
                    "actual": actual,
                    "plan_user": username,
                    "plan_updated": now,
                    "root_lot_id": root,
                    "wafer_id": wafer,
                    "column": column,
                })
            _notify_plan_actual_mismatches_once(product, save_mismatches, actor="flow")
            # v8.8.33: notify 이벤트 — 본인이 아닌 원 소유자에게만.
            try:
                from core.notify import emit_event
                for ck, old, val in changed_entries:
                    if old == val:
                        continue
                    target = original_owners.get(ck)
                    if not target or target == username:
                        continue
                    parts = (ck or "").split("|")
                    emit_event(
                        "my_plan_changed",
                        actor=username,
                        target_user=target,
                        title="[plan 변경]",
                        body=f"{username} 가 {product}/{parts[0] if parts else ''} plan 을 변경",
                        payload={
                            "product": product,
                            "cell": ck,
                            "root_lot_id": root_lot_id_req or (parts[0] if parts else ""),
                            "wafer_id": parts[1] if len(parts) > 1 else "",
                            "column": parts[2] if len(parts) > 2 else "",
                            "old": old, "new": val,
                        },
                    )
            except Exception:
                pass
        except Exception as exc:
            logger.warning(f"plan post-save background work failed for {product}: {exc}")

    global _PLAN_POST_SAVE_LAST_THREAD
    _PLAN_POST_SAVE_LAST_THREAD = threading.Thread(target=_plan_post_save, daemon=True, name="splittable-plan-postsave")
    _PLAN_POST_SAVE_LAST_THREAD.start()
    # Plan saves stay in SplitTable history/notifications only; Inform snapshots
    # are attached explicitly from Inform so users do not get extra auto cards.
    _audit_user(req.username, "splittable:plan_save",
                detail=f"product={req.product} saved={len(req.plans)} rejected={len(rejected)}",
                tab="splittable")
    return {"ok": True, "saved": len(req.plans), "rejected": rejected}


class PlanDeleteReq(BaseModel):
    product: str
    cell_keys: list
    username: str = "unknown"
    reason: str = ""


@router.post("/plan/delete")
def delete_plan(req: PlanDeleteReq, request: Request = None):
    if request is not None:
        try:
            me = current_user(request)
            req.username = me.get("username") or req.username or "unknown"
        except Exception:
            raise
    pf = _plan_history_path(req.product)
    if not any(p.exists() for p in _plan_alias_paths(req.product)):
        raise HTTPException(404)
    data = _load_plan_data(req.product)
    now = datetime.datetime.now().isoformat()
    deleted = []
    prior_history = list(data.get("history") or [])
    batch_id = _new_history_batch_id()
    reason = _clean_plan_reason(req.reason)
    new_history: list[dict] = []
    for ck in req.cell_keys:
        if ck in data.get("plans", {}):
            old = data["plans"][ck].get("value")
            prev_owner = str((data["plans"][ck] or {}).get("user") or "")
            del data["plans"][ck]
            cell_root, cell_wafer, cell_column = _split_plan_cell_key(ck)
            entry = {
                "cell": ck, "old": old, "new": None,
                "user": req.username, "time": now, "action": "delete",
                # v10: delete 엔트리에도 root/wafer/column 을 남긴다 — 예전에는
                # root_lot_id 가 없어 cell prefix 매칭에만 의존했다.
                "root_lot_id": cell_root, "wafer_id": cell_wafer,
                "column": cell_column, "batch": batch_id,
                "batch_size": len(req.cell_keys), "prev_user": prev_owner,
                "reason": reason,
            }
            data.setdefault("history", []).append(entry)
            new_history.append(entry)
            deleted.append((ck, old))
    data["history"] = data.get("history", [])[-1000:]
    save_json(pf, data)
    _archive_plan_history(req.product, new_history, prior_history)
    _invalidate_plan_risk_cache(req.product)

    # v9.1.x: knowledge 적재는 백그라운드 — 삭제 응답도 즉시 반환.
    product = req.product
    username = req.username

    def _plan_delete_post():
        try:
            for ck, old in deleted:
                _append_splittable_plan_knowledge(
                    product=product,
                    cell_key=ck,
                    old=old,
                    new=None,
                    actor=username,
                    changed_at=now,
                    conflicting=bool(old not in (None, "")),
                )
        except Exception as exc:
            logger.warning(f"plan delete post work failed for {product}: {exc}")

    global _PLAN_POST_SAVE_LAST_THREAD
    _PLAN_POST_SAVE_LAST_THREAD = threading.Thread(target=_plan_delete_post, daemon=True, name="splittable-plan-delpost")
    _PLAN_POST_SAVE_LAST_THREAD.start()
    # SplitTable plan deletes stay in SplitTable history/notifications only.
    _audit_user(req.username, "splittable:plan_delete",
                detail=f"product={req.product} deleted={len(deleted)}",
                tab="splittable")
    return {"ok": True}


@router.get("/history")
def get_history(product: str = Query(...), root_lot_id: str = Query(""),
                limit: int = Query(500), offset: int = Query(0),
                user: str = Query(""), action: str = Query(""),
                column: str = Query(""), wafer_id: str = Query(""),
                q: str = Query(""), since: str = Query(""), until: str = Query(""),
                has_reason: bool = Query(False)):
    """plan 변경 이력. 응답의 `history` 는 예전과 같이 **시간 오름차순**이고
    기본 동작(필터 없음)도 예전과 같다 — 필터/페이지 필드만 추가됐다.

    `offset` 은 최신 쪽에서부터 건너뛰는 개수다 (offset=0 이 가장 최근 묶음).
    """
    empty_facets = {"users": [], "actions": [], "columns": []}
    if (not any(p.exists() for p in _plan_alias_paths(product))
            and not _plan_history_log_path(product).exists()):
        return {"history": [], "total": 0, "scope_total": 0,
                "returned": 0, "has_more": False, "facets": empty_facets}
    hist = _plan_history_entries(product)
    scoped = [h for h in hist if _history_matches_root(h, root_lot_id)]
    facets = _plan_history_facets(scoped)
    filtered = [h for h in scoped
                if _history_matches_filters(h, user=user, action=action,
                                            column=column, wafer_id=wafer_id,
                                            q=q, since=since, until=until,
                                            has_reason=bool(has_reason))]
    take = max(1, int(limit or 500))
    skip = max(0, int(offset or 0))
    end = max(0, len(filtered) - skip)
    start = max(0, end - take)
    page = filtered[start:end]
    return {"history": page, "total": len(filtered), "scope_total": len(scoped),
            "returned": len(page), "has_more": start > 0, "facets": facets}


@router.get("/operational-history")
def get_operational_history(request: Request, product: str = Query(...),
                            root_lot_id: str = Query(""), wafer_ids: str = Query("")):
    me = current_user(request)
    items = _load_operational_history(
        product=product,
        root_lot_id=root_lot_id,
        wafer_ids=wafer_ids,
        username=me.get("username", ""),
        role=me.get("role", "user"),
    )
    return {"items": items, "total": len(items)}


@router.get("/history/final")
def get_history_final(request: Request, product: str = Query(...), root_lot_id: str = Query(""),
                      include_deleted: bool = Query(False)):
    # v8.8.33 보안: 세션 토큰 필수 (plan history 내 username 노출 방지).
    from core.auth import current_user
    _ = current_user(request)
    """v8.8.33: final-plan-only 뷰.
    각 cell 의 최종 상태(가장 최근 set 또는 delete)만 반환 + plan drift 경고.

    drift 판정:
      - 같은 cell 에 set 이 2회 이상이고 old != new 가 섞임 → drift_level="multi"
      - 서로 다른 user 가 set → drift_level="multi_user"
      - 둘 다 → "multi_user_multi_change"
    """
    payload = _get_plan_risk_payload(product, include_deleted=include_deleted)
    return _copy_plan_risk_payload(payload, root_lot_id=root_lot_id)


@router.get("/history-csv")
def download_history_csv(product: str = Query(...), root_lot_id: str = Query(""),
                         user: str = Query(""), action: str = Query(""),
                         column: str = Query(""), wafer_id: str = Query(""),
                         q: str = Query(""), since: str = Query(""),
                         until: str = Query(""), has_reason: bool = Query(False)):
    """Admin: 화면과 같은 필터가 걸린 전체 이력을 CSV 로. 인자가 없으면 전량."""
    if (not any(p.exists() for p in _plan_alias_paths(product))
            and not _plan_history_log_path(product).exists()):
        raise HTTPException(404, "No history")
    hist = [h for h in _plan_history_entries(product)
            if _history_matches_root(h, root_lot_id)
            and _history_matches_filters(h, user=user, action=action,
                                         column=column, wafer_id=wafer_id,
                                         q=q, since=since, until=until,
                                         has_reason=bool(has_reason))]
    if not hist:
        raise HTTPException(404, "No history entries")

    header = ["time", "user", "action", "root_lot_id", "wafer_id",
              "column", "old_value", "new_value", "reason", "prev_plan_user", "batch"]

    def _rows():
        for h in hist:
            lot, wf, col = _history_cell_parts(h)
            yield [h.get("time", ""), h.get("user", ""), h.get("action", ""),
                   lot, wf, col, h.get("old", ""), h.get("new", ""),
                   h.get("reason", ""), h.get("prev_user", ""), h.get("batch", "")]

    return csv_response(csv_writer_bytes(header, _rows()), f"{product}_history.csv")


def _log_split_table_download(username: str, product: str, root_lot_id: str,
                              prefix: str, custom_name: str, fmt: str,
                              rows: int, cols: int, size_bytes: int,
                              selected: list[str] | None = None) -> None:
    """SplitTable CSV/XLSX 내보내기를 downloads.jsonl(관리자 다운로드 모니터)에 기록."""
    try:
        from core.utils import jsonl_append
        scope = f"custom={custom_name}" if custom_name else f"prefix={prefix or 'all'}"
        sel = selected or []
        jsonl_append(PATHS.download_log, {
            "source": "splittable",
            "username": username or "",
            "product": product,
            "sql": f"root_lot_id={root_lot_id or 'all'}, {scope}, format={fmt}",
            "rows": int(rows or 0), "cols": int(cols or 0),
            "select_cols": ",".join(sel[:8]) + ("…" if len(sel) > 8 else ""),
            "size_mb": round((size_bytes or 0) / 1e6, 2),
        })
    except Exception:
        pass


# ── Transposed CSV ──
@router.get("/download-csv")
def download_csv(product: str = Query(...), root_lot_id: str = Query(""),
                 wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
                 custom_name: str = Query(""), transposed: str = Query("true"),
                 username: str = Query(""),
                 custom_cols: str = Query(""),
                 step_labels: str = Query(""),
                 exclude_not_null: str = Query("1")):
    fp = _product_path(product)
    lf = _scan_product(product, root_lot_id=root_lot_id, wafer_ids=wafer_ids)
    lot_col, wf_col = _detect_lot_wafer(lf)
    lf = _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id, wafer_ids)
    df = lf.collect()

    all_data_cols = _view_data_columns(df.columns, lot_col, wf_col)
    tag_labels = _custom_tag_label_map(product)
    for tag_col in tag_labels:
        if tag_col not in all_data_cols:
            all_data_cols.append(tag_col)
    management_labels = _management_row_label_map(product)
    if custom_name or custom_cols:
        for mgmt_col in management_labels:
            if mgmt_col not in all_data_cols:
                all_data_cols.append(mgmt_col)
    selected = _select_columns(all_data_cols, custom_name, prefix,
                               max_fallback=200, custom_cols=custom_cols)
    if not custom_name and not custom_cols:
        for raw_pref in [p.strip() for p in str(prefix or "").split(",") if p.strip()]:
            for virt in _virtual_columns_for_prefix(product, raw_pref):
                if virt not in selected:
                    selected.append(virt)
    if not custom_name and not custom_cols:
        for raw_pref in [p.strip() for p in str(prefix or "").split(",") if p.strip()]:
            for virt in _virtual_columns_for_prefix(product, raw_pref):
                if virt not in selected:
                    selected.append(virt)
    # v8.8.14: display rename (rule_order + step_desc) 적용.
    # 정렬은 view 와 동일 — ppid_knob step_desc → step_id 공정 순서 우선.
    col_rename = _build_col_rename_map(selected, product)
    col_rename.update({col: f"{CUSTOM_TAG_PREFIX}_{label}" for col, label in tag_labels.items()})
    col_rename.update({col: label for col, label in management_labels.items()})
    _dl_step_rank = _split_step_order_context(product).get("param_rank") or {}
    selected = sorted(selected, key=lambda c: _step_order_sort_key(c, col_rename.get(c, c), _dl_step_rank))
    # 화면에서 "적용 공정 정보"를 켠 채로 받으면 내보내기 항목명도 연결 공정 표기가 된다.
    if _truthy_value(step_labels):
        selected, col_rename = _apply_step_label_columns(
            product, selected, col_rename, exclude_not_null=_truthy_value(exclude_not_null))

    if transposed.lower() == "true" and wf_col and wf_col in df.columns:
        # Resolve wafer values (handle W01 format)
        wf_raw_int = df[wf_col].cast(pl.Int64, strict=False).to_list()
        non_null = [v for v in wf_raw_int if v is not None]
        if non_null:
            wf_vals = wf_raw_int
        else:
            wf_vals = [str(v) for v in df[wf_col].to_list()]
        # v8.4.4: fab_lot_id 로 1차 정렬, wafer 로 2차 정렬 — UI 그룹 순서와 일치
        fab_col = "fab_lot_id" if "fab_lot_id" in df.columns else None
        wf2fab: dict = {}
        if fab_col:
            fab_vals = [(None if v is None else str(v)) for v in df[fab_col].to_list()]
            for w, f in zip(wf_vals, fab_vals):
                if w is None: continue
                if w not in wf2fab and f and f not in ("None","null"):
                    wf2fab[w] = f
        wf_uniq = [w for w in dict.fromkeys(wf_vals) if w is not None and w != "None" and w != "null"]
        # v8.8.3: fab_lot 그룹 → wafer_id 숫자-aware 정렬 (view 와 동일 로직).
        def _wf_sort_key2(w):
            primary = wf2fab.get(w, "~")
            try:
                return (primary, 0, int(w))
            except (TypeError, ValueError):
                s = str(w)
                if s.upper().startswith("W"):
                    try:
                        return (primary, 0, int(s[1:]))
                    except ValueError:
                        pass
                return (primary, 1, s)
        wf_sorted = sorted(wf_uniq, key=_wf_sort_key2)
        headers = [f"#{v}" for v in wf_sorted]
        fab_row = [wf2fab.get(w, "") for w in wf_sorted]
        wf_idx = {v: i for i, v in enumerate(wf_sorted)}

        plans = _load_plan_data(product).get("plans", {})
        tag_values = _custom_tag_values_for_root(product, root_lot_id)
        management_values = _management_row_values_for_root(product, root_lot_id)

        output = io.StringIO()
        writer = csv_mod.writer(output)
        # Header rows (v8.4.4b): downloaded_at, username, root_lot_id, fab_lot_id, Parameter
        download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        writer.writerow(["downloaded_at", download_ts])
        writer.writerow(["username", username or ""])
        writer.writerow(["root_lot_id", root_lot_id or ""])
        if fab_col:
            writer.writerow(["fab_lot_id"] + fab_row)
        writer.writerow(["Parameter"] + headers)
        for col_name in selected:
            row_data = [""] * len(wf_sorted)
            if col_name in tag_labels:
                for idx, wk in enumerate(wf_sorted):
                    row_data[idx] = tag_values.get(f"{root_lot_id}|{wk}|{col_name}", "")
            elif col_name in management_labels:
                for idx, wk in enumerate(wf_sorted):
                    row_data[idx] = management_values.get(f"{root_lot_id}|{wk}|{col_name}", "")
            elif col_name in df.columns:
                vals = df[col_name].to_list()
                for i, v in enumerate(vals):
                    wk = wf_vals[i] if i < len(wf_vals) else None
                    idx = wf_idx.get(wk)
                    if idx is not None:
                        sv = str(v) if v is not None and str(v) not in ("None", "null") else ""
                        ck = f"{root_lot_id}|{wk}|{col_name}"
                        pv = plans.get(ck, {}).get("value")
                        row_data[idx] = pv if pv and not sv else sv
            else:
                for idx, wk in enumerate(wf_sorted):
                    ck = f"{root_lot_id}|{wk}|{col_name}"
                    pv = plans.get(ck, {}).get("value")
                    row_data[idx] = "" if pv is None else str(pv)
            writer.writerow([col_rename.get(col_name, col_name)] + row_data)
        # v8.4.4: Excel 한글 깨짐 방지 — UTF-8 BOM prefix
        csv_bytes = b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
        _log_rows, _log_cols = len(selected), len(wf_sorted)
    else:
        csv_bytes = b"\xef\xbb\xbf" + df.write_csv().encode("utf-8")
        _log_rows, _log_cols = df.height, df.width

    _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                              "csv", _log_rows, _log_cols, len(csv_bytes), selected)
    return csv_response(csv_bytes, f"{product}_{root_lot_id or 'all'}.csv")


SPLIT_CHECK_XLSX_PREFIX_COLUMNS = ["항목", "값", "Split"]


def _export_has_value(value: Any) -> bool:
    text = "" if value is None else str(value).strip()
    return bool(text and text not in {"None", "null"})


def _split_check_export_supported(selected: list[str]) -> bool:
    for column in selected or []:
        up = str(column or "").strip().upper()
        if up in {"INLINE", "VM"} or up.startswith("INLINE_") or up.startswith("VM_"):
            return False
    return True


def _build_split_check_export_rows(
    selected: list[str],
    wafer_count: int,
    value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
    col_rename: dict[str, str] | None = None,
) -> list[list[str]]:
    rows: list[list[str]] = []
    rename = col_rename or {}
    for column in selected or []:
        display_name = str(rename.get(column, column) or column)
        actual_by_idx, plan_by_idx = value_maps.get(column, ({}, {}))
        values_by_idx: dict[int, str] = {}
        order: list[str] = []
        seen: set[str] = set()
        for idx in range(max(0, int(wafer_count or 0))):
            plan_value = plan_by_idx.get(idx, "")
            actual_value = actual_by_idx.get(idx, "")
            value = plan_value if _export_has_value(plan_value) else actual_value
            if not _export_has_value(value):
                continue
            text = str(value)
            values_by_idx[idx] = text
            if text not in seen:
                seen.add(text)
                order.append(text)
        for split_idx, value in enumerate(order):
            label = f"S{split_idx}"
            checks = ["✓" if values_by_idx.get(idx) == value else "" for idx in range(max(0, int(wafer_count or 0)))]
            rows.append([display_name, value, label, *checks])
    return rows


def _split_check_export_param_keys(
    selected: list[str],
    wafer_count: int,
    value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
) -> list[str]:
    """Raw parameter key for every row emitted by _build_split_check_export_rows."""
    keys: list[str] = []
    for column in selected or []:
        actual_by_idx, plan_by_idx = value_maps.get(column, ({}, {}))
        seen: set[str] = set()
        for idx in range(max(0, int(wafer_count or 0))):
            plan_value = plan_by_idx.get(idx, "")
            actual_value = actual_by_idx.get(idx, "")
            value = plan_value if _export_has_value(plan_value) else actual_value
            if not _export_has_value(value):
                continue
            text = str(value)
            if text in seen:
                continue
            seen.add(text)
            keys.append(str(column))
    return keys


def _split_check_param_merges(rows: list[list[str]], start_row: int) -> list[tuple[int, int, int, int]]:
    merges: list[tuple[int, int, int, int]] = []
    current = ""
    run_start = 0
    for idx, row in enumerate([*(rows or []), ["__flow_end__"]]):
        param = str(row[0] if row else "")
        if idx == 0:
            current = param
            run_start = 0
            continue
        if param == current:
            continue
        if current and idx - run_start > 1:
            merges.append((start_row + run_start, 1, start_row + idx - 1, 1))
        current = param
        run_start = idx
    return merges


def _step_progress_not_reached_cells(step_progress: dict, selected: list[str],
                                     wafer_keys: list) -> set[tuple[str, int]]:
    """Return (raw parameter, wafer index) cells shaded as not reached.

    This is the export counterpart of the browser's per-wafer latest-step
    shading.  Prefer by_wafer whenever it exists; root-level metadata is only
    a compatibility fallback.
    """
    progress = step_progress if isinstance(step_progress, dict) else {}
    by_wafer_raw = progress.get("by_wafer") if isinstance(progress.get("by_wafer"), dict) else {}
    by_wafer = {
        _normalize_wafer_id(wafer): {str(v) for v in (meta.get("not_reached") or [])}
        for wafer, meta in by_wafer_raw.items() if isinstance(meta, dict)
    }
    root_set = {str(v) for v in (progress.get("not_reached") or [])}
    out: set[tuple[str, int]] = set()
    for idx, wafer in enumerate(wafer_keys or []):
        cell_set = by_wafer.get(_normalize_wafer_id(wafer), set()) if by_wafer else root_set
        for param in selected or []:
            if str(param) in cell_set:
                out.add((str(param), idx))
    return out


def _exclude_populated_not_reached_cells(
        not_reached_cells: set[tuple[str, int]],
        value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
) -> set[tuple[str, int]]:
    """Grey only cells that are both not reached and empty.

    ML_TABLE can already carry KNOB/INLINE/VM values even when the root is not
    present in FAB.  A visible actual or plan value is stronger evidence for
    the cell display and must keep its normal value/palette styling.
    """
    out: set[tuple[str, int]] = set()
    for raw_param, wafer_idx in not_reached_cells or set():
        actual_by_idx, plan_by_idx = value_maps.get(str(raw_param), ({}, {}))
        if _export_has_value(actual_by_idx.get(wafer_idx)):
            continue
        if _export_has_value(plan_by_idx.get(wafer_idx)):
            continue
        out.add((str(raw_param), wafer_idx))
    return out


@router.get("/download-xlsx")
def download_xlsx(product: str = Query(...), root_lot_id: str = Query(""),
                  wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
                  custom_name: str = Query(""), username: str = Query(""),
                  custom_cols: str = Query(""),
                  display_mode: str = Query(""),
                  step_labels: str = Query(""),
                  exclude_not_null: str = Query("1")):
    """v8.4.4 — XLSX 내보내기. fab_lot_id 행이 동일 값 구간별로 셀 병합되어
    UI 의 그룹 헤더와 동일하게 표시.
    v8.8.33: custom_cols 추가 — save 없이 체크만 한 ad-hoc 컬럼.
    v8.8.34: display_mode=split_check 이면 화면의 Split 체크 표시 행 형식으로 export.
    """
    openpyxl_error = None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        openpyxl_error = e

    lf = _scan_product(product, root_lot_id=root_lot_id, wafer_ids=wafer_ids)
    lot_col, wf_col = _detect_lot_wafer(lf, product)
    lf = _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id, wafer_ids)
    df = lf.collect()

    all_data_cols = _view_data_columns(df.columns, lot_col, wf_col)
    tag_labels = _custom_tag_label_map(product)
    for tag_col in tag_labels:
        if tag_col not in all_data_cols:
            all_data_cols.append(tag_col)
    management_labels = _management_row_label_map(product)
    if custom_name or custom_cols:
        for mgmt_col in management_labels:
            if mgmt_col not in all_data_cols:
                all_data_cols.append(mgmt_col)
    selected = _select_columns(all_data_cols, custom_name, prefix,
                               max_fallback=200, custom_cols=custom_cols)
    # v8.8.14: display rename (rule_order + step_desc) 적용.
    # 정렬은 view 와 동일 — ppid_knob step_desc → step_id 공정 순서 우선.
    col_rename = _build_col_rename_map(selected, product)
    col_rename.update({col: f"{CUSTOM_TAG_PREFIX}_{label}" for col, label in tag_labels.items()})
    col_rename.update({col: label for col, label in management_labels.items()})
    _dl_step_rank = _split_step_order_context(product).get("param_rank") or {}
    selected = sorted(selected, key=lambda c: _step_order_sort_key(c, col_rename.get(c, c), _dl_step_rank))
    # 화면에서 "적용 공정 정보"를 켠 채로 받으면 내보내기 항목명도 연결 공정 표기가 된다.
    step_label_mode = _truthy_value(step_labels)
    if step_label_mode:
        selected, col_rename = _apply_step_label_columns(
            product, selected, col_rename, exclude_not_null=_truthy_value(exclude_not_null))

    wf_raw_int = df[wf_col].cast(pl.Int64, strict=False).to_list() if wf_col else []
    non_null = [v for v in wf_raw_int if v is not None]
    if non_null:
        wf_vals = wf_raw_int
    else:
        wf_vals = [str(v) for v in df[wf_col].to_list()] if wf_col else []
    fab_col = "fab_lot_id" if "fab_lot_id" in df.columns else None
    wf2fab: dict = {}
    if fab_col:
        fab_vals = [(None if v is None else str(v)) for v in df[fab_col].to_list()]
        for w, f in zip(wf_vals, fab_vals):
            if w is None: continue
            if w not in wf2fab and f and f not in ("None","null"):
                wf2fab[w] = f
    wf_uniq = [w for w in dict.fromkeys(wf_vals) if w is not None and w != "None" and w != "null"]
    wf_sorted = sorted(wf_uniq, key=lambda w: (wf2fab.get(w, "~"), w))
    wf_idx = {v: i for i, v in enumerate(wf_sorted)}

    plans = _load_plan_data(product).get("plans", {})
    tag_values = _custom_tag_values_for_root(product, root_lot_id)
    management_values = _management_row_values_for_root(product, root_lot_id)
    split_check_mode = (
        str(display_mode or "").strip().lower() == "split_check"
        and _split_check_export_supported(selected)
    )
    # v9.1.x: 제3 표시형식 — 행에서 왼쪽 값과 같은 칸을 셀 병합해 export (UI 병합 표시와 동일).
    merged_mode = (
        str(display_mode or "").strip().lower() == "merged"
        and not split_check_mode
    )

    def _xlsx_value_maps_for_col(col_name: str) -> tuple[dict[int, str], dict[int, str]]:
        actual_by_idx: dict[int, str] = {}
        plan_by_idx: dict[int, str] = {}
        if col_name in tag_labels:
            for idx, wk in enumerate(wf_sorted):
                tv = tag_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if _export_has_value(tv):
                    actual_by_idx[idx] = str(tv)
        elif col_name in management_labels:
            for idx, wk in enumerate(wf_sorted):
                mv = management_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if _export_has_value(mv):
                    actual_by_idx[idx] = str(mv)
        elif col_name in df.columns:
            vals = df[col_name].to_list()
            for i, v in enumerate(vals):
                wk = wf_vals[i] if i < len(wf_vals) else None
                idx = wf_idx.get(wk)
                if idx is None:
                    continue
                sv = str(v) if _export_has_value(v) else ""
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if _export_has_value(sv):
                    actual_by_idx[idx] = sv
                if _export_has_value(pv):
                    plan_by_idx[idx] = str(pv)
        else:
            for idx, wk in enumerate(wf_sorted):
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if _export_has_value(pv):
                    plan_by_idx[idx] = str(pv)
        return actual_by_idx, plan_by_idx

    if wf2fab:
        fab_present: bool | None = True
    else:
        fab_scope = _fab_history_scope(product, root_lot_id=root_lot_id, limit=1)
        fab_present = (
            True if fab_scope.get("candidates")
            else False if fab_scope.get("query_ok")
            else None
        )
    step_progress = _split_step_progress(
        product, root_lot_id, selected, wf_sorted, fab_present=fab_present)
    not_reached_cells = _step_progress_not_reached_cells(step_progress, selected, wf_sorted)
    value_maps = {col_name: _xlsx_value_maps_for_col(col_name) for col_name in selected}
    not_reached_cells = _exclude_populated_not_reached_cells(not_reached_cells, value_maps)
    split_check_rows: list[list[str]] = []
    split_check_param_keys: list[str] = []
    if split_check_mode:
        split_check_rows = _build_split_check_export_rows(
            selected,
            len(wf_sorted),
            value_maps,
            col_rename,
        )
        split_check_param_keys = _split_check_export_param_keys(selected, len(wf_sorted), value_maps)

    if openpyxl_error is not None:
        try:
            from core.simple_xlsx import build_workbook
            from fastapi.responses import StreamingResponse
        except Exception as e:
            import sys
            raise HTTPException(
                500,
                f"XLSX export unavailable at {sys.executable}: openpyxl={openpyxl_error}; fallback={e}",
            )

        download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        n_wafers = len(wf_sorted)
        prefix_count = len(SPLIT_CHECK_XLSX_PREFIX_COLUMNS) if split_check_mode else 1
        last_col = max(prefix_count + n_wafers, prefix_count + 1) if split_check_mode else prefix_count + n_wafers
        rows = [["downloaded_at", download_ts], ["username", username or ""]]
        merges = []

        if split_check_mode:
            root_row = ["root_lot_id", "", "", root_lot_id or "", *["" for _ in range(max(0, n_wafers - 1))]]
            rows.append(root_row)
            merges.append((3, 1, 3, prefix_count))
            if n_wafers > 1:
                merges.append((3, prefix_count + 1, 3, last_col))

            has_fab_row = bool(fab_col and wf_sorted)
            if has_fab_row:
                fab_row = ["fab_lot_id", "", "", *["" for _ in wf_sorted]]
                cur = None
                start = 0
                row_no = len(rows) + 1
                merges.append((row_no, 1, row_no, prefix_count))
                for i, w in enumerate(wf_sorted):
                    f = wf2fab.get(w, "")
                    if f != cur:
                        if cur is not None and i - start > 0:
                            fab_row[prefix_count + start] = cur
                            if i - start > 1:
                                merges.append((row_no, prefix_count + 1 + start, row_no, prefix_count + i))
                        cur = f
                        start = i
                if cur is not None and len(wf_sorted) - start > 0:
                    fab_row[prefix_count + start] = cur
                    if len(wf_sorted) - start > 1:
                        merges.append((row_no, prefix_count + 1 + start, row_no, prefix_count + len(wf_sorted)))
                rows.append(fab_row)

            header_row_no = len(rows) + 1
            rows.append([*SPLIT_CHECK_XLSX_PREFIX_COLUMNS, *[f"#{w}" for w in wf_sorted]])
            data_start_row = header_row_no + 1
            rows.extend(split_check_rows)
            merges.extend(_split_check_param_merges(split_check_rows, data_start_row))
        else:
            rows.append(["root_lot_id", root_lot_id or "", *["" for _ in range(max(0, n_wafers - 1))]])
            if n_wafers > 1:
                merges.append((3, 2, 3, last_col))

            has_fab_row = bool(fab_col and wf_sorted)
            if has_fab_row:
                fab_row = ["fab_lot_id", *["" for _ in wf_sorted]]
                cur = None
                start = 0
                for i, w in enumerate(wf_sorted):
                    f = wf2fab.get(w, "")
                    if f != cur:
                        if cur is not None and i - start > 0:
                            fab_row[1 + start] = cur
                            if i - start > 1:
                                merges.append((4, 2 + start, 4, 2 + i - 1))
                        cur = f
                        start = i
                if cur is not None and len(wf_sorted) - start > 0:
                    fab_row[1 + start] = cur
                    if len(wf_sorted) - start > 1:
                        merges.append((4, 2 + start, 4, 2 + len(wf_sorted) - 1))
                rows.append(fab_row)

            rows.append(["Parameter", *[f"#{w}" for w in wf_sorted]])
            for col_name in selected:
                display_name = col_rename.get(col_name, col_name)
                actual_by_idx, plan_by_idx = value_maps.get(col_name, ({}, {}))
                out = [display_name, *["" for _ in wf_sorted]]
                for idx in sorted(set(list(actual_by_idx.keys()) + list(plan_by_idx.keys()))):
                    sv = actual_by_idx.get(idx, "")
                    pv = plan_by_idx.get(idx, "")
                    if sv and pv and sv != pv:
                        out[1 + idx] = f"{sv} != {pv}"
                    elif pv and not sv:
                        out[1 + idx] = f"PLAN: {pv}"
                    else:
                        out[1 + idx] = sv or pv
                rows.append(out)
                if merged_mode and n_wafers > 1 and _merge_view_allowed_param(col_name):
                    row_no = len(rows)
                    start = 0
                    for j in range(1, n_wafers + 1):
                        if j == n_wafers or str(out[1 + j]) != str(out[1 + start]):
                            if j - start > 1:
                                merges.append((row_no, 2 + start, row_no, 2 + j - 1))
                            start = j

        data = build_workbook([{"title": product[:31], "rows": rows, "merges": merges}])
        fmt_suffix = "_split_check" if split_check_mode else ("_merged" if merged_mode else "")
        fname = f"{product}_{root_lot_id or 'all'}{fmt_suffix}.xlsx"
        _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                                  f"xlsx{fmt_suffix}",
                                  len(split_check_rows) if split_check_mode else len(selected),
                                  len(wf_sorted), len(data), selected)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = product[:31]
    hdr_fill = PatternFill("solid", fgColor="1f2937")
    fab_fill = PatternFill("solid", fgColor="374151")
    param_fill = PatternFill("solid", fgColor="374151")
    not_reached_fill = PatternFill("solid", fgColor="9CA3AF")
    white = Font(color="FFFFFF", bold=True)
    # fab_lot_id 헤더는 어두운 배경 + 흰 글자로 고정해 노란색 대비 문제를 피한다.
    fab_font = Font(color="FFFFFF", bold=True, name="Consolas", size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _outline_span(row: int, col_start: int, span: int, box: Side) -> None:
        """병합 구간 전체에 테두리 '상자'를 그린다.

        openpyxl 은 값/채우기는 병합 범위의 좌상단 셀만 봐도 되지만 **테두리는
        셀마다 따로 그려진다.** 앵커 셀에만 넣으면 상자가 첫 칸에서 끊기고
        나머지 칸은 아래 그리드 보강 패스가 기본 테두리로 덮어써서, 병합
        다운로드에서 mismatch 빨간 테두리가 반 칸만 나오거나 아예 안 보였다.
        구간 안쪽 세로선은 병합되면 어차피 렌더되지 않으므로 기본 thin 을 둔다."""
        span = max(1, int(span or 1))
        for k in range(span):
            ws.cell(row=row, column=col_start + k).border = Border(
                left=box if k == 0 else thin,
                right=box if k == span - 1 else thin,
                top=box, bottom=box,
            )

    if split_check_mode:
        prefix_count = len(SPLIT_CHECK_XLSX_PREFIX_COLUMNS)
        n_wafers = len(wf_sorted)
        first_wafer_col = prefix_count + 1
        last_col = max(prefix_count + n_wafers, first_wafer_col)
        prefix_fill = PatternFill("solid", fgColor="F9FAFB")
        mark_font = Font(color="000000", bold=True, name="Consolas", size=11)
        prefix_font = Font(color="000000", bold=True, name="Consolas", size=11)
        value_font = Font(color="000000", name="Consolas", size=11)
        palette = [
            ("C6EFCE", "000000"),
            ("FFEB9C", "000000"),
            ("FBE5D6", "000000"),
            ("BDD7EE", "000000"),
            ("E2BFEE", "000000"),
            ("B4DED4", "000000"),
            ("F4CCCC", "000000"),
        ]

        def _split_fill(label: str):
            import re as _re
            m = _re.fullmatch(r"S(\d+)", str(label or "").strip(), flags=_re.I)
            if not m:
                return None
            bg, _fg = palette[int(m.group(1)) % len(palette)]
            return PatternFill("solid", fgColor=bg)

        def _style_cell(cell, *, fill=None, font=None, alignment=None):
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            cell.border = border

        ws.cell(row=1, column=1, value="downloaded_at")
        _style_cell(ws.cell(row=1, column=1), fill=hdr_fill, font=white)
        ws.cell(row=1, column=2, value=download_ts)
        ws.cell(row=2, column=1, value="username")
        _style_cell(ws.cell(row=2, column=1), fill=hdr_fill, font=white)
        ws.cell(row=2, column=2, value=username or "")

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=prefix_count)
        _style_cell(ws.cell(row=3, column=1, value="root_lot_id"), fill=hdr_fill, font=white, alignment=center)
        root_value_col = first_wafer_col
        ws.cell(row=3, column=root_value_col, value=root_lot_id or "")
        if n_wafers > 1:
            ws.merge_cells(start_row=3, start_column=first_wafer_col, end_row=3, end_column=prefix_count + n_wafers)
        for col_idx in range(root_value_col, (prefix_count + n_wafers if n_wafers else root_value_col) + 1):
            _style_cell(ws.cell(row=3, column=col_idx), fill=hdr_fill, font=Font(color="FBBF24", bold=True, name="Consolas", size=13), alignment=center)

        has_fab_row = bool(fab_col and wf_sorted)
        header_row = 5 if has_fab_row else 4
        if has_fab_row:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=prefix_count)
            _style_cell(ws.cell(row=4, column=1, value="fab_lot_id"), fill=hdr_fill, font=white, alignment=center)
            cur = None
            start = 0
            for i, w in enumerate(wf_sorted):
                f = wf2fab.get(w, "")
                if f != cur:
                    if cur is not None and i - start > 0:
                        c = ws.cell(row=4, column=first_wafer_col + start, value=cur)
                        _style_cell(c, fill=fab_fill, font=fab_font, alignment=center)
                        if i - start > 1:
                            ws.merge_cells(start_row=4, start_column=first_wafer_col + start, end_row=4, end_column=first_wafer_col + i - 1)
                    cur = f
                    start = i
            if cur is not None and len(wf_sorted) - start > 0:
                c = ws.cell(row=4, column=first_wafer_col + start, value=cur)
                _style_cell(c, fill=fab_fill, font=fab_font, alignment=center)
                if len(wf_sorted) - start > 1:
                    ws.merge_cells(start_row=4, start_column=first_wafer_col + start, end_row=4, end_column=first_wafer_col + len(wf_sorted) - 1)

        for i, label in enumerate(SPLIT_CHECK_XLSX_PREFIX_COLUMNS, start=1):
            c = ws.cell(row=header_row, column=i, value=label)
            _style_cell(c, fill=param_fill, font=white, alignment=center)
        for i, w in enumerate(wf_sorted):
            c = ws.cell(row=header_row, column=first_wafer_col + i, value=f"#{w}")
            _style_cell(c, fill=param_fill, font=white, alignment=center)

        data_start = header_row + 1
        for r_idx, row in enumerate(split_check_rows, start=data_start):
            raw_param = split_check_param_keys[r_idx - data_start] if r_idx - data_start < len(split_check_param_keys) else ""
            all_not_reached = bool(wf_sorted) and all((raw_param, idx) in not_reached_cells for idx in range(len(wf_sorted)))
            label = str(row[2] if len(row) > 2 else "")
            fill = _split_fill(label)
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if c_idx <= prefix_count:
                    prefix_bg = not_reached_fill if all_not_reached else (fill if c_idx == 3 and fill else prefix_fill)
                    # 적용 공정 표기는 줄바꿈이 섞이므로 항목 칸만 wrap 을 켠다.
                    left_top = Alignment(horizontal="left", vertical="top",
                                         wrap_text=bool(step_label_mode and c_idx == 1))
                    _style_cell(cell, fill=prefix_bg, font=(mark_font if c_idx == 3 else prefix_font), alignment=center if c_idx == 3 else left_top)
                else:
                    wafer_idx = c_idx - first_wafer_col
                    mark_fill = not_reached_fill if (raw_param, wafer_idx) in not_reached_cells else (fill if value else None)
                    _style_cell(cell, fill=mark_fill, font=mark_font if value else value_font, alignment=center)
        for r1, c1, r2, c2 in _split_check_param_merges(split_check_rows, data_start):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            ws.cell(row=r1, column=c1).alignment = Alignment(horizontal="left", vertical="top")

        ws.column_dimensions["A"].width = 28
        ws.column_dimensions["B"].width = 18
        ws.column_dimensions["C"].width = 10
        for i in range(len(wf_sorted)):
            ws.column_dimensions[get_column_letter(first_wafer_col + i)].width = 12
        ws.freeze_panes = f"{get_column_letter(first_wafer_col)}{data_start}"

        last_row = header_row + len(split_check_rows)
        for row_cells in ws.iter_rows(min_row=1, max_row=max(last_row, header_row), min_col=1, max_col=last_col):
            for c in row_cells:
                b = c.border
                if not (b and b.left and b.left.style):
                    c.border = border

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from fastapi.responses import StreamingResponse
        fname = f"{product}_{root_lot_id or 'all'}_split_check.xlsx"
        _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                                  "xlsx_split_check", len(split_check_rows),
                                  len(wf_sorted), buf.getbuffer().nbytes, selected)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    n_wafers = len(wf_sorted)
    last_col = 1 + n_wafers
    # v8.4.4c — downloaded_at / username: 병합하지 않고 label+value 2칸만 표시
    c_ts = ws.cell(row=1, column=1, value="downloaded_at"); c_ts.font = white; c_ts.fill = hdr_fill
    ws.cell(row=1, column=2, value=download_ts)
    # username
    c1 = ws.cell(row=2, column=1, value="username"); c1.font = white; c1.fill = hdr_fill
    ws.cell(row=2, column=2, value=username or "")
    # root_lot_id (v8.4.5c — 병합 복원: wafer 컬럼 전체 colspan)
    c2 = ws.cell(row=3, column=1, value="root_lot_id"); c2.font = white; c2.fill = hdr_fill
    c2v = ws.cell(row=3, column=2, value=root_lot_id or "")
    c2v.alignment = center; c2v.fill = hdr_fill
    c2v.font = Font(color="fbbf24", bold=True, name="Consolas", size=13)
    if n_wafers > 1:
        ws.merge_cells(start_row=3, start_column=2, end_row=3, end_column=last_col)
    # Row 4: fab_lot_id (merged by contiguous groups)
    FAB_ROW = 4
    if fab_col and wf_sorted:
        ws.cell(row=FAB_ROW, column=1, value="fab_lot_id").font = white
        ws.cell(row=FAB_ROW, column=1).fill = hdr_fill
        cur = None; start = 0
        for i, w in enumerate(wf_sorted):
            f = wf2fab.get(w, "")
            if f != cur:
                if cur is not None and i - start > 0:
                    c = ws.cell(row=FAB_ROW, column=2+start, value=cur)
                    c.font = fab_font; c.fill = fab_fill; c.alignment = center; c.border = border
                    if i - start > 1:
                        ws.merge_cells(start_row=FAB_ROW, start_column=2+start, end_row=FAB_ROW, end_column=2+i-1)
                cur = f; start = i
        if cur is not None and len(wf_sorted) - start > 0:
            c = ws.cell(row=FAB_ROW, column=2+start, value=cur)
            c.font = fab_font; c.fill = fab_fill; c.alignment = center; c.border = border
            if len(wf_sorted) - start > 1:
                ws.merge_cells(start_row=FAB_ROW, start_column=2+start,
                               end_row=FAB_ROW, end_column=2+len(wf_sorted)-1)

    # Row 5: Parameter | #1 #2 ...
    param_row = 5 if fab_col else 4
    ws.cell(row=param_row, column=1, value="Parameter").font = white
    ws.cell(row=param_row, column=1).fill = param_fill
    for i, w in enumerate(wf_sorted):
        c = ws.cell(row=param_row, column=2+i, value=f"#{w}")
        c.font = white; c.fill = param_fill; c.alignment = center; c.border = border

    # v8.4.4c: UI 와 동일한 7-색 팔레트 (CELL_COLORS). KNOB_ / MASK_ prefix 행만 컬러링.
    CELL_PALETTE = [
        ("C6EFCE", "006100"),  # green
        ("FFEB9C", "9C5700"),  # yellow
        ("FBE5D6", "BF4E00"),  # orange
        ("BDD7EE", "1F4E79"),  # blue
        ("E2BFEE", "7030A0"),  # purple
        ("B4DED4", "0B5345"),  # teal
        ("F4CCCC", "75194C"),  # pink
    ]
    COLOR_PREFIXES = ("KNOB_", "MASK_")

    for r_off, col_name in enumerate(selected):
        rr = param_row + 1 + r_off
        # v8.8.14: display rename 된 이름을 표기 (원본 col_name 으로는 여전히 df 조회).
        display_name = col_rename.get(col_name, col_name)
        label_cell = ws.cell(row=rr, column=1, value=display_name)
        label_cell.font = Font(bold=True)
        # 적용 공정 표기는 여러 step_id 를 줄바꿈으로 쌓는다 — wrap 없이는 한 줄만 보인다.
        if step_label_mode and "\n" in str(display_name or ""):
            label_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
        if wf_sorted and all((str(col_name), idx) in not_reached_cells for idx in range(len(wf_sorted))):
            label_cell.fill = not_reached_fill
        up = (col_name or "").upper()
        should_color = any(up.startswith(p) for p in COLOR_PREFIXES)
        vals = df[col_name].to_list() if col_name in df.columns else []
        # Build unique-value map — include plan values in palette assignment
        row_values_ordered = []  # preserve column order for uniq index
        actual_by_idx = {}
        plan_by_idx = {}
        if col_name in tag_labels:
            for idx, wk in enumerate(wf_sorted):
                tv = tag_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if tv:
                    actual_by_idx[idx] = str(tv)
        elif col_name in management_labels:
            for idx, wk in enumerate(wf_sorted):
                mv = management_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if mv:
                    actual_by_idx[idx] = str(mv)
        else:
            for i, v in enumerate(vals):
                wk = wf_vals[i] if i < len(wf_vals) else None
                idx = wf_idx.get(wk)
                if idx is None: continue
                sv = str(v) if v is not None and str(v) not in ("None","null") else ""
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if sv: actual_by_idx[idx] = sv
                if pv: plan_by_idx[idx] = str(pv)
        if col_name not in df.columns and col_name not in tag_labels and col_name not in management_labels:
            for idx, wk in enumerate(wf_sorted):
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if pv:
                    plan_by_idx[idx] = str(pv)
        for idx in sorted(set(list(actual_by_idx.keys()) + list(plan_by_idx.keys()))):
            if idx in actual_by_idx: row_values_ordered.append(actual_by_idx[idx])
            elif idx in plan_by_idx: row_values_ordered.append(plan_by_idx[idx])
        uniq_vals = list(dict.fromkeys(row_values_ordered))
        uniq_map = {v: i for i, v in enumerate(uniq_vals)}

        # v8.4.5b: plan 전용 — 진한 주황 테두리 4면 + 이탤릭
        orange_side = Side(style="medium", color="ea580c")
        plan_border = Border(left=orange_side, right=orange_side,
                             top=orange_side, bottom=orange_side)
        red_side = Side(style="medium", color="ef4444")
        mismatch_border = Border(left=red_side, right=red_side,
                                 top=red_side, bottom=red_side)
        if merged_mode and _merge_view_allowed_param(col_name):
            # v9.1.x: 병합 표시 형식 — 왼쪽 칸과 같은 값이면 연속 구간을 셀 병합.
            # KNOB/FAB/MASK 만 대상이다. INLINE/VM/TAG 는 아래 일반 경로로 떨어진다.
            groups = []
            for idx in range(len(wf_sorted)):
                sv = actual_by_idx.get(idx, "")
                pv = plan_by_idx.get(idx, "")
                cell_val = sv or pv
                cell_not_reached = (str(col_name), idx) in not_reached_cells
                if groups and cell_val == groups[-1]["val"] and cell_not_reached == groups[-1]["not_reached"]:
                    groups[-1]["span"] += 1
                else:
                    groups.append({"val": cell_val, "sv": sv, "pv": pv, "start": idx, "span": 1, "not_reached": cell_not_reached})
            for g in groups:
                idx = g["start"]; sv = g["sv"]; pv = g["pv"]; cell_val = g["val"]
                if not cell_val and g["span"] == 1 and not g["not_reached"]:
                    continue
                is_plan_only = (not sv) and bool(pv)
                is_mismatch = bool(sv) and bool(pv) and sv != pv
                cell = ws.cell(row=rr, column=2 + idx, value=cell_val)
                cell.alignment = center
                cell.border = border
                if should_color and cell_val and cell_val in uniq_map:
                    bg, fg = CELL_PALETTE[uniq_map[cell_val] % len(CELL_PALETTE)]
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(color=fg, bold=True, italic=is_plan_only, size=11, name="Consolas")
                elif is_plan_only:
                    cell.fill = PatternFill("solid", fgColor="fef3c7")
                    cell.font = Font(color="ea580c", bold=True, italic=True, name="Consolas")
                if g["not_reached"]:
                    cell.fill = not_reached_fill
                if is_plan_only:
                    _outline_span(rr, 2 + idx, g["span"], orange_side)
                    if cell_val and not str(cell_val).startswith("📌 "):
                        cell.value = "📌 " + str(cell_val)
                elif is_mismatch:
                    # 계획값과 실제값이 다른 셀 — 병합 구간 전체를 빨간 상자로.
                    _outline_span(rr, 2 + idx, g["span"], red_side)
                if g["span"] > 1:
                    ws.merge_cells(start_row=rr, start_column=2 + idx,
                                   end_row=rr, end_column=2 + idx + g["span"] - 1)
            continue
        # 값이 없는 미진행 셀도 실제 cell 객체를 만들어야 회색 fill이 보인다.
        # 값이 있는 셀은 위에서 not_reached_cells에서 제외됐으므로 팔레트/plan
        # 스타일을 그대로 유지한다.
        for idx in range(len(wf_sorted)):
            sv = actual_by_idx.get(idx, "")
            pv = plan_by_idx.get(idx, "")
            cell_val = sv or pv
            is_plan_only = (not sv) and bool(pv)
            is_mismatch = bool(sv) and bool(pv) and sv != pv
            cell = ws.cell(row=rr, column=2+idx, value=cell_val)
            cell.alignment = center
            cell.border = border
            if should_color and cell_val in uniq_map:
                bg, fg = CELL_PALETTE[uniq_map[cell_val] % len(CELL_PALETTE)]
                cell.fill = PatternFill("solid", fgColor=bg)
                if is_plan_only:
                    cell.font = Font(color=fg, italic=True, bold=True, size=11, name="Consolas")
                else:
                    cell.font = Font(color=fg, bold=True, size=11, name="Consolas")
            elif is_plan_only:
                cell.fill = PatternFill("solid", fgColor="fef3c7")
                cell.font = Font(color="ea580c", bold=True, italic=True, name="Consolas")
            if (str(col_name), idx) in not_reached_cells:
                cell.fill = not_reached_fill
            # Plan-only: 진한 주황 테두리 4면 — 눈에 확 띄도록
            if is_plan_only:
                cell.border = plan_border
                # 📌 prefix 접두로 plan 임을 한 번 더 명시
                if not str(cell_val).startswith("📌 "):
                    cell.value = "📌 " + str(cell_val)
            elif is_mismatch:
                cell.border = mismatch_border

    # Column widths
    ws.column_dimensions["A"].width = 28
    for i in range(len(wf_sorted)):
        ws.column_dimensions[get_column_letter(2+i)].width = 14

    # Freeze panes at param_row+1, B
    ws.freeze_panes = f"B{param_row+1}"

    # v8.8.13: 전체 그리드 테두리 보강 — 값 없는 빈 셀·헤더 셀까지 기본 border 적용.
    # plan_border / mismatch_border 처럼 특수 스타일이 이미 들어간 셀은 건너뜀.
    last_row = param_row + len(selected)
    for row_cells in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for c in row_cells:
            b = c.border
            if not (b and b.left and b.left.style):
                c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    fname = f"{product}_{root_lot_id or 'all'}{'_merged' if merged_mode else ''}.xlsx"
    _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                              "xlsx_merged" if merged_mode else "xlsx", len(selected),
                              len(wf_sorted), buf.getbuffer().nbytes, selected)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/plans-csv")
def export_plans_csv(product: str = Query(...)):
    if not any(p.exists() for p in _plan_alias_paths(product)):
        raise HTTPException(404, "No plans")
    plans = _load_plan_data(product).get("plans", {})
    if not plans:
        raise HTTPException(404, "No plans saved")

    header = ["root_lot_id", "wafer_id", "column", "plan_value", "user", "updated"]

    def _rows():
        for cell_key, info in plans.items():
            parts = cell_key.split("|")
            lot = parts[0] if len(parts) > 0 else ""
            wf = parts[1] if len(parts) > 1 else ""
            col = parts[2] if len(parts) > 2 else cell_key
            yield [lot, wf, col, info.get("value", ""),
                   info.get("user", ""), info.get("updated", "")]

    return csv_response(csv_writer_bytes(header, _rows()), f"{product}_plans.csv")
