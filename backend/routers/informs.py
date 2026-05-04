"""routers/informs.py v8.7.0 — 모듈 인폼 시스템 (역할 뷰 + 체크 + flow 상태 + SplitTable 연동 + 이미지 첨부 + 설정형 모듈/사유 + SplitTable 자동기록).

스키마 ({data_root}/informs/informs.json):
  [{
    id, parent_id, wafer_id, lot_id, product,
    module, reason, text, author, created_at,
    checked, checked_by, checked_at,
    flow_status, status_history:[{status, actor, at, note}],
    splittable_change: {column, old_value, new_value, applied} | null
  }]

규약:
  - parent_id 가 null 이면 루트 인폼. 답글/재인폼은 parent_id 로 트리 구성.
  - 수정은 작성자 본인 또는 admin 가능. 삭제는 작성자/admin/모듈 담당자 가능.
  - 체크·상태변경은 해당 인폼 module 을 담당하는 유저 또는 admin.
  - flow_status: received | reviewing | in_progress | completed (순서 강제는 안 함).
  - splittable_change 는 자유형 메타. FE 에서 plan 변경 요약 카드로 렌더.

엔드포인트:
  GET  /api/informs?wafer_id=...        — 특정 wafer 스레드
  GET  /api/informs/recent              — 최근 루트 (role 필터 적용)
  GET  /api/informs/wafers              — 인폼 있는 wafer 목록
  GET  /api/informs/by-lot?lot_id=...   — 해당 lot 의 모든 스레드 (root+전체뷰)
  GET  /api/informs/by-product?product= — 해당 product 인폼 목록
  GET  /api/informs/my                  — 내 모듈 범위 인폼 (담당자용)
  GET  /api/informs/products            — 인폼 기록된 product 목록
  GET  /api/informs/lots                — 인폼 기록된 lot 목록
  GET  /api/informs/modules             — 모듈 드롭다운 옵션 (constants)
  POST /api/informs                     — 생성
  POST /api/informs/delete?id=          — 삭제 (작성자/admin/모듈 담당자)
  POST /api/informs/check?id=           — 체크 토글
  POST /api/informs/status?id=          — flow_status 변경
  POST /api/informs/splittable?id=      — SplitTable 변경요청 attach
"""
import datetime
import html as _html
import json as _json
import mimetypes
import re
import statistics
import sys
import time
import urllib.error
import urllib.request
import uuid
from pathlib import Path
from typing import Any, Dict, List, Optional

_BACKEND_ROOT = Path(__file__).resolve().parents[1]
_APP_ROOT = _BACKEND_ROOT.parent
for _path in (_APP_ROOT, _BACKEND_ROOT):
    _raw = str(_path)
    sys.path[:] = [p for p in sys.path if p != _raw]
    sys.path.insert(0, _raw)

from fastapi import APIRouter, HTTPException, Query, Request, Depends
from fastapi.responses import FileResponse
from pydantic import BaseModel

from core.paths import PATHS
from core.product_dedup import canonical_product, find_duplicate_product, normalize_products
from core.utils import load_json, save_json
from core.auth import current_user, require_admin, require_page_admin
from core.audit import record as _audit
from core.splittable_sets_cache import list_sets as list_cached_splittable_sets
from app_v2.shared.source_adapter import resolve_column
from app_v2.modules.informs.splittable_embed import build_splittable_embed
from routers.groups import user_modules

router = APIRouter(prefix="/api/informs", tags=["informs"])

INFORMS_DIR = PATHS.data_root / "informs"
INFORMS_DIR.mkdir(parents=True, exist_ok=True)
INFORMS_FILE = INFORMS_DIR / "informs.json"
INFORM_AUDIT_FILE = INFORMS_DIR / "audit_log.json"
CONFIG_FILE = INFORMS_DIR / "config.json"
UPLOADS_DIR = INFORMS_DIR / "uploads"
UPLOADS_DIR.mkdir(parents=True, exist_ok=True)
ADMIN_SETTINGS_FILE = PATHS.data_root / "admin_settings.json"
MODULE_KNOB_MAP_FILE = PATHS.data_root / "inform_module_knob_map.json"

# Default 모듈·사유. config.json 에 저장된 값이 있으면 그것을 우선.
DEFAULT_MODULES = ["GATE", "STI", "PC", "MOL", "BEOL", "ET", "EDS", "S-D Epi", "Spacer", "Well", "기타"]
DEFAULT_REASONS = ["PEMS"]  # v9.0.1: 단일 사유 — split table plan 인폼이 주 용도. 추가 사유는 admin 에서 등록.
# v9.x: 등록 → 메일완료 → 등록적용확인 3단계.
# 과거 received/reviewing/in_progress/completed 는 읽기/상태변경에서 호환 수용.
FLOW_STATUSES = ["registered", "mail_completed", "apply_confirmed"]
FLOW_STATUSES_LEGACY = [
    *FLOW_STATUSES,
    "received", "reviewing", "in_progress", "completed",
]
ALLOWED_IMAGE_EXTS = {".png", ".jpg", ".jpeg", ".gif", ".webp", ".bmp"}
MAX_UPLOAD_BYTES = 8 * 1024 * 1024  # 8 MB/이미지
_INFORMS_CACHE_SIG: tuple[float, int] | None = None
_INFORMS_CACHE_ITEMS: list | None = None
INFORM_DASHBOARD_CACHE_TTL = 60.0
_INFORM_DASHBOARD_CACHE: dict[str, tuple[float, dict[str, Any]]] = {}


def _load_config() -> dict:
    data = load_json(CONFIG_FILE, {})
    if not isinstance(data, dict):
        data = {}
    mods = data.get("modules")
    reas = data.get("reasons")
    prods = data.get("products")
    raw_root = data.get("raw_db_root")
    # v8.8.17: 사유별 메일 템플릿. schema: { "<reason>": {"subject": "...", "body": "..."} }
    rt = data.get("reason_templates")
    if not isinstance(mods, list) or not mods:
        mods = list(DEFAULT_MODULES)
    if not isinstance(reas, list) or not reas:
        reas = list(DEFAULT_REASONS)
    if not isinstance(prods, list):
        prods = []
    if not isinstance(raw_root, str):
        raw_root = ""
    if not isinstance(rt, dict):
        rt = {}
    # sanitize — drop entries where subject/body not str.
    rt_clean = {}
    for k, v in rt.items():
        if isinstance(v, dict):
            s = str(v.get("subject", "") or "")
            b = str(v.get("body", "") or "")
            rt_clean[str(k)] = {"subject": s, "body": b}
    return {"modules": mods, "reasons": reas, "products": prods,
            "raw_db_root": raw_root, "reason_templates": rt_clean}


def _fab_db_products() -> list[str]:
    out: list[str] = []
    seen: set[str] = set()
    root = PATHS.db_root / "1.RAWDATA_DB_FAB"
    try:
        if not root.is_dir():
            return out
        for sub in sorted(root.iterdir()):
            if not sub.is_dir():
                continue
            name = _canonical_product(sub.name)
            if not name:
                continue
            key = name.lower()
            if key in seen:
                continue
            seen.add(key)
            out.append(name)
    except Exception:
        return out
    return out


def _merged_catalog_products(extra: list[str] | None = None) -> list[str]:
    merged: list[str] = []
    seen: set[str] = set()
    for src in [*(_load_config().get("products") or []), *_fab_db_products(), *(extra or [])]:
        name = _canonical_product(str(src or ""))
        if not name:
            continue
        key = name.lower()
        if key in seen:
            continue
        seen.add(key)
        merged.append(name)
    return merged


def _save_config(cfg: dict) -> None:
    save_json(CONFIG_FILE, cfg, indent=2)


DEFAULT_MODULE_KNOB_MAP = {
    "GATE": ["GATE_DOSE", "GATE_TIME"],
    "STI": ["STI_DEPTH", "STI_CD"],
    "PC": ["PC_DOSE"],
    "MOL": [],
    "BEOL": [],
    "ET": [],
    "EDS": [],
    "S-D Epi": [],
    "Spacer": [],
    "Well": [],
}


def _normalize_knob_map(raw: Any) -> dict[str, list[str]]:
    if not isinstance(raw, dict):
        return {}
    out: dict[str, list[str]] = {}
    for mod, knobs in raw.items():
        m = str(mod or "").strip()
        if not m:
            continue
        vals = knobs if isinstance(knobs, list) else []
        clean: list[str] = []
        seen: set[str] = set()
        for k in vals:
            name = str(k or "").strip()
            if not name or name in seen:
                continue
            seen.add(name)
            clean.append(name)
        out[m] = clean
    return out


def _load_module_knob_map() -> dict[str, list[str]]:
    data = load_json(MODULE_KNOB_MAP_FILE, {})
    return _normalize_knob_map(data)


def _save_module_knob_map(mapping: dict[str, list[str]]) -> None:
    save_json(MODULE_KNOB_MAP_FILE, _normalize_knob_map(mapping), indent=2)


def _module_highlight_knobs(module: str) -> set[str]:
    mod = str(module or "").strip()
    if not mod:
        return set()
    mapping = _load_module_knob_map()
    knobs = mapping.get(mod) or []
    return {str(k or "").strip().upper() for k in knobs if str(k or "").strip()}


# v8.8.13: 유저별 인폼 모듈 조회 권한. admin_settings.json 의 `inform_user_modules` 에 저장.
#   스키마: { username: [module, ...] }.
#   - admin 은 항상 전체(all_rounder) — 설정값과 무관.
#   - username 이 키에 없으면 기존 `/api/groups/my-modules` 동작 fallback.
#   - 빈 배열은 "아무 모듈도 조회 못함" 으로 해석.
_INFORM_USER_MODS_KEY = "inform_user_modules"


def _inform_user_mods_path():
    return ADMIN_SETTINGS_FILE


def _read_admin_settings() -> dict:
    p = _inform_user_mods_path()
    try:
        if p.is_file():
            with open(p, "r", encoding="utf-8") as f:
                d = _json.load(f)
                return d if isinstance(d, dict) else {}
    except Exception:
        return {}
    return {}


def _write_admin_settings(cfg: dict) -> None:
    p = _inform_user_mods_path()
    p.parent.mkdir(parents=True, exist_ok=True)
    tmp = p.with_suffix(".json.tmp")
    tmp.write_text(_json.dumps(cfg, ensure_ascii=False, indent=2), encoding="utf-8")
    import os as _os
    _os.replace(tmp, p)


def _get_inform_user_mods() -> dict:
    d = _read_admin_settings()
    um = d.get(_INFORM_USER_MODS_KEY) or {}
    return um if isinstance(um, dict) else {}


def build_inform_module_user_index(user_modules_map: Optional[dict] = None) -> dict[str, list[str]]:
    """Return module -> usernames from admin_settings.inform_user_modules."""
    src = user_modules_map if isinstance(user_modules_map, dict) else _get_inform_user_mods()
    out: dict[str, list[str]] = {}
    seen: dict[str, set[str]] = {}
    for username, modules in (src or {}).items():
        un = str(username or "").strip()
        if not un or not isinstance(modules, list):
            continue
        for raw_mod in modules:
            mod = str(raw_mod or "").strip()
            if not mod:
                continue
            bucket = out.setdefault(mod, [])
            used = seen.setdefault(mod, set())
            if un not in used:
                used.add(un)
                bucket.append(un)
    return out


def _module_usernames(module: str) -> list[str]:
    mod = str(module or "").strip()
    if not mod:
        return []
    return build_inform_module_user_index().get(mod, [])


def _module_recipient_rows(module: str) -> list[dict]:
    out: list[dict] = []
    seen_emails: set[str] = set()
    for un in _module_usernames(module):
        emails = _resolve_users_to_emails([un])
        if not emails:
            continue
        em = emails[0]
        key = em.lower()
        if key in seen_emails:
            continue
        seen_emails.add(key)
        out.append({"username": un, "email": em})
    return out


def _user_module_scope(username: str, role: str):
    """인폼 목록 필터링용 모듈 scope 반환.
      - None          : 필터 off (admin 또는 권한 설정 없음 → 기존 group 기반).
      - set({...})    : 이 모듈들만 통과. module 비어있는 인폼은 항상 통과(legacy 보호).
    """
    if role == "admin":
        return None
    um = _get_inform_user_mods()
    if username and username in um:
        return set([str(m) for m in (um[username] or [])])
    return None


def _effective_modules(username: str, role: str) -> set:
    """admin → {"__all__"} sentinel.
    inform_user_modules 에 지정이 있으면 그 set 을 사용(빈 set 포함 = 아무것도 못 봄).
    없으면 groups 기반 user_modules fallback."""
    from routers.groups import user_modules as _um
    if role == "admin":
        return {"__all__"}
    um = _get_inform_user_mods()
    if username and username in um:
        return set(um[username] or [])
    return _um(username, role)


# legacy 변수 — 다른 모듈에서 import 해도 기본값 세트로 동작.
MODULES = list(DEFAULT_MODULES)
REASONS = list(DEFAULT_REASONS)


# ── helpers ────────────────────────────────────────────────────────────
def _load() -> list:
    global _INFORMS_CACHE_SIG, _INFORMS_CACHE_ITEMS
    try:
        st = INFORMS_FILE.stat()
        sig = (st.st_mtime, st.st_size)
    except Exception:
        sig = (0.0, 0)
    if _INFORMS_CACHE_ITEMS is not None and _INFORMS_CACHE_SIG == sig:
        return _INFORMS_CACHE_ITEMS
    data = load_json(INFORMS_FILE, [])
    items = data if isinstance(data, list) else []
    _INFORMS_CACHE_SIG = sig
    _INFORMS_CACHE_ITEMS = items
    return items


def _save(items: list) -> None:
    global _INFORMS_CACHE_SIG, _INFORMS_CACHE_ITEMS
    save_json(INFORMS_FILE, items, indent=2)
    try:
        st = INFORMS_FILE.stat()
        _INFORMS_CACHE_SIG = (st.st_mtime, st.st_size)
        _INFORMS_CACHE_ITEMS = items
    except Exception:
        _INFORMS_CACHE_SIG = None
        _INFORMS_CACHE_ITEMS = None


def _new_id() -> str:
    return f"inf_{datetime.datetime.now().strftime('%y%m%d')}_{uuid.uuid4().hex[:6]}"


def _now() -> str:
    return datetime.datetime.now().isoformat(timespec="seconds")


def _find(items: list, iid: str) -> Optional[dict]:
    return next((x for x in items if x.get("id") == iid), None)


def _is_deleted(entry: dict) -> bool:
    return bool((entry or {}).get("deleted") or (entry or {}).get("deleted_at"))


def _without_deleted(items: list, include_deleted: bool = False) -> list:
    if include_deleted:
        return list(items or [])
    return [x for x in (items or []) if not _is_deleted(x)]


def _actor_from_subject(subject, fallback: str = "") -> str:
    try:
        state_user = getattr(getattr(subject, "state", None), "user", None)
        if isinstance(state_user, dict) and state_user.get("username"):
            return str(state_user["username"])
        if isinstance(subject, str) and subject:
            return subject
    except Exception:
        pass
    return fallback or "anonymous"


def _load_inform_audit() -> list:
    rows = load_json(INFORM_AUDIT_FILE, [])
    return rows if isinstance(rows, list) else []


def _save_inform_audit(rows: list) -> None:
    save_json(INFORM_AUDIT_FILE, list(rows or [])[-10000:], indent=2)


def _audit_record(subject, typ: str, target: Optional[dict], payload: Optional[dict] = None,
                  summary: str = "", at: str = "") -> dict:
    """Append an inform-scoped audit row and mirror it to the global activity log."""
    target = target or {}
    payload = payload if isinstance(payload, dict) else {}
    row = {
        "id": f"aud_{datetime.datetime.now().strftime('%y%m%d')}_{uuid.uuid4().hex[:8]}",
        "type": str(typ or "").strip() or "event",
        "actor": _actor_from_subject(subject, str(payload.get("actor") or "")),
        "target_id": target.get("id") or payload.get("target_id") or "",
        "inform_id": target.get("id") or payload.get("target_id") or "",
        "product": _canonical_product(target.get("product") or payload.get("product") or ""),
        "root_lot_id": target.get("root_lot_id") or payload.get("root_lot_id") or "",
        "lot_id": target.get("lot_id") or payload.get("lot_id") or "",
        "fab_lot_id_at_save": target.get("fab_lot_id_at_save") or payload.get("fab_lot_id_at_save") or "",
        "module": target.get("module") or payload.get("module") or "",
        "summary": str(summary or payload.get("summary") or "")[:500],
        "payload": payload,
        "at": at or _now(),
    }
    try:
        rows = _load_inform_audit()
        rows.append(row)
        _save_inform_audit(rows)
    except Exception:
        pass
    try:
        _audit(subject, f"inform:{row['type']}", detail=row["summary"] or f"id={row['inform_id']}", tab="inform")
    except Exception:
        pass
    return row


def _looks_like_fab_lot(lot_id: str) -> bool:
    """FAB lot 은 보통 split suffix(. / _ / -) 를 포함한다.

    root_lot_id 길이를 5자로 가정하면 LOT029AA 같은 DB 원본 root 가 잘리므로
    구분자가 있는 경우에만 fab lot 으로 보고 root fallback 을 만든다.
    """
    s = str(lot_id or "").strip()
    return bool(s and re.search(r"[._\-/]", s))


def _root_lot_from_values(lot_id: str, embed: Optional[dict] = None) -> str:
    """Inform root key. Prefer the SplitTable snapshot root when available."""
    try:
        st_view = (embed or {}).get("st_view") or {}
        snap_root = str(st_view.get("root_lot_id") or "").strip()
        if snap_root:
            return snap_root
    except Exception:
        pass
    s = str(lot_id or "").strip()
    if not s:
        return ""
    return s[:5] if _looks_like_fab_lot(s) else s


def _wafer_key(value) -> str:
    text = str(value or "").strip()
    if not text:
        return ""
    core = re.sub(r"^(?:WAFER|WF|W)", "", text, flags=re.I).strip()
    try:
        return str(int(core))
    except Exception:
        return text.upper()


def _split_saved_ids(value) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(raw_value) -> None:
        for raw in re.split(r"[,;/\s]+", str(raw_value or "")):
            text = raw.strip()
            if not text:
                continue
            key = text.casefold()
            if key in seen:
                continue
            seen.add(key)
            out.append(text)

    if isinstance(value, (list, tuple, set)):
        for item in value:
            add(item)
    else:
        add(value)
    return out


def _entry_wafer_ids(entry: dict) -> list[str]:
    values = _split_saved_ids((entry or {}).get("wafer_ids_at_save"))
    if values:
        return values
    return _split_saved_ids((entry or {}).get("wafer_id"))


def _wafer_matches(value, query) -> bool:
    target = _wafer_key(query)
    if not target:
        return False
    for part in _split_saved_ids(value):
        key = _wafer_key(part)
        if key and key == target:
            return True
    return False


def _entry_wafer_matches(entry: dict, query) -> bool:
    target = _wafer_key(query)
    if not target:
        return False
    return any(_wafer_key(w) == target for w in _entry_wafer_ids(entry))


def _module_progress_summary(items: list, modules: Optional[list[str]] = None) -> dict:
    configured = [str(m or "").strip() for m in (modules if modules is not None else _load_config().get("modules") or [])]
    configured = [m for m in dict.fromkeys(configured) if m]
    extras = []
    for x in items or []:
        mod = str((x or {}).get("module") or "").strip() or "(미지정)"
        if mod not in configured and mod not in extras:
            extras.append(mod)
    ordered = [*configured, *extras]
    rows = []
    for mod in ordered:
        entries = [
            x for x in (items or [])
            if (str((x or {}).get("module") or "").strip() or "(미지정)") == mod
        ]
        last_at = ""
        completed_at = ""
        checked_at = ""
        mail_count = 0
        for entry in entries:
            ts = str(entry.get("created_at") or "")
            if ts > last_at:
                last_at = ts
            if entry.get("checked"):
                ca = str(entry.get("checked_at") or ts)
                if ca > checked_at:
                    checked_at = ca
            if _is_apply_confirmed(entry):
                hist_times = _status_history_times(entry, {"apply_confirmed"})
                ca = str((hist_times[-1] if hist_times else "") or ts)
                if ca > completed_at:
                    completed_at = ca
            for mh in entry.get("mail_history") or []:
                if isinstance(mh, dict):
                    mail_count += 1
        has_inform = bool(entries)
        completed = bool(completed_at)
        rows.append({
            "module": mod,
            "status": "missing" if not has_inform else ("apply_confirmed" if completed else "registered"),
            "has_inform": has_inform,
            "completed": completed,
            "count": len(entries),
            "mail_count": mail_count,
            "last_at": last_at,
            "completed_at": completed_at,
            "checked_at": checked_at,
        })
    active = [r for r in rows if r["has_inform"]]
    completed_rows = [r for r in rows if r["completed"]]
    missing = [r["module"] for r in rows if not r["has_inform"]]
    pending = [r["module"] for r in rows if r["has_inform"] and not r["completed"]]
    return {
        "modules": rows,
        "total_modules": len(rows),
        "active_modules": len(active),
        "completed_modules": len(completed_rows),
        "missing_modules": missing,
        "pending_modules": pending,
        "inform_count": len(items or []),
        "open_count": sum(1 for x in (items or []) if not _is_apply_confirmed(x or {})),
    }


def _parse_iso_datetime(value) -> Optional[datetime.datetime]:
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        dt = datetime.datetime.fromisoformat(raw.replace("Z", "+00:00"))
        if dt.tzinfo is not None:
            dt = dt.astimezone().replace(tzinfo=None)
        return dt
    except Exception:
        return None


_INFORM_ALL_DAYS = 36500


def _normalize_inform_days(days, default: int = _INFORM_ALL_DAYS) -> int:
    raw = getattr(days, "default", days)
    try:
        n = int(raw)
    except Exception:
        return default
    if n <= 0:
        return 0
    return min(n, default)


def _entry_last_update(entry: dict) -> str:
    candidates = [
        str(entry.get("created_at") or ""),
        str(entry.get("checked_at") or ""),
    ]
    for h in entry.get("status_history") or []:
        if isinstance(h, dict):
            candidates.append(str(h.get("at") or ""))
    for h in entry.get("edit_history") or []:
        if isinstance(h, dict):
            candidates.append(str(h.get("at") or ""))
    return max([x for x in candidates if x] or [""])


def _canonical_flow_status(value: str, entry: Optional[dict] = None) -> str:
    raw = str(value or "").strip()
    if raw in FLOW_STATUSES:
        return raw
    if raw == "completed":
        return "apply_confirmed"
    if raw in ("reviewing", "in_progress"):
        return "mail_completed"
    if raw in ("received", ""):
        if entry and (entry.get("mail_history") or []):
            return "mail_completed"
        return "registered"
    return "registered"


def _is_apply_confirmed(entry: dict) -> bool:
    return _canonical_flow_status((entry or {}).get("flow_status"), entry) == "apply_confirmed"


def _status_history_times(entry: dict, statuses: set[str]) -> list[str]:
    out: list[str] = []
    for h in (entry or {}).get("status_history") or []:
        if not isinstance(h, dict):
            continue
        if _canonical_flow_status(h.get("status")) in statuses:
            at = str(h.get("at") or "")
            if at:
                out.append(at)
    return out


def _lot_matrix_state(entry: dict) -> str:
    return _canonical_flow_status((entry or {}).get("flow_status"), entry) or "registered"


_LOT_MATRIX_STATE_RANK = {
    "pending": 0,
    "registered": 1,
    "mail_completed": 2,
    "apply_confirmed": 3,
}


def _lot_matrix_module_order() -> list[str]:
    admin_order: list[str] = []
    seen: set[str] = set()

    def add(target: list[str], raw) -> None:
        mod = str(raw or "").strip()
        if not mod or mod in seen:
            return
        seen.add(mod)
        target.append(mod)

    for modules in (_get_inform_user_mods() or {}).values():
        if not isinstance(modules, list):
            continue
        for mod in modules:
            add(admin_order, mod)

    cfg_modules = _load_config().get("modules") or []
    if admin_order:
        ordered = list(admin_order)
        for mod in cfg_modules:
            add(ordered, mod)
        return ordered

    ordered: list[str] = []
    for mod in (cfg_modules or DEFAULT_MODULES):
        add(ordered, mod)
    return ordered or list(DEFAULT_MODULES)


def _lot_matrix_cell(entry: dict) -> dict:
    state = _lot_matrix_state(entry)
    return {
        "state": state,
        "inform_id": entry.get("id") or "",
        "checked_at": entry.get("checked_at") or "",
        "created_at": entry.get("created_at") or "",
        "updated_at": _entry_last_update(entry),
        "author": entry.get("author") or "",
        "reason": entry.get("reason") or "",
        "text": entry.get("text") or "",
    }


def _lot_matrix_recent_item(entry: dict) -> dict:
    return {
        "inform_id": entry.get("id") or "",
        "author": entry.get("author") or "",
        "updated_at": _entry_last_update(entry) or entry.get("created_at") or "",
        "state": _lot_matrix_state(entry),
        "reason": entry.get("reason") or "",
        "body_preview": str(entry.get("text") or "").strip().replace("\n", " ")[:80],
    }


def _merge_lot_matrix_cell(prev: Optional[dict], entry: dict) -> dict:
    cell = _lot_matrix_cell(entry)
    recent = [_lot_matrix_recent_item(entry)]
    if prev:
        count = int(prev.get("inform_count") or 1) + 1
        recent = [*(prev.get("recent") or []), *recent]
        recent.sort(key=lambda x: str(x.get("updated_at") or ""), reverse=True)
        # Keep legacy state priority for existing callers while exposing latest rows in recent[].
        base = dict(cell if _lot_matrix_cell_key(cell) >= _lot_matrix_cell_key(prev) else prev)
    else:
        count = 1
        base = cell
    base["inform_count"] = count
    base["recent"] = recent[:5]
    return base


def _lot_matrix_cell_key(cell: dict) -> tuple[int, str]:
    return (
        _LOT_MATRIX_STATE_RANK.get(str(cell.get("state") or ""), 0),
        str(cell.get("updated_at") or cell.get("created_at") or ""),
    )


def _root_lot_module_count_map(items: list[dict]) -> dict[str, dict[str, int]]:
    out: dict[str, dict[str, int]] = {}
    for entry in items or []:
        if entry.get("parent_id"):
            continue
        mod = str(entry.get("module") or "").strip() or "기타"
        for lot_key in _inform_fab_lots(entry):
            if not lot_key:
                continue
            bucket = out.setdefault(lot_key, {})
            bucket[mod] = bucket.get(mod, 0) + 1
    return out


def _attach_root_lot_module_counts(roots: list[dict], visible_roots: list[dict]) -> list[dict]:
    counts_by_lot = _root_lot_module_count_map(visible_roots)
    out = []
    for root in roots or []:
        row = dict(root)
        fab_lots = _inform_fab_lots(row)
        lot_key = fab_lots[0] if fab_lots else str(row.get("lot_id") or row.get("root_lot_id") or "").strip()
        counts = dict(counts_by_lot.get(lot_key) or {})
        row["fab_lot_id"] = lot_key
        row["matrix_lot_id"] = lot_key
        row["root_lot_module_counts"] = counts
        row["informed_modules"] = sorted(counts, key=lambda m: (-counts[m], m))
        row["attachments"] = _attachment_view(row)
        out.append(row)
    return out


def _thread_stats(items: list[dict]) -> dict[str, dict[str, Any]]:
    by_id = {str(x.get("id")): x for x in items or [] if x.get("id")}
    roots: dict[str, str] = {}

    def root_id(entry: dict) -> str:
        eid = str(entry.get("id") or "")
        if eid in roots:
            return roots[eid]
        cur = entry
        seen: set[str] = set()
        while cur.get("parent_id"):
            pid = str(cur.get("parent_id") or "")
            if not pid or pid in seen or pid not in by_id:
                break
            seen.add(pid)
            cur = by_id[pid]
        rid = str(cur.get("id") or eid)
        for sid in seen:
            roots[sid] = rid
        roots[eid] = rid
        return rid

    stats: dict[str, dict[str, Any]] = {}
    for entry in items or []:
        rid = root_id(entry)
        bucket = stats.setdefault(rid, {"reply_count": 0, "thread_updated_at": ""})
        if entry.get("parent_id"):
            bucket["reply_count"] += 1
        for ts in [
            entry.get("created_at"),
            entry.get("updated_at"),
            entry.get("checked_at"),
            *[h.get("at") for h in (entry.get("status_history") or []) if isinstance(h, dict)],
            *[h.get("at") or h.get("sent_at") for h in (entry.get("mail_history") or []) if isinstance(h, dict)],
            *[h.get("at") for h in (entry.get("edit_history") or []) if isinstance(h, dict)],
        ]:
            if ts and str(ts) > bucket["thread_updated_at"]:
                bucket["thread_updated_at"] = str(ts)
    return stats


def _attach_thread_stats(roots: list[dict], items: list[dict]) -> list[dict]:
    stats = _thread_stats(items)
    out = []
    for root in roots or []:
        row = dict(root)
        stat = stats.get(str(row.get("id") or "")) or {}
        row["reply_count"] = int(stat.get("reply_count") or 0)
        row["thread_updated_at"] = stat.get("thread_updated_at") or row.get("updated_at") or row.get("created_at") or ""
        row["attachments"] = _attachment_view(row)
        out.append(row)
    return out


def _add_unique_text(out: list[str], seen: set[str], value) -> None:
    s = str(value or "").strip()
    if not s or s in ("—", "-", "None", "null"):
        return
    key = s.casefold()
    if key in seen:
        return
    seen.add(key)
    out.append(s)


def _extract_fab_lots_from_embed(embed: Optional[dict]) -> list[str]:
    """Read fab_lot_id labels from a saved SplitTable snapshot as rendered in DB view."""
    if not isinstance(embed, dict):
        return []
    st_view = embed.get("st_view") or {}
    if not isinstance(st_view, dict):
        return []
    out: list[str] = []
    seen: set[str] = set()
    for g in st_view.get("header_groups") or []:
        if isinstance(g, dict):
            _add_unique_text(out, seen, g.get("label"))
    for v in st_view.get("wafer_fab_list") or []:
        _add_unique_text(out, seen, v)
    return out


def _split_table_header_groups(st_view: Optional[dict]) -> list[dict]:
    """Return SplitTable-style fab_lot_id header groups aligned to wafer headers."""
    if not isinstance(st_view, dict):
        return []
    headers = list(st_view.get("headers") or [])
    raw_groups = st_view.get("header_groups") or []
    groups: list[dict] = []
    for g in raw_groups:
        if not isinstance(g, dict):
            continue
        label = str(g.get("label") or "").strip()
        try:
            span = max(1, int(g.get("span") or 0))
        except Exception:
            span = 1
        if label and span > 0:
            groups.append({"label": label, "span": span})
    if groups and sum(int(g.get("span") or 0) for g in groups) == len(headers):
        return groups

    wafer_fabs = [str(v or "").strip() for v in (st_view.get("wafer_fab_list") or [])]
    if len(wafer_fabs) != len(headers) or not any(wafer_fabs):
        return []
    out: list[dict] = []
    for value in wafer_fabs:
        label = value or "—"
        if out and out[-1]["label"] == label:
            out[-1]["span"] += 1
        else:
            out.append({"label": label, "span": 1})
    return out


def _upgrade(entry: dict) -> dict:
    """Legacy v8.5.1 레코드에 v8.7.0 필드를 채워 넣는다 (in-place safe copy)."""
    entry.setdefault("lot_id", "")
    # v9.0.x: root_lot_id 는 DB/SplitTable 값을 그대로 보존. 구분자가 있는 fab_lot_id 만
    # legacy fallback 으로 앞 5자를 root 후보로 사용한다.
    if not entry.get("root_lot_id"):
        entry["root_lot_id"] = _root_lot_from_values(entry.get("lot_id") or "", entry.get("embed_table"))
    entry.setdefault("product", "")
    entry.setdefault("checked", False)
    entry.setdefault("checked_by", "")
    entry.setdefault("checked_at", "")
    entry.setdefault("flow_status", "registered" if not entry.get("parent_id") else "")
    entry.setdefault("status_history", [])
    entry.setdefault("splittable_change", None)
    entry.setdefault("images", [])
    entry.setdefault("embed_table", None)
    entry.setdefault("auto_generated", False)
    entry.setdefault("deadline", "")  # v8.7.1: 이슈 마감일 (YYYY-MM-DD 또는 "")
    entry.setdefault("group_ids", [])  # v8.7.6: 그룹 가시성
    entry.setdefault("deleted", False)
    entry.setdefault("deleted_at", "")
    entry.setdefault("deleted_by", "")
    # v8.8.2: status_history 의 `prev` 필드 backfill — legacy 엔트리는
    # prev 가 없어 "확인 취소" 이벤트가 TimelineLog 에서 사라졌다.
    hist = entry.get("status_history") or []
    last_status = ""
    dirty = False
    for h in hist:
        if not isinstance(h, dict):
            continue
        if "prev" not in h:
            h["prev"] = last_status
            dirty = True
        # received/registered 이면서 이전이 completed/apply_confirmed 였다면 자동으로 "확인 취소" note 부여.
        if (_canonical_flow_status(h.get("status")) == "registered"
                and _canonical_flow_status(last_status) == "apply_confirmed"
                and not h.get("note")):
            h["note"] = "확인 취소"
            dirty = True
        last_status = h.get("status") or last_status
    if dirty:
        entry["status_history"] = hist
    return entry


def _group_visible(entry: dict, username: str, role: str) -> bool:
    """v8.7.6: group_ids 기반 가시성. 비어 있으면 public."""
    gids = entry.get("group_ids") or []
    if not gids:
        return True
    if role == "admin":
        return True
    try:
        from routers.groups import user_group_ids as _ugids
        my = _ugids(username, role)
    except Exception:
        my = set()
    return any(g in my for g in gids)


def _load_upgraded() -> list:
    items = _load()
    changed = False
    for x in items:
        before_keys = set(x.keys())
        _upgrade(x)
        if set(x.keys()) != before_keys:
            changed = True
    if changed:
        _save(items)
    return items


def _visible_to(entry: dict, username: str, role: str, my_mods: set) -> bool:
    """admin/all-rounder 전부 통과. 그 외에는 본인이 작성했거나 모듈 담당인 경우.
    v8.7.6: group_ids 가 설정된 인폼은 해당 그룹에 속해야만 추가로 통과."""
    if role == "admin" or "__all__" in my_mods:
        return True
    if not _group_visible(entry, username, role):
        return False
    if entry.get("author") == username:
        return True
    mod = entry.get("module") or ""
    if mod and mod in my_mods:
        return True
    return False


def _can_moderate(entry: dict, username: str, role: str, my_mods: set) -> bool:
    """체크·상태변경 권한: admin 또는 해당 module 담당자 또는 작성자."""
    if role == "admin":
        return True
    if entry.get("author") == username:
        return True
    mod = entry.get("module") or ""
    return bool(mod and mod in my_mods)


def _root_id(items: list, entry: dict) -> str:
    """entry 가 속한 루트 인폼의 id 반환."""
    cur = entry
    seen: set = set()
    while cur and cur.get("parent_id"):
        if cur["id"] in seen:
            break
        seen.add(cur["id"])
        parent = _find(items, cur.get("parent_id"))
        if not parent:
            break
        cur = parent
    return cur.get("id", "") if cur else ""


# ── Dashboard data adapter ─────────────────────────────────────────────
_INFORM_DASHBOARD_METRICS = {
    "count",
    "resolution_rate",
    "first_reply_h",
    "mail_rate",
    "attach_rate",
    "pending_age",
}
_INFORM_DASHBOARD_GROUPBYS = {
    "module",
    "product",
    "root_lot",
    "fab_lot",
    "author",
    "status",
    "date_day",
    "date_week",
    "date_month",
    "hour_of_day",
    "day_of_week",
}
_INFORM_DASHBOARD_STATUSES = ["registered", "mail_completed", "apply_confirmed"]
_INFORM_DASHBOARD_DAYS = ["Mon", "Tue", "Wed", "Thu", "Fri", "Sat", "Sun"]
_INFORM_DASHBOARD_COLORS = [
    "#2563eb", "#dc2626", "#16a34a", "#ca8a04", "#7c3aed", "#0891b2",
    "#db2777", "#4b5563", "#ea580c", "#059669", "#9333ea", "#0f766e",
]


def _inform_dashboard_file_sig() -> tuple[float, int]:
    try:
        st = INFORMS_FILE.stat()
        return (st.st_mtime, st.st_size)
    except Exception:
        return (0.0, 0)


def _inform_dashboard_clone(payload: dict) -> dict:
    try:
        return _json.loads(_json.dumps(payload, ensure_ascii=False))
    except Exception:
        return dict(payload or {})


def _inform_dashboard_color(label: str) -> str:
    text = str(label or "")
    if not text:
        return _INFORM_DASHBOARD_COLORS[0]
    idx = sum(ord(ch) for ch in text) % len(_INFORM_DASHBOARD_COLORS)
    return _INFORM_DASHBOARD_COLORS[idx]


def _inform_period_cutoff(period: str, now: datetime.datetime) -> Optional[datetime.datetime]:
    p = str(period or "all").strip().lower()
    days = {"7d": 7, "30d": 30, "90d": 90}.get(p)
    if not days:
        return None
    return now - datetime.timedelta(days=days)


def _inform_root_id_fast(by_id: dict[str, dict], entry: dict) -> str:
    eid = str((entry or {}).get("id") or "")
    cur = entry or {}
    seen: set[str] = set()
    while cur.get("parent_id"):
        pid = str(cur.get("parent_id") or "")
        if not pid or pid in seen or pid not in by_id:
            break
        seen.add(pid)
        cur = by_id[pid]
    return str(cur.get("id") or eid)


def _inform_visible_threads(items: list[dict], request: Optional[Request]) -> tuple[list[dict], dict[str, list[dict]], dict[str, Any]]:
    by_id = {str(x.get("id") or ""): x for x in (items or []) if x.get("id")}
    root_ids = {str(x.get("id") or "") for x in items or [] if x.get("id") and not x.get("parent_id")}
    if request is not None:
        me = current_user(request)
        role = me.get("role", "user")
        username = me.get("username", "")
        my_mods = _effective_modules(username, role)
        visible_ids = {
            rid
            for rid in root_ids
            if _visible_to(by_id.get(rid) or {}, username, role, my_mods)
        }
        scope = {
            "username": username,
            "role": role,
            "modules": sorted(my_mods),
        }
    else:
        visible_ids = set(root_ids)
        scope = {"username": "__dashboard__", "role": "system", "modules": ["__all__"]}

    roots: list[dict] = []
    children_by_root: dict[str, list[dict]] = {}
    for entry in items or []:
        rid = _inform_root_id_fast(by_id, entry)
        if not rid or rid not in visible_ids:
            continue
        if entry.get("parent_id"):
            children_by_root.setdefault(rid, []).append(entry)
        elif str(entry.get("id") or "") in visible_ids:
            roots.append(entry)
    for children in children_by_root.values():
        children.sort(key=lambda x: str(x.get("created_at") or ""))
    return roots, children_by_root, scope


def _inform_fab_lots(entry: dict) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(value: Any) -> None:
        val = str(value or "").strip()
        if not val:
            return
        key = val.casefold()
        if key in seen:
            return
        seen.add(key)
        out.append(val)

    def add_many(value: Any) -> None:
        for raw in _split_saved_ids(value):
            add(raw)

    add_many((entry or {}).get("fab_lot_id_at_save"))
    if out:
        return out
    for val in _extract_fab_lots_from_embed((entry or {}).get("embed_table")):
        add(val)
    if not out:
        fallback = str((entry or {}).get("lot_id") or "").strip()
        if fallback:
            add(fallback)
    if not out:
        add((entry or {}).get("root_lot_id"))
    return out


def _inform_has_attachment(entry: dict) -> bool:
    if (entry or {}).get("images"):
        return True
    embed = (entry or {}).get("embed_table")
    if isinstance(embed, dict):
        if embed.get("attached_sets"):
            return True
        if embed.get("columns") or embed.get("rows") or embed.get("st_view"):
            return True
    attachments = (entry or {}).get("attachments")
    return bool(attachments)


def _inform_thread_has_mail(root: dict, children: list[dict]) -> bool:
    return any((entry or {}).get("mail_history") for entry in [root, *(children or [])])


def _inform_thread_has_attachment(root: dict, children: list[dict]) -> bool:
    return any(_inform_has_attachment(entry) for entry in [root, *(children or [])])


def _inform_first_reply_hours(root: dict, children: list[dict]) -> Optional[float]:
    start = _parse_iso_datetime((root or {}).get("created_at"))
    if start is None:
        return None
    first: Optional[datetime.datetime] = None
    for child in children or []:
        dt = _parse_iso_datetime(child.get("created_at"))
        if dt is None or dt < start:
            continue
        if first is None or dt < first:
            first = dt
    if first is None:
        return None
    return max(0.0, (first - start).total_seconds() / 3600.0)


def _inform_pending_age_hours(root: dict, now: datetime.datetime) -> Optional[float]:
    if _is_apply_confirmed(root or {}):
        return None
    start = _parse_iso_datetime((root or {}).get("created_at"))
    if start is None:
        return None
    return max(0.0, (now - start).total_seconds() / 3600.0)


def _inform_group_values(root: dict, groupby: str, dt: Optional[datetime.datetime]) -> list[str]:
    g = str(groupby or "module").strip() or "module"
    if g == "module":
        value = str(root.get("module") or "").strip() or "기타"
    elif g == "product":
        value = _canonical_product(root.get("product") or "") or "미지정"
    elif g == "root_lot":
        value = str(root.get("root_lot_id") or _root_lot_from_values(root.get("lot_id") or "", root.get("embed_table"))).strip() or "미지정"
    elif g == "fab_lot":
        vals = _inform_fab_lots(root)
        return vals or ["미지정"]
    elif g == "author":
        value = str(root.get("author") or "").strip() or "미지정"
    elif g == "status":
        value = _canonical_flow_status(root.get("flow_status"), root)
    elif g == "date_day":
        value = dt.strftime("%Y-%m-%d") if dt else "미지정"
    elif g == "date_week":
        if dt:
            iso = dt.isocalendar()
            value = f"{iso.year}-W{iso.week:02d}"
        else:
            value = "미지정"
    elif g == "date_month":
        value = dt.strftime("%Y-%m") if dt else "미지정"
    elif g == "hour_of_day":
        value = f"{dt.hour:02d}" if dt else "미지정"
    elif g == "day_of_week":
        value = _INFORM_DASHBOARD_DAYS[dt.weekday()] if dt else "미지정"
    else:
        value = "미지정"
    return [value]


def _inform_group_sort_key(groupby: str, value: str) -> tuple:
    v = str(value or "")
    if groupby in {"date_day", "date_week", "date_month"}:
        return (0, v)
    if groupby == "hour_of_day":
        try:
            return (0, int(v))
        except Exception:
            return (1, v)
    if groupby == "day_of_week":
        if v in _INFORM_DASHBOARD_DAYS:
            return (0, _INFORM_DASHBOARD_DAYS.index(v))
        return (1, v)
    if groupby == "status":
        if v in _INFORM_DASHBOARD_STATUSES:
            return (0, _INFORM_DASHBOARD_STATUSES.index(v))
        return (1, v)
    return (0, v.casefold())


def _inform_dashboard_filtered_roots(
    roots: list[dict],
    product: str,
    module: str,
    period: str,
    now: datetime.datetime,
) -> list[tuple[dict, Optional[datetime.datetime]]]:
    cutoff = _inform_period_cutoff(period, now)
    want_product = _canonical_product(product or "").casefold()
    want_module = str(module or "").strip().casefold()
    out: list[tuple[dict, Optional[datetime.datetime]]] = []
    for root in roots or []:
        prod = _canonical_product(root.get("product") or "")
        if want_product and prod.casefold() != want_product:
            continue
        mod = str(root.get("module") or "").strip()
        if want_module and mod.casefold() != want_module:
            continue
        dt = _parse_iso_datetime(root.get("created_at"))
        if cutoff is not None and dt is not None and dt < cutoff:
            continue
        out.append((root, dt))
    return out


def _inform_dashboard_pending_table(
    roots_with_dt: list[tuple[dict, Optional[datetime.datetime]]],
    now: datetime.datetime,
    top_n: Optional[int],
) -> dict:
    rows: list[dict[str, Any]] = []
    for root, dt in roots_with_dt:
        age = _inform_pending_age_hours(root, now)
        if age is None:
            continue
        root_lot = str(root.get("root_lot_id") or _root_lot_from_values(root.get("lot_id") or "", root.get("embed_table"))).strip()
        rows.append({
            "root_lot": root_lot,
            "fab_lot": ", ".join(_inform_fab_lots(root)),
            "모듈": str(root.get("module") or "").strip() or "기타",
            "작성자": str(root.get("author") or "").strip() or "미지정",
            "경과시간(h)": round(age, 1),
            "제목": str(root.get("text") or "").strip().replace("\n", " ")[:120],
            "상태": _canonical_flow_status(root.get("flow_status"), root),
            "created_at": (dt.isoformat(timespec="seconds") if dt else str(root.get("created_at") or "")),
            "inform_id": root.get("id") or "",
        })
    rows.sort(key=lambda r: float(r.get("경과시간(h)") or 0.0), reverse=True)
    if top_n:
        rows = rows[:max(1, int(top_n))]
    columns = ["root_lot", "fab_lot", "모듈", "작성자", "경과시간(h)", "제목", "상태", "created_at"]
    return {
        "points": rows,
        "series_order": [],
        "meta": {
            "metric": "pending_age",
            "groupby": "pending_table",
            "unit": "hour",
            "table_columns": columns,
            "total": len(rows),
        },
    }


def _inform_dashboard_reply_buckets(
    roots_with_dt: list[tuple[dict, Optional[datetime.datetime]]],
    children_by_root: dict[str, list[dict]],
) -> dict:
    buckets = [
        ("0-1h", 0.0, 1.0),
        ("1-4h", 1.0, 4.0),
        ("4-24h", 4.0, 24.0),
        ("24-72h", 24.0, 72.0),
        ("72h+", 72.0, None),
        ("답글 없음", None, None),
    ]
    counts = {label: 0 for label, _lo, _hi in buckets}
    for root, _dt in roots_with_dt:
        h = _inform_first_reply_hours(root, children_by_root.get(str(root.get("id") or "")) or [])
        label = "답글 없음"
        if h is not None:
            for cand, lo, hi in buckets:
                if lo is None:
                    continue
                if h >= lo and (hi is None or h < hi):
                    label = cand
                    break
        counts[label] += 1
    points = [{"x": label, "label": label, "y": counts[label]} for label, _lo, _hi in buckets]
    return {
        "points": points,
        "series_order": [],
        "meta": {
            "metric": "first_reply_h",
            "groupby": "first_reply_bucket",
            "unit": "count",
            "total": sum(counts.values()),
        },
    }


def _inform_dashboard_attach_mail_rate(
    roots_with_dt: list[tuple[dict, Optional[datetime.datetime]]],
    children_by_root: dict[str, list[dict]],
) -> dict:
    total = len(roots_with_dt)
    attach = 0
    mail = 0
    for root, _dt in roots_with_dt:
        children = children_by_root.get(str(root.get("id") or "")) or []
        if _inform_thread_has_attachment(root, children):
            attach += 1
        if _inform_thread_has_mail(root, children):
            mail += 1
    denom = total or 1
    points = [
        {"x": "첨부 있음", "label": "첨부 있음", "y": round(attach / denom * 100.0, 2), "count": attach, "total": total, "series": "attach_rate", "color": "attach_rate"},
        {"x": "메일 발송", "label": "메일 발송", "y": round(mail / denom * 100.0, 2), "count": mail, "total": total, "series": "mail_rate", "color": "mail_rate"},
    ]
    return {
        "points": points,
        "series_order": ["attach_rate", "mail_rate"],
        "meta": {
            "metric": "attach_mail_rate",
            "groupby": "rate_kind",
            "unit": "percent",
            "total": total,
        },
    }


def _inform_dashboard_categorical_heatmap(
    roots_with_dt: list[tuple[dict, Optional[datetime.datetime]]],
    x_groupby: str,
    y_groupby: str,
    metric: str,
    children_by_root: dict[str, list[dict]],
    now: datetime.datetime,
    top_n: Optional[int],
) -> dict:
    xg = x_groupby if x_groupby in _INFORM_DASHBOARD_GROUPBYS else "module"
    yg = y_groupby if y_groupby in _INFORM_DASHBOARD_GROUPBYS else "root_lot"
    buckets: dict[tuple[str, str], dict[str, Any]] = {}
    y_totals: dict[str, float] = {}
    for root, dt in roots_with_dt:
        rid = str(root.get("id") or "")
        children = children_by_root.get(rid) or []
        xs = _inform_group_values(root, xg, dt)
        ys = _inform_group_values(root, yg, dt)
        for xv in xs:
            for yv in ys:
                bucket = buckets.setdefault((xv, yv), {
                    "total": 0,
                    "completed": 0,
                    "mail": 0,
                    "attach": 0,
                    "reply_hours": [],
                    "pending_hours": [],
                })
                _inform_dashboard_add_root(bucket, root, children, now)
                y_totals[yv] = y_totals.get(yv, 0.0) + 1
    keep_y = set(y_totals)
    if top_n and len(keep_y) > int(top_n):
        keep_y = {
            y for y, _cnt in sorted(y_totals.items(), key=lambda kv: (-kv[1], _inform_group_sort_key(yg, kv[0])))[:int(top_n)]
        }
    x_values = sorted({x for x, y in buckets if y in keep_y}, key=lambda v: _inform_group_sort_key(xg, v))
    y_values = sorted(keep_y, key=lambda v: (-y_totals.get(v, 0), _inform_group_sort_key(yg, v))) if top_n else sorted(keep_y, key=lambda v: _inform_group_sort_key(yg, v))
    points = []
    for (xv, yv), bucket in buckets.items():
        if yv not in keep_y:
            continue
        val = _inform_dashboard_metric_value(bucket, metric)
        points.append({
            "x": xv,
            "y": yv,
            "cnt": round(float(val), 2) if isinstance(val, float) else val,
            "label": f"{yv} / {xv}",
        })
    points.sort(key=lambda p: (_inform_group_sort_key(yg, p["y"]), _inform_group_sort_key(xg, p["x"])))
    return {
        "points": points,
        "series_order": [],
        "meta": {
            "metric": metric,
            "groupby": f"{xg}x{yg}",
            "unit": _inform_dashboard_metric_unit(metric),
            "heatmap_meta": {
                "kind": "categorical",
                "x_values": x_values,
                "y_values": y_values,
                "x_label": xg,
                "y_label": yg,
            },
        },
    }


def _inform_dashboard_add_root(bucket: dict[str, Any], root: dict, children: list[dict], now: datetime.datetime) -> None:
    bucket["total"] = int(bucket.get("total") or 0) + 1
    if _is_apply_confirmed(root):
        bucket["completed"] = int(bucket.get("completed") or 0) + 1
    if _inform_thread_has_mail(root, children):
        bucket["mail"] = int(bucket.get("mail") or 0) + 1
    if _inform_thread_has_attachment(root, children):
        bucket["attach"] = int(bucket.get("attach") or 0) + 1
    reply_h = _inform_first_reply_hours(root, children)
    if reply_h is not None:
        bucket.setdefault("reply_hours", []).append(reply_h)
    pending_h = _inform_pending_age_hours(root, now)
    if pending_h is not None:
        bucket.setdefault("pending_hours", []).append(pending_h)


def _inform_dashboard_metric_unit(metric: str) -> str:
    if metric in {"resolution_rate", "mail_rate", "attach_rate"}:
        return "percent"
    if metric in {"first_reply_h", "pending_age"}:
        return "hour"
    return "count"


def _inform_dashboard_metric_value(bucket: dict[str, Any], metric: str) -> float | int:
    total = int(bucket.get("total") or 0)
    if metric == "count":
        return total
    if metric == "resolution_rate":
        return round((int(bucket.get("completed") or 0) / total * 100.0) if total else 0.0, 2)
    if metric == "mail_rate":
        return round((int(bucket.get("mail") or 0) / total * 100.0) if total else 0.0, 2)
    if metric == "attach_rate":
        return round((int(bucket.get("attach") or 0) / total * 100.0) if total else 0.0, 2)
    if metric == "first_reply_h":
        vals = [float(v) for v in bucket.get("reply_hours") or []]
        return round((sum(vals) / len(vals)) if vals else 0.0, 2)
    if metric == "pending_age":
        vals = [float(v) for v in bucket.get("pending_hours") or []]
        return round((sum(vals) / len(vals)) if vals else 0.0, 2)
    return total


def build_inform_dashboard_data(
    metric: str = "count",
    groupby: str = "module",
    period: str = "all",
    product: str = "",
    module: str = "",
    request: Optional[Request] = None,
    items: Optional[list[dict]] = None,
    x_groupby: str = "",
    y_groupby: str = "",
    series_groupby: str = "",
    top_n: Optional[int] = None,
    chart_type: str = "",
) -> dict:
    """Return dashboard-ready Inform aggregate data.

    The public endpoint exposes the stable metric/groupby/period/product/module
    contract. Extra parameters are used by dashboard presets for series and
    categorical heatmaps while preserving the same points/series/meta shape.
    """
    metric = str(metric or "count").strip() or "count"
    groupby = str(groupby or "module").strip() or "module"
    period = str(period or "all").strip().lower() or "all"
    x_groupby = str(x_groupby or "").strip()
    y_groupby = str(y_groupby or "").strip()
    series_groupby = str(series_groupby or "").strip()
    chart_type = str(chart_type or "").strip()
    if metric not in _INFORM_DASHBOARD_METRICS and metric not in {"attach_mail_rate"}:
        raise HTTPException(400, f"unsupported metric: {metric}")
    if groupby not in _INFORM_DASHBOARD_GROUPBYS and groupby not in {"first_reply_bucket", "pending_table", "rate_kind"}:
        raise HTTPException(400, f"unsupported groupby: {groupby}")
    if period not in {"all", "7d", "30d", "90d"}:
        raise HTTPException(400, "period must be all/7d/30d/90d")
    try:
        top_n_int = int(top_n) if top_n not in (None, "", 0) else None
    except Exception:
        top_n_int = None

    raw_items = _without_deleted(items if items is not None else _load_upgraded())
    roots, children_by_root, scope = _inform_visible_threads(raw_items, request)
    cache_key = ""
    if items is None:
        cache_key = _json.dumps({
            "sig": _inform_dashboard_file_sig(),
            "metric": metric,
            "groupby": groupby,
            "period": period,
            "product": _canonical_product(product or ""),
            "module": module or "",
            "x_groupby": x_groupby,
            "y_groupby": y_groupby,
            "series_groupby": series_groupby,
            "top_n": top_n_int,
            "chart_type": chart_type,
            "scope": scope,
        }, ensure_ascii=False, sort_keys=True)
        cached = _INFORM_DASHBOARD_CACHE.get(cache_key)
        if cached and time.time() - cached[0] <= INFORM_DASHBOARD_CACHE_TTL:
            return _inform_dashboard_clone(cached[1])

    now = datetime.datetime.now()
    roots_with_dt = _inform_dashboard_filtered_roots(roots, product, module, period, now)

    if metric == "attach_mail_rate" or groupby == "rate_kind":
        payload = _inform_dashboard_attach_mail_rate(roots_with_dt, children_by_root)
    elif groupby == "first_reply_bucket":
        payload = _inform_dashboard_reply_buckets(roots_with_dt, children_by_root)
    elif groupby == "pending_table" or chart_type == "table":
        payload = _inform_dashboard_pending_table(roots_with_dt, now, top_n_int)
    elif x_groupby and y_groupby:
        payload = _inform_dashboard_categorical_heatmap(
            roots_with_dt, x_groupby, y_groupby, metric, children_by_root, now, top_n_int
        )
    else:
        buckets: dict[tuple[str, str], dict[str, Any]] = {}
        for root, dt in roots_with_dt:
            rid = str(root.get("id") or "")
            children = children_by_root.get(rid) or []
            series_values = _inform_group_values(root, series_groupby, dt) if series_groupby else [""]
            for xv in _inform_group_values(root, groupby, dt):
                for sv in series_values:
                    bucket = buckets.setdefault((xv, sv), {
                        "total": 0,
                        "completed": 0,
                        "mail": 0,
                        "attach": 0,
                        "reply_hours": [],
                        "pending_hours": [],
                    })
                    _inform_dashboard_add_root(bucket, root, children, now)

        points: list[dict[str, Any]] = []
        series_seen: list[str] = []
        for (xv, sv), bucket in buckets.items():
            value = _inform_dashboard_metric_value(bucket, metric)
            point: dict[str, Any] = {
                "x": xv,
                "label": xv,
                "y": value,
                "count": int(bucket.get("total") or 0),
                "color": sv or (xv if groupby == "module" else ""),
            }
            if sv:
                point["series"] = sv
                if sv not in series_seen:
                    series_seen.append(sv)
            if metric == "first_reply_h":
                vals = [float(v) for v in bucket.get("reply_hours") or []]
                point["median"] = round(statistics.median(vals), 2) if vals else 0.0
            points.append(point)

        if top_n_int:
            points.sort(key=lambda p: (-float(p.get("y") or 0.0), _inform_group_sort_key(groupby, p.get("x") or "")))
            points = points[:max(1, top_n_int)]
        else:
            points.sort(key=lambda p: (_inform_group_sort_key(groupby, p.get("x") or ""), _inform_group_sort_key(series_groupby, p.get("series") or "")))
        series_order = sorted(series_seen, key=lambda s: _inform_group_sort_key(series_groupby, s)) if series_groupby else []
        payload = {
            "points": points,
            "series_order": series_order,
            "meta": {
                "metric": metric,
                "groupby": groupby,
                "series_groupby": series_groupby,
                "unit": _inform_dashboard_metric_unit(metric),
            },
        }

    payload.setdefault("points", [])
    payload.setdefault("series_order", [])
    payload.setdefault("meta", {})
    payload["meta"] = {
        **(payload.get("meta") or {}),
        "period": period,
        "product": _canonical_product(product or ""),
        "module": module or "",
        "total_roots": len(roots_with_dt),
        "generated_at": now.isoformat(timespec="seconds"),
    }
    if cache_key:
        _INFORM_DASHBOARD_CACHE[cache_key] = (time.time(), _inform_dashboard_clone(payload))
    return payload


@router.get("/dashboard-data")
def dashboard_data(
    request: Request,
    metric: str = Query("count"),
    groupby: str = Query("module"),
    period: str = Query("all"),
    product: str = Query(""),
    module: str = Query(""),
    x_groupby: str = Query(""),
    y_groupby: str = Query(""),
    series_groupby: str = Query(""),
    top_n: int = Query(0, ge=0, le=500),
):
    return build_inform_dashboard_data(
        metric=metric,
        groupby=groupby,
        period=period,
        product=product,
        module=module,
        request=request,
        x_groupby=x_groupby,
        y_groupby=y_groupby,
        series_groupby=series_groupby,
        top_n=top_n or None,
    )


# ── Pydantic ───────────────────────────────────────────────────────────
class SplitChange(BaseModel):
    column: str = ""
    old_value: str = ""
    new_value: str = ""
    applied: bool = False


# v8.8.23: CI (case-insensitive) 컬럼 해석 헬퍼 — splittable.py 와 동일 정신.
#   ML_TABLE_<PROD> 가 대문자 (ROOT_LOT_ID / LOT_ID / FAB_LOT_ID / WAFER_ID) 로 찍히거나,
#   hive 원천이 소문자로 찍혀도 동일 컬럼으로 인식.
def _ci_resolve(name: str, pool) -> str:
    """pool 에서 name 과 alias/casefold 매칭되는 실제 컬럼명 반환. 없으면 '' ."""
    hit = resolve_column(list(pool or []), name)
    return (hit.matched if hit else "") or ""


def _ci_pick_first(candidates, pool) -> str:
    """candidates 중 pool 에 CI 존재하는 첫 이름(실제 casing) 반환."""
    for c in candidates:
        got = _ci_resolve(c, pool)
        if got:
            return got
    return ""


# v8.8.15: 인폼 저장 시점의 fab_lot_id 를 ML_TABLE 에서 resolve.
#   FE 가 명시적으로 보내지 않았을 때만 호출. splittable 의 /view 가 fab_lot_id 를 join 해 주는 로직을
#   재구현하면 비용/결합이 크므로, 여기서는 ML_TABLE_<PRODUCT>.parquet 에서 lot_id==root_lot_id 인
#   가장 최근 행의 fab_lot_id 하나만 싸게 조회 (실패해도 "" 반환).
# v8.8.23: 컬럼명을 case-insensitive 로 매칭 — ML_TABLE 이 대문자(ROOT_LOT_ID/FAB_LOT_ID) 로
#   찍혀도 정상 추출. 기존엔 literal 비교라 대/소문자 혼재 시 "" 반환 → fab_lot_id 누락.
def _resolve_fab_lot_snapshot(product: str, lot_id: str, wafer_id: str) -> str:
    try:
        if not product or not (lot_id or wafer_id):
            return ""
        # Prefer the same coalesced SplitTable pipeline users see in DB/SplitTable.
        try:
            from routers.splittable import resolve_fab_lot_snapshot as _split_fab_snapshot
            root_key = _root_lot_from_values(lot_id)
            got = _split_fab_snapshot(product, root_key, wafer_id)
            if got:
                return got
        except Exception:
            pass
        import polars as pl  # runtime optional
        from core.roots import get_base_root as _get_base_root
        base_dir = _get_base_root()
        if not base_dir or not Path(base_dir).exists():
            return ""
        base_dir = Path(base_dir)
        # product 는 "ML_TABLE_<X>" 형태이거나 순수 제품명일 수 있음.
        stem = product if product.startswith("ML_TABLE_") else f"ML_TABLE_{product}"
        candidates = [base_dir / f"{stem}.parquet", base_dir / f"{stem.upper()}.parquet"]
        fp = next((c for c in candidates if c.exists()), None)
        if not fp:
            return ""
        lf = pl.scan_parquet(fp)
        try:
            from core.utils import filter_valid_wafer_ids_lazy
            lf = filter_valid_wafer_ids_lazy(lf)
        except Exception:
            pass
        names = list(lf.collect_schema().names()) if hasattr(lf, "collect_schema") else list(lf.schema.keys())
        fab_col = _ci_resolve("fab_lot_id", names)
        if not fab_col:
            return ""
        # lot_id / wafer_id 컬럼 CI 감지 — root_lot_id 우선.
        lot_col = _ci_pick_first(("root_lot_id", "lot_id"), names)
        wf_col = _ci_resolve("wafer_id", names)
        if not lot_col and not wf_col:
            return ""
        key = _root_lot_from_values(lot_id)
        expr = None
        if lot_col and key:
            expr = pl.col(lot_col).cast(pl.Utf8) == key
        elif wf_col and wafer_id:
            expr = pl.col(wf_col).cast(pl.Utf8) == str(wafer_id)
        if expr is None:
            return ""
        # 최신 fab_lot_id 하나만 뽑기 (order-by 없이 임의의 첫 값 — 스냅샷 용도로 충분).
        df = lf.filter(expr).select(pl.col(fab_col).cast(pl.Utf8)).head(1).collect()
        if df.height == 0 and lot_col and key and len(key) > 5:
            df = (
                lf.filter(pl.col(lot_col).cast(pl.Utf8).str.starts_with(key[:5]))
                .select(pl.col(fab_col).cast(pl.Utf8))
                .head(1)
                .collect()
            )
        if df.height == 0:
            return ""
        v = df.item(0, 0)
        return str(v or "").strip()
    except Exception:
        return ""


def _first_saved_id(value) -> str:
    vals = _split_saved_ids(value)
    return vals[0] if vals else ""


def _embed_lot_identity(embed: Optional[dict], fab_lot_id: str = "") -> dict[str, Any]:
    if not isinstance(embed, dict):
        return {"root_lot_id": "", "wafer_ids": [], "matched_fab": False}
    st = embed.get("st_view") or {}
    if not isinstance(st, dict):
        return {"root_lot_id": "", "wafer_ids": [], "matched_fab": False}
    root_lot = str(st.get("root_lot_id") or "").strip()
    headers = [str(v or "").strip() for v in (st.get("headers") or [])]
    wafer_fabs = [str(v or "").strip() for v in (st.get("wafer_fab_list") or [])]
    fab = str(fab_lot_id or "").strip()
    wafers: list[str] = []
    matched = False
    if fab and len(headers) == len(wafer_fabs):
        for idx, wafer_fab in enumerate(wafer_fabs):
            if wafer_fab.casefold() != fab.casefold():
                continue
            matched = True
            wafer = headers[idx].strip()
            if wafer:
                wafers.append(wafer)
    elif not fab:
        wafers = [h for h in headers if h]
    return {
        "root_lot_id": root_lot,
        "wafer_ids": list(dict.fromkeys(wafers)),
        "matched_fab": matched,
    }


def _fab_lot_scope_snapshot(product: str, fab_lot_id: str) -> dict[str, Any]:
    fab = str(fab_lot_id or "").strip()
    if not product or not fab:
        return {}
    try:
        from routers.splittable import _fab_history_scope
        scope = _fab_history_scope(product, fab_lot_id=fab, limit=1000) or {}
        return {
            "root_ids": [str(v).strip() for v in (scope.get("root_ids") or []) if str(v or "").strip()],
            "wafer_ids": [str(v).strip() for v in (scope.get("wafer_ids") or []) if str(v or "").strip()],
            "source": scope.get("source") or ("match_cache" if scope.get("cache") else "fab_history"),
        }
    except Exception:
        return {}


def _resolve_lot_identity_snapshot(
    product: str,
    lot_id: str,
    requested_wafer_id: str,
    embed: Optional[dict],
    explicit_fab_lot_id: str = "",
) -> dict[str, Any]:
    """Resolve and freeze the LOT_ID -> ROOT_LOT_ID/WAFER_ID mapping at create time."""
    lot = str(lot_id or "").strip()
    requested_wafers = _split_saved_ids(requested_wafer_id)
    explicit_fab = _first_saved_id(explicit_fab_lot_id)
    fab = explicit_fab or (lot if _looks_like_fab_lot(lot) else "")
    root_hint = "" if fab else _root_lot_from_values(lot, embed)
    if not fab and root_hint:
        fab = _resolve_fab_lot_snapshot(product, root_hint, requested_wafers[0] if requested_wafers else "")

    scope = _fab_lot_scope_snapshot(product, fab) if fab else {}
    embed_meta = _embed_lot_identity(embed, fab)
    root_ids = list(dict.fromkeys(scope.get("root_ids") or []))
    wafer_ids = list(dict.fromkeys(requested_wafers or (scope.get("wafer_ids") or [])))
    source = str(scope.get("source") or "")

    if not root_ids and (not fab or embed_meta.get("matched_fab")) and embed_meta.get("root_lot_id"):
        root_ids = [embed_meta["root_lot_id"]]
        source = source or "embed"
    if not wafer_ids and (not fab or embed_meta.get("matched_fab")):
        wafer_ids = list(dict.fromkeys(embed_meta.get("wafer_ids") or []))
        if wafer_ids:
            source = source or "embed"
    if not root_ids and root_hint:
        root_ids = [root_hint]
        source = source or "fallback"
    if not root_ids and lot:
        root_ids = [_root_lot_from_values(lot)]
        source = source or "fallback"
    if not fab:
        embed_fabs = _extract_fab_lots_from_embed(embed)
        fab = ", ".join(embed_fabs)

    return {
        "root_lot_id": root_ids[0] if root_ids else "",
        "root_lot_ids": root_ids,
        "wafer_id": ", ".join(wafer_ids),
        "wafer_ids": wafer_ids,
        "fab_lot_id": explicit_fab_lot_id.strip() or fab,
        "source": source or "unknown",
    }


class ImageRef(BaseModel):
    filename: str
    url: str
    size: int = 0


class EmbedTable(BaseModel):
    source: str = ""          # 예: "SplitTable/PROD_A"
    columns: List[str] = []
    rows: List[List] = []
    note: str = ""
    # v8.8.33: SplitTable /view 원형 응답 보존 — EmbedTableView 가 컬러링/plan-pin 동일 렌더.
    #   st_view: {headers, rows, wafer_fab_list, header_groups}
    #   st_scope: {prefix, custom_name, inline_cols} — 어떤 범위로 찍혔는지 재생 가능.
    st_view: Optional[dict] = None
    st_scope: Optional[dict] = None
    attached_sets: List[dict] = []


class InformCreate(BaseModel):
    # v8.7.9: wafer_id 선택 필드. 없으면 lot_id 로 자동 채움 (스레드 묶기 용).
    wafer_id: str = ""
    lot_id: str = ""
    product: str = ""
    module: str = ""
    reason: str = ""
    text: str = ""
    parent_id: Optional[str] = None
    splittable_change: Optional[SplitChange] = None
    images: List[ImageRef] = []
    embed_table: Optional[EmbedTable] = None
    attached_sets: List[dict] = []
    # v8.7.9: deadline 필드 폐기. 호환을 위해 스키마에 남겨 두되 저장하지 않음.
    deadline: str = ""
    group_ids: List[str] = []  # v8.7.6: 그룹 가시성 필터. 비어 있으면 public (모듈 규칙만 적용)
    # v8.8.15: 저장 시점의 fab_lot_id 스냅샷 — FE SplitTable 맥락에서 resolve 된 값.
    #   이후 ML_TABLE 이 재빌드되어 fab_lot_id 매핑이 바뀌어도, 인폼이 가리키던 fab_lot_id 는 보존된다.
    fab_lot_id_at_save: str = ""


class ConfigReq(BaseModel):
    modules: Optional[List[str]] = None
    reasons: Optional[List[str]] = None
    products: Optional[List[str]] = None
    raw_db_root: Optional[str] = None
    # v8.8.17: { "<reason>": {"subject": "...", "body": "..."} } — admin 만 편집.
    reason_templates: Optional[dict] = None


class ModuleKnobMapReq(BaseModel):
    module: str
    knobs: List[str] = []


class SplitTableSnapshotReq(BaseModel):
    product: str
    lot_id: str
    custom_cols: List[str] = []
    is_fab_lot: Optional[bool] = None


class ProductReq(BaseModel):
    product: str


class StatusReq(BaseModel):
    status: str
    note: str = ""


def _sanitize_attached_sets(rows: list[dict] | None) -> list[dict]:
    out = []
    for row in rows or []:
        if not isinstance(row, dict):
            continue
        columns = [str(c) for c in (row.get("columns") or [])][:80]
        data_rows = []
        for values in (row.get("rows") or [])[:200]:
            if isinstance(values, list):
                data_rows.append([str(v) if v is not None else "" for v in values[:80]])
        out.append({
            "id": str(row.get("id") or "")[:160],
            "name": str(row.get("name") or "")[:160],
            "source": str(row.get("source") or "")[:40],
            "columns_count": int(row.get("columns_count") or len(columns) or 0),
            "wafer_count": int(row.get("wafer_count") or len(data_rows) or 0),
            "updated_at": str(row.get("updated_at") or "")[:80],
            "owner": str(row.get("owner") or "")[:80],
            "columns": columns,
            "rows": data_rows,
        })
    return out[:20]


def _sanitize_embed_table(embed_table: Optional[EmbedTable]) -> Optional[dict]:
    if not embed_table:
        return None
    attached_sets = _sanitize_attached_sets(embed_table.attached_sets or [])
    if not (embed_table.columns or embed_table.rows or embed_table.st_view or embed_table.st_scope or attached_sets):
        return None
    cols = [str(c) for c in (embed_table.columns or [])][:80]
    rows = []
    for r in (embed_table.rows or [])[:200]:
        rows.append([str(x) if x is not None else "" for x in (r or [])[:80]])
    return {
        "source": (embed_table.source or "").strip()[:160],
        "columns": cols,
        "rows": rows,
        "note": (embed_table.note or "").strip()[:500],
        "st_view": embed_table.st_view or None,
        "st_scope": embed_table.st_scope or None,
        "attached_sets": attached_sets,
    }


def _embed_with_attached_sets(embed: Optional[dict], attached_sets: list[dict] | None) -> Optional[dict]:
    attached = _sanitize_attached_sets(attached_sets or [])
    if not attached:
        return embed
    base = dict(embed or {})
    current = [x for x in (base.get("attached_sets") or []) if isinstance(x, dict)]
    seen = {(str(x.get("source") or ""), str(x.get("id") or x.get("name") or "")) for x in current}
    for item in attached:
        key = (str(item.get("source") or ""), str(item.get("id") or item.get("name") or ""))
        if key not in seen:
            current.append(item)
            seen.add(key)
    first = next((x for x in current if x.get("columns") and x.get("rows")), None)
    if first and not base.get("columns") and not base.get("rows"):
        base["columns"] = list(first.get("columns") or [])
        base["rows"] = list(first.get("rows") or [])
    base["attached_sets"] = current
    base["source"] = str(base.get("source") or "SplitTable selected sets")
    base["note"] = str(base.get("note") or f"{len(current)} selected set(s) attached")
    return base


def _attachment_view(entry: dict) -> list[dict]:
    embed = entry.get("embed_table") if isinstance(entry, dict) else {}
    if not isinstance(embed, dict):
        return []
    return [dict(x) for x in (embed.get("attached_sets") or []) if isinstance(x, dict)]


class CheckReq(BaseModel):
    checked: bool


class DeadlineReq(BaseModel):
    deadline: str = ""  # YYYY-MM-DD 또는 "" (해제)


def _validate_deadline(s: str) -> str:
    s = (s or "").strip()
    if not s:
        return ""
    try:
        datetime.date.fromisoformat(s[:10])
        return s[:10]
    except Exception:
        raise HTTPException(400, "deadline 포맷: YYYY-MM-DD")


# ── Endpoints ──────────────────────────────────────────────────────────
@router.get("/modules")
def list_modules():
    cfg = _load_config()
    return {
        "modules": cfg["modules"],
        "reasons": cfg["reasons"],
        "flow_statuses": FLOW_STATUSES,
    }


@router.get("/modules/recipients")
def module_recipients(request: Request, module: str = Query("")):
    _ = current_user(request)
    mod = (module or "").strip()
    if not mod:
        raise HTTPException(400, "module required")
    return {"module": mod, "recipients": _module_recipient_rows(mod)}


@router.get("/modules/knob-map")
def get_module_knob_map(request: Request):
    _ = current_user(request)
    return {"knob_map": _load_module_knob_map()}


@router.get("/modules/summary")
def module_summary(request: Request, days: int = Query(_INFORM_ALL_DAYS, ge=0, le=_INFORM_ALL_DAYS)):
    """Module-wise Inform counts for the left summary pane.

    Shape is intentionally a plain list for the frontend contract:
    [{module, registered, mail_completed, apply_confirmed, pending}].
    """
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    days = _normalize_inform_days(days)
    now = datetime.datetime.now()
    cutoff = None if days <= 0 or days >= _INFORM_ALL_DAYS else now - datetime.timedelta(days=days)

    def in_window(entry: dict) -> bool:
        raw = str(entry.get("created_at") or "")
        if not raw:
            return True
        if cutoff is None:
            return True
        try:
            ts = datetime.datetime.fromisoformat(raw.replace("Z", "+00:00"))
            if ts.tzinfo is not None:
                ts = ts.astimezone().replace(tzinfo=None)
            return ts >= cutoff
        except Exception:
            return True

    cfg_modules = [
        str(m or "").strip()
        for m in (_load_config().get("modules") or DEFAULT_MODULES)
        if str(m or "").strip()
    ]
    ordered = list(dict.fromkeys(cfg_modules))
    def empty_bucket(mod: str) -> dict[str, int | str]:
        return {
            "module": mod,
            "registered": 0,
            "mail_completed": 0,
            "apply_confirmed": 0,
            "pending": 0,
            # Legacy aliases for older UI/tests.
            "received": 0,
            "in_progress": 0,
            "completed": 0,
        }

    buckets: dict[str, dict[str, int | str]] = {mod: empty_bucket(mod) for mod in ordered}
    items = _without_deleted(_load_upgraded())
    for entry in items:
        if entry.get("parent_id"):
            continue
        if not in_window(entry):
            continue
        if not _visible_to(entry, me["username"], me.get("role", "user"), my_mods):
            continue
        mod = str(entry.get("module") or "").strip() or "기타"
        if mod not in buckets:
            buckets[mod] = empty_bucket(mod)
            ordered.append(mod)
        status = _canonical_flow_status(entry.get("flow_status"), entry)
        if status == "apply_confirmed":
            buckets[mod]["apply_confirmed"] = int(buckets[mod]["apply_confirmed"]) + 1
            buckets[mod]["completed"] = int(buckets[mod]["completed"]) + 1
        elif status == "mail_completed":
            buckets[mod]["mail_completed"] = int(buckets[mod]["mail_completed"]) + 1
            buckets[mod]["in_progress"] = int(buckets[mod]["in_progress"]) + 1
            buckets[mod]["pending"] = int(buckets[mod]["pending"]) + 1
        else:
            buckets[mod]["registered"] = int(buckets[mod]["registered"]) + 1
            buckets[mod]["received"] = int(buckets[mod]["received"]) + 1
            buckets[mod]["pending"] = int(buckets[mod]["pending"]) + 1
    return [buckets[m] for m in ordered]


@router.get("/lot-matrix")
def lot_matrix(
    request: Request,
    product: str = Query(""),
    days: int = Query(_INFORM_ALL_DAYS, ge=0, le=_INFORM_ALL_DAYS),
    search: str = Query(""),
):
    """Product/fab-lot x module progress matrix for the left Inform pane."""
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    days = _normalize_inform_days(days)
    cutoff = None if days <= 0 or days >= _INFORM_ALL_DAYS else datetime.datetime.now() - datetime.timedelta(days=days)
    want_product = _canonical_product(product or "")
    lot_query = str(search or "").strip().casefold()

    def in_window(entry: dict) -> bool:
        dt = _parse_iso_datetime(_entry_last_update(entry) or entry.get("created_at"))
        if dt is None:
            return True
        if cutoff is None:
            return True
        return dt >= cutoff

    module_order = _lot_matrix_module_order()
    module_seen = set(module_order)
    products_by_key: dict[str, dict] = {}

    for entry in _without_deleted(_load_upgraded()):
        if entry.get("parent_id"):
            continue
        if not _visible_to(entry, me["username"], me.get("role", "user"), my_mods):
            continue
        if not in_window(entry):
            continue

        prod = _canonical_product(entry.get("product") or "")
        if want_product and prod.casefold() != want_product.casefold():
            continue

        lot_keys = _inform_fab_lots(entry)
        if not lot_keys:
            continue
        if lot_query and not any(lot_query in lot_key.casefold() for lot_key in lot_keys):
            continue

        mod = str(entry.get("module") or "").strip() or "기타"
        if mod not in module_seen:
            module_seen.add(mod)
            module_order.append(mod)

        product_label = prod or "미지정"
        product_key = product_label.casefold()
        product_bucket = products_by_key.setdefault(product_key, {
            "product": product_label,
            "lots_by_key": {},
            "module_totals": {},
        })
        root_lot = str(entry.get("root_lot_id") or _root_lot_from_values(entry.get("lot_id") or "", entry.get("embed_table"))).strip()
        latest_update = str((_lot_matrix_recent_item(entry)).get("updated_at") or "")
        for lot_key in lot_keys:
            lot_bucket = product_bucket["lots_by_key"].setdefault(lot_key, {
                "root_lot_id": lot_key,  # legacy FE key; value is now the FAB lot matrix key.
                "lot_id": lot_key,
                "fab_lot_id": lot_key,
                "source_root_lot_id": root_lot,
                "modules": {},
                "progress": {"done": 0, "total": 0},
                "last_update": "",
            })
            if root_lot and not lot_bucket.get("source_root_lot_id"):
                lot_bucket["source_root_lot_id"] = root_lot

            prev = lot_bucket["modules"].get(mod)
            cell = _merge_lot_matrix_cell(prev, entry)
            lot_bucket["modules"][mod] = cell
            product_bucket["module_totals"][mod] = int(product_bucket["module_totals"].get(mod) or 0) + 1
            if latest_update > lot_bucket["last_update"]:
                lot_bucket["last_update"] = latest_update

    products_out: list[dict] = []
    for product_bucket in products_by_key.values():
        lots = []
        for lot_bucket in product_bucket["lots_by_key"].values():
            done = sum(
                1
                for mod in module_order
                if (lot_bucket["modules"].get(mod) or {}).get("state") == "apply_confirmed"
            )
            lot_bucket["progress"] = {"done": done, "total": len(module_order)}
            lots.append(lot_bucket)
        lots.sort(key=lambda x: (x.get("last_update") or "", x.get("fab_lot_id") or x.get("lot_id") or ""), reverse=True)
        products_out.append({
            "product": product_bucket["product"],
            "lots": lots,
            "module_totals": product_bucket.get("module_totals") or {},
        })

    products_out.sort(key=lambda x: str(x.get("product") or "").casefold())
    return {"products": products_out, "module_order": module_order}


_AUDIT_TYPE_ALIASES = {
    "상태변경": "status_change",
    "상태": "status_change",
    "메일": "mail",
    "댓글": "comment",
    "수정": "edit",
    "생성": "create",
    "삭제": "delete",
    "status": "status_change",
    "status_change": "status_change",
    "mail": "mail",
    "comment": "comment",
    "edit": "edit",
    "create": "create",
    "delete": "delete",
}


@router.get("/audit-log")
def audit_log(
    request: Request,
    products: List[str] = Query(default=[]),
    products_bracket: List[str] = Query(default=[], alias="products[]"),
    modules: List[str] = Query(default=[]),
    modules_bracket: List[str] = Query(default=[], alias="modules[]"),
    lot_search: str = Query(""),
    days: int = Query(_INFORM_ALL_DAYS, ge=0, le=_INFORM_ALL_DAYS),
    types: List[str] = Query(default=[]),
    types_bracket: List[str] = Query(default=[], alias="types[]"),
    start: str = Query(""),
    end: str = Query(""),
):
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _load_upgraded()
    targets = {
        str(x.get("id") or ""): x
        for x in items
        if x.get("id") and _visible_to(x, me["username"], me.get("role", "user"), my_mods)
    }

    product_filter = {
        _canonical_product(str(p or "")).casefold()
        for p in [*products, *products_bracket]
        if str(p or "").strip()
    }
    module_filter = {str(m or "").strip() for m in [*modules, *modules_bracket] if str(m or "").strip()}
    type_filter = {
        _AUDIT_TYPE_ALIASES.get(str(t or "").strip(), str(t or "").strip())
        for t in [*types, *types_bracket]
        if str(t or "").strip()
    }
    lot_q = str(lot_search or "").strip().casefold()
    days = _normalize_inform_days(days)
    now = datetime.datetime.now()
    cutoff = None if days <= 0 or days >= _INFORM_ALL_DAYS else now - datetime.timedelta(days=days)
    start_day = str(start or "").strip()[:10]
    end_day = str(end or "").strip()[:10]

    def in_window(row: dict) -> bool:
        raw = str(row.get("at") or "")
        day = raw[:10]
        if start_day and day and day < start_day:
            return False
        if end_day and day and day > end_day:
            return False
        if start_day or end_day:
            return True
        if cutoff is None:
            return True
        dt = _parse_iso_datetime(raw)
        return True if dt is None else dt >= cutoff

    out = []
    for idx, row in enumerate(_load_inform_audit()):
        if not isinstance(row, dict):
            continue
        inform_id = str(row.get("inform_id") or row.get("target_id") or "")
        target = targets.get(inform_id)
        if inform_id and target is None:
            continue
        enriched = {**row, "_idx": idx}
        if target:
            for key in ("product", "root_lot_id", "lot_id", "fab_lot_id_at_save", "module"):
                enriched[key] = enriched.get(key) or target.get(key) or ""
            enriched["deleted"] = _is_deleted(target)
        typ = _AUDIT_TYPE_ALIASES.get(str(enriched.get("type") or "").strip(), str(enriched.get("type") or "").strip())
        enriched["type"] = typ
        prod = _canonical_product(enriched.get("product") or "").casefold()
        if product_filter and prod not in product_filter:
            continue
        if module_filter and str(enriched.get("module") or "").strip() not in module_filter:
            continue
        if type_filter and typ not in type_filter:
            continue
        if lot_q:
            hay = " ".join(str(enriched.get(k) or "") for k in ("root_lot_id", "lot_id", "fab_lot_id_at_save")).casefold()
            if lot_q not in hay:
                continue
        if not in_window(enriched):
            continue
        out.append(enriched)

    out.sort(key=lambda x: (str(x.get("at") or ""), int(x.get("_idx") or 0)), reverse=True)
    for row in out:
        row.pop("_idx", None)
    return {"audit": out, "logs": out, "count": len(out)}


@router.post("/modules/knob-map")
def set_module_knob_map(
    req: ModuleKnobMapReq,
    request: Request,
    _perm=Depends(require_page_admin("informs")),
):
    me = current_user(request)
    from core.auth import is_page_admin
    if me.get("role") != "admin" and not is_page_admin(me.get("username") or "", "informs"):
        raise HTTPException(403, "admin or informs page_admin only")
    mod = (req.module or "").strip()
    if not mod:
        raise HTTPException(400, "module required")
    mapping = _load_module_knob_map()
    clean: list[str] = []
    seen: set[str] = set()
    for raw in req.knobs or []:
        knob = str(raw or "").strip()
        if not knob or knob in seen:
            continue
        seen.add(knob)
        clean.append(knob)
    mapping[mod] = clean
    _save_module_knob_map(mapping)
    _audit(request, "inform:module-knob-map", detail=f"module={mod} knobs={len(clean)}", tab="inform")
    return {"ok": True, "module": mod, "knobs": clean, "knob_map": mapping}


@router.get("/config")
def get_config():
    """v8.8.33: products 필드를 읽는 시점에 normalize (trim + case-insensitive dedup).
    기존 admin_settings 에 'PRODA' / 'PRODA ' 같은 잉여가 있어도 FE 에서는 1개로 노출."""
    cfg = _load_config()
    cfg = dict(cfg)
    cfg["products"] = _merged_catalog_products()
    return cfg


@router.get("/settings")
def get_settings_compat():
    """Compatibility alias for older PageGear builds."""
    return get_config()


@router.post("/config")
def save_config_endpoint(req: ConfigReq, _admin=Depends(require_page_admin("informs"))):
    """Admin 또는 informs page_admin 전용 — 모듈/사유 옵션 목록 편집."""
    cfg = _load_config()
    if req.modules is not None:
        cfg["modules"] = [m.strip() for m in req.modules if m and m.strip()]
    if req.reasons is not None:
        cfg["reasons"] = [r.strip() for r in req.reasons if r and r.strip()]
    if req.products is not None:
        cfg["products"] = [p.strip() for p in req.products if p and p.strip()]
    if req.raw_db_root is not None:
        cfg["raw_db_root"] = req.raw_db_root.strip()
    # v8.8.17: 사유별 메일 템플릿 upsert. None = 변경 없음, {} = 전체 비움.
    if req.reason_templates is not None:
        rt_in = req.reason_templates or {}
        rt_out = {}
        if isinstance(rt_in, dict):
            for k, v in rt_in.items():
                if isinstance(v, dict):
                    s = str(v.get("subject", "") or "")[:500]
                    b = str(v.get("body", "") or "")[:10000]
                    if s or b:
                        rt_out[str(k)[:100]] = {"subject": s, "body": b}
        cfg["reason_templates"] = rt_out
    # de-dup 유지 순서
    cfg["modules"] = list(dict.fromkeys(cfg["modules"]))
    cfg["reasons"] = list(dict.fromkeys(cfg["reasons"]))
    cfg["products"] = list(dict.fromkeys(cfg.get("products") or []))
    if not cfg["modules"]:
        cfg["modules"] = list(DEFAULT_MODULES)
    if not cfg["reasons"]:
        cfg["reasons"] = list(DEFAULT_REASONS)
    _save_config(cfg)
    resp = dict(cfg)
    resp["products"] = _merged_catalog_products(resp.get("products") or [])
    return {"ok": True, "config": resp}


@router.post("/settings")
def save_settings_compat(req: ConfigReq, _admin=Depends(require_page_admin("informs"))):
    """Compatibility alias for older PageGear builds."""
    return save_config_endpoint(req, _admin)


@router.post("/splittable-snapshot")
def splittable_snapshot(req: SplitTableSnapshotReq, request: Request):
    """Build the Inform SplitTable embed via the app_v2 service layer."""
    current_user(request)
    embed = build_splittable_embed(
        product=req.product,
        lot_id=req.lot_id,
        custom_cols=req.custom_cols,
        is_fab_lot=req.is_fab_lot,
    )
    return {"ok": True, "embed": embed}


@router.get("/splittable-sets")
def splittable_sets(request: Request, product: str = Query("")):
    current_user(request)
    return list_cached_splittable_sets(product)


# v8.8.13: 유저별 인폼 모듈 조회 권한 엔드포인트 ────────────────────────
class UserModulesSaveReq(BaseModel):
    username: str
    modules: List[str] = []


@router.get("/user-modules")
def list_user_modules(request: Request):
    """Admin: 인폼 탭 접근 가능한 유저 + 각자의 현재 모듈 권한.
    인폼 탭 권한이 있는 유저(tabs 에 'inform' 또는 '__all__') 만 노출."""
    me = current_user(request)
    if me.get("role") != "admin":
        raise HTTPException(403, "admin only")
    from routers.auth import read_users
    um = _get_inform_user_mods()
    out = []
    for u in read_users():
        if u.get("status") != "approved":
            continue
        tabs = (u.get("tabs") or "").strip()
        has_inform = (tabs == "__all__") or ("inform" in [t.strip() for t in tabs.split(",")])
        if u.get("role") != "admin" and not has_inform:
            continue
        un = u.get("username") or ""
        out.append({
            "username": un,
            "role": u.get("role", "user"),
            "email": u.get("email") or "",
            "modules": list(um.get(un, [])),
            "has_setting": un in um,
        })
    return {"users": out}


@router.post("/user-modules/save")
def save_user_modules(req: UserModulesSaveReq, request: Request):
    """Admin: 특정 유저의 인폼 모듈 조회 권한 저장. 빈 배열 = '아무 모듈도 조회 못함'."""
    me = current_user(request)
    if me.get("role") != "admin":
        raise HTTPException(403, "admin only")
    uname = (req.username or "").strip()
    if not uname:
        raise HTTPException(400, "username required")
    cfg = _read_admin_settings()
    um = dict(cfg.get(_INFORM_USER_MODS_KEY) or {})
    mods = [str(m).strip() for m in (req.modules or []) if str(m).strip()]
    um[uname] = list(dict.fromkeys(mods))
    cfg[_INFORM_USER_MODS_KEY] = um
    _write_admin_settings(cfg)
    _audit(request, "inform:user-modules",
           detail=f"user={uname} modules={','.join(um[uname])}", tab="inform")
    return {"ok": True, "username": uname, "modules": um[uname]}


@router.post("/user-modules/clear")
def clear_user_modules(req: UserModulesSaveReq, request: Request):
    """Admin: 특정 유저의 권한 설정 완전 제거 → group 기반 fallback 으로 복귀."""
    me = current_user(request)
    if me.get("role") != "admin":
        raise HTTPException(403, "admin only")
    uname = (req.username or "").strip()
    if not uname:
        raise HTTPException(400, "username required")
    cfg = _read_admin_settings()
    um = dict(cfg.get(_INFORM_USER_MODS_KEY) or {})
    um.pop(uname, None)
    cfg[_INFORM_USER_MODS_KEY] = um
    _write_admin_settings(cfg)
    return {"ok": True, "username": uname, "cleared": True}


@router.get("/my-modules")
def my_inform_modules(request: Request):
    """현재 유저의 인폼 모듈 조회 권한.
      - admin → all_rounder=True
      - inform_user_modules 에 저장된 값 있으면 그걸 사용
      - 그 외엔 /api/groups/my-modules 값으로 fallback
    """
    me = current_user(request)
    uname = me.get("username") or ""
    role = me.get("role") or "user"
    if role == "admin":
        return {"modules": [], "all_rounder": True, "source": "admin"}
    um = _get_inform_user_mods()
    if uname in um:
        return {"modules": list(um[uname]), "all_rounder": False, "source": "inform_user_modules"}
    # fallback: groups.user_modules 에서 compute
    try:
        from routers.groups import user_modules
        mods = user_modules(uname, role) or set()
        # "__all__" sentinel 은 admin 경로에서만 나오므로 여기선 없음.
        return {"modules": list(mods), "all_rounder": False, "source": "groups"}
    except Exception:
        return {"modules": [], "all_rounder": False, "source": "fallback"}


# v8.8.1: 제품 카탈로그 CRUD (모든 로그인 유저 — 등록된 제품 선택용).
# v8.8.33: 모든 요청 시 catalog 자체를 normalize (trim + case-insensitive dedup) —
#   기존 데이터에 "PRODA", "PRODA " 같은 중복이 있으면 이 경로에서 한번에 정리.
def _normalize_products(products: list) -> list:
    """Back-compat wrapper around core.product_dedup.normalize_products."""
    return normalize_products(products)


@router.post("/products/add")
@router.put("/products/add")
@router.patch("/products/add")
@router.post("/product/add")
@router.put("/product/add")
@router.patch("/product/add")
def add_product(req: ProductReq, request: Request):
    # v8.8.33 보안: admin 또는 page_admin('informs') 만 카탈로그 변경.
    from core.auth import is_page_admin
    me = current_user(request)
    if me.get("role") != "admin" and not is_page_admin(me.get("username") or "", "informs"):
        raise HTTPException(403, "admin or informs page_admin only")
    p = _canonical_product(req.product or "")
    if not p:
        raise HTTPException(400, "product required")
    cfg = _load_config()
    products = _normalize_products(list(cfg.get("products") or []))
    dup = find_duplicate_product(products, p)
    if dup:
        raise HTTPException(409, {"code": "duplicate_product", "existing_product": dup})
    products.append(p)
    cfg["products"] = products
    _save_config(cfg)
    _audit(request, "inform:product_add", detail=f"product={p} by={me['username']}", tab="inform")
    return {"ok": True, "products": products}


@router.post("/products")
@router.put("/products")
@router.patch("/products")
def add_product_collection_compat(req: ProductReq, request: Request):
    """Compatibility for cached clients that POST to the products collection."""
    return add_product(req, request)


@router.get("/products/add")
@router.get("/product/add")
def add_product_get_compat(request: Request, product: str = Query("")):
    """Back-compat for older Inform UI builds that used a query-string add call."""
    return add_product(ProductReq(product=product), request)


@router.post("/products/dedup")
def dedup_products_all_sources(request: Request):
    """v8.8.33: admin 전용 one-shot — config.json.products, product_contacts.json.products 키,
    심지어 informs.json 레코드의 product 값까지 trim + case-insensitive 로 정규화하고 저장.
    사이드바 'PRODA 2개' 같은 유령 중복을 근본 청소.
    """
    me = current_user(request)
    if me.get("role") != "admin":
        raise HTTPException(403, "admin only")
    report = {}
    # 1) config.products
    cfg = _load_config()
    before = list(cfg.get("products") or [])
    after = _normalize_products(before)
    if before != after:
        cfg["products"] = after
        _save_config(cfg)
    report["catalog"] = {"before": len(before), "after": len(after), "values": after}
    # 2) product_contacts.json
    pc = _load_product_contacts()   # 로드 단계에서 이미 병합됨
    _save_product_contacts(pc)       # 병합 결과 디스크 기록
    report["contacts"] = {"products": list(pc.get("products", {}).keys())}
    # 3) informs records product 값
    try:
        items = _load_upgraded()
        updated = 0
        for it in items:
            p = it.get("product")
            if isinstance(p, str):
                trimmed = p.strip()
                if trimmed != p:
                    it["product"] = trimmed
                    updated += 1
        if updated:
            _save(items)
        report["records_trimmed"] = updated
    except Exception as e:
        report["records_error"] = str(e)
    return {"ok": True, "report": report}


@router.post("/products/delete")
def delete_product(req: ProductReq, request: Request):
    """v8.8.1: 카탈로그에서 제품 삭제. admin 또는 등록자(추적불가시 admin) 권한.
    실제 인폼 레코드(product 필드)는 건드리지 않음 — 드롭다운에서만 제외.
    v8.8.33: case-insensitive 매칭 + 중복 전부 제거 → 기존 PRODA/"PRODA " 같은 유령 제거.
    v8.8.33 보안: admin 또는 page_admin('informs') 만 삭제 가능."""
    from core.auth import is_page_admin
    me = current_user(request)
    if me.get("role") != "admin" and not is_page_admin(me.get("username") or "", "informs"):
        raise HTTPException(403, "admin or informs page_admin only")
    p = (req.product or "").strip()
    if not p:
        raise HTTPException(400, "product required")
    cfg = _load_config()
    before = _normalize_products(list(cfg.get("products") or []))
    target_key = p.lower()
    after = [x for x in before if x.lower() != target_key]
    if len(after) == len(before):
        raise HTTPException(404, "product not in catalog")
    cfg["products"] = after
    _save_config(cfg)
    _audit(request, "inform:product_delete", detail=f"product={p} by={me['username']}", tab="inform")
    return {"ok": True, "products": after}


@router.get("/product-lots")
def list_product_lots(request: Request, product: str = Query(...)):
    """v8.8.1: Admin 이 설정한 raw_db_root 에서 제품별 Lot 후보 스캔.
    스캔 위치: {raw_db_root}/1.RAWDATA_DB/{product}/  (서브폴더 이름을 lot 으로 간주).
    폴더가 없거나 설정 안 된 경우 빈 리스트."""
    _ = current_user(request)
    cfg = _load_config()
    root = (cfg.get("raw_db_root") or "").strip()
    product = (product or "").strip()
    if not root or not product:
        return {"product": product, "lots": [], "source": ""}
    try:
        # 표준 경로. 필요시 여러 후보 검색.
        candidates = [
            Path(root) / "1.RAWDATA_DB" / product,
            Path(root) / product,
        ]
        target = next((c for c in candidates if c.exists() and c.is_dir()), None)
        if not target:
            return {"product": product, "lots": [], "source": str(candidates[0])}
        lots = sorted({d.name for d in target.iterdir() if d.is_dir() and not d.name.startswith(".")})
        return {"product": product, "lots": lots, "source": str(target)}
    except Exception as e:
        return {"product": product, "lots": [], "source": root, "error": str(e)}


# ── Image upload / serving ────────────────────────────────────────────
_SAFE_NAME_RE = re.compile(r"[^A-Za-z0-9._-]")


def _safe_filename(name: str) -> str:
    name = Path(name).name  # strip dirs
    name = _SAFE_NAME_RE.sub("_", name)
    return name[-120:] or "file"


async def _read_upload_payload(request: Request) -> tuple[str, bytes]:
    """Read multipart field `file` without a FastAPI File dependency.

    Using `UploadFile = File(...)` makes FastAPI validate python-multipart at
    router import time. If that package is missing in an operator environment,
    the entire informs router fails to load. Keep the dependency runtime-only
    so non-upload informs APIs stay available.
    """
    try:
        form = await request.form()
    except Exception as exc:
        raise HTTPException(
            500,
            "파일 업로드 파서가 준비되지 않았습니다. 서버에서 `python setup.py install-deps` "
            f"또는 `{sys.executable} -m pip install python-multipart` 실행이 필요합니다: {exc}",
        )
    file = form.get("file")
    if file is None or not hasattr(file, "read"):
        raise HTTPException(400, "file 필드가 필요합니다.")
    filename = str(getattr(file, "filename", "") or "")
    data_or_coro = file.read()
    data = await data_or_coro if hasattr(data_or_coro, "__await__") else data_or_coro
    if isinstance(data, str):
        data = data.encode("utf-8")
    return filename, bytes(data or b"")


@router.post("/upload")
async def upload_image(request: Request):
    """인폼용 이미지 업로드. 유저당 세션으로만 가능 (current_user 검증)."""
    me = current_user(request)
    filename, data = await _read_upload_payload(request)
    ext = Path(filename or "").suffix.lower()
    if ext not in ALLOWED_IMAGE_EXTS:
        raise HTTPException(400, f"이미지 형식만 업로드 가능합니다 ({', '.join(sorted(ALLOWED_IMAGE_EXTS))}).")
    if len(data) > MAX_UPLOAD_BYTES:
        raise HTTPException(413, "파일이 너무 큽니다 (최대 8MB).")
    if not data:
        raise HTTPException(400, "빈 파일입니다.")

    uid = uuid.uuid4().hex[:12]
    safe = _safe_filename(filename or ("image" + ext))
    if not safe.lower().endswith(ext):
        safe += ext
    subdir = UPLOADS_DIR / uid
    subdir.mkdir(parents=True, exist_ok=True)
    dst = subdir / safe
    dst.write_bytes(data)

    url = f"/api/informs/files/{uid}/{safe}"
    return {"ok": True, "filename": safe, "url": url, "size": len(data),
            "uploaded_by": me["username"]}


# v8.8.18: 메일 첨부용 범용 업로드 — 이미지뿐 아니라 xlsx/pptx/pdf 등 모든 파일.
#   인폼 메일 다이얼로그에서 사용자가 선택한 파일을 먼저 이 엔드포인트로 올려
#   반환 URL 을 send-mail 의 attachments 에 포함. `_resolve_inform_attachment`
#   가 동일 경로 규약(/api/informs/files/<uid>/<name>)으로 이미 해석 가능.
_ATTACHMENT_MAX_BYTES = 10 * 1024 * 1024   # 10MB 개별 / send-mail 에서 총합 10MB 한도 재검사
_ATTACHMENT_BLOCKED_EXTS = {".exe", ".bat", ".cmd", ".com", ".scr", ".msi",
                             ".dll", ".vbs", ".ps1", ".sh", ".js"}


@router.post("/upload-attachment")
async def upload_attachment(request: Request):
    """메일 첨부용 범용 업로드. 실행 가능한 확장자(.exe/.bat/...)는 차단.
    반환 URL 은 이미지 업로드와 동일 경로 규약 → 기존 send-mail attachment resolver 재사용."""
    me = current_user(request)
    filename, data = await _read_upload_payload(request)
    ext = Path(filename or "").suffix.lower()
    if ext in _ATTACHMENT_BLOCKED_EXTS:
        raise HTTPException(400, f"보안상 업로드 차단된 확장자: {ext}")
    if len(data) > _ATTACHMENT_MAX_BYTES:
        raise HTTPException(413, f"파일이 너무 큽니다 (최대 {_ATTACHMENT_MAX_BYTES // (1024*1024)}MB).")
    if not data:
        raise HTTPException(400, "빈 파일입니다.")

    uid = uuid.uuid4().hex[:12]
    safe = _safe_filename(filename or ("attachment" + ext))
    if ext and not safe.lower().endswith(ext):
        safe += ext
    subdir = UPLOADS_DIR / uid
    subdir.mkdir(parents=True, exist_ok=True)
    dst = subdir / safe
    dst.write_bytes(data)

    url = f"/api/informs/files/{uid}/{safe}"
    mime = mimetypes.guess_type(safe)[0] or "application/octet-stream"
    return {"ok": True, "filename": safe, "url": url, "size": len(data),
            "mime": mime, "uploaded_by": me["username"]}


@router.get("/files/{uid}/{name}")
def serve_image(request: Request, uid: str, name: str):
    """업로드 이미지 서빙. path traversal 차단."""
    # 인증은 전역 미들웨어가 처리하지만 방어적 검증.
    _ = current_user(request)
    if not re.fullmatch(r"[A-Za-z0-9]+", uid):
        raise HTTPException(400, "bad uid")
    safe = _safe_filename(name)
    dst = (UPLOADS_DIR / uid / safe).resolve()
    try:
        dst.relative_to(UPLOADS_DIR.resolve())
    except Exception:
        raise HTTPException(403, "path traversal")
    if not dst.is_file():
        raise HTTPException(404)
    mime, _ = mimetypes.guess_type(str(dst))
    return FileResponse(str(dst), media_type=mime or "application/octet-stream")


@router.get("")
def list_by_wafer(wafer_id: str = Query(..., min_length=1),
                  include_deleted: bool = Query(False)):
    include_deleted = include_deleted if isinstance(include_deleted, bool) else False
    items = [x for x in _without_deleted(_load_upgraded(), include_deleted) if _entry_wafer_matches(x, wafer_id)]
    items.sort(key=lambda x: x.get("created_at", ""))
    return {"informs": items, "module_summary": _module_progress_summary(items)}


def _sidebar_payload(items: list, me: dict, my_mods: set,
                     wafer_limit: int = 500, product_limit: int = 500, lot_limit: int = 1000) -> dict:
    role = me.get("role", "user")
    username = me.get("username", "")
    wafers_seen: dict = {}
    products_seen: dict = {}
    lots_seen: dict = {}

    for x in items:
        if not _visible_to(x, username, role, my_mods):
            continue
        ts = x.get("created_at", "")

        for w in _entry_wafer_ids(x):
            wk = _wafer_key(w)
            if not wk:
                continue
            cur = wafers_seen.get(wk)
            if cur is None:
                wafers_seen[wk] = {
                    "wafer_id": w, "wafer_key": wk, "last": ts, "count": 0,
                    "lot_id": x.get("lot_id", ""), "product": x.get("product", ""),
                }
                cur = wafers_seen[wk]
            elif ts > cur.get("last", ""):
                cur["last"] = ts
                cur["wafer_id"] = w
                if x.get("lot_id"):
                    cur["lot_id"] = x.get("lot_id")
                if x.get("product"):
                    cur["product"] = x.get("product")
            cur["count"] = cur.get("count", 0) + 1

        p = x.get("product")
        if p and isinstance(p, str):
            canon = _canonical_product(p)
            if canon:
                key = canon.lower()
                ps = products_seen.setdefault(key, {"product": canon, "count": 0, "last": ""})
                ps["count"] += 1
                if ts > ps["last"]:
                    ps["last"] = ts

        source_root = str(x.get("root_lot_id") or _root_lot_from_values(x.get("lot_id") or "", x.get("embed_table"))).strip()
        for lot_key in _inform_fab_lots(x):
            ls = lots_seen.setdefault(lot_key, {
                "lot_id": lot_key,
                "fab_lot_id": lot_key,
                "root_lot_id": lot_key,
                "source_root_lot_id": source_root,
                "count": 0,
                "last": "",
                "product": x.get("product", ""),
                "fab_lots": set(),
            })
            ls["count"] += 1
            if ts > ls["last"]:
                ls["last"] = ts
            if x.get("product"):
                ls["product"] = x.get("product")
            if source_root:
                ls["source_root_lot_id"] = source_root
            ls["fab_lots"].add(lot_key)

    for p in _fab_db_products():
        key = p.lower()
        products_seen.setdefault(key, {"product": p, "count": 0, "last": ""})

    wafers = sorted(wafers_seen.values(), key=lambda v: v.get("last", ""), reverse=True)[:wafer_limit]
    products = sorted(products_seen.values(), key=lambda v: (v["last"], v["product"]), reverse=True)[:product_limit]
    lots = []
    for s in lots_seen.values():
        s["fab_lots"] = sorted(s["fab_lots"])
        lots.append(s)
    lots.sort(key=lambda v: v["last"], reverse=True)
    return {"wafers": wafers, "products": products, "lots": lots[:lot_limit]}


@router.get("/sidebar")
def sidebar(request: Request,
            wafer_limit: int = Query(500, ge=1, le=5000),
            product_limit: int = Query(500, ge=1, le=5000),
            lot_limit: int = Query(1000, ge=1, le=10000)):
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    return _sidebar_payload(_without_deleted(_load_upgraded()), me, my_mods, wafer_limit, product_limit, lot_limit)


@router.get("/recent")
def recent_roots(request: Request, limit: int = Query(50, ge=1, le=500),
                 include_deleted: bool = Query(False)):
    include_deleted = include_deleted if isinstance(include_deleted, bool) else False
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded(), include_deleted)
    roots = [x for x in items if not x.get("parent_id")]
    roots = [x for x in roots if _visible_to(x, me["username"], me.get("role", "user"), my_mods)]
    roots = _attach_thread_stats(roots, items)
    roots.sort(key=lambda x: x.get("thread_updated_at") or x.get("created_at", ""), reverse=True)
    return {"informs": _attach_root_lot_module_counts(roots[:limit], roots)}


@router.get("/wafers")
def list_wafers(request: Request, limit: int = Query(500, ge=1, le=5000)):
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded())
    seen: dict = {}
    for x in items:
        if not _visible_to(x, me["username"], me.get("role", "user"), my_mods):
            continue
        ts = x.get("created_at", "")
        for w in _entry_wafer_ids(x):
            wk = _wafer_key(w)
            if not wk:
                continue
            cur = seen.get(wk)
            if cur is None or ts > cur.get("last", ""):
                if cur is None:
                    seen[wk] = {"wafer_id": w, "wafer_key": wk, "last": ts, "count": 0, "lot_id": x.get("lot_id", ""),
                                "product": x.get("product", "")}
                else:
                    cur["last"] = ts
                    cur["wafer_id"] = w
                    if x.get("lot_id"):
                        cur["lot_id"] = x.get("lot_id")
                    if x.get("product"):
                        cur["product"] = x.get("product")
            seen[wk]["count"] = seen[wk].get("count", 0) + 1
    arr = sorted(seen.values(), key=lambda v: v.get("last", ""), reverse=True)
    return {"wafers": arr[:limit]}


@router.get("/by-lot")
def by_lot(request: Request, lot_id: str = Query(..., min_length=1),
           include_deleted: bool = Query(False)):
    """FAB lot id 매칭. root_lot_id 는 저장 메타로만 보존한다."""
    include_deleted = include_deleted if isinstance(include_deleted, bool) else False
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded(), include_deleted)
    query = (lot_id or "").strip()

    def matches(entry: dict) -> bool:
        if not query:
            return False
        q = query.casefold()
        return any(str(v or "").strip().casefold() == q for v in _inform_fab_lots(entry))

    hits = [x for x in items if matches(x)]
    hits = [x for x in hits if _visible_to(x, me["username"], me.get("role", "user"), my_mods)]
    hits = _attach_thread_stats(hits, hits)
    hits.sort(key=lambda x: x.get("created_at", ""))
    wafers = sorted({wafer for x in hits for wafer in _entry_wafer_ids(x)}, key=_wafer_key)
    lots = sorted({lot for x in hits for lot in _inform_fab_lots(x)})
    module_counts: dict[str, int] = {}
    for x in hits:
        if x.get("parent_id"):
            continue
        mod = str(x.get("module") or "").strip() or "기타"
        module_counts[mod] = module_counts.get(mod, 0) + 1
    informed_modules = sorted(module_counts, key=lambda m: (-module_counts[m], m))
    available_modules = _lot_matrix_module_order()
    for mod in informed_modules:
        if mod not in available_modules:
            available_modules.append(mod)
    return {
        "informs": hits,
        "wafers": wafers,
        "lots": lots,
        "lot_id": query,
        "fab_lot_id": query,
        "root_lot_id": query,
        "count": len(hits),
        "module_summary": _module_progress_summary(hits),
        "informed_modules": informed_modules,
        "available_modules": available_modules,
        "module_counts": module_counts,
    }


@router.get("/by-product")
def by_product(request: Request, product: str = Query(..., min_length=1),
               limit: int = Query(500, ge=1, le=5000),
               include_deleted: bool = Query(False)):
    include_deleted = include_deleted if isinstance(include_deleted, bool) else False
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded(), include_deleted)
    want = _canonical_product(product)
    hits = [x for x in items if _canonical_product(x.get("product") or "") == want]
    hits = [x for x in hits if _visible_to(x, me["username"], me.get("role", "user"), my_mods)]
    # 루트 우선 최근순
    roots = [x for x in hits if not x.get("parent_id")]
    roots = _attach_thread_stats(roots, hits)
    roots.sort(key=lambda x: x.get("thread_updated_at") or x.get("created_at", ""), reverse=True)
    return {"informs": _attach_root_lot_module_counts(roots[:limit], roots), "count": len(roots)}


@router.get("/my")
def my_informs(request: Request, limit: int = Query(200, ge=1, le=2000)):
    """현재 유저 모듈 범위의 인폼 루트 (담당자 대시보드)."""
    me = current_user(request)
    role = me.get("role", "user")
    my_mods = user_modules(me["username"], role)
    items = _without_deleted(_load_upgraded())
    roots = [x for x in items if not x.get("parent_id")]
    if role == "admin" or "__all__" in my_mods:
        vis = roots
    else:
        vis = [
            x for x in roots
            if (x.get("module") in my_mods) or x.get("author") == me["username"]
        ]
    vis = _attach_thread_stats(vis, items)
    vis.sort(key=lambda x: x.get("thread_updated_at") or x.get("created_at", ""), reverse=True)
    return {"informs": _attach_root_lot_module_counts(vis[:limit], vis), "all_rounder": role == "admin" or "__all__" in my_mods,
            "my_modules": [] if "__all__" in my_mods else sorted(my_mods)}


def _canonical_product(s: str) -> str:
    """Back-compat wrapper around core.product_dedup.canonical_product."""
    return canonical_product(s)


@router.get("/products")
def list_products(request: Request):
    """v9.0.0: product 키를 canonical 로 병합 — ML_TABLE_PRODA + PRODA + 'PRODA ' 변형 전부 1개로.
    레코드 count 는 canonical 기준 합산.
    """
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded())
    seen: dict = {}
    for x in items:
        p = x.get("product")
        if not p or not isinstance(p, str):
            continue
        if not _visible_to(x, me["username"], me.get("role", "user"), my_mods):
            continue
        canon = _canonical_product(p)
        if not canon:
            continue
        key = canon.lower()
        s = seen.setdefault(key, {"product": canon, "count": 0, "last": ""})
        s["count"] += 1
        ts = x.get("created_at", "")
        if ts > s["last"]:
            s["last"] = ts
    for p in _fab_db_products():
        key = p.lower()
        seen.setdefault(key, {"product": p, "count": 0, "last": ""})
    arr = sorted(seen.values(), key=lambda v: (v["last"], v["product"]), reverse=True)
    return {"products": arr}


@router.get("/lots")
def list_lots(request: Request):
    """FAB_LOT_ID/LOT_ID 기준으로 그룹핑한다."""
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _without_deleted(_load_upgraded())
    seen: dict = {}
    for x in items:
        if not _visible_to(x, me["username"], me.get("role", "user"), my_mods):
            continue
        ts = x.get("created_at", "")
        source_root = str(x.get("root_lot_id") or _root_lot_from_values(x.get("lot_id") or "", x.get("embed_table"))).strip()
        for lot_key in _inform_fab_lots(x):
            s = seen.setdefault(lot_key, {
                "lot_id": lot_key,
                "fab_lot_id": lot_key,
                "root_lot_id": lot_key,
                "source_root_lot_id": source_root,
                "count": 0, "last": "",
                "product": x.get("product", ""),
                "fab_lots": set(),
            })
            s["count"] += 1
            if ts > s["last"]:
                s["last"] = ts
            if x.get("product"):
                s["product"] = x.get("product")
            if source_root:
                s["source_root_lot_id"] = source_root
            s["fab_lots"].add(lot_key)
    arr = []
    for s in seen.values():
        s["fab_lots"] = sorted(s["fab_lots"])
        arr.append(s)
    arr.sort(key=lambda v: v["last"], reverse=True)
    return {"lots": arr}


@router.post("")
def create_inform(req: InformCreate, request: Request):
    me = current_user(request)
    # wafer_id 는 생성 시점의 LOT_ID 매핑 snapshot 이다. 없다고 lot_id 로 대체하지 않는다.
    requested_wafer_id = (req.wafer_id or "").strip()
    wid = requested_wafer_id
    lot_for_fallback = (req.lot_id or "").strip()
    if not lot_for_fallback and not wid and not req.parent_id:
        raise HTTPException(400, "lot_id (또는 wafer_id) 가 필요합니다.")
    items = _load_upgraded()

    # parent 검증 + 상속 (lot_id / product).
    inherit_lot = (req.lot_id or "").strip()
    inherit_product = (req.product or "").strip()
    if req.parent_id:
        parent = _find(items, req.parent_id)
        if not parent:
            raise HTTPException(404, "parent not found")
        if not wid:
            wid = str(parent.get("wafer_id") or "").strip()
        if parent.get("wafer_id") != wid:
            raise HTTPException(400, "parent wafer mismatch")
        # 자식은 부모 lot/product 상속 (입력 없을 때)
        inherit_lot = inherit_lot or parent.get("lot_id", "")
        inherit_product = inherit_product or parent.get("product", "")

    sc = None
    if req.splittable_change and (req.splittable_change.column or req.splittable_change.new_value):
        sc = {
            "column": (req.splittable_change.column or "").strip(),
            "old_value": (req.splittable_change.old_value or "").strip(),
            "new_value": (req.splittable_change.new_value or "").strip(),
            "applied": bool(req.splittable_change.applied),
        }

    # 이미지 화이트리스트: 서버에 저장된 업로드 경로만 허용 (URL 필터링).
    imgs = []
    for im in (req.images or []):
        if not im.url or not im.url.startswith("/api/informs/files/"):
            continue
        imgs.append({
            "filename": _safe_filename(im.filename or "image"),
            "url": im.url,
            "size": max(0, int(im.size or 0)),
        })

    embed = _embed_with_attached_sets(_sanitize_embed_table(req.embed_table), req.attached_sets)

    now = _now()
    is_root = not req.parent_id
    identity_snapshot = _resolve_lot_identity_snapshot(
        inherit_product,
        inherit_lot,
        requested_wafer_id,
        embed,
        req.fab_lot_id_at_save,
    )
    root_lot = identity_snapshot.get("root_lot_id") or _root_lot_from_values(inherit_lot, None)
    wid = identity_snapshot.get("wafer_id") or wid
    fab_snapshot = (
        str(identity_snapshot.get("fab_lot_id") or "").strip()
        or (req.fab_lot_id_at_save or "").strip()
        or _resolve_fab_lot_snapshot(inherit_product, root_lot, requested_wafer_id)
        or ", ".join(_extract_fab_lots_from_embed(embed))
    )
    entry = {
        "id": _new_id(),
        "parent_id": req.parent_id or None,
        "wafer_id": wid,
        "wafer_ids_at_save": list(identity_snapshot.get("wafer_ids") or _split_saved_ids(wid)),
        "lot_id": inherit_lot,
        "root_lot_id": root_lot,
        "root_lot_ids_at_save": list(identity_snapshot.get("root_lot_ids") or ([root_lot] if root_lot else [])),
        "product": inherit_product,
        "module": (req.module or "").strip(),
        "reason": (req.reason or "").strip() or "PEMS",
        "text": (req.text or "").strip(),
        "author": me["username"],
        "created_at": now,
        "checked": False,
        "checked_by": "",
        "checked_at": "",
        "flow_status": "registered" if is_root else "",
        "status_history": (
            [{"status": "registered", "actor": me["username"], "at": now, "note": "created"}]
            if is_root else []
        ),
        "splittable_change": sc,
        "images": imgs,
        "embed_table": embed,
        "attachments": _attachment_view({"embed_table": embed}),
        "auto_generated": False,
        # v8.7.9: deadline 필드 폐기 — 저장하지 않음.
        "group_ids": [str(g).strip() for g in (req.group_ids or []) if g and str(g).strip()],
        # v8.8.15: fab_lot_id 스냅샷 — FE 가 SplitTable 맥락에서 resolve 해서 보내준 값.
        #   없으면 BE 에서 resolve 시도 (product + lot_id 기준으로 ML_TABLE 의 최신 fab_lot_id 조회).
        "fab_lot_id_at_save": fab_snapshot,
        "lot_identity_snapshot": {
            "lot_id": inherit_lot,
            "fab_lot_id": fab_snapshot,
            "root_lot_id": root_lot,
            "root_lot_ids": list(identity_snapshot.get("root_lot_ids") or ([root_lot] if root_lot else [])),
            "wafer_id": wid,
            "wafer_ids": list(identity_snapshot.get("wafer_ids") or _split_saved_ids(wid)),
            "source": identity_snapshot.get("source") or "",
            "created_at": now,
        },
    }
    items.append(entry)
    _save(items)
    _audit_record(
        request,
        "comment" if req.parent_id else "create",
        entry,
        {"parent_id": req.parent_id or "", "root": bool(is_root)},
        f"{'댓글' if req.parent_id else '생성'} · {entry['module'] or '-'} · {inherit_lot or wid}",
        at=now,
    )
    return {"ok": True, "inform": entry}


# ── Auto-log helper (다른 라우터가 import) ─────────────────────────────
def auto_log_splittable_change(author: str, product: str, lot_id: str,
                               cell_key: str, old_value, new_value, action: str = "set",
                               fab_lot_id: str = "") -> None:
    """SplitTable plan 변경이 일어나면 해당 lot 에 자동 인폼 루트를 남긴다.

    - wafer_id 가 없으면 lot_id 를 placeholder 로 사용 (스레드는 lot 뷰에서 묶여 보임).
    - module 은 cell_key prefix 로 추정 (KNOB/MASK/FAB → 기타). 추후 룰 확장.
    - auto_generated=True 로 표시 → FE 에서 시스템 발행 카드로 렌더.
    """
    try:
        items = _load_upgraded()
        parts = str(cell_key or "").split("|")
        cell_root = parts[0].strip() if len(parts) >= 3 else ""
        cell_wafer = parts[1].strip() if len(parts) >= 3 else ""
        root_lot = _root_lot_from_values(lot_id or cell_root)
        col = parts[-1] if len(parts) >= 3 else cell_key
        upper = col.upper()
        if upper.startswith("MASK_"):
            mod = "MASK"
        elif upper.startswith("FAB_"):
            mod = "FAB"
        elif upper.startswith("KNOB_") or "_" not in upper:
            mod = "KNOB"
        else:
            mod = ""
        now = _now()
        text = f"[SplitTable 자동기록] {action} · {col} · {old_value!r} → {new_value!r}"
        fab_snapshot = (fab_lot_id or "").strip() or _resolve_fab_lot_snapshot(product, root_lot, cell_wafer)
        embed = None
        try:
            if root_lot and col:
                embed = build_splittable_embed(
                    product=product,
                    lot_id=root_lot,
                    custom_cols=[col],
                    is_fab_lot=False,
                )
        except Exception:
            embed = None
        entry = {
            "id": _new_id(),
            "parent_id": None,
            "wafer_id": cell_wafer or root_lot or product or "auto",
            "lot_id": root_lot or "",
            "root_lot_id": root_lot or "",
            "product": product or "",
            "module": mod,
            "reason": "레시피 변경",
            "text": text,
            "author": author or "system",
            "created_at": now,
            "checked": False, "checked_by": "", "checked_at": "",
            "flow_status": "registered",
            "status_history": [{"status": "registered", "actor": author or "system",
                                "at": now, "note": "auto from SplitTable"}],
            "splittable_change": {
                "column": col,
                "old_value": ("" if old_value is None else str(old_value)),
                "new_value": ("" if new_value is None else str(new_value)),
                "applied": (action == "set"),
            },
            "images": [],
            "embed_table": embed,
            "auto_generated": True,
            "fab_lot_id_at_save": fab_snapshot,
        }
        items.append(entry)
        _save(items)
        _audit_record(author or "system", "create", entry,
                      {"auto_generated": True, "cell_key": cell_key, "action": action},
                      f"SplitTable 자동기록 · {col}", at=now)
    except Exception:
        # 자동기록 실패로 인해 plan 저장까지 실패시키면 안 됨.
        pass


@router.post("/delete")
def delete_inform(request: Request, id: str = Query(...)):
    return _delete_inform_by_id(id, request)


@router.delete("/{inform_id}")
def delete_inform_rest(inform_id: str, request: Request):
    return _delete_inform_by_id(inform_id, request)


def _delete_inform_by_id(id: str, request: Request):
    """Soft-delete an inform. Only the original author or admin can delete."""
    me = current_user(request)
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        return {"ok": True, "noop": True, "inform": target}
    role = me.get("role", "user")
    allowed = target.get("author") == me["username"] or role == "admin"
    if not allowed:
        raise HTTPException(403, "삭제 권한이 없습니다 (작성자/admin).")
    now = _now()
    target["deleted"] = True
    target["deleted_at"] = now
    target["deleted_by"] = me["username"]
    target["updated_at"] = now
    _save(items)
    _audit_record(request, "delete", target, {"deleted_at": now}, f"삭제 · id={id}", at=now)
    return {"ok": True, "inform": target}


class InformEditReq(BaseModel):
    text: Optional[str] = None
    module: Optional[str] = None
    reason: Optional[str] = None
    embed_table: Optional[EmbedTable] = None
    # wafer_id / lot_id / product 는 변경 불가 (스레드/매칭 깨짐 방지).


@router.post("/edit")
def edit_inform(req: InformEditReq, request: Request, id: str = Query(...)):
    return _edit_inform_by_id(id, req, request)


@router.post("/{inform_id}/edit")
@router.put("/{inform_id}/edit")
def edit_inform_rest(inform_id: str, req: InformEditReq, request: Request):
    return _edit_inform_by_id(inform_id, req, request)


def _edit_inform_by_id(id: str, req: InformEditReq, request: Request):
    """v8.8.12: 등록된 인폼의 본문/모듈/사유 수정.
    작성자 본인 또는 admin 은 수정 가능. 다른 사용자는 답글로 보완한다."""
    me = current_user(request)
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    if me.get("role") != "admin" and target.get("author") != me["username"]:
        raise HTTPException(403, "작성자 또는 admin만 수정할 수 있습니다. 내용 보완은 답글로 추가하세요.")
    now = _now()
    hist = target.get("edit_history") or []
    changed = []
    if req.text is not None and (req.text or "").strip() != (target.get("text") or ""):
        before = target.get("text") or ""
        target["text"] = (req.text or "").strip()
        hist.append({"at": now, "actor": me["username"], "field": "text",
                     "before": before[:400], "after": target["text"][:400],
                     "kind": "edit"})
        changed.append("text")
    if req.module is not None and (req.module or "").strip() != (target.get("module") or ""):
        before = target.get("module") or ""
        target["module"] = (req.module or "").strip()
        hist.append({"at": now, "actor": me["username"], "field": "module",
                     "before": before, "after": target["module"], "kind": "edit"})
        changed.append("module")
    if req.reason is not None and (req.reason or "").strip() != (target.get("reason") or ""):
        before = target.get("reason") or ""
        target["reason"] = (req.reason or "").strip()
        hist.append({"at": now, "actor": me["username"], "field": "reason",
                     "before": before, "after": target["reason"], "kind": "edit"})
        changed.append("reason")
    fields_set = getattr(req, "model_fields_set", None)
    if fields_set is None:
        fields_set = getattr(req, "__fields_set__", set())
    if "embed_table" in fields_set:
        before_count = len((target.get("embed_table") or {}).get("attached_sets") or []) if isinstance(target.get("embed_table"), dict) else 0
        target["embed_table"] = _sanitize_embed_table(req.embed_table)
        target["attachments"] = _attachment_view(target)
        after_count = len((target.get("embed_table") or {}).get("attached_sets") or []) if isinstance(target.get("embed_table"), dict) else 0
        hist.append({"at": now, "actor": me["username"], "field": "embed_table",
                     "before": f"{before_count} attached sets", "after": f"{after_count} attached sets",
                     "kind": "edit"})
        changed.append("embed_table")
    if not changed:
        return {"ok": True, "noop": True}
    target["edit_history"] = hist[-200:]
    target["updated_at"] = now
    _save(items)
    diff = [
        h for h in hist[-len(changed):]
        if isinstance(h, dict) and h.get("field") in changed
    ]
    _audit_record(request, "edit", target, {"fields": changed, "diff": diff},
                  f"수정 · {','.join(changed)}", at=now)
    return {"ok": True, "changed": changed, "inform": target}


@router.post("/check")
def check_inform(req: CheckReq, request: Request, id: str = Query(...)):
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    if not _can_moderate(target, me["username"], me.get("role", "user"), my_mods):
        raise HTTPException(403, "모듈 담당자만 체크할 수 있습니다.")
    target["checked"] = bool(req.checked)
    target["checked_by"] = me["username"] if req.checked else ""
    target["checked_at"] = _now() if req.checked else ""
    _save(items)
    _audit_record(request, "status_change", target, {"checked": target["checked"]},
                  f"체크 {'완료' if target['checked'] else '해제'} · id={id}",
                  at=target["checked_at"] or _now())
    return {"ok": True, "inform": target}


@router.post("/status")
def set_status(req: StatusReq, request: Request, id: str = Query(...)):
    raw_status = (req.status or "").strip()
    if raw_status not in FLOW_STATUSES_LEGACY:
        raise HTTPException(400, f"invalid status; must be one of {FLOW_STATUSES_LEGACY}")
    st = _canonical_flow_status(raw_status)
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    if target.get("parent_id"):
        raise HTTPException(400, "status 는 루트 인폼에만 적용됩니다.")
    if not _can_moderate(target, me["username"], me.get("role", "user"), my_mods):
        raise HTTPException(403, "모듈 담당자만 상태를 변경할 수 있습니다.")
    prev_status = _canonical_flow_status(target.get("flow_status"), target)
    if raw_status == "registered" and prev_status == "apply_confirmed" and (target.get("mail_history") or []):
        st = "mail_completed"
    if prev_status == st:
        return {"ok": True, "inform": target}
    target["flow_status"] = st
    hist = target.get("status_history") or []
    note = (req.note or "").strip()
    if prev_status == "apply_confirmed" and st != "apply_confirmed" and not note:
        note = "확인 취소"
    hist.append({"status": st, "prev": prev_status, "actor": me["username"],
                 "at": _now(), "note": note})
    target["status_history"] = hist
    _save(items)
    _audit_record(request, "status_change", target,
                  {"prev": prev_status, "status": st, "note": note},
                  f"상태변경 · {prev_status or '-'} → {st}", at=hist[-1]["at"])
    return {"ok": True, "inform": target}


@router.post("/deadline")
def set_deadline(req: DeadlineReq, request: Request, id: str = Query(...)):
    """루트 인폼의 마감일을 설정/해제. 작성자/모듈 담당자/admin 만 가능."""
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    if target.get("parent_id"):
        raise HTTPException(400, "deadline 은 루트 인폼에만 설정 가능합니다.")
    if not _can_moderate(target, me["username"], me.get("role", "user"), my_mods):
        raise HTTPException(403, "작성자/모듈 담당자/관리자만 변경 가능합니다.")
    dl = _validate_deadline(req.deadline)
    target["deadline"] = dl
    _save(items)
    _audit_record(request, "edit", target, {"field": "deadline", "deadline": dl},
                  f"마감일 변경 · {dl or '해제'}")
    return {"ok": True, "inform": target}


# ── v8.8.0: 제품별 담당자 (product contacts) ───────────────────────
# 좌측 사이드바 + 메일 본문 자동 삽입용. 모든 로그인 유저가 CRUD 가능.
PRODUCT_CONTACTS_FILE = INFORMS_DIR / "product_contacts.json"


def _load_product_contacts() -> dict:
    data = load_json(PRODUCT_CONTACTS_FILE, {"products": {}})
    if isinstance(data, dict) and isinstance(data.get("products"), dict):
        # v8.8.33: trim + case-insensitive dedup 을 로드 타임에 적용 —
        #   유령 중복 키 ("PRODA", "PRODA ", "proda") 를 하나로 병합.
        prods = data.get("products") or {}
        seen: dict = {}
        for raw_key, contacts in prods.items():
            if not isinstance(raw_key, str):
                continue
            trimmed = raw_key.strip()
            if not trimmed:
                continue
            canon = None
            for existing in seen:
                if existing.lower() == trimmed.lower():
                    canon = existing
                    break
            if canon is None:
                seen[trimmed] = list(contacts or [])
            else:
                # merge — 기존 list 에 병합, id 기준 dedup.
                existing_ids = {c.get("id") for c in seen[canon] if isinstance(c, dict)}
                for c in (contacts or []):
                    if isinstance(c, dict) and c.get("id") not in existing_ids:
                        seen[canon].append(c)
        return {"products": seen}
    return {"products": {}}


def _save_product_contacts(data: dict) -> None:
    save_json(PRODUCT_CONTACTS_FILE, data)


def _new_contact_id() -> str:
    import secrets as _secrets
    return "pc_" + _secrets.token_hex(5)


class ProductContactReq(BaseModel):
    product: str
    name: str
    role: str = ""           # 직책/역할 (예: "PIE", "측정")
    email: str = ""
    phone: str = ""
    note: str = ""


@router.get("/product-contacts")
def list_product_contacts(product: str = Query("")):
    data = _load_product_contacts()
    products = data.get("products") or {}
    if product:
        return {"product": product, "contacts": products.get(product, [])}
    return {"products": products}


@router.post("/product-contacts")
def add_product_contact(req: ProductContactReq, request: Request):
    me = current_user(request)
    prod = (req.product or "").strip()
    name = (req.name or "").strip()
    if not prod or not name:
        raise HTTPException(400, "product/name required")
    data = _load_product_contacts()
    products = data.setdefault("products", {})
    contact = {
        "id": _new_contact_id(),
        "name": name,
        "role": (req.role or "").strip(),
        "email": (req.email or "").strip(),
        "phone": (req.phone or "").strip(),
        "note": (req.note or "").strip(),
        "added_by": me.get("username", ""),
        "added_at": datetime.datetime.now().isoformat(timespec="seconds"),
    }
    products.setdefault(prod, []).append(contact)
    _save_product_contacts(data)
    _audit(request, "inform:product-contact-add", detail=f"product={prod} name={name}", tab="inform")
    return {"ok": True, "contact": contact}


@router.post("/product-contacts/update")
def update_product_contact(req: ProductContactReq, request: Request, id: str = Query(...)):
    me = current_user(request)
    prod = (req.product or "").strip()
    if not prod:
        raise HTTPException(400, "product required")
    data = _load_product_contacts()
    arr = data.get("products", {}).get(prod) or []
    target = next((c for c in arr if c.get("id") == id), None)
    if not target:
        raise HTTPException(404, "contact not found")
    target["name"] = (req.name or target.get("name", "")).strip()
    target["role"] = (req.role or "").strip()
    target["email"] = (req.email or "").strip()
    target["phone"] = (req.phone or "").strip()
    target["note"] = (req.note or "").strip()
    target["updated_by"] = me.get("username", "")
    target["updated_at"] = datetime.datetime.now().isoformat(timespec="seconds")
    _save_product_contacts(data)
    return {"ok": True, "contact": target}


@router.post("/product-contacts/delete")
def delete_product_contact(request: Request, id: str = Query(...), product: str = Query(...)):
    _ = current_user(request)
    data = _load_product_contacts()
    arr = data.get("products", {}).get(product) or []
    new_arr = [c for c in arr if c.get("id") != id]
    if len(new_arr) == len(arr):
        raise HTTPException(404, "contact not found")
    data["products"][product] = new_arr
    _save_product_contacts(data)
    _audit(request, "inform:product-contact-del", detail=f"product={product} id={id}", tab="inform")
    return {"ok": True}


# v8.8.19: 인폼 제품 담당자 전용 필터 — admin 역할 / `admin`·`hol`·`test` 포함
#   username 을 제외한다. 그룹(`_is_blocked_member`) 는 v8.8.5 에서 admin 을 허용하도록
#   완화됐지만, 인폼 제품 담당자는 "메일 수신 대상 + 시스템 계정 배제" 요건이 강해
#   별도 규칙을 쓴다. eligible-users 엔드포인트와 bulk-add 양쪽에서 동일 기준 적용.
_BLOCKED_UNAME_TOKENS = ("admin", "hol", "test")


def _is_blocked_contact(username: str, user_obj: dict | None = None) -> bool:
    un = (username or "").strip()
    if not un:
        return True
    low = un.lower()
    for tok in _BLOCKED_UNAME_TOKENS:
        if tok in low:
            return True
    if isinstance(user_obj, dict):
        role = (user_obj.get("role") or "").strip().lower()
        if role == "admin":
            return True
    return False


@router.get("/eligible-contacts")
def eligible_contacts(request: Request):
    """v8.8.19: 인폼 제품 담당자로 등록 가능한 유저 목록.
    admin 역할 / admin·hol·test 포함 username 은 제외."""
    _me = current_user(request)
    try:
        from routers.auth import read_users
        users = read_users()
    except Exception:
        users = []
    out = []
    for u in users:
        un = (u.get("username") or "").strip() if isinstance(u, dict) else ""
        if not un:
            continue
        if _is_blocked_contact(un, u):
            continue
        out.append({
            "username": un,
            "email": (u.get("email") or "").strip() if isinstance(u, dict) else "",
            "role": (u.get("role") or "user") if isinstance(u, dict) else "user",
            # v8.8.27: 이름(실명) 라벨. FE picker 가 `{name} ({username})` 로 표시.
            "name": ((u.get("name") or "").strip()) if isinstance(u, dict) else "",
        })
    out.sort(key=lambda x: ((x.get("name") or "").lower(), x["username"].lower()))
    return {"users": out}


# v8.8.2: bulk add — 개별 유저 / 그룹 멤버 혼합 추가.
class ProductContactBulkReq(BaseModel):
    product: str
    usernames: List[str] = []         # 개별 유저 선택 결과
    group_ids: List[str] = []         # 선택한 그룹(들) — 멤버 전체 풀
    role: str = ""                    # 일괄 적용할 역할 (선택)


@router.post("/product-contacts/bulk-add")
def bulk_add_product_contacts(req: ProductContactBulkReq, request: Request):
    """유저 / 그룹 혼합 일괄 추가.

    - usernames 에 적힌 각 유저를 contacts 로 등록.
    - group_ids 의 모든 그룹 members 도 pool 에 합류.
    - admin/test 계정은 서버측에서 한 번 더 필터.
    - 이미 같은 product 에 동일 username(혹은 email) 이 등록돼 있으면 dedup.
    """
    me = current_user(request)
    prod = (req.product or "").strip()
    if not prod:
        raise HTTPException(400, "product required")
    from routers.groups import _load_users_by_name, _load as _load_groups
    users_by_name = _load_users_by_name()
    # pool: 유니크 username 모음
    pool: List[str] = []
    for un in (req.usernames or []):
        un = (un or "").strip()
        if un and un not in pool:
            pool.append(un)
    if req.group_ids:
        gids = set(req.group_ids)
        for g in _load_groups():
            if g.get("id") in gids:
                for m in (g.get("members") or []):
                    if m and m not in pool:
                        pool.append(m)
    # 필터 + 유저 프로필 resolve
    data = _load_product_contacts()
    products = data.setdefault("products", {})
    existing = products.setdefault(prod, [])
    existing_keys = set()
    for c in existing:
        uname = (c.get("source_username") or "").strip().lower()
        email = (c.get("email") or "").strip().lower()
        if uname:
            existing_keys.add(("u", uname))
        if email:
            existing_keys.add(("e", email))
    added: List[dict] = []
    skipped: List[str] = []
    for un in pool:
        u = users_by_name.get(un) or {}
        # v8.8.19: 인폼 담당자는 admin/hol/test 완전 차단 (_is_blocked_contact).
        if _is_blocked_contact(un, u):
            skipped.append(un)
            continue
        email = (u.get("email") or "").strip() if isinstance(u, dict) else ""
        name = (u.get("display_name") or u.get("name") or un) if isinstance(u, dict) else un
        key_u = ("u", un.lower())
        key_e = ("e", email.lower()) if email else None
        if key_u in existing_keys or (key_e and key_e in existing_keys):
            skipped.append(un)
            continue
        contact = {
            "id": _new_contact_id(),
            "name": name,
            "role": (req.role or (u.get("role", "") if isinstance(u, dict) else "") or "").strip(),
            "email": email,
            "phone": (u.get("phone", "") if isinstance(u, dict) else "").strip(),
            "note": "",
            "source_username": un,
            "added_by": me.get("username", ""),
            "added_at": datetime.datetime.now().isoformat(timespec="seconds"),
        }
        existing.append(contact)
        added.append(contact)
        existing_keys.add(key_u)
        if key_e:
            existing_keys.add(key_e)
    _save_product_contacts(data)
    _audit(request, "inform:product-contact-bulk",
           detail=f"product={prod} added={len(added)} skipped={len(skipped)}", tab="inform")
    return {"ok": True, "added": added, "skipped": skipped, "total": len(existing)}


# ── v8.7.2: Mail relay ─────────────────────────────────────────────
def _load_mail_cfg() -> dict:
    data = load_json(ADMIN_SETTINGS_FILE, {})
    if not isinstance(data, dict):
        return {}
    m = data.get("mail") or {}
    return m if isinstance(m, dict) else {}


class SendMailReq(BaseModel):
    to: List[str] = []              # resolved email addresses (fallback)
    to_users: List[str] = []        # usernames — also resolved to emails via users.csv
    groups: List[str] = []          # recipient group names (resolved via admin settings)
    subject: str = ""               # → title
    body: str = ""                  # optional extra prose prepended to HTML body
    include_thread: bool = True     # include full thread HTML in content
    status_code: str = ""           # per-send override; else admin default
    attachments: List[str] = []     # inform image URLs to attach


def _resolve_mail_recipients(
    target: dict,
    *,
    to: Optional[List[str]] = None,
    to_users: Optional[List[str]] = None,
    groups: Optional[List[str]] = None,
    cfg: Optional[dict] = None,
) -> tuple[list[str], list[dict], bool]:
    """Resolve explicit recipients, or module recipients when no explicit target is given."""
    cfg = cfg or {}
    explicit_to = [str(x or "").strip() for x in (to or []) if str(x or "").strip()]
    explicit_users = [str(x or "").strip() for x in (to_users or []) if str(x or "").strip()]
    explicit_groups = [str(x or "").strip() for x in (groups or []) if str(x or "").strip()]
    use_auto_module = not explicit_to and not explicit_users and not explicit_groups and bool(str(target.get("module") or "").strip())

    to_addrs: List[str] = []
    seen_addrs: set = set()

    def _push(em: str):
        em = (em or "").strip()
        if not em or "@" not in em:
            return
        key = em.lower()
        if key in seen_addrs:
            return
        seen_addrs.add(key)
        to_addrs.append(em)

    auto_rows: list[dict] = []
    if use_auto_module:
        auto_rows = _module_recipient_rows(str(target.get("module") or ""))
        for row in auto_rows:
            _push(row.get("email", ""))
        return to_addrs, auto_rows, True

    for a in explicit_to:
        _push(a)
    for em in _resolve_users_to_emails(explicit_users):
        _push(em)
    rg_cfg = (cfg or {}).get("recipient_groups") or {}
    for gname in explicit_groups:
        members = rg_cfg.get(gname) if isinstance(rg_cfg, dict) else None
        if isinstance(members, list):
            for em in members:
                _push(str(em))
    return to_addrs, auto_rows, False


@router.get("/mail-groups")
def list_mail_groups(request: Request):
    """v8.8.23: 단일 진실원 = groups.json (Admin 그룹).
    이름 → [이메일] 리스트 포맷으로 resolve (members 의 사내 이메일 + extra_emails).
    legacy 엔드포인트 — recipient_groups 와 동일 응답 모양을 유지해 FE 호환성 확보.
    """
    _ = current_user(request)
    try:
        from routers.groups import _load as _groups_load
        groups = _groups_load()
    except Exception:
        groups = []
    # username → email 매핑 (users.csv + admin 도메인 자동합성).
    try:
        from routers.auth import read_users
        users = read_users()
    except Exception:
        users = []
    try:
        from core.mail import load_mail_cfg as _load_mcfg
        _domain = (_load_mcfg().get("domain") or "").strip().lstrip("@")
    except Exception:
        _domain = ""
    un2em = {}
    for u in users or []:
        un = (u.get("username") or "").strip()
        if not un:
            continue
        em = (u.get("email") or "").strip()
        if em and "@" in em:
            un2em[un] = em
        elif "@" in un and "." in un.split("@", 1)[1]:
            un2em[un] = un
        elif _domain:
            un2em[un] = f"{un}@{_domain}"
    out: dict = {}
    for g in groups or []:
        if not isinstance(g, dict):
            continue
        name = (g.get("name") or "").strip()
        if not name:
            continue
        emails = []
        seen = set()
        for un in (g.get("members") or []):
            em = un2em.get(un, "")
            if em and em not in seen:
                seen.add(em); emails.append(em)
        for em in (g.get("extra_emails") or []):
            em = str(em).strip()
            if em and "@" in em and em not in seen:
                seen.add(em); emails.append(em)
        out[name] = emails
    return {"groups": out}


@router.get("/recipients")
def list_recipients(request: Request):
    """모든 승인 유저 + email. 인폼 메일 수신자 선택용 (로그인 유저 누구나 조회).

    v8.8.17: username 이 곧 사내 email 인 케이스를 effective_email 로 보조 노출.
    v8.8.22: admin 메일 도메인(admin_settings.mail.domain) 이 설정돼 있으면
      `<username>@<domain>` 으로 자동 합성. users.csv 에 email 이 안 적혀 있어도
      개별 유저가 picker 에 정상 노출되어 "리스트가 비어 있음" 문제 해결.
    """
    _ = current_user(request)  # enforce login
    from routers.auth import read_users
    # v8.8.22: admin 메일 도메인 가져와 CI 합성.
    try:
        from core.mail import load_mail_cfg as _load_mcfg
        _domain = (_load_mcfg().get("domain") or "").strip().lstrip("@")
    except Exception:
        _domain = ""
    out = []
    for u in read_users():
        if u.get("status") != "approved":
            continue
        un = u.get("username", "") or ""
        em = (u.get("email", "") or "").strip()
        if em and "@" in em:
            eff = em
        elif "@" in un and "." in un.split("@", 1)[1]:
            eff = un
        elif _domain and un:
            eff = f"{un}@{_domain}"
        else:
            eff = ""
        # v8.8.21: 인폼 메일 수신자 picker 에서는 admin/hol/test 계정 + 이메일 해결 불가 계정
        #   (_is_blocked_contact 기준) 은 아예 제외. FE 가 "(no email)" 을 표시할 상황이 없어진다.
        if _is_blocked_contact(un, u):
            continue
        if not eff:
            continue
        out.append({
            "username": un,
            "email": em,
            "effective_email": eff,
            "role": u.get("role", ""),
            # v8.8.27: 이름(실명) 라벨. FE 가 `{name} ({username})` 로 표시.
            "name": (u.get("name") or "").strip(),
        })
    # v8.8.27: 이름(있으면) → username 순서로 정렬해서 picker 에서 한글 이름 기준 탐색 쉽게.
    out.sort(key=lambda x: ((x.get("name") or "").lower(), (x.get("username") or "").lower()))
    return {"recipients": out}


def _thread_text(items: list, root_id: str) -> str:
    """작성 시각 순으로 root+children 본문을 평탄화 (plain text fallback)."""
    root = next((x for x in items if x.get("id") == root_id), None)
    if not root:
        return ""
    lines: List[str] = []

    def dump(node: dict, depth: int):
        prefix = "  " * depth
        ts = (node.get("created_at") or "")[:16].replace("T", " ")
        lines.append(f"{prefix}[{ts}] {node.get('author','?')} · {node.get('module','')}")
        body = (node.get("text") or "").strip()
        for ln in body.splitlines() or [""]:
            lines.append(f"{prefix}  {ln}")
        kids = sorted(
            [x for x in items if x.get("parent_id") == node.get("id")],
            key=lambda x: x.get("created_at", ""),
        )
        for k in kids:
            dump(k, depth + 1)

    dump(root, 0)
    return "\n".join(lines)


def _thread_html(items: list, root_id: str) -> str:
    """Render the root + its children as a nested HTML block."""
    root = next((x for x in items if x.get("id") == root_id), None)
    if not root:
        return ""

    def esc(s):
        return _html.escape(str(s or ""))

    parts: List[str] = []

    def render(node: dict, depth: int):
        bg = "#fff" if depth == 0 else "#fafafa"
        border = "#f97316" if depth == 0 else "#d1d5db"
        left_pad = 14 + depth * 14
        ts = (node.get("created_at") or "")[:16].replace("T", " ")
        author = esc(node.get("author", "?"))
        module = esc(node.get("module", ""))
        status = esc(node.get("flow_status", ""))
        body_lines = (node.get("text") or "").splitlines()
        body_html = "<br/>".join(esc(ln) for ln in body_lines) or "<i style='color:#999'>(본문 없음)</i>"
        sc = node.get("splittable_change") or None
        sc_block = ""
        if sc and (sc.get("column") or sc.get("new_value")):
            sc_block = (
                f"<div style='margin-top:6px;padding:6px 8px;background:#fff7ed;border-left:3px solid #f97316;font-family:monospace;font-size:{_MAIL_MIN_FONT};'>"
                f"▸ <b>{esc(sc.get('column',''))}</b>: "
                f"<span style='color:#6b7280;text-decoration:line-through'>{esc(sc.get('old_value','-'))}</span>"
                f" → <span style='color:#16a34a;font-weight:700'>{esc(sc.get('new_value','-'))}</span>"
                "</div>"
            )
        parts.append(
            f"<div style='margin-left:{left_pad}px;margin-bottom:8px;padding:10px 12px;"
            f"background:{bg};border:1px solid {border};border-left:4px solid {border};"
            f"border-radius:6px;font-family:-apple-system,Segoe UI,Arial,sans-serif;font-size:{_MAIL_MIN_FONT};color:#1f2937;'>"
            f"<div style='font-size:{_MAIL_MIN_FONT};color:#6b7280;margin-bottom:4px;'>"
            f"<b style='color:#1f2937'>{author}</b> · {esc(ts)} · "
            f"<span style='color:#f97316'>{module}</span>"
            + (f" · <span style='padding:1px 6px;border-radius:10px;background:#e0f2fe;color:#0369a1;font-size:{_MAIL_MIN_FONT};'>{status}</span>" if status else "")
            + f"</div>"
            f"<div style='line-height:1.55'>{body_html}</div>"
            f"{sc_block}"
            f"</div>"
        )
        kids = sorted(
            [x for x in items if x.get("parent_id") == node.get("id")],
            key=lambda x: x.get("created_at", ""),
        )
        for k in kids:
            render(k, depth + 1)

    render(root, 0)
    return "\n".join(parts)


# v8.8.30: SplitTable HTML 인라인 렌더 시 FE 와 동일한 팔레트로 컬러링.
#   기준: My_Inform.jsx stCellBg — KNOB_* / MASK_* prefix 만 컬러 적용.
#   FE 팔레트 ST_CELL_COLORS 를 BE 에 동일 복제. HTML 메일 클라이언트 호환을 위해 inline style.
_ST_COLOR_PREFIXES = ("KNOB", "MASK")
_ST_CELL_COLORS = [
    {"bg": "#C6EFCE", "fg": "#006100"},
    {"bg": "#FFEB9C", "fg": "#9C5700"},
    {"bg": "#FBE5D6", "fg": "#BF4E00"},
    {"bg": "#BDD7EE", "fg": "#1F4E79"},
    {"bg": "#E2BFEE", "fg": "#7030A0"},
    {"bg": "#B4DED4", "fg": "#0B5345"},
    {"bg": "#F4CCCC", "fg": "#75194C"},
]
_GO_FLOW_URL = "http://go/flow"
_MAIL_MIN_FONT = "11px"


def _mail_fit_col_styles(data_columns: int) -> tuple[str, str]:
    """Column widths for mail tables that must fit without horizontal scrolling."""
    n = max(1, int(data_columns or 0))
    first_pct = 42.0 if n == 1 else (34.0 if n <= 3 else (28.0 if n <= 6 else (24.0 if n <= 10 else 22.0)))
    data_pct = (100.0 - first_pct) / n
    break_style = "white-space:normal;word-break:break-word;overflow-wrap:anywhere;vertical-align:top;box-sizing:border-box;"
    return (
        f"width:{first_pct:.2f}%;max-width:{first_pct:.2f}%;{break_style}",
        f"width:{data_pct:.2f}%;max-width:{data_pct:.2f}%;{break_style}",
    )


def _mail_table_style(data_columns: int) -> str:
    width_style = "width:auto;min-width:360px;" if max(1, int(data_columns or 0)) <= 3 else "width:100%;"
    return (
        f"border-collapse:collapse;font-size:{_MAIL_MIN_FONT};{width_style}max-width:100%;"
        "table-layout:fixed;mso-table-lspace:0pt;mso-table-rspace:0pt;"
    )


def _mail_data_col_style(data_col_style: str, span: int = 1) -> str:
    m = re.search(r"width:([0-9.]+)(%|px)", data_col_style or "")
    if not m:
        return data_col_style
    width = float(m.group(1)) * max(1, int(span or 1))
    unit = m.group(2)
    return re.sub(r"width:[^;]+;max-width:[^;]+;", f"width:{width:.2f}{unit};max-width:{width:.2f}{unit};", data_col_style, count=1)


def _mail_scroll_col_styles(data_columns: int) -> tuple[str, str]:
    """Fixed pixel columns for horizontally scrollable SplitTable mail snapshots."""
    n = max(1, int(data_columns or 0))
    if n <= 30:
        first_pct = 18.0 if n >= 16 else 22.0
        data_pct = (100.0 - first_pct) / n
        break_style = "white-space:normal;word-break:break-word;overflow-wrap:anywhere;vertical-align:top;box-sizing:border-box;"
        return (
            f"width:{first_pct:.2f}%;max-width:{first_pct:.2f}%;{break_style}",
            f"width:{data_pct:.2f}%;max-width:{data_pct:.2f}%;{break_style}",
        )
    first_px = 180
    data_px = 72 if n >= 16 else 86
    break_style = "white-space:normal;word-break:break-word;overflow-wrap:anywhere;vertical-align:top;box-sizing:border-box;"
    return (
        f"width:{first_px}px;max-width:{first_px}px;{break_style}",
        f"width:{data_px}px;max-width:{data_px}px;{break_style}",
    )


def _mail_scroll_table_style(data_columns: int) -> str:
    n = max(1, int(data_columns or 0))
    if n <= 30:
        return (
            f"border-collapse:collapse;font-size:{_MAIL_MIN_FONT};width:100%;max-width:100%;"
            "table-layout:fixed;mso-table-lspace:0pt;mso-table-rspace:0pt;"
        )
    data_px = 72 if n >= 16 else 86
    total_px = 180 + data_px * n
    return (
        f"border-collapse:collapse;font-size:{_MAIL_MIN_FONT};width:{total_px}px;"
        "min-width:100%;max-width:none;table-layout:fixed;"
        "mso-table-lspace:0pt;mso-table-rspace:0pt;"
    )


def _mail_colgroup_html(first_col_style: str, data_col_style: str, data_columns: int) -> str:
    """Make fixed-layout mail tables keep every wafer column the same width."""
    def _width(style: str) -> str:
        m = re.search(r"width:([0-9.]+)(%|px)", style or "")
        return f"{float(m.group(1)):.2f}{m.group(2)}" if m else ""

    first_w = _width(first_col_style)
    data_w = _width(data_col_style)
    if not first_w or not data_w:
        return ""
    cols = [f"<col style='width:{first_w};max-width:{first_w};'/>"]
    cols.extend(f"<col style='width:{data_w};max-width:{data_w};'/>" for _ in range(max(0, int(data_columns or 0))))
    return "<colgroup>" + "".join(cols) + "</colgroup>"


def _st_cell_bg(val: str, uniq_map: dict, pname: str) -> str:
    """FE stCellBg 와 동일 규약. value → color index (prefix 별). CSS inline style string 반환."""
    if val is None or val == "" or str(val) in ("None", "null"):
        return ""
    pn = (pname or "").upper()
    if not any(pn.startswith(p + "_") for p in _ST_COLOR_PREFIXES):
        return ""
    s = str(val)
    m = uniq_map.get(pn) or {}
    idx = m.get(s)
    if idx is None:
        return ""
    c = _ST_CELL_COLORS[idx % len(_ST_CELL_COLORS)]
    return f"background:{c['bg']};color:{c['fg']};"


def _st_build_uniq_map(rows_st: list, headers: list) -> dict:
    """각 KNOB/MASK parameter 행에 대해 value → 등장 순서 인덱스를 기록. stCellBg 조회용."""
    out: dict = {}
    for r in rows_st or []:
        pn = str(r.get("_param", "")).upper()
        if not any(pn.startswith(p + "_") for p in _ST_COLOR_PREFIXES):
            continue
        seen: dict = {}
        cells = r.get("_cells") or {}
        for i in range(len(headers or [])):
            cell = cells.get(i) or cells.get(str(i)) or {}
            for v in (cell.get("actual"), cell.get("plan")):
                if v is None or v == "":
                    continue
                s = str(v)
                if s not in seen:
                    seen[s] = len(seen)
        out[pn] = seen
    return out


_MODULE_KNOB_HILITE = ""


def _is_module_highlight_param(param: str, highlight_knobs: set[str]) -> bool:
    if not highlight_knobs:
        return False
    return str(param or "").strip().upper() in highlight_knobs


def _inform_lot_ids(root: dict) -> list[str]:
    out: list[str] = []
    seen: set[str] = set()

    def add(v: Any) -> None:
        s = str(v or "").strip()
        if not s:
            return
        key = s.lower()
        if key in seen:
            return
        seen.add(key)
        out.append(s)

    embed = root.get("embed_table") or {}
    st = embed.get("st_view") if isinstance(embed, dict) else {}
    if isinstance(st, dict):
        for g in _split_table_header_groups(st):
            if isinstance(g, dict):
                add(g.get("label"))
        for lot in st.get("wafer_fab_list") or []:
            add(lot)
    for raw in str(root.get("fab_lot_id_at_save") or "").split(","):
        add(raw)
    add(root.get("root_lot_id"))
    add(root.get("lot_id"))
    add(root.get("wafer_id"))
    return out


def _default_mail_subject(root: dict) -> str:
    product = str(root.get("product") or "").strip()
    module = str(root.get("module") or "").strip()
    root_lots = [x.strip() for x in str(root.get("root_lot_id") or "").split(",") if x.strip()]
    if not root_lots:
        root_lots = [x.strip() for x in str(root.get("lot_id") or "").split(",") if x.strip()]
    lots = list(dict.fromkeys(root_lots)) or _inform_lot_ids(root)
    if len(lots) <= 1:
        lot = lots[0] if lots else str(root.get("root_lot_id") or root.get("lot_id") or "").strip()
        core = " ".join([x for x in [product, lot] if x])
    else:
        core = f"{product} ({len(lots)}lots)" if product else f"({len(lots)}lots)"
    suffix = f" - {module}" if module else ""
    return f"[plan 적용 통보] {core}{suffix}".strip()


def _default_mail_prose(root: dict, sender_username: str = "") -> str:
    return str((root or {}).get("text") or "").strip()


def _render_embed_table_html(embed: Optional[dict], max_rows: int = 60, module: str = "") -> str:
    """v8.8.22: 인폼에 붙어있는 SplitTable 스냅샷을 메일 본문 HTML 테이블로 인라인 렌더.
    - st_view (SplitTable /view 응답) 가 있으면 parameter x wafer 매트릭스로 렌더.
    - 없으면 legacy 2D (columns/rows) 를 그대로 렌더.
    max_rows 초과 시 하단에 잘림 표시.
    v8.8.30: KNOB/MASK 행에 FE 와 동일한 7색 팔레트 컬러링 (unique-value 기반).
    """
    if not embed or not isinstance(embed, dict):
        return ""
    esc = _html.escape
    # Mail recipients need the business table, not internal Flow/SplitTable source labels.
    # Keep source/note out of the HTML body to avoid exposing IDs, scope strings, or cache details.
    attached_sets = [x for x in (embed.get("attached_sets") or []) if isinstance(x, dict)]

    def _attached_sets_html() -> str:
        # Attached custom-set metadata is kept on the Inform detail record, but
        # recipients should only see the LOT-specific SplitTable snapshot.
        return ""
        if not attached_sets:
            return ""
        chunks = []
        for item in attached_sets[:8]:
            name = esc(str(item.get("name") or "set"))
            source = esc(str(item.get("source") or "set"))
            cols = [str(c) for c in (item.get("columns") or [])][:80]
            rows = [r for r in (item.get("rows") or []) if isinstance(r, list)][:max_rows]
            if not cols:
                chunks.append(
                    f"<div style='margin:8px 0;padding:6px 8px;border:1px solid #d1d5db;border-radius:4px;'>"
                    f"<div style='font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;'>📋 {name} <span style='color:#6b7280;font-weight:400;'>({source})</span></div>"
                    f"<div style='font-size:{_MAIL_MIN_FONT};color:#6b7280;'>columns {int(item.get('columns_count') or 0)} · rows {int(item.get('wafer_count') or 0)}</div>"
                    "</div>"
                )
                continue
            data_cols = max(0, len(cols) - 1)
            first_col_style, data_col_style = _mail_fit_col_styles(data_cols)
            colgroup = _mail_colgroup_html(first_col_style, data_col_style, data_cols)
            th = (f"border:1px solid #d1d5db;padding:2px 3px;background:#f3f4f6;font-size:{_MAIL_MIN_FONT};"
                  "font-family:monospace;text-align:left;white-space:normal;word-break:break-all;overflow-wrap:anywhere;")
            td = (f"border:1px solid #d1d5db;padding:2px 3px;font-size:{_MAIL_MIN_FONT};"
                  "font-family:monospace;text-align:left;white-space:normal;word-break:break-all;overflow-wrap:anywhere;")
            head = "<tr>" + "".join(
                f"<th style='{th}{first_col_style if i == 0 else data_col_style}'>{esc(c)}</th>"
                for i, c in enumerate(cols)
            ) + "</tr>"
            body = "".join(
                "<tr>" + "".join(
                    f"<td style='{td}{first_col_style if i == 0 else data_col_style}'>{esc(str(v if v is not None else ''))}</td>"
                    for i, v in enumerate(row[:len(cols)])
                ) + "</tr>"
                for row in rows
            )
            chunks.append(
                f"<div style='margin:10px 0;'>"
                f"<div style='font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;margin-bottom:4px;'>📋 {name} <span style='color:#6b7280;font-weight:400;'>({source})</span></div>"
                f"<table style='border-collapse:collapse;font-size:{_MAIL_MIN_FONT};width:100%;max-width:100%;table-layout:fixed;'>{colgroup}<thead>{head}</thead><tbody>{body}</tbody></table>"
                "</div>"
            )
        return "".join(chunks)

    # Try st_view first.
    st = embed.get("st_view") or {}
    rows_st = st.get("rows") or []
    headers = st.get("headers") or []
    header_groups = _split_table_header_groups(st)
    root_lot_id = str(st.get("root_lot_id") or "").strip()
    lot_id_values = []
    for group in header_groups:
        label = str(group.get("label") or "").strip() if isinstance(group, dict) else ""
        if label and label not in lot_id_values:
            lot_id_values.append(label)
    lot_id_label = ", ".join(lot_id_values[:3])
    if len(lot_id_values) > 3:
        lot_id_label += f" +{len(lot_id_values) - 3}"
    src = str(embed.get("source") or "")
    m = re.search(r"SplitTable/([^ @·]+)", src)
    product = (m.group(1).strip() if m else "").strip()
    highlight_knobs = _module_highlight_knobs(module)

    def _lineage_summary_html() -> str:
        if not rows_st or not product:
            return ""
        def _step_note(step_ids: list[str]) -> str:
            ids = [str(x).strip() for x in (step_ids or []) if str(x).strip()]
            if len(ids) <= 1:
                return ""
            manual_like = [sid for sid in ids if re.search(r"[A-Z]{2}\d{6}[A-Z]{2}$", sid)]
            parts = ["복수 step_id 이므로 적용 전 담당 엔지니어가 실제 사용 step_id를 확인해 주세요."]
            if manual_like:
                parts.append("수동/예외 step 후보가 포함됐을 수 있습니다.")
            return " ".join(parts)
        try:
            from routers.splittable import _build_knob_meta, _build_inline_meta, _build_vm_meta
            knob_meta = _build_knob_meta(product) or {}
            inline_meta = _build_inline_meta(product) or {}
            vm_meta = _build_vm_meta(product) or {}
        except Exception:
            return ""
        out = []
        for r in rows_st:
            param = str(r.get("_param") or "").strip()
            if not param:
                continue
            km = knob_meta.get(param) or {}
            if km.get("groups"):
                for gi, g in enumerate(km.get("groups") or []):
                    out.append({
                        "key": f"{param}-k-{gi}",
                        "parameter": param,
                        "function_step": str(g.get("func_step") or ""),
                        "step_ids": [str(x) for x in (g.get("step_ids") or []) if str(x or "").strip()],
                    })
                continue
            tail_vm = param.replace("VM_", "", 1)
            vm = vm_meta.get(param) or vm_meta.get(tail_vm) or {}
            if param.startswith("VM_") and (vm.get("groups") or vm.get("step_id") or vm.get("function_step")):
                if vm.get("groups"):
                    for gi, g in enumerate(vm.get("groups") or []):
                        sid = str(g.get("step_id") or "").strip()
                        out.append({
                            "key": f"{param}-v-{gi}",
                            "parameter": param,
                            "function_step": str(g.get("function_step") or vm.get("function_step") or ""),
                            "step_ids": [sid] if sid else ([str(vm.get("step_id"))] if vm.get("step_id") else []),
                        })
                else:
                    sid = str(vm.get("step_id") or "").strip()
                    out.append({
                        "key": f"{param}-v",
                        "parameter": param,
                        "function_step": str(vm.get("function_step") or ""),
                        "step_ids": [sid] if sid else [],
                    })
                continue
            tail_in = param.replace("INLINE_", "", 1)
            im = inline_meta.get(param) or inline_meta.get(tail_in) or {}
            if param.startswith("INLINE_") and (im.get("groups") or im.get("step_id") or im.get("function_step")):
                if im.get("groups"):
                    for gi, g in enumerate(im.get("groups") or []):
                        sid = str(g.get("step_id") or "").strip()
                        out.append({
                            "key": f"{param}-i-{gi}",
                            "parameter": param,
                            "function_step": str(g.get("function_step") or im.get("function_step") or ""),
                            "step_ids": [sid] if sid else [str(x) for x in (im.get("step_ids") or []) if str(x or "").strip()],
                        })
                else:
                    out.append({
                        "key": f"{param}-i",
                        "parameter": param,
                        "function_step": str(im.get("function_step") or ""),
                        "step_ids": [str(x) for x in (im.get("step_ids") or ([im.get("step_id")] if im.get("step_id") else [])) if str(x or "").strip()],
                    })
        if not out:
            return ""
        th = ("border:1px solid #d1d5db;padding:4px 8px;background:#f3f4f6;"
              f"font-size:{_MAIL_MIN_FONT};color:#1f2937;text-align:left;font-family:monospace;"
              "white-space:normal;word-break:break-word;overflow-wrap:anywhere;")
        body = []
        for row in out:
            note = _step_note(row["step_ids"])
            body.append(
                "<tr>"
                f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(row['parameter'])}</td>"
                f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;color:#6b7280;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(row['function_step'] or '—')}</td>"
                f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;color:#2563eb;font-weight:700;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(', '.join(row['step_ids']) if row['step_ids'] else '—')}"
                + (f"<div style='margin-top:4px;font-size:{_MAIL_MIN_FONT};line-height:1.35;color:#dc2626;font-family:system-ui,sans-serif;font-weight:600;'>{esc(note)}</div>" if note else "")
                + "</td>"
                "</tr>"
            )
        return (
            "<div style='margin-top:8px;'>"
            f"<div style='font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;margin-bottom:4px;'>🧭 Parameter별 적용 step 요약</div>"
            f"<table style='border-collapse:collapse;font-size:{_MAIL_MIN_FONT};max-width:100%;table-layout:fixed;'>"
            "<thead><tr>"
            f"<th style='{th}'>parameter</th><th style='{th}'>function_step</th><th style='{th}'>step_id</th>"
            "</tr></thead>"
            f"<tbody>{''.join(body)}</tbody></table>"
            f"<div style='margin-top:6px;font-size:{_MAIL_MIN_FONT};color:#6b7280;'>function_step 에 여러 step_id 가 연결되면 메일 수신/적용 엔지니어가 현재 제품의 유효 step_id 를 확인한 뒤 적용해야 합니다.</div>"
            "</div>"
        )

    def _plan_summary_html() -> str:
        if not rows_st or not headers:
            return ""
        out = []
        for idx, wafer in enumerate(headers):
            plans = []
            for r in rows_st:
                cells = r.get("_cells") or {}
                cell = cells.get(idx) or cells.get(str(idx)) or {}
                plan = cell.get("plan")
                if plan in (None, ""):
                    continue
                plans.append({
                    "parameter": str(r.get("_param") or ""),
                    "actual": "" if cell.get("actual") in (None, "") else str(cell.get("actual")),
                    "plan": str(plan),
                })
            for p in plans:
                out.append(
                    "<tr>"
                    f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;font-weight:700;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(str(wafer or ''))}</td>"
                    f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(p['parameter'])}</td>"
                    f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;color:#6b7280;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(p['actual'] or '—')}</td>"
                    f"<td style='border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};font-family:monospace;color:#ea580c;font-weight:700;white-space:normal;word-break:break-word;overflow-wrap:anywhere;'>{esc(p['plan'])}</td>"
                    "</tr>"
                )
        if not out:
            return ""
        th = ("border:1px solid #d1d5db;padding:4px 8px;background:#f3f4f6;"
              f"font-size:{_MAIL_MIN_FONT};color:#1f2937;text-align:left;font-family:monospace;"
              "white-space:normal;word-break:break-word;overflow-wrap:anywhere;")
        return (
            "<div style='margin-top:8px;'>"
            f"<div style='font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;margin-bottom:4px;'>📋 Wafer별 적용 plan 요약</div>"
            f"<table style='border-collapse:collapse;font-size:{_MAIL_MIN_FONT};max-width:100%;table-layout:fixed;'>"
            "<thead><tr>"
            f"<th style='{th}'>wafer</th><th style='{th}'>parameter</th><th style='{th}'>actual</th><th style='{th}'>plan</th>"
            "</tr></thead>"
            f"<tbody>{''.join(out)}</tbody></table>"
            "</div>"
        )

    def _wrap(body_rows_html: str, head_cells: list[str], truncated: bool,
              highlight_col_indices: Optional[set[int]] = None) -> str:
        first_col_style, data_col_style = _mail_fit_col_styles(max(0, len(head_cells) - 1))
        th_style = ("border:1px solid #d1d5db;padding:4px 8px;background:#f3f4f6;"
                    f"font-size:{_MAIL_MIN_FONT};color:#1f2937;text-align:center;font-family:monospace;line-height:1.25;")
        table_style = _mail_table_style(max(0, len(head_cells) - 1))
        colgroup = _mail_colgroup_html(first_col_style, data_col_style, max(0, len(head_cells) - 1))
        highlight_col_indices = highlight_col_indices or set()
        thead = "<tr>" + "".join(
            f"<th style='{th_style}{first_col_style if i == 0 else data_col_style}{_MODULE_KNOB_HILITE if i in highlight_col_indices else ''}'>{c}</th>"
            for i, c in enumerate(head_cells)
        ) + "</tr>"
        hdr = (
            f"<div style='margin:12px 0 4px 0;font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;'>"
            f"Split table"
            + "</div>"
        )
        note_html = ""
        trunc_html = (f"<div style='font-size:{_MAIL_MIN_FONT};color:#b91c1c;margin-top:4px;'>"
                      f"⚠ {max_rows}행으로 잘림 — 전체 데이터는 첨부 xlsx 참고</div>") if truncated else ""
        return (
            f"{hdr}{note_html}"
            f"<div style='width:100%;max-width:100%;'>"
            f"<table style='{table_style}'>"
            f"{colgroup}"
            f"<thead>{thead}</thead>"
            f"<tbody>{body_rows_html}</tbody>"
            f"</table></div>{trunc_html}{_attached_sets_html()}{_plan_summary_html()}{_lineage_summary_html()}"
        )

    if rows_st and headers:
        truncated = len(rows_st) > max_rows
        shown = rows_st[:max_rows]
        # v8.8.30: KNOB/MASK 행용 unique-value → color index 맵.
        uniq_map = _st_build_uniq_map(shown, headers)
        dense = len(headers) >= 12
        data_pad = "4px 5px" if dense else "4px 8px"
        first_pad = "4px 6px" if dense else "4px 8px"
        font_sz = _MAIL_MIN_FONT
        line_h = "1.25"
        first_col_style, data_col_style = _mail_scroll_col_styles(len(headers))
        colgroup = _mail_colgroup_html(first_col_style, data_col_style, len(headers))
        td_first = (f"border:1px solid #d1d5db;padding:{first_pad};background:#f9fafb;"
                    f"font-size:{font_sz};font-weight:700;font-family:monospace;line-height:{line_h};")
        td_cell_base = (f"border:1px solid #d1d5db;padding:{data_pad};text-align:center;"
                        f"font-size:{font_sz};font-family:monospace;line-height:{line_h};")
        th_style = (f"border:1px solid #d1d5db;padding:{data_pad};background:#f3f4f6;"
                    f"font-size:{font_sz};color:#1f2937;text-align:center;font-family:monospace;line-height:{line_h};")
        th_label = th_style + first_col_style + "text-align:left;font-weight:700;color:#6b7280;"
        th_root = ("border:1px solid #d1d5db;padding:5px 8px;background:#f3f4f6;"
                   f"font-size:{_MAIL_MIN_FONT};color:#ea580c;text-align:center;font-family:monospace;font-weight:700;")
        th_group = (f"border:1px solid #d1d5db;padding:{data_pad};background:#f9fafb;"
                    f"font-size:{font_sz};color:#b45309;text-align:center;font-family:monospace;font-weight:700;")
        hdr = (
            f"<div style='margin:12px 0 4px 0;font-size:{_MAIL_MIN_FONT};font-weight:700;color:#ea580c;'>"
            f"Split table"
            + "</div>"
        )
        note_html = ""
        trunc_html = (f"<div style='font-size:{_MAIL_MIN_FONT};color:#b91c1c;margin-top:4px;'>"
                      f"⚠ {max_rows}행으로 잘림 — 전체 데이터는 첨부 xlsx 참고</div>") if truncated else ""
        body_parts = []
        for r in shown:
            param_raw = str(r.get("_param", ""))
            param = esc(str(r.get("_display") or param_raw or "").replace("KNOB_", "").replace("MASK_", "").replace("INLINE_", "").replace("VM_", ""))
            cells = r.get("_cells") or {}
            row_highlight = _is_module_highlight_param(param_raw, highlight_knobs)
            hilite = _MODULE_KNOB_HILITE if row_highlight else ""
            tds = [f"<td style='{td_first}{first_col_style}{hilite}'>{param}</td>"]
            for i in range(len(headers)):
                cell = cells.get(i) or cells.get(str(i)) or {}
                actual = cell.get("actual")
                plan = cell.get("plan")
                disp = "" if actual is None or actual == "" else str(actual)
                plan_diff = plan is not None and plan != "" and plan != actual
                if plan_diff:
                    disp_html = f"{esc(disp)} <span style='color:#ea580c;font-weight:700'>→ {esc(str(plan))}</span>"
                else:
                    disp_html = esc(disp)
                # plan 이 있으면 plan 기준으로 컬러링해 SplitTable unique 색상과 맞춘다.
                paint_val = plan if plan not in (None, "") else actual
                cell_bg = _st_cell_bg(paint_val, uniq_map, param_raw)
                tds.append(f"<td style='{td_cell_base}{data_col_style}{cell_bg}{hilite}'>{disp_html}</td>")
            body_parts.append("<tr>" + "".join(tds) + "</tr>")

        thead_parts = [
            "<tr>"
            f"<th style='{th_label}'>root_lot_id</th>"
            f"<th colspan='{max(1, len(headers))}' style='{th_root}{_mail_data_col_style(data_col_style, len(headers))}'>{esc(root_lot_id or '—')}</th>"
            "</tr>"
        ]
        if header_groups:
            cells = [f"<th style='{th_label}'>lot_id</th>"]
            for g in header_groups:
                span = max(1, int(g.get("span") or 1))
                cells.append(f"<th colspan='{span}' style='{th_group}{_mail_data_col_style(data_col_style, span)}'>{esc(str(g.get('label') or '—'))}</th>")
            thead_parts.append("<tr>" + "".join(cells) + "</tr>")
        else:
            thead_parts.append(
                "<tr>"
                f"<th style='{th_label}'>lot_id</th>"
                f"<th colspan='{max(1, len(headers))}' style='{th_group}{_mail_data_col_style(data_col_style, len(headers))}'>{esc(lot_id_label or '—')}</th>"
                "</tr>"
            )
        head_cells = ["항목"] + [esc(h or "") for h in headers]
        thead_parts.append("<tr>" + "".join(
            f"<th style='{th_style}{first_col_style if i == 0 else data_col_style}'>{c}</th>"
            for i, c in enumerate(head_cells)
        ) + "</tr>")
        table_style = _mail_scroll_table_style(len(headers))
        table_html = (
            "<div style='overflow-x:auto;-webkit-overflow-scrolling:touch;max-width:100%'>"
            f"<table style='{table_style}'>"
            f"{colgroup}"
            f"<thead>{''.join(thead_parts)}</thead>"
            f"<tbody>{''.join(body_parts)}</tbody>"
            "</table></div>"
        )
        return (
            f"{hdr}{note_html}"
            f"{table_html}{trunc_html}{_attached_sets_html()}{_plan_summary_html()}{_lineage_summary_html()}"
        )

    # Legacy 2D path.
    if attached_sets and not rows_st and str(embed.get("source") or "").startswith("SplitTable selected sets"):
        return ""
    cols = embed.get("columns") or []
    rows2d = embed.get("rows") or []
    if not cols and not rows2d:
        return _attached_sets_html()
    truncated = len(rows2d) > max_rows
    shown = rows2d[:max_rows]
    first_col_style, data_col_style = _mail_fit_col_styles(max(0, len(cols) - 1))
    td_cell = (f"border:1px solid #d1d5db;padding:4px 8px;font-size:{_MAIL_MIN_FONT};"
               "font-family:monospace;line-height:1.2;")
    highlight_col_indices = {
        i for i, c in enumerate(cols)
        if _is_module_highlight_param(str(c or ""), highlight_knobs)
    }
    body_parts = []
    for r in shown:
        if not isinstance(r, (list, tuple)):
            continue
        row_highlight = bool(r) and _is_module_highlight_param(str(r[0] or ""), highlight_knobs)
        tds = "".join(
            f"<td style='{td_cell}{first_col_style if i == 0 else data_col_style}{_MODULE_KNOB_HILITE if (row_highlight or i in highlight_col_indices) else ''}'>{esc('' if v is None else str(v))}</td>"
            for i, v in enumerate(r)
        )
        body_parts.append(f"<tr>{tds}</tr>")
    return _wrap("".join(body_parts), [esc(c or "") for c in cols], truncated, highlight_col_indices)


def _build_html_body(root: dict, thread_html: str, extra_prose: str,
                     sender_username: str = "", product_contacts: Optional[list] = None,
                     embed_table: Optional[dict] = None) -> str:
    """최상위 루트 메타 + 사용자 prose + 스레드 HTML 을 한 문서로.

    v8.8.0: 메일은 Admin 계정으로 발송되더라도 본문에 실제 요청자(sender_username) 를
    명시. 또한 해당 product 의 담당자 그룹(product_contacts) 이 있으면 표로 첨부.
    v8.8.22: embed_table 이 있으면 SplitTable 스냅샷을 HTML 테이블로 본문에 인라인 주입.
    v8.8.30: 메타 테이블을 4필드로 축약 — 제품 / Lot / 작성자 / 작성시간.
      (모듈/사유/wafer/deadline/진행상태는 SplitTable 스냅샷 + 본문으로 충분, 중복 제거.)
    """
    esc = _html.escape
    # v8.8.30: created_at 을 "YYYY-MM-DD HH:MM" 로 정돈. 없으면 스킵.
    created_raw = (root.get("created_at") or "").strip()
    created_disp = created_raw.replace("T", " ")[:16] if created_raw else ""
    meta_rows = []
    for k, label, disp in [
        ("product",    "제품",     root.get("product", "")),
        ("lot_id",     "Lot",      root.get("lot_id") or root.get("root_lot_id") or ""),
        ("author",     "작성자",   root.get("author", "")),
        ("__created__","작성시간", created_disp),
    ]:
        val = disp
        if not val:
            continue
        meta_rows.append(
            f"<tr><td style='padding:4px 10px;font-size:{_MAIL_MIN_FONT};color:#6b7280;background:#f3f4f6;width:90px;'>{esc(label)}</td>"
            f"<td style='padding:4px 10px;font-size:{_MAIL_MIN_FONT};color:#1f2937;font-family:monospace;'>{esc(val)}</td></tr>"
        )
    meta_tbl = "<table style='border-collapse:collapse;border:1px solid #d1d5db;margin:10px 0;width:100%;max-width:560px;'>" + "".join(meta_rows) + "</table>"
    prose_block = ""
    # Keep the mail body anchored to the inform note. If the caller explicitly
    # supplies prose use it; otherwise use only the saved note, not a generated
    # greeting/template.
    prose_text = extra_prose if (extra_prose or "").strip() else str(root.get("text") or "")
    if prose_text.strip():
        safe = _html.escape(prose_text).replace("\n", "<br/>")
        prose_block = (
            "<div style='margin:0 0 12px 0;padding:0;background:transparent;border:none;"
            "font-size:12pt;line-height:1.45;color:#1f2937;'>"
            f"{safe}</div>"
        )
    # v8.8.1: 발송 요청자(hol) 자동 명시 제거.
    contacts_block = ""
    if product_contacts:
        names = []
        for c in product_contacts:
            nm = (c.get("name") or "").strip()
            em = (c.get("email") or "").strip()
            if nm and em:
                names.append(f"{esc(nm)} &lt;{esc(em)}&gt;")
            elif nm:
                names.append(esc(nm))
            elif em:
                names.append(esc(em))
        if names:
            contacts_block = (
                f"<div style='margin:10px 0;padding:10px 12px;background:#f0fdf4;border-left:4px solid #16a34a;"
                f"border-radius:4px;font-size:14px;line-height:1.45;color:#14532d;font-weight:700;'>"
                f"제품 담당자 : " + ", ".join(names)
                + "</div>"
            )
    embed_html = _render_embed_table_html(embed_table, module=root.get("module", "")) if embed_table else ""
    # v8.8.30: 스레드 섹션 조건부 — thread_html 이 비어있으면 "스레드" 헤더 자체를 넣지 않음.
    #   FE include_thread 체크박스 제거에 맞춰, 기본 발송은 스레드 없는 간결 본문.
    thread_block = (
        f"<h3 style='font-size:{_MAIL_MIN_FONT};margin:14px 0 6px 0;color:#374151;'>스레드</h3>{thread_html}"
        if (thread_html or "").strip() else ""
    )
    return (
        "<div style='font-family:-apple-system,Segoe UI,Arial,sans-serif;color:#1f2937;width:100%;max-width:none;margin:0;'>"
        f"{prose_block}"
        f"{contacts_block}"
        f"{meta_tbl}"
        f"{embed_html}"
        f"{thread_block}"
        "<hr style='border:none;border-top:1px solid #e5e7eb;margin:18px 0 8px 0;'/>"
        f"<div style='font-size:{_MAIL_MIN_FONT};color:#6b7280;margin-bottom:4px;'>상세 확인 및 후속 조치는 "
        f"<a href='{esc(_GO_FLOW_URL)}' target='_blank' rel='noopener noreferrer' "
        f"style='color:#ea580c;text-decoration:underline;font-weight:700;'>go/flow</a> 에서 진행해 주세요.</div>"
        "</div>"
    )


def _resolve_users_to_emails(usernames: List[str]) -> List[str]:
    """v8.8.17: 사내 username 이 곧 사내 email 인 환경 대응.
      우선순위:
        1) users.csv 에 email 필드가 있고 '@' 가 포함되면 그 값.
        2) username 자체가 '@' 를 포함하면 email 로 취급 (가입 시 사내 메일 입력 규약).
        3) admin mail domain 이 있으면 <username>@<domain> 합성.
        4) 둘 다 아니면 제외.
    """
    if not usernames:
        return []
    from routers.auth import read_users
    all_users = {u.get("username", ""): u for u in read_users()}
    out = []
    seen: set[str] = set()
    try:
        domain = str((_load_mail_cfg().get("domain") or "")).strip().lstrip("@")
    except Exception:
        domain = ""

    def _looks_like_email(v: str) -> bool:
        if not v or "@" not in v:
            return False
        _local, _, _domain = v.partition("@")
        return bool(_local) and "." in _domain

    def _append(v: str) -> bool:
        em = (v or "").strip()
        if not _looks_like_email(em):
            return False
        key = em.lower()
        if key in seen:
            return True
        seen.add(key)
        out.append(em)
        return True

    for un in usernames:
        un = str(un or "").strip()
        if not un:
            continue
        u = all_users.get(un) or {}
        em = (u.get("email") or "").strip()
        if _append(em):
            continue
        # username 자체가 이메일 포맷이면 그대로 사용.
        if _append(un):
            continue
        if domain:
            _append(f"{un}@{domain}")
    return out


MAIL_CONTENT_MAX = 2 * 1024 * 1024          # 2 MB HTML body
MAIL_ATTACH_MAX  = 10 * 1024 * 1024         # 10 MB total attachments


def _inform_snapshot_simple_sheets(target: dict) -> list[dict]:
    rows1 = [["Field", "Value"]]
    meta_keys = ("id", "product", "module", "reason", "root_lot_id", "lot_id",
                 "wafer_id", "fab_lot_id_at_save", "flow_status", "author",
                 "created_at", "updated_at")
    for k in meta_keys:
        v = target.get(k, "")
        if v not in ("", None):
            rows1.append([k, str(v)])
    sc = target.get("splittable_change") or {}
    if sc:
        rows1.extend([[], ["SplitTable Change", ""]])
        for k in ("column", "old_value", "new_value", "applied"):
            if k in sc:
                rows1.append([k, str(sc.get(k, ""))])
    body = (target.get("text") or "").strip()
    if body:
        rows1.extend([[], ["Body", ""]])
        for ln in body.splitlines():
            rows1.append(["", ln])

    sheets = [{"title": "Inform Snapshot", "rows": rows1}]
    embed = target.get("embed_table") or {}
    st = embed.get("st_view") or {}
    headers = [str(h or "") for h in (st.get("headers") or [])]
    rows = st.get("rows") or []
    if headers or rows:
        rows2 = []
        merges = []
        root_lot_id = str(st.get("root_lot_id") or "").strip()
        if root_lot_id:
            rows2.append(["", root_lot_id, *["" for _ in headers[1:]]])
            if headers:
                merges.append((len(rows2), 2, len(rows2), len(headers) + 1))
        header_groups = _split_table_header_groups(st)
        if header_groups:
            out = [""]
            col = 2
            merge_row = len(rows2) + 1
            for group in header_groups:
                span = max(1, int(group.get("span") or 1))
                out.extend([str(group.get("label") or ""), *["" for _ in range(span - 1)]])
                if span > 1:
                    merges.append((merge_row, col, merge_row, col + span - 1))
                col += span
            rows2.append(out[:len(headers) + 1])
        rows2.append(["parameter", *headers])
        for row in rows:
            label = str(row.get("_param") or row.get("_display") or "")
            cells = row.get("_cells") or {}
            out = [label]
            for i, _ in enumerate(headers):
                cell = cells.get(i) or cells.get(str(i)) or {}
                actual = cell.get("actual")
                plan = cell.get("plan")
                if plan not in (None, "") and str(plan) != str(actual if actual is not None else ""):
                    out.append(f"{'' if actual is None else actual} -> {plan}")
                else:
                    out.append("" if actual is None else str(actual))
            rows2.append(out)
        sheets.append({"title": "SplitTable Snapshot", "rows": rows2, "merges": merges})
    elif embed.get("columns") or embed.get("rows"):
        rows2 = []
        cols = [str(c or "") for c in (embed.get("columns") or [])]
        if cols:
            rows2.append(cols)
        for row in (embed.get("rows") or []):
            rows2.append([("" if v is None else str(v)) for v in row])
        sheets.append({"title": "Embedded Table", "rows": rows2})
    return sheets


# v8.8.21: 인폼 → SplitTable 스냅샷 xlsx 자동 첨부.
#   사용자가 직접 파일 업로드하지 않아도, 해당 인폼에 담긴 제품/lot/wafer 스냅샷을
#   SplitTable 엑셀 내보내기와 동일 형식으로 렌더해 첨부한다.
def _build_inform_snapshot_xlsx(target: dict) -> Optional[tuple]:
    """Return (filename, bytes, mime) or None if no snapshot available."""
    try:
        import io as _io
        mime = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        product = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(target.get("product") or "product")).strip("_") or "product"
        lot = re.sub(r"[^A-Za-z0-9_.-]+", "_", str(target.get("lot_id") or target.get("root_lot_id") or target.get("wafer_id") or "lot")).strip("_") or "lot"
        fn = f"inform_{product}_{lot}.xlsx"
        try:
            from openpyxl import Workbook
        except Exception:
            try:
                from core.simple_xlsx import build_workbook
                data = build_workbook(_inform_snapshot_simple_sheets(target))
                return (fn, data, mime) if data else None
            except Exception:
                return None
        wb = Workbook()
        ws = wb.active
        ws.title = "Inform Snapshot"
        ws.append(["Field", "Value"])
        meta_keys = ("id", "product", "module", "reason", "root_lot_id", "lot_id",
                     "wafer_id", "fab_lot_id_at_save", "flow_status", "author",
                     "created_at", "updated_at")
        for k in meta_keys:
            v = target.get(k, "")
            if v not in ("", None):
                ws.append([k, str(v)])
        sc = target.get("splittable_change") or {}
        if sc:
            ws.append([])
            ws.append(["SplitTable Change", ""])
            for k in ("column", "old_value", "new_value", "applied"):
                if k in sc:
                    ws.append([k, str(sc.get(k, ""))])
        # body text 도 같이 싣기 (읽기 편하게).
        body = (target.get("text") or "").strip()
        if body:
            ws.append([])
            ws.append(["Body", ""])
            for ln in body.splitlines():
                ws.append(["", ln])
        embed = target.get("embed_table") or {}
        st = embed.get("st_view") or {}
        headers = [str(h or "") for h in (st.get("headers") or [])]
        rows = st.get("rows") or []
        if headers or rows:
            ws2 = wb.create_sheet("SplitTable Snapshot")
            header_groups = _split_table_header_groups(st)
            root_lot_id = str(st.get("root_lot_id") or "").strip()
            if root_lot_id:
                ws2.append(["", root_lot_id, *["" for _ in headers[1:]]])
                if headers:
                    ws2.merge_cells(start_row=ws2.max_row, start_column=2, end_row=ws2.max_row, end_column=len(headers) + 1)
            if header_groups:
                row = [""]
                for group in header_groups:
                    span = max(1, int(group.get("span") or 1))
                    row.extend([str(group.get("label") or ""), *["" for _ in range(span - 1)]])
                ws2.append(row[:len(headers) + 1])
                merge_row = ws2.max_row
                col = 2
                for group in header_groups:
                    span = max(1, int(group.get("span") or 1))
                    if span > 1:
                        ws2.merge_cells(start_row=merge_row, start_column=col, end_row=merge_row, end_column=col + span - 1)
                    col += span
            ws2.append(["parameter", *headers])
            for row in rows:
                label = str(row.get("_param") or row.get("_display") or "")
                cells = row.get("_cells") or {}
                out = [label]
                for i, _ in enumerate(headers):
                    cell = cells.get(i) or cells.get(str(i)) or {}
                    actual = cell.get("actual")
                    plan = cell.get("plan")
                    if plan not in (None, "") and str(plan) != str(actual if actual is not None else ""):
                        out.append(f"{'' if actual is None else actual} -> {plan}")
                    else:
                        out.append("" if actual is None else str(actual))
                ws2.append(out)
        elif embed.get("columns") or embed.get("rows"):
            ws2 = wb.create_sheet("Embedded Table")
            cols = [str(c or "") for c in (embed.get("columns") or [])]
            if cols:
                ws2.append(cols)
            for row in (embed.get("rows") or []):
                ws2.append([("" if v is None else str(v)) for v in row])
        buf = _io.BytesIO()
        wb.save(buf)
        data = buf.getvalue()
        if not data:
            return None
        return (fn, data, mime)
    except Exception:
        return None


@router.get("/{inform_id}/mail-preview")
def mail_preview(inform_id: str, request: Request, body: str = Query(""),
                 subject: str = Query(""),
                 to: Optional[List[str]] = Query(None),
                 to_users: Optional[List[str]] = Query(None),
                 groups: Optional[List[str]] = Query(None)):
    """v8.8.21: FE MailDialog 용 실시간 프리뷰.
    실제로 send-mail 이 호출할 HTML body / 수신자 그룹 해석 / 담당자 라인 / 자동 첨부 목록을 반환.
    v8.8.30: HTML 용량 계산 + 2MB 한도 초과 경고 플래그를 응답에 포함 (FE 가 실제 발송 전 경고).
    스레드는 기본 미포함(간결 본문) — FE 체크박스 제거와 정합.
    """
    me = current_user(request)
    items = _load_upgraded()
    target = _find(items, inform_id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    pc_data = _load_product_contacts()
    pc_list = (pc_data.get("products") or {}).get(target.get("product", ""), []) or []
    cfg = _load_mail_cfg()
    body_text = body if isinstance(body, str) else ""
    subject_text = subject if isinstance(subject, str) else ""
    q_to = list(to) if isinstance(to, list) else []
    q_to_users = list(to_users) if isinstance(to_users, list) else []
    q_groups = list(groups) if isinstance(groups, list) else []
    to_addrs, auto_rows, used_auto = _resolve_mail_recipients(
        target,
        to=q_to,
        to_users=q_to_users,
        groups=q_groups,
        cfg=cfg,
    )
    # v8.8.30: 프리뷰는 스레드 섹션 없이 보여줘 — 발송 시 include_thread 기본 off 와 일치.
    html = _build_html_body(target, "", body_text, sender_username=me.get("username", ""),
                             product_contacts=pc_list,
                             embed_table=target.get("embed_table"))
    size_bytes = len(html.encode("utf-8"))
    snap = _build_inform_snapshot_xlsx(target)
    owners_line = ", ".join([(c.get("name") or c.get("email") or "").strip()
                             for c in pc_list if c.get("name") or c.get("email")])
    return {
        "inform_id": inform_id,
        "product": target.get("product", ""),
        "subject": subject_text.strip() or _default_mail_subject(target),
        "resolved_recipients": to_addrs,
        "auto_module_recipients": auto_rows,
        "auto_module_used": used_auto,
        "owners_line": owners_line,
        "product_contacts": pc_list,
        "html_body": html,
        "html_size_bytes": size_bytes,
        "html_size_kb": round(size_bytes / 1024, 1),
        "html_size_limit_bytes": MAIL_CONTENT_MAX,
        "html_over_limit": size_bytes > MAIL_CONTENT_MAX,
        "auto_attachments": [{"name": snap[0], "bytes": len(snap[1])}] if snap else [],
        "attachment_total_bytes": (len(snap[1]) if snap else 0),
        "attachment_total_kb": round((len(snap[1]) if snap else 0) / 1024, 1),
    }


def _resolve_inform_attachment(url: str) -> Optional[Path]:
    """Map /api/informs/files/{uid}/{name} → local UPLOADS_DIR/{uid}/{name}."""
    if not url:
        return None
    m = re.match(r"^/?api/informs/files/([A-Za-z0-9_\-]+)/([^/\\?#]+)", url)
    if not m:
        return None
    uid, name = m.group(1), m.group(2)
    if ".." in name or "/" in name or "\\" in name:
        return None
    fp = UPLOADS_DIR / uid / name
    try:
        fp_res = fp.resolve()
        root_res = UPLOADS_DIR.resolve()
        fp_res.relative_to(root_res)  # traversal guard
    except Exception:
        return None
    return fp if fp.is_file() else None


def _encode_multipart(fields: Dict[str, str], files: List[tuple]) -> tuple:
    """Encode form fields + files as multipart/form-data.
    fields: {name: string_value}
    files:  [(field_name, filename, bytes, mime)]
    Returns (body_bytes, content_type_header).
    """
    boundary = "----flowInform" + uuid.uuid4().hex
    chunks: List[bytes] = []
    for name, value in fields.items():
        chunks.append(f"--{boundary}\r\n".encode())
        chunks.append(f'Content-Disposition: form-data; name="{name}"\r\n'.encode())
        chunks.append(b"Content-Type: text/plain; charset=utf-8\r\n\r\n")
        chunks.append(str(value).encode("utf-8"))
        chunks.append(b"\r\n")
    for (fname_field, filename, content, mime) in files:
        chunks.append(f"--{boundary}\r\n".encode())
        safe_fn = filename.replace('"', '').replace("\r", "").replace("\n", "")
        chunks.append(
            f'Content-Disposition: form-data; name="{fname_field}"; filename="{safe_fn}"\r\n'.encode()
        )
        chunks.append(f"Content-Type: {mime}\r\n\r\n".encode())
        chunks.append(content)
        chunks.append(b"\r\n")
    chunks.append(f"--{boundary}--\r\n".encode())
    return b"".join(chunks), f"multipart/form-data; boundary={boundary}"


# v8.8.27: 메일 폭탄 방지 — 인폼 한 건당 연속 발송 rate limit.
#   철학: 인폼은 "LOT 단위 split 설정 plan 을 각 모듈에 알림" + 변경 발생 시 재통지.
#     하지만 한 인폼을 초당 수십번 발송하거나 유저가 전체 인폼을 폭발적으로 쏘는 상황은
#     사내 메일 서버에 부담 + 수신자 피로 → 서버에서 단단히 막는다.
#   룰:
#     1) 인폼별 쿨다운 `MAIL_COOLDOWN_SEC` (기본 30s) — 같은 inform 에 직전 발송 시점부터.
#     2) 인폼별 버스트 `MAIL_BURST_PER_INFORM` (10분 5건) — mail_history 기준.
#     3) 유저별 글로벌 버스트 `MAIL_BURST_PER_USER` (10분 30건) — 전 인폼 누적.
#   위반 시 429 + `detail` + `retry_after_sec`. 관리자라도 동일 적용 (loud > silent).
MAIL_COOLDOWN_SEC = 30
MAIL_BURST_PER_INFORM_N = 5
MAIL_BURST_PER_INFORM_WINDOW = 600  # 10min
MAIL_BURST_PER_USER_N = 30
MAIL_BURST_PER_USER_WINDOW = 600    # 10min


def _parse_iso(ts: str) -> float:
    """ISO string → epoch seconds. 실패 시 0."""
    try:
        return datetime.datetime.fromisoformat(str(ts).replace("Z", "")).timestamp()
    except Exception:
        return 0.0


def _check_mail_throttle(items: list, inform_id: str, target: dict, username: str) -> None:
    """v8.8.27: 메일 발송 전 rate limit 검사. 위반 시 HTTPException(429) raise."""
    now = datetime.datetime.now().timestamp()
    hist = list(target.get("mail_history") or [])
    # rule 1: 인폼 쿨다운
    if hist:
        last_ts = _parse_iso(hist[-1].get("at", "")) if hist else 0.0
        if last_ts and (now - last_ts) < MAIL_COOLDOWN_SEC:
            wait = int(MAIL_COOLDOWN_SEC - (now - last_ts)) + 1
            raise HTTPException(429, f"이 인폼에 방금 메일을 보냈습니다. {wait}초 후 다시 시도하세요 (쿨다운 {MAIL_COOLDOWN_SEC}초).")
    # rule 2: 인폼별 버스트
    per_inform_recent = [h for h in hist
                         if (now - _parse_iso(h.get("at", ""))) < MAIL_BURST_PER_INFORM_WINDOW]
    if len(per_inform_recent) >= MAIL_BURST_PER_INFORM_N:
        raise HTTPException(429,
            f"이 인폼의 최근 10분 발송 횟수가 한도({MAIL_BURST_PER_INFORM_N}회)를 초과했습니다. 수신자에게 피로가 누적되지 않도록 잠시 후 다시 시도하세요.")
    # rule 3: 유저 글로벌 버스트 — 전 인폼 누적.
    if username:
        total_recent = 0
        for it in items:
            for h in (it.get("mail_history") or []):
                if h.get("by") != username:
                    continue
                if (now - _parse_iso(h.get("at", ""))) < MAIL_BURST_PER_USER_WINDOW:
                    total_recent += 1
        if total_recent >= MAIL_BURST_PER_USER_N:
            raise HTTPException(429,
                f"최근 10분 사이 {total_recent}건의 인폼 메일을 보냈습니다 (한도 {MAIL_BURST_PER_USER_N}). 메일 폭주 방지를 위해 잠시 대기해주세요.")


@router.post("/{inform_id}/send-mail")
def send_mail(inform_id: str, req: SendMailReq, request: Request):
    """인폼 HTML 본문 + 선택 수신자로 사내 메일 API 호출.

    수신자 확정 순서:
      1) req.to (이메일 직접 지정)
      2) req.to_users (username → email 매핑)
      3) req.groups → admin 설정 recipient_groups[group] (username 리스트) → email

    Admin 설정의 mail.api_url/headers/from_addr/extra_data 를 사용. enabled=False
    이거나 api_url 이 비어있으면 400. api_url=='dry-run' 이면 실제 전송 없이 payload
    를 그대로 반환 (구성 검증용).

    v8.8.27: 메일 폭탄 방지 rate limit 검사 추가 (인폼 쿨다운·인폼 버스트·유저 글로벌 버스트).
    """
    me = current_user(request)
    cfg = _load_mail_cfg()
    if not cfg.get("enabled") or not (cfg.get("api_url") or "").strip():
        raise HTTPException(400, "메일 API 가 설정되지 않았습니다. Admin > 메일 API 에서 활성화하세요.")

    items = _load_upgraded()
    target = _find(items, inform_id)
    if not target:
        raise HTTPException(404, "인폼을 찾을 수 없습니다.")
    if _is_deleted(target):
        raise HTTPException(404, "인폼을 찾을 수 없습니다.")
    if me.get("role") != "admin" and target.get("author") != me.get("username"):
        raise HTTPException(403, "작성자 또는 admin 만 인폼 메일을 발송할 수 있습니다.")

    # v8.8.27: 발송 rate limit 검사 (dry-run 도 동일 적용 — preview 스팸도 방지).
    _check_mail_throttle(items, inform_id, target, me.get("username", ""))

    # Resolve recipients. If the sender did not explicitly choose people or
    # groups, a module inform goes to users assigned to that module.
    to_addrs, auto_module_recipients, used_auto_module = _resolve_mail_recipients(
        target,
        to=list(req.to or []),
        to_users=list(req.to_users or []),
        groups=list(req.groups or []),
        cfg=cfg,
    )

    if not to_addrs:
        raise HTTPException(400, "수신자 이메일이 없습니다 (유저 email 또는 group 을 먼저 설정하세요).")
    if len(to_addrs) > 199:
        raise HTTPException(400, f"수신자는 최대 199명까지 지정할 수 있습니다 (현재 {len(to_addrs)}명).")
    # receiverList object form per mail API spec.
    receiver_list = [{"email": em, "recipientType": "TO", "seq": i + 1}
                     for i, em in enumerate(to_addrs)]
    to_list = to_addrs  # kept for audit (plain list of emails)

    subject = (req.subject or "").strip() or _default_mail_subject(target)
    # HTML body (content)
    thread_html = _thread_html(items, inform_id) if req.include_thread else ""
    # v8.8.0: sender = 실제 요청 유저(me), 제품 담당자 자동 첨부.
    pc_data = _load_product_contacts()
    pc_list = (pc_data.get("products") or {}).get(target.get("product", ""), []) or []
    html_body = _build_html_body(target, thread_html, (req.body or ""),
                                 sender_username=me.get("username", ""),
                                 product_contacts=pc_list,
                                 embed_table=target.get("embed_table"))
    content_bytes_len = len(html_body.encode("utf-8"))
    if content_bytes_len > MAIL_CONTENT_MAX:
        raise HTTPException(400, f"메일 본문이 2MB 한도를 초과했습니다 ({content_bytes_len // 1024}KB). 스레드 첨부를 끄거나 본문을 줄여주세요.")

    # Collect attachments (optional)
    attach_files: List[tuple] = []
    attach_total = 0
    # v8.8.21: 인폼 스냅샷 xlsx 자동 첨부 — FE 에서 별도 업로드 불필요.
    snap = _build_inform_snapshot_xlsx(target)
    if snap:
        _fn, _bytes, _mime = snap
        attach_total += len(_bytes)
        if attach_total <= MAIL_ATTACH_MAX:
            attach_files.append(("files", _fn, _bytes, _mime))
    for url_ in (req.attachments or []):
        fp = _resolve_inform_attachment(url_)
        if not fp:
            continue
        content = fp.read_bytes()
        attach_total += len(content)
        if attach_total > MAIL_ATTACH_MAX:
            raise HTTPException(400, f"첨부파일 총 용량이 10MB 한도를 초과했습니다 ({attach_total // 1024}KB).")
        mime = mimetypes.guess_type(fp.name)[0] or "application/octet-stream"
        attach_files.append(("files", fp.name, content, mime))

    # Build `data` object per spec.
    # v8.8.21: senderMailAddress(camelCase) 를 표준으로 쓰고 legacy 소문자 변형도 함께 주입.
    _sender_addr = (cfg.get("from_addr") or "").strip()
    data_obj: Dict[str, Any] = {
        "content":           html_body,
        "receiverList":      receiver_list,
        "senderMailAddress": _sender_addr,
        "senderMailaddress": _sender_addr,
        "statusCode":        (req.status_code or cfg.get("status_code") or "").strip(),
        "title":             subject,
    }
    # Merge admin extra_data without clobbering reserved keys.
    extra = cfg.get("extra_data") or {}
    if isinstance(extra, dict):
        for k, v in extra.items():
            if k and k not in data_obj:
                data_obj[k] = v

    headers = {}
    cfg_headers = cfg.get("headers") or {}
    if isinstance(cfg_headers, dict):
        for k, v in cfg_headers.items():
            if k:
                headers[str(k)] = str(v)

    url = cfg.get("api_url").strip()
    dry_run = url.lower() == "dry-run"
    # v9.0.0 (Q4 확정): 사내 API — data = {"mailSendString": "<JSON of data_obj>"} 단일 래핑.
    #   camelCase `mailSendString` 한 키로 통일.
    mail_send_string = _json.dumps(data_obj, ensure_ascii=False)
    wrapped = {"mailSendString": mail_send_string}
    if dry_run:
        result_info = {
            "status": 200, "dry_run": True,
            "preview_data": data_obj,
            "preview_data_wrapped": wrapped,
            "preview_attachments": [{"name": f[1], "bytes": len(f[2])} for f in attach_files],
            "preview_headers": headers,
        }
    else:
        # multipart/form-data top-level form field: "mailSendString" 만.
        fields = {"mailSendString": mail_send_string}
        body_bytes, content_type = _encode_multipart(fields, attach_files)
        hdrs_out = dict(headers)
        hdrs_out["Content-Type"] = content_type
        try:
            r = urllib.request.Request(url, data=body_bytes, headers=hdrs_out, method="POST")
            with urllib.request.urlopen(r, timeout=15) as resp:
                status = resp.status
                text = resp.read(2048).decode("utf-8", errors="replace")
            result_info = {"status": status, "response": text[:512]}
        except urllib.error.HTTPError as e:
            detail_text = ""
            try:
                detail_text = e.read(512).decode("utf-8", errors="replace")
            except Exception:
                pass
            _audit(request, "inform:mail-fail", detail=f"id={inform_id} http={e.code}", tab="inform")
            raise HTTPException(502, f"메일 API 오류: HTTP {e.code} {detail_text[:200]}")
        except Exception as e:
            _audit(request, "inform:mail-fail", detail=f"id={inform_id} err={e}", tab="inform")
            raise HTTPException(502, f"메일 전송 실패: {e}")

    # Best-effort audit log on the inform itself.
    hist = target.get("mail_history") or []
    hist.append({
        "at": _now(),
        "by": me.get("username", ""),
        "to": to_list,
        "to_users": list(req.to_users or []),
        "groups": list(req.groups or []),
        "auto_module": target.get("module", "") if used_auto_module else "",
        "auto_module_recipients": auto_module_recipients,
        "subject": subject,
    })
    target["mail_history"] = hist[-20:]  # keep last 20
    mail_at = hist[-1]["at"]
    prev_status = _canonical_flow_status(target.get("flow_status"), target)
    if prev_status != "apply_confirmed":
        target["flow_status"] = "mail_completed"
        status_hist = target.get("status_history") or []
        if not status_hist or _canonical_flow_status(status_hist[-1].get("status")) != "mail_completed":
            status_hist.append({
                "status": "mail_completed",
                "prev": prev_status,
                "actor": me.get("username", ""),
                "at": mail_at,
                "note": "mail sent",
            })
            target["status_history"] = status_hist
    _save(items)
    _audit_record(request, "mail", target,
                  {"to_count": len(to_list), "subject": subject, "dry_run": dry_run},
                  f"메일 · {subject}", at=mail_at)
    return {
        "ok": True,
        "to": to_list,
        "subject": subject,
        "auto_module_recipients": auto_module_recipients,
        "auto_module_used": used_auto_module,
        **result_info,
    }


@router.post("/splittable")
def attach_splittable(req: SplitChange, request: Request, id: str = Query(...)):
    """해당 인폼에 SplitTable 변경요청 메타 attach (작성자/담당자/admin)."""
    me = current_user(request)
    my_mods = _effective_modules(me["username"], me.get("role", "user"))
    items = _load_upgraded()
    target = _find(items, id)
    if not target:
        raise HTTPException(404)
    if _is_deleted(target):
        raise HTTPException(404)
    if not _can_moderate(target, me["username"], me.get("role", "user"), my_mods):
        raise HTTPException(403)
    target["splittable_change"] = {
        "column": (req.column or "").strip(),
        "old_value": (req.old_value or "").strip(),
        "new_value": (req.new_value or "").strip(),
        "applied": bool(req.applied),
    }
    _save(items)
    _audit_record(request, "edit", target,
                  {"field": "splittable_change", "value": target["splittable_change"]},
                  f"SplitTable 변경요청 · {target['splittable_change']['column']}")
    return {"ok": True, "inform": target}
