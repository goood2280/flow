class PivotRefreshReq(BaseModel):
    product: str
    username: str = ""


@router.post("/cache/pivot/refresh")
def refresh_pivot_cache(req: PivotRefreshReq, _perm=Depends(require_page_manager("splittable"))):
    """수동 pivot cache 재빌드 트리거. 빌드는 백그라운드에서 돌고 완료 시
    view payload cache 를 비워 다음 조회부터 최신 데이터가 보인다."""
    queued = _enqueue_pivot_cache_build(req.product, reason="manual_refresh", immediate=True)
    return {
        "ok": True,
        "queued": queued,
        "state": _pivot_cache_build_state(req.product),
    }


@router.get("/cache/pivot/status")
def pivot_cache_status(product: str = Query(...), username: str = Query("")):
    canonical = _canonical_mltable_product_name(product, allow_bare=True) or str(product or "").strip().upper()
    cache_dir = _pivot_cache_path(canonical, "_probe").parent
    files = 0
    latest_mtime = 0.0
    try:
        if cache_dir.exists():
            for fp_ in cache_dir.glob("*.parquet"):
                files += 1
                latest_mtime = max(latest_mtime, fp_.stat().st_mtime)
    except Exception:
        pass
    return {
        "product": canonical,
        "state": _pivot_cache_build_state(canonical),
        "files": files,
        "last_built": datetime.datetime.fromtimestamp(latest_mtime).isoformat(timespec="seconds") if latest_mtime else None,
    }


# ── Plans ──
class PlanReq(BaseModel):
    product: str
    plans: dict
    username: str = "unknown"
    root_lot_id: str = ""
    # 선택 입력 — 왜 바꿨는지. 비워 두는 게 기본이고, 남기면 이력에 그대로 붙는다.
    reason: str = ""


@router.post("/plan")
def save_plan(req: PlanReq, request: Request = None):
    if request is not None:
        try:
            me = current_user(request)
            req.username = me.get("username") or req.username or "unknown"
        except Exception:
            raise
    # Validate: only KNOB/MASK/FAB columns can have plans
    rejected = []
    for ck in list(req.plans.keys()):
        col_name = ck.split("|")[-1] if "|" in ck else ck
        col_upper = col_name.upper()
        if not any(col_upper.startswith(p + "_") for p in PLAN_ALLOWED_PREFIXES):
            rejected.append(col_name)
            del req.plans[ck]
    if rejected and not req.plans:
        raise HTTPException(400, f"Plan not allowed for: {', '.join(rejected)}. Only {'/'.join(PLAN_ALLOWED_PREFIXES)} columns.")

    pf = _plan_history_path(req.product)
    data = _load_plan_data(req.product)
    data.setdefault("history", [])
    now = datetime.datetime.now().isoformat()
    changed_entries = []
    # v8.8.33: my_plan_changed 이벤트 대상자 수집.
    #   같은 cell 에 과거 plan 이 있었으면 그 plan 을 만든 user 에게 "내 plan 이 변경됨" 알림.
    original_owners: dict[str, str] = {}
    for ck in req.plans.keys():
        prev_user = (data["plans"].get(ck) or {}).get("user")
        if prev_user:
            original_owners[ck] = prev_user
    prior_history = list(data["history"])
    batch_id = _new_history_batch_id()
    reason = _clean_plan_reason(req.reason)
    new_history: list[dict] = []
    s0_bases = _knob_current_s0_for_product(req.product, list({
        _split_plan_cell_key(ck)[2] for ck in req.plans
    }))
    for ck, val in req.plans.items():
        old = data["plans"].get(ck, {}).get("value")
        prev_owner = original_owners.get(ck) or ""
        data["plans"][ck] = {"value": val, "user": req.username, "updated": now}
        cell_root, cell_wafer, cell_column = _split_plan_cell_key(ck)
        s0_basis = s0_bases.get(cell_column)
        if s0_basis:
            data["plans"][ck]["s0_basis"] = {**s0_basis, "captured_at": now}
        entry = {
            "cell": ck, "old": old, "new": val, "user": req.username,
            "time": now, "action": "set", "root_lot_id": req.root_lot_id or cell_root,
            # 아래는 화면 필터·CSV 를 위해 풀어 둔 것 — cell key 파싱을 소비자마다
            # 반복하지 않게 한다. 예전 엔트리는 없으므로 읽는 쪽이 폴백을 갖는다.
            "wafer_id": cell_wafer, "column": cell_column,
            "batch": batch_id, "batch_size": len(req.plans),
            "prev_user": prev_owner, "reason": reason,
        }
        if s0_basis:
            entry["s0_basis"] = {**s0_basis, "captured_at": now}
        data["history"].append(entry)
        new_history.append(entry)
        changed_entries.append((ck, old, val))
    data["history"] = data["history"][-1000:]
    save_json(pf, data)
    _archive_plan_history(req.product, new_history, prior_history)
    _invalidate_plan_risk_cache(req.product)

    # v9.1.x: plan 저장은 여기서 즉시 완료 — actual 대조(셀당 파케이 스캔)·knowledge
    # 적재·알림은 백그라운드로 옮겨 저장 응답 지연을 없앤다 (가장 사용 빈도 높은 경로).
    product = req.product
    username = req.username
    root_lot_id_req = req.root_lot_id

    def _plan_post_save():
        try:
            for ck, old, val in changed_entries:
                _append_splittable_plan_knowledge(
                    product=product,
                    cell_key=ck,
                    old=old,
                    new=val,
                    actor=username,
                    changed_at=now,
                    conflicting=bool(old not in (None, "") and old != val),
                )
            save_mismatches = []
            for ck, _old, val in changed_entries:
                actual = _actual_value_for_plan_cell(product, ck)
                if not _plan_actual_mismatch(val, actual):
                    continue
                root, wafer, column = _split_plan_cell_key(ck)
                save_mismatches.append({
                    "key": ck,
                    "plan": val,
                    "actual": actual,
                    "plan_user": username,
                    "plan_updated": now,
                    "root_lot_id": root,
                    "wafer_id": wafer,
                    "column": column,
                })
            _notify_plan_actual_mismatches_once(product, save_mismatches, actor="flow")
            # v8.8.33: notify 이벤트 — 본인이 아닌 원 소유자 및 관심랏 등록 유저에게 전달.
            try:
                from core.notify import emit_event
                from core import watchlist as _wl
                notified_users = set()
                notified_watch_lots = set()

                for ck, old, val in changed_entries:
                    if old == val:
                        continue
                    parts = (ck or "").split("|")
                    r_lot = root_lot_id_req or (parts[0] if parts else "")
                    target = original_owners.get(ck)
                    if target and target != username and (target, r_lot) not in notified_users:
                        notified_users.add((target, r_lot))
                        emit_event(
                            "my_plan_changed",
                            actor=username,
                            target_user=target,
                            title="[plan 변경]",
                            body=f"{username} 가 {product}/{parts[0] if parts else ''} plan 을 변경",
                            payload={
                                "product": product,
                                "cell": ck,
                                "root_lot_id": r_lot,
                                "wafer_id": parts[1] if len(parts) > 1 else "",
                                "column": parts[2] if len(parts) > 2 else "",
                                "old": old, "new": val,
                            },
                        )

                    # 관심랏 등록 유저에게 알림
                    if r_lot:
                        watchers = _wl.get_users_watching_lot(r_lot)
                        is_plan_add = not bool(str(old or "").strip()) and bool(str(val or "").strip())
                        is_plan_del = bool(str(old or "").strip()) and not bool(str(val or "").strip())
                        badge_name = "Plan 추가" if is_plan_add else ("Plan 삭제" if is_plan_del else "Plan 변경")
                        action_text = "추가" if is_plan_add else ("삭제" if is_plan_del else "변경")
                        for watcher in watchers:
                            if (watcher, r_lot) not in notified_watch_lots:
                                notified_watch_lots.add((watcher, r_lot))
                                emit_event(
                                    "watched_lot_split_changed",
                                    actor=username,
                                    target_user=watcher,
                                    title=f"[관심랏 {badge_name}] {r_lot}",
                                    body=f"{username} 님이 {product}/{r_lot} 의 Split Plan 을 {action_text}했습니다.",
                                    payload={
                                        "product": product,
                                        "root_lot_id": r_lot,
                                        "category": "관심랏",
                                        "badge": badge_name,
                                        "column": parts[2] if len(parts) > 2 else "",
                                        "allow_self": True,
                                    },
                                    allow_self=True,
                                )
            except Exception:
                pass
        except Exception as exc:
            logger.warning(f"plan post-save background work failed for {product}: {exc}")

    global _PLAN_POST_SAVE_LAST_THREAD
    _PLAN_POST_SAVE_LAST_THREAD = threading.Thread(target=_plan_post_save, daemon=True, name="splittable-plan-postsave")
    _PLAN_POST_SAVE_LAST_THREAD.start()
    # Plan saves stay in SplitTable history/notifications only; Inform snapshots
    # are attached explicitly from Inform so users do not get extra auto cards.
    _audit_user(req.username, "splittable:plan_save",
                detail=f"product={req.product} saved={len(req.plans)} rejected={len(rejected)}",
                tab="splittable")
    return {"ok": True, "saved": len(req.plans), "rejected": rejected}


class PlanDeleteReq(BaseModel):
    product: str
    cell_keys: list
    username: str = "unknown"
    reason: str = ""


@router.post("/plan/delete")
def delete_plan(req: PlanDeleteReq, request: Request = None):
    if request is not None:
        try:
            me = current_user(request)
            req.username = me.get("username") or req.username or "unknown"
        except Exception:
            raise
    pf = _plan_history_path(req.product)
    if not any(p.exists() for p in _plan_alias_paths(req.product)):
        raise HTTPException(404)
    data = _load_plan_data(req.product)
    now = datetime.datetime.now().isoformat()
    deleted = []
    prior_history = list(data.get("history") or [])
    batch_id = _new_history_batch_id()
    reason = _clean_plan_reason(req.reason)
    new_history: list[dict] = []
    for ck in req.cell_keys:
        if ck in data.get("plans", {}):
            old = data["plans"][ck].get("value")
            prev_owner = str((data["plans"][ck] or {}).get("user") or "")
            del data["plans"][ck]
            cell_root, cell_wafer, cell_column = _split_plan_cell_key(ck)
            entry = {
                "cell": ck, "old": old, "new": None,
                "user": req.username, "time": now, "action": "delete",
                # v10: delete 엔트리에도 root/wafer/column 을 남긴다 — 예전에는
                # root_lot_id 가 없어 cell prefix 매칭에만 의존했다.
                "root_lot_id": cell_root, "wafer_id": cell_wafer,
                "column": cell_column, "batch": batch_id,
                "batch_size": len(req.cell_keys), "prev_user": prev_owner,
                "reason": reason,
            }
            data.setdefault("history", []).append(entry)
            new_history.append(entry)
            deleted.append((ck, old))
    data["history"] = data.get("history", [])[-1000:]
    save_json(pf, data)
    _archive_plan_history(req.product, new_history, prior_history)
    _invalidate_plan_risk_cache(req.product)

    # v9.1.x: knowledge 적재는 백그라운드 — 삭제 응답도 즉시 반환.
    product = req.product
    username = req.username

    def _plan_delete_post():
        try:
            for ck, old in deleted:
                _append_splittable_plan_knowledge(
                    product=product,
                    cell_key=ck,
                    old=old,
                    new=None,
                    actor=username,
                    changed_at=now,
                    conflicting=bool(old not in (None, "")),
                )
        except Exception as exc:
            logger.warning(f"plan delete post work failed for {product}: {exc}")

    global _PLAN_POST_SAVE_LAST_THREAD
    _PLAN_POST_SAVE_LAST_THREAD = threading.Thread(target=_plan_delete_post, daemon=True, name="splittable-plan-delpost")
    _PLAN_POST_SAVE_LAST_THREAD.start()
    # SplitTable plan deletes stay in SplitTable history/notifications only.
    _audit_user(req.username, "splittable:plan_delete",
                detail=f"product={req.product} deleted={len(deleted)}",
                tab="splittable")
    return {"ok": True}


@router.get("/history")
def get_history(product: str = Query(...), root_lot_id: str = Query(""),
                limit: int = Query(500), offset: int = Query(0),
                user: str = Query(""), action: str = Query(""),
                column: str = Query(""), wafer_id: str = Query(""),
                q: str = Query(""), since: str = Query(""), until: str = Query(""),
                has_reason: bool = Query(False)):
    """plan 변경 이력. 응답의 `history` 는 예전과 같이 **시간 오름차순**이고
    기본 동작(필터 없음)도 예전과 같다 — 필터/페이지 필드만 추가됐다.

    `offset` 은 최신 쪽에서부터 건너뛰는 개수다 (offset=0 이 가장 최근 묶음).
    """
    empty_facets = {"users": [], "actions": [], "columns": []}
    if (not any(p.exists() for p in _plan_alias_paths(product))
            and not _plan_history_log_path(product).exists()):
        return {"history": [], "total": 0, "scope_total": 0,
                "returned": 0, "has_more": False, "facets": empty_facets}
    hist = _plan_history_entries(product)
    scoped = [h for h in hist if _history_matches_root(h, root_lot_id)]
    facets = _plan_history_facets(scoped)
    filtered = [h for h in scoped
                if _history_matches_filters(h, user=user, action=action,
                                            column=column, wafer_id=wafer_id,
                                            q=q, since=since, until=until,
                                            has_reason=bool(has_reason))]
    take = max(1, int(limit or 500))
    skip = max(0, int(offset or 0))
    end = max(0, len(filtered) - skip)
    start = max(0, end - take)
    page = filtered[start:end]
    return {"history": page, "total": len(filtered), "scope_total": len(scoped),
            "returned": len(page), "has_more": start > 0, "facets": facets}


@router.get("/operational-history")
def get_operational_history(request: Request, product: str = Query(...),
                            root_lot_id: str = Query(""), wafer_ids: str = Query("")):
    me = current_user(request)
    items = _load_operational_history(
        product=product,
        root_lot_id=root_lot_id,
        wafer_ids=wafer_ids,
        username=me.get("username", ""),
        role=me.get("role", "user"),
    )
    return {"items": items, "total": len(items)}


@router.get("/history/final")
def get_history_final(request: Request, product: str = Query(...), root_lot_id: str = Query(""),
                      include_deleted: bool = Query(False)):
    # v8.8.33 보안: 세션 토큰 필수 (plan history 내 username 노출 방지).
    from core.auth import current_user
    _ = current_user(request)
    """v8.8.33: final-plan-only 뷰.
    각 cell 의 최종 상태(가장 최근 set 또는 delete)만 반환 + plan drift 경고.

    drift 판정:
      - 같은 cell 에 set 이 2회 이상이고 old != new 가 섞임 → drift_level="multi"
      - 서로 다른 user 가 set → drift_level="multi_user"
      - 둘 다 → "multi_user_multi_change"
    """
    payload = _get_plan_risk_payload(product, include_deleted=include_deleted)
    return _copy_plan_risk_payload(payload, root_lot_id=root_lot_id)


@router.get("/history-csv")
def download_history_csv(product: str = Query(...), root_lot_id: str = Query(""),
                         user: str = Query(""), action: str = Query(""),
                         column: str = Query(""), wafer_id: str = Query(""),
                         q: str = Query(""), since: str = Query(""),
                         until: str = Query(""), has_reason: bool = Query(False)):
    """Admin: 화면과 같은 필터가 걸린 전체 이력을 CSV 로. 인자가 없으면 전량."""
    if (not any(p.exists() for p in _plan_alias_paths(product))
            and not _plan_history_log_path(product).exists()):
        raise HTTPException(404, "No history")
    hist = [h for h in _plan_history_entries(product)
            if _history_matches_root(h, root_lot_id)
            and _history_matches_filters(h, user=user, action=action,
                                         column=column, wafer_id=wafer_id,
                                         q=q, since=since, until=until,
                                         has_reason=bool(has_reason))]
    if not hist:
        raise HTTPException(404, "No history entries")

    header = ["time", "user", "action", "root_lot_id", "wafer_id",
              "column", "old_value", "new_value", "reason", "prev_plan_user", "batch"]

    def _rows():
        for h in hist:
            lot, wf, col = _history_cell_parts(h)
            yield [h.get("time", ""), h.get("user", ""), h.get("action", ""),
                   lot, wf, col, h.get("old", ""), h.get("new", ""),
                   h.get("reason", ""), h.get("prev_user", ""), h.get("batch", "")]

    return csv_response(csv_writer_bytes(header, _rows()), f"{product}_history.csv")


def _log_split_table_download(username: str, product: str, root_lot_id: str,
                              prefix: str, custom_name: str, fmt: str,
                              rows: int, cols: int, size_bytes: int,
                              selected: list[str] | None = None) -> None:
    """SplitTable CSV/XLSX 내보내기를 downloads.jsonl(관리자 다운로드 모니터)에 기록."""
    try:
        from core.utils import jsonl_append
        scope = f"custom={custom_name}" if custom_name else f"prefix={prefix or 'all'}"
        sel = selected or []
        jsonl_append(PATHS.download_log, {
            "source": "splittable",
            "username": username or "",
            "product": product,
            "sql": f"root_lot_id={root_lot_id or 'all'}, {scope}, format={fmt}",
            "rows": int(rows or 0), "cols": int(cols or 0),
            "select_cols": ",".join(sel[:8]) + ("…" if len(sel) > 8 else ""),
            "size_mb": round((size_bytes or 0) / 1e6, 2),
        })
    except Exception:
        pass


# ── Transposed CSV ──
@router.get("/download-csv")
def download_csv(product: str = Query(...), root_lot_id: str = Query(""),
                 wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
                 custom_name: str = Query(""), transposed: str = Query("true"),
                 username: str = Query(""),
                 custom_cols: str = Query(""),
                 step_labels: str = Query(""),
                 exclude_not_null: str = Query("1")):
    fp = _product_path(product)
    lf = _scan_product(product, root_lot_id=root_lot_id, wafer_ids=wafer_ids)
    lot_col, wf_col = _detect_lot_wafer(lf)
    lf = _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id, wafer_ids)
    df = lf.collect()

    all_data_cols = _view_data_columns(df.columns, lot_col, wf_col)
    tag_labels = _custom_tag_label_map(product)
    tag_labels[DEFAULT_CUSTOM_TAG_COLUMN] = DEFAULT_CUSTOM_TAG_LABEL
    for tag_col in tag_labels:
        if tag_col not in all_data_cols:
            all_data_cols.append(tag_col)
    management_labels = _management_row_label_map(product)
    if custom_name or custom_cols:
        for mgmt_col in management_labels:
            if mgmt_col not in all_data_cols:
                all_data_cols.append(mgmt_col)
    selected = _select_columns(all_data_cols, custom_name, prefix,
                               max_fallback=200, custom_cols=custom_cols)
    selected = _with_default_custom_tag(selected)
    if not custom_name and not custom_cols:
        for raw_pref in [p.strip() for p in str(prefix or "").split(",") if p.strip()]:
            for virt in _virtual_columns_for_prefix(product, raw_pref, existing_columns=selected):
                if virt not in selected:
                    selected.append(virt)
    # v8.8.14: display rename (rule_order + step_desc) 적용.
    # 정렬은 view와 동일 — prefix 구분 없이 parameter별 step_id 공정 순서 우선.
    col_rename = _build_col_rename_map(selected, product)
    col_rename.update({col: f"{CUSTOM_TAG_PREFIX}_{label}" for col, label in tag_labels.items()})
    col_rename.update({col: label for col, label in management_labels.items()})
    _dl_step_rank = _split_step_order_context(product).get("param_rank") or {}
    selected = sorted(selected, key=lambda c: _step_order_sort_key(c, col_rename.get(c, c), _dl_step_rank))
    # 적용 공정 정보는 항목명을 바꾸지 않고 왼쪽 step_id / step_desc 열로 내보낸다.
    step_label_mode = _truthy_value(step_labels)
    process_columns = _build_step_process_columns(
        product, selected, exclude_not_null=_truthy_value(exclude_not_null)
    ) if step_label_mode else {}

    if transposed.lower() == "true" and wf_col and wf_col in df.columns:
        # Resolve wafer values (handle W01 format)
        wf_raw_int = df[wf_col].cast(pl.Int64, strict=False).to_list()
        non_null = [v for v in wf_raw_int if v is not None]
        if non_null:
            wf_vals = wf_raw_int
        else:
            wf_vals = [str(v) for v in df[wf_col].to_list()]
        # v8.4.4: fab_lot_id 로 1차 정렬, wafer 로 2차 정렬 — UI 그룹 순서와 일치
        fab_col = "fab_lot_id" if "fab_lot_id" in df.columns else None
        wf2fab: dict = {}
        if fab_col:
            fab_vals = [(None if v is None else str(v)) for v in df[fab_col].to_list()]
            for w, f in zip(wf_vals, fab_vals):
                if w is None: continue
                if w not in wf2fab and f and f not in ("None","null"):
                    wf2fab[w] = f
        wf_uniq = [w for w in dict.fromkeys(wf_vals) if w is not None and w != "None" and w != "null"]
        # v8.8.3: fab_lot 그룹 → wafer_id 숫자-aware 정렬 (view 와 동일 로직).
        def _wf_sort_key2(w):
            primary = wf2fab.get(w, "~")
            try:
                return (primary, 0, int(w))
            except (TypeError, ValueError):
                s = str(w)
                if s.upper().startswith("W"):
                    try:
                        return (primary, 0, int(s[1:]))
                    except ValueError:
                        pass
                return (primary, 1, s)
        wf_sorted = sorted(wf_uniq, key=_wf_sort_key2)
        headers = [f"#{v}" for v in wf_sorted]
        fab_row = [wf2fab.get(w, "") for w in wf_sorted]
        wf_idx = {v: i for i, v in enumerate(wf_sorted)}

        plans = _load_plan_data(product).get("plans", {})
        tag_values = _custom_tag_values_for_root(product, root_lot_id)
        management_values = _management_row_values_for_root(product, root_lot_id)

        output = io.StringIO()
        writer = csv_mod.writer(output)
        # Header rows (v8.4.4b): downloaded_at, username, root_lot_id, fab_lot_id, Parameter
        download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        writer.writerow(["downloaded_at", download_ts])
        writer.writerow(["username", username or ""])
        writer.writerow(["root_lot_id", root_lot_id or ""])
        if fab_col:
            writer.writerow(["fab_lot_id"] + fab_row)
        writer.writerow((["step_id", "step_desc"] if step_label_mode else []) + ["Parameter"] + headers)
        for col_name in selected:
            row_data = [""] * len(wf_sorted)
            if col_name in tag_labels:
                for idx, wk in enumerate(wf_sorted):
                    row_data[idx] = tag_values.get(f"{root_lot_id}|{wk}|{col_name}", "")
            elif col_name in management_labels:
                for idx, wk in enumerate(wf_sorted):
                    row_data[idx] = management_values.get(f"{root_lot_id}|{wk}|{col_name}", "")
            elif col_name in df.columns:
                vals = df[col_name].to_list()
                for i, v in enumerate(vals):
                    wk = wf_vals[i] if i < len(wf_vals) else None
                    idx = wf_idx.get(wk)
                    if idx is not None:
                        sv = str(v) if v is not None and str(v) not in ("None", "null") else ""
                        ck = f"{root_lot_id}|{wk}|{col_name}"
                        pv = plans.get(ck, {}).get("value")
                        row_data[idx] = pv if pv and not sv else sv
            else:
                for idx, wk in enumerate(wf_sorted):
                    ck = f"{root_lot_id}|{wk}|{col_name}"
                    pv = plans.get(ck, {}).get("value")
                    row_data[idx] = "" if pv is None else str(pv)
            process = process_columns.get(str(col_name), {})
            writer.writerow(
                ([process.get("step_id", ""), process.get("step_desc", "")] if step_label_mode else [])
                + [col_rename.get(col_name, col_name)] + row_data
            )
        # v8.4.4: Excel 한글 깨짐 방지 — UTF-8 BOM prefix
        csv_bytes = b"\xef\xbb\xbf" + output.getvalue().encode("utf-8")
        _log_rows, _log_cols = len(selected), len(wf_sorted)
    else:
        csv_bytes = b"\xef\xbb\xbf" + df.write_csv().encode("utf-8")
        _log_rows, _log_cols = df.height, df.width

    _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                              "csv", _log_rows, _log_cols, len(csv_bytes), selected)
    return csv_response(csv_bytes, f"{product}_{root_lot_id or 'all'}.csv")


SPLIT_CHECK_XLSX_PREFIX_COLUMNS = ["항목", "값", "Split"]


def _export_has_value(value: Any) -> bool:
    text = "" if value is None else str(value).strip()
    return bool(text and text not in {"None", "null"})


def _split_check_export_supported(selected: list[str]) -> bool:
    for column in selected or []:
        up = str(column or "").strip().upper()
        if up in {"INLINE", "VM"} or up.startswith("INLINE_") or up.startswith("VM_"):
            return False
    return True


def _build_split_check_export_rows(
    selected: list[str],
    wafer_count: int,
    value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
    col_rename: dict[str, str] | None = None,
    s0_by_param: dict[str, str] | None = None,
) -> list[list[str]]:
    rows: list[list[str]] = []
    rename = col_rename or {}
    s0_values = s0_by_param or {}
    for column in selected or []:
        display_name = str(rename.get(column, column) or column)
        actual_by_idx, plan_by_idx = value_maps.get(column, ({}, {}))
        values_by_idx: dict[int, str] = {}
        order: list[str] = []
        seen: set[str] = set()
        preferred = str(s0_values.get(column) or "").strip()
        if preferred:
            seen.add(preferred)
            order.append(preferred)
        for idx in range(max(0, int(wafer_count or 0))):
            plan_value = plan_by_idx.get(idx, "")
            actual_value = actual_by_idx.get(idx, "")
            value = plan_value if _export_has_value(plan_value) else actual_value
            if not _export_has_value(value):
                continue
            text = str(value)
            values_by_idx[idx] = text
            if text not in seen:
                seen.add(text)
                order.append(text)
        for split_idx, value in enumerate(order):
            label = f"S{split_idx}"
            checks = ["✓" if values_by_idx.get(idx) == value else "" for idx in range(max(0, int(wafer_count or 0)))]
            rows.append([display_name, value, label, *checks])
    return rows


def _build_pems_export_rows(
    selected: list[str],
    value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
    col_rename: dict[str, str] | None = None,
    s0_by_param: dict[str, str] | None = None,
) -> tuple[list[list[str]], list[str]]:
    """Build the browser PEMS matrix for physical wafers 1..25.

    Every wafer belongs to one S group per parameter. Empty/missing wafer
    values are deliberately assigned to S0 so a Chrome extension can consume
    the sheet without having to infer omitted physical wafer columns.
    """
    rows: list[list[str]] = []
    param_keys: list[str] = []
    rename = col_rename or {}
    s0_values = s0_by_param or {}
    wafer_count = 25
    for column in selected or []:
        import re as _re
        raw_display_name = str(rename.get(column, column) or column)
        display_name = _re.sub(r"^[A-Za-z]+_", "", raw_display_name)
        if str(column or "").upper().startswith("KNOB_"):
            display_name = _re.sub(r"_Split$", "", display_name, flags=_re.I)
        display_name = display_name.strip() or raw_display_name
        actual_by_idx, plan_by_idx = value_maps.get(column, ({}, {}))
        values_by_idx: dict[int, str] = {}
        order: list[str] = []
        seen: set[str] = set()
        preferred = str(s0_values.get(column) or "").strip()
        if preferred:
            seen.add(preferred)
            order.append(preferred)
        for idx in range(wafer_count):
            plan_value = plan_by_idx.get(idx, "")
            actual_value = actual_by_idx.get(idx, "")
            value = plan_value if _export_has_value(plan_value) else actual_value
            if not _export_has_value(value):
                continue
            text = str(value)
            values_by_idx[idx] = text
            if text not in seen:
                seen.add(text)
                order.append(text)
        if not order:
            order.append("")
        for split_idx, value in enumerate(order):
            label = f"S{split_idx}"
            marks = [
                label if (values_by_idx.get(idx, "") == value or (split_idx == 0 and idx not in values_by_idx)) else ""
                for idx in range(wafer_count)
            ]
            rows.append([display_name, value, label, *marks])
            param_keys.append(str(column))
    return rows, param_keys


def _split_check_export_param_keys(
    selected: list[str],
    wafer_count: int,
    value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
    s0_by_param: dict[str, str] | None = None,
) -> list[str]:
    """Raw parameter key for every row emitted by _build_split_check_export_rows."""
    keys: list[str] = []
    s0_values = s0_by_param or {}
    for column in selected or []:
        actual_by_idx, plan_by_idx = value_maps.get(column, ({}, {}))
        seen: set[str] = set()
        preferred = str(s0_values.get(column) or "").strip()
        if preferred:
            seen.add(preferred)
            keys.append(str(column))
        for idx in range(max(0, int(wafer_count or 0))):
            plan_value = plan_by_idx.get(idx, "")
            actual_value = actual_by_idx.get(idx, "")
            value = plan_value if _export_has_value(plan_value) else actual_value
            if not _export_has_value(value):
                continue
            text = str(value)
            if text in seen:
                continue
            seen.add(text)
            keys.append(str(column))
    return keys


def _split_check_param_merges(rows: list[list[str]], start_row: int,
                              parameter_col: int = 1) -> list[tuple[int, int, int, int]]:
    merges: list[tuple[int, int, int, int]] = []
    current = ""
    run_start = 0
    for idx, row in enumerate([*(rows or []), ["__flow_end__"]]):
        param_idx = max(0, int(parameter_col or 1) - 1)
        param = str(row[param_idx] if row and len(row) > param_idx else "")
        if idx == 0:
            current = param
            run_start = 0
            continue
        if param == current:
            continue
        if current and idx - run_start > 1:
            for col in range(1, param_idx + 2):
                merges.append((start_row + run_start, col, start_row + idx - 1, col))
        current = param
        run_start = idx
    return merges


def _step_progress_not_reached_cells(step_progress: dict, selected: list[str],
                                     wafer_keys: list) -> set[tuple[str, int]]:
    """Return (raw parameter, wafer index) cells shaded as not reached.

    This is the export counterpart of the browser's per-wafer latest-step
    shading.  Prefer by_wafer whenever it exists; root-level metadata is only
    a compatibility fallback.
    """
    progress = step_progress if isinstance(step_progress, dict) else {}
    by_wafer_raw = progress.get("by_wafer") if isinstance(progress.get("by_wafer"), dict) else {}
    by_wafer = {
        _normalize_wafer_id(wafer): {str(v) for v in (meta.get("not_reached") or [])}
        for wafer, meta in by_wafer_raw.items() if isinstance(meta, dict)
    }
    root_set = {str(v) for v in (progress.get("not_reached") or [])}
    out: set[tuple[str, int]] = set()
    for idx, wafer in enumerate(wafer_keys or []):
        cell_set = by_wafer.get(_normalize_wafer_id(wafer), set()) if by_wafer else root_set
        for param in selected or []:
            if str(param) in cell_set:
                out.add((str(param), idx))
    return out


def _exclude_populated_not_reached_cells(
        not_reached_cells: set[tuple[str, int]],
        value_maps: dict[str, tuple[dict[int, str], dict[int, str]]],
) -> set[tuple[str, int]]:
    """Grey only cells that are both not reached and empty.

    ML_TABLE can already carry KNOB/INLINE/VM values even when the root is not
    present in FAB.  A visible actual or plan value is stronger evidence for
    the cell display and must keep its normal value/palette styling.
    """
    out: set[tuple[str, int]] = set()
    for raw_param, wafer_idx in not_reached_cells or set():
        actual_by_idx, plan_by_idx = value_maps.get(str(raw_param), ({}, {}))
        if _export_has_value(actual_by_idx.get(wafer_idx)):
            continue
        if _export_has_value(plan_by_idx.get(wafer_idx)):
            continue
        out.add((str(raw_param), wafer_idx))
    return out


@router.get("/download-xlsx")
def download_xlsx(product: str = Query(...), root_lot_id: str = Query(""),
                  wafer_ids: str = Query(""), prefix: str = Query("KNOB"),
                  custom_name: str = Query(""), username: str = Query(""),
                  custom_cols: str = Query(""),
                  display_mode: str = Query(""),
                  step_labels: str = Query(""),
                  exclude_not_null: str = Query("1")):
    """v8.4.4 — XLSX 내보내기. fab_lot_id 행이 동일 값 구간별로 셀 병합되어
    UI 의 그룹 헤더와 동일하게 표시.
    v8.8.33: custom_cols 추가 — save 없이 체크만 한 ad-hoc 컬럼.
    v8.8.34: display_mode=split_check 이면 화면의 Split 체크 표시 행 형식으로 export.
    PEMS: root lot 전용 1..25 고정 열 + S0/S1 직접 표기 형식으로 export.
    """
    openpyxl_error = None
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except Exception as e:
        openpyxl_error = e

    requested_display_mode = str(display_mode or "").strip().lower()
    pems_requested = requested_display_mode == "pems"
    if pems_requested and not str(root_lot_id or "").strip():
        raise HTTPException(400, "PEMS export requires root_lot_id")
    # PEMS는 물리 wafer 1..25 전체가 계약이므로 전달된 wafer 필터를 무시한다.
    effective_wafer_ids = "" if pems_requested else wafer_ids
    lf = _scan_product(product, root_lot_id=root_lot_id, wafer_ids=effective_wafer_ids)
    lot_col, wf_col = _detect_lot_wafer(lf, product)
    lf = _filter_lot_wafer(lf, lot_col, wf_col, root_lot_id, effective_wafer_ids)
    df = lf.collect()

    all_data_cols = _view_data_columns(df.columns, lot_col, wf_col)
    tag_labels = _custom_tag_label_map(product)
    tag_labels[DEFAULT_CUSTOM_TAG_COLUMN] = DEFAULT_CUSTOM_TAG_LABEL
    for tag_col in tag_labels:
        if tag_col not in all_data_cols:
            all_data_cols.append(tag_col)
    management_labels = _management_row_label_map(product)
    if custom_name or custom_cols:
        for mgmt_col in management_labels:
            if mgmt_col not in all_data_cols:
                all_data_cols.append(mgmt_col)
    selected = _select_columns(all_data_cols, custom_name, prefix,
                               max_fallback=200, custom_cols=custom_cols)
    selected = _with_default_custom_tag(selected)
    # v8.8.14: display rename (rule_order + step_desc) 적용.
    # 정렬은 view와 동일 — prefix 구분 없이 parameter별 step_id 공정 순서 우선.
    col_rename = _build_col_rename_map(selected, product)
    col_rename.update({col: f"{CUSTOM_TAG_PREFIX}_{label}" for col, label in tag_labels.items()})
    col_rename.update({col: label for col, label in management_labels.items()})
    _dl_step_rank = _split_step_order_context(product).get("param_rank") or {}
    selected = sorted(selected, key=lambda c: _step_order_sort_key(c, col_rename.get(c, c), _dl_step_rank))
    # 화면과 동일하게 기존 항목은 보존하고 별도 step_id / step_desc 열을 만든다.
    step_label_mode = _truthy_value(step_labels)
    process_columns = _build_step_process_columns(
        product, selected, exclude_not_null=_truthy_value(exclude_not_null)
    ) if step_label_mode else {}

    wf_raw_int = df[wf_col].cast(pl.Int64, strict=False).to_list() if wf_col else []
    non_null = [v for v in wf_raw_int if v is not None]
    if non_null:
        wf_vals = wf_raw_int
    else:
        wf_vals = [str(v) for v in df[wf_col].to_list()] if wf_col else []
    fab_col = "fab_lot_id" if "fab_lot_id" in df.columns else None
    wf2fab: dict = {}
    if fab_col:
        fab_vals = [(None if v is None else str(v)) for v in df[fab_col].to_list()]
        for w, f in zip(wf_vals, fab_vals):
            if w is None: continue
            if w not in wf2fab and f and f not in ("None","null"):
                wf2fab[w] = f
    wf_uniq = [w for w in dict.fromkeys(wf_vals) if w is not None and w != "None" and w != "null"]

    pems_mode = pems_requested and _split_check_export_supported(selected)
    if pems_requested and not pems_mode:
        raise HTTPException(400, "PEMS export does not support INLINE/VM columns")
    if pems_mode:
        wf_sorted = list(range(1, 26))
        wf_idx = {str(w): i for i, w in enumerate(wf_sorted)}
    else:
        wf_sorted = sorted(wf_uniq, key=lambda w: (wf2fab.get(w, "~"), w))
        wf_idx = {v: i for i, v in enumerate(wf_sorted)}

    plans = _load_plan_data(product).get("plans", {})
    tag_values = _custom_tag_values_for_root(product, root_lot_id)
    tag_colors = _custom_tag_colors_for_root(product, root_lot_id)
    management_values = _management_row_values_for_root(product, root_lot_id)
    split_check_mode = (
        requested_display_mode == "split_check"
        and _split_check_export_supported(selected)
    )
    # v9.1.x: 제3 표시형식 — 행에서 왼쪽 값과 같은 칸을 셀 병합해 export (UI 병합 표시와 동일).
    merged_mode = (
        requested_display_mode == "merged"
        and not split_check_mode and not pems_mode
    )

    def _export_wafer_index(wafer) -> int | None:
        if pems_mode:
            normalized = _normalize_wafer_id(wafer)
            return wf_idx.get(normalized) if normalized else None
        return wf_idx.get(wafer)

    def _xlsx_value_maps_for_col(col_name: str) -> tuple[dict[int, str], dict[int, str]]:
        actual_by_idx: dict[int, str] = {}
        plan_by_idx: dict[int, str] = {}
        if col_name in tag_labels:
            for idx, wk in enumerate(wf_sorted):
                tv = tag_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if _export_has_value(tv):
                    actual_by_idx[idx] = str(tv)
        elif col_name in management_labels:
            for idx, wk in enumerate(wf_sorted):
                mv = management_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if _export_has_value(mv):
                    actual_by_idx[idx] = str(mv)
        elif col_name in df.columns:
            vals = df[col_name].to_list()
            for i, v in enumerate(vals):
                wk = wf_vals[i] if i < len(wf_vals) else None
                idx = _export_wafer_index(wk)
                if idx is None:
                    continue
                sv = str(v) if _export_has_value(v) else ""
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if _export_has_value(sv):
                    actual_by_idx[idx] = sv
                if _export_has_value(pv):
                    plan_by_idx[idx] = str(pv)
        else:
            for idx, wk in enumerate(wf_sorted):
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if _export_has_value(pv):
                    plan_by_idx[idx] = str(pv)
        return actual_by_idx, plan_by_idx

    if wf2fab:
        fab_present: bool | None = True
    else:
        fab_scope = _fab_history_scope(product, root_lot_id=root_lot_id, limit=1)
        fab_present = (
            True if fab_scope.get("candidates")
            else False if fab_scope.get("query_ok")
            else None
        )
    step_progress = _split_step_progress(
        product, root_lot_id, selected, wf_sorted, fab_present=fab_present)
    not_reached_cells = _step_progress_not_reached_cells(step_progress, selected, wf_sorted)
    value_maps = {col_name: _xlsx_value_maps_for_col(col_name) for col_name in selected}
    s0_by_param = {
        column: str(meta.get("ppid") or "")
        for column, meta in _knob_s0_for_root(product, root_lot_id, selected).items()
        if isinstance(meta, dict) and str(meta.get("ppid") or "").strip()
    }
    pems_missing_wafer_indices: set[int] = set()
    if pems_mode:
        source_wafers = {_normalize_wafer_id(w) for w in wf_uniq}
        source_wafers.discard("")
        pems_missing_wafer_indices = {idx for idx in range(25) if str(idx + 1) not in source_wafers}
        # 웹 PEMS와 동일: 실제 wafer가 없거나 해당 항목 값이 비어 있으면 S0로
        # 표기하면서 회색을 유지한다. step_progress 미진행 셀도 값 유무와 무관하게 회색이다.
        for raw_param in selected:
            actual_by_idx, plan_by_idx = value_maps.get(raw_param, ({}, {}))
            for idx in range(25):
                wafer = str(idx + 1)
                has_value = _export_has_value(plan_by_idx.get(idx)) or _export_has_value(actual_by_idx.get(idx))
                if wafer not in source_wafers or not has_value:
                    not_reached_cells.add((str(raw_param), idx))
    else:
        not_reached_cells = _exclude_populated_not_reached_cells(not_reached_cells, value_maps)
    split_check_rows: list[list[str]] = []
    split_check_param_keys: list[str] = []
    if pems_mode:
        split_check_rows, split_check_param_keys = _build_pems_export_rows(
            selected,
            value_maps,
            col_rename,
            s0_by_param,
        )
    elif split_check_mode:
        split_check_rows = _build_split_check_export_rows(
            selected,
            len(wf_sorted),
            value_maps,
            col_rename,
            s0_by_param,
        )
        split_check_param_keys = _split_check_export_param_keys(
            selected, len(wf_sorted), value_maps, s0_by_param,
        )
    if step_label_mode and split_check_rows:
        split_check_rows = [
            [
                process_columns.get(str(param), {}).get("step_id", ""),
                process_columns.get(str(param), {}).get("step_desc", ""),
                *row,
            ]
            for row, param in zip(split_check_rows, split_check_param_keys)
        ]
    split_like_mode = split_check_mode or pems_mode
    export_prefix_columns = (["step_id", "step_desc"] if step_label_mode else []) + SPLIT_CHECK_XLSX_PREFIX_COLUMNS
    parameter_prefix_col = 3 if step_label_mode else 1
    regular_prefix_columns = (["step_id", "step_desc"] if step_label_mode else []) + ["Parameter"]
    regular_prefix_count = len(regular_prefix_columns)

    if openpyxl_error is not None:
        try:
            from core.simple_xlsx import build_workbook
            from fastapi.responses import StreamingResponse
        except Exception as e:
            import sys
            raise HTTPException(
                500,
                f"XLSX export unavailable at {sys.executable}: openpyxl={openpyxl_error}; fallback={e}",
            )

        download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")
        n_wafers = len(wf_sorted)
        prefix_count = len(export_prefix_columns) if split_like_mode else (3 if step_label_mode else 1)
        last_col = max(prefix_count + n_wafers, prefix_count + 1) if split_like_mode else prefix_count + n_wafers
        rows = [["downloaded_at", download_ts], ["username", username or ""]]
        merges = []

        if split_like_mode:
            root_row = ["root_lot_id", *([""] * (prefix_count - 1)), root_lot_id or "", *["" for _ in range(max(0, n_wafers - 1))]]
            rows.append(root_row)
            merges.append((3, 1, 3, prefix_count))
            if n_wafers > 1:
                merges.append((3, prefix_count + 1, 3, last_col))

            has_fab_row = bool(not pems_mode and fab_col and wf_sorted)
            if has_fab_row:
                fab_row = ["fab_lot_id", *([""] * (prefix_count - 1)), *["" for _ in wf_sorted]]
                cur = None
                start = 0
                row_no = len(rows) + 1
                merges.append((row_no, 1, row_no, prefix_count))
                for i, w in enumerate(wf_sorted):
                    f = wf2fab.get(w, "")
                    if f != cur:
                        if cur is not None and i - start > 0:
                            fab_row[prefix_count + start] = cur
                            if i - start > 1:
                                merges.append((row_no, prefix_count + 1 + start, row_no, prefix_count + i))
                        cur = f
                        start = i
                if cur is not None and len(wf_sorted) - start > 0:
                    fab_row[prefix_count + start] = cur
                    if len(wf_sorted) - start > 1:
                        merges.append((row_no, prefix_count + 1 + start, row_no, prefix_count + len(wf_sorted)))
                rows.append(fab_row)

            header_row_no = len(rows) + 1
            wafer_headers = [str(w) if pems_mode else f"#{w}" for w in wf_sorted]
            rows.append([*export_prefix_columns, *wafer_headers])
            data_start_row = header_row_no + 1
            rows.extend(split_check_rows)
            merges.extend(_split_check_param_merges(split_check_rows, data_start_row, parameter_prefix_col))
        else:
            rows.append(["root_lot_id", *([""] * (prefix_count - 1)), root_lot_id or "", *["" for _ in range(max(0, n_wafers - 1))]])
            merges.append((3, 1, 3, prefix_count))
            if n_wafers > 1:
                merges.append((3, prefix_count + 1, 3, last_col))

            has_fab_row = bool(fab_col and wf_sorted)
            if has_fab_row:
                fab_row = ["fab_lot_id", *([""] * (prefix_count - 1)), *["" for _ in wf_sorted]]
                merges.append((4, 1, 4, prefix_count))
                cur = None
                start = 0
                for i, w in enumerate(wf_sorted):
                    f = wf2fab.get(w, "")
                    if f != cur:
                        if cur is not None and i - start > 0:
                            fab_row[prefix_count + start] = cur
                            if i - start > 1:
                                merges.append((4, prefix_count + 1 + start, 4, prefix_count + i))
                        cur = f
                        start = i
                if cur is not None and len(wf_sorted) - start > 0:
                    fab_row[prefix_count + start] = cur
                    if len(wf_sorted) - start > 1:
                        merges.append((4, prefix_count + 1 + start, 4, prefix_count + len(wf_sorted)))
                rows.append(fab_row)

            regular_headers = (["step_id", "step_desc"] if step_label_mode else []) + ["Parameter"]
            rows.append([*regular_headers, *[f"#{w}" for w in wf_sorted]])
            for col_name in selected:
                display_name = col_rename.get(col_name, col_name)
                actual_by_idx, plan_by_idx = value_maps.get(col_name, ({}, {}))
                process = process_columns.get(str(col_name), {})
                out = ([process.get("step_id", ""), process.get("step_desc", "")] if step_label_mode else []) + [display_name, *["" for _ in wf_sorted]]
                for idx in sorted(set(list(actual_by_idx.keys()) + list(plan_by_idx.keys()))):
                    sv = actual_by_idx.get(idx, "")
                    pv = plan_by_idx.get(idx, "")
                    if sv and pv and sv != pv:
                        out[prefix_count + idx] = f"{sv} != {pv}"
                    elif pv and not sv:
                        out[prefix_count + idx] = f"PLAN: {pv}"
                    else:
                        out[prefix_count + idx] = sv or pv
                rows.append(out)
                if merged_mode and n_wafers > 1 and _merge_view_allowed_param(col_name):
                    row_no = len(rows)
                    start = 0
                    for j in range(1, n_wafers + 1):
                        if j == n_wafers or str(out[prefix_count + j]) != str(out[prefix_count + start]):
                            if j - start > 1:
                                merges.append((row_no, prefix_count + 1 + start, row_no, prefix_count + j))
                            start = j

        data = build_workbook([{"title": product[:31], "rows": rows, "merges": merges}])
        fmt_suffix = "_pems" if pems_mode else ("_split_check" if split_check_mode else ("_merged" if merged_mode else ""))
        fname = f"{product}_{root_lot_id or 'all'}{fmt_suffix}.xlsx"
        _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                                  f"xlsx{fmt_suffix}",
                                  len(split_check_rows) if split_like_mode else len(selected),
                                  len(wf_sorted), len(data), selected)
        return StreamingResponse(
            iter([data]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    wb = Workbook()
    ws = wb.active
    ws.title = product[:31]
    hdr_fill = PatternFill("solid", fgColor="1f2937")
    fab_fill = PatternFill("solid", fgColor="374151")
    param_fill = PatternFill("solid", fgColor="374151")
    not_reached_fill = PatternFill("solid", fgColor="9CA3AF")
    white = Font(color="FFFFFF", bold=True)
    # fab_lot_id 헤더는 어두운 배경 + 흰 글자로 고정해 노란색 대비 문제를 피한다.
    fab_font = Font(color="FFFFFF", bold=True, name="Consolas", size=12)
    center = Alignment(horizontal="center", vertical="center")
    thin = Side(style="thin", color="555555")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    download_ts = datetime.datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    def _outline_span(row: int, col_start: int, span: int, box: Side) -> None:
        """병합 구간 전체에 테두리 '상자'를 그린다.

        openpyxl 은 값/채우기는 병합 범위의 좌상단 셀만 봐도 되지만 **테두리는
        셀마다 따로 그려진다.** 앵커 셀에만 넣으면 상자가 첫 칸에서 끊기고
        나머지 칸은 아래 그리드 보강 패스가 기본 테두리로 덮어써서, 병합
        다운로드에서 mismatch 빨간 테두리가 반 칸만 나오거나 아예 안 보였다.
        구간 안쪽 세로선은 병합되면 어차피 렌더되지 않으므로 기본 thin 을 둔다."""
        span = max(1, int(span or 1))
        for k in range(span):
            ws.cell(row=row, column=col_start + k).border = Border(
                left=box if k == 0 else thin,
                right=box if k == span - 1 else thin,
                top=box, bottom=box,
            )

    if split_like_mode:
        prefix_count = len(export_prefix_columns)
        n_wafers = len(wf_sorted)
        first_wafer_col = prefix_count + 1
        last_col = max(prefix_count + n_wafers, first_wafer_col)
        prefix_fill = PatternFill("solid", fgColor="F9FAFB")
        mark_font = Font(color="000000", bold=True, name="Consolas", size=11)
        prefix_font = Font(color="000000", bold=True, name="Consolas", size=11)
        value_font = Font(color="000000", name="Consolas", size=11)
        palette = [
            ("C6EFCE", "000000"),
            ("FFEB9C", "000000"),
            ("FBE5D6", "000000"),
            ("BDD7EE", "000000"),
            ("E2BFEE", "000000"),
            ("B4DED4", "000000"),
            ("F4CCCC", "000000"),
        ]

        def _split_fill(label: str):
            import re as _re
            m = _re.fullmatch(r"S(\d+)", str(label or "").strip(), flags=_re.I)
            if not m:
                return None
            bg, _fg = palette[int(m.group(1)) % len(palette)]
            return PatternFill("solid", fgColor=bg)

        def _style_cell(cell, *, fill=None, font=None, alignment=None):
            if fill is not None:
                cell.fill = fill
            if font is not None:
                cell.font = font
            if alignment is not None:
                cell.alignment = alignment
            cell.border = border

        ws.cell(row=1, column=1, value="downloaded_at")
        _style_cell(ws.cell(row=1, column=1), fill=hdr_fill, font=white)
        ws.cell(row=1, column=2, value=download_ts)
        ws.cell(row=2, column=1, value="username")
        _style_cell(ws.cell(row=2, column=1), fill=hdr_fill, font=white)
        ws.cell(row=2, column=2, value=username or "")

        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=prefix_count)
        _style_cell(ws.cell(row=3, column=1, value="root_lot_id"), fill=hdr_fill, font=white, alignment=center)
        root_value_col = first_wafer_col
        ws.cell(row=3, column=root_value_col, value=root_lot_id or "")
        if n_wafers > 1:
            ws.merge_cells(start_row=3, start_column=first_wafer_col, end_row=3, end_column=prefix_count + n_wafers)
        for col_idx in range(root_value_col, (prefix_count + n_wafers if n_wafers else root_value_col) + 1):
            _style_cell(ws.cell(row=3, column=col_idx), fill=hdr_fill, font=Font(color="FBBF24", bold=True, name="Consolas", size=13), alignment=center)

        has_fab_row = bool(not pems_mode and fab_col and wf_sorted)
        header_row = 5 if has_fab_row else 4
        if has_fab_row:
            ws.merge_cells(start_row=4, start_column=1, end_row=4, end_column=prefix_count)
            _style_cell(ws.cell(row=4, column=1, value="fab_lot_id"), fill=hdr_fill, font=white, alignment=center)
            cur = None
            start = 0
            for i, w in enumerate(wf_sorted):
                f = wf2fab.get(w, "")
                if f != cur:
                    if cur is not None and i - start > 0:
                        c = ws.cell(row=4, column=first_wafer_col + start, value=cur)
                        _style_cell(c, fill=fab_fill, font=fab_font, alignment=center)
                        if i - start > 1:
                            ws.merge_cells(start_row=4, start_column=first_wafer_col + start, end_row=4, end_column=first_wafer_col + i - 1)
                    cur = f
                    start = i
            if cur is not None and len(wf_sorted) - start > 0:
                c = ws.cell(row=4, column=first_wafer_col + start, value=cur)
                _style_cell(c, fill=fab_fill, font=fab_font, alignment=center)
                if len(wf_sorted) - start > 1:
                    ws.merge_cells(start_row=4, start_column=first_wafer_col + start, end_row=4, end_column=first_wafer_col + len(wf_sorted) - 1)

        for i, label in enumerate(export_prefix_columns, start=1):
            c = ws.cell(row=header_row, column=i, value=label)
            _style_cell(c, fill=param_fill, font=white, alignment=center)
        for i, w in enumerate(wf_sorted):
            c = ws.cell(row=header_row, column=first_wafer_col + i, value=str(w) if pems_mode else f"#{w}")
            _style_cell(c, fill=not_reached_fill if i in pems_missing_wafer_indices else param_fill, font=white, alignment=center)

        data_start = header_row + 1
        for r_idx, row in enumerate(split_check_rows, start=data_start):
            raw_param = split_check_param_keys[r_idx - data_start] if r_idx - data_start < len(split_check_param_keys) else ""
            all_not_reached = bool(wf_sorted) and all((raw_param, idx) in not_reached_cells for idx in range(len(wf_sorted)))
            label_idx = parameter_prefix_col + 1
            label = str(row[label_idx] if len(row) > label_idx else "")
            fill = _split_fill(label)
            for c_idx, value in enumerate(row, start=1):
                cell = ws.cell(row=r_idx, column=c_idx, value=value)
                if c_idx <= prefix_count:
                    split_col = parameter_prefix_col + 2
                    prefix_bg = not_reached_fill if all_not_reached else (fill if c_idx == split_col and fill else prefix_fill)
                    # 공정 두 열과 항목 칸은 여러 줄을 보존한다.
                    left_top = Alignment(horizontal="left", vertical="top",
                                         wrap_text=bool(step_label_mode and c_idx <= parameter_prefix_col))
                    _style_cell(cell, fill=prefix_bg, font=(mark_font if c_idx == split_col else prefix_font), alignment=center if c_idx == split_col else left_top)
                else:
                    wafer_idx = c_idx - first_wafer_col
                    mark_fill = not_reached_fill if (raw_param, wafer_idx) in not_reached_cells else (fill if value else None)
                    _style_cell(cell, fill=mark_fill, font=mark_font if value else value_font, alignment=center)
        for r1, c1, r2, c2 in _split_check_param_merges(split_check_rows, data_start, parameter_prefix_col):
            ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)
            ws.cell(row=r1, column=c1).alignment = Alignment(horizontal="left", vertical="top")

        for idx, label in enumerate(export_prefix_columns, start=1):
            width = 22 if label == "step_id" else 24 if label == "step_desc" else 28 if label == "항목" else 18 if label == "값" else 10
            ws.column_dimensions[get_column_letter(idx)].width = width
        for i in range(len(wf_sorted)):
            ws.column_dimensions[get_column_letter(first_wafer_col + i)].width = 12
        ws.freeze_panes = f"{get_column_letter(first_wafer_col)}{data_start}"

        last_row = header_row + len(split_check_rows)
        for row_cells in ws.iter_rows(min_row=1, max_row=max(last_row, header_row), min_col=1, max_col=last_col):
            for c in row_cells:
                b = c.border
                if not (b and b.left and b.left.style):
                    c.border = border

        buf = io.BytesIO()
        wb.save(buf)
        buf.seek(0)

        from fastapi.responses import StreamingResponse
        fmt_name = "pems" if pems_mode else "split_check"
        fname = f"{product}_{root_lot_id or 'all'}_{fmt_name}.xlsx"
        _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                                  f"xlsx_{fmt_name}", len(split_check_rows),
                                  len(wf_sorted), buf.getbuffer().nbytes, selected)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            headers={"Content-Disposition": f'attachment; filename="{fname}"'},
        )

    n_wafers = len(wf_sorted)
    first_wafer_col = regular_prefix_count + 1
    last_col = regular_prefix_count + n_wafers
    # v8.4.4c — downloaded_at / username: 병합하지 않고 label+value 2칸만 표시
    c_ts = ws.cell(row=1, column=1, value="downloaded_at"); c_ts.font = white; c_ts.fill = hdr_fill
    ws.cell(row=1, column=2, value=download_ts)
    # username
    c1 = ws.cell(row=2, column=1, value="username"); c1.font = white; c1.fill = hdr_fill
    ws.cell(row=2, column=2, value=username or "")
    # root_lot_id (v8.4.5c — 병합 복원: wafer 컬럼 전체 colspan)
    c2 = ws.cell(row=3, column=1, value="root_lot_id"); c2.font = white; c2.fill = hdr_fill
    if regular_prefix_count > 1:
        ws.merge_cells(start_row=3, start_column=1, end_row=3, end_column=regular_prefix_count)
    c2v = ws.cell(row=3, column=first_wafer_col, value=root_lot_id or "")
    c2v.alignment = center; c2v.fill = hdr_fill
    c2v.font = Font(color="fbbf24", bold=True, name="Consolas", size=13)
    if n_wafers > 1:
        ws.merge_cells(start_row=3, start_column=first_wafer_col, end_row=3, end_column=last_col)
    # Row 4: fab_lot_id (merged by contiguous groups)
    FAB_ROW = 4
    if fab_col and wf_sorted:
        ws.cell(row=FAB_ROW, column=1, value="fab_lot_id").font = white
        ws.cell(row=FAB_ROW, column=1).fill = hdr_fill
        if regular_prefix_count > 1:
            ws.merge_cells(start_row=FAB_ROW, start_column=1, end_row=FAB_ROW, end_column=regular_prefix_count)
        cur = None; start = 0
        for i, w in enumerate(wf_sorted):
            f = wf2fab.get(w, "")
            if f != cur:
                if cur is not None and i - start > 0:
                    c = ws.cell(row=FAB_ROW, column=first_wafer_col+start, value=cur)
                    c.font = fab_font; c.fill = fab_fill; c.alignment = center; c.border = border
                    if i - start > 1:
                        ws.merge_cells(start_row=FAB_ROW, start_column=first_wafer_col+start, end_row=FAB_ROW, end_column=first_wafer_col+i-1)
                cur = f; start = i
        if cur is not None and len(wf_sorted) - start > 0:
            c = ws.cell(row=FAB_ROW, column=first_wafer_col+start, value=cur)
            c.font = fab_font; c.fill = fab_fill; c.alignment = center; c.border = border
            if len(wf_sorted) - start > 1:
                ws.merge_cells(start_row=FAB_ROW, start_column=first_wafer_col+start,
                               end_row=FAB_ROW, end_column=first_wafer_col+len(wf_sorted)-1)

    # Row 5: Parameter | #1 #2 ...
    param_row = 5 if fab_col else 4
    for prefix_idx, prefix_label in enumerate(regular_prefix_columns, start=1):
        prefix_cell = ws.cell(row=param_row, column=prefix_idx, value=prefix_label)
        prefix_cell.font = white
        prefix_cell.fill = param_fill
        prefix_cell.border = border
    for i, w in enumerate(wf_sorted):
        c = ws.cell(row=param_row, column=first_wafer_col+i, value=f"#{w}")
        c.font = white; c.fill = param_fill; c.alignment = center; c.border = border

    # v8.4.4c: UI 와 동일한 7-색 팔레트 (CELL_COLORS). KNOB_ / MASK_ prefix 행만 컬러링.
    CELL_PALETTE = [
        ("C6EFCE", "006100"),  # green
        ("FFEB9C", "9C5700"),  # yellow
        ("FBE5D6", "BF4E00"),  # orange
        ("BDD7EE", "1F4E79"),  # blue
        ("E2BFEE", "7030A0"),  # purple
        ("B4DED4", "0B5345"),  # teal
        ("F4CCCC", "75194C"),  # pink
    ]
    COLOR_PREFIXES = ("KNOB_", "MASK_")

    for r_off, col_name in enumerate(selected):
        rr = param_row + 1 + r_off
        # v8.8.14: display rename 된 이름을 표기 (원본 col_name 으로는 여전히 df 조회).
        display_name = col_rename.get(col_name, col_name)
        process = process_columns.get(str(col_name), {})
        if step_label_mode:
            for process_idx, key in enumerate(("step_id", "step_desc"), start=1):
                process_cell = ws.cell(row=rr, column=process_idx, value=process.get(key, ""))
                process_cell.alignment = Alignment(horizontal="left", vertical="top", wrap_text=True)
                process_cell.border = border
        label_cell = ws.cell(row=rr, column=regular_prefix_count, value=display_name)
        label_cell.font = Font(bold=True)
        if wf_sorted and all((str(col_name), idx) in not_reached_cells for idx in range(len(wf_sorted))):
            label_cell.fill = not_reached_fill
        up = (col_name or "").upper()
        should_color = any(up.startswith(p) for p in COLOR_PREFIXES)
        vals = df[col_name].to_list() if col_name in df.columns else []
        # Build unique-value map — include plan values in palette assignment
        row_values_ordered = []  # preserve column order for uniq index
        actual_by_idx = {}
        plan_by_idx = {}
        if col_name in tag_labels:
            for idx, wk in enumerate(wf_sorted):
                tv = tag_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if tv:
                    actual_by_idx[idx] = str(tv)
        elif col_name in management_labels:
            for idx, wk in enumerate(wf_sorted):
                mv = management_values.get(f"{root_lot_id}|{wk}|{col_name}")
                if mv:
                    actual_by_idx[idx] = str(mv)
        else:
            for i, v in enumerate(vals):
                wk = wf_vals[i] if i < len(wf_vals) else None
                idx = wf_idx.get(wk)
                if idx is None: continue
                sv = str(v) if v is not None and str(v) not in ("None","null") else ""
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if sv: actual_by_idx[idx] = sv
                if pv: plan_by_idx[idx] = str(pv)
        if col_name not in df.columns and col_name not in tag_labels and col_name not in management_labels:
            for idx, wk in enumerate(wf_sorted):
                ck = f"{root_lot_id}|{wk}|{col_name}"
                pv = plans.get(ck, {}).get("value")
                if pv:
                    plan_by_idx[idx] = str(pv)
        for idx in sorted(set(list(actual_by_idx.keys()) + list(plan_by_idx.keys()))):
            if idx in actual_by_idx: row_values_ordered.append(actual_by_idx[idx])
            elif idx in plan_by_idx: row_values_ordered.append(plan_by_idx[idx])
        uniq_vals = list(dict.fromkeys(row_values_ordered))
        uniq_map = {v: i for i, v in enumerate(uniq_vals)}

        # v8.4.5b: plan 전용 — 진한 주황 테두리 4면 + 이탤릭
        orange_side = Side(style="medium", color="ea580c")
        plan_border = Border(left=orange_side, right=orange_side,
                             top=orange_side, bottom=orange_side)
        red_side = Side(style="medium", color="ef4444")
        mismatch_border = Border(left=red_side, right=red_side,
                                 top=red_side, bottom=red_side)
        if merged_mode and _merge_view_allowed_param(col_name):
            # v9.1.x: 병합 표시 형식 — 왼쪽 칸과 같은 값이면 연속 구간을 셀 병합.
            # KNOB/FAB/MASK 만 대상이다. INLINE/VM/TAG 는 아래 일반 경로로 떨어진다.
            groups = []
            for idx in range(len(wf_sorted)):
                sv = actual_by_idx.get(idx, "")
                pv = plan_by_idx.get(idx, "")
                cell_val = sv or pv
                cell_not_reached = (str(col_name), idx) in not_reached_cells
                if groups and cell_val == groups[-1]["val"] and cell_not_reached == groups[-1]["not_reached"]:
                    groups[-1]["span"] += 1
                else:
                    groups.append({"val": cell_val, "sv": sv, "pv": pv, "start": idx, "span": 1, "not_reached": cell_not_reached})
            for g in groups:
                idx = g["start"]; sv = g["sv"]; pv = g["pv"]; cell_val = g["val"]
                if not cell_val and g["span"] == 1 and not g["not_reached"]:
                    continue
                is_plan_only = (not sv) and bool(pv)
                is_mismatch = bool(sv) and bool(pv) and sv != pv
                cell = ws.cell(row=rr, column=first_wafer_col + idx, value=cell_val)
                cell.alignment = center
                cell.border = border
                if should_color and cell_val and cell_val in uniq_map:
                    bg, fg = CELL_PALETTE[uniq_map[cell_val] % len(CELL_PALETTE)]
                    cell.fill = PatternFill("solid", fgColor=bg)
                    cell.font = Font(color=fg, bold=True, italic=is_plan_only, size=11, name="Consolas")
                elif is_plan_only:
                    cell.fill = PatternFill("solid", fgColor="fef3c7")
                    cell.font = Font(color="ea580c", bold=True, italic=True, name="Consolas")
                if g["not_reached"]:
                    cell.fill = not_reached_fill
                if is_plan_only:
                    _outline_span(rr, first_wafer_col + idx, g["span"], orange_side)
                    if cell_val and not str(cell_val).startswith("📌 "):
                        cell.value = "📌 " + str(cell_val)
                elif is_mismatch:
                    # 계획값과 실제값이 다른 셀 — 병합 구간 전체를 빨간 상자로.
                    _outline_span(rr, first_wafer_col + idx, g["span"], red_side)
                if g["span"] > 1:
                    ws.merge_cells(start_row=rr, start_column=first_wafer_col + idx,
                                   end_row=rr, end_column=first_wafer_col + idx + g["span"] - 1)
            continue
        # 값이 없는 미진행 셀도 실제 cell 객체를 만들어야 회색 fill이 보인다.
        # 값이 있는 셀은 위에서 not_reached_cells에서 제외됐으므로 팔레트/plan
        # 스타일을 그대로 유지한다.
        for idx in range(len(wf_sorted)):
            sv = actual_by_idx.get(idx, "")
            pv = plan_by_idx.get(idx, "")
            cell_val = sv or pv
            is_plan_only = (not sv) and bool(pv)
            is_mismatch = bool(sv) and bool(pv) and sv != pv
            cell = ws.cell(row=rr, column=first_wafer_col+idx, value=cell_val)
            cell.alignment = center
            cell.border = border
            if should_color and cell_val in uniq_map:
                bg, fg = CELL_PALETTE[uniq_map[cell_val] % len(CELL_PALETTE)]
                cell.fill = PatternFill("solid", fgColor=bg)
                if is_plan_only:
                    cell.font = Font(color=fg, italic=True, bold=True, size=11, name="Consolas")
                else:
                    cell.font = Font(color=fg, bold=True, size=11, name="Consolas")
            elif is_plan_only:
                cell.fill = PatternFill("solid", fgColor="fef3c7")
                cell.font = Font(color="ea580c", bold=True, italic=True, name="Consolas")
            if col_name in tag_labels:
                custom_color = tag_colors.get(f"{root_lot_id}|{wf_sorted[idx]}|{col_name}")
                if custom_color:
                    cell.fill = PatternFill("solid", fgColor=custom_color.lstrip("#").upper())
            if (str(col_name), idx) in not_reached_cells:
                cell.fill = not_reached_fill
            # Plan-only: 진한 주황 테두리 4면 — 눈에 확 띄도록
            if is_plan_only:
                cell.border = plan_border
                # 📌 prefix 접두로 plan 임을 한 번 더 명시
                if not str(cell_val).startswith("📌 "):
                    cell.value = "📌 " + str(cell_val)
            elif is_mismatch:
                cell.border = mismatch_border

    # Column widths
    for prefix_idx, prefix_label in enumerate(regular_prefix_columns, start=1):
        width = 22 if prefix_label == "step_id" else 24 if prefix_label == "step_desc" else 28
        ws.column_dimensions[get_column_letter(prefix_idx)].width = width
    for i in range(len(wf_sorted)):
        ws.column_dimensions[get_column_letter(first_wafer_col+i)].width = 14

    # Freeze panes at param_row+1, B
    ws.freeze_panes = f"{get_column_letter(first_wafer_col)}{param_row+1}"

    # v8.8.13: 전체 그리드 테두리 보강 — 값 없는 빈 셀·헤더 셀까지 기본 border 적용.
    # plan_border / mismatch_border 처럼 특수 스타일이 이미 들어간 셀은 건너뜀.
    last_row = param_row + len(selected)
    for row_cells in ws.iter_rows(min_row=1, max_row=last_row, min_col=1, max_col=last_col):
        for c in row_cells:
            b = c.border
            if not (b and b.left and b.left.style):
                c.border = border

    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)

    from fastapi.responses import StreamingResponse
    fname = f"{product}_{root_lot_id or 'all'}{'_merged' if merged_mode else ''}.xlsx"
    _log_split_table_download(username, product, root_lot_id, prefix, custom_name,
                              "xlsx_merged" if merged_mode else "xlsx", len(selected),
                              len(wf_sorted), buf.getbuffer().nbytes, selected)
    return StreamingResponse(
        iter([buf.getvalue()]),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{fname}"'},
    )


@router.get("/plans-csv")
def export_plans_csv(product: str = Query(...)):
    if not any(p.exists() for p in _plan_alias_paths(product)):
        raise HTTPException(404, "No plans")
    plans = _load_plan_data(product).get("plans", {})
    if not plans:
        raise HTTPException(404, "No plans saved")

    header = ["root_lot_id", "wafer_id", "column", "plan_value", "user", "updated"]

    def _rows():
        for cell_key, info in plans.items():
            parts = cell_key.split("|")
            lot = parts[0] if len(parts) > 0 else ""
            wf = parts[1] if len(parts) > 1 else ""
            col = parts[2] if len(parts) > 2 else cell_key
            yield [lot, wf, col, info.get("value", ""),
                   info.get("user", ""), info.get("updated", "")]

    return csv_response(csv_writer_bytes(header, _rows()), f"{product}_plans.csv")
