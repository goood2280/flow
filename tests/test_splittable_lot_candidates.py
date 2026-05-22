from __future__ import annotations

import asyncio
import io
import json
import sys
from pathlib import Path

import polars as pl
import pytest
from fastapi import HTTPException

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))
if str(ROOT / "backend") not in sys.path:
    sys.path.insert(0, str(ROOT / "backend"))

from core import auth as auth_core  # noqa: E402
from routers import informs, splittable  # noqa: E402


def _read_streaming_response(response) -> bytes:
    async def read_body():
        chunks = []
        async for chunk in response.body_iterator:
            chunks.append(chunk)
        return b"".join(chunks)

    return asyncio.run(read_body())


class _State:
    def __init__(self, user: dict):
        self.user = user


class _Request:
    headers = {}

    def __init__(self, username: str = "alice", role: str = "user"):
        self.state = _State({"username": username, "role": role})


def test_root_lot_candidates_prefer_renderable_mltable_roots(tmp_path, monkeypatch):
    monkeypatch.setattr(splittable, "MATCH_CACHE_DIR", tmp_path / "match_cache")
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()

    result = splittable.get_lot_candidates(
        product="ML_TABLE_PRODA",
        col="root_lot_id",
        prefix="A10",
        limit=20,
        source="auto",
        root_lot_id="",
    )

    assert result["source"] == "mltable"
    assert result["fab_source"] == "lot_progress_latest_cache"
    assert "A1000" in result["candidates"]
    assert "A0001" not in result["candidates"]


def test_operational_history_matches_saved_full_inform_root(tmp_path, monkeypatch):
    informs_file = tmp_path / "informs.json"
    tracker_file = tmp_path / "issues.json"
    informs_file.write_text(json.dumps([{
        "id": "inf_1",
        "root_lot_id": "LOT029AA",
        "lot_id": "",
        "wafer_id": "7",
        "product": "PRODA",
        "module": "KNOB",
        "reason": "PEMS",
        "text": "plan saved",
        "author": "tester",
        "created_at": "2026-04-28T10:00:00",
        "flow_status": "received",
        "group_ids": [],
    }]), encoding="utf-8")
    tracker_file.write_text("[]", encoding="utf-8")
    monkeypatch.setattr(splittable, "INFORMS_FILE", informs_file)
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tracker_file)

    items = splittable._load_operational_history(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT029AA",
        wafer_ids="",
        username="tester",
        role="admin",
    )

    assert len(items) == 1
    assert items[0]["source"] == "inform"
    assert items[0]["detail"] == "plan saved"


def test_save_plan_does_not_auto_log_inform(tmp_path, monkeypatch):
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "_audit_user", lambda *_args, **_kwargs: None)
    monkeypatch.setattr(splittable, "_resolve_fab_lot_for_cell", lambda *_args, **_kwargs: "FAB1000.1")

    calls = []

    def fake_auto_log(*args, **kwargs):
        calls.append((args, kwargs))

    monkeypatch.setattr(informs, "auto_log_splittable_change", fake_auto_log)

    result = splittable.save_plan(splittable.PlanReq(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        username="tester",
        plans={"A1000|1|KNOB_GATE": "R2"},
    ))

    assert result == {"ok": True, "saved": 1, "rejected": []}
    saved = json.loads((plan_dir / "ML_TABLE_PRODA.json").read_text(encoding="utf-8"))
    assert saved["plans"]["A1000|1|KNOB_GATE"]["value"] == "R2"
    assert calls == []


def test_save_plan_uses_canonical_flow_data_plan_file(tmp_path, monkeypatch):
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "_audit_user", lambda *_args, **_kwargs: None)

    result = splittable.save_plan(splittable.PlanReq(
        product="PRODA",
        root_lot_id="A1000",
        username="tester",
        plans={"A1000|1|KNOB_GATE": "R2"},
    ))

    assert result["saved"] == 1
    assert not (plan_dir / "PRODA.json").exists()
    saved = splittable._load_plan_data("ML_TABLE_PRODA")
    assert saved["plans"]["A1000|1|KNOB_GATE"]["value"] == "R2"


def test_save_plan_notifies_owner_once_when_existing_actual_mismatches(tmp_path, monkeypatch):
    import core.notify as flow_notify

    pl.DataFrame({
        "root_lot_id": ["A1000"],
        "wafer_id": ["1"],
        "KNOB_GATE": ["R1"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "_audit_user", lambda *_args, **_kwargs: None)
    splittable._LOT_LOOKUP_CACHE.clear()

    events = []
    monkeypatch.setattr(flow_notify, "emit_event", lambda *args, **kwargs: events.append((args, kwargs)) or True)

    result = splittable.save_plan(splittable.PlanReq(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        username="plan_owner",
        plans={"A1000|1|KNOB_GATE": "R2"},
    ))
    assert result == {"ok": True, "saved": 1, "rejected": []}
    assert len(events) == 1
    assert events[0][0][0] == "my_plan_actual_mismatch"
    assert events[0][1]["target_user"] == "plan_owner"
    assert events[0][1]["actor"] == "flow"
    assert "WF1" in events[0][1]["body"]
    assert "[plan] R2 → [actual] R1" in events[0][1]["body"]
    assert events[0][1]["payload"]["actual"] == "R1"
    assert events[0][1]["payload"]["plan"] == "R2"

    splittable.save_plan(splittable.PlanReq(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        username="plan_owner",
        plans={"A1000|1|KNOB_GATE": "R2"},
    ))
    assert len(events) == 1


def test_view_plan_actual_mismatch_notification_is_deduped(tmp_path, monkeypatch):
    import core.notify as flow_notify

    pl.DataFrame({
        "root_lot_id": ["A1000"],
        "wafer_id": ["1"],
        "KNOB_GATE": ["R1"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    (plan_dir / "ML_TABLE_PRODA.json").write_text(json.dumps({
        "plans": {
            "A1000|1|KNOB_GATE": {
                "value": "R2",
                "user": "plan_owner",
                "updated": "2026-05-22T10:00:00",
            },
        },
        "history": [],
    }), encoding="utf-8")
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")
    splittable._LOT_LOOKUP_CACHE.clear()

    events = []
    monkeypatch.setattr(flow_notify, "emit_event", lambda *args, **kwargs: events.append((args, kwargs)) or True)

    first = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )
    second = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert first["mismatch_count"] == 1
    assert second["mismatch_count"] == 1
    assert len(events) == 1
    assert events[0][0][0] == "my_plan_actual_mismatch"
    assert events[0][1]["target_user"] == "plan_owner"
    assert "WF1" in events[0][1]["body"]
    saved = json.loads((plan_dir / "ML_TABLE_PRODA.json").read_text(encoding="utf-8"))
    assert len(saved["mismatch_alerts"]) == 1


def test_plan_actual_mismatch_notification_uses_warning_tone():
    import core.notify as flow_notify

    assert flow_notify._EVENT_META["my_plan_actual_mismatch"][1] == "warning"


def test_view_includes_related_tracker_issues_for_root_lot(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT900AA", "LOT900AA"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["A", "B"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    tracker_file = tmp_path / "issues.json"
    tracker_file.write_text(json.dumps([{
        "id": "iss_related",
        "title": "Related lot issue",
        "status": "in_progress",
        "category": "Monitor",
        "priority": "normal",
        "username": "owner",
        "updated_at": "2026-04-28T10:00:00",
        "group_ids": [],
        "lots": [{"product": "PRODA", "root_lot_id": "LOT900AA", "wafer_id": "1", "lot_id": "LOT900AA.1"}],
        "comments": [{"username": "owner", "text": "check", "replies": [{"username": "peer", "text": "ok"}]}],
    }, {
        "id": "iss_other",
        "title": "Other lot issue",
        "status": "in_progress",
        "category": "Monitor",
        "group_ids": [],
        "lots": [{"product": "PRODA", "root_lot_id": "LOT901AA"}],
        "comments": [],
    }]), encoding="utf-8")
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tracker_file)

    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT900AA",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert [x["id"] for x in result["related_issues"]] == ["iss_related"]
    assert result["related_issues"][0]["matched_wafers"] == ["1"]
    assert result["related_issues"][0]["comment_count"] == 2


def test_view_normalizes_wafer_ids_above_25(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT925AA", "LOT925AA", "LOT925AA"],
        "wafer_id": ["1", "1000", "0"],
        "KNOB_ALPHA": ["A", "MAPPED", "BAD"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")

    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT925AA",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert result["headers"] == ["#1", "#25"]
    row = result["rows"][0]
    assert row["_cells"]["0"]["actual"] == "A"
    assert row["_cells"]["1"]["actual"] == "MAPPED"
    assert all("1000" not in cell["key"] for cell in row["_cells"].values())


def test_custom_tag_columns_overlay_view_and_custom_set(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT926AA", "LOT926AA"],
        "wafer_id": ["1", "2"],
        "KNOB_ALPHA": ["A", "B"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")

    created = splittable.save_custom_tag_column(
        splittable.CustomTagColumnReq(product="ML_TABLE_PRODA", name="review_flag", username="owner")
    )
    column = created["column"]
    splittable.save_custom_tag_values(
        splittable.CustomTagValuesReq(
            product="ML_TABLE_PRODA",
            root_lot_id="LOT926AA",
            values={
                f"LOT926AA|1|{column}": "hold",
                f"LOT926AA|2|{column}": "pass",
            },
            username="owner",
        )
    )
    splittable.save_custom(
        splittable.CustomSaveReq(name="tag_check", username="owner", columns=[column], expected_version=0)
    )

    schema = splittable.get_schema(product="ML_TABLE_PRODA")
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT926AA",
        wafer_ids="",
        prefix="",
        custom_name="tag_check",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert any(c["name"] == column and c["dtype"] == "custom_tag" for c in schema["columns"])
    assert result["all_columns"].count(column) == 1
    assert result["rows"][0]["_param"] == column
    assert result["rows"][0]["_display"] == "TAG_review_flag"
    assert result["rows"][0]["_cells"]["0"]["actual"] == "hold"
    assert result["rows"][0]["_cells"]["1"]["actual"] == "pass"
    assert result["rows"][0]["_cells"]["0"]["can_plan"] is False
    assert result["rows"][0]["_cells"]["0"]["is_custom_tag"] is True
    assert column not in pl.read_parquet(tmp_path / "ML_TABLE_PRODA.parquet").columns


def test_custom_tag_column_delete_removes_definition_values_and_preserves_source(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT926BB", "LOT926BB"],
        "wafer_id": ["1", "2"],
        "KNOB_ALPHA": ["A", "B"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    monkeypatch.setattr(splittable, "_audit_user", lambda *_args, **_kwargs: None)
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")

    created = splittable.save_custom_tag_column(
        splittable.CustomTagColumnReq(product="ML_TABLE_PRODA", name="review_flag", username="owner")
    )
    column = created["column"]
    splittable.save_custom_tag_values(
        splittable.CustomTagValuesReq(
            product="ML_TABLE_PRODA",
            root_lot_id="LOT926BB",
            values={
                f"LOT926BB|1|{column}": "hold",
                f"LOT926BB|2|{column}": "pass",
            },
            username="owner",
        )
    )

    deleted = splittable.delete_custom_tag_column(
        splittable.CustomTagColumnDeleteReq(product="ML_TABLE_PRODA", column=column, username="root"),
        _perm={"username": "root", "role": "admin"},
    )

    overlay = json.loads((plan_dir / "custom_tags.json").read_text(encoding="utf-8"))
    assert deleted["deleted_columns"] == 1
    assert deleted["deleted_values"] == 2
    assert all(c.get("column") != column for c in overlay["columns"])
    assert all(not key.endswith("|" + column) for key in overlay["values"])
    assert column not in pl.read_parquet(tmp_path / "ML_TABLE_PRODA.parquet").columns


def test_custom_tag_rows_keep_numeric_sort_position(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT926CC"],
        "wafer_id": ["1"],
        "KNOB_1.0 AA": ["A"],
        "KNOB_2.0 BB": ["B"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")

    created = splittable.save_custom_tag_column(
        splittable.CustomTagColumnReq(product="ML_TABLE_PRODA", name="1.5 review", username="owner")
    )
    column = created["column"]
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT926CC",
        wafer_ids="",
        prefix="",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols=f"KNOB_1.0 AA,{column},KNOB_2.0 BB",
    )

    assert column == "TAG_1.5_review"
    assert [row["_param"] for row in result["rows"]] == ["KNOB_1.0 AA", column, "KNOB_2.0 BB"]


def test_custom_tag_delete_requires_splittable_manager(monkeypatch):
    monkeypatch.setattr(auth_core, "get_page_admins", lambda: {})

    with pytest.raises(HTTPException) as exc:
        auth_core.require_page_manager("splittable")(_Request("alice", "user"))

    assert exc.value.status_code == 403

    monkeypatch.setattr(auth_core, "get_page_admins", lambda: {"splittable": ["alice"]})
    assert auth_core.require_page_manager("splittable")(_Request("alice", "user"))["username"] == "alice"
    assert auth_core.require_page_manager("splittable")(_Request("root", "admin"))["username"] == "root"


def test_custom_tag_delete_routes_are_registered():
    paths = {getattr(route, "path", "") for route in splittable.router.routes}

    assert "/api/splittable/custom-tags/delete" in paths
    assert "/api/splittable/custom-tags/columns/delete" in paths


def test_management_rows_overlay_view_custom_set_and_exports(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["LOT927AA", "LOT927AA"],
        "wafer_id": ["1", "2"],
        "KNOB_ALPHA": ["A", "B"],
    }).write_parquet(tmp_path / "ML_TABLE_PRODA.parquet")

    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "PREFIX_CFG", plan_dir / "prefix_config.json")
    monkeypatch.setattr(splittable, "SOURCE_CFG", plan_dir / "source_config.json")
    monkeypatch.setattr(splittable, "PRECISION_CFG", plan_dir / "precision_config.json")
    monkeypatch.setattr(splittable, "TRACKER_ISSUES_FILE", tmp_path / "issues.json")
    (tmp_path / "issues.json").write_text("[]", encoding="utf-8")

    created = splittable.save_management_row_column(
        splittable.ManagementRowColumnReq(product="ML_TABLE_PRODA", name="Purpose", username="owner")
    )
    column = created["column"]
    assert column == "MGMT_Purpose"

    splittable.save_management_row_values(
        splittable.ManagementRowValuesReq(
            product="ML_TABLE_PRODA",
            root_lot_id="LOT927AA",
            values={
                f"LOT927AA|1|{column}": "Reliability",
                f"LOT927AA|2|{column}": "Monitor",
            },
            username="owner",
        )
    )
    splittable.save_custom(
        splittable.CustomSaveReq(name="mgmt_check", username="owner", columns=[column], expected_version=0)
    )

    schema = splittable.get_schema(product="ML_TABLE_PRODA")
    default_result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT927AA",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT927AA",
        wafer_ids="",
        prefix="",
        custom_name="mgmt_check",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert any(c["name"] == column and c["dtype"] == "management_row" for c in schema["columns"])
    assert all(r["_param"] != column for r in default_result["rows"])
    assert result["all_columns"].count(column) == 1
    assert result["rows"][0]["_param"] == column
    assert result["rows"][0]["_display"] == "Purpose"
    assert result["rows"][0]["_cells"]["0"]["actual"] == "Reliability"
    assert result["rows"][0]["_cells"]["1"]["actual"] == "Monitor"
    assert result["rows"][0]["_cells"]["0"]["can_plan"] is False
    assert result["rows"][0]["_cells"]["0"]["is_management_row"] is True
    assert result["rows"][0]["_cells"]["0"]["can_management_edit"] is True

    overlay = json.loads((plan_dir / "management_rows.json").read_text(encoding="utf-8"))
    assert overlay["values"][f"ML_TABLE_PRODA|LOT927AA|1|{column}"]["value"] == "Reliability"
    assert column not in pl.read_parquet(tmp_path / "ML_TABLE_PRODA.parquet").columns

    csv_body = _read_streaming_response(splittable.download_csv(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT927AA",
        wafer_ids="",
        prefix="",
        custom_name="mgmt_check",
        transposed="true",
        username="owner",
        custom_cols="",
    )).decode("utf-8-sig")
    assert "Purpose" in csv_body
    assert "Reliability" in csv_body
    assert "Monitor" in csv_body
    assert "MGMT_Purpose" not in csv_body

    from openpyxl import load_workbook

    xlsx_body = _read_streaming_response(splittable.download_xlsx(
        product="ML_TABLE_PRODA",
        root_lot_id="LOT927AA",
        wafer_ids="",
        prefix="",
        custom_name="mgmt_check",
        username="owner",
        custom_cols="",
    ))
    workbook = load_workbook(io.BytesIO(xlsx_body), data_only=True)
    rows = list(workbook.active.iter_rows(values_only=True))
    purpose_row = next(row for row in rows if row and row[0] == "Purpose")
    assert "Reliability" in purpose_row
    assert "Monitor" in purpose_row


def test_lot_ids_do_not_suggest_fab_roots_that_cannot_render():
    result = splittable.get_lot_ids(product="ML_TABLE_PRODA", limit=20)

    assert result["fallback"] == ""
    assert result["fab_source"] == "lot_progress_latest_cache"
    assert "A1000" in result["lot_ids"]
    assert "A0001" not in result["lot_ids"]


def _setup_knob_meta_fixture(tmp_path, monkeypatch):
    plan_dir = tmp_path / "flow-data" / "splittable"
    plan_dir.mkdir(parents=True)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "PLAN_DIR", plan_dir)
    monkeypatch.setattr(splittable, "RULEBOOK_SCHEMA_FILE", plan_dir / "rulebook_schema.json")
    splittable._CSV_ROWS_CACHE.clear()
    splittable._SCHEMA_COLUMNS_CACHE.clear()


def test_knob_meta_accepts_ppid_knob_without_product_column_and_scopes_vehicle_steps(tmp_path, monkeypatch):
    _setup_knob_meta_fixture(tmp_path, monkeypatch)
    (tmp_path / "ppid_knob.csv").write_text(
        "feature_name,function_step,rule_order,operator,category\n"
        "5.0 PC,PC,R1,=,PPID_A\n",
        encoding="utf-8",
    )
    (tmp_path / "Vehicle_matching.csv").write_text(
        "product,step_id,function_step,module\n"
        "PRODA,STEP_PRODA,PC,PC\n"
        "PRODB,STEP_PRODB,PC,PC\n",
        encoding="utf-8",
    )

    meta = splittable._build_knob_meta("ML_TABLE_PRODA")

    assert meta["5.0 PC"]["groups"][0]["step_ids"] == ["STEP_PRODA"]
    assert "STEP_PRODB" not in meta["5.0 PC"]["groups"][0]["step_ids"]


def test_knob_meta_treats_legacy_ppid_product_column_as_common_rule(tmp_path, monkeypatch):
    _setup_knob_meta_fixture(tmp_path, monkeypatch)
    (tmp_path / "ppid_knob.csv").write_text(
        "product,feature_name,function_step,rule_order,operator,category\n"
        "PRODB,7.0 PC,PC,R1,=,PPID_B\n",
        encoding="utf-8",
    )
    (tmp_path / "Vehicle_matching.csv").write_text(
        "product,step_id,function_step,module\n"
        "PRODA,STEP_PRODA,PC,PC\n",
        encoding="utf-8",
    )

    meta = splittable._build_knob_meta("ML_TABLE_PRODA")

    assert "7.0 PC" in meta
    assert meta["7.0 PC"]["groups"][0]["step_ids"] == ["STEP_PRODA"]


def test_knob_meta_falls_back_to_step_matching_when_vehicle_matching_is_missing(tmp_path, monkeypatch):
    _setup_knob_meta_fixture(tmp_path, monkeypatch)
    (tmp_path / "ppid_knob.csv").write_text(
        "feature_name,function_step,rule_order,operator,category\n"
        "9.0 PC,PC,R1,=,PPID_C\n",
        encoding="utf-8",
    )
    (tmp_path / "step_matching.csv").write_text(
        "product,step_id,function_step,module\n"
        "PRODA,STEP_FALLBACK,PC,PC\n"
        "PRODB,STEP_OTHER,PC,PC\n",
        encoding="utf-8",
    )

    meta = splittable._build_knob_meta("ML_TABLE_PRODA")

    assert meta["9.0 PC"]["groups"][0]["step_ids"] == ["STEP_FALLBACK"]


def test_view_accepts_fab_lot_pasted_into_root_field():
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000A.1",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert result["root_lot_id"] == "A1000"
    assert result["headers"]
    assert result["rows"]
    assert "fab_lot_id" in result["lot_warn"]


def test_view_validates_root_and_fab_scope_together():
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="A1001A.1",
        custom_cols="",
    )

    assert result["root_lot_id"] == "A1000"
    assert result["headers"]
    assert result["rows"]
    assert "Root Lot ID 기준" in result["lot_warn"]
    assert all(
        group["label"] == "—" or str(group["label"]).startswith("A1000")
        for group in result["header_groups"]
    )


def test_view_keeps_matching_root_and_fab_scope_narrow():
    result = splittable.view_split(
        product="ML_TABLE_PRODA",
        root_lot_id="A1000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="A1000A.1",
        custom_cols="",
    )

    assert result["root_lot_id"] == "A1000"
    assert result["headers"]
    assert result["header_groups"] == [{"label": "A1000A.1", "span": len(result["headers"])}]
    assert result["lot_warn"] == ""


def test_mltable_product_files_are_discovered_case_insensitively(tmp_path, monkeypatch):
    fp = tmp_path / "ml_table_mixed.PARQUET"
    pl.DataFrame({
        "root_lot_id": ["R1000"],
        "wafer_id": [1],
        "KNOB_ALPHA": ["ON"],
    }).write_parquet(fp)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)

    products = splittable.list_products()["products"]
    assert products == [{
        "name": "ML_TABLE_MIXED",
        "file": "ml_table_mixed.PARQUET",
        "size": fp.stat().st_size,
        "root": "Base",
        "type": "parquet",
        "source_type": "base_file",
    }]
    assert splittable._product_path("ML_TABLE_MIXED") == fp

    result = splittable.view_split(
        product="ML_TABLE_MIXED",
        root_lot_id="R1000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )
    assert result["headers"] == ["#1"]
    assert result["rows"][0]["_param"] == "KNOB_ALPHA"


def test_root_lot_candidates_fall_back_to_detected_uppercase_column(tmp_path, monkeypatch):
    fp = tmp_path / "ML_TABLE_REAL.parquet"
    pl.DataFrame({
        "ROOT_LOT_ID": ["R2000", "R2001"],
        "WAFER_ID": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(fp)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_main_table_candidates", lambda *args, **kwargs: {"candidates": []})
    monkeypatch.setattr(splittable, "_fab_history_root_candidates", lambda *args, **kwargs: {"candidates": [], "source": ""})

    result = splittable.get_lot_candidates(
        product="ML_TABLE_REAL",
        col="root_lot_id",
        prefix="R20",
        limit=20,
        source="auto",
        root_lot_id="",
    )

    assert result["match_mode"] == "detected_lot_col_fallback"
    assert result["source_col"] == "ROOT_LOT_ID"
    assert result["candidates"] == ["R2000", "R2001"]


def test_root_lot_candidate_search_reaches_beyond_empty_preview(tmp_path, monkeypatch):
    fp = tmp_path / "ML_TABLE_BIG.parquet"
    pl.DataFrame({
        "root_lot_id": [f"R{i:04d}" for i in range(1200)],
        "wafer_id": [1 for _ in range(1200)],
        "KNOB_ALPHA": ["ON" for _ in range(1200)],
    }).write_parquet(fp)
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)

    preview = splittable.get_lot_candidates(
        product="ML_TABLE_BIG",
        col="root_lot_id",
        prefix="",
        limit=20,
        source="auto",
        root_lot_id="",
    )
    searched = splittable.get_lot_candidates(
        product="ML_TABLE_BIG",
        col="root_lot_id",
        prefix="R1199",
        limit=20,
        source="auto",
        root_lot_id="",
    )

    assert len(preview["candidates"]) <= 20
    assert "R1199" not in preview["candidates"]
    assert searched["candidates"] == ["R1199"]


def test_root_scoped_fab_candidates_keep_all_fab_history_lots(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["A1003", "A1003", "A1003", "A1003", "A1003"],
        "lot_id": ["A1003A.2", "A1003A.3", "A1003B.2", "A1003C.1", "A1003C.2"],
        "wafer_id": [1, 2, 3, 4, 5],
        "KNOB_ALPHA": ["ON", "OFF", "ON", "OFF", "ON"],
    }).write_parquet(tmp_path / "ML_TABLE_MULTI.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "MULTI" / "date=20260430"
    fab_root.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["A1003", "A1003", "A1003", "A1003", "A1003", "A1003"],
        "fab_lot_id": ["A1003A.1", "A1003A.2", "A1003A.3", "A1003B.2", "A1003C.1", "A1003C.2"],
        "wafer_id": [1, 1, 2, 3, 4, 5],
        "tkout_time": [
            "2026-04-30T10:00:00",
            "2026-04-30T11:00:00",
            "2026-04-30T12:00:00",
            "2026-04-30T13:00:00",
            "2026-04-30T14:00:00",
            "2026-04-30T15:00:00",
        ],
    }).write_parquet(fab_root / "part_0.parquet")

    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()

    result = splittable.get_lot_candidates(
        product="ML_TABLE_MULTI",
        col="fab_lot_id",
        prefix="A1003A",
        limit=20,
        source="auto",
        root_lot_id="A1003",
    )

    assert result["candidates"] == ["A1003A.2", "A1003A.3"]

    unscoped = splittable.get_lot_candidates(
        product="ML_TABLE_MULTI",
        col="fab_lot_id",
        prefix="A1003",
        limit=20,
        source="auto",
        root_lot_id="",
    )

    assert unscoped["candidates"] == ["A1003A.2", "A1003A.3", "A1003B.2", "A1003C.1", "A1003C.2"]

    root_all = splittable.get_lot_candidates(
        product="ML_TABLE_MULTI",
        col="fab_lot_id",
        prefix="A1003",
        limit=20,
        source="auto",
        root_lot_id="A1003",
    )

    assert root_all["candidates"] == ["A1003A.2", "A1003A.3", "A1003B.2", "A1003C.1", "A1003C.2"]


def test_root_lot_view_handles_categorical_fab_partitions(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9000", "R9000"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_MIXCAT.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "MIXCAT"
    part_a = fab_root / "date=20240418"
    part_b = fab_root / "date=20240419"
    part_a.mkdir(parents=True)
    part_b.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["R9000"],
        "fab_lot_id": ["F9000A.1"],
        "wafer_id": [1],
        "tkout_time": ["2024-04-18T10:00:00"],
    }).write_parquet(part_a / "part_0.parquet")
    pl.DataFrame({
        "root_lot_id": ["R9000"],
        "fab_lot_id": ["F9000A.2"],
        "wafer_id": [2],
        "tkout_time": ["2024-04-19T10:00:00"],
    }).with_columns(
        pl.col("root_lot_id").cast(pl.Categorical)
    ).write_parquet(part_b / "part_0.parquet")

    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)

    searched = splittable.get_lot_candidates(
        product="ML_TABLE_MIXCAT",
        col="root_lot_id",
        prefix="R9000",
        limit=20,
        source="auto",
        root_lot_id="",
    )
    result = splittable.view_split(
        product="ML_TABLE_MIXCAT",
        root_lot_id="R9000",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert searched["candidates"] == ["R9000"]
    assert result["headers"] == ["#1", "#2"]
    assert [g["label"] for g in result["header_groups"]] == ["F9000A.1", "F9000A.2"]
    assert result["available_fab_lots"] == ["F9000A.1", "F9000A.2"]


def test_fab_lot_id_is_exposed_when_fab_source_uses_lot_id(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9100", "R9100"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_STD.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "STD" / "date=20240420"
    fab_root.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["R9100", "R9100", "R9100"],
        "lot_id": ["F9100_OLD", "F9100A.1", "F9100A.2"],
        "wafer_id": [1, 1, 2],
        "tkout_time": ["2024-04-20T09:00:00", "2024-04-20T10:00:00", "2024-04-20T10:01:00"],
    }).write_parquet(fab_root / "part_0.parquet")

    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)

    schema_names = splittable._scan_product(
        "ML_TABLE_STD",
        root_lot_id="R9100",
    ).collect_schema().names()
    result = splittable.view_split(
        product="ML_TABLE_STD",
        root_lot_id="R9100",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert "fab_lot_id" in schema_names
    assert result["header_groups"] == [
        {"label": "F9100A.1", "span": 1},
        {"label": "F9100A.2", "span": 1},
    ]
    assert result["available_fab_lots"] == ["F9100A.1", "F9100A.2"]

    candidates = splittable.get_lot_candidates(
        product="ML_TABLE_STD",
        col="fab_lot_id",
        prefix="F9100",
        limit=20,
        source="auto",
        root_lot_id="R9100",
    )
    assert candidates["candidates"] == ["F9100A.1", "F9100A.2"]


def test_match_cache_supplies_fab_lot_without_rescanning_source(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9200", "R9200"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_CACHE.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "CACHE" / "date=20240420"
    fab_root.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["R9200", "R9200"],
        "lot_id": ["F9200A.1", "F9200A.1"],
        "wafer_id": [1, 2],
        "tkout_time": ["2024-04-20T10:00:00", "2024-04-20T10:01:00"],
    }).write_parquet(fab_root / "part_0.parquet")

    cache_dir = tmp_path / "flow-data" / "splittable" / "match_cache"
    source_cfg = tmp_path / "flow-data" / "splittable" / "source_config.json"
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "MATCH_CACHE_DIR", cache_dir)
    monkeypatch.setattr(splittable, "SOURCE_CFG", source_cfg)
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()

    built = splittable.refresh_match_cache(product="ML_TABLE_CACHE", force=True)
    assert built["products"][0]["ok"] is True
    assert built["products"][0]["row_count"] == 2
    fab_candidates = splittable.get_lot_candidates(
        product="ML_TABLE_CACHE",
        col="fab_lot_id",
        prefix="F9200",
        limit=20,
        source="auto",
        root_lot_id="R9200",
    )

    def fail_scan(_source):
        raise AssertionError("raw FAB source should not be scanned when cache exists")

    monkeypatch.setattr(splittable, "_scan_fab_source", fail_scan)

    result = splittable.view_split(
        product="ML_TABLE_CACHE",
        root_lot_id="R9200",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert result["header_groups"] == [{"label": "F9200A.1", "span": 2}]
    assert result["available_fab_lots"] == ["F9200A.1"]
    assert fab_candidates["candidates"] == ["F9200A.1"]
    assert fab_candidates["fab_source"] == "1.RAWDATA_DB_FAB/CACHE"


def test_view_uses_lot_progress_cache_before_raw_override_fallback(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9250", "R9250"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_AUTOCACHE.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "AUTOCACHE" / "date=20240420"
    fab_root.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["R9250", "R9250"],
        "lot_id": ["F9250A.1", "F9250A.2"],
        "wafer_id": [1, 2],
        "tkout_time": ["2024-04-20T10:00:00", "2024-04-20T10:01:00"],
    }).write_parquet(fab_root / "part_0.parquet")

    cache_dir = tmp_path / "flow-data" / "splittable" / "match_cache"
    source_cfg = tmp_path / "flow-data" / "splittable" / "source_config.json"
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "MATCH_CACHE_DIR", cache_dir)
    monkeypatch.setattr(splittable, "SOURCE_CFG", source_cfg)
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()
    splittable._MATCH_CACHE_AUTO_BUILD_MISS.clear()
    cache_root = tmp_path / "cache"
    cache_root.mkdir()
    pl.DataFrame({
        "product": ["ML_TABLE_AUTOCACHE", "ML_TABLE_AUTOCACHE"],
        "root_lot_id": ["R9250", "R9250"],
        "wafer_id": ["1", "2"],
        "lot_id": ["F9250A.1", "F9250A.2"],
        "step_id": ["", ""],
        "function_step": ["", ""],
        "tkout_time": ["2024-04-20T10:00:00", "2024-04-20T10:01:00"],
        "update_time": ["2024-04-20T10:02:00", "2024-04-20T10:02:00"],
    }).write_parquet(cache_root / "lot_progress_latest_lot_by_root_wafer.parquet")

    result = splittable.view_split(
        product="ML_TABLE_AUTOCACHE",
        root_lot_id="R9250",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert result["match_cache"]["hit"] is True
    assert result["match_cache"]["source"] == "lot_progress_latest_cache"
    assert result["header_groups"] == [
        {"label": "F9250A.1", "span": 1},
        {"label": "F9250A.2", "span": 1},
    ]

    def fail_scan(_source):
        raise AssertionError("raw FAB source should not be scanned after auto cache build")

    monkeypatch.setattr(splittable, "_scan_fab_source", fail_scan)
    cached_result = splittable.view_split(
        product="ML_TABLE_AUTOCACHE",
        root_lot_id="R9250",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )
    assert cached_result["match_cache"]["hit"] is True
    assert cached_result["header_groups"] == result["header_groups"]


def test_match_cache_searches_entire_fab_db_when_product_folder_is_missing(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9300", "R9300"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_NOMATCH.parquet")

    other_fab = tmp_path / "1.RAWDATA_DB_FAB" / "OTHER" / "date=20240421"
    other_fab.mkdir(parents=True)
    pl.DataFrame({
        "root_lot_id": ["R9300", "R9300"],
        "fab_lot_id": ["F9300A.1", "F9300A.1"],
        "wafer_id": [1, 2],
        "tkout_time": ["2024-04-21T10:00:00", "2024-04-21T10:01:00"],
    }).write_parquet(other_fab / "part_0.parquet")

    cache_dir = tmp_path / "flow-data" / "splittable" / "match_cache"
    source_cfg = tmp_path / "flow-data" / "splittable" / "source_config.json"
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "MATCH_CACHE_DIR", cache_dir)
    monkeypatch.setattr(splittable, "SOURCE_CFG", source_cfg)
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()

    built = splittable.refresh_match_cache(product="ML_TABLE_NOMATCH", force=True)
    result = splittable.view_split(
        product="ML_TABLE_NOMATCH",
        root_lot_id="R9300",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )

    assert built["products"][0]["ok"] is True
    assert built["products"][0]["fab_source"] == ""
    assert built["products"][0]["fab_sources"] == ["1.RAWDATA_DB_FAB/OTHER"]
    assert result["header_groups"] == [{"label": "F9300A.1", "span": 2}]
    assert result["available_fab_lots"] == ["F9300A.1"]


def test_match_cache_keeps_latest_fab_lot_per_root_wafer(tmp_path, monkeypatch):
    pl.DataFrame({
        "root_lot_id": ["R9400", "R9400"],
        "wafer_id": [1, 2],
        "KNOB_ALPHA": ["ON", "OFF"],
    }).write_parquet(tmp_path / "ML_TABLE_LATEST.parquet")

    fab_root = tmp_path / "1.RAWDATA_DB_FAB" / "LATEST" / "date=20240422"
    fab_root.mkdir(parents=True)
    (tmp_path / "step_matching.csv").write_text(
        "product,step_id,function_step\nLATEST,STEP_OLD,OLD_FUNC\nLATEST,STEP_NEW,GATE\nLATEST,STEP_W2,ETCH\n",
        encoding="utf-8",
    )
    pl.DataFrame({
        "root_lot_id": ["R9400", "R9400", "R9400"],
        "lot_id": ["F9400_OLD", "F9400_NEW", "F9400_W2"],
        "wafer_id": [1, 1, 2],
        "step_id": ["STEP_OLD", "STEP_NEW", "STEP_W2"],
        "tkout_time": [
            "2024-04-22T08:00:00",
            "2024-04-22T12:00:00",
            "2024-04-22T09:00:00",
        ],
    }).write_parquet(fab_root / "part_0.parquet")

    cache_dir = tmp_path / "flow-data" / "splittable" / "match_cache"
    source_cfg = tmp_path / "flow-data" / "splittable" / "source_config.json"
    monkeypatch.setattr(splittable, "_base_root", lambda: tmp_path)
    monkeypatch.setattr(splittable, "_db_base", lambda: tmp_path)
    monkeypatch.setattr(splittable, "MATCH_CACHE_DIR", cache_dir)
    monkeypatch.setattr(splittable, "SOURCE_CFG", source_cfg)
    import core.lot_step as lot_step
    monkeypatch.setattr(lot_step, "_get_db_root", lambda: tmp_path)
    splittable._LOT_LOOKUP_CACHE.clear()
    splittable._RGLOB_CACHE.clear()
    splittable._DB_ROOTS_CACHE.clear()

    built = splittable.refresh_match_cache(product="ML_TABLE_LATEST", force=True)
    cache_df = pl.read_parquet(cache_dir / "ML_TABLE_LATEST.parquet")
    latest_path = tmp_path / "cache" / "lot_progress_latest_lot_by_root_wafer.parquet"
    latest_df = pl.read_parquet(latest_path).sort("wafer_id")
    result = splittable.view_split(
        product="ML_TABLE_LATEST",
        root_lot_id="R9400",
        wafer_ids="",
        prefix="KNOB",
        custom_name="",
        view_mode="all",
        history_mode="all",
        fab_lot_id="",
        custom_cols="",
    )
    fab_candidates = splittable.get_lot_candidates(
        product="ML_TABLE_LATEST",
        col="fab_lot_id",
        prefix="F9400",
        limit=20,
        source="auto",
        root_lot_id="R9400",
    )
    wafer_snapshot = splittable.resolve_fab_lot_snapshot("LATEST", "R9400", "1")

    assert built["products"][0]["ok"] is True
    assert built["products"][0]["row_count"] == 2
    assert built["latest_cache"]["row_count"] == 2
    assert built["latest_cache"]["path"] == str(latest_path)
    assert not (tmp_path / "cache" / "splittable_latest_lot_step.parquet").exists()
    assert latest_df.columns == [
        "product", "root_lot_id", "wafer_id", "lot_id",
        "step_id", "function_step", "tkout_time", "update_time",
    ]
    assert latest_df.select(["root_lot_id", "wafer_id", "lot_id", "step_id", "function_step", "tkout_time"]).to_dicts() == [
        {
            "root_lot_id": "R9400",
            "wafer_id": "1",
            "lot_id": "F9400_NEW",
            "step_id": "STEP_NEW",
            "function_step": "GATE",
            "tkout_time": "2024-04-22T12:00:00",
        },
        {
            "root_lot_id": "R9400",
            "wafer_id": "2",
            "lot_id": "F9400_W2",
            "step_id": "STEP_W2",
            "function_step": "ETCH",
            "tkout_time": "2024-04-22T09:00:00",
        },
    ]
    assert all(row["update_time"] for row in latest_df.select("update_time").to_dicts())
    assert cache_df["__cache_fab_lot_id"].to_list() == ["F9400_NEW", "F9400_W2"]
    assert result["header_groups"] == [
        {"label": "F9400_NEW", "span": 1},
        {"label": "F9400_W2", "span": 1},
    ]
    assert fab_candidates["candidates"] == ["F9400_NEW", "F9400_W2"]
    assert wafer_snapshot == "F9400_NEW"
