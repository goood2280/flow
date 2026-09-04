import asyncio
import io
from pathlib import Path

import polars as pl
import pytest
from fastapi import HTTPException


def _response_bytes(response):
    async def collect():
        return b"".join([chunk async for chunk in response.body_iterator])

    return asyncio.run(collect())


def test_purpose_is_builtin_first_and_cannot_be_deleted(monkeypatch):
    from routers import splittable

    store = {
        "columns": [
            {"product": "P1", "column": "TAG_zeta", "label": "zeta"},
            {"product": "P1", "column": "TAG_alpha", "label": "alpha"},
        ],
        "values": {},
        "colors": {},
    }
    monkeypatch.setattr(splittable, "_load_custom_tags_data", lambda: store)

    columns = splittable._custom_tag_columns_for_product("P1")

    assert columns[0]["column"] == "TAG_purpose"
    assert columns[0]["label"] == "purpose"
    assert columns[0]["builtin"] is True
    assert [column["column"] for column in columns[1:]] == ["TAG_zeta", "TAG_alpha"]
    ordered = sorted(
        ["TAG_zeta", "KNOB_A", "TAG_purpose", "TAG_alpha"],
        key=lambda column: splittable._step_order_sort_key(column, column, {}),
    )
    assert ordered[:3] == ["TAG_purpose", "TAG_alpha", "TAG_zeta"]
    assert splittable._with_default_custom_tag(["KNOB_A"]) == ["TAG_purpose", "KNOB_A"]

    with pytest.raises(HTTPException) as error:
        splittable.delete_custom_tag_column(
            splittable.CustomTagColumnDeleteReq(product="P1", column="TAG_purpose")
        )
    assert error.value.status_code == 400


def test_purpose_value_and_color_are_saved_per_wafer_and_expanded(monkeypatch):
    from routers import splittable

    store = {"columns": [], "values": {}, "colors": {}}
    monkeypatch.setattr(splittable, "_load_custom_tags_data", lambda: store)
    monkeypatch.setattr(splittable, "_save_custom_tags_data", lambda data: None)

    result = splittable.save_custom_tag_values(
        splittable.CustomTagValuesReq(
            product="P1",
            values={"L1|1|TAG_purpose": "DOE"},
            colors={"L1|1|TAG_purpose": "#fecaca"},
            username="tester",
        )
    )

    assert result["saved"] == 1
    assert result["colors_saved"] == 1
    assert splittable._custom_tag_values_for_root("P1", "L1") == {
        "L1|1|TAG_purpose": "DOE"
    }
    assert splittable._custom_tag_colors_for_root("P1", "L1") == {
        "L1|1|TAG_purpose": "#fecaca"
    }

    expanded = splittable._expand_view_rows({
        "root_lot_id": "L1",
        "wafer_keys": ["1"],
        "rows_compact": [{
            "_param": "TAG_purpose",
            "_display": "TAG_purpose",
            "a": ["DOE"],
            "tag": True,
            "tc": {"0": "#fecaca"},
        }],
    })
    cell = expanded["rows"][0]["_cells"]["0"]
    assert cell["actual"] == "DOE"
    assert cell["tag_color"] == "#fecaca"
    assert cell["is_custom_tag"] is True


def test_purpose_is_in_csv_and_xlsx_keeps_its_background(monkeypatch):
    from openpyxl import load_workbook
    from routers import splittable

    frame = pl.DataFrame({
        "root_lot_id": ["L1"],
        "fab_lot_id": ["L1.1"],
        "wafer_id": [1],
        "KNOB_A": ["PP_A"],
    })
    monkeypatch.setattr(splittable, "_product_path", lambda *args, **kwargs: None)
    monkeypatch.setattr(splittable, "_scan_product", lambda *args, **kwargs: frame.lazy())
    monkeypatch.setattr(splittable, "_load_plan_data", lambda *args, **kwargs: {"plans": {}})
    # Even a legacy/empty TAG catalog must not make the built-in purpose row
    # disappear from a KNOB-only selection or its exports.
    monkeypatch.setattr(splittable, "_custom_tag_label_map", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_custom_tag_values_for_root", lambda *args, **kwargs: {"L1|1|TAG_purpose": "DOE"})
    monkeypatch.setattr(splittable, "_custom_tag_colors_for_root", lambda *args, **kwargs: {"L1|1|TAG_purpose": "#fecaca"})
    monkeypatch.setattr(splittable, "_management_row_label_map", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_management_row_values_for_root", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_split_step_order_context", lambda *args, **kwargs: {"param_rank": {}})
    monkeypatch.setattr(splittable, "_split_step_progress", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_log_split_table_download", lambda *args, **kwargs: None)

    csv_response = splittable.download_csv(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", transposed="true", username="u", custom_cols="",
        step_labels="", exclude_not_null="1",
    )
    csv_text = _response_bytes(csv_response).decode("utf-8-sig")
    assert "TAG_purpose,DOE" in csv_text

    xlsx_response = splittable.download_xlsx(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", username="u", custom_cols="", display_mode="",
        step_labels="", exclude_not_null="1",
    )
    sheet = load_workbook(io.BytesIO(_response_bytes(xlsx_response))).active
    assert sheet.cell(6, 1).value == "TAG_purpose"
    assert sheet.cell(6, 2).value == "DOE"
    assert sheet.cell(6, 2).fill.fill_type == "solid"
    assert sheet.cell(6, 2).fill.fgColor.rgb.endswith("FECACA")


def test_split_and_pems_views_keep_purpose_as_a_fixed_header_row():
    root = Path(__file__).resolve().parents[1]
    page = (root / "frontend/src/features/splittable/My_SplitTable.jsx").read_text("utf-8")
    snapshot = (root / "frontend/src/components/SplitTableSnapshotView.jsx").read_text("utf-8")

    assert "purpose_row:purposeViewRow" in page
    assert "purpose_row: purposeRow || undefined" in snapshot
    assert "{hasPurposeRow && (" in snapshot
    assert "const splitSourceRows = rows.filter(row => !isPurposeTagRow(row));" in snapshot
