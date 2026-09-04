import asyncio
import io
from pathlib import Path

import polars as pl


def _response_bytes(response):
    async def collect():
        return b"".join([chunk async for chunk in response.body_iterator])

    return asyncio.run(collect())


def test_process_columns_keep_step_id_and_step_desc_separate():
    from routers import splittable

    metas = {
        "knob": {
            "KNOB_A": {
                "groups": [
                    {"rule_order": "R1", "step_desc": "ETCH", "step_ids": ["S10", "S11"]},
                    {"rule_order": "R1", "step_desc": "CLEAN", "step_ids": ["S20"]},
                    {"rule_order": "R2", "step_desc": "SKIP", "step_ids": ["S30"], "operator": "not_null"},
                ]
            }
        },
        "inline": {},
        "vm": {},
    }

    columns = splittable._step_process_columns_for_param("KNOB_A", metas, exclude_not_null=True)

    assert columns == {
        "step_id": "S10\nS11\nS20",
        "step_desc": "ETCH\nCLEAN",
    }


def test_xlsx_process_columns_dedupe_across_knob_rule_orders():
    from routers import splittable

    repeated = {
        "knob": {
            "KNOB_A": {
                "groups": [
                    {"rule_order": "R1", "step_desc": "ETCH", "step_ids": ["S10"]},
                    {"rule_order": "R2", "step_desc": "ETCH", "step_ids": ["S10"]},
                    {"rule_order": "R3", "step_desc": "ETCH", "step_ids": ["S10"]},
                ]
            }
        },
        "inline": {},
        "vm": {},
    }

    columns = splittable._step_process_columns_for_param("KNOB_A", repeated)

    assert columns == {"step_id": "S10", "step_desc": "ETCH"}


def test_step_order_interleaves_mapped_prefixes_and_leaves_mask_unranked(monkeypatch):
    from routers import splittable

    splittable._STEP_ORDER_CTX_CACHE.clear()
    monkeypatch.setattr(splittable, "_s0_sop_catalog", lambda: {})
    monkeypatch.setattr(splittable, "_load_knob_step_matching_rows", lambda *args, **kwargs: [
        {"product": "P1", "step_id": "ST300", "step_desc": "FAB STEP"},
        {"product": "P1", "step_id": "ST100", "step_desc": "KNOB STEP"},
        {"product": "P1", "step_id": "ST200", "step_desc": "INLINE STEP"},
        {"product": "P1", "step_id": "ST500", "step_desc": "MASK STEP"},
        {"product": "P1", "step_id": "ST400", "step_desc": "VM STEP"},
    ])
    monkeypatch.setattr(splittable, "_sch", lambda name: {
        "product_col": "product", "step_id_col": "step_id", "step_desc_col": "step_desc",
    } if name == "step_matching" else {})
    monkeypatch.setattr(splittable, "_load_prefixes", lambda: ["KNOB", "FAB", "MASK", "INLINE", "VM"])
    monkeypatch.setattr(splittable, "_mltable_schema_columns", lambda *args, **kwargs: [
        "KNOB_K", "FAB_F", "MASK_M", "INLINE_I", "VM_V",
    ])
    inferred = {
        "FAB": {"FAB_F": {"groups": [{"step_ids": ["ST300"]}]}},
    }
    monkeypatch.setattr(
        splittable, "_inferred_stage_meta",
        lambda product, prefix: inferred.get(str(prefix).upper(), {}),
    )
    monkeypatch.setattr(splittable, "_build_knob_meta", lambda *args, **kwargs: {
        "K": {"groups": [{"step_ids": ["ST100"]}]},
    })
    monkeypatch.setattr(splittable, "_build_inline_meta", lambda *args, **kwargs: {
        "I": {"groups": [{"step_ids": ["ST200"]}]},
    })
    monkeypatch.setattr(splittable, "_build_vm_meta", lambda *args, **kwargs: {
        "V": {"groups": [{"step_ids": ["ST400"]}]},
    })

    context = splittable._split_step_order_context("P1")
    source = ["VM_V", "MASK_M", "KNOB_K", "FAB_F", "INLINE_I"]
    ordered = sorted(
        source,
        key=lambda column: splittable._step_order_sort_key(
            column, column, context["param_rank"],
        ),
    )

    # f_step이 없을 때는 기존 Vehicle_matching 행 순서를 유지한다.
    assert ordered == ["FAB_F", "KNOB_K", "INLINE_I", "VM_V", "MASK_M"]
    assert context["param_step"]["FAB_F"] == "ST300"
    assert "MASK_M" not in context["param_rank"]
    splittable._STEP_ORDER_CTX_CACHE.clear()


def test_f_step_route_tracks_unmatched_current_step_and_excludes_unmapped_rows(monkeypatch):
    from routers import splittable

    splittable._STEP_ORDER_CTX_CACHE.clear()
    monkeypatch.setattr(splittable, "_s0_sop_catalog", lambda: {
        "p1": {
            "product": "P1",
            "step_order": ["ST100", "ST200", "ST300"],
            "rows": {
                "st100": {"step_id": "ST100", "ppid": "PP1"},
                "st200": {"step_id": "ST200", "ppid": "PP2"},
                "st300": {"step_id": "ST300", "ppid": "PP3"},
            },
        }
    })
    monkeypatch.setattr(splittable, "_load_knob_step_matching_rows", lambda *args, **kwargs: [
        {"product": "P1", "step_id": "ST100", "step_desc": "UPPER"},
        {"product": "P1", "step_id": "ST300", "step_desc": "LOWER"},
        {"product": "P1", "step_id": "ST999", "step_desc": "NOT_IN_ROUTE"},
    ])
    monkeypatch.setattr(splittable, "_sch", lambda name: {
        "product_col": "product", "step_id_col": "step_id", "step_desc_col": "step_desc",
    } if name == "step_matching" else {})
    monkeypatch.setattr(splittable, "_load_prefixes", lambda: ["KNOB"])
    monkeypatch.setattr(splittable, "_mltable_schema_columns", lambda *args, **kwargs: [
        "KNOB_UPPER", "KNOB_LOWER", "KNOB_NOT_IN_ROUTE", "KNOB_NO_STEP",
    ])
    monkeypatch.setattr(splittable, "_inferred_stage_meta", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_build_knob_meta", lambda *args, **kwargs: {
        "UPPER": {"groups": [{"step_ids": ["ST100"]}]},
        "LOWER": {"groups": [{"step_ids": ["ST300"]}]},
        "NOT_IN_ROUTE": {"groups": [{"step_ids": ["ST999"]}]},
        "NO_STEP": {"groups": []},
    })
    monkeypatch.setattr(splittable, "_build_inline_meta", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_build_vm_meta", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_root_latest_step_state", lambda *args, **kwargs: {
        "step_id": "ST200",
        "by_wafer": {"1": "ST200"},
    })

    progress = splittable._split_step_progress(
        "P1", "ROOT1",
        ["KNOB_UPPER", "KNOB_LOWER", "KNOB_NOT_IN_ROUTE", "KNOB_NO_STEP"],
        [1],
    )

    assert progress["tracked"] == ["KNOB_UPPER", "KNOB_LOWER"]
    assert progress["not_reached"] == ["KNOB_LOWER"]
    assert progress["by_wafer"]["1"]["not_reached"] == ["KNOB_LOWER"]
    context = splittable._split_step_order_context("P1")
    assert context["seq_rank"]["ST200"] == 1
    assert "KNOB_NOT_IN_ROUTE" not in context["progress_param_rank"]
    splittable._STEP_ORDER_CTX_CACHE.clear()


def test_fab_missing_greys_only_f_step_tracked_parameters(monkeypatch):
    from routers import splittable

    monkeypatch.setattr(splittable, "_split_step_order_context", lambda product: {
        "progress_param_rank": {"KNOB_MAPPED": 1},
    })

    progress = splittable._split_step_progress(
        "P1", "ROOT1", ["KNOB_MAPPED", "KNOB_NO_STEP"], [1], fab_present=False,
    )

    assert progress["tracked"] == ["KNOB_MAPPED"]
    assert progress["not_reached"] == ["KNOB_MAPPED"]
    assert progress["by_wafer"]["1"]["not_reached"] == ["KNOB_MAPPED"]


def test_split_table_unmatched_steps_do_not_move_grey_boundary():
    source = (
        Path(__file__).parents[1]
        / "frontend"
        / "src"
        / "features"
        / "splittable"
        / "My_SplitTable.jsx"
    ).read_text(encoding="utf-8")

    assert "trackedProgressParams" in source
    assert "if(rowTracksStepProgress[ri])lastFilledRowByCol[ci]=ri" in source
    assert "if(!rowTracksStepProgress[ri])return false" in source


def test_display_settings_save_normalizes_shared_column_widths(tmp_path, monkeypatch):
    from routers import splittable

    settings_file = tmp_path / "display_settings.json"
    monkeypatch.setattr(splittable, "DISPLAY_SETTINGS_CFG", settings_file)

    saved = splittable.save_display_settings(
        splittable.DisplaySettingsReq(column_widths={
            "module": 120,
            "step_id": 220,
            "step_desc": 9999,
            "item": 360,
            "value": 40,
        }),
        _perm={"role": "admin"},
    )

    assert saved["column_widths"] == {
        "module": 120,
        "step_id": 220,
        "step_desc": 640,
        "item": 360,
        "value": 48,
        "split": 80,
        "wafer": 115,
    }
    assert splittable.get_display_settings() == {
        "column_widths": saved["column_widths"]
    }


def test_default_view_merges_context_labels_and_uses_configurable_widths():
    root = Path(__file__).parents[1]
    page = (root / "frontend/src/features/splittable/My_SplitTable.jsx").read_text(encoding="utf-8")
    snapshot = (root / "frontend/src/components/SplitTableSnapshotView.jsx").read_text(encoding="utf-8")

    assert "const leftPrefixColumnCount=1+(showModuleCol?1:0)+(showParamMeta?2:0)" in page
    assert "colSpan={leftPrefixColumnCount} title={lotContextTitle}" in page
    assert "colSpan={leftPrefixColumnCount} style={{boxSizing:\"border-box\",height:purposeHeaderHeight" in page
    assert "columnWidths={columnWidths}" in page
    assert 'sf(API+"/display-settings")' in page
    assert 'sf(API+"/display-settings/save"' in page
    for label in ("module", "step_id", "step_desc", "항목", "값", "Split", "wafer"):
        assert f'\"{label}\"' in page
    assert "normalizeSplitTableColumnWidths" in snapshot
    assert "effectiveColumnWidths.value" in snapshot


def test_merged_view_keeps_live_and_snapshot_context_headers_visually_unified():
    root = Path(__file__).parents[1]
    page = (root / "frontend/src/features/splittable/My_SplitTable.jsx").read_text(encoding="utf-8")
    snapshot = (root / "frontend/src/components/SplitTableSnapshotView.jsx").read_text(encoding="utf-8")

    assert "const mergedContextLeftStyle=mergedViewActive?" in page
    assert page.count('stm-context-left--merged') >= 3
    assert "maxWidth:leftPrefixWidth" in page
    assert "const mergedContextLeftStyle = mergedMode ?" in snapshot
    assert snapshot.count('stm-context-left--merged') >= 3
    assert "maxWidth: prefixTotalWidth" in snapshot


def test_stage_inference_keeps_vehicle_steps_without_numeric_step_desc(monkeypatch):
    from routers import splittable

    monkeypatch.setattr(splittable, "_load_knob_step_matching_rows", lambda *args, **kwargs: [
        {"product": "P1", "step_id": "CC942300", "step_desc": "GATE_ETCH"},
    ])
    monkeypatch.setattr(splittable, "_sch", lambda name: {
        "product_col": "product", "step_id_col": "step_id", "step_desc_col": "step_desc",
    } if name == "step_matching" else {})
    monkeypatch.setattr(splittable, "_mltable_schema_columns", lambda product, prefix="": [
        "FAB_4.0 GATE_OX",
    ] if str(prefix).upper() == "FAB" else [])

    meta = splittable._inferred_stage_meta("P1", "FAB")

    assert meta["FAB_4.0 GATE_OX"]["step_ids"] == ["CC942300"]


def test_knob_virtual_columns_emit_one_row_per_rulebook_feature(monkeypatch):
    from routers import splittable

    pc = {"feature_name": "5.0 PC", "groups": []}
    ldd = {"feature_name": "6.0 LDD", "groups": []}
    monkeypatch.setattr(splittable, "_build_knob_meta", lambda *args, **kwargs: {
        "5.0 PC": pc,
        "KNOB_5.0 PC": pc,
        "5.0_PC_Split": pc,
        "KNOB_5.0_PC_Split": pc,
        "6.0 LDD": ldd,
        "KNOB_6.0_LDD_Split": ldd,
    })

    virtual = splittable._virtual_columns_for_prefix("P1", "KNOB")

    assert virtual == ["KNOB_5.0 PC", "KNOB_6.0 LDD"]


def test_knob_virtual_columns_do_not_duplicate_a_physical_alias(monkeypatch):
    from routers import splittable

    pc = {"feature_name": "5.0 PC", "groups": []}
    ldd = {"feature_name": "6.0 LDD", "groups": []}
    monkeypatch.setattr(splittable, "_build_knob_meta", lambda *args, **kwargs: {
        "5.0 PC": pc,
        "KNOB_5.0_PC_Split": pc,
        "6.0 LDD": ldd,
        "KNOB_6.0_LDD_Split": ldd,
    })

    virtual = splittable._virtual_columns_for_prefix(
        "P1", "KNOB", existing_columns=["KNOB_5.0_PC_Split"],
    )

    assert virtual == ["KNOB_6.0 LDD"]


def _patch_export_source(monkeypatch):
    from routers import splittable

    frame = pl.DataFrame({
        "root_lot_id": ["L1"],
        "fab_lot_id": ["L1.1"],
        "wafer_id": [1],
        "KNOB_A": ["PP_A"],
    })
    monkeypatch.setattr(splittable, "_product_path", lambda *args, **kwargs: None)
    monkeypatch.setattr(splittable, "_scan_product", lambda *args, **kwargs: frame.lazy())
    monkeypatch.setattr(splittable, "_build_step_process_columns", lambda *args, **kwargs: {
        "KNOB_A": {"step_id": "S10", "step_desc": "ETCH"},
    })
    monkeypatch.setattr(splittable, "_load_plan_data", lambda *args, **kwargs: {"plans": {}})
    monkeypatch.setattr(splittable, "_custom_tag_label_map", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_custom_tag_colors_for_root", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_management_row_label_map", lambda *args, **kwargs: {})
    monkeypatch.setattr(splittable, "_split_step_order_context", lambda *args, **kwargs: {"param_rank": {}})
    monkeypatch.setattr(splittable, "_log_split_table_download", lambda *args, **kwargs: None)


def test_csv_process_columns_precede_preserved_parameter(monkeypatch):
    from routers import splittable

    _patch_export_source(monkeypatch)
    response = splittable.download_csv(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", transposed="true", username="u", custom_cols="",
        step_labels="1", exclude_not_null="1",
    )
    text = _response_bytes(response).decode("utf-8-sig")

    assert "step_id,step_desc,Parameter,#1" in text
    assert "S10,ETCH,KNOB_A,PP_A" in text


def test_xlsx_process_columns_precede_preserved_parameter(monkeypatch):
    from openpyxl import load_workbook
    from routers import splittable

    _patch_export_source(monkeypatch)
    monkeypatch.setattr(splittable, "_split_step_progress", lambda *args, **kwargs: {})
    response = splittable.download_xlsx(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", username="u", custom_cols="", display_mode="",
        step_labels="1", exclude_not_null="1",
    )

    workbook = load_workbook(io.BytesIO(_response_bytes(response)))
    sheet = workbook.active

    assert [sheet.cell(5, col).value for col in range(1, 5)] == ["step_id", "step_desc", "Parameter", "#1"]
    assert [sheet.cell(6, col).value for col in range(1, 5)] == [None, None, "TAG_purpose", None]
    assert [sheet.cell(7, col).value for col in range(1, 5)] == ["S10", "ETCH", "KNOB_A", "PP_A"]


def test_xlsx_applied_process_columns_match_web_unique_values(monkeypatch):
    from openpyxl import load_workbook
    from routers import splittable

    _patch_export_source(monkeypatch)
    repeated = {
        "knob": {
            "KNOB_A": {
                "groups": [
                    {"rule_order": "R1", "step_desc": "ETCH", "step_ids": ["S10"]},
                    {"rule_order": "R2", "step_desc": "ETCH", "step_ids": ["S10"]},
                ]
            }
        },
        "inline": {},
        "vm": {},
    }
    monkeypatch.setattr(splittable, "_step_label_metas", lambda *args, **kwargs: repeated)
    monkeypatch.setattr(splittable, "_split_step_progress", lambda *args, **kwargs: {})

    response = splittable.download_xlsx(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", username="u", custom_cols="", display_mode="",
        step_labels="1", exclude_not_null="1",
    )
    sheet = load_workbook(io.BytesIO(_response_bytes(response))).active

    assert sheet.cell(7, 1).value == "S10"
    assert sheet.cell(7, 2).value == "ETCH"


def test_split_check_xlsx_uses_the_same_process_prefix(monkeypatch):
    from openpyxl import load_workbook
    from routers import splittable

    _patch_export_source(monkeypatch)
    monkeypatch.setattr(splittable, "_split_step_progress", lambda *args, **kwargs: {})
    response = splittable.download_xlsx(
        product="P1", root_lot_id="L1", wafer_ids="", prefix="KNOB",
        custom_name="", username="u", custom_cols="", display_mode="split_check",
        step_labels="1", exclude_not_null="1",
    )

    sheet = load_workbook(io.BytesIO(_response_bytes(response))).active

    assert [sheet.cell(5, col).value for col in range(1, 7)] == [
        "step_id", "step_desc", "항목", "값", "Split", "#1",
    ]
    assert [sheet.cell(6, col).value for col in range(1, 7)] == [
        "S10", "ETCH", "KNOB_A", "PP_A", "S0", "✓",
    ]


def test_inform_snapshot_preserves_parameter_beside_process_columns():
    from routers import informs

    embed = {
        "st_view": {
            "headers": ["#1"],
            "root_lot_id": "L1",
            "step_labels": True,
            "rows": [{
                "_param": "KNOB_A",
                "_display": "KNOB_A",
                "_process_columns": {"step_id": "S10", "step_desc": "ETCH"},
                "_cells": {"0": {"actual": "PP_A", "plan": ""}},
            }],
        },
        "step_labels": True,
    }

    result = informs._apply_step_labels_to_embed(embed, "P1")
    html = informs._render_embed_table_html(result)

    assert result["columns"][:3] == ["step_id", "step_desc", "parameter"]
    assert result["rows"][0][:4] == ["S10", "ETCH", "KNOB_A", "PP_A"]
    assert all(value in html for value in ["step_id", "step_desc", "S10", "ETCH"])
    assert ">A</td>" in html


def test_inform_split_check_keeps_process_prefix_columns():
    from routers import informs

    embed = {
        "st_view": {
            "headers": ["#1"],
            "root_lot_id": "L1",
            "step_labels": True,
            "rows": [{
                "_param": "KNOB_A",
                "_display": "KNOB_A",
                "_process_columns": {"step_id": "S10", "step_desc": "ETCH"},
                "_cells": {"0": {"actual": "PP_A", "plan": ""}},
            }],
        },
        "step_labels": True,
    }

    result = informs._convert_splittable_embed_to_split_check(embed)

    assert result["st_view"]["prefix_columns"] == ["step_id", "step_desc", "항목", "값", "Split"]
    assert result["st_view"]["parameter_prefix_index"] == 2
    assert result["rows"][0][:6] == ["S10", "ETCH", "A", "PP_A", "S0", "✓"]
